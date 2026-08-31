import re
from collections.abc import Mapping
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any
from urllib.parse import urlsplit
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.model_configuration import (
    MODEL_CONFIGURATION_CATEGORY_NAME,
    MODEL_CONFIGURATION_CATEGORY_SOURCE_ID,
    ModelConfigurationRecord,
    ModelConfigurationSyncError,
    parse_model_configuration_payload,
)


MAX_IMPORT_ROWS = 500
MAX_MODEL_CONFIGURATION_IMPORT_ROWS = 5000
MAX_IMPORT_FILE_BYTES = 5 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
IMPORT_SHEET_NAME = "知识导入"
KNOWLEDGE_UPDATE_SHEET_NAME = "知识批量修改"
MODEL_CONFIGURATION_IMPORT_SHEET_NAME = "机型配置信息"
MODEL_CONFIGURATION_LEGACY_SHEET_NAME = "个性化配置信息"
EXPORT_SHEET_NAME = "知识库主表"

MODEL_CONFIGURATION_HEADER_ALIASES = {
    "source_record_id": {"来源知识ID", "知识ID", "记录ID"},
    "title": {"标题", "主标题"},
    "category_id": {"品类ID"},
    "category_name": {"品类", "品类名称"},
    "brand_id": {"品牌ID"},
    "brand_name": {"品牌", "品牌名称"},
    "model_id": {"型号ID"},
    "model_name": {"型号", "机型", "机型名称"},
    "content": {"综合内容", "正文", "知识内容"},
}
MODEL_CONFIGURATION_REQUIRED_HEADER_LABELS = {
    "title": "标题",
    "brand_id": "品牌ID",
    "brand_name": "品牌",
    "model_id": "型号ID",
    "model_name": "型号",
    "content": "综合内容",
}

EXPORT_HEADERS = [
    "知识ID",
    "主题键",
    "记录ID",
    "知识键",
    "主标题",
    "副标题",
    "知识内容",
    "知识来源",
    "业务类型",
    "知识分类",
    "录入方式",
    "关联标准项",
    "适用范围",
    "生效状态",
    "来源版本",
    "变更类型",
    "创建类型",
    "失效类型",
    "失效原因",
    "来源追溯",
    "校验备注",
]

EXPORT_STATUS_LABELS = {
    "draft": "草稿",
    "review": "待审核",
    "published": "生效中",
    "deprecated": "已失效",
}

EXPORT_SOURCE_LABELS = {
    "manual": "手工录入",
    "excel": "Excel 批量导入",
    "integration": "接口导入",
    "automation": "接口导入",
    "candidate": "候选知识",
}

HEADER_ALIASES = {
    "knowledge_id": {"知识ID", "知识库ID", "中台知识ID"},
    "title": {"标题", "知识标题", "主标题"},
    "knowledge_origin": {"知识来源", "业务来源"},
    "business_type": {"业务类型", "所属业务类型"},
    "category": {"知识分类", "所属分类", "分类", "知识分类ID", "分类ID"},
    "content": {"正文", "知识正文", "知识内容", "内容"},
    "subtitles": {"副标题", "副标题列表"},
    "scenes": {"场景标签", "适用场景"},
    "scope": {"适用范围"},
    "source_status": {"生效状态"},
    "applicable_categories": {"适用类目"},
    "brands": {"适用品牌", "品牌"},
    "models": {"适用机型", "机型"},
    "related_standard_items": {"关联标准项", "关联标准", "标准项"},
    "source_topic_key": {"主题键"},
    "source_record_id": {"记录ID"},
    "source_knowledge_key": {"知识键"},
}

EXPORT_HEADER_IMPORT_FIELDS = {
    "主题键": "source_topic_key",
    "记录ID": "source_record_id",
    "知识键": "source_knowledge_key",
    "主标题": "title",
    "副标题": "subtitles",
    "知识内容": "content",
    "业务类型": "business_type",
    "知识分类": "category",
    "知识来源": "knowledge_origin",
    "业务来源": "knowledge_origin",
    "关联标准项": "related_standard_items",
    "适用范围": "scope",
    "生效状态": "source_status",
}

CATEGORY_VALUE_ALIASES = {
    "场景判定": "质检标准",
    "标准定义": "质检标准",
    "检测方法": "操作流程",
}

BUSINESS_TYPE_LABELS = {
    "self_operated": "自营回收",
    "aggregated": "聚合回收",
}
BUSINESS_TYPE_VALUE_ALIASES = {
    "自营回收": "self_operated",
    "聚合回收": "aggregated",
    "self_operated": "self_operated",
    "aggregated": "aggregated",
}

KNOWLEDGE_ORIGIN_LABELS = {
    "headquarters_standard": "总部标准",
    "business_accumulation": "业务沉淀",
    "model_configuration": "机型配置信息",
}
KNOWLEDGE_ORIGIN_VALUE_ALIASES = {
    "总部标准": "headquarters_standard",
    "业务沉淀": "business_accumulation",
    "机型配置信息": "model_configuration",
    "headquarters_standard": "headquarters_standard",
    "business_accumulation": "business_accumulation",
    "model_configuration": "model_configuration",
}

VALID_SOURCE_STATUSES = {"生效中", "待审核", "已禁用"}
IMPORTABLE_SOURCE_STATUS = "生效中"
REVIEW_SOURCE_STATUS = "待审核"
DEPRECATED_SOURCE_STATUS = "已禁用"
UNRESTRICTED_SCOPES = {"通用"}
LEGACY_SCOPE_BRAND_VALUES = {
    "苹果",
    "华为",
    "OPPO",
    "VIVO",
    "vivo",
    "三星",
    "其他品牌",
}
LEGACY_SCOPE_BRAND_ALIASES = {
    "小米/红米": ["小米", "红米"],
    "小米／红米": ["小米", "红米"],
}
EXTERNAL_MEDIA_TOKEN_PATTERN = re.compile(
    r"\[(?P<kind>img|video):[ \t]*"
    r"(?P<url>https://[^\s\[\]<>\"']+)\]",
    re.IGNORECASE,
)
MEDIA_PLACEHOLDER_LINE_PATTERN = re.compile(
    r"^[+\-‐‑‒–—―*•·●○▪▫]$"
)
LOCAL_MEDIA_EXPORT_PLACEHOLDER_PATTERN = re.compile(
    r"^\[(?:图片|视频)(?:：[^\]]*)?\]$"
)


class KnowledgeExcelError(ValueError):
    """工作簿级错误，整个文件无法继续解析。

    ``issues`` is optional structured context used by the upload preflight to
    report row/field/code/reason details for workbook-level validation errors
    (for example, model-configuration rows whose payload parser fails).  The
    string form remains unchanged for existing callers and task history.
    """

    def __init__(self, message: str, *, issues: list[dict[str, Any]] | None = None):
        super().__init__(message)
        self.issues = list(issues or [])


class KnowledgeExcelRowError(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


@dataclass
class ExcelKnowledgeRow:
    row_number: int
    title: str
    knowledge_id: str = ""
    knowledge_origin: str = ""
    business_type: str = ""
    category_id: str = ""
    content: Any = ""
    subtitles: list[str] | None = None
    applicable_scenes: list[str] | None = None
    applicable_categories: list[str] | None = None
    applicable_brands: list[str] | None = None
    applicable_models: list[str] | None = None
    related_standard_items: list[str] | None = None
    source_topic_key: str = ""
    source_record_id: str = ""
    source_knowledge_key: str = ""
    source_fields: dict[str, str] = field(default_factory=dict)
    source_status: str = ""
    source_scope: str = ""
    provided_fields: set[str] = field(default_factory=set)
    import_mode: str = "knowledge"
    error_code: str | None = None
    error_message: str | None = None

    @property
    def is_valid(self) -> bool:
        return self.error_code is None


@dataclass(frozen=True)
class ParsedModelConfigurationWorkbook:
    records: list[ModelConfigurationRecord]
    row_numbers: list[int]


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value) -> str:
    text = _cell_text(value)
    text = re.sub(r"[（(].*?[）)]", "", text)
    return text.replace("*", "").replace(" ", "").strip()


def _split_values(value) -> list[str]:
    text = _cell_text(value)
    if not text:
        return []
    return [item.strip() for item in re.split(r"[；;|\n]+", text) if item.strip()]


def _merge_values(*groups: list[str]) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for group in groups:
        for value in group:
            if value not in seen:
                seen.add(value)
                merged.append(value)
    return merged


def _legacy_scope_brands(value) -> list[str]:
    """兼容旧主表：仅将明确品牌值从“适用范围”补入适用品牌。"""
    brands: list[str] = []
    for scope in _split_values(value):
        if scope in LEGACY_SCOPE_BRAND_ALIASES:
            brands.extend(LEGACY_SCOPE_BRAND_ALIASES[scope])
        elif scope in LEGACY_SCOPE_BRAND_VALUES:
            brands.append(scope)
    return _merge_values(brands)


def _is_safe_external_url(value: str) -> bool:
    try:
        parsed = urlsplit(value)
        return (
            parsed.scheme.lower() == "https"
            and bool(parsed.hostname)
            and not parsed.username
            and not parsed.password
        )
    except ValueError:
        return False


def _trim_media_placeholder_lines(
    segment: str,
    *,
    has_media_before: bool,
    has_media_after: bool,
) -> str:
    """仅清理紧邻媒体标记、且独占一行的列表占位符。"""
    lines = segment.replace("\r\n", "\n").replace("\r", "\n").split("\n")

    if has_media_before:
        while lines and not lines[0].strip():
            lines.pop(0)
        while (
            lines
            and MEDIA_PLACEHOLDER_LINE_PATTERN.fullmatch(lines[0].strip())
        ):
            lines.pop(0)
            while lines and not lines[0].strip():
                lines.pop(0)

    if has_media_after:
        while lines and not lines[-1].strip():
            lines.pop()
        while (
            lines
            and MEDIA_PLACEHOLDER_LINE_PATTERN.fullmatch(lines[-1].strip())
        ):
            lines.pop()
            while lines and not lines[-1].strip():
                lines.pop()

    return "\n".join(lines)


def _content_with_external_media(value) -> str | dict[str, Any]:
    text = _cell_text(value)
    blocks: list[dict[str, str]] = []
    found_media = False
    cursor = 0

    def append_text(
        segment: str,
        *,
        has_media_before: bool,
        has_media_after: bool,
    ) -> None:
        segment = _trim_media_placeholder_lines(
            segment,
            has_media_before=has_media_before,
            has_media_after=has_media_after,
        )
        if segment.strip():
            blocks.append({"type": "text", "value": segment})

    for match in EXTERNAL_MEDIA_TOKEN_PATTERN.finditer(text):
        external_url = match.group("url").strip()
        if not _is_safe_external_url(external_url):
            continue
        media_type = "image" if match.group("kind").lower() == "img" else "video"
        append_text(
            text[cursor : match.start()],
            has_media_before=found_media,
            has_media_after=True,
        )
        blocks.append(
            {
                "type": media_type,
                "external_url": external_url,
                "alt": "",
                "caption": "",
            }
        )
        cursor = match.end()
        found_media = True
    append_text(
        text[cursor:],
        has_media_before=found_media,
        has_media_after=False,
    )

    return {"blocks": blocks} if found_media else text


def _category_records(categories) -> tuple[dict[str, object], dict[str, list[str]], dict[str, str]]:
    by_id = {str(category.id): category for category in categories}
    by_name: dict[str, list[str]] = {}
    path_by_id: dict[str, str] = {}

    def build_path(category_id: str, visited: set[str] | None = None) -> str:
        if category_id in path_by_id:
            return path_by_id[category_id]
        category = by_id[category_id]
        visited = set(visited or ())
        if category_id in visited:
            return str(category.name)
        visited.add(category_id)
        parent_id = str(category.parent_id) if category.parent_id else ""
        if parent_id and parent_id in by_id:
            path = f"{build_path(parent_id, visited)}/{category.name}"
        else:
            path = str(category.name)
        path_by_id[category_id] = path
        return path

    for category_id, category in by_id.items():
        name = str(category.name).strip()
        by_name.setdefault(name, []).append(category_id)
        build_path(category_id)

    return by_id, by_name, path_by_id


def _resolve_category(
    value,
    category_records: tuple[dict[str, object], dict[str, list[str]], dict[str, str]],
) -> str:
    text = _cell_text(value)
    if not text:
        raise KnowledgeExcelRowError("CATEGORY_REQUIRED", "知识分类不能为空。")

    by_id, by_name, path_by_id = category_records
    if text in by_id:
        return text

    path_matches = [
        category_id
        for category_id, path in path_by_id.items()
        if path == text
    ]
    if len(path_matches) == 1:
        return path_matches[0]

    name_matches = by_name.get(text, [])
    if len(name_matches) == 1:
        return name_matches[0]
    if len(name_matches) > 1:
        raise KnowledgeExcelRowError(
            "CATEGORY_AMBIGUOUS",
            f"分类名称“{text}”存在重名，请填写分类ID或完整分类路径。",
        )

    mapped_name = CATEGORY_VALUE_ALIASES.get(text)
    if mapped_name:
        mapped_matches = by_name.get(mapped_name, [])
        if len(mapped_matches) == 1:
            return mapped_matches[0]
        if len(mapped_matches) > 1:
            raise KnowledgeExcelRowError(
                "CATEGORY_AMBIGUOUS",
                f"兼容分类“{text}”映射到“{mapped_name}”后存在重名，"
                "请填写分类ID或完整分类路径。",
            )
    raise KnowledgeExcelRowError(
        "CATEGORY_NOT_FOUND",
        f"分类“{text}”不存在，请从模板的“分类字典”工作表中选择。",
    )


def _resolve_business_type(value) -> str:
    text = _cell_text(value)
    if not text:
        raise KnowledgeExcelRowError(
            "BUSINESS_TYPE_REQUIRED",
            "业务类型不能为空。",
        )
    normalized = text.lower()
    business_type = BUSINESS_TYPE_VALUE_ALIASES.get(normalized)
    if business_type:
        return business_type
    raise KnowledgeExcelRowError(
        "BUSINESS_TYPE_INVALID",
        f"业务类型“{text}”不受支持，仅允许自营回收、聚合回收、"
        "self_operated 或 aggregated。",
    )


def _resolve_knowledge_origin(value) -> str:
    text = _cell_text(value)
    if not text:
        raise KnowledgeExcelRowError(
            "KNOWLEDGE_ORIGIN_REQUIRED",
            "知识来源不能为空。",
        )
    normalized = text.lower()
    knowledge_origin = KNOWLEDGE_ORIGIN_VALUE_ALIASES.get(normalized)
    if knowledge_origin == "model_configuration":
        raise KnowledgeExcelRowError(
            "KNOWLEDGE_ORIGIN_MANAGED",
            "机型配置信息由飞书专用同步维护，不能通过普通 Excel 导入。",
        )
    if knowledge_origin:
        return knowledge_origin
    raise KnowledgeExcelRowError(
        "KNOWLEDGE_ORIGIN_INVALID",
        f"知识来源“{text}”不受支持，仅允许总部标准、业务沉淀、"
        "headquarters_standard 或 business_accumulation；"
        "机型配置信息由飞书专用同步维护。",
    )


def _header_indexes(
    header_row,
    *,
    require_knowledge_id: bool = False,
) -> dict[str, int]:
    if require_knowledge_id:
        for field, aliases in HEADER_ALIASES.items():
            matching_headers = [
                _normalize_header(value)
                for value in header_row
                if _normalize_header(value) in aliases
            ]
            if len(matching_headers) <= 1:
                continue
            if field == "knowledge_id":
                raise KnowledgeExcelError(
                    "批量修改文件只能保留一个知识ID列，"
                    "不能同时使用知识ID、知识库ID或中台知识ID。"
                )
            raise KnowledgeExcelError(
                "批量修改文件中同一字段只能保留一个兼容列："
                + "、".join(matching_headers)
                + "。"
            )
    normalized = {
        _normalize_header(value): index
        for index, value in enumerate(header_row)
        if _normalize_header(value)
    }
    indexes: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                indexes[field] = normalized[alias]
                break

    is_source_deprecation_sheet = (
        not require_knowledge_id
        and
        "source_status" in indexes
        and any(
            field in indexes
            for field in (
                "source_knowledge_key",
                "source_topic_key",
                "source_record_id",
            )
        )
    )
    missing = []
    if require_knowledge_id and "knowledge_id" not in indexes:
        missing.append("知识ID")
    if "knowledge_origin" not in indexes:
        missing.append("知识来源")
    if "business_type" not in indexes:
        missing.append("业务类型")
    if require_knowledge_id or not is_source_deprecation_sheet:
        missing.extend(
            label
            for field, label in (
                ("title", "标题"),
                ("category", "知识分类"),
                ("content", "正文"),
            )
            if field not in indexes
        )
    if missing:
        raise KnowledgeExcelError(
            f"缺少必填列：{'、'.join(missing)}。请使用系统下载的最新模板。"
        )
    return indexes


def _source_fields(header_row, values) -> dict[str, str]:
    """保留上传表中所有非空表头对应的原始单元格值，供后续导出还原。"""
    fields: dict[str, str] = {}
    for index, header in enumerate(header_row):
        normalized_header = _normalize_header(header)
        if not normalized_header:
            continue
        value = values[index] if index < len(values) else None
        fields[normalized_header] = _cell_text(value)
    return fields


def _validate_xlsx_container(data: bytes) -> None:
    if not data.startswith(b"PK"):
        raise KnowledgeExcelError("文件不是有效的 .xlsx 工作簿。")
    try:
        with ZipFile(BytesIO(data)) as archive:
            total_size = sum(entry.file_size for entry in archive.infolist())
    except BadZipFile as exc:
        raise KnowledgeExcelError("文件不是有效的 .xlsx 工作簿。") from exc
    if total_size > MAX_UNCOMPRESSED_BYTES:
        raise KnowledgeExcelError("Excel 解压后体积过大，请拆分后导入。")


def _model_configuration_header_indexes(header_row) -> dict[str, int]:
    normalized_headers: dict[str, list[int]] = {}
    for index, value in enumerate(header_row):
        normalized = _normalize_header(value)
        if normalized:
            normalized_headers.setdefault(normalized, []).append(index)
    duplicated_headers = [
        header
        for header, indexes in normalized_headers.items()
        if len(indexes) > 1
    ]
    if duplicated_headers:
        raise KnowledgeExcelError(
            "机型配置信息存在重复表头："
            + "、".join(duplicated_headers[:10])
            + "。"
        )

    indexes: dict[str, int] = {}
    for field, aliases in MODEL_CONFIGURATION_HEADER_ALIASES.items():
        matches = [
            normalized_headers[alias][0]
            for alias in aliases
            if alias in normalized_headers
        ]
        if len(matches) > 1:
            raise KnowledgeExcelError(
                f"机型配置信息中字段“{field}”存在多个兼容列，请只保留一列。"
            )
        if matches:
            indexes[field] = matches[0]

    missing = [
        label
        for field, label in MODEL_CONFIGURATION_REQUIRED_HEADER_LABELS.items()
        if field not in indexes
    ]
    if missing:
        raise KnowledgeExcelError(
            "机型配置信息缺少必填列："
            + "、".join(missing)
            + "。请使用系统下载的最新模板。"
        )
    has_category_id = "category_id" in indexes
    has_category_name = "category_name" in indexes
    if has_category_id != has_category_name:
        raise KnowledgeExcelError(
            "品类ID和品类必须同时提供；原始飞书表可同时省略这两列。"
        )
    return indexes


def _worksheet_has_data(sheet) -> bool:
    return any(
        any(_cell_text(value) for value in values)
        for values in sheet.iter_rows(min_row=2, values_only=True)
    )


def parse_model_configuration_workbook(
    data: bytes,
) -> ParsedModelConfigurationWorkbook:
    if not data:
        raise KnowledgeExcelError("Excel 文件为空。")
    if len(data) > MAX_IMPORT_FILE_BYTES:
        raise KnowledgeExcelError("Excel 文件不能超过 5MB。")
    _validate_xlsx_container(data)

    try:
        workbook = load_workbook(
            BytesIO(data),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise KnowledgeExcelError("Excel 文件损坏或无法读取。") from exc

    matching_sheets = [
        name
        for name in (
            MODEL_CONFIGURATION_IMPORT_SHEET_NAME,
            MODEL_CONFIGURATION_LEGACY_SHEET_NAME,
        )
        if name in workbook.sheetnames
    ]
    if not matching_sheets:
        raise KnowledgeExcelError(
            "未找到“机型配置信息”工作表；也未找到兼容的"
            "“个性化配置信息”工作表。"
        )
    populated_matching_sheets = [
        name
        for name in matching_sheets
        if _worksheet_has_data(workbook[name])
    ]
    if len(populated_matching_sheets) > 1:
        raise KnowledgeExcelError(
            "工作簿同时包含“机型配置信息”和“个性化配置信息”，"
            "请一次只保留一个待导入工作表。"
        )
    selected_sheet_name = (
        populated_matching_sheets[0]
        if populated_matching_sheets
        else matching_sheets[0]
    )
    if (
        IMPORT_SHEET_NAME in workbook.sheetnames
        and _worksheet_has_data(workbook[IMPORT_SHEET_NAME])
    ):
        raise KnowledgeExcelError(
            "工作簿同时包含普通知识和机型配置信息数据，"
            "请拆分为两个文件分别导入。"
        )

    sheet = workbook[selected_sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise KnowledgeExcelError("机型配置信息中没有可读取的表头。") from exc
    indexes = _model_configuration_header_indexes(header_row)
    data_column_indexes = [
        index
        for index, header in enumerate(header_row)
        if _normalize_header(header)
    ]

    def value_at(values, field: str):
        index = indexes.get(field)
        return values[index] if index is not None and index < len(values) else None

    raw_records: list[dict[str, Any]] = []
    row_numbers: list[int] = []
    seen_model_key_rows: dict[tuple[str, str, str], int] = {}
    validation_errors: list[dict[str, Any]] = []
    data_row_count = 0
    category_columns_present = (
        "category_id" in indexes and "category_name" in indexes
    )
    for row_number, values in enumerate(rows, start=2):
        if not any(
            _cell_text(values[index])
            for index in data_column_indexes
            if index < len(values)
        ):
            continue
        data_row_count += 1
        source_fields = _source_fields(header_row, values)
        if data_row_count > MAX_MODEL_CONFIGURATION_IMPORT_ROWS:
            raise KnowledgeExcelError(
                "机型配置信息单次最多导入 "
                f"{MAX_MODEL_CONFIGURATION_IMPORT_ROWS} 条，请拆分文件后重试。"
            )

        category_id = (
            _cell_text(value_at(values, "category_id"))
            if category_columns_present
            else MODEL_CONFIGURATION_CATEGORY_SOURCE_ID
        )
        category_name = (
            _cell_text(value_at(values, "category_name"))
            if category_columns_present
            else MODEL_CONFIGURATION_CATEGORY_NAME
        )
        if category_columns_present and (not category_id or not category_name):
            validation_errors.append(
                {
                    "row": row_number,
                    "title": _cell_text(value_at(values, "title")),
                    "field": "品类ID/品类",
                    "code": "MODEL_CONFIGURATION_FIELD_REQUIRED",
                    "message": "品类ID和品类不能为空。",
                }
            )
            continue

        source_fields["来源工作表"] = sheet.title
        source_fields["来源行号"] = str(row_number)
        raw_record = {
            "source_record_id": _cell_text(
                value_at(values, "source_record_id")
            ),
            "title": _cell_text(value_at(values, "title")),
            "category_id": category_id,
            "category_name": category_name,
            "brand_id": _cell_text(value_at(values, "brand_id")),
            "brand_name": _cell_text(value_at(values, "brand_name")),
            "model_id": _cell_text(value_at(values, "model_id")),
            "model_name": _cell_text(value_at(values, "model_name")),
            "content": _cell_text(value_at(values, "content")),
            "source_fields": source_fields,
        }
        try:
            record = parse_model_configuration_payload(
                {"records": [raw_record]}
            )[0]
        except ModelConfigurationSyncError as exc:
            reason = str(exc)
            field = next(
                (
                    label
                    for label in MODEL_CONFIGURATION_REQUIRED_HEADER_LABELS.values()
                    if f"“{label}”" in reason
                ),
                "数据",
            )
            validation_errors.append(
                {
                    "row": row_number,
                    "title": raw_record["title"],
                    "field": field,
                    "code": exc.code,
                    "message": reason,
                }
            )
            continue

        model_key = (
            record.category_id,
            record.brand_id,
            record.model_id,
        )
        previous_model_row = seen_model_key_rows.get(model_key)
        if previous_model_row is not None:
            validation_errors.append(
                {
                    "row": row_number,
                    "title": record.title,
                    "field": "品类/品牌/型号ID",
                    "code": "MODEL_CONFIGURATION_MODEL_ID_DUPLICATED",
                    "message": (
                        "组合 "
                        f"{record.category_id}/{record.brand_id}/{record.model_id} "
                        f"与第 {previous_model_row} 行重复。"
                    ),
                }
            )
            continue
        seen_model_key_rows[model_key] = row_number
        raw_records.append(raw_record)
        row_numbers.append(row_number)

    if validation_errors:
        preview = "；".join(
            f"第 {issue['row']} 行【{issue['field']}】：{issue['message']}"
            for issue in validation_errors[:100]
        )
        suffix = (
            f"；另有 {len(validation_errors) - 100} 条错误未展开。"
            if len(validation_errors) > 100
            else ""
        )
        raise KnowledgeExcelError(
            "机型配置信息全表校验失败：" + preview + suffix,
            issues=validation_errors,
        )
    if not raw_records:
        raise KnowledgeExcelError("机型配置信息中没有可导入的数据行。")

    try:
        records = parse_model_configuration_payload(
            {"records": raw_records}
        )
    except ModelConfigurationSyncError as exc:
        raise KnowledgeExcelError(f"机型配置工作簿校验失败：{exc}") from exc
    return ParsedModelConfigurationWorkbook(
        records=records,
        row_numbers=row_numbers,
    )


def parse_knowledge_workbook(
    data: bytes,
    categories,
    *,
    update_mode: bool = False,
) -> list[ExcelKnowledgeRow]:
    if not data:
        raise KnowledgeExcelError("Excel 文件为空。")
    if len(data) > MAX_IMPORT_FILE_BYTES:
        raise KnowledgeExcelError("Excel 文件不能超过 5MB。")
    _validate_xlsx_container(data)

    try:
        workbook = load_workbook(
            BytesIO(data),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise KnowledgeExcelError("Excel 文件损坏或无法读取。") from exc

    populated_model_configuration_sheets = [
        name
        for name in (
            MODEL_CONFIGURATION_IMPORT_SHEET_NAME,
            MODEL_CONFIGURATION_LEGACY_SHEET_NAME,
        )
        if (
            name in workbook.sheetnames
            and _worksheet_has_data(workbook[name])
        )
    ]
    if (
        IMPORT_SHEET_NAME in workbook.sheetnames
        and _worksheet_has_data(workbook[IMPORT_SHEET_NAME])
        and populated_model_configuration_sheets
    ):
        raise KnowledgeExcelError(
            "工作簿同时包含普通知识和机型配置信息数据，"
            "请拆分为两个文件分别导入。"
        )

    if update_mode and KNOWLEDGE_UPDATE_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[KNOWLEDGE_UPDATE_SHEET_NAME]
    elif not update_mode and IMPORT_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[IMPORT_SHEET_NAME]
    else:
        sheet = workbook.active
    if update_mode and sheet.title == EXPORT_SHEET_NAME:
        raise KnowledgeExcelError(
            "“知识库主表”普通导出文件不能直接用于批量修改；"
            "请下载“知识批量修改”专用模板后填写。"
        )
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise KnowledgeExcelError("Excel 中没有可读取的表头。") from exc
    indexes = _header_indexes(
        header_row,
        require_knowledge_id=update_mode,
    )
    category_records = _category_records(categories)

    def value_at(values, field: str):
        index = indexes.get(field)
        return values[index] if index is not None and index < len(values) else None

    parsed_rows: list[ExcelKnowledgeRow] = []
    seen_knowledge_id_rows: dict[str, int] = {}
    duplicate_knowledge_ids: list[tuple[str, int, int]] = []
    regular_import_contains_knowledge_id = False
    for row_number, values in enumerate(rows, start=2):
        if not any(_cell_text(value) for value in values):
            continue
        if len(parsed_rows) >= MAX_IMPORT_ROWS:
            raise KnowledgeExcelError(
                f"单次最多导入 {MAX_IMPORT_ROWS} 条知识，请拆分文件后重试。"
            )

        title = _cell_text(value_at(values, "title"))
        knowledge_id = _cell_text(value_at(values, "knowledge_id"))
        if not update_mode and knowledge_id:
            regular_import_contains_knowledge_id = True
        knowledge_origin = _cell_text(value_at(values, "knowledge_origin"))
        business_type = _cell_text(value_at(values, "business_type"))
        source_status = _cell_text(value_at(values, "source_status"))
        source_scope = _cell_text(value_at(values, "scope"))
        source_topic_key = _cell_text(value_at(values, "source_topic_key"))
        source_record_id = _cell_text(value_at(values, "source_record_id"))
        source_knowledge_key = _cell_text(value_at(values, "source_knowledge_key"))
        result = ExcelKnowledgeRow(
            row_number=row_number,
            title=title,
            knowledge_id=knowledge_id,
            source_fields=_source_fields(header_row, values),
            source_status=source_status,
            source_scope=source_scope,
            source_topic_key=source_topic_key,
            source_record_id=source_record_id,
            source_knowledge_key=source_knowledge_key,
            provided_fields=set(indexes),
            import_mode="knowledge_update" if update_mode else "knowledge",
        )
        try:
            if update_mode:
                if not knowledge_id:
                    raise KnowledgeExcelRowError(
                        "KNOWLEDGE_ID_REQUIRED",
                        "知识ID不能为空。",
                    )
                if len(knowledge_id) > 64:
                    raise KnowledgeExcelRowError(
                        "KNOWLEDGE_ID_TOO_LONG",
                        "知识ID不能超过 64 个字符。",
                    )
                previous_row = seen_knowledge_id_rows.get(knowledge_id)
                if previous_row is not None:
                    duplicate_knowledge_ids.append(
                        (knowledge_id, previous_row, row_number)
                    )
                else:
                    seen_knowledge_id_rows[knowledge_id] = row_number
            result.knowledge_origin = _resolve_knowledge_origin(knowledge_origin)
            result.business_type = _resolve_business_type(business_type)
            if not update_mode and "source_status" in indexes:
                if not source_status:
                    raise KnowledgeExcelRowError(
                        "SOURCE_STATUS_REQUIRED",
                        "生效状态不能为空。",
                    )
                if source_status not in VALID_SOURCE_STATUSES:
                    raise KnowledgeExcelRowError(
                        "SOURCE_STATUS_INVALID",
                        f"生效状态“{source_status}”不受支持，"
                        "仅允许生效中、待审核或已禁用。",
                    )
                if source_status == DEPRECATED_SOURCE_STATUS:
                    if not any(
                        (
                            source_knowledge_key,
                            source_topic_key,
                            source_record_id,
                        )
                    ):
                        raise KnowledgeExcelRowError(
                            "SOURCE_IDENTIFIER_REQUIRED",
                            "“已禁用”记录至少需要填写知识键、主题键或记录ID，"
                            "以定位需要废弃的原知识。",
                        )
                    parsed_rows.append(result)
                    continue

                if source_status not in {
                    IMPORTABLE_SOURCE_STATUS,
                    REVIEW_SOURCE_STATUS,
                }:
                    raise KnowledgeExcelRowError(
                        "SOURCE_STATUS_NOT_IMPORTABLE",
                        f"该记录为“{source_status}”，无法处理。",
                    )

            if not title:
                raise KnowledgeExcelRowError("TITLE_REQUIRED", "标题不能为空。")
            if len(title) > 256:
                raise KnowledgeExcelRowError("TITLE_TOO_LONG", "标题不能超过 256 个字符。")

            content = _cell_text(value_at(values, "content"))
            if not content:
                raise KnowledgeExcelRowError("CONTENT_REQUIRED", "正文不能为空。")
            if len(content) > 100_000:
                raise KnowledgeExcelRowError(
                    "CONTENT_TOO_LONG",
                    "单条正文不能超过 100000 个字符。",
                )
            if update_mode and any(
                LOCAL_MEDIA_EXPORT_PLACEHOLDER_PATTERN.fullmatch(
                    line.strip()
                )
                for line in content.splitlines()
                if line.strip()
            ):
                raise KnowledgeExcelRowError(
                    "LOCAL_MEDIA_PLACEHOLDER_UNSUPPORTED",
                    "检测到普通导出的本地图片或视频占位文本；"
                    "该格式不能安全回填媒体，请使用专用模板，"
                    "含本地媒体的知识请在页面中单条编辑。",
                )

            result.category_id = _resolve_category(
                value_at(values, "category"),
                category_records,
            )
            result.content = _content_with_external_media(content)
            result.subtitles = _split_values(value_at(values, "subtitles"))
            scope_tags = [
                f"适用范围：{scope}"
                for scope in _split_values(source_scope)
                if scope not in UNRESTRICTED_SCOPES
            ]
            result.applicable_scenes = _merge_values(
                _split_values(value_at(values, "scenes")),
                scope_tags,
            )
            result.applicable_categories = _split_values(
                value_at(values, "applicable_categories")
            )
            explicit_brands = _split_values(value_at(values, "brands"))
            result.applicable_brands = (
                explicit_brands
                if explicit_brands
                else _legacy_scope_brands(source_scope)
            )
            result.applicable_models = _split_values(value_at(values, "models"))
            result.related_standard_items = _split_values(
                value_at(values, "related_standard_items")
            )
        except KnowledgeExcelRowError as exc:
            result.error_code = exc.code
            result.error_message = str(exc)
        parsed_rows.append(result)

    if not parsed_rows:
        raise KnowledgeExcelError("Excel 中没有可导入的数据行。")
    if regular_import_contains_knowledge_id:
        raise KnowledgeExcelError(
            "检测到非空“知识ID”列；该文件属于批量修改数据，"
            "请切换到“批量修改”模式，系统不会按知识ID执行普通新增导入。",
            issues=[
                {
                    "row": row.row_number,
                    "title": row.title,
                    "field": "知识ID",
                    "code": "KNOWLEDGE_ID_NOT_ALLOWED",
                    "message": (
                        "普通新增导入不能填写知识ID；如需按ID覆盖，"
                        "请切换到“批量修改”模式。"
                    ),
                }
                for row in parsed_rows
                if row.knowledge_id
            ],
        )
    if duplicate_knowledge_ids:
        details = "；".join(
            f"{knowledge_id}（第 {first_row}、{duplicate_row} 行）"
            for knowledge_id, first_row, duplicate_row
            in duplicate_knowledge_ids[:10]
        )
        raise KnowledgeExcelError(
            f"同一 Excel 中知识ID不能重复：{details}。",
            issues=[
                {
                    "row": duplicate_row,
                    "title": next(
                        (
                            row.title
                            for row in parsed_rows
                            if row.row_number == duplicate_row
                        ),
                        "",
                    ),
                    "field": "知识ID",
                    "code": "KNOWLEDGE_ID_DUPLICATED",
                    "message": (
                        f"知识ID“{knowledge_id}”与第 {first_row} 行重复。"
                    ),
                }
                for knowledge_id, first_row, duplicate_row
                in duplicate_knowledge_ids
            ],
        )
    return parsed_rows


def parse_knowledge_update_workbook(
    data: bytes,
    categories,
) -> list[ExcelKnowledgeRow]:
    return parse_knowledge_workbook(
        data,
        categories,
        update_mode=True,
    )


def _export_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Mapping):
        for key in ("name", "label", "value", "title", "id"):
            item = value.get(key)
            if item not in (None, ""):
                return _export_cell_text(item)
        return ""
    return str(value).strip()


def _export_join(values: Any, separator: str = "；") -> str:
    if not isinstance(values, (list, tuple, set)):
        return _export_cell_text(values)
    result: list[str] = []
    for value in values:
        text = _export_cell_text(value)
        if text and text not in result:
            result.append(text)
    return separator.join(result)


def _export_content(value: Any) -> str:
    if isinstance(value, str):
        return value.strip()
    if not isinstance(value, Mapping):
        return _export_cell_text(value)

    blocks = value.get("blocks")
    if not isinstance(blocks, list):
        return _export_cell_text(value)

    pieces: list[str] = []
    for block in blocks:
        if not isinstance(block, Mapping):
            text = _export_cell_text(block)
            if text:
                pieces.append(text)
            continue
        block_type = _export_cell_text(block.get("type")).lower()
        if block_type == "text":
            text = _export_cell_text(block.get("value") or block.get("text"))
        elif block_type in {"image", "video"}:
            external_url = _export_cell_text(block.get("external_url"))
            if external_url:
                token = "img" if block_type == "image" else "video"
                text = f"[{token}:{external_url}]"
            else:
                label = _export_cell_text(block.get("caption") or block.get("alt"))
                media_label = "图片" if block_type == "image" else "视频"
                text = f"[{media_label}{'：' + label if label else ''}]"
        else:
            text = _export_cell_text(block.get("value") or block.get("text"))
        if text:
            pieces.append(text)
    return "\n".join(pieces)


def _export_scope(item: Any) -> str:
    groups = (
        ("场景", getattr(item, "applicable_scenes", None)),
        ("适用类目", getattr(item, "applicable_categories", None)),
        ("适用品牌", getattr(item, "applicable_brands", None)),
        ("适用机型", getattr(item, "applicable_models", None)),
    )
    parts = [
        f"{label}：{values}"
        for label, raw_values in groups
        if (values := _export_join(raw_values))
    ]
    return "；".join(parts) if parts else "通用"


def _export_status(value: Any) -> str:
    raw = _export_cell_text(getattr(value, "value", value)).lower()
    return EXPORT_STATUS_LABELS.get(raw, _export_cell_text(value))


def _export_source(value: Any) -> str:
    raw = _export_cell_text(value)
    return EXPORT_SOURCE_LABELS.get(raw, raw)


def _export_business_type(value: Any) -> str:
    raw = _export_cell_text(getattr(value, "value", value))
    normalized = raw.lower()
    canonical = BUSINESS_TYPE_VALUE_ALIASES.get(normalized, normalized)
    return BUSINESS_TYPE_LABELS.get(canonical, raw)


def _export_knowledge_origin(value: Any) -> str:
    raw = _export_cell_text(getattr(value, "value", value))
    normalized = raw.lower()
    canonical = KNOWLEDGE_ORIGIN_VALUE_ALIASES.get(normalized, normalized)
    return KNOWLEDGE_ORIGIN_LABELS.get(canonical, raw)


def _export_source_field(source_fields: Any, header: str, fallback: str) -> str:
    """导出时优先使用上传时保留的原始字段，旧数据则使用系统字段兜底。"""
    if not isinstance(source_fields, dict):
        return fallback
    candidates = [_normalize_header(header)]
    import_field = EXPORT_HEADER_IMPORT_FIELDS.get(header)
    if import_field:
        candidates.extend(
            _normalize_header(alias)
            for alias in HEADER_ALIASES.get(import_field, set())
        )
    for candidate in candidates:
        if candidate in source_fields:
            return _export_cell_text(source_fields[candidate])
    return fallback


def build_knowledge_export_workbook(items) -> bytes:
    """生成与历史知识库主表字段一致的只读导出工作簿。"""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = EXPORT_SHEET_NAME
    sheet.append(EXPORT_HEADERS)

    for item in items:
        category = getattr(item, "category", None)
        category_name = _export_cell_text(getattr(category, "name", None))
        if not category_name:
            category_name = _export_cell_text(getattr(item, "category_id", None))
        source_fields = getattr(item, "source_fields", None)
        business_type = getattr(item, "business_type", None)
        if not _export_cell_text(business_type):
            business_type = _export_source_field(
                source_fields,
                "业务类型",
                "",
            )
        knowledge_origin = getattr(item, "knowledge_origin", None)
        if not _export_cell_text(knowledge_origin):
            knowledge_origin = _export_source_field(
                source_fields,
                "知识来源",
                "",
            )
        sheet.append(
            [
                _export_cell_text(getattr(item, "id", None)),
                _export_source_field(source_fields, "主题键", _export_cell_text(getattr(item, "source_topic_key", None))),
                _export_source_field(source_fields, "记录ID", _export_cell_text(getattr(item, "source_record_id", None))),
                _export_source_field(source_fields, "知识键", _export_cell_text(getattr(item, "source_knowledge_key", None))),
                _export_source_field(source_fields, "主标题", _export_cell_text(getattr(item, "title", None))),
                _export_source_field(source_fields, "副标题", _export_join(getattr(item, "subtitles", None), separator="\n")),
                _export_source_field(source_fields, "知识内容", _export_content(getattr(item, "content", None))),
                _export_knowledge_origin(knowledge_origin),
                _export_business_type(business_type),
                _export_source_field(source_fields, "知识分类", category_name),
                _export_source(getattr(item, "source", None)),
                _export_source_field(source_fields, "关联标准项", _export_join(getattr(item, "related_standard_items", None)),),
                _export_source_field(source_fields, "适用范围", _export_scope(item)),
                _export_source_field(source_fields, "生效状态", _export_status(getattr(item, "status", None))),
                _export_source_field(source_fields, "来源版本", ""),
                _export_source_field(source_fields, "变更类型", ""),
                _export_source_field(source_fields, "创建类型", ""),
                _export_source_field(source_fields, "失效类型", ""),
                _export_source_field(source_fields, "失效原因", ""),
                _export_source_field(source_fields, "来源追溯", ""),
                _export_source_field(source_fields, "校验备注", ""),
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9E8FB")
    header_font = Font(name="宋体", size=11, bold=True)
    body_font = Font(name="宋体", size=11)
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    body_alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:U{max(sheet.max_row, 1)}"
    sheet.row_dimensions[1].height = 30
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = body_alignment

    column_widths = {
        "A": 14,
        "B": 18,
        "C": 16,
        "D": 16,
        "E": 32,
        "F": 36,
        "G": 72,
        "H": 16,
        "I": 16,
        "J": 18,
        "K": 18,
        "L": 28,
        "M": 42,
        "N": 14,
        "O": 16,
        "P": 16,
        "Q": 16,
        "R": 16,
        "S": 22,
        "T": 28,
        "U": 28,
    }
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_knowledge_import_template(categories) -> bytes:
    workbook = Workbook()
    import_sheet = workbook.active
    import_sheet.title = IMPORT_SHEET_NAME
    knowledge_origin_sheet = workbook.create_sheet("知识来源字典")
    business_type_sheet = workbook.create_sheet("业务类型字典")
    dictionary_sheet = workbook.create_sheet("分类字典")
    instructions_sheet = workbook.create_sheet("填写说明")

    headers = [
        "标题（必填）",
        "知识来源（必填）",
        "业务类型（必填）",
        "知识分类（必填）",
        "正文（必填）",
        "副标题（选填）",
        "场景标签（选填）",
        "关联标准项（选填）",
        "适用类目（选填）",
        "适用品牌（选填）",
        "适用机型（选填）",
    ]
    import_sheet.append(headers)
    import_sheet.freeze_panes = "A2"
    import_sheet.auto_filter.ref = "A1:K1"
    import_sheet.row_dimensions[1].height = 28
    import_sheet.column_dimensions["A"].width = 32
    import_sheet.column_dimensions["B"].width = 22
    import_sheet.column_dimensions["C"].width = 22
    import_sheet.column_dimensions["D"].width = 28
    import_sheet.column_dimensions["E"].width = 70
    for column in ("F", "G", "H", "I", "J", "K"):
        import_sheet.column_dimensions[column].width = 24

    required_fill = PatternFill("solid", fgColor="0F766E")
    optional_fill = PatternFill("solid", fgColor="475569")
    for index, cell in enumerate(import_sheet[1], start=1):
        cell.fill = required_fill if index <= 5 else optional_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    import_sheet["A1"].comment = Comment("必填，最多 256 个字符。", "知识库")
    import_sheet["B1"].comment = Comment(
        "必填。可填写总部标准、业务沉淀，或对应代码 "
        "headquarters_standard、business_accumulation。"
        "机型配置信息由飞书专用同步维护。",
        "知识库",
    )
    import_sheet["C1"].comment = Comment(
        "必填。可填写自营回收、聚合回收，或对应代码 "
        "self_operated、aggregated。",
        "知识库",
    )
    import_sheet["D1"].comment = Comment(
        "必填。优先填写分类ID，也支持唯一分类名称或完整分类路径。",
        "知识库",
    )
    import_sheet["E1"].comment = Comment(
        "必填。仅处理插件自动回填的 [img:https://...] 或 "
        "[video:https://...] 标记；其他 URL 保持原文。",
        "知识库",
    )
    import_sheet["F1"].comment = Comment("多项请使用中文分号“；”分隔。", "知识库")
    for cell_ref in ("G1", "H1", "I1", "J1", "K1"):
        import_sheet[cell_ref].comment = Comment(
            "多项请使用中文分号“；”分隔。",
            "知识库",
        )

    knowledge_origin_sheet.append(["知识来源代码", "知识来源名称"])
    for code, label in KNOWLEDGE_ORIGIN_LABELS.items():
        if code == "model_configuration":
            continue
        knowledge_origin_sheet.append([code, label])
    knowledge_origin_sheet.freeze_panes = "A2"
    knowledge_origin_sheet.auto_filter.ref = (
        f"A1:B{max(knowledge_origin_sheet.max_row, 1)}"
    )
    knowledge_origin_sheet.column_dimensions["A"].width = 28
    knowledge_origin_sheet.column_dimensions["B"].width = 24
    for cell in knowledge_origin_sheet[1]:
        cell.fill = required_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    business_type_sheet.append(["业务类型代码", "业务类型名称"])
    for code, label in BUSINESS_TYPE_LABELS.items():
        business_type_sheet.append([code, label])
    business_type_sheet.freeze_panes = "A2"
    business_type_sheet.auto_filter.ref = (
        f"A1:B{max(business_type_sheet.max_row, 1)}"
    )
    business_type_sheet.column_dimensions["A"].width = 24
    business_type_sheet.column_dimensions["B"].width = 24
    for cell in business_type_sheet[1]:
        cell.fill = required_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    dictionary_sheet.append(["分类ID", "分类名称", "完整分类路径"])
    by_id, _, path_by_id = _category_records(categories)
    sorted_categories = sorted(
        by_id.values(),
        key=lambda item: (
            int(getattr(item, "level", 1) or 1),
            int(getattr(item, "sort_order", 0) or 0),
            str(item.name),
        ),
    )
    for category in sorted_categories:
        category_id = str(category.id)
        dictionary_sheet.append(
            [category_id, str(category.name), path_by_id[category_id]]
        )
    dictionary_sheet.freeze_panes = "A2"
    dictionary_sheet.auto_filter.ref = (
        f"A1:C{max(dictionary_sheet.max_row, 1)}"
    )
    dictionary_sheet.column_dimensions["A"].width = 24
    dictionary_sheet.column_dimensions["B"].width = 28
    dictionary_sheet.column_dimensions["C"].width = 42
    for cell in dictionary_sheet[1]:
        cell.fill = required_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    instructions = [
        ("必填列", "标题、知识来源、业务类型、知识分类、正文。"),
        (
            "知识来源",
            "从“知识来源字典”复制中文名称或代码；"
            "普通导入仅允许总部标准、业务沉淀、"
            "headquarters_standard、business_accumulation；"
            "机型配置信息由飞书专用同步维护。",
        ),
        (
            "业务类型",
            "从“业务类型字典”复制中文名称或代码；"
            "仅允许自营回收、聚合回收、self_operated、aggregated。",
        ),
        ("知识分类", "推荐从“分类字典”复制分类ID；也可填写唯一分类名称或完整分类路径。"),
        ("多值字段", "副标题、场景标签、关联标准项等多项内容使用中文分号“；”分隔。"),
        (
            "兼容格式",
            "支持“知识库主表”的主标题、知识内容、适用范围和生效状态列；"
            "主表也必须提供知识来源（兼容列名“业务来源”）；"
            "存在生效状态列时：生效中直接发布，待审核进入审核，"
            "已禁用按知识键、主题键或记录ID同步废弃原知识。",
        ),
        (
            "正文媒体",
            "仅识别插件标记 [img:https://...] 和 [video:https://...]；"
            "导入后会在原位置显示缩略图或视频卡片，官网、文档及正文原有 URL 不转换。",
        ),
        (
            "疑似重复确认",
            "疑似重复行会自动进入系统“候选价值复核”的“疑似重复确认”筛选，"
            "由审核人核对命中知识后决定是否送审；无需修改 Excel 或重复上传。",
        ),
        (
            "导入结果",
            "来源表的“生效中”行直接发布，“待审核”行进入待审核，"
            "“已禁用”行同步废弃原知识；普通模板成功行进入待审核；"
            "疑似重复、格式错误、分类不存在或完全重复的行会逐行返回原因。",
        ),
        ("单次上限", f"每个文件最多 {MAX_IMPORT_ROWS} 条、文件最大 5MB，仅支持 .xlsx。"),
        (
            "示例",
            "标题：设备无法开机；知识来源：总部标准；业务类型：自营回收；"
            "知识分类：cat-qc-process；正文：先检查电量，再长按电源键。",
        ),
    ]
    instructions_sheet.append(["项目", "说明"])
    for item in instructions:
        instructions_sheet.append(item)
    instructions_sheet.column_dimensions["A"].width = 18
    instructions_sheet.column_dimensions["B"].width = 90
    instructions_sheet.freeze_panes = "A2"
    for cell in instructions_sheet[1]:
        cell.fill = required_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in instructions_sheet.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_knowledge_update_template(categories) -> bytes:
    """生成仅按中台知识ID定位、整行覆盖可编辑字段的修改模板。"""

    workbook = load_workbook(
        BytesIO(build_knowledge_import_template(categories))
    )
    import_sheet = workbook[IMPORT_SHEET_NAME]
    import_sheet.title = KNOWLEDGE_UPDATE_SHEET_NAME
    import_sheet.insert_cols(1)
    import_sheet["A1"] = "知识ID（必填）"
    import_sheet["A1"].fill = PatternFill("solid", fgColor="0F766E")
    import_sheet["A1"].font = Font(color="FFFFFF", bold=True)
    import_sheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    import_sheet["A1"].comment = Comment(
        "必填。仅使用中台知识ID精确匹配，例如 A-00001；"
        "不存在的ID不会新增知识，同一文件内ID不能重复。",
        "知识库",
    )
    import_sheet.freeze_panes = "A2"
    import_sheet.auto_filter.ref = "A1:L1"
    import_sheet.column_dimensions["A"].width = 18
    import_sheet.column_dimensions["B"].width = 32
    import_sheet.column_dimensions["C"].width = 22
    import_sheet.column_dimensions["D"].width = 22
    import_sheet.column_dimensions["E"].width = 28
    import_sheet.column_dimensions["F"].width = 70
    for column in ("G", "H", "I", "J", "K", "L"):
        import_sheet.column_dimensions[column].width = 24

    instructions_sheet = workbook["填写说明"]
    if instructions_sheet.max_row > 1:
        instructions_sheet.delete_rows(2, instructions_sheet.max_row - 1)
    instructions = [
        (
            "唯一匹配规则",
            "系统只按“知识ID”精确匹配现有知识，不按标题、主题键、"
            "记录ID或知识键兜底；ID不存在时该行失败，不会自动新增。",
        ),
        (
            "覆盖规则",
            "标题、知识来源、业务类型、知识分类和正文必须填写；"
            "副标题、场景标签、关联标准项、适用类目、品牌、机型为空时会清空原值。",
        ),
        (
            "保留字段",
            "知识ID、当前状态、录入方式、创建人、创建时间、"
            "质量分、使用统计和系统标签保持不变。",
        ),
        (
            "知识来源",
            "仅支持总部标准和业务沉淀；机型配置信息继续使用"
            "专用 Excel 同步，不能通过本模板修改。",
        ),
        (
            "正文媒体",
            "正文会整体替换，仅识别 [img:https://...] 和 "
            "[video:https://...] 外链媒体标记。含本地上传图片或视频的知识"
            "不支持 Excel 批量修改，请在页面中单条编辑；普通知识导出文件"
            "中的 [图片：说明]、[视频：说明] 不能回填。",
        ),
        (
            "修改结果",
            "实际变化显示“已修改”，内容完全一致显示“未变化”，"
            "校验失败会逐行展示原因；修改直接作用于原知识，不进入新增知识审核流程。",
        ),
        (
            "单次上限",
            f"每个文件最多 {MAX_IMPORT_ROWS} 条、文件最大 5MB，仅支持 .xlsx。",
        ),
    ]
    for item in instructions:
        instructions_sheet.append(item)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_model_configuration_import_template() -> bytes:
    workbook = Workbook()
    import_sheet = workbook.active
    import_sheet.title = MODEL_CONFIGURATION_IMPORT_SHEET_NAME
    instructions_sheet = workbook.create_sheet("填写说明")

    headers = [
        "来源知识ID（选填）",
        "标题（必填）",
        "品类ID（必填）",
        "品类（必填）",
        "品牌ID（必填）",
        "品牌（必填）",
        "型号ID（必填）",
        "型号（必填）",
        "综合内容（必填）",
    ]
    import_sheet.append(headers)
    import_sheet.freeze_panes = "A2"
    import_sheet.auto_filter.ref = "A1:I1"
    import_sheet.row_dimensions[1].height = 28

    required_columns = {2, 3, 4, 5, 6, 7, 8, 9}
    required_fill = PatternFill("solid", fgColor="0F766E")
    optional_fill = PatternFill("solid", fgColor="475569")
    for index, cell in enumerate(import_sheet[1], start=1):
        cell.fill = (
            required_fill if index in required_columns else optional_fill
        )
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    widths = {
        "A": 20,
        "B": 42,
        "C": 14,
        "D": 18,
        "E": 14,
        "F": 18,
        "G": 14,
        "H": 34,
        "I": 72,
    }
    for column, width in widths.items():
        import_sheet.column_dimensions[column].width = width

    import_sheet["A1"].comment = Comment(
        "其他知识库的追溯ID，可留空；不是中台 A-xxxxx 知识ID，"
        "也不作为机型配置信息的唯一标识。",
        "知识库",
    )
    import_sheet["C1"].comment = Comment(
        f"原始“个性化配置信息”工作表没有品类列时，系统默认 "
        f"{MODEL_CONFIGURATION_CATEGORY_SOURCE_ID}。",
        "知识库",
    )
    import_sheet["D1"].comment = Comment(
        f"原始“个性化配置信息”工作表没有品类列时，系统默认"
        f"“{MODEL_CONFIGURATION_CATEGORY_NAME}”。",
        "知识库",
    )
    import_sheet["I1"].comment = Comment(
        "完整机型配置正文，最多 100000 个字符。",
        "知识库",
    )

    instructions = [
        ("导入类型", "上传时请选择“机型配置信息”。"),
        (
            "必填列",
            "标题、品类ID、品类、品牌ID、品牌、型号ID、型号、综合内容。",
        ),
        (
            "来源知识ID",
            "选填。该字段来自其他知识库，仅用于追溯；留空不影响导入，"
            "中台以品类ID、品牌ID、型号ID组合识别同一条机型配置信息。",
        ),
        (
            "原表兼容",
            "兼容工作表名“个性化配置信息”和第一列“知识ID”；"
            "原表同时缺少品类ID、品类时默认使用 119 / 平板电脑；"
            "旧表中的卡槽、Home键等个性属性列会被忽略，只同步综合内容。",
        ),
        (
            "同步规则",
            "整本工作簿先完整校验，再在单个事务中同步；任一字段、"
            "重复键或数据库冲突都会整批回滚。",
        ),
        (
            "幂等规则",
            "同一品类ID、品牌ID、型号ID组合重复上传不会新增重复知识；"
            "内容变化会保留中台知识ID并原地更新。",
        ),
        (
            "发布规则",
            "成功记录直接发布为机型配置信息，不进入普通审核、查重或向量流程。",
        ),
        (
            "单次上限",
            f"每个文件最多 {MAX_MODEL_CONFIGURATION_IMPORT_ROWS} 条、"
            "文件最大 5MB，仅支持 .xlsx。",
        ),
    ]
    instructions_sheet.append(["项目", "说明"])
    for item in instructions:
        instructions_sheet.append(item)
    instructions_sheet.column_dimensions["A"].width = 18
    instructions_sheet.column_dimensions["B"].width = 96
    instructions_sheet.freeze_panes = "A2"
    for cell in instructions_sheet[1]:
        cell.fill = required_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in instructions_sheet.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
