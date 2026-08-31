import asyncio
from io import BytesIO
from types import SimpleNamespace
import unittest

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.models.knowledge import Category, KnowledgeImportTask
from app.routes.knowledge import (
    _validate_excel_before_queue,
    import_knowledge_excel,
)


class ExcelUploadPreflightTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite+pysqlite:///:memory:")
        KnowledgeImportTask.__table__.create(self.engine)
        Category.__table__.create(self.engine)
        self.session_factory = sessionmaker(bind=self.engine)
        with self.session_factory() as db:
            db.add(
                Category(
                    id="cat-qc",
                    name="质检标准",
                    parent_id=None,
                    level=1,
                    sort_order=1,
                )
            )
            db.commit()
        self.user = SimpleNamespace(role="super_admin", username="tester")

    def tearDown(self):
        self.engine.dispose()

    @staticmethod
    def _knowledge_workbook(rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "知识导入"
        sheet.append(["标题", "知识来源", "业务类型", "知识分类", "正文"])
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _model_configuration_workbook(rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "机型配置信息"
        sheet.append(
            [
                "来源知识ID",
                "标题",
                "品类ID",
                "品类",
                "品牌ID",
                "品牌",
                "型号ID",
                "型号",
                "综合内容",
            ]
        )
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_invalid_rows_are_rejected_before_task_creation(self):
        payload = self._knowledge_workbook(
            [
                ["有效知识", "总部标准", "自营回收", "cat-qc", "正文"],
                ["缺少来源", "", "自营回收", "cat-qc", "正文"],
                ["缺少正文", "业务沉淀", "自营回收", "cat-qc", ""],
            ]
        )

        with self.session_factory() as db:
            upload = UploadFile(file=BytesIO(payload), filename="invalid.xlsx")
            with self.assertRaises(HTTPException) as raised:
                asyncio.run(
                    import_knowledge_excel(
                        file=upload,
                        import_type="knowledge",
                        db=db,
                        current_user=self.user,
                    )
                )

            self.assertEqual(raised.exception.status_code, 422)
            detail = raised.exception.detail
            self.assertEqual(detail["code"], "EXCEL_VALIDATION_FAILED")
            self.assertEqual(detail["error_count"], 2)
            self.assertEqual(
                [(item["row"], item["field"]) for item in detail["errors"]],
                [(3, "知识来源"), (4, "正文")],
            )
            self.assertIn("未创建导入任务", detail["message"])
            self.assertEqual(db.query(KnowledgeImportTask).count(), 0)

    def test_valid_workbook_is_queued_after_preflight(self):
        payload = self._knowledge_workbook(
            [["有效知识", "总部标准", "自营回收", "cat-qc", "正文"]]
        )

        with self.session_factory() as db:
            upload = UploadFile(file=BytesIO(payload), filename="valid.xlsx")
            response = asyncio.run(
                import_knowledge_excel(
                    file=upload,
                    import_type="knowledge",
                    db=db,
                    current_user=self.user,
                )
            )

            self.assertEqual(response.status, "queued")
            self.assertEqual(response.total_rows, 1)
            self.assertEqual(db.query(KnowledgeImportTask).count(), 1)

    def test_model_configuration_reports_all_invalid_rows_before_queue(self):
        payload = self._model_configuration_workbook(
            [
                [
                    "source-1",
                    "有效机型",
                    "119",
                    "平板电脑",
                    "10530",
                    "苹果",
                    "97519",
                    "iPad 10",
                    "配置正文",
                ],
                [
                    "source-2",
                    "缺少品牌ID",
                    "119",
                    "平板电脑",
                    "",
                    "苹果",
                    "97520",
                    "iPad 11",
                    "配置正文",
                ],
                [
                    "source-3",
                    "缺少综合内容",
                    "119",
                    "平板电脑",
                    "10530",
                    "苹果",
                    "97521",
                    "iPad 12",
                    "",
                ],
            ]
        )

        with self.session_factory() as db:
            with self.assertRaises(HTTPException) as raised:
                _validate_excel_before_queue(
                    payload,
                    "model_configuration",
                    db,
                )

            detail = raised.exception.detail
            self.assertEqual(detail["code"], "EXCEL_VALIDATION_FAILED")
            self.assertEqual(detail["error_count"], 2)
            self.assertEqual(
                [item["row"] for item in detail["errors"]],
                [3, 4],
            )
            self.assertEqual(
                [item["field"] for item in detail["errors"]],
                ["品牌ID", "综合内容"],
            )
            self.assertEqual(db.query(KnowledgeImportTask).count(), 0)

    def test_workbook_level_error_has_a_visible_reason(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "知识导入"
        # 缺少知识来源、业务类型、知识分类等必填表头。
        sheet.append(["标题", "正文"])
        sheet.append(["标题", "正文"])
        output = BytesIO()
        workbook.save(output)

        with self.session_factory() as db:
            with self.assertRaises(HTTPException) as raised:
                _validate_excel_before_queue(
                    output.getvalue(),
                    "knowledge",
                    db,
                )

        detail = raised.exception.detail
        self.assertEqual(detail["code"], "EXCEL_VALIDATION_FAILED")
        self.assertGreaterEqual(detail["error_count"], 1)
        self.assertTrue(detail["errors"][0]["message"])
        self.assertEqual(detail["errors"][0]["row"], 1)


if __name__ == "__main__":
    unittest.main()
