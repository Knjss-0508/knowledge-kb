import hashlib
import unittest
from datetime import datetime, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.models.knowledge import (
    Category,
    Knowledge,
    KnowledgeChangeLog,
    KnowledgeImportTask,
    KnowledgeMedia,
    KnowledgeStatus,
    MediaDeletionTask,
)
from app.models.user import User
from app.routes.knowledge import process_knowledge_import_task
from app.services.knowledge_excel import (
    KnowledgeExcelError,
    build_knowledge_update_template,
    parse_knowledge_workbook,
    parse_knowledge_update_workbook,
)


UPDATE_HEADERS = [
    "知识ID（必填）",
    "知识来源（必填）",
    "业务类型（必填）",
    "标题（必填）",
    "知识分类（必填）",
    "正文（必填）",
    "副标题（选填）",
    "场景标签（选填）",
    "关联标准项（选填）",
    "适用类目（选填）",
    "适用品牌（选填）",
    "适用机型（选填）",
]


def _workbook_bytes(rows, *, headers=None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "知识批量修改"
    sheet.append(headers or UPDATE_HEADERS)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _content_text(content) -> str:
    if isinstance(content, str):
        return content
    if isinstance(content, dict):
        return "\n".join(
            str(block.get("value") or "")
            for block in content.get("blocks", [])
            if isinstance(block, dict) and block.get("type") == "text"
        )
    return str(content or "")


class KnowledgeUpdateWorkbookTests(unittest.TestCase):
    def setUp(self):
        self.categories = [
            SimpleNamespace(
                id="cat-old",
                name="旧分类",
                parent_id=None,
                level=1,
                sort_order=1,
            ),
            SimpleNamespace(
                id="cat-new",
                name="新分类",
                parent_id=None,
                level=1,
                sort_order=2,
            ),
        ]

    def test_template_makes_id_only_update_contract_explicit(self):
        payload = build_knowledge_update_template(self.categories)
        workbook = load_workbook(BytesIO(payload), read_only=True)

        self.assertIn("知识批量修改", workbook.sheetnames)
        self.assertIn("填写说明", workbook.sheetnames)
        headers = list(
            next(
                workbook["知识批量修改"].iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                )
            )
        )
        self.assertEqual(headers[0], "知识ID（必填）")
        for required_header in UPDATE_HEADERS[1:6]:
            self.assertIn(required_header, headers)

        instructions = "\n".join(
            str(value or "")
            for row in workbook["填写说明"].iter_rows(values_only=True)
            for value in row
        )
        self.assertIn("只按", instructions)
        self.assertIn("知识ID", instructions)
        self.assertIn("不会自动新增", instructions)
        self.assertIn("机型配置信息", instructions)

    def test_parser_exposes_knowledge_id_and_rejects_a_missing_id(self):
        payload = _workbook_bytes(
            [
                [
                    "A-00001",
                    "总部标准",
                    "自营回收",
                    "有ID",
                    "cat-new",
                    "新正文",
                ],
                [
                    "",
                    "业务沉淀",
                    "自营回收",
                    "无ID",
                    "cat-new",
                    "不会被更新",
                ],
            ]
        )

        rows = parse_knowledge_update_workbook(payload, self.categories)

        self.assertEqual(len(rows), 2)
        self.assertTrue(rows[0].is_valid)
        self.assertEqual(rows[0].knowledge_id, "A-00001")
        self.assertFalse(rows[1].is_valid)
        self.assertEqual(rows[1].error_code, "KNOWLEDGE_ID_REQUIRED")

    def test_parser_rejects_duplicate_ids_as_a_workbook_error(self):
        payload = _workbook_bytes(
            [
                [
                    "A-00001",
                    "总部标准",
                    "自营回收",
                    "第一次修改",
                    "cat-new",
                    "正文一",
                ],
                [
                    "A-00001",
                    "业务沉淀",
                    "聚合回收",
                    "第二次修改",
                    "cat-old",
                    "正文二",
                ],
            ]
        )

        with self.assertRaises(KnowledgeExcelError) as caught:
            parse_knowledge_update_workbook(payload, self.categories)

        self.assertIn("重复", str(caught.exception))
        self.assertIn("A-00001", str(caught.exception))

    def test_parser_rejects_multiple_compatible_knowledge_id_columns(self):
        headers = ["知识ID", "知识库ID", *UPDATE_HEADERS[1:]]
        payload = _workbook_bytes(
            [
                [
                    "A-00001",
                    "A-00002",
                    "总部标准",
                    "自营回收",
                    "冲突ID列",
                    "cat-new",
                    "正文",
                ],
            ],
            headers=headers,
        )

        with self.assertRaises(KnowledgeExcelError) as caught:
            parse_knowledge_update_workbook(payload, self.categories)

        self.assertIn("只能保留一个知识ID列", str(caught.exception))

    def test_parser_rejects_multiple_alias_columns_for_the_same_field(self):
        headers = [
            "知识ID",
            "知识来源",
            "业务类型",
            "标题",
            "主标题",
            "知识分类",
            "正文",
        ]
        payload = _workbook_bytes(
            [
                [
                    "A-00001",
                    "总部标准",
                    "自营回收",
                    "标题一",
                    "标题二",
                    "cat-new",
                    "正文",
                ],
            ],
            headers=headers,
        )

        with self.assertRaises(KnowledgeExcelError) as caught:
            parse_knowledge_update_workbook(payload, self.categories)

        self.assertIn("同一字段只能保留一个兼容列", str(caught.exception))

    def test_parser_rejects_regular_export_and_local_media_placeholders(self):
        export_payload = _workbook_bytes(
            [
                [
                    "A-00001",
                    "总部标准",
                    "自营回收",
                    "标题",
                    "cat-new",
                    "正文",
                ],
            ]
        )
        export_workbook = load_workbook(BytesIO(export_payload))
        export_workbook.active.title = "知识库主表"
        output = BytesIO()
        export_workbook.save(output)

        with self.assertRaises(KnowledgeExcelError) as caught:
            parse_knowledge_update_workbook(
                output.getvalue(),
                self.categories,
            )
        self.assertIn("普通导出文件不能直接用于批量修改", str(caught.exception))

        placeholder_payload = _workbook_bytes(
            [
                [
                    "A-00001",
                    "总部标准",
                    "自营回收",
                    "标题",
                    "cat-new",
                    "正文\n[图片：旧图片]",
                ],
            ]
        )
        rows = parse_knowledge_update_workbook(
            placeholder_payload,
            self.categories,
        )
        self.assertFalse(rows[0].is_valid)
        self.assertEqual(
            rows[0].error_code,
            "LOCAL_MEDIA_PLACEHOLDER_UNSUPPORTED",
        )

    def test_regular_import_rejects_a_non_empty_knowledge_id_column(self):
        payload = _workbook_bytes(
            [
                [
                    "A-00001",
                    "总部标准",
                    "自营回收",
                    "不应新增",
                    "cat-new",
                    "修改文件不能走新增模式",
                ],
            ]
        )

        with self.assertRaises(KnowledgeExcelError) as caught:
            parse_knowledge_workbook(payload, self.categories)

        self.assertIn("批量修改", str(caught.exception))
        self.assertIn("不会按知识ID执行普通新增导入", str(caught.exception))


class KnowledgeUpdateImportTaskTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite+pysqlite:///:memory:")
        for table in (
            Category.__table__,
            Knowledge.__table__,
            KnowledgeChangeLog.__table__,
            KnowledgeMedia.__table__,
            MediaDeletionTask.__table__,
            KnowledgeImportTask.__table__,
            User.__table__,
        ):
            table.create(self.engine)
        self.session_factory = sessionmaker(bind=self.engine)
        with self.session_factory() as db:
            db.add_all(
                [
                    Category(
                        id="cat-old",
                        name="旧分类",
                        parent_id=None,
                        level=1,
                        sort_order=1,
                    ),
                    Category(
                        id="cat-new",
                        name="新分类",
                        parent_id=None,
                        level=1,
                        sort_order=2,
                    ),
                    User(
                        id="user-batch-editor",
                        username="batch-editor",
                        password_hash="test-only",
                        role="super_support",
                        is_active=True,
                    ),
                ]
            )
            db.commit()

    def tearDown(self):
        self.engine.dispose()

    def _add_knowledge(
        self,
        knowledge_id: str,
        *,
        title: str = "旧标题",
        knowledge_origin: str = "business_accumulation",
        status: KnowledgeStatus = KnowledgeStatus.PUBLISHED,
        created_by: str = "original-author",
        created_at: datetime | None = None,
    ) -> datetime:
        created_at = created_at or datetime(2026, 1, 2, 3, 4, 5)
        with self.session_factory() as db:
            db.add(
                Knowledge(
                    id=knowledge_id,
                    knowledge_origin=knowledge_origin,
                    business_type="self_operated",
                    title=title,
                    subtitles=["旧副标题"],
                    content={
                        "blocks": [
                            {"type": "text", "value": "旧正文"},
                        ]
                    },
                    category_id="cat-old",
                    status=status,
                    source="manual",
                    source_session_id="session-original",
                    source_topic_key="topic-original",
                    source_record_id="record-original",
                    source_knowledge_key="knowledge-key-original",
                    quality_score=0.75,
                    applicable_scenes=["旧场景"],
                    applicable_categories=["旧类目"],
                    applicable_brands=["旧品牌"],
                    applicable_models=["旧机型"],
                    related_standard_items=["旧标准"],
                    source_fields={"追溯字段": "原值"},
                    deduplication_metadata={"keep": True},
                    created_by=created_by,
                    updated_by="previous-editor",
                    created_at=created_at,
                    updated_at=datetime(2026, 2, 3, 4, 5, 6),
                )
            )
            db.commit()
        return created_at

    def _add_task(
        self,
        task_id: str,
        rows,
    ) -> None:
        content = _workbook_bytes(rows)
        with self.session_factory() as db:
            db.add(
                KnowledgeImportTask(
                    id=task_id,
                    import_type="knowledge_update",
                    created_by="batch-editor",
                    original_filename="knowledge-update.xlsx",
                    file_size=len(content),
                    file_sha256=hashlib.sha256(content).hexdigest(),
                    file_content=content,
                    status="running",
                    attempt_count=1,
                    next_attempt_at=datetime.utcnow() - timedelta(seconds=1),
                    lease_expires_at=datetime.utcnow() + timedelta(minutes=5),
                )
            )
            db.commit()

    @staticmethod
    def _updated_row(knowledge_id: str, *, title: str = "新标题"):
        return [
            knowledge_id,
            "总部标准",
            "聚合回收",
            title,
            "cat-new",
            "新正文",
            "新副标题一；新副标题二",
            "",
            "新标准一；新标准二",
            "",
            "",
            "",
        ]

    def test_task_updates_only_the_exact_id_and_rebuilds_search_vectors(self):
        created_at = self._add_knowledge("A-00001")
        self._add_knowledge(
            "A-00002",
            title="新标题",
            created_by="other-author",
        )
        self._add_task(
            "import-update-by-id",
            [self._updated_row("A-00001")],
        )

        with (
            patch(
                "app.routes.knowledge.ensure_embedding"
            ) as ensure_embedding,
            patch(
                "app.routes.knowledge.ensure_search_embeddings"
            ) as ensure_search_embeddings,
            patch(
                "app.routes.knowledge._create_knowledge_item"
            ) as create_knowledge,
            patch(
                "app.routes.knowledge._generate_knowledge_id"
            ) as generate_knowledge_id,
        ):
            process_knowledge_import_task(
                "import-update-by-id",
                session_factory=self.session_factory,
            )

        create_knowledge.assert_not_called()
        generate_knowledge_id.assert_not_called()
        ensure_embedding.assert_called_once()
        ensure_search_embeddings.assert_called_once()

        with self.session_factory() as db:
            self.assertEqual(db.query(Knowledge).count(), 2)
            updated = db.get(Knowledge, "A-00001")
            untouched = db.get(Knowledge, "A-00002")
            task = db.get(KnowledgeImportTask, "import-update-by-id")

            self.assertEqual(updated.id, "A-00001")
            self.assertEqual(updated.title, "新标题")
            self.assertEqual(_content_text(updated.content), "新正文")
            self.assertEqual(
                updated.subtitles,
                ["新副标题一", "新副标题二"],
            )
            self.assertEqual(updated.knowledge_origin, "headquarters_standard")
            self.assertEqual(updated.business_type, "aggregated")
            self.assertEqual(updated.category_id, "cat-new")
            self.assertEqual(updated.applicable_scenes, [])
            self.assertEqual(updated.applicable_categories, [])
            self.assertEqual(updated.applicable_brands, [])
            self.assertEqual(updated.applicable_models, [])
            self.assertEqual(
                updated.related_standard_items,
                ["新标准一", "新标准二"],
            )
            self.assertEqual(updated.status, KnowledgeStatus.PUBLISHED)
            self.assertEqual(updated.created_by, "original-author")
            self.assertEqual(updated.created_at, created_at)
            self.assertEqual(updated.source, "manual")
            self.assertEqual(updated.source_session_id, "session-original")
            self.assertEqual(updated.source_topic_key, "topic-original")
            self.assertEqual(updated.source_record_id, "record-original")
            self.assertEqual(
                updated.source_knowledge_key,
                "knowledge-key-original",
            )
            self.assertEqual(updated.quality_score, 0.75)
            self.assertEqual(
                updated.deduplication_metadata,
                {"keep": True},
            )
            self.assertEqual(updated.updated_by, "batch-editor")

            self.assertEqual(untouched.title, "新标题")
            self.assertEqual(_content_text(untouched.content), "旧正文")
            self.assertEqual(untouched.created_by, "other-author")

            self.assertEqual(task.status, "completed")
            self.assertEqual(task.total_rows, 1)
            self.assertEqual(task.processed_rows, 1)
            self.assertEqual(task.imported, 1)
            self.assertEqual(task.updated, 1)
            self.assertEqual(task.unchanged, 0)
            self.assertEqual(task.failed, 0)
            self.assertEqual(task.results[0]["knowledge_id"], "A-00001")
            self.assertEqual(task.results[0]["operation"], "updated")

            logs = (
                db.query(KnowledgeChangeLog)
                .filter(KnowledgeChangeLog.knowledge_id == "A-00001")
                .all()
            )
            self.assertEqual(len(logs), 1)
            self.assertEqual(logs[0].changed_by, "batch-editor")
            self.assertTrue(
                {
                    "title",
                    "subtitles",
                    "content",
                    "knowledge_origin",
                    "business_type",
                    "category_id",
                    "applicable_scenes",
                    "applicable_categories",
                    "applicable_brands",
                    "applicable_models",
                    "related_standard_items",
                }.issubset(set(logs[0].changed_fields))
            )
            self.assertEqual(logs[0].before_data["title"], "旧标题")
            self.assertEqual(logs[0].after_data["title"], "新标题")

    def test_missing_target_reports_failure_and_never_creates_knowledge(self):
        self._add_task(
            "import-update-missing",
            [self._updated_row("A-40404")],
        )

        with (
            patch(
                "app.routes.knowledge._create_knowledge_item"
            ) as create_knowledge,
            patch(
                "app.routes.knowledge._generate_knowledge_id"
            ) as generate_knowledge_id,
            patch(
                "app.routes.knowledge.ensure_embedding"
            ) as ensure_embedding,
            patch(
                "app.routes.knowledge.ensure_search_embeddings"
            ) as ensure_search_embeddings,
        ):
            process_knowledge_import_task(
                "import-update-missing",
                session_factory=self.session_factory,
            )

        create_knowledge.assert_not_called()
        generate_knowledge_id.assert_not_called()
        ensure_embedding.assert_not_called()
        ensure_search_embeddings.assert_not_called()
        with self.session_factory() as db:
            self.assertEqual(db.query(Knowledge).count(), 0)
            task = db.get(KnowledgeImportTask, "import-update-missing")
            self.assertEqual(task.status, "completed_with_errors")
            self.assertEqual(task.processed_rows, 1)
            self.assertEqual(task.imported, 0)
            self.assertEqual(task.updated, 0)
            self.assertEqual(task.unchanged, 0)
            self.assertEqual(task.failed, 1)
            self.assertEqual(
                task.results[0]["error_code"],
                "KNOWLEDGE_ID_NOT_FOUND",
            )
            self.assertEqual(task.results[0]["knowledge_id"], "A-40404")

    def test_identical_row_is_unchanged_without_log_or_vector_rebuild(self):
        created_at = self._add_knowledge(
            "A-00003",
            title="原样标题",
            knowledge_origin="headquarters_standard",
            status=KnowledgeStatus.REVIEW,
        )
        with self.session_factory() as db:
            item = db.get(Knowledge, "A-00003")
            item.business_type = "self_operated"
            item.title = "原样标题"
            item.subtitles = ["旧副标题"]
            item.content = {
                "blocks": [{"type": "text", "value": "旧正文"}]
            }
            item.category_id = "cat-old"
            item.applicable_scenes = []
            item.applicable_categories = []
            item.applicable_brands = []
            item.applicable_models = []
            item.related_standard_items = ["旧标准"]
            db.commit()
            db.refresh(item)
            original_updated_at = item.updated_at

        same_row = [
            "A-00003",
            "总部标准",
            "自营回收",
            "原样标题",
            "cat-old",
            "旧正文",
            "旧副标题",
            "",
            "旧标准",
            "",
            "",
            "",
        ]
        self._add_task("import-update-unchanged", [same_row])

        with (
            patch(
                "app.routes.knowledge.ensure_embedding"
            ) as ensure_embedding,
            patch(
                "app.routes.knowledge.ensure_search_embeddings"
            ) as ensure_search_embeddings,
        ):
            process_knowledge_import_task(
                "import-update-unchanged",
                session_factory=self.session_factory,
            )

        ensure_embedding.assert_not_called()
        ensure_search_embeddings.assert_not_called()
        with self.session_factory() as db:
            item = db.get(Knowledge, "A-00003")
            task = db.get(KnowledgeImportTask, "import-update-unchanged")
            self.assertEqual(item.status, KnowledgeStatus.REVIEW)
            self.assertEqual(item.created_at, created_at)
            self.assertEqual(item.updated_at, original_updated_at)
            self.assertEqual(
                db.query(KnowledgeChangeLog)
                .filter(KnowledgeChangeLog.knowledge_id == "A-00003")
                .count(),
                0,
            )
            self.assertEqual(task.status, "completed")
            self.assertEqual(task.updated, 0)
            self.assertEqual(task.unchanged, 1)
            self.assertEqual(task.results[0]["operation"], "unchanged")

    def test_model_configuration_target_is_rejected_without_mutation(self):
        self._add_knowledge(
            "A-00999",
            title="受管机型配置",
            knowledge_origin="model_configuration",
            created_by="model-sync",
        )
        self._add_task(
            "import-update-model-configuration",
            [self._updated_row("A-00999", title="不应覆盖")],
        )

        with (
            patch(
                "app.routes.knowledge.ensure_embedding"
            ) as ensure_embedding,
            patch(
                "app.routes.knowledge.ensure_search_embeddings"
            ) as ensure_search_embeddings,
        ):
            process_knowledge_import_task(
                "import-update-model-configuration",
                session_factory=self.session_factory,
            )

        ensure_embedding.assert_not_called()
        ensure_search_embeddings.assert_not_called()
        with self.session_factory() as db:
            item = db.get(Knowledge, "A-00999")
            task = db.get(
                KnowledgeImportTask,
                "import-update-model-configuration",
            )
            self.assertEqual(item.title, "受管机型配置")
            self.assertEqual(item.knowledge_origin, "model_configuration")
            self.assertEqual(item.created_by, "model-sync")
            self.assertEqual(
                db.query(KnowledgeChangeLog)
                .filter(KnowledgeChangeLog.knowledge_id == "A-00999")
                .count(),
                0,
            )
            self.assertEqual(task.status, "completed_with_errors")
            self.assertEqual(task.updated, 0)
            self.assertEqual(task.failed, 1)
            self.assertEqual(
                task.results[0]["error_code"],
                "KNOWLEDGE_ORIGIN_MANAGED",
            )
            self.assertEqual(task.results[0]["knowledge_id"], "A-00999")

    def test_local_media_target_is_rejected_without_deletion(self):
        self._add_knowledge("A-00004")
        with self.session_factory() as db:
            item = db.get(Knowledge, "A-00004")
            item.content = {
                "blocks": [
                    {
                        "type": "image",
                        "media_id": "old-image.png",
                        "alt": "旧图片",
                    }
                ]
            }
            db.add(
                KnowledgeMedia(
                    id="media-old-image",
                    knowledge_id=item.id,
                    media_type="image",
                    filename="old-image.png",
                    original_name="old-image.png",
                    file_path="knowledge/A-00004/old-image.png",
                    file_size=128,
                    mime_type="image/png",
                )
            )
            db.commit()
        self._add_task(
            "import-update-media",
            [self._updated_row("A-00004")],
        )

        with (
            patch(
                "app.routes.knowledge.ensure_embedding"
            ) as ensure_embedding,
            patch(
                "app.routes.knowledge.ensure_search_embeddings"
            ) as ensure_search_embeddings,
        ):
            process_knowledge_import_task(
                "import-update-media",
                session_factory=self.session_factory,
            )

        ensure_embedding.assert_not_called()
        ensure_search_embeddings.assert_not_called()
        with self.session_factory() as db:
            task = db.get(KnowledgeImportTask, "import-update-media")
            self.assertEqual(task.status, "completed_with_errors")
            self.assertEqual(task.updated, 0)
            self.assertEqual(task.failed, 1)
            self.assertEqual(
                task.results[0]["error_code"],
                "LOCAL_MEDIA_BATCH_UPDATE_UNSUPPORTED",
            )
            self.assertEqual(
                db.query(KnowledgeMedia)
                .filter(KnowledgeMedia.knowledge_id == "A-00004")
                .count(),
                1,
            )
            self.assertEqual(db.query(MediaDeletionTask).count(), 0)
            item = db.get(Knowledge, "A-00004")
            self.assertEqual(item.title, "旧标题")
            self.assertEqual(item.content["blocks"][0]["media_id"], "old-image.png")


if __name__ == "__main__":
    unittest.main()
