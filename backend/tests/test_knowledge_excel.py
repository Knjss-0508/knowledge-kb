import unittest
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.services.knowledge_excel import (
    KnowledgeExcelError,
    build_model_configuration_import_template,
    build_knowledge_export_workbook,
    build_knowledge_import_template,
    parse_model_configuration_workbook,
    parse_knowledge_workbook,
)


class KnowledgeExcelTests(unittest.TestCase):
    def setUp(self):
        self.categories = [
            SimpleNamespace(
                id="cat-parent",
                name="质检",
                parent_id=None,
                level=1,
                sort_order=10,
            ),
            SimpleNamespace(
                id="cat-process",
                name="操作流程",
                parent_id="cat-parent",
                level=2,
                sort_order=20,
            ),
            SimpleNamespace(
                id="cat-qc-standard",
                name="质检标准",
                parent_id=None,
                level=1,
                sort_order=30,
            ),
        ]

    @staticmethod
    def workbook_bytes(
        headers,
        rows,
        *,
        include_knowledge_origin=True,
        knowledge_origin="总部标准",
        include_business_type=True,
        business_type="自营回收",
    ):
        headers = list(headers)
        rows = [list(row) for row in rows]
        prefix_headers = []
        prefix_values = []
        has_knowledge_origin = any(
            str(header).startswith("知识来源") or header == "业务来源"
            for header in headers
        )
        has_business_type = any(
            str(header).startswith("业务类型") or header == "所属业务类型"
            for header in headers
        )
        if include_knowledge_origin and not has_knowledge_origin:
            prefix_headers.append("知识来源")
            prefix_values.append(knowledge_origin)
        if include_business_type and not has_business_type:
            prefix_headers.append("业务类型")
            prefix_values.append(business_type)
        headers = [*prefix_headers, *headers]
        rows = [[*prefix_values, *row] for row in rows]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "知识导入"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def model_configuration_workbook_bytes(
        headers,
        rows,
        *,
        sheet_name="机型配置信息",
        generic_rows=None,
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        if generic_rows is not None:
            generic_sheet = workbook.create_sheet("知识导入")
            generic_sheet.append(
                ["知识来源", "业务类型", "标题", "知识分类", "正文"]
            )
            for row in generic_rows:
                generic_sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_template_contains_category_dictionary_and_no_layer_column(self):
        payload = build_knowledge_import_template(self.categories)
        workbook = load_workbook(BytesIO(payload), read_only=True)

        self.assertEqual(
            workbook.sheetnames,
            ["知识导入", "知识来源字典", "业务类型字典", "分类字典", "填写说明"],
        )
        headers = [
            cell.value
            for cell in next(workbook["知识导入"].iter_rows(max_row=1))
        ]
        self.assertIn("标题（必填）", headers)
        self.assertIn("知识来源（必填）", headers)
        self.assertIn("业务类型（必填）", headers)
        self.assertIn("知识分类（必填）", headers)
        self.assertIn("正文（必填）", headers)
        self.assertIn("副标题（选填）", headers)
        self.assertIn("场景标签（选填）", headers)
        self.assertIn("适用类目（选填）", headers)
        self.assertIn("适用品牌（选填）", headers)
        self.assertIn("适用机型（选填）", headers)
        self.assertIn("关联标准项（选填）", headers)
        self.assertNotIn("知识层级", headers)
        self.assertNotIn("适用业务", headers)
        self.assertNotIn("机型个性化", headers)
        self.assertNotIn("确认疑似重复（选填）", headers)
        self.assertLess(
            headers.index("知识来源（必填）"),
            headers.index("业务类型（必填）"),
        )
        knowledge_origin_rows = list(
            workbook["知识来源字典"].iter_rows(min_row=2, values_only=True)
        )
        self.assertEqual(
            knowledge_origin_rows,
            [
                ("headquarters_standard", "总部标准"),
                ("business_accumulation", "业务沉淀"),
            ],
        )
        dictionary_rows = list(
            workbook["分类字典"].iter_rows(min_row=2, values_only=True)
        )
        self.assertIn(("cat-process", "操作流程", "质检/操作流程"), dictionary_rows)
        business_type_rows = list(
            workbook["业务类型字典"].iter_rows(min_row=2, values_only=True)
        )
        self.assertEqual(
            business_type_rows,
            [
                ("self_operated", "自营回收"),
                ("aggregated", "聚合回收"),
            ],
        )

    def test_model_configuration_template_has_dedicated_contract(self):
        payload = build_model_configuration_import_template()
        workbook = load_workbook(BytesIO(payload), read_only=True)

        self.assertEqual(
            workbook.sheetnames,
            ["机型配置信息", "填写说明"],
        )
        headers = [
            cell.value
            for cell in next(
                workbook["机型配置信息"].iter_rows(max_row=1)
            )
        ]
        self.assertEqual(
            headers,
            [
                "来源知识ID（必填）",
                "标题（必填）",
                "品类ID（必填）",
                "品类（必填）",
                "品牌ID（必填）",
                "品牌（必填）",
                "型号ID（必填）",
                "型号（必填）",
                "是否有卡槽（选填）",
                "Home键（选填）",
                "指纹识别（选填）",
                "3D面容（选填）",
                "内置手写笔（选填）",
                "闪光灯（选填）",
                "蜂窝网络（选填）",
                "光线传感器（选填）",
                "综合内容（必填）",
            ],
        )

    def test_parse_model_configuration_template_preserves_source_fields(self):
        payload = self.model_configuration_workbook_bytes(
            [
                "来源知识ID（必填）",
                "标题（必填）",
                "品类ID（必填）",
                "品类（必填）",
                "品牌ID（必填）",
                "品牌（必填）",
                "型号ID（必填）",
                "型号（必填）",
                "是否有卡槽（选填）",
                "Home键（选填）",
                "综合内容（必填）",
            ],
            [
                [
                    "56383",
                    "iPad 10 配置信息",
                    "119",
                    "平板电脑",
                    "10530",
                    "苹果",
                    "97519",
                    "iPad 10 (2022) 10.9英寸",
                    "蜂窝版有卡槽",
                    "不支持",
                    "Home键：不支持；",
                ]
            ],
        )

        parsed = parse_model_configuration_workbook(payload)

        self.assertEqual(parsed.row_numbers, [2])
        self.assertEqual(len(parsed.records), 1)
        record = parsed.records[0]
        self.assertEqual(record.source_record_id, "56383")
        self.assertEqual(record.category_id, "119")
        self.assertEqual(record.brand_id, "10530")
        self.assertEqual(record.model_id, "97519")
        self.assertEqual(record.source_fields["是否有卡槽"], "蜂窝版有卡槽")
        self.assertEqual(record.source_fields["来源工作表"], "机型配置信息")
        self.assertEqual(record.source_fields["来源行号"], "2")

    def test_parse_legacy_model_configuration_defaults_category_and_ignores_unheaded_cells(self):
        payload = self.model_configuration_workbook_bytes(
            [
                "知识ID",
                "标题",
                "品牌ID",
                "品牌",
                "型号ID",
                "型号",
                "综合内容",
                "是否更新",
                None,
            ],
            [
                [
                    "56362",
                    "Acer A510 配置信息",
                    "11427",
                    "宏碁",
                    "252285",
                    "Acer A510 10.4英寸",
                    "指纹识别：不支持；",
                    "",
                    "有表头数据行上的残留",
                ],
                [None, None, None, None, None, None, None, None, "仅无表头残留"],
                [None] * 9,
            ],
            sheet_name="个性化配置信息",
        )

        parsed = parse_model_configuration_workbook(payload)

        self.assertEqual(len(parsed.records), 1)
        record = parsed.records[0]
        self.assertEqual(record.category_id, "119")
        self.assertEqual(record.category_name, "平板电脑")
        self.assertNotIn("", record.source_fields)
        self.assertNotIn("有表头数据行上的残留", record.source_fields.values())

    def test_model_configuration_rejects_mixed_or_duplicate_rows_before_sync(self):
        mixed_payload = self.model_configuration_workbook_bytes(
            [
                "知识ID",
                "标题",
                "品牌ID",
                "品牌",
                "型号ID",
                "型号",
                "综合内容",
            ],
            [["1", "配置", "2", "品牌", "3", "型号", "正文"]],
            generic_rows=[
                ["业务沉淀", "自营回收", "普通", "cat-process", "正文"]
            ],
        )
        with self.assertRaisesRegex(
            KnowledgeExcelError,
            "同时包含普通知识和机型配置信息",
        ):
            parse_model_configuration_workbook(mixed_payload)

        duplicate_payload = self.model_configuration_workbook_bytes(
            [
                "知识ID",
                "标题",
                "品牌ID",
                "品牌",
                "型号ID",
                "型号",
                "综合内容",
            ],
            [
                ["1", "配置一", "2", "品牌", "3", "型号一", "正文一"],
                ["1", "配置二", "2", "品牌", "4", "型号二", "正文二"],
            ],
        )
        with self.assertRaisesRegex(
            KnowledgeExcelError,
            "第 3 行.*第 2 行重复",
        ):
            parse_model_configuration_workbook(duplicate_payload)

    def test_regular_import_rejects_workbook_with_model_configuration_data(self):
        payload = self.model_configuration_workbook_bytes(
            [
                "知识ID",
                "标题",
                "品牌ID",
                "品牌",
                "型号ID",
                "型号",
                "综合内容",
            ],
            [["1", "配置", "2", "品牌", "3", "型号", "正文"]],
            generic_rows=[
                ["业务沉淀", "自营回收", "普通", "cat-process", "正文"]
            ],
        )

        with self.assertRaisesRegex(
            KnowledgeExcelError,
            "同时包含普通知识和机型配置信息",
        ):
            parse_knowledge_workbook(payload, self.categories)

    def test_model_configuration_has_independent_row_limit(self):
        payload = self.model_configuration_workbook_bytes(
            [
                "知识ID",
                "标题",
                "品牌ID",
                "品牌",
                "型号ID",
                "型号",
                "综合内容",
            ],
            [
                ["1", "配置一", "2", "品牌", "3", "型号一", "正文一"],
                ["2", "配置二", "2", "品牌", "4", "型号二", "正文二"],
            ],
        )
        with patch(
            "app.services.knowledge_excel.MAX_MODEL_CONFIGURATION_IMPORT_ROWS",
            1,
        ), self.assertRaisesRegex(
            KnowledgeExcelError,
            "单次最多导入 1 条",
        ):
            parse_model_configuration_workbook(payload)

    def test_export_matches_knowledge_main_sheet_format_and_maps_supported_fields(self):
        item = SimpleNamespace(
            id="K-00001",
            knowledge_origin="headquarters_standard",
            business_type="self_operated",
            title="平板外观检查",
            subtitles=["平板外观如何检查", "屏幕划痕如何判定"],
            content={
                "blocks": [
                    {"type": "text", "value": "按标准检查外观。"},
                    {
                        "type": "image",
                        "external_url": "https://cdn.example.com/front.png",
                    },
                    {
                        "type": "video",
                        "external_url": "https://cdn.example.com/check.mp4",
                    },
                    {"type": "image", "caption": "本地拍摄图片"},
                ]
            },
            category_id="cat-qc-standard",
            category=SimpleNamespace(name="质检标准"),
            source="excel",
            related_standard_items=["屏幕检测", "外观检测"],
            applicable_scenes=["验机"],
            applicable_categories=["平板"],
            applicable_brands=["Apple"],
            applicable_models=["iPad Pro"],
            status="published",
            source_topic_key="基本情况-外观::平板",
            source_record_id="rec-001",
            source_knowledge_key="基本情况-外观::通用",
            source_fields={
                "来源版本": "2026.07",
                "变更类型": "新增",
                "创建类型": "标准同步",
                "失效类型": "不适用",
                "失效原因": "-",
                "来源追溯": "平板质检标准主表",
                "校验备注": "已完成字段校验",
            },
        )

        payload = build_knowledge_export_workbook([item])
        workbook = load_workbook(BytesIO(payload), data_only=True)
        sheet = workbook["知识库主表"]

        self.assertEqual(sheet.max_column, 21)
        self.assertEqual(sheet.title, "知识库主表")
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            [
                "知识ID", "主题键", "记录ID", "知识键", "主标题", "副标题", "知识内容",
                "知识来源", "业务类型", "知识分类", "录入方式", "关联标准项",
                "适用范围", "生效状态", "来源版本", "变更类型", "创建类型",
                "失效类型", "失效原因", "来源追溯", "校验备注",
            ],
        )
        self.assertEqual(sheet["A2"].value, "K-00001")
        self.assertEqual(sheet["B2"].value, "基本情况-外观::平板")
        self.assertEqual(sheet["C2"].value, "rec-001")
        self.assertEqual(sheet["D2"].value, "基本情况-外观::通用")
        self.assertEqual(sheet["E2"].value, "平板外观检查")
        self.assertEqual(sheet["F2"].value, "平板外观如何检查\n屏幕划痕如何判定")
        self.assertEqual(
            sheet["G2"].value,
            "按标准检查外观。\n[img:https://cdn.example.com/front.png]\n"
            "[video:https://cdn.example.com/check.mp4]\n[图片：本地拍摄图片]",
        )
        self.assertEqual(sheet["H2"].value, "总部标准")
        self.assertEqual(sheet["I2"].value, "自营回收")
        self.assertEqual(sheet["J2"].value, "质检标准")
        self.assertEqual(sheet["K2"].value, "Excel 批量导入")
        self.assertEqual(sheet["L2"].value, "屏幕检测；外观检测")
        self.assertEqual(
            sheet["M2"].value,
            "场景：验机；适用类目：平板；适用品牌：Apple；适用机型：iPad Pro",
        )
        self.assertEqual(sheet["N2"].value, "生效中")
        self.assertEqual(sheet["O2"].value, "2026.07")
        self.assertEqual(sheet["P2"].value, "新增")
        self.assertEqual(sheet["Q2"].value, "标准同步")
        self.assertEqual(sheet["R2"].value, "不适用")
        self.assertEqual(sheet["S2"].value, "-")
        self.assertEqual(sheet["T2"].value, "平板质检标准主表")
        self.assertEqual(sheet["U2"].value, "已完成字段校验")
        self.assertEqual(sheet["A1"].font.name, "宋体")
        self.assertTrue(sheet["A1"].font.bold)
        self.assertTrue(sheet["A1"].alignment.wrap_text)
        self.assertTrue(sheet["A2"].alignment.wrap_text)
        self.assertTrue(sheet["A1"].fill.fgColor.rgb.endswith("D9E8FB"))
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, "A1:U2")

    def test_parse_accepts_category_id_and_splits_multi_value_fields(self):
        payload = self.workbook_bytes(
            ["标题", "知识分类", "正文", "副标题", "场景标签"],
            [
                [
                    "设备无法开机",
                    "cat-process",
                    "先检查电量，再执行强制重启。",
                    "黑屏怎么办；无法启动",
                    "无法开机；售后咨询",
                ]
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(len(rows), 1)
        self.assertTrue(rows[0].is_valid)
        self.assertEqual(rows[0].knowledge_origin, "headquarters_standard")
        self.assertEqual(rows[0].business_type, "self_operated")
        self.assertEqual(rows[0].category_id, "cat-process")
        self.assertEqual(rows[0].subtitles, ["黑屏怎么办", "无法启动"])
        self.assertEqual(rows[0].applicable_scenes, ["无法开机", "售后咨询"])

    def test_export_maps_aggregated_business_type_to_chinese_label(self):
        item = SimpleNamespace(
            id="K-00002",
            knowledge_origin="business_accumulation",
            business_type="aggregated",
            title="聚合回收说明",
            content="聚合回收正文。",
            category_id="cat-process",
            category=SimpleNamespace(name="操作流程"),
            status="review",
        )

        payload = build_knowledge_export_workbook([item])
        workbook = load_workbook(BytesIO(payload), data_only=True)

        self.assertEqual(workbook["知识库主表"]["H2"].value, "业务沉淀")
        self.assertEqual(workbook["知识库主表"]["I2"].value, "聚合回收")

    def test_parse_splits_related_standard_items(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容", "关联标准项"],
            [
                [
                    "设备外观检查",
                    "标准定义",
                    "按标准完成外观检查。",
                    "屏幕检测；边框检测\n摄像头检测",
                ]
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertTrue(rows[0].is_valid)
        self.assertEqual(
            rows[0].related_standard_items,
            ["屏幕检测", "边框检测", "摄像头检测"],
        )

    def test_parse_accepts_full_category_path(self):
        payload = self.workbook_bytes(
            ["标题", "知识分类", "正文"],
            [["流程说明", "质检/操作流程", "按流程逐项检查。"]],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertTrue(rows[0].is_valid)
        self.assertEqual(rows[0].category_id, "cat-process")

    def test_invalid_rows_are_reported_without_hiding_valid_rows(self):
        payload = self.workbook_bytes(
            ["标题", "知识分类", "正文"],
            [
                ["有效知识", "cat-process", "有效正文内容。"],
                ["分类错误", "不存在的分类", "正文内容。"],
                ["", "cat-process", "正文内容。"],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertTrue(rows[0].is_valid)
        self.assertEqual(rows[1].error_code, "CATEGORY_NOT_FOUND")
        self.assertEqual(rows[2].error_code, "TITLE_REQUIRED")

    def test_parse_accepts_governed_qc_workbook_format(self):
        payload = self.workbook_bytes(
            ["主标题", "副标题", "知识内容", "知识分类", "适用范围", "生效状态"],
            [
                [
                    "苹果设备外观判定",
                    "苹果设备外观怎么检查\n苹果外观如何判定",
                    "按质检标准逐项检查。\n"
                    "[img:https://cdn.example.com/qc/apple.png]",
                    "场景判定",
                    "苹果",
                    "生效中",
                ],
                [
                    "安卓设备检测方法",
                    "安卓设备怎么检测",
                    "按步骤执行检测。",
                    "检测方法",
                    "通用",
                    "待审核",
                ],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(len(rows), 2)
        self.assertTrue(rows[0].is_valid)
        self.assertEqual(rows[0].category_id, "cat-qc-standard")
        self.assertEqual(
            rows[0].subtitles,
            ["苹果设备外观怎么检查", "苹果外观如何判定"],
        )
        self.assertEqual(
            rows[0].content,
            {
                "blocks": [
                    {"type": "text", "value": "按质检标准逐项检查。"},
                    {
                        "type": "image",
                        "external_url": "https://cdn.example.com/qc/apple.png",
                        "alt": "",
                        "caption": "",
                    },
                ]
            },
        )
        self.assertEqual(rows[0].applicable_scenes, ["适用范围：苹果"])
        self.assertEqual(rows[0].applicable_brands, ["苹果"])
        self.assertEqual(rows[0].source_status, "生效中")
        self.assertEqual(rows[0].source_scope, "苹果")
        self.assertTrue(rows[1].is_valid)
        self.assertEqual(rows[1].applicable_brands, [])
        self.assertEqual(rows[1].source_status, "待审核")

    def test_legacy_scope_maps_only_explicit_brand_names(self):
        payload = self.workbook_bytes(
            ["主标题", "知识内容", "知识分类", "适用范围", "生效状态"],
            [
                ["组合品牌", "正文。", "场景判定", "小米/红米", "生效中"],
                ["安卓范围", "正文。", "场景判定", "安卓", "生效中"],
                ["其他品牌", "正文。", "场景判定", "其他品牌", "生效中"],
                ["普通场景", "正文。", "场景判定", "线下质检", "生效中"],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(rows[0].applicable_brands, ["小米", "红米"])
        self.assertEqual(rows[1].applicable_brands, [])
        self.assertEqual(rows[2].applicable_brands, ["其他品牌"])
        self.assertEqual(rows[3].applicable_brands, [])

    def test_parse_binds_optional_source_identifiers(self):
        payload = self.workbook_bytes(
            ["主题键", "记录ID", "知识键", "主标题", "知识分类", "知识内容", "生效状态", "来源版本", "校验备注"],
            [[
                "基本情况-外观::通用",
                "rec-source-001",
                "基本情况-外观::平板",
                "平板外观检查",
                "质检标准",
                "按标准检查。",
                "生效中",
                "2026.07",
                "来源主表已核验",
            ]],
        )

        row = parse_knowledge_workbook(payload, self.categories)[0]

        self.assertTrue(row.is_valid)
        self.assertEqual(row.source_topic_key, "基本情况-外观::通用")
        self.assertEqual(row.source_record_id, "rec-source-001")
        self.assertEqual(row.source_knowledge_key, "基本情况-外观::平板")
        self.assertEqual(
            row.source_fields,
            {
                "知识来源": "总部标准",
                "业务类型": "自营回收",
                "主题键": "基本情况-外观::通用",
                "记录ID": "rec-source-001",
                "知识键": "基本情况-外观::平板",
                "主标题": "平板外观检查",
                "知识分类": "质检标准",
                "知识内容": "按标准检查。",
                "生效状态": "生效中",
                "来源版本": "2026.07",
                "校验备注": "来源主表已核验",
            },
        )

    def test_parse_disabled_source_row_requires_identifier_but_not_content_fields(self):
        payload = self.workbook_bytes(
            ["知识键", "生效状态"],
            [
                ["基本情况-外观::通用", "已禁用"],
                ["", "已禁用"],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertTrue(rows[0].is_valid)
        self.assertEqual(rows[0].source_knowledge_key, "基本情况-外观::通用")
        self.assertEqual(rows[1].error_code, "SOURCE_IDENTIFIER_REQUIRED")

    def test_parse_maps_qc_category_values_to_system_categories(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容"],
            [
                ["标准定义示例", "标准定义", "标准正文。"],
                ["检测方法示例", "检测方法", "检测步骤。"],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(rows[0].category_id, "cat-qc-standard")
        self.assertEqual(rows[1].category_id, "cat-process")

    def test_parse_accepts_business_type_labels_and_codes(self):
        payload = self.workbook_bytes(
            ["业务类型", "标题", "知识分类", "正文"],
            [
                ["自营回收", "自营中文", "cat-process", "正文。"],
                ["self_operated", "自营代码", "cat-process", "正文。"],
                ["聚合回收", "聚合中文", "cat-process", "正文。"],
                ["aggregated", "聚合代码", "cat-process", "正文。"],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            [row.business_type for row in rows],
            [
                "self_operated",
                "self_operated",
                "aggregated",
                "aggregated",
            ],
        )
        self.assertTrue(all(row.is_valid for row in rows))

    def test_parse_accepts_knowledge_origin_labels_codes_and_business_source_alias(self):
        payload = self.workbook_bytes(
            ["知识来源", "业务类型", "标题", "知识分类", "正文"],
            [
                ["总部标准", "自营回收", "总部中文", "cat-process", "正文。"],
                [
                    "headquarters_standard",
                    "self_operated",
                    "总部代码",
                    "cat-process",
                    "正文。",
                ],
                ["业务沉淀", "聚合回收", "业务中文", "cat-process", "正文。"],
                [
                    "business_accumulation",
                    "aggregated",
                    "业务代码",
                    "cat-process",
                    "正文。",
                ],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            [row.knowledge_origin for row in rows],
            [
                "headquarters_standard",
                "headquarters_standard",
                "business_accumulation",
                "business_accumulation",
            ],
        )
        self.assertTrue(all(row.is_valid for row in rows))

        alias_payload = self.workbook_bytes(
            ["业务来源", "业务类型", "标题", "知识分类", "正文"],
            [["总部标准", "自营回收", "兼容别名", "cat-process", "正文。"]],
        )
        alias_row = parse_knowledge_workbook(alias_payload, self.categories)[0]
        self.assertEqual(alias_row.knowledge_origin, "headquarters_standard")

    def test_parse_reports_missing_and_invalid_business_type_per_row(self):
        payload = self.workbook_bytes(
            ["业务类型", "标题", "知识分类", "正文"],
            [
                ["", "缺失业务类型", "cat-process", "正文。"],
                ["寄售回收", "非法业务类型", "cat-process", "正文。"],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(rows[0].error_code, "BUSINESS_TYPE_REQUIRED")
        self.assertEqual(rows[0].error_message, "业务类型不能为空。")
        self.assertEqual(rows[1].error_code, "BUSINESS_TYPE_INVALID")
        self.assertIn("仅允许自营回收、聚合回收", rows[1].error_message)

    def test_parse_reports_missing_and_invalid_knowledge_origin_per_row(self):
        payload = self.workbook_bytes(
            ["知识来源", "业务类型", "标题", "知识分类", "正文"],
            [
                ["", "自营回收", "缺失来源", "cat-process", "正文。"],
                ["内部草稿", "自营回收", "非法来源", "cat-process", "正文。"],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(rows[0].error_code, "KNOWLEDGE_ORIGIN_REQUIRED")
        self.assertEqual(rows[0].error_message, "知识来源不能为空。")
        self.assertEqual(rows[1].error_code, "KNOWLEDGE_ORIGIN_INVALID")
        self.assertIn("仅允许总部标准、业务沉淀", rows[1].error_message)

    def test_managed_model_configuration_origin_rejects_generic_excel_import(self):
        payload = self.workbook_bytes(
            ["知识来源", "业务类型", "标题", "知识分类", "正文"],
            [
                [
                    "机型配置信息",
                    "自营回收",
                    "iPad 机型配置",
                    "cat-process",
                    "配置正文。",
                ],
                [
                    "model_configuration",
                    "self_operated",
                    "iPad 机型配置代码",
                    "cat-process",
                    "配置正文。",
                ],
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            [row.error_code for row in rows],
            ["KNOWLEDGE_ORIGIN_MANAGED", "KNOWLEDGE_ORIGIN_MANAGED"],
        )
        self.assertTrue(
            all("飞书专用同步" in row.error_message for row in rows)
        )

    def test_parse_only_promotes_prefixed_media_tokens(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容"],
            [
                [
                    "链接与图片示例",
                    "标准定义",
                    "帮助地址：https://example.com/help\n"
                    "[img:https://cdn.example.com/image-resource?version=2]\n"
                    "后续说明。",
                ]
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            rows[0].content,
            {
                "blocks": [
                    {
                        "type": "text",
                        "value": "帮助地址：https://example.com/help",
                    },
                    {
                        "type": "image",
                        "external_url": "https://cdn.example.com/image-resource?version=2",
                        "alt": "",
                        "caption": "",
                    },
                    {"type": "text", "value": "后续说明。"},
                ]
            },
        )

    def test_parse_prefixed_media_token_preserves_inline_position(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容"],
            [
                [
                    "行内媒体示例",
                    "标准定义",
                    "图片前文[img:https://cdn.example.com/image-resource]图片后文",
                ]
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            rows[0].content,
            {
                "blocks": [
                    {"type": "text", "value": "图片前文"},
                    {
                        "type": "image",
                        "external_url": "https://cdn.example.com/image-resource",
                        "alt": "",
                        "caption": "",
                    },
                    {"type": "text", "value": "图片后文"},
                ]
            },
        )

    def test_parse_drops_standalone_list_markers_adjacent_to_media(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容"],
            [
                [
                    "媒体列表占位符",
                    "标准定义",
                    "【检测方法】\n"
                    "- 通过官网查询。\n"
                    "-\n"
                    "[img:https://cdn.example.com/first.png]\n"
                    "-\n"
                    "[img:https://cdn.example.com/second.png]\n"
                    "- 安卓系统按其他版本判断。",
                ]
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            rows[0].content,
            {
                "blocks": [
                    {
                        "type": "text",
                        "value": "【检测方法】\n- 通过官网查询。",
                    },
                    {
                        "type": "image",
                        "external_url": "https://cdn.example.com/first.png",
                        "alt": "",
                        "caption": "",
                    },
                    {
                        "type": "image",
                        "external_url": "https://cdn.example.com/second.png",
                        "alt": "",
                        "caption": "",
                    },
                    {
                        "type": "text",
                        "value": "- 安卓系统按其他版本判断。",
                    },
                ]
            },
        )

    def test_parse_keeps_standalone_marker_without_media(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容"],
            [["保留普通横杠", "标准定义", "说明文字\n-"]],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(rows[0].content, "说明文字\n-")

    def test_parse_preserves_multi_character_separator_before_media(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容"],
            [
                [
                    "保留正文分隔线",
                    "标准定义",
                    "说明文字\n---\n"
                    "[img:https://cdn.example.com/separator.png]",
                ]
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            rows[0].content,
            {
                "blocks": [
                    {"type": "text", "value": "说明文字\n---"},
                    {
                        "type": "image",
                        "external_url": "https://cdn.example.com/separator.png",
                        "alt": "",
                        "caption": "",
                    },
                ]
            },
        )

    def test_parse_preserves_external_media_order_and_duplicate_references(self):
        repeated_image = "https://cdn.example.com/psn-resource"
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容"],
            [
                [
                    "序列号查看说明",
                    "标准定义",
                    "【苹果】补充：\n"
                    "[img:https://cdn.example.com/apple-resource]\n"
                    "【安卓】补充：\n"
                    "[video:https://cdn.example.com/android-stream?version=2]\n"
                    "【小米/红米】PSN码查看：\n"
                    f"[img:{repeated_image}]\n"
                    "【小米/红米】补充：\n"
                    f"[img:{repeated_image}]",
                ]
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            rows[0].content,
            {
                "blocks": [
                    {"type": "text", "value": "【苹果】补充："},
                    {
                        "type": "image",
                        "external_url": "https://cdn.example.com/apple-resource",
                        "alt": "",
                        "caption": "",
                    },
                    {"type": "text", "value": "【安卓】补充："},
                    {
                        "type": "video",
                        "external_url": "https://cdn.example.com/android-stream?version=2",
                        "alt": "",
                        "caption": "",
                    },
                    {"type": "text", "value": "【小米/红米】PSN码查看："},
                    {
                        "type": "image",
                        "external_url": repeated_image,
                        "alt": "",
                        "caption": "",
                    },
                    {"type": "text", "value": "【小米/红米】补充："},
                    {
                        "type": "image",
                        "external_url": repeated_image,
                        "alt": "",
                        "caption": "",
                    },
                ]
            },
        )

    def test_parse_keeps_unprefixed_urls_as_plain_text(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容"],
            [
                [
                    "原有链接示例",
                    "标准定义",
                    "官网：https://example.com\n"
                    "https://cdn.example.com/raw-image.png\n"
                    "https://cdn.example.com/raw-video.mp4",
                ]
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            rows[0].content,
            "官网：https://example.com\n"
            "https://cdn.example.com/raw-image.png\n"
            "https://cdn.example.com/raw-video.mp4",
        )

    def test_parse_keeps_unsafe_prefixed_url_as_plain_text(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容"],
            [
                [
                    "不安全媒体标记",
                    "标准定义",
                    "[img:https://user@example.com/private-resource]",
                ]
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(
            rows[0].content,
            "[img:https://user@example.com/private-resource]",
        )

    def test_parse_accepts_195_rows(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容", "生效状态"],
            [
                [f"知识 {index}", "标准定义", f"正文 {index}", "生效中"]
                for index in range(195)
            ],
        )

        rows = parse_knowledge_workbook(payload, self.categories)

        self.assertEqual(len(rows), 195)
        self.assertTrue(all(row.is_valid for row in rows))

    def test_parse_rejects_more_than_500_rows(self):
        payload = self.workbook_bytes(
            ["标题", "知识分类", "正文"],
            [
                [f"知识 {index}", "cat-process", f"正文 {index}"]
                for index in range(501)
            ],
        )

        with self.assertRaisesRegex(KnowledgeExcelError, "单次最多导入 500 条"):
            parse_knowledge_workbook(payload, self.categories)

    def test_missing_required_headers_rejects_workbook(self):
        payload = self.workbook_bytes(
            ["标题", "知识分类"],
            [["缺少正文列", "cat-process"]],
        )

        with self.assertRaisesRegex(KnowledgeExcelError, "缺少必填列"):
            parse_knowledge_workbook(payload, self.categories)

    def test_old_template_without_business_type_header_is_rejected(self):
        payload = self.workbook_bytes(
            ["标题", "知识分类", "正文"],
            [["旧模板知识", "cat-process", "旧模板正文。"]],
            include_business_type=False,
        )

        with self.assertRaisesRegex(
            KnowledgeExcelError,
            "缺少必填列：业务类型",
        ):
            parse_knowledge_workbook(payload, self.categories)

    def test_main_sheet_without_knowledge_origin_header_is_rejected(self):
        payload = self.workbook_bytes(
            ["主标题", "知识分类", "知识内容", "生效状态"],
            [["主表缺少来源", "cat-process", "正文。", "生效中"]],
            include_knowledge_origin=False,
        )

        with self.assertRaisesRegex(
            KnowledgeExcelError,
            "缺少必填列：知识来源",
        ):
            parse_knowledge_workbook(payload, self.categories)

    def test_export_uses_business_source_fallback_and_separate_entry_method(self):
        item = SimpleNamespace(
            id="K-00003",
            knowledge_origin=None,
            business_type="self_operated",
            title="来源兼容导出",
            content="正文。",
            category_id="cat-process",
            category=SimpleNamespace(name="操作流程"),
            source="manual",
            source_fields={"业务来源": "业务沉淀"},
            status="review",
        )

        payload = build_knowledge_export_workbook([item])
        workbook = load_workbook(BytesIO(payload), data_only=True)
        sheet = workbook["知识库主表"]

        self.assertEqual(sheet["H2"].value, "业务沉淀")
        self.assertEqual(sheet["K2"].value, "手工录入")

    def test_export_labels_automation_as_api_entry(self):
        item = SimpleNamespace(
            id="K-00004",
            knowledge_origin="business_accumulation",
            business_type="aggregated",
            title="自动化入库",
            content="正文。",
            category_id="cat-process",
            category=SimpleNamespace(name="操作流程"),
            source="automation",
            source_fields={},
            status="review",
        )

        payload = build_knowledge_export_workbook([item])
        workbook = load_workbook(BytesIO(payload), data_only=True)
        sheet = workbook["知识库主表"]

        self.assertEqual(sheet["K2"].value, "接口导入")


if __name__ == "__main__":
    unittest.main()
