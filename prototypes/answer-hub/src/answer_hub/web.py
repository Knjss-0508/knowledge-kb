from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from email.parser import BytesParser
from email.policy import default as email_policy
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
import json
import os

try:
    from flask import Flask, jsonify, render_template, request, send_file
except ModuleNotFoundError:
    Flask = None  # type: ignore[assignment,misc]
from openpyxl import load_workbook

from .audit import AuditStore
from .catalog import load_standard_catalog
from .excel_io import read_workbook_rows
from .mimo import MimoClient
from .workflow import (
    filter_source_rows_by_product_type,
    generate_phone_candidate_rows,
    initial_label_from_workbook,
    preprocess_source_rows,
    REVIEW_COLUMNS,
    REVIEW_DECISIONS,
)


MAX_PREVIEW_ROWS = 200
MAX_REVIEW_ROWS = 500
PACKAGE_DIR = Path(__file__).resolve().parent


def _as_json(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, list):
        return [_as_json(item) for item in value]
    if isinstance(value, dict):
        return {key: _as_json(item) for key, item in value.items()}
    return str(value)


def _threshold(value: str | None) -> float:
    try:
        parsed = float(value or 0.75)
    except ValueError:
        parsed = 0.75
    return max(0.1, min(parsed, 0.98))


def _save_upload(upload, directory: Path, name: str) -> Path | None:
    if upload is None or not upload.filename:
        return None
    suffix = Path(upload.filename).suffix.lower()
    if suffix not in {".xlsx", ".json"}:
        raise ValueError("只支持 .xlsx 或 .json 文件")
    target = directory / f"{name}{suffix}"
    upload.save(target)
    return target


def _review_rows_from_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    if "review_queue" not in workbook.sheetnames:
        raise ValueError("复核工作簿缺少 review_queue 工作表")
    worksheet = workbook["review_queue"]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return []
    columns = [str(value).strip() if value is not None else "" for value in values[0]]
    rows: list[dict[str, Any]] = []
    for row_index, values_row in enumerate(values[1:], start=2):
        row = {
            column: values_row[index] if index < len(values_row) else None
            for index, column in enumerate(columns)
            if column
        }
        if any(value not in (None, "") for value in row.values()):
            row["_review_row_index"] = row_index
            rows.append(row)
    return rows


def _review_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    decisions = Counter(str(row.get("CZ复核结论") or "").strip() for row in rows)
    return {
        "total_rows": len(rows),
        "reviewed_rows": sum(decisions[decision] for decision in REVIEW_DECISIONS),
        "pending_rows": decisions.get("", 0),
        "focus_review_rows": sum(str(row.get("是否重点复核") or "").strip() == "是" for row in rows),
        "decision_counts": {decision: decisions.get(decision, 0) for decision in REVIEW_DECISIONS},
    }


def _review_upload(upload, directory: Path) -> Path:
    path = _save_upload(upload, directory, "review_queue")
    if path is None:
        raise ValueError("请上传 review_queue.xlsx")
    if path.suffix.lower() != ".xlsx":
        raise ValueError("复标工作台只支持 .xlsx 复核工作簿")
    return path


def _build_preview(
    source_path: Path,
    standards_path: Path | None,
    product_type: str,
    min_confidence: float,
    use_mimo: bool = True,
) -> dict[str, Any]:
    _, source_rows = read_workbook_rows(source_path)
    selected_rows, excluded_rows = filter_source_rows_by_product_type(source_rows, product_type)
    preprocessed_rows = preprocess_source_rows(selected_rows)
    standard_catalog = load_standard_catalog(standards_path)
    audit_store = AuditStore.from_env()
    candidates, run_id = generate_phone_candidate_rows(
        preprocessed_rows,
        standard_catalog,
        min_confidence=min_confidence,
        raw_source_rows=selected_rows,
        use_mimo=use_mimo,
        audit_store=audit_store,
    )
    evidence_counter = Counter(str(row.get("证据等级") or "") for row in candidates)
    stage_counter = Counter(row["模型阶段状态"] for row in candidates)
    image_unavailable_rows = sum("不可用:" in str(row.get("图片处理状态") or "") for row in candidates)
    missing_counter = Counter()
    for row in preprocessed_rows:
        missing = str(row.get("缺失字段") or "").strip()
        if not missing:
            missing_counter["无缺失"] += 1
        else:
            missing_counter.update(item for item in missing.splitlines() if item)

    return {
        "product_type": product_type,
        "standard_count": len(standard_catalog),
        "source_total_rows": len(source_rows),
        "selected_rows": len(selected_rows),
        "excluded_rows": len(excluded_rows),
        "focus_review_rows": 0,
        "model_failed_rows": stage_counter.get("topic_model_failed", 0),
        "evidence_level_counts": dict(evidence_counter),
        "image_unavailable_rows": image_unavailable_rows,
        "run_id": run_id,
        "mimo_configured": bool(MimoClient.from_env()),
        "mimo_enabled": use_mimo,
        "audit_db": str(audit_store.path),
        "missing_field_counts": dict(missing_counter),
        "candidates": _as_json(
            [
                {
                    **row,
                    "模型一级分类": row.get("一级分类", ""),
                    "模型二级分类": row.get("二级分类", ""),
                }
                for row in candidates[:MAX_PREVIEW_ROWS]
            ]
        ),
        "preprocessed": _as_json(preprocessed_rows[:MAX_PREVIEW_ROWS]),
        "excluded": _as_json(excluded_rows[:MAX_PREVIEW_ROWS]),
        "truncated": len(candidates) > MAX_PREVIEW_ROWS,
    }


@dataclass(frozen=True)
class _MemoryUpload:
    filename: str
    payload: bytes

    def save(self, path: str | Path) -> None:
        Path(path).write_bytes(self.payload)


def _parse_multipart(content_type: str, body: bytes) -> tuple[dict[str, str], dict[str, _MemoryUpload]]:
    if not content_type.lower().startswith("multipart/form-data"):
        raise ValueError("请求必须使用 multipart/form-data")
    envelope = (
        f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8")
        + body
    )
    message = BytesParser(policy=email_policy).parsebytes(envelope)
    if not message.is_multipart():
        raise ValueError("无法解析上传表单")
    fields: dict[str, str] = {}
    files: dict[str, _MemoryUpload] = {}
    for part in message.iter_parts():
        name = part.get_param("name", header="content-disposition")
        if not name:
            continue
        payload = part.get_payload(decode=True) or b""
        filename = part.get_filename()
        if filename:
            files[name] = _MemoryUpload(filename=filename, payload=payload)
        else:
            fields[name] = payload.decode(part.get_content_charset() or "utf-8", errors="replace")
    return fields, files


def _review_queue_payload(upload: _MemoryUpload) -> dict[str, Any]:
    with TemporaryDirectory(prefix="answer-hub-review-view-") as temp_dir:
        path = _review_upload(upload, Path(temp_dir))
        rows = _review_rows_from_workbook(path)
    payload = _review_summary(rows)
    payload.update(
        {
            "rows": _as_json(rows[:MAX_REVIEW_ROWS]),
            "truncated": len(rows) > MAX_REVIEW_ROWS,
        }
    )
    return payload


def _export_review_workbook(upload: _MemoryUpload, raw_changes: str) -> bytes:
    try:
        changes = json.loads(raw_changes or "[]")
    except json.JSONDecodeError as exc:
        raise ValueError("复标修改内容不是有效 JSON") from exc
    if not isinstance(changes, list):
        raise ValueError("复标修改内容必须是数组")

    with TemporaryDirectory(prefix="answer-hub-review-export-") as temp_dir:
        path = _review_upload(upload, Path(temp_dir))
        workbook = load_workbook(path)
        if "review_queue" not in workbook.sheetnames:
            raise ValueError("复核工作簿缺少 review_queue 工作表")
        worksheet = workbook["review_queue"]
        header_map = {
            str(cell.value).strip(): cell.column
            for cell in worksheet[1]
            if cell.value is not None and str(cell.value).strip()
        }
        missing_columns = [column for column in REVIEW_COLUMNS if column not in header_map]
        if missing_columns:
            raise ValueError(f"复核工作簿缺少 CZ 复标列：{', '.join(missing_columns)}")

        for change in changes:
            if not isinstance(change, dict):
                raise ValueError("每条复标修改必须是对象")
            row_index = change.get("row_index")
            updates = change.get("updates")
            if not isinstance(row_index, int) or not 2 <= row_index <= worksheet.max_row:
                raise ValueError("复标修改包含无效行号")
            if not isinstance(updates, dict):
                raise ValueError("复标修改缺少 updates 对象")
            decision = str(updates.get("CZ复核结论") or "").strip()
            if decision and decision not in REVIEW_DECISIONS:
                raise ValueError(f"不支持的复核结论：{decision}")
            for field, value in updates.items():
                if field not in REVIEW_COLUMNS:
                    raise ValueError(f"不允许修改字段：{field}")
                if isinstance(value, list):
                    value = "\n".join(str(item) for item in value if item not in (None, ""))
                worksheet.cell(row=row_index, column=header_map[field], value="" if value is None else str(value))

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


def _review_workbook_bytes(
    source_upload: _MemoryUpload,
    standards_upload: _MemoryUpload | None,
    product_type: str,
    min_confidence: float,
    use_mimo: bool,
) -> bytes:
    with TemporaryDirectory(prefix="answer-hub-review-") as temp_dir:
        directory = Path(temp_dir)
        source_path = _save_upload(source_upload, directory, "source")
        standards_path = _save_upload(standards_upload, directory, "standards")
        output_dir = directory / "outputs"
        initial_label_from_workbook(
            source_path=source_path,
            standards_path=standards_path,
            output_dir=output_dir,
            min_confidence=min_confidence,
            product_type=product_type,
            use_mimo=use_mimo,
        )
        return (output_dir / "review_queue.xlsx").read_bytes()


class _FallbackRequestHandler(BaseHTTPRequestHandler):
    server_version = "AnswerHubWeb/1.0"

    def log_message(self, _format: str, *_args: Any) -> None:
        return

    def _send_json(self, payload: dict[str, Any], status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False, default=_as_json).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_bytes(self, body: bytes, content_type: str, filename: str = "") -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        if filename:
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:  # noqa: N802
        if self.path == "/":
            template = (PACKAGE_DIR / "templates" / "index.html").read_text(encoding="utf-8")
            template = template.replace(
                "{{ url_for('static', filename='app.css') }}",
                "/static/app.css",
            ).replace(
                "{{ url_for('static', filename='app.js') }}",
                "/static/app.js",
            )
            self._send_bytes(template.encode("utf-8"), "text/html; charset=utf-8")
            return
        static_files = {
            "/static/app.css": ("app.css", "text/css; charset=utf-8"),
            "/static/app.js": ("app.js", "application/javascript; charset=utf-8"),
        }
        target = static_files.get(self.path)
        if target:
            self._send_bytes((PACKAGE_DIR / "static" / target[0]).read_bytes(), target[1])
            return
        self._send_json({"error": "未找到页面"}, 404)

    def do_POST(self) -> None:  # noqa: N802
        try:
            content_length = int(self.headers.get("Content-Length") or "0")
            if content_length > 40 * 1024 * 1024:
                raise ValueError("文件过大，单次上传上限为 40MB")
            body = self.rfile.read(content_length)
            fields, files = _parse_multipart(self.headers.get("Content-Type") or "", body)
            if self.path == "/api/preview":
                source_upload = files.get("source")
                if source_upload is None:
                    raise ValueError("请上传第二部分数据表")
                product_type = (fields.get("product_type") or "手机").strip() or "手机"
                min_confidence = _threshold(fields.get("min_confidence"))
                use_mimo = (fields.get("use_mimo") or "true").lower() not in {"false", "0", "off"}
                with TemporaryDirectory(prefix="answer-hub-preview-") as temp_dir:
                    directory = Path(temp_dir)
                    source_path = _save_upload(source_upload, directory, "source")
                    standards_path = _save_upload(files.get("standards"), directory, "standards")
                    result = _build_preview(source_path, standards_path, product_type, min_confidence, use_mimo)
                self._send_json(result)
                return
            if self.path == "/api/review-workbook":
                source_upload = files.get("source")
                if source_upload is None:
                    raise ValueError("请上传第二部分数据表")
                product_type = (fields.get("product_type") or "手机").strip() or "手机"
                workbook = _review_workbook_bytes(
                    source_upload,
                    files.get("standards"),
                    product_type,
                    _threshold(fields.get("min_confidence")),
                    (fields.get("use_mimo") or "true").lower() not in {"false", "0", "off"},
                )
                self._send_bytes(
                    workbook,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    f"{product_type}-候选知识复核表.xlsx",
                )
                return
            if self.path == "/api/review-queue":
                upload = files.get("review_file")
                if upload is None:
                    raise ValueError("请上传 review_queue.xlsx")
                self._send_json(_review_queue_payload(upload))
                return
            if self.path == "/api/review-export":
                upload = files.get("review_file")
                if upload is None:
                    raise ValueError("请上传 review_queue.xlsx")
                workbook = _export_review_workbook(upload, fields.get("changes") or "[]")
                self._send_bytes(
                    workbook,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "手机-已复标候选知识表.xlsx",
                )
                return
            self._send_json({"error": "未找到接口"}, 404)
        except ValueError as exc:
            self._send_json({"error": str(exc)}, 400)
        except Exception as exc:
            self._send_json({"error": str(exc)}, 500)


def _serve_without_flask(port: int) -> None:
    server = ThreadingHTTPServer(("127.0.0.1", port), _FallbackRequestHandler)
    print(f"Answer Hub review workbench: http://127.0.0.1:{port}")
    server.serve_forever()


def create_app() -> Flask:
    if Flask is None:
        raise RuntimeError("未安装 Flask；请直接运行 python -m answer_hub.web 使用内置本地服务。")
    app = Flask(__name__)
    app.config["MAX_CONTENT_LENGTH"] = 40 * 1024 * 1024

    @app.get("/")
    def index():
        return render_template("index.html")

    @app.post("/api/preview")
    def preview():
        source_upload = request.files.get("source")
        if source_upload is None or not source_upload.filename:
            return jsonify({"error": "请上传第二部分数据表"}), 400

        product_type = (request.form.get("product_type") or "手机").strip() or "手机"
        min_confidence = _threshold(request.form.get("min_confidence"))
        use_mimo = request.form.get("use_mimo", "true").lower() not in {"false", "0", "off"}
        with TemporaryDirectory(prefix="answer-hub-preview-") as temp_dir:
            directory = Path(temp_dir)
            source_path = _save_upload(source_upload, directory, "source")
            standards_path = _save_upload(request.files.get("standards"), directory, "standards")
            result = _build_preview(source_path, standards_path, product_type, min_confidence, use_mimo)
        return jsonify(result)

    @app.post("/api/review-workbook")
    def review_workbook():
        source_upload = request.files.get("source")
        if source_upload is None or not source_upload.filename:
            return jsonify({"error": "请上传第二部分数据表"}), 400

        product_type = (request.form.get("product_type") or "手机").strip() or "手机"
        min_confidence = _threshold(request.form.get("min_confidence"))
        use_mimo = request.form.get("use_mimo", "true").lower() not in {"false", "0", "off"}
        with TemporaryDirectory(prefix="answer-hub-review-") as temp_dir:
            directory = Path(temp_dir)
            source_path = _save_upload(source_upload, directory, "source")
            standards_path = _save_upload(request.files.get("standards"), directory, "standards")
            output_dir = directory / "outputs"
            initial_label_from_workbook(
                source_path=source_path,
                standards_path=standards_path,
                output_dir=output_dir,
                min_confidence=min_confidence,
                product_type=product_type,
                use_mimo=use_mimo,
            )
            workbook = (output_dir / "review_queue.xlsx").read_bytes()

        return send_file(
            BytesIO(workbook),
            as_attachment=True,
            download_name=f"{product_type}-候选知识复核表.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.post("/api/review-queue")
    def review_queue():
        with TemporaryDirectory(prefix="answer-hub-review-view-") as temp_dir:
            path = _review_upload(request.files.get("review_file"), Path(temp_dir))
            rows = _review_rows_from_workbook(path)
        payload = _review_summary(rows)
        payload.update(
            {
                "rows": _as_json(rows[:MAX_REVIEW_ROWS]),
                "truncated": len(rows) > MAX_REVIEW_ROWS,
            }
        )
        return jsonify(payload)

    @app.post("/api/review-export")
    def review_export():
        raw_changes = request.form.get("changes") or "[]"
        try:
            changes = json.loads(raw_changes)
        except json.JSONDecodeError as exc:
            raise ValueError("复标修改内容不是有效 JSON") from exc
        if not isinstance(changes, list):
            raise ValueError("复标修改内容必须是数组")

        with TemporaryDirectory(prefix="answer-hub-review-export-") as temp_dir:
            path = _review_upload(request.files.get("review_file"), Path(temp_dir))
            workbook = load_workbook(path)
            if "review_queue" not in workbook.sheetnames:
                raise ValueError("复核工作簿缺少 review_queue 工作表")
            worksheet = workbook["review_queue"]
            header_map = {
                str(cell.value).strip(): cell.column
                for cell in worksheet[1]
                if cell.value is not None and str(cell.value).strip()
            }
            missing_columns = [column for column in REVIEW_COLUMNS if column not in header_map]
            if missing_columns:
                raise ValueError(f"复核工作簿缺少 CZ 复标列：{', '.join(missing_columns)}")

            for change in changes:
                if not isinstance(change, dict):
                    raise ValueError("每条复标修改必须是对象")
                row_index = change.get("row_index")
                updates = change.get("updates")
                if not isinstance(row_index, int) or not 2 <= row_index <= worksheet.max_row:
                    raise ValueError("复标修改包含无效行号")
                if not isinstance(updates, dict):
                    raise ValueError("复标修改缺少 updates 对象")
                decision = str(updates.get("CZ复核结论") or "").strip()
                if decision and decision not in REVIEW_DECISIONS:
                    raise ValueError(f"不支持的复核结论：{decision}")
                for field, value in updates.items():
                    if field not in REVIEW_COLUMNS:
                        raise ValueError(f"不允许修改字段：{field}")
                    if isinstance(value, list):
                        value = "\n".join(str(item) for item in value if item not in (None, ""))
                    worksheet.cell(row=row_index, column=header_map[field], value="" if value is None else str(value))

            output = BytesIO()
            workbook.save(output)
            output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="手机-已复标候选知识表.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.errorhandler(413)
    def request_too_large(_error):
        return jsonify({"error": "文件过大，单次上传上限为 40MB"}), 413

    @app.errorhandler(Exception)
    def unexpected_error(error):
        app.logger.exception("validation page error")
        return jsonify({"error": str(error)}), 500

    return app


def main() -> None:
    port = int(os.getenv("ANSWER_HUB_WEB_PORT", "8765"))
    if Flask is None:
        _serve_without_flask(port)
    else:
        create_app().run(host="127.0.0.1", port=port, debug=False)


if __name__ == "__main__":
    main()
