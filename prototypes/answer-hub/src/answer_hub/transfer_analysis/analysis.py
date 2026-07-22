from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from random import Random
from typing import Any, Callable, Iterable
import csv
import json
import math
import re
import uuid

from openpyxl import load_workbook

from answer_hub.catalog import StandardCatalogItem, load_standard_catalog
from answer_hub.excel_io import write_rows_to_workbook
from answer_hub.mimo import MimoClient, MimoError, MimoLabelResult
from answer_hub.workflow import retrieve_topic_signal_matches

from .collectors import collect_with_endpoint_profile
from .schema import (
    ANALYSIS_COLUMNS,
    BAD_CASE_COLUMNS,
    DEFAULT_CAPABILITY_REGISTRY,
    KNOWLEDGE_GAP_COLUMNS,
    OWNER_COLUMNS,
    RETRIEVAL_COLUMNS,
    TOOL_ISSUE_COLUMNS,
    TRANSFER_ANALYSIS_PROMPT_VERSION,
    TRANSFER_REASON_OPTIONS,
    CapabilityRegistry,
)
from .store import TransferAnalysisStore


ProgressCallback = Callable[[dict[str, Any]], None]

RISK_REASONS = {"答非所问", "回答内容无法理解", "其他"}
SUCCESS_TOOL_STATUSES = {"success", "succeeded", "completed", "ok", "成功", "已完成", "调用成功"}
FAILED_TOOL_STATUSES = {"failed", "error", "timeout", "失败", "调用失败", "超时"}
USER_ROLE_MARKERS = ("user", "human", "engineer", "工程师", "用户", "提问")
ASSISTANT_ROLE_MARKERS = ("assistant", "bot", "ai", "百晓生", "机器人", "模型")
VAGUE_QUESTIONS = {
    "怎么办",
    "怎么弄",
    "怎么处理",
    "有问题",
    "不行",
    "帮忙看下",
    "帮看一下",
    "看一下",
    "这个呢",
    "怎么回事",
}
INVALID_QUESTIONS = {"", "你好", "您好", "在吗", "转人工", "人工", "找人工", "人工客服"}
MANUAL_OPERATION_MARKERS = (
    "查后台",
    "查询后台",
    "查订单",
    "查物流",
    "修改订单",
    "取消订单",
    "退款",
    "人工审核",
    "人工操作",
    "开权限",
    "实时库存",
    "实时价格",
)
IMAGE_REQUEST_MARKERS = (
    "看图",
    "图片",
    "照片",
    "截图",
    "识别一下",
    "帮我看看这个",
    "这是什么",
)

FIELD_ALIASES = {
    "transfer_id": ("转人工ID", "转人工会话ID", "会话id", "会话ID", "会话Id", "id", "ID"),
    "conversation_id": ("会话id", "会话ID", "会话Id", "conversation_id"),
    "source_id": ("会话id", "会话ID", "会话Id", "conversation_id", "id", "ID"),
    "work_order_id": ("工单ID", "工单Id", "工单id", "work_order_id"),
    "event_time": ("转人工时间", "创建时间", "分析时间", "会话时间", "时间"),
    "started_at": ("开始时间", "会话开始时间", "创建时间", "分析时间"),
    "ended_at": ("结束时间", "会话结束时间", "更新时间", "转人工时间", "分析时间"),
    "engineer": ("工程师", "上传者", "提问人", "用户", "创建人"),
    "transfer_reason": ("转人工原因", "转人工选择原因", "原因"),
    "category": ("类目", "产品类型", "一级分类"),
    "model": ("机型", "型号"),
    "order_status": ("订单所处状态", "订单状态", "状态"),
    "conversation_text": ("聊天内容", "会话内容", "完整会话", "转人工后完整会话"),
    "first_question": ("问题", "首轮问题", "核心问题", "用户问题"),
    "last_answer": ("回答", "大模型回答", "最后一轮回复", "机器人回复"),
    "intent_result": ("意图识别结果", "意图结果", "识别意图"),
    "retrievals": ("召回知识", "召回结果", "知识召回", "检索结果"),
    "top_similarity": ("Top相似度", "最高相似度", "知识相似度", "相似度"),
    "production_threshold": ("生产阈值", "相似度阈值", "阈值"),
    "tools": ("工具调用", "工具调用记录", "工具结果", "调用工具"),
    "attachments": ("图片链接", "附件", "附件链接", "图片"),
}


def _text(value: Any, limit: int = 12000) -> str:
    if value is None:
        return ""
    return str(value).strip()[:limit]


def _pick(row: dict[str, Any], canonical: str, default: Any = "") -> Any:
    for alias in FIELD_ALIASES.get(canonical, (canonical,)):
        if alias in row and row[alias] not in (None, "", [], {}):
            return row[alias]
    return default


def _parse_datetime(value: Any) -> datetime | None:
    text = _text(value)
    if not text:
        return None
    normalized = text.replace("Z", "+00:00").replace("/", "-")
    candidates = [
        normalized,
        normalized.replace("T", " "),
    ]
    for candidate in candidates:
        try:
            return datetime.fromisoformat(candidate)
        except ValueError:
            pass
    for pattern in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y%m%d%H%M%S",
    ):
        try:
            return datetime.strptime(normalized, pattern)
        except ValueError:
            continue
    return None


def _iso_datetime(value: Any) -> str:
    parsed = _parse_datetime(value)
    return parsed.isoformat(timespec="seconds") if parsed else _text(value)


def _role(item: dict[str, Any]) -> str:
    return _text(
        item.get("role")
        or item.get("senderType")
        or item.get("speaker")
        or item.get("sender")
        or item.get("name")
    ).lower()


def _message_content(item: dict[str, Any]) -> str:
    return _text(
        item.get("content")
        or item.get("text")
        or item.get("message")
        or item.get("body"),
        6000,
    )


def _is_user_role(role: str) -> bool:
    return any(marker in role for marker in USER_ROLE_MARKERS)


def _is_assistant_role(role: str) -> bool:
    return any(marker in role for marker in ASSISTANT_ROLE_MARKERS)


def _parse_conversation_text(value: Any) -> list[dict[str, str]]:
    if isinstance(value, list):
        results: list[dict[str, str]] = []
        for item in value:
            if isinstance(item, dict):
                content = _message_content(item)
                if content:
                    results.append({"role": _role(item) or "unknown", "content": content})
            elif _text(item):
                results.append({"role": "unknown", "content": _text(item)})
        return results
    text = _text(value)
    if not text:
        return []
    messages: list[dict[str, str]] = []
    pattern = re.compile(
        r"^\s*(工程师|用户|提问人|百晓生|机器人|助手|AI|人工客服|客服)\s*[:：]\s*(.*)$",
        re.IGNORECASE,
    )
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        match = pattern.match(line)
        if match:
            speaker, content = match.groups()
            role = "assistant" if speaker in {"百晓生", "机器人", "助手", "AI"} else "user"
            messages.append({"role": role, "content": content.strip()})
        elif messages:
            messages[-1]["content"] = f"{messages[-1]['content']}\n{line}".strip()
        else:
            messages.append({"role": "unknown", "content": line})
    return messages


def _conversation_text(messages: list[dict[str, Any]], fallback: str = "") -> str:
    lines: list[str] = []
    for message in messages:
        content = _message_content(message)
        if not content:
            continue
        role = _role(message)
        label = "工程师" if _is_user_role(role) else ("百晓生" if _is_assistant_role(role) else "未知")
        lines.append(f"{label}：{content}")
    return "\n".join(lines) or _text(fallback)


def _first_question(messages: list[dict[str, Any]], fallback: Any = "") -> str:
    for message in messages:
        if _is_user_role(_role(message)) and _message_content(message):
            return _message_content(message)
    return _text(fallback)


def _last_answer(messages: list[dict[str, Any]], fallback: Any = "") -> str:
    answer = ""
    for message in messages:
        if _is_assistant_role(_role(message)) and _message_content(message):
            answer = _message_content(message)
    return answer or _text(fallback)


def _parse_json_or_lines(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        return [value]
    text = _text(value)
    if not text:
        return []
    if text[:1] in {"[", "{"}:
        try:
            parsed = json.loads(text)
            return parsed if isinstance(parsed, list) else [parsed]
        except json.JSONDecodeError:
            pass
    return [line.strip() for line in text.splitlines() if line.strip()]


def _normalize_retrievals(row: dict[str, Any]) -> list[dict[str, Any]]:
    raw = _parse_json_or_lines(_pick(row, "retrievals"))
    top_similarity = _pick(row, "top_similarity")
    threshold = _pick(row, "production_threshold")
    results: list[dict[str, Any]] = []
    for index, item in enumerate(raw):
        if isinstance(item, dict):
            normalized = dict(item)
        else:
            normalized = {"title": _text(item)}
        if index == 0 and top_similarity not in (None, ""):
            normalized.setdefault("similarity", top_similarity)
        if threshold not in (None, ""):
            normalized.setdefault("threshold", threshold)
        results.append(normalized)
    if not results and top_similarity not in (None, ""):
        results.append({"similarity": top_similarity, "threshold": threshold})
    return results


def _normalize_tools(row: dict[str, Any]) -> list[dict[str, Any]]:
    raw = _parse_json_or_lines(_pick(row, "tools"))
    results: list[dict[str, Any]] = []
    for item in raw:
        if isinstance(item, dict):
            results.append(item)
        else:
            results.append({"name": _text(item), "status": "unknown"})
    return results


def _normalize_attachments(row: dict[str, Any]) -> list[str]:
    value = _pick(row, "attachments")
    if isinstance(value, list):
        return [_text(item) for item in value if _text(item)]
    return [
        item.strip()
        for item in re.split(r"[\n,，;；]+", _text(value))
        if item.strip()
    ]


def _rows_from_file(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(f"数据文件不存在：{source}")
    if source.suffix.lower() == ".json":
        payload = json.loads(source.read_text(encoding="utf-8-sig"))
        rows = payload.get("items") or payload.get("rows") or [] if isinstance(payload, dict) else payload
        if not isinstance(rows, list):
            raise ValueError("JSON 数据必须是数组，或包含 items/rows 数组")
        return [dict(row) for row in rows if isinstance(row, dict)]
    if source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    workbook = load_workbook(source, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(iterator, ())]
    rows: list[dict[str, Any]] = []
    for raw in iterator:
        row = {
            header: raw[index] if index < len(raw) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    workbook.close()
    return rows


def _normalize_transfer_row(row: dict[str, Any], index: int) -> dict[str, Any]:
    transfer_id = _text(_pick(row, "transfer_id")) or f"manhattan-import-{index:06d}"
    return {
        "transfer_id": transfer_id,
        "conversation_id": _text(_pick(row, "conversation_id")),
        "work_order_id": _text(_pick(row, "work_order_id")),
        "event_time": _iso_datetime(_pick(row, "event_time")),
        "engineer": _text(_pick(row, "engineer")),
        "transfer_reason": _text(_pick(row, "transfer_reason")),
        "category": _text(_pick(row, "category")),
        "model": _text(_pick(row, "model")),
        "order_status": _text(_pick(row, "order_status")),
        "source": row,
    }


def _normalize_conversation_row(
    row: dict[str, Any],
    index: int,
    system: str,
) -> dict[str, Any]:
    messages = _parse_conversation_text(_pick(row, "conversation_text"))
    first_question = _text(_pick(row, "first_question")) or _first_question(messages)
    last_answer = _text(_pick(row, "last_answer")) or _last_answer(messages)
    source_id = _text(_pick(row, "source_id")) or f"{system}-import-{index:06d}"
    return {
        "source_id": source_id,
        "work_order_id": _text(_pick(row, "work_order_id")),
        "engineer": _text(_pick(row, "engineer")),
        "started_at": _iso_datetime(_pick(row, "started_at")),
        "ended_at": _iso_datetime(_pick(row, "ended_at")),
        "first_question": first_question,
        "last_answer": last_answer,
        "intent_result": _text(_pick(row, "intent_result")),
        "conversation_text": _conversation_text(messages, _pick(row, "conversation_text")),
        "messages": messages,
        "retrievals": _normalize_retrievals(row),
        "tools": _normalize_tools(row),
        "attachments": _normalize_attachments(row),
        "source": row,
    }


def import_source_file(
    path: str | Path,
    system: str,
    store: TransferAnalysisStore,
) -> dict[str, Any]:
    system_name = system.strip().lower()
    if system_name not in {"manhattan", "baixiaosheng"}:
        raise ValueError("system 仅支持 manhattan 或 baixiaosheng")
    rows = _rows_from_file(path)
    run_id = f"{system_name}-import-{uuid.uuid4().hex}"
    store.save_collection_run(run_id, system_name, "", "", "running")
    transfers = 0
    conversations = 0
    try:
        for index, row in enumerate(rows, start=1):
            if system_name == "manhattan":
                transfer = _normalize_transfer_row(row, index)
                store.upsert_transfer(transfer)
                transfers += 1
                if not _text(_pick(row, "conversation_text")):
                    continue
                detail = _normalize_conversation_row(row, index, system_name)
                detail["source_id"] = transfer["transfer_id"]
                detail["work_order_id"] = detail["work_order_id"] or transfer["work_order_id"]
                store.upsert_conversation(system_name, detail)
                conversations += 1
            else:
                detail = _normalize_conversation_row(row, index, system_name)
                store.upsert_conversation(system_name, detail)
                conversations += 1
    except Exception as exc:
        store.save_collection_run(
            run_id,
            system_name,
            "",
            "",
            "failed",
            {"source_rows": len(rows), "transfer_records": transfers, "detail_records": conversations},
            error=str(exc),
        )
        raise
    store.save_collection_run(
        run_id,
        system_name,
        "",
        "",
        "completed",
        {"source_rows": len(rows), "transfer_records": transfers, "detail_records": conversations},
    )
    return {
        "system": system_name,
        "source_file": str(Path(path)),
        "source_rows": len(rows),
        "transfer_records": transfers,
        "conversation_records": conversations,
    }


def _largest_remainder_quotas(
    groups: dict[tuple[str, str, str], list[dict[str, Any]]],
    target: int,
) -> dict[tuple[str, str, str], int]:
    total = sum(len(rows) for rows in groups.values())
    if total <= 0 or target <= 0:
        return {key: 0 for key in groups}
    exact = {key: target * len(rows) / total for key, rows in groups.items()}
    quotas = {key: min(len(groups[key]), int(math.floor(value))) for key, value in exact.items()}
    remaining = target - sum(quotas.values())
    order = sorted(
        groups,
        key=lambda key: (exact[key] - quotas[key], len(groups[key]), key),
        reverse=True,
    )
    while remaining > 0:
        progressed = False
        for key in order:
            if quotas[key] >= len(groups[key]):
                continue
            quotas[key] += 1
            remaining -= 1
            progressed = True
            if remaining == 0:
                break
        if not progressed:
            break
    return quotas


def stratified_sample(
    records: list[dict[str, Any]],
    sample_size: int = 350,
    *,
    seed: str | int = "transfer-analysis-v1",
) -> list[dict[str, Any]]:
    if sample_size <= 0 or not records:
        return []
    target = min(len(records), int(sample_size))
    randomizer = Random(str(seed))
    pool = [dict(record) for record in records]
    randomizer.shuffle(pool)

    proportional_target = min(target, round(target * 0.7))
    risk_target = min(target - proportional_target, round(target * 0.2))
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in pool:
        event_time = _parse_datetime(record.get("event_time"))
        event_date = event_time.date().isoformat() if event_time else _text(record.get("event_time"))[:10]
        key = (
            event_date,
            _text(record.get("transfer_reason")) or "未选择",
            _text(record.get("category")) or "未分类",
        )
        groups[key].append(record)
    quotas = _largest_remainder_quotas(groups, proportional_target)
    selected: list[dict[str, Any]] = []
    selected_ids: set[str] = set()
    for key in sorted(groups):
        randomizer.shuffle(groups[key])
        for record in groups[key][: quotas[key]]:
            record["_sample_bucket"] = "比例分层"
            selected.append(record)
            selected_ids.add(_text(record.get("transfer_id")))

    remaining = [record for record in pool if _text(record.get("transfer_id")) not in selected_ids]
    risk = [record for record in remaining if _text(record.get("transfer_reason")) in RISK_REASONS]
    randomizer.shuffle(risk)
    for record in risk[:risk_target]:
        record["_sample_bucket"] = "重点场景"
        selected.append(record)
        selected_ids.add(_text(record.get("transfer_id")))

    remaining = [record for record in pool if _text(record.get("transfer_id")) not in selected_ids]
    randomizer.shuffle(remaining)
    for record in remaining[: target - len(selected)]:
        record["_sample_bucket"] = "随机补充"
        selected.append(record)
    return selected[:target]


def _hours_between(left: datetime | None, right: datetime | None) -> float:
    if left is None or right is None:
        return float("inf")
    if left.tzinfo is not None and right.tzinfo is None:
        right = right.replace(tzinfo=left.tzinfo)
    if right.tzinfo is not None and left.tzinfo is None:
        left = left.replace(tzinfo=right.tzinfo)
    return abs((left - right).total_seconds()) / 3600


def associate_conversation(
    transfer: dict[str, Any],
    candidates: list[dict[str, Any]],
) -> dict[str, Any]:
    if not candidates:
        return {
            "conversation": None,
            "confidence": "低",
            "reason": "同一工单ID下没有百晓生会话。",
            "candidate_count": 0,
        }
    transfer_time = _parse_datetime(transfer.get("event_time"))
    engineer = _text(transfer.get("engineer")).lower()
    ranked: list[tuple[tuple[int, int, float, str], dict[str, Any]]] = []
    for candidate in candidates:
        candidate_engineer = _text(candidate.get("engineer")).lower()
        ended_at = _parse_datetime(candidate.get("ended_at") or candidate.get("started_at"))
        same_engineer = bool(engineer and candidate_engineer and engineer == candidate_engineer)
        is_prior = bool(transfer_time and ended_at and ended_at <= transfer_time)
        distance = _hours_between(transfer_time, ended_at)
        within_day = distance <= 24
        rank = (
            0 if same_engineer else 1,
            0 if is_prior else 1,
            distance if within_day else distance + 10000,
            _text(candidate.get("source_id")),
        )
        ranked.append((rank, candidate))
    ranked.sort(key=lambda item: item[0])
    chosen = ranked[0][1]
    chosen_rank = ranked[0][0]
    if len(candidates) == 1:
        confidence = "高"
        reason = "工单ID唯一匹配到一条百晓生会话。"
    elif chosen_rank[0] == 0 and chosen_rank[1] == 0 and chosen_rank[2] <= 6:
        confidence = "中"
        reason = "同工单存在多会话，按同一工程师且转人工前6小时内最近会话匹配。"
    else:
        confidence = "低"
        reason = "同工单存在多个候选，按时间最近记录匹配，需人工确认。"
    return {
        "conversation": chosen,
        "confidence": confidence,
        "reason": reason,
        "candidate_count": len(candidates),
    }


def _char_ngrams(value: str, size: int = 2) -> set[str]:
    normalized = re.sub(r"\s+", "", value)
    if len(normalized) < size:
        return {normalized} if normalized else set()
    return {normalized[index : index + size] for index in range(len(normalized) - size + 1)}


def _text_similarity(left: str, right: str) -> float:
    left_set = _char_ngrams(left)
    right_set = _char_ngrams(right)
    if not left_set or not right_set:
        return 0.0
    return len(left_set & right_set) / len(left_set | right_set)


def _first_manual_question(detail: dict[str, Any] | None) -> str:
    if not detail:
        return ""
    messages = detail.get("messages") or []
    question = _first_question(messages)
    return question or _text(detail.get("first_question")) or _text(detail.get("conversation_text"))


def _is_valid_question(question: str) -> bool:
    normalized = re.sub(r"[\s，。！？!?、,.]+", "", question)
    return bool(normalized and normalized not in INVALID_QUESTIONS and len(normalized) >= 3)


def _is_clear_intent(question: str) -> bool:
    normalized = re.sub(r"\s+", "", question)
    if not _is_valid_question(question):
        return False
    if normalized in VAGUE_QUESTIONS:
        return False
    return len(normalized) >= 8


def _engineer_issue(question: str, manual_question: str) -> tuple[bool, list[str]]:
    tags: list[str] = []
    if not _is_valid_question(question):
        tags.extend(["无效问", "工程师提问质量差"])
    elif not _is_clear_intent(question):
        tags.extend(["意图不明确", "工程师提问质量差"])
    if manual_question and question:
        similarity = _text_similarity(question, manual_question)
        if len(manual_question) >= 8 and similarity < 0.15:
            tags.extend(["前后问题不一致", "转人工后才补充信息"])
        elif len(manual_question) > len(question) * 1.8 and similarity < 0.35:
            tags.extend(["缺少必要信息", "转人工后才补充信息"])
    return bool(tags), list(dict.fromkeys(tags))


def _retrieval_summary(retrievals: Any) -> tuple[str, float | None, float | None]:
    if not isinstance(retrievals, list):
        return _text(retrievals), None, None
    lines: list[str] = []
    top_score: float | None = None
    threshold: float | None = None
    for index, item in enumerate(retrievals):
        if not isinstance(item, dict):
            lines.append(_text(item))
            continue
        title = _text(
            item.get("title")
            or item.get("knowledgeTitle")
            or item.get("name")
            or item.get("knowledge_id")
            or item.get("id")
        )
        score_value = (
            item.get("similarity")
            if item.get("similarity") is not None
            else item.get("score", item.get("rerank_score"))
        )
        try:
            score = float(score_value) if score_value not in (None, "") else None
        except (TypeError, ValueError):
            score = None
        threshold_value = item.get("threshold", item.get("score_threshold"))
        try:
            current_threshold = (
                float(threshold_value) if threshold_value not in (None, "") else None
            )
        except (TypeError, ValueError):
            current_threshold = None
        if index == 0:
            top_score = score
        if current_threshold is not None:
            threshold = current_threshold
        score_text = f"（{score:.4f}）" if score is not None else ""
        selected = item.get("selected")
        selected_text = "，已采用" if selected in {True, "true", "是", 1} else ""
        if title or score_text:
            lines.append(f"{title or '未命名知识'}{score_text}{selected_text}")
    return "\n".join(lines), top_score, threshold


def _standard_summary(matches: list[tuple[StandardCatalogItem, float]]) -> str:
    return "\n".join(
        f"{item.standard_id or item.title}｜{item.title}｜检索分 {score:.3f}"
        for item, score in matches
    )


def _tool_name(tool: dict[str, Any]) -> str:
    return _text(tool.get("name") or tool.get("tool_name") or tool.get("tool") or tool.get("type"))


def _tool_status(tool: dict[str, Any]) -> str:
    return _text(tool.get("status") or tool.get("state") or tool.get("result_status")).lower()


def _tool_result(tool: dict[str, Any]) -> str:
    value = tool.get("result") or tool.get("output") or tool.get("response") or tool.get("error")
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)[:1500]
    return _text(value, 1500)


def _assess_capability(
    question: str,
    tools: list[dict[str, Any]],
    attachments: list[Any],
    knowledge_matches: list[tuple[StandardCatalogItem, float]],
    registry: CapabilityRegistry,
) -> dict[str, Any]:
    manual_operation = any(marker in question for marker in MANUAL_OPERATION_MARKERS)
    image_request = bool(attachments) and any(marker in question for marker in IMAGE_REQUEST_MARKERS)
    tool_candidates = registry.tools_for_question(question)
    called_names = [_tool_name(tool) for tool in tools if _tool_name(tool)]
    called = "是" if called_names else "否"
    attribution: list[str] = []
    result_text = "；".join(
        filter(None, (_tool_result(tool) for tool in tools))
    )[:2000]

    successful_capability = None
    failed_capability = None
    for tool in tools:
        capability = registry.tool_for_name(_tool_name(tool))
        if capability is None:
            continue
        status = _tool_status(tool)
        if status in SUCCESS_TOOL_STATUSES:
            successful_capability = capability
        elif status in FAILED_TOOL_STATUSES:
            failed_capability = capability

    required_tool = tool_candidates[0].name if tool_candidates else ""
    if successful_capability:
        required_tool = successful_capability.name
        return {
            "solvable": "是",
            "solution_method": "调用工具",
            "required_tool": required_tool,
            "tool_called": "是",
            "tool_result": result_text or "工具调用成功",
            "tool_attribution": "",
            "uncertain_tool_scope": False,
        }
    if failed_capability:
        attribution.append("工具调用失败")
        return {
            "solvable": "是",
            "solution_method": "调用工具",
            "required_tool": failed_capability.name,
            "tool_called": "是",
            "tool_result": result_text or "工具调用失败",
            "tool_attribution": "；".join(attribution),
            "uncertain_tool_scope": False,
        }
    if manual_operation:
        return {
            "solvable": "否",
            "solution_method": "不可解决",
            "required_tool": "",
            "tool_called": called,
            "tool_result": result_text,
            "tool_attribution": "百晓生能力不支持",
            "uncertain_tool_scope": False,
        }
    if tool_candidates:
        capability = tool_candidates[0]
        if not capability.confirmed_scope:
            return {
                "solvable": "不确定",
                "solution_method": "调用工具",
                "required_tool": capability.name,
                "tool_called": called,
                "tool_result": result_text,
                "tool_attribution": "工具能力不覆盖",
                "uncertain_tool_scope": True,
            }
        if capability.requires_attachment and not attachments:
            attribution.append("缺少工具必需输入")
            return {
                "solvable": "是",
                "solution_method": "追问后回答",
                "required_tool": capability.name,
                "tool_called": called,
                "tool_result": result_text,
                "tool_attribution": "；".join(attribution),
                "uncertain_tool_scope": False,
            }
        attribution.append("应调用工具未调用")
        return {
            "solvable": "是",
            "solution_method": "调用工具",
            "required_tool": capability.name,
            "tool_called": called,
            "tool_result": result_text,
            "tool_attribution": "；".join(attribution),
            "uncertain_tool_scope": False,
        }
    if image_request and not registry.supports_general_multimodal:
        return {
            "solvable": "否",
            "solution_method": "不可解决",
            "required_tool": "",
            "tool_called": called,
            "tool_result": result_text,
            "tool_attribution": "百晓生能力不支持",
            "uncertain_tool_scope": False,
        }
    if knowledge_matches:
        return {
            "solvable": "是",
            "solution_method": "直接回答",
            "required_tool": "",
            "tool_called": called,
            "tool_result": result_text,
            "tool_attribution": "",
            "uncertain_tool_scope": False,
        }
    return {
        "solvable": "否",
        "solution_method": "不可解决",
        "required_tool": "",
        "tool_called": called,
        "tool_result": result_text,
        "tool_attribution": "",
        "uncertain_tool_scope": False,
    }


def _diagnostic_remark(
    tags: list[str],
    fact: str,
    suggestion: str,
) -> str:
    normalized_tags = list(dict.fromkeys(tag for tag in tags if tag))
    tag_text = "、".join(normalized_tags) or "待人工确认"
    return f"【诊断】{tag_text}；【事实】{fact or '证据不足'}；【建议】{suggestion or '人工复核后确定优化动作'}"


def _rule_annotation(
    transfer: dict[str, Any],
    bxs: dict[str, Any] | None,
    manual: dict[str, Any] | None,
    link: dict[str, Any],
    matches: list[tuple[StandardCatalogItem, float]],
    capability: dict[str, Any],
) -> dict[str, Any]:
    question = _text((bxs or {}).get("first_question"))
    answer = _text((bxs or {}).get("last_answer"))
    manual_question = _first_manual_question(manual)
    true_intent = manual_question or question
    valid = _is_valid_question(question)
    clear = _is_clear_intent(question)
    engineer_issue, engineer_tags = _engineer_issue(question, manual_question)
    retrieval_text, top_score, threshold = _retrieval_summary((bxs or {}).get("retrievals") or [])
    tags = list(engineer_tags)
    original_reason = _text(transfer.get("transfer_reason"))
    corrected_reason = original_reason if original_reason in TRANSFER_REASON_OPTIONS else "其他"
    responsibility = "算法"
    suggest_knowledge = "否"

    if engineer_issue:
        corrected_reason = "其他"
        responsibility = "运营"
    elif capability["solvable"] == "否" and capability["tool_attribution"]:
        corrected_reason = "其他"
        tags.append(capability["tool_attribution"])
        responsibility = "产品/系统"
    elif not matches and valid:
        corrected_reason = "该问题没有相关知识"
        tags.append("知识缺失可新增")
        responsibility = "知识库"
        suggest_knowledge = "是"
    elif matches and not retrieval_text:
        corrected_reason = "答非所问"
        tags.extend(["无召回", "正确知识未召回"])
        responsibility = "算法"
    elif (
        top_score is not None
        and threshold is not None
        and top_score < threshold
        and answer
    ):
        corrected_reason = "答非所问"
        tags.append("低相似度仍输出")
        responsibility = "算法"
    elif capability["tool_attribution"]:
        tags.append(capability["tool_attribution"])
        responsibility = "算法" if "调用" in capability["tool_attribution"] else "产品/系统"

    if original_reason == "更信任人工" and capability["solvable"] == "是" and answer:
        corrected_reason = "更信任人工"
        responsibility = "运营"

    fact = (
        f"首轮问题：{question or '空'}；转人工后问题：{manual_question or '未取得'}；"
        f"知识主表匹配{len(matches)}条；百晓生召回{len((bxs or {}).get('retrievals') or [])}条。"
    )
    suggestions = {
        "算法": "检查意图识别、召回阈值、回答生成或工具路由。",
        "知识库": "评估补充或更新相关知识，并补充检索关键词。",
        "运营": "优化工程师提问指引，要求在转人工前补充对象、现象和必要信息。",
        "产品/系统": "确认能力边界、工具覆盖范围或后台人工操作流程。",
    }
    confidence = {
        "意图是否明确": 0.88 if clear else 0.72,
        "真实意图": 0.86 if manual_question else 0.7,
        "转人工原因(校正)": 0.84 if not engineer_issue else 0.76,
        "大模型是否可以解决": 0.82 if capability["solvable"] != "不确定" else 0.55,
        "是否有效问": 0.9,
    }
    return {
        "意图是否明确": "是" if clear else "否",
        "真实意图": true_intent,
        "转人工原因(校正)": corrected_reason,
        "备注": _diagnostic_remark(tags, fact, suggestions.get(responsibility, "")),
        "大模型是否可以解决": capability["solvable"],
        "是否有效问": "是" if valid else "否",
        "是否建议补充知识": suggest_knowledge,
        "建议优化责任方": responsibility,
        "解决方式": capability["solution_method"],
        "所需工具": capability["required_tool"],
        "工具是否调用": capability["tool_called"],
        "工具调用结果": capability["tool_result"],
        "工具归因标签": capability["tool_attribution"],
        "confidence": confidence,
        "evidence": [
            item
            for item in (
                f"百晓生首轮：{question}" if question else "",
                f"转人工后首轮：{manual_question}" if manual_question else "",
                f"召回：{retrieval_text}" if retrieval_text else "",
            )
            if item
        ],
        "rule_flags": {
            "engineer_issue": engineer_issue,
            "uncertain_tool_scope": capability["uncertain_tool_scope"],
            "link_confidence": link["confidence"],
        },
    }


def _transfer_prompt(
    transfer: dict[str, Any],
    bxs: dict[str, Any] | None,
    manual: dict[str, Any] | None,
    link: dict[str, Any],
    matches: list[tuple[StandardCatalogItem, float]],
    capability: dict[str, Any],
    rule_result: dict[str, Any],
    retry_reason: str = "",
) -> str:
    retry = f"\n上次输出错误：{retry_reason}\n请修正后只输出JSON。" if retry_reason else ""
    payload = {
        "转人工记录": {
            "工单ID": transfer.get("work_order_id"),
            "工程师选择原因": transfer.get("transfer_reason"),
            "类目": transfer.get("category"),
            "机型": transfer.get("model"),
            "订单状态": transfer.get("order_status"),
        },
        "关联": {
            "置信度": link.get("confidence"),
            "依据": link.get("reason"),
            "候选数": link.get("candidate_count"),
        },
        "百晓生": {
            "首轮问题": (bxs or {}).get("first_question"),
            "最后回答": (bxs or {}).get("last_answer"),
            "完整会话": (bxs or {}).get("conversation_text"),
            "意图识别": (bxs or {}).get("intent_result"),
            "召回": (bxs or {}).get("retrievals"),
            "工具调用": (bxs or {}).get("tools"),
            "附件数量": len((bxs or {}).get("attachments") or []),
        },
        "转人工后": {
            "完整会话": (manual or {}).get("conversation_text"),
            "消息": (manual or {}).get("messages"),
        },
        "知识主表检索": [
            {
                "id": item.standard_id,
                "标题": item.title,
                "范围": item.scope,
                "参考话术": item.response_snippet,
                "检索分": round(score, 3),
            }
            for item, score in matches
        ],
        "百晓生能力判断": capability,
        "规则初判": {
            key: value
            for key, value in rule_result.items()
            if key not in {"confidence", "evidence", "rule_flags"}
        },
        "能力注册表": DEFAULT_CAPABILITY_REGISTRY.to_dict(),
    }
    return f"""
你是百晓生转人工会话质检员。请结合百晓生会话、转人工后会话、实际召回、
当前知识主表和百晓生能力边界分析转人工原因。

必须遵守：
1. 百晓生不支持通用多模态，不得因你能理解图片而扩大百晓生能力。
2. 内存硬盘品牌识别工具属于已确认能力；笔记本识别工具只有实际成功调用
   或明确能力说明时才能确认覆盖。
3. 工程师前后问题不一致、信息未提供、不会描述、转人工后才补充信息，
   校正原因必须为“其他”，并在备注中写明工程师侧诊断。
4. 诊断标签不要单独输出列，全部写进备注，格式：
   【诊断】标签1、标签2；【事实】证据；【建议】动作。
5. 转人工原因(校正)只能是：
   {json.dumps(list(TRANSFER_REASON_OPTIONS), ensure_ascii=False)}
6. 大模型是否可以解决只能是“是”“否”“不确定”。
7. 证据不足时不要编造，降低对应字段置信度。

只输出以下JSON：
{{
  "意图是否明确": "是或否",
  "真实意图": "对象＋现象＋期望处理",
  "转人工原因(校正)": "六个原因之一",
  "备注": "【诊断】...；【事实】...；【建议】...",
  "大模型是否可以解决": "是/否/不确定",
  "是否有效问": "是或否",
  "是否建议补充知识": "是/否/待确认",
  "建议优化责任方": "算法/知识库/运营/产品系统/多方",
  "解决方式": "直接回答/追问后回答/调用工具/不可解决",
  "所需工具": "",
  "工具是否调用": "是/否/不适用/未知",
  "工具调用结果": "",
  "工具归因标签": "",
  "confidence": {{
    "意图是否明确": 0.0,
    "真实意图": 0.0,
    "转人工原因(校正)": 0.0,
    "大模型是否可以解决": 0.0,
    "是否有效问": 0.0
  }},
  "evidence": ["引用的消息或召回证据"]
}}

输入：
{json.dumps(payload, ensure_ascii=False, default=str)}
{retry}
""".strip()


def _content_from_response(response: dict[str, Any]) -> str:
    choices = response.get("choices")
    if not isinstance(choices, list) or not choices:
        raise MimoError("MiMo 返回缺少 choices")
    message = choices[0].get("message") or {}
    content = message.get("content")
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        return "".join(
            _text(item.get("text"))
            for item in content
            if isinstance(item, dict)
        )
    raise MimoError("MiMo 返回缺少 message.content")


def _strip_json_fence(value: str) -> str:
    text = value.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)
    return text.strip()


def _validate_model_annotation(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise MimoError("转人工分析输出必须是JSON对象")
    reason = _text(value.get("转人工原因(校正)"))
    if reason not in TRANSFER_REASON_OPTIONS:
        raise MimoError("转人工原因(校正)不在允许枚举中")
    for key in ("意图是否明确", "是否有效问"):
        if _text(value.get(key)) not in {"是", "否"}:
            raise MimoError(f"{key}必须是是或否")
    if _text(value.get("大模型是否可以解决")) not in {"是", "否", "不确定"}:
        raise MimoError("大模型是否可以解决必须是是、否或不确定")
    remark = _text(value.get("备注"), 3000)
    if "【诊断】" not in remark or "【事实】" not in remark or "【建议】" not in remark:
        raise MimoError("备注必须包含诊断、事实和建议")
    confidence_raw = value.get("confidence")
    if not isinstance(confidence_raw, dict):
        raise MimoError("confidence必须是对象")
    confidence: dict[str, float] = {}
    for field_name in (
        "意图是否明确",
        "真实意图",
        "转人工原因(校正)",
        "大模型是否可以解决",
        "是否有效问",
    ):
        try:
            score = float(confidence_raw.get(field_name))
        except (TypeError, ValueError) as exc:
            raise MimoError(f"{field_name}置信度必须是0到1") from exc
        if not 0 <= score <= 1:
            raise MimoError(f"{field_name}置信度必须是0到1")
        confidence[field_name] = round(score, 3)
    evidence = value.get("evidence")
    if not isinstance(evidence, list):
        evidence = []
    return {
        "意图是否明确": _text(value.get("意图是否明确")),
        "真实意图": _text(value.get("真实意图"), 1500),
        "转人工原因(校正)": reason,
        "备注": remark,
        "大模型是否可以解决": _text(value.get("大模型是否可以解决")),
        "是否有效问": _text(value.get("是否有效问")),
        "是否建议补充知识": _text(value.get("是否建议补充知识")) or "待确认",
        "建议优化责任方": _text(value.get("建议优化责任方")) or "多方",
        "解决方式": _text(value.get("解决方式")) or "不可解决",
        "所需工具": _text(value.get("所需工具")),
        "工具是否调用": _text(value.get("工具是否调用")) or "未知",
        "工具调用结果": _text(value.get("工具调用结果"), 2000),
        "工具归因标签": _text(value.get("工具归因标签")),
        "confidence": confidence,
        "evidence": [_text(item, 1000) for item in evidence if _text(item)],
    }


class TransferMimoClient(MimoClient):
    def analyze_transfer(
        self,
        transfer: dict[str, Any],
        bxs: dict[str, Any] | None,
        manual: dict[str, Any] | None,
        link: dict[str, Any],
        matches: list[tuple[StandardCatalogItem, float]],
        capability: dict[str, Any],
        rule_result: dict[str, Any],
    ) -> MimoLabelResult:
        validation_error = ""
        last_response: dict[str, Any] = {}
        for attempt in range(2):
            prompt = _transfer_prompt(
                transfer,
                bxs,
                manual,
                link,
                matches,
                capability,
                rule_result,
                validation_error,
            )
            payload = {
                "model": self.config.model,
                "messages": [
                    {
                        "role": "system",
                        "content": "你是严谨的百晓生转人工会话质检员。",
                    },
                    {
                        "role": "user",
                        "content": [{"type": "text", "text": prompt}],
                    },
                ],
                "temperature": 0.1,
                "response_format": {"type": "json_object"},
            }
            request_audit = {
                "endpoint": self.config.chat_completions_url(),
                "model": self.config.model,
                "prompt_version": TRANSFER_ANALYSIS_PROMPT_VERSION,
                "attempt": attempt + 1,
                "transfer_id": transfer.get("transfer_id"),
                "work_order_id": transfer.get("work_order_id"),
            }
            try:
                response = self._post(payload)
                last_response = response
                candidate = _validate_model_annotation(
                    json.loads(_strip_json_fence(_content_from_response(response)))
                )
                return MimoLabelResult(
                    candidate=candidate,
                    request_audit=request_audit,
                    response_audit=response,
                )
            except (json.JSONDecodeError, MimoError) as exc:
                validation_error = str(exc)
                if attempt == 1:
                    raise MimoError(
                        f"MiMo转人工分析JSON校验失败（已重试一次）：{validation_error}"
                    ) from exc
            except Exception as exc:
                raise MimoError(f"MiMo转人工分析调用失败：{exc}") from exc
        raise MimoError(f"MiMo转人工分析未产生有效结果：{last_response}")


def _apply_guardrails(
    model_result: dict[str, Any],
    rule_result: dict[str, Any],
    capability: dict[str, Any],
) -> tuple[dict[str, Any], bool]:
    guarded = dict(model_result)
    flags = rule_result.get("rule_flags") or {}
    conflict = False
    if flags.get("engineer_issue") and guarded.get("转人工原因(校正)") != "其他":
        guarded["转人工原因(校正)"] = "其他"
        conflict = True
    if flags.get("engineer_issue"):
        rule_remark = _text(rule_result.get("备注"))
        rule_diagnosis = re.search(r"【诊断】([^；]+)", rule_remark)
        if rule_diagnosis:
            for tag in rule_diagnosis.group(1).split("、"):
                tag = tag.strip()
                if tag and tag not in guarded.get("备注", ""):
                    guarded["备注"] = guarded["备注"].replace(
                        "【诊断】",
                        f"【诊断】{tag}、",
                        1,
                    )
    if capability["solvable"] == "否" and guarded.get("大模型是否可以解决") == "是":
        guarded["大模型是否可以解决"] = "否"
        guarded["解决方式"] = "不可解决"
        conflict = True
    if capability["uncertain_tool_scope"]:
        guarded["大模型是否可以解决"] = "不确定"
        guarded["解决方式"] = "调用工具"
        conflict = True
    if capability["required_tool"]:
        guarded["所需工具"] = capability["required_tool"]
    if capability["tool_attribution"]:
        guarded["工具归因标签"] = capability["tool_attribution"]
        if capability["tool_attribution"] not in guarded.get("备注", ""):
            guarded["备注"] = guarded["备注"].replace(
                "【诊断】",
                f"【诊断】{capability['tool_attribution']}、",
                1,
            )
    return guarded, conflict


def _build_analysis_row(
    transfer: dict[str, Any],
    bxs: dict[str, Any] | None,
    manual: dict[str, Any] | None,
    link: dict[str, Any],
    matches: list[tuple[StandardCatalogItem, float]],
    annotation: dict[str, Any],
    *,
    model_name: str,
    model_status: str,
    needs_review: bool,
) -> dict[str, Any]:
    retrieval_text, top_score, threshold = _retrieval_summary((bxs or {}).get("retrievals") or [])
    confidence = annotation.get("confidence") or {}
    evidence = annotation.get("evidence") or []
    source = (bxs or {}).get("source") or {}
    row = {
        "会话id": _text((bxs or {}).get("source_id") or transfer.get("conversation_id")),
        "工单ID": _text(transfer.get("work_order_id")),
        "问题": _text((bxs or {}).get("first_question")),
        "大模型回答": _text(source.get("大模型回答") or (bxs or {}).get("last_answer")),
        "意图是否明确": annotation.get("意图是否明确", ""),
        "真实意图": annotation.get("真实意图", ""),
        "回答": _text((bxs or {}).get("last_answer")),
        "转人工原因": _text(transfer.get("transfer_reason")),
        "转人工原因(校正)": annotation.get("转人工原因(校正)", ""),
        "备注": annotation.get("备注", ""),
        "大模型是否可以解决": annotation.get("大模型是否可以解决", ""),
        "是否有效问": annotation.get("是否有效问", ""),
        "类目": _text(transfer.get("category")),
        "机型": _text(transfer.get("model")),
        "订单所处状态": _text(transfer.get("order_status")),
        "意图识别结果": _text((bxs or {}).get("intent_result")),
        "曼哈顿转人工ID": _text(transfer.get("transfer_id")),
        "关联置信度": link.get("confidence", ""),
        "关联依据": link.get("reason", ""),
        "百晓生完整会话": _text((bxs or {}).get("conversation_text"), 30000),
        "转人工后完整会话": _text((manual or {}).get("conversation_text"), 30000),
        "召回知识": retrieval_text or _standard_summary(matches),
        "Top相似度": top_score if top_score is not None else "",
        "生产阈值": threshold if threshold is not None else "",
        "是否建议补充知识": annotation.get("是否建议补充知识", ""),
        "建议优化责任方": annotation.get("建议优化责任方", ""),
        "解决方式": annotation.get("解决方式", ""),
        "所需工具": annotation.get("所需工具", ""),
        "工具是否调用": annotation.get("工具是否调用", ""),
        "工具调用结果": annotation.get("工具调用结果", ""),
        "工具归因标签": annotation.get("工具归因标签", ""),
        "字段置信度": json.dumps(confidence, ensure_ascii=False),
        "证据引用": "\n".join(_text(item) for item in evidence if _text(item)),
        "是否需要人工复核": "是" if needs_review else "否",
        "模型名称": model_name,
        "Prompt版本": TRANSFER_ANALYSIS_PROMPT_VERSION,
        "模型状态": model_status,
        "审核状态": "待审核" if needs_review else "无需复核",
        "审核人": "",
        "审核时间": "",
    }
    return {column: row.get(column, "") for column in ANALYSIS_COLUMNS}


def _week_dates(week_start: str | date | datetime) -> tuple[str, str]:
    if isinstance(week_start, datetime):
        start = week_start.date()
    elif isinstance(week_start, date):
        start = week_start
    else:
        start = date.fromisoformat(str(week_start))
    return start.isoformat(), (start + timedelta(days=7)).isoformat()


def _notify(callback: ProgressCallback | None, **payload: Any) -> None:
    if callback:
        callback(payload)


def run_weekly_analysis(
    store: TransferAnalysisStore,
    week_start: str | date | datetime,
    standards_path: str | Path,
    output_dir: str | Path,
    *,
    sample_size: int = 350,
    use_mimo: bool = True,
    manhattan_profile: str | Path | None = None,
    baixiaosheng_profile: str | Path | None = None,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    start_date, end_date = _week_dates(week_start)
    transfers = store.list_transfers(start_date, end_date)
    if not transfers:
        raise ValueError(f"{start_date} 至 {end_date} 没有曼哈顿转人工列表数据")
    samples = stratified_sample(transfers, sample_size, seed=start_date)
    store.replace_samples(
        start_date,
        [
            (_text(record.get("transfer_id")), _text(record.get("_sample_bucket")))
            for record in samples
        ],
    )
    _notify(progress_callback, stage="sample", current=len(samples), total=len(samples))

    if manhattan_profile:
        collect_with_endpoint_profile(
            manhattan_profile,
            store,
            start_date=start_date,
            end_date=end_date,
            transfer_ids=[record["transfer_id"] for record in samples],
        )
    if baixiaosheng_profile:
        collect_with_endpoint_profile(
            baixiaosheng_profile,
            store,
            start_date=start_date,
            end_date=end_date,
            work_order_ids=[record["work_order_id"] for record in samples],
        )

    catalog = load_standard_catalog(standards_path)
    client = TransferMimoClient.from_env() if use_mimo else None
    completed = 0
    fallback_count = 0
    for index, transfer in enumerate(samples, start=1):
        candidates = store.conversations_for_work_order(
            "baixiaosheng",
            _text(transfer.get("work_order_id")),
        )
        link = associate_conversation(transfer, candidates)
        bxs = link["conversation"]
        manual = store.get_conversation("manhattan", _text(transfer.get("transfer_id")))
        if manual is None and transfer.get("conversation_id"):
            manual = store.get_conversation("manhattan", _text(transfer.get("conversation_id")))
        store.save_link(
            _text(transfer.get("transfer_id")),
            _text((bxs or {}).get("source_id")),
            link["confidence"],
            link["reason"],
            link["candidate_count"],
        )

        query = {
            "聊天内容": _text((bxs or {}).get("conversation_text")),
            "核心问题": _text((bxs or {}).get("first_question")),
            "产品类型": _text(transfer.get("category")),
        }
        matches = retrieve_topic_signal_matches(query, catalog, top_k=5)
        capability = _assess_capability(
            _text((bxs or {}).get("first_question")),
            list((bxs or {}).get("tools") or []),
            list((bxs or {}).get("attachments") or []),
            matches,
            DEFAULT_CAPABILITY_REGISTRY,
        )
        rule_result = _rule_annotation(
            transfer,
            bxs,
            manual,
            link,
            matches,
            capability,
        )
        annotation = rule_result
        model_name = "transfer-analysis-rule-v1"
        model_status = "rule"
        error = ""
        conflict = False
        if client is not None:
            model_name = client.config.model
            try:
                result = client.analyze_transfer(
                    transfer,
                    bxs,
                    manual,
                    link,
                    matches,
                    capability,
                    rule_result,
                )
                annotation, conflict = _apply_guardrails(
                    result.candidate,
                    rule_result,
                    capability,
                )
                model_status = "completed"
            except Exception as exc:
                annotation = rule_result
                model_status = "fallback"
                error = str(exc)
                fallback_count += 1
        confidence = annotation.get("confidence") or {}
        low_confidence = any(float(value) < 0.8 for value in confidence.values())
        needs_review = any(
            (
                low_confidence,
                link["confidence"] == "低",
                annotation.get("转人工原因(校正)") != transfer.get("transfer_reason"),
                annotation.get("大模型是否可以解决") == "不确定",
                capability["uncertain_tool_scope"],
                conflict,
                model_status == "fallback",
            )
        )
        row = _build_analysis_row(
            transfer,
            bxs,
            manual,
            link,
            matches,
            annotation,
            model_name=model_name,
            model_status=model_status,
            needs_review=needs_review,
        )
        store.save_annotation(
            start_date,
            _text(transfer.get("transfer_id")),
            row,
            confidence,
            annotation.get("evidence") or [],
            status="completed" if model_status != "fallback" else "fallback",
            model_name=model_name,
            prompt_version=TRANSFER_ANALYSIS_PROMPT_VERSION,
            needs_review=needs_review,
            error=error,
        )
        completed += 1
        _notify(
            progress_callback,
            stage="annotation",
            current=index,
            total=len(samples),
            transfer_id=transfer.get("transfer_id"),
        )

    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    report_path = output_root / f"转人工分析周报_{start_date}.xlsx"
    report_summary = build_weekly_report(store, start_date, report_path)
    summary = {
        "week_start": start_date,
        "week_end": end_date,
        "source_records": len(transfers),
        "sample_records": len(samples),
        "annotation_records": completed,
        "fallback_records": fallback_count,
        "review_records": report_summary["review_records"],
        "report_file": str(report_path),
        "audit_db": str(store.path),
    }
    summary_path = output_root / f"转人工分析周报_{start_date}.json"
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    summary["summary_file"] = str(summary_path)
    return summary


def _statistics_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    total = len(rows)
    output: list[dict[str, Any]] = []

    def add_distribution(kind: str, field: str) -> None:
        counts = Counter(_text(row.get(field)) or "空" for row in rows)
        for item, count in counts.most_common():
            output.append(
                {
                    "统计类型": kind,
                    "项目": item,
                    "数量": count,
                    "占比": round(count / total, 4) if total else 0,
                }
            )

    output.extend(
        [
            {"统计类型": "总体", "项目": "样本量", "数量": total, "占比": 1 if total else 0},
            {
                "统计类型": "总体",
                "项目": "需要人工复核",
                "数量": sum(row.get("是否需要人工复核") == "是" for row in rows),
                "占比": round(
                    sum(row.get("是否需要人工复核") == "是" for row in rows) / total,
                    4,
                )
                if total
                else 0,
            },
            {
                "统计类型": "总体",
                "项目": "原因发生校正",
                "数量": sum(
                    _text(row.get("转人工原因"))
                    != _text(row.get("转人工原因(校正)"))
                    for row in rows
                ),
                "占比": round(
                    sum(
                        _text(row.get("转人工原因"))
                        != _text(row.get("转人工原因(校正)"))
                        for row in rows
                    )
                    / total,
                    4,
                )
                if total
                else 0,
            },
        ]
    )
    add_distribution("校正原因", "转人工原因(校正)")
    add_distribution("责任方", "建议优化责任方")
    add_distribution("可解决性", "大模型是否可以解决")
    add_distribution("类目", "类目")
    return output


def _owner_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        owner = _text(row.get("建议优化责任方")) or "多方"
        remark = _text(row.get("备注"))
        match = re.search(r"【诊断】([^；]+)", remark)
        issue = (
            match.group(1)
            if match
            else _text(row.get("转人工原因(校正)"))
        )
        grouped[(owner, issue)].append(row)
    suggestions = {
        "算法": "检查意图识别、召回阈值、生成回答和工具路由。",
        "知识库": "补充或更新知识，并增加真实提问表达作为检索关键词。",
        "运营": "优化工程师提问培训和转人工原因选择指引。",
        "产品/系统": "确认能力边界、工具稳定性和必须人工操作的流程。",
        "多方": "由算法、知识库和运营联合复盘代表性会话。",
    }
    output: list[dict[str, Any]] = []
    for (owner, issue), values in sorted(
        grouped.items(),
        key=lambda item: (-len(item[1]), item[0]),
    ):
        output.append(
            {
                "建议优化责任方": owner,
                "问题类型": issue,
                "数量": len(values),
                "示例工单ID": "\n".join(
                    dict.fromkeys(_text(row.get("工单ID")) for row in values if row.get("工单ID"))
                )[:1000],
                "建议动作": suggestions.get(owner, suggestions["多方"]),
            }
        )
    return output


def build_weekly_report(
    store: TransferAnalysisStore,
    week_start: str,
    output_path: str | Path,
) -> dict[str, Any]:
    rows = store.list_annotation_rows(week_start)
    review_rows = [row for row in rows if row.get("是否需要人工复核") == "是"]
    bad_cases = [
        row
        for row in rows
        if row.get("是否需要人工复核") == "是"
        or _text(row.get("转人工原因")) != _text(row.get("转人工原因(校正)"))
    ]
    knowledge_gaps = [
        row for row in rows if _text(row.get("是否建议补充知识")) == "是"
    ]
    retrieval_rows = [
        row
        for row in rows
        if _text(row.get("召回知识")) or row.get("Top相似度") not in (None, "")
    ]
    tool_rows = [
        row
        for row in rows
        if _text(row.get("所需工具")) or _text(row.get("工具归因标签"))
    ]
    statistics = _statistics_rows(rows)
    owners = _owner_rows(bad_cases)
    write_rows_to_workbook(
        {
            "转人工分析明细": (ANALYSIS_COLUMNS, rows),
            "人工复核队列": (ANALYSIS_COLUMNS, review_rows),
            "badcase清单": (BAD_CASE_COLUMNS, bad_cases),
            "知识补充候选": (KNOWLEDGE_GAP_COLUMNS, knowledge_gaps),
            "召回质量分析": (RETRIEVAL_COLUMNS, retrieval_rows),
            "工具调用问题": (TOOL_ISSUE_COLUMNS, tool_rows),
            "周度统计": (["统计类型", "项目", "数量", "占比"], statistics),
            "责任方优化清单": (OWNER_COLUMNS, owners),
        },
        output_path,
    )
    return {
        "week_start": week_start,
        "total_records": len(rows),
        "review_records": len(review_rows),
        "badcase_records": len(bad_cases),
        "knowledge_gap_records": len(knowledge_gaps),
        "tool_issue_records": len(tool_rows),
        "output_file": str(Path(output_path)),
    }
