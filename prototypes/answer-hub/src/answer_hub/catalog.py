from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import hashlib
import json
import re

from openpyxl import load_workbook


ACTIVE_STANDARD_STATUSES = {
    "active",
    "published",
    "生效",
    "生效中",
    "已发布",
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_keywords(value: Any) -> list[str]:
    text = _clean(value)
    if not text:
        return []
    parts = re.split(r"[,\n，;；/|、\s]+", text)
    return [part.strip() for part in parts if part.strip()]


def _expand_keyword_aliases(keywords: list[str]) -> list[str]:
    expanded: list[str] = []
    for keyword in keywords:
        if not keyword:
            continue
        expanded.append(keyword)
        if keyword.endswith("更换") and len(keyword) > 2:
            expanded.append(keyword[:-2])
        if "后摄" in keyword:
            expanded.append(keyword.replace("后摄", "后置摄像头"))
            if keyword.endswith("更换") and len(keyword) > 2:
                expanded.append(keyword[:-2].replace("后摄", "后置摄像头"))
        if "前摄" in keyword:
            expanded.append(keyword.replace("前摄", "前置摄像头"))
            if keyword.endswith("更换") and len(keyword) > 2:
                expanded.append(keyword[:-2].replace("前摄", "前置摄像头"))
    return list(dict.fromkeys(expanded))


def _pick(data: dict[str, Any], aliases: list[str]) -> str:
    for key in aliases:
        value = data.get(key)
        if value not in (None, ""):
            return _clean(value)
    return ""


def _path_categories(value: str) -> tuple[str, str]:
    parts = [
        re.sub(r"[*（(][^）)]*[）)]", "", part).strip()
        for part in re.findall(r"【([^】]+)】", value)
    ]
    parts = [part for part in parts if part]
    return (parts[0] if parts else "", parts[1] if len(parts) > 1 else "")


def is_active_standard(status: str) -> bool:
    return _clean(status).lower() in ACTIVE_STANDARD_STATUSES


def _catalog_scope(data: dict[str, Any]) -> str:
    raw_scope = _pick(data, ["scope", "适用范围", "适用场景"])
    product_scope = _pick(data, ["适用类目", "产品类型", "产品品类"])
    if product_scope in {"平板", "平板电脑"}:
        product_scope = "平板电脑"
    if not product_scope:
        return raw_scope
    if raw_scope.startswith(product_scope):
        return raw_scope
    if not raw_scope or raw_scope == "通用":
        return f"{product_scope}-通用"
    return f"{product_scope}-{raw_scope}"


def _stable_catalog_standard_id(data: dict[str, Any]) -> str:
    explicit_id = _pick(data, ["standard_id", "标准ID", "ID", "id"])
    if explicit_id:
        return explicit_id
    if _pick(data, ["知识来源", "source_type"]) != "总部标准":
        return ""
    product_scope = _pick(data, ["适用类目", "产品类型", "产品品类"])
    prefix = "CZ-HQ-TABLET" if product_scope in {"平板", "平板电脑"} else "CZ-HQ"
    identity = "|".join(
        [
            _pick(data, ["主题键", "topic_key"]),
            _pick(data, ["主标题", "title", "标准标题", "知识标题"]),
            _pick(data, ["关联标准项", "standard_path", "关联标准路径"]),
            _pick(data, ["适用范围", "scope", "适用场景"]),
            _pick(data, ["来源版本", "version", "版本", "source_version"]),
        ]
    )
    if not identity.replace("|", "").strip():
        return ""
    digest = hashlib.sha1(identity.encode("utf-8")).hexdigest()[:12].upper()
    return f"{prefix}-{digest}"


@dataclass(frozen=True)
class StandardCatalogItem:
    standard_id: str
    title: str
    category_l1: str
    category_l2: str
    knowledge_type: str
    standard_path: str
    keywords: list[str]
    scope: str
    response_snippet: str
    status: str
    version: str
    source_file: str = ""
    source_sheet: str = ""
    source_rows: tuple[int, ...] = ()

    def searchable_text(self) -> str:
        parts = [
            self.title,
            self.category_l1,
            self.category_l2,
            self.knowledge_type,
            self.standard_path,
            self.scope,
            self.response_snippet,
            " ".join(self.keywords),
        ]
        return " ".join(part for part in parts if part).lower()


def _normalize_row(data: dict[str, Any]) -> StandardCatalogItem:
    explicit_keywords = _split_keywords(_pick(data, ["keywords", "检索关键词", "关键词", "标签"]))
    standard_path = _pick(data, ["standard_path", "关联标准项", "关联标准路径"])
    path_keywords = _split_keywords(
        re.sub(r"[【】\[\]()（）*\\\-]+", " ", standard_path)
    )
    normalized_keywords = _expand_keyword_aliases(
        list(dict.fromkeys(explicit_keywords + path_keywords))
    )
    path_l1, path_l2 = _path_categories(standard_path)
    raw_source_rows = data.get("source_rows") or []
    if not isinstance(raw_source_rows, (list, tuple)):
        raw_source_rows = [raw_source_rows]
    source_rows = tuple(
        row_number
        for row_number in (
            int(value)
            for value in raw_source_rows
            if str(value).strip().isdigit()
        )
        if row_number > 0
    )
    return StandardCatalogItem(
        standard_id=_stable_catalog_standard_id(data),
        title=_pick(data, ["title", "主标题", "标准标题", "知识标题"]),
        category_l1=_pick(data, ["category_l1", "一级分类", "分类一级", "一级类目"]) or path_l1,
        category_l2=_pick(data, ["category_l2", "二级分类", "分类二级", "二级类目"]) or path_l2,
        knowledge_type=_pick(data, ["knowledge_type", "知识分类"]),
        standard_path=standard_path,
        keywords=normalized_keywords,
        scope=_catalog_scope(data),
        response_snippet=_pick(data, ["response_snippet", "参考话术", "话术", "回复话术", "知识内容"]),
        status=_pick(data, ["status", "状态", "生效状态"]) or "published",
        version=_pick(data, ["version", "版本", "source_version", "来源版本"]) or "v1",
        source_file=_pick(data, ["source_file", "来源文件"]),
        source_sheet=_pick(data, ["source_sheet", "来源工作表"]),
        source_rows=source_rows,
    )


def _rows_from_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    results: list[dict[str, Any]] = []
    for row in rows[1:]:
        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            record[header] = value
        if any(value not in (None, "") for value in record.values()):
            results.append(record)
    return results


def load_standard_catalog(path: str | Path | None) -> list[StandardCatalogItem]:
    if not path:
        return []
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Standard catalog not found: {file_path}")

    if file_path.suffix.lower() == ".json":
        payload = json.loads(file_path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            rows = payload.get("items") or payload.get("rows") or []
        else:
            rows = payload
    else:
        rows = _rows_from_xlsx(file_path)
    return [
        item
        for item in (_normalize_row(dict(row)) for row in rows)
        if is_active_standard(item.status)
    ]
