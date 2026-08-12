from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
import json
import math
from pathlib import Path
import random
import re
from typing import Any, Iterable, Sequence
from zipfile import BadZipFile

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .knowledge_categories import (
    SUPPORTED_KNOWLEDGE_CATEGORIES,
    UNCERTAIN_CATEGORY,
    normalize_knowledge_category,
)
from .operations import BLOCKING_PATTERNS, WARNING_PATTERNS


MODEL_VERSION = "topic-label-hash-nb-v1"
PSEUDO_LABEL_SOURCE = "mimo_pseudo_labels"
HUMAN_LABEL_SOURCE = "human_review"
KNOWLEDGE_VALUE_CLASSES = ("值得沉淀", "不值得沉淀")
DEFAULT_HASH_SIZE = 8192
DEFAULT_CONFIDENCE_THRESHOLD = 0.72
DEFAULT_NGRAM_MIN = 1
DEFAULT_NGRAM_MAX = 3
MIN_SAMPLES_PER_LABEL = 2

_YES_VALUES = {"是", "yes", "true", "1", "进入", "纳入"}
_NO_VALUES = {"否", "no", "false", "0", "不进入", "不纳入"}
_SENSITIVE_PATTERNS = {
    **BLOCKING_PATTERNS,
    "bank_card_like": WARNING_PATTERNS["bank_card_like"],
    "long_number": re.compile(r"(?<!\d)\d{8,}(?!\d)"),
}
_WHITESPACE_RE = re.compile(r"\s+")

_FEATURE_FIELDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    (
        "回收业务层级",
        (
            "business_lines",
            "scope_types",
            "回收业务层级",
        ),
    ),
    (
        "产品类型",
        (
            "product_categories",
            "product_category",
            "产品类型",
        ),
    ),
    (
        "主题问题",
        (
            "normalized_issues",
            "聚类主题",
            "主题问题",
            "成员核心问题",
        ),
    ),
    (
        "问题意图",
        (
            "intents",
            "主题问题意图",
        ),
    ),
    (
        "对象部位",
        (
            "subjects",
            "judgment_targets",
            "主题对象/部位",
        ),
    ),
    (
        "异常现象",
        (
            "phenomena",
            "主题异常现象",
        ),
    ),
    (
        "解题方式",
        (
            "resolution_modes",
            "主题解题方式",
        ),
    ),
    (
        "阈值例外",
        (
            "thresholds_or_exceptions",
        ),
    ),
    (
        "会话证据",
        (
            "conversation_evidence",
            "evidence_summaries",
            "historical_replies",
        ),
    ),
    (
        "会话类型",
        (
            "conversation_types",
        ),
    ),
)


@dataclass(frozen=True)
class TopicLabelTrainingSample:
    sample_id: str
    topic: dict[str, Any]
    topic_stage: str
    knowledge_value: str
    label_source: str


@dataclass(frozen=True)
class _ModelHead:
    classes: tuple[str, ...]
    log_prior: np.ndarray
    feature_log_probability: np.ndarray


class TopicLabelModel:
    def __init__(
        self,
        *,
        metadata: dict[str, Any],
        stage_head: _ModelHead,
        value_head: _ModelHead,
    ) -> None:
        self.metadata = metadata
        self.stage_head = stage_head
        self.value_head = value_head

    def predict(self, topic: dict[str, Any]) -> dict[str, Any]:
        feature_text = build_topic_feature_text(topic)
        if not feature_text:
            raise ValueError("主题缺少可用于本地分类的文本证据")
        feature_indices = _hashed_feature_indices(
            feature_text,
            hash_size=int(self.metadata["hash_size"]),
            ngram_min=int(self.metadata["ngram_min"]),
            ngram_max=int(self.metadata["ngram_max"]),
        )
        stage_probabilities = _predict_probabilities(
            self.stage_head,
            feature_indices,
        )
        value_probabilities = _predict_probabilities(
            self.value_head,
            feature_indices,
        )
        stage_index = int(np.argmax(stage_probabilities))
        value_index = int(np.argmax(value_probabilities))
        topic_stage = self.stage_head.classes[stage_index]
        knowledge_value = self.value_head.classes[value_index]
        stage_confidence = float(stage_probabilities[stage_index])
        value_confidence = float(value_probabilities[value_index])
        threshold = float(
            self.metadata.get(
                "confidence_threshold",
                DEFAULT_CONFIDENCE_THRESHOLD,
            )
        )
        review_reasons: list[str] = []
        if bool(self.metadata.get("force_human_review")):
            review_reasons.append(
                "当前模型由 MiMo 伪标签训练，尚未使用人工真值验收"
            )
        if not bool(self.metadata.get("production_eligible")):
            review_reasons.append("当前模型尚未通过生产验收")
        if any(
            _as_bool(topic.get(key))
            for key in (
                "upstream_requires_review",
                "requires_review",
                "是否重点复核",
                "主题分类重点复核",
            )
        ):
            review_reasons.append("上游证据冲突或风险标记要求人工复核")
        if topic_stage == UNCERTAIN_CATEGORY:
            review_reasons.append("主题类别预测为不确定")
        if stage_confidence < threshold:
            review_reasons.append(
                f"主题类别置信度低于 {threshold:.2f}"
            )
        if value_confidence < threshold:
            review_reasons.append(
                f"沉淀价值置信度低于 {threshold:.2f}"
            )
        needs_human_review = bool(review_reasons)
        combined_confidence = min(stage_confidence, value_confidence)
        return {
            "topic_stage": topic_stage,
            "knowledge_value": knowledge_value,
            "topic_stage_confidence": round(stage_confidence, 4),
            "knowledge_value_confidence": round(value_confidence, 4),
            "confidence": round(combined_confidence, 4),
            "needs_human_review": needs_human_review,
            "review_reasons": review_reasons,
            "stage_reason": (
                "本地实验模型根据主题证据的哈希字符特征预测问题分类。"
            ),
            "value_reason": (
                "本地实验模型根据主题证据预测沉淀价值；"
                "结果不替代人工复核。"
            ),
            "reusable_knowledge": (
                "该模型只输出分类标签，不生成或补写知识内容。"
            ),
            "topic_stage_probabilities": {
                label: round(float(probability), 4)
                for label, probability in zip(
                    self.stage_head.classes,
                    stage_probabilities,
                    strict=True,
                )
            },
            "knowledge_value_probabilities": {
                label: round(float(probability), 4)
                for label, probability in zip(
                    self.value_head.classes,
                    value_probabilities,
                    strict=True,
                )
            },
            "model_version": self.metadata.get(
                "model_version",
                MODEL_VERSION,
            ),
            "label_source": self.metadata.get("label_source", ""),
        }


def _text(value: Any, *, limit: int = 1200) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text[:limit]


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = _text(value, limit=32).lower()
    return text in {"1", "true", "yes", "是", "需要", "重点复核"}


def _flatten_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        result: list[str] = []
        for item in value:
            result.extend(_flatten_values(item))
        return result
    text = _text(value)
    return [text] if text else []


def _unique(values: Iterable[str], *, limit: int = 16) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
        if len(result) >= limit:
            break
    return result


def _sanitize_feature_value(value: str) -> str:
    text = _WHITESPACE_RE.sub(" ", _text(value)).strip()
    for name, pattern in _SENSITIVE_PATTERNS.items():
        text = pattern.sub(f"[{name}]", text)
    return text


def build_topic_feature_text(topic: dict[str, Any]) -> str:
    """Build evidence-only text without IDs, old categories or teacher reasons."""
    lines: list[str] = []
    for label, keys in _FEATURE_FIELDS:
        values: list[str] = []
        for key in keys:
            values.extend(_flatten_values(topic.get(key)))
        sanitized = _unique(
            _sanitize_feature_value(value)
            for value in values
        )
        if sanitized:
            lines.append(f"{label}：" + "；".join(sanitized))

    try:
        member_count = int(topic.get("member_count") or topic.get("主题样本数") or 0)
    except (TypeError, ValueError):
        member_count = 0
    if member_count:
        lines.append("主题规模：" + ("单案例" if member_count == 1 else "多案例"))
    if _as_bool(topic.get("upstream_requires_review")):
        lines.append("上游状态：需要人工复核")
    return "\n".join(lines)[:12000]


def _normalize_knowledge_value(value: Any) -> str:
    text = _text(value, limit=32).lower()
    if text in {"值得沉淀", "值得", "是", "yes", "true", "1"}:
        return "值得沉淀"
    if text in {
        "不值得沉淀",
        "不值得",
        "否",
        "no",
        "false",
        "0",
    }:
        return "不值得沉淀"
    return ""


def _load_pseudo_label_samples_with_stats(
    source_path: str | Path,
    *,
    minimum_confidence: float,
    allow_upstream_risk: bool,
) -> tuple[list[TopicLabelTrainingSample], dict[str, int | float]]:
    path = Path(source_path)
    payload = json.loads(path.read_text(encoding="utf-8"))
    themes = payload.get("themes") if isinstance(payload, dict) else None
    if not isinstance(themes, list):
        raise ValueError("伪标签 JSON 缺少 themes 数组")
    stats: dict[str, int | float] = {
        "minimum_confidence": minimum_confidence,
        "source_theme_count": len(themes),
        "selected_count": 0,
        "skipped_invalid": 0,
        "skipped_low_confidence": 0,
        "skipped_risky": 0,
        "skipped_upstream_risk": 0,
        "selected_upstream_risk": 0,
        "selected_uncertain_review": 0,
    }
    samples: list[TopicLabelTrainingSample] = []
    for index, theme in enumerate(themes, start=1):
        if not isinstance(theme, dict):
            stats["skipped_invalid"] += 1
            continue
        if _text(theme.get("classification_status"), limit=32) not in {
            "",
            "ok",
        }:
            stats["skipped_invalid"] += 1
            continue
        prediction = theme.get("prediction")
        if not isinstance(prediction, dict):
            stats["skipped_invalid"] += 1
            continue
        topic_stage = normalize_knowledge_category(
            prediction.get("topic_stage"),
            default="",
        )
        knowledge_value = _normalize_knowledge_value(
            prediction.get("knowledge_value")
        )
        if not topic_stage or not knowledge_value:
            stats["skipped_invalid"] += 1
            continue
        try:
            teacher_confidence = float(prediction.get("confidence"))
        except (TypeError, ValueError):
            stats["skipped_invalid"] += 1
            continue
        if not 0 <= teacher_confidence <= 1:
            stats["skipped_invalid"] += 1
            continue
        if teacher_confidence < minimum_confidence:
            stats["skipped_low_confidence"] += 1
            continue
        teacher_requires_review = _as_bool(
            prediction.get("needs_human_review")
        )
        if (
            teacher_requires_review
            and topic_stage != UNCERTAIN_CATEGORY
        ):
            stats["skipped_risky"] += 1
            continue
        upstream_requires_review = any(
            _as_bool(theme.get(key))
            for key in (
                "upstream_requires_review",
                "requires_review",
                "是否重点复核",
                "主题分类重点复核",
            )
        )
        if (
            upstream_requires_review
            and topic_stage != UNCERTAIN_CATEGORY
            and not allow_upstream_risk
        ):
            stats["skipped_upstream_risk"] += 1
            continue
        if not build_topic_feature_text(theme):
            stats["skipped_invalid"] += 1
            continue
        if teacher_requires_review:
            stats["selected_uncertain_review"] += 1
        if upstream_requires_review:
            stats["selected_upstream_risk"] += 1
        samples.append(
            TopicLabelTrainingSample(
                sample_id=(
                    _text(theme.get("theme_id"), limit=120)
                    or f"theme-{index:05d}"
                ),
                topic=theme,
                topic_stage=topic_stage,
                knowledge_value=knowledge_value,
                label_source=PSEUDO_LABEL_SOURCE,
            )
        )
    stats["selected_count"] = len(samples)
    if not samples:
        raise ValueError("伪标签 JSON 中没有可训练的主题")
    return samples, stats


def load_pseudo_label_samples(
    source_path: str | Path,
    *,
    minimum_confidence: float = DEFAULT_CONFIDENCE_THRESHOLD,
    allow_upstream_risk: bool = False,
) -> list[TopicLabelTrainingSample]:
    samples, _ = _load_pseudo_label_samples_with_stats(
        source_path,
        minimum_confidence=minimum_confidence,
        allow_upstream_risk=allow_upstream_risk,
    )
    return samples


def _header_map(values: Sequence[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(values):
        name = _text(value, limit=120)
        if name and name not in result:
            result[name] = index
    return result


def _row_value(
    row: Sequence[Any],
    headers: dict[str, int],
    *names: str,
) -> Any:
    for name in names:
        index = headers.get(name)
        if index is not None and index < len(row):
            return row[index]
    return ""


def _split_cell(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return _unique(
        part.strip()
        for part in re.split(r"[\r\n；]+", text)
        if part.strip()
    )


def load_human_review_samples(
    source_path: str | Path,
) -> list[TopicLabelTrainingSample]:
    path = Path(source_path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ValueError(
            f"无法读取人工复核工作簿，文件可能损坏或格式不正确：{path.name}"
        ) from exc
    selected_sheet = None
    headers: dict[str, int] = {}
    for sheet in workbook.worksheets:
        try:
            first_row = next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )
        except StopIteration:
            continue
        candidate_headers = _header_map(first_row)
        if (
            "人工主题问题分类" in candidate_headers
            and (
                "人工是否值得沉淀" in candidate_headers
                or "是否值得沉淀" in candidate_headers
            )
        ):
            selected_sheet = sheet
            headers = candidate_headers
            break
    if selected_sheet is None:
        raise ValueError(
            "人工复核工作簿缺少“人工主题问题分类”和“人工是否值得沉淀”列"
        )

    raw_rows = list(selected_sheet.iter_rows(min_row=2, values_only=True))
    training_flags = [
        _text(_row_value(row, headers, "是否进入训练集"), limit=32).lower()
        for row in raw_rows
    ]
    require_training_flag = any(training_flags)
    samples: list[TopicLabelTrainingSample] = []
    for index, row in enumerate(raw_rows, start=2):
        flag = _text(
            _row_value(row, headers, "是否进入训练集"),
            limit=32,
        ).lower()
        if require_training_flag and flag not in _YES_VALUES:
            continue
        if flag in _NO_VALUES:
            continue
        topic_stage = normalize_knowledge_category(
            _row_value(row, headers, "人工主题问题分类"),
            default="",
        )
        knowledge_value = _normalize_knowledge_value(
            _row_value(
                row,
                headers,
                "人工是否值得沉淀",
                "是否值得沉淀",
            )
        )
        if not topic_stage or not knowledge_value:
            continue
        topic = {
            "member_count": _row_value(row, headers, "主题样本数"),
            "product_categories": _split_cell(
                _row_value(row, headers, "产品类型")
            ),
            "normalized_issues": _split_cell(
                _row_value(row, headers, "聚类主题", "主题问题")
            )
            + _split_cell(
                _row_value(row, headers, "成员核心问题")
            ),
            "subjects": _split_cell(
                _row_value(row, headers, "主题对象/部位")
            ),
            "phenomena": _split_cell(
                _row_value(row, headers, "主题异常现象")
            ),
            "resolution_modes": _split_cell(
                _row_value(row, headers, "主题解题方式")
            ),
        }
        if not build_topic_feature_text(topic):
            continue
        samples.append(
            TopicLabelTrainingSample(
                sample_id=(
                    _text(
                        _row_value(
                            row,
                            headers,
                            "聚类主题ID",
                            "主题ID",
                        ),
                        limit=120,
                    )
                    or f"row-{index:05d}"
                ),
                topic=topic,
                topic_stage=topic_stage,
                knowledge_value=knowledge_value,
                label_source=HUMAN_LABEL_SOURCE,
            )
        )
    if not samples:
        raise ValueError("人工复核工作簿中没有可训练且标签完整的主题")
    return samples


def _hashed_feature_indices(
    text: str,
    *,
    hash_size: int,
    ngram_min: int,
    ngram_max: int,
) -> np.ndarray:
    indices: set[int] = set()
    for raw_line in text.splitlines():
        line = _WHITESPACE_RE.sub("", raw_line.lower())
        if not line:
            continue
        bounded = f"^{line}$"
        for size in range(ngram_min, ngram_max + 1):
            if len(bounded) < size:
                continue
            for start in range(len(bounded) - size + 1):
                token = bounded[start : start + size]
                digest = hashlib.blake2b(
                    token.encode("utf-8"),
                    digest_size=8,
                    person=b"topic-label-v1",
                ).digest()
                indices.add(
                    int.from_bytes(digest, "little", signed=False)
                    % hash_size
                )
    return np.asarray(sorted(indices), dtype=np.int32)


def _fit_head(
    feature_texts: Sequence[str],
    labels: Sequence[str],
    classes: Sequence[str],
    *,
    hash_size: int,
    ngram_min: int,
    ngram_max: int,
    alpha: float,
) -> _ModelHead:
    class_tuple = tuple(classes)
    class_index = {label: index for index, label in enumerate(class_tuple)}
    feature_counts = np.zeros(
        (len(class_tuple), hash_size),
        dtype=np.float64,
    )
    class_feature_totals = np.zeros(len(class_tuple), dtype=np.float64)
    for text, label in zip(feature_texts, labels, strict=True):
        label_index = class_index[label]
        indices = _hashed_feature_indices(
            text,
            hash_size=hash_size,
            ngram_min=ngram_min,
            ngram_max=ngram_max,
        )
        if indices.size:
            feature_counts[label_index, indices] += 1.0
            class_feature_totals[label_index] += float(indices.size)
    feature_log_probability = np.log(
        (feature_counts + alpha)
        / (
            class_feature_totals[:, None]
            + alpha * hash_size
        )
    )
    log_prior = np.full(
        len(class_tuple),
        -math.log(len(class_tuple)),
        dtype=np.float64,
    )
    return _ModelHead(
        classes=class_tuple,
        log_prior=log_prior,
        feature_log_probability=feature_log_probability,
    )


def _predict_probabilities(
    head: _ModelHead,
    feature_indices: np.ndarray,
) -> np.ndarray:
    scores = head.log_prior.copy()
    if feature_indices.size:
        scores += head.feature_log_probability[:, feature_indices].sum(axis=1)
    scores -= np.max(scores)
    probabilities = np.exp(scores)
    total = float(probabilities.sum())
    if not total:
        return np.full(
            len(head.classes),
            1.0 / len(head.classes),
            dtype=np.float64,
        )
    return probabilities / total


def _stratified_holdout(
    labels: Sequence[str],
    *,
    seed: int,
    test_ratio: float = 0.2,
) -> tuple[list[int], list[int]]:
    grouped: dict[str, list[int]] = {}
    for index, label in enumerate(labels):
        grouped.setdefault(label, []).append(index)
    randomizer = random.Random(seed)
    train_indices: list[int] = []
    test_indices: list[int] = []
    for label in sorted(grouped):
        indices = list(grouped[label])
        randomizer.shuffle(indices)
        if len(indices) < 2:
            train_indices.extend(indices)
            continue
        test_count = max(1, int(round(len(indices) * test_ratio)))
        test_count = min(test_count, len(indices) - 1)
        test_indices.extend(indices[:test_count])
        train_indices.extend(indices[test_count:])
    return sorted(train_indices), sorted(test_indices)


def _classification_metrics(
    expected: Sequence[str],
    predicted: Sequence[str],
    classes: Sequence[str],
) -> dict[str, Any]:
    total = len(expected)
    correct = sum(
        actual == guess
        for actual, guess in zip(expected, predicted, strict=True)
    )
    per_class: dict[str, dict[str, Any]] = {}
    f1_values: list[float] = []
    confusion: dict[str, dict[str, int]] = {
        label: {other: 0 for other in classes}
        for label in classes
    }
    for actual, guess in zip(expected, predicted, strict=True):
        confusion[actual][guess] += 1
    for label in classes:
        true_positive = sum(
            actual == label and guess == label
            for actual, guess in zip(expected, predicted, strict=True)
        )
        false_positive = sum(
            actual != label and guess == label
            for actual, guess in zip(expected, predicted, strict=True)
        )
        false_negative = sum(
            actual == label and guess != label
            for actual, guess in zip(expected, predicted, strict=True)
        )
        support = sum(actual == label for actual in expected)
        precision = (
            true_positive / (true_positive + false_positive)
            if true_positive + false_positive
            else 0.0
        )
        recall = (
            true_positive / (true_positive + false_negative)
            if true_positive + false_negative
            else 0.0
        )
        f1 = (
            2 * precision * recall / (precision + recall)
            if precision + recall
            else 0.0
        )
        if support:
            f1_values.append(f1)
        per_class[label] = {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "support": support,
        }
    return {
        "sample_count": total,
        "accuracy": round(correct / total, 4) if total else 0.0,
        "macro_f1": (
            round(sum(f1_values) / len(f1_values), 4)
            if f1_values
            else 0.0
        ),
        "per_class": per_class,
        "confusion_matrix": confusion,
    }


def _evaluate_head(
    feature_texts: Sequence[str],
    labels: Sequence[str],
    classes: Sequence[str],
    *,
    hash_size: int,
    ngram_min: int,
    ngram_max: int,
    alpha: float,
    seed: int,
) -> dict[str, Any]:
    train_indices, test_indices = _stratified_holdout(
        labels,
        seed=seed,
    )
    if not test_indices:
        return {
            "sample_count": 0,
            "accuracy": 0.0,
            "macro_f1": 0.0,
            "per_class": {},
            "confusion_matrix": {},
        }
    head = _fit_head(
        [feature_texts[index] for index in train_indices],
        [labels[index] for index in train_indices],
        classes,
        hash_size=hash_size,
        ngram_min=ngram_min,
        ngram_max=ngram_max,
        alpha=alpha,
    )
    predicted: list[str] = []
    for index in test_indices:
        feature_indices = _hashed_feature_indices(
            feature_texts[index],
            hash_size=hash_size,
            ngram_min=ngram_min,
            ngram_max=ngram_max,
        )
        probabilities = _predict_probabilities(head, feature_indices)
        predicted.append(head.classes[int(np.argmax(probabilities))])
    expected = [labels[index] for index in test_indices]
    result = _classification_metrics(expected, predicted, classes)
    majority_label = Counter(
        labels[index] for index in train_indices
    ).most_common(1)[0][0]
    result["majority_baseline_accuracy"] = round(
        sum(label == majority_label for label in expected)
        / len(expected),
        4,
    )
    return result


def _deduplicate_samples(
    samples: Sequence[TopicLabelTrainingSample],
) -> tuple[list[TopicLabelTrainingSample], int]:
    result: list[TopicLabelTrainingSample] = []
    seen: dict[str, tuple[str, str]] = {}
    duplicates_removed = 0
    for sample in samples:
        feature_text = build_topic_feature_text(sample.topic)
        fingerprint = hashlib.sha256(
            feature_text.encode("utf-8")
        ).hexdigest()
        labels = (sample.topic_stage, sample.knowledge_value)
        previous = seen.get(fingerprint)
        if previous is None:
            seen[fingerprint] = labels
            result.append(sample)
            continue
        if previous != labels:
            raise ValueError(
                "相同主题证据出现互相冲突的分类或沉淀价值标签，"
                "请先完成人工复核"
            )
        duplicates_removed += 1
    return result, duplicates_removed


def _validate_label_coverage(
    samples: Sequence[TopicLabelTrainingSample],
) -> tuple[Counter[str], Counter[str]]:
    stage_counts = Counter(sample.topic_stage for sample in samples)
    value_counts = Counter(sample.knowledge_value for sample in samples)
    missing_or_small = [
        f"{label}={stage_counts.get(label, 0)}"
        for label in SUPPORTED_KNOWLEDGE_CATEGORIES
        if stage_counts.get(label, 0) < MIN_SAMPLES_PER_LABEL
    ]
    missing_or_small.extend(
        f"{label}={value_counts.get(label, 0)}"
        for label in KNOWLEDGE_VALUE_CLASSES
        if value_counts.get(label, 0) < MIN_SAMPLES_PER_LABEL
    )
    if missing_or_small:
        raise ValueError(
            "训练样本不足，每个标签至少需要 "
            f"{MIN_SAMPLES_PER_LABEL} 条："
            + "，".join(missing_or_small)
        )
    return stage_counts, value_counts


def _training_fingerprint(
    samples: Sequence[TopicLabelTrainingSample],
) -> str:
    digest = hashlib.sha256()
    for sample in sorted(samples, key=lambda item: item.sample_id):
        digest.update(build_topic_feature_text(sample.topic).encode("utf-8"))
        digest.update(b"\0")
        digest.update(sample.topic_stage.encode("utf-8"))
        digest.update(b"\0")
        digest.update(sample.knowledge_value.encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def _imbalance_warnings(
    stage_counts: Counter[str],
    value_counts: Counter[str],
    *,
    label_source: str,
) -> list[str]:
    warnings: list[str] = []
    if label_source == PSEUDO_LABEL_SOURCE:
        warnings.append(
            "训练标签来自 MiMo 初标，不是人工真值；模型仅可用于实验和辅助标注。"
        )
    for label, count in [*stage_counts.items(), *value_counts.items()]:
        if count < 20:
            warnings.append(
                f"标签“{label}”只有 {count} 条，评估结果不稳定。"
            )
    for name, counts in (
        ("主题类别", stage_counts),
        ("沉淀价值", value_counts),
    ):
        nonzero = [count for count in counts.values() if count]
        if nonzero and max(nonzero) / min(nonzero) >= 5:
            warnings.append(
                f"{name}标签分布严重不均衡，最大类别是最小类别的"
                f" {max(nonzero) / min(nonzero):.1f} 倍。"
            )
    return warnings


def _evaluation_warnings(
    stage_evaluation: dict[str, Any],
    value_evaluation: dict[str, Any],
) -> list[str]:
    warnings: list[str] = []
    for label, evaluation in (
        ("主题类别", stage_evaluation),
        ("沉淀价值", value_evaluation),
    ):
        accuracy = float(evaluation.get("accuracy") or 0)
        baseline = float(
            evaluation.get("majority_baseline_accuracy") or 0
        )
        if accuracy <= baseline:
            warnings.append(
                f"{label}留出准确率 {accuracy:.4f} 未超过多数类基线 "
                f"{baseline:.4f}，不能证明模型优于恒定预测。"
            )
        zero_recall_labels = [
            class_name
            for class_name, metrics in (
                evaluation.get("per_class") or {}
            ).items()
            if int(metrics.get("support") or 0) > 0
            and float(metrics.get("recall") or 0) == 0
        ]
        if zero_recall_labels:
            warnings.append(
                f"{label}留出评估中以下标签召回率为 0："
                + "、".join(zero_recall_labels)
                + "。"
            )
    return warnings


def _atomic_write_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(path.suffix + ".tmp")
    temporary_path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary_path.replace(path)


def train_topic_label_model(
    source_path: str | Path,
    output_dir: str | Path,
    *,
    allow_pseudo_labels: bool = False,
    hash_size: int = DEFAULT_HASH_SIZE,
    ngram_min: int = DEFAULT_NGRAM_MIN,
    ngram_max: int = DEFAULT_NGRAM_MAX,
    alpha: float = 1.0,
    confidence_threshold: float = DEFAULT_CONFIDENCE_THRESHOLD,
    pseudo_min_confidence: float = DEFAULT_CONFIDENCE_THRESHOLD,
    allow_upstream_risk_pseudo_labels: bool = False,
    seed: int = 42,
) -> dict[str, Any]:
    source = Path(source_path)
    if not source.exists():
        raise FileNotFoundError(source)
    if hash_size < 256:
        raise ValueError("hash_size 至少为 256")
    if not 1 <= ngram_min <= ngram_max <= 6:
        raise ValueError("字符 n-gram 范围必须满足 1 <= min <= max <= 6")
    if alpha <= 0:
        raise ValueError("alpha 必须大于 0")
    if not 0 <= confidence_threshold <= 1:
        raise ValueError("confidence_threshold 必须在 0~1")
    if not 0 <= pseudo_min_confidence <= 1:
        raise ValueError("pseudo_min_confidence 必须在 0~1")

    pseudo_label_filter: dict[str, int | float] | None = None
    if source.suffix.lower() == ".json":
        if not allow_pseudo_labels:
            raise ValueError(
                "JSON 中是模型伪标签；如确认只训练实验基线，"
                "请显式设置 allow_pseudo_labels=True"
            )
        samples, pseudo_label_filter = (
            _load_pseudo_label_samples_with_stats(
                source,
                minimum_confidence=pseudo_min_confidence,
                allow_upstream_risk=(
                    allow_upstream_risk_pseudo_labels
                ),
            )
        )
    elif source.suffix.lower() in {".xlsx", ".xlsm"}:
        samples = load_human_review_samples(source)
    else:
        raise ValueError("训练源仅支持 topic_stage_predictions.json 或人工复核 Excel")

    samples, duplicates_removed = _deduplicate_samples(samples)
    stage_counts, value_counts = _validate_label_coverage(samples)
    label_sources = {sample.label_source for sample in samples}
    if len(label_sources) != 1:
        raise ValueError("同一次训练不能混用人工标签和模型伪标签")
    label_source = next(iter(label_sources))
    feature_texts = [
        build_topic_feature_text(sample.topic)
        for sample in samples
    ]
    stage_labels = [sample.topic_stage for sample in samples]
    value_labels = [sample.knowledge_value for sample in samples]

    stage_evaluation = _evaluate_head(
        feature_texts,
        stage_labels,
        SUPPORTED_KNOWLEDGE_CATEGORIES,
        hash_size=hash_size,
        ngram_min=ngram_min,
        ngram_max=ngram_max,
        alpha=alpha,
        seed=seed,
    )
    value_evaluation = _evaluate_head(
        feature_texts,
        value_labels,
        KNOWLEDGE_VALUE_CLASSES,
        hash_size=hash_size,
        ngram_min=ngram_min,
        ngram_max=ngram_max,
        alpha=alpha,
        seed=seed + 1,
    )
    stage_head = _fit_head(
        feature_texts,
        stage_labels,
        SUPPORTED_KNOWLEDGE_CATEGORIES,
        hash_size=hash_size,
        ngram_min=ngram_min,
        ngram_max=ngram_max,
        alpha=alpha,
    )
    value_head = _fit_head(
        feature_texts,
        value_labels,
        KNOWLEDGE_VALUE_CLASSES,
        hash_size=hash_size,
        ngram_min=ngram_min,
        ngram_max=ngram_max,
        alpha=alpha,
    )

    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    model_path = output / "model.npz"
    temporary_model_path = output / "model.npz.tmp"
    with temporary_model_path.open("wb") as model_file:
        np.savez_compressed(
            model_file,
            stage_log_prior=stage_head.log_prior,
            stage_feature_log_probability=(
                stage_head.feature_log_probability
            ),
            value_log_prior=value_head.log_prior,
            value_feature_log_probability=(
                value_head.feature_log_probability
            ),
        )
    temporary_model_path.replace(model_path)

    created_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    force_human_review = label_source == PSEUDO_LABEL_SOURCE
    metadata = {
        "model_version": MODEL_VERSION,
        "created_at": created_at,
        "label_source": label_source,
        "force_human_review": force_human_review,
        "production_eligible": False,
        "sample_count": len(samples),
        "hash_size": hash_size,
        "ngram_min": ngram_min,
        "ngram_max": ngram_max,
        "alpha": alpha,
        "confidence_threshold": confidence_threshold,
        "pseudo_min_confidence": (
            pseudo_min_confidence
            if label_source == PSEUDO_LABEL_SOURCE
            else None
        ),
        "allow_upstream_risk_pseudo_labels": (
            allow_upstream_risk_pseudo_labels
            if label_source == PSEUDO_LABEL_SOURCE
            else False
        ),
        "stage_classes": list(SUPPORTED_KNOWLEDGE_CATEGORIES),
        "value_classes": list(KNOWLEDGE_VALUE_CLASSES),
        "training_fingerprint": _training_fingerprint(samples),
        "raw_text_stored": False,
        "feature_vocabulary_stored": False,
        "confidence_calibrated": False,
        "old_category_fields_used": False,
        "standard_reference_fields_used": False,
    }
    warnings = _imbalance_warnings(
        stage_counts,
        value_counts,
        label_source=label_source,
    )
    warnings.extend(
        _evaluation_warnings(
            stage_evaluation,
            value_evaluation,
        )
    )
    if (
        pseudo_label_filter
        and int(
            pseudo_label_filter.get("selected_upstream_risk") or 0
        )
    ):
        warnings.append(
            "本次训练显式纳入了带上游风险标记的伪标签；"
            "模型只能用于实验辅助标注，所有结果必须人工复核。"
        )
    report = {
        **metadata,
        "source_file": source.name,
        "duplicates_removed": duplicates_removed,
        "pseudo_label_filter": pseudo_label_filter,
        "topic_stage_distribution": dict(stage_counts),
        "knowledge_value_distribution": dict(value_counts),
        "topic_stage_holdout": stage_evaluation,
        "knowledge_value_holdout": value_evaluation,
        "warnings": warnings,
        "evaluation_reference": (
            "MiMo 伪标签，仅表示对教师标签的拟合能力，不是人工准确率。"
            if label_source == PSEUDO_LABEL_SOURCE
            else "人工复核标签"
        ),
        "recommended_next_step": (
            "完成人工复核工作簿中的“人工主题问题分类、"
            "人工是否值得沉淀、是否进入训练集”，"
            "再用人工真值重新训练并按品类做独立验收。"
        ),
    }
    _atomic_write_json(output / "metadata.json", metadata)
    _atomic_write_json(output / "training_report.json", report)

    return {
        "status": "trained_experimental_baseline",
        "model_version": MODEL_VERSION,
        "sample_count": len(samples),
        "label_source": label_source,
        "production_eligible": False,
        "force_human_review": force_human_review,
        "model_path": str(model_path),
        "metadata_path": str(output / "metadata.json"),
        "report_path": str(output / "training_report.json"),
        "topic_stage_accuracy": stage_evaluation["accuracy"],
        "topic_stage_macro_f1": stage_evaluation["macro_f1"],
        "knowledge_value_accuracy": value_evaluation["accuracy"],
        "knowledge_value_macro_f1": value_evaluation["macro_f1"],
        "warnings": warnings,
    }


def load_topic_label_model(
    model_dir: str | Path,
) -> TopicLabelModel:
    directory = Path(model_dir)
    metadata_path = directory / "metadata.json"
    model_path = directory / "model.npz"
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        raise ValueError("模型元数据损坏或不完整") from exc
    if metadata.get("model_version") != MODEL_VERSION:
        raise ValueError(
            "不支持的主题自动标注模型版本："
            f"{metadata.get('model_version', '')}"
        )
    try:
        with np.load(model_path, allow_pickle=False) as arrays:
            stage_head = _ModelHead(
                classes=tuple(metadata["stage_classes"]),
                log_prior=np.asarray(
                    arrays["stage_log_prior"],
                    dtype=np.float64,
                ),
                feature_log_probability=np.asarray(
                    arrays["stage_feature_log_probability"],
                    dtype=np.float64,
                ),
            )
            value_head = _ModelHead(
                classes=tuple(metadata["value_classes"]),
                log_prior=np.asarray(
                    arrays["value_log_prior"],
                    dtype=np.float64,
                ),
                feature_log_probability=np.asarray(
                    arrays["value_feature_log_probability"],
                    dtype=np.float64,
                ),
            )
    except (BadZipFile, KeyError, OSError, TypeError, ValueError) as exc:
        raise ValueError("模型权重损坏或不完整") from exc
    return TopicLabelModel(
        metadata=metadata,
        stage_head=stage_head,
        value_head=value_head,
    )


def predict_topic_labels_from_json(
    model_dir: str | Path,
    source_path: str | Path,
    output_path: str | Path,
) -> dict[str, Any]:
    source = Path(source_path)
    payload = json.loads(source.read_text(encoding="utf-8"))
    themes = payload.get("themes") if isinstance(payload, dict) else None
    if not isinstance(themes, list):
        raise ValueError("待标注 JSON 缺少 themes 数组")
    model = load_topic_label_model(model_dir)
    predictions: list[dict[str, Any]] = []
    successful_count = 0
    failed_count = 0
    human_review_count = 0
    for index, theme in enumerate(themes, start=1):
        theme_id = (
            _text(
                theme.get("theme_id") if isinstance(theme, dict) else "",
                limit=120,
            )
            or f"theme-{index:05d}"
        )
        if not isinstance(theme, dict):
            predictions.append(
                {
                    "theme_id": theme_id,
                    "status": "error",
                    "error": "主题不是 JSON 对象",
                }
            )
            failed_count += 1
            continue
        try:
            prediction = model.predict(theme)
        except (TypeError, ValueError) as exc:
            predictions.append(
                {
                    "theme_id": theme_id,
                    "status": "error",
                    "error": str(exc),
                }
            )
            failed_count += 1
            continue
        predictions.append(
            {
                "theme_id": theme_id,
                "status": "ok",
                "prediction": prediction,
            }
        )
        successful_count += 1
        human_review_count += int(
            bool(prediction.get("needs_human_review"))
        )
    output = {
        "metadata": {
            "model_version": model.metadata.get("model_version", ""),
            "label_source": model.metadata.get("label_source", ""),
            "production_eligible": bool(
                model.metadata.get("production_eligible")
            ),
            "source_file": source.name,
            "theme_count": len(themes),
            "successful_count": successful_count,
            "failed_count": failed_count,
            "human_review_count": human_review_count,
            "raw_topic_text_copied": False,
        },
        "predictions": predictions,
    }
    destination = Path(output_path)
    _atomic_write_json(destination, output)
    return {
        "status": "predicted",
        "model_version": model.metadata.get("model_version", ""),
        "theme_count": len(themes),
        "successful_count": successful_count,
        "failed_count": failed_count,
        "human_review_count": human_review_count,
        "output_path": str(destination),
    }
