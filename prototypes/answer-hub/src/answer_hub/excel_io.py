from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def write_rows_to_workbook(
    sheets: dict[str, tuple[list[str], Iterable[dict[str, Any]]]],
    path: str | Path,
) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    for sheet_name, (columns, rows) in sheets.items():
        worksheet = workbook.create_sheet(title=sheet_name)
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)

        for col_index, column_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        for row_index, row in enumerate(rows, start=2):
            for col_index, column_name in enumerate(columns, start=1):
                value = row.get(column_name, "")
                if isinstance(value, list):
                    value = "\n".join(str(item) for item in value if item not in (None, ""))
                worksheet.cell(row=row_index, column=col_index, value=value)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_index, column_name in enumerate(columns, start=1):
            max_len = len(column_name)
            for row in worksheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                value = row[0].value
                if value is None:
                    continue
                max_len = max(max_len, min(len(str(value)), 50))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_len + 2, 42)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def read_workbook_rows(path: str | Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    columns = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        record: dict[str, Any] = {}
        for index, column in enumerate(columns):
            if not column:
                continue
            record[column] = raw_row[index] if index < len(raw_row) else None
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return columns, records

