from __future__ import annotations

from collections import Counter
from dataclasses import asdict, dataclass
from math import sqrt
from pathlib import Path
from typing import Any, Callable, Protocol
import re

from openpyxl import load_workbook

from .catalog import StandardCatalogItem
from .embedding import EmbeddingError
from .product_taxonomy import (
    canonical_product_name,
    configured_product_names,
    product_from_scope,
)


OFFICIAL_CATEGORY_ALIASES = {
    "信息查询": {"基本情况"},
    "成色与回收标准": {"基本情况", "包装及配件"},
    "功能问题": {"设备功能情况"},
    "显示问题": {"屏幕显示情况", "副屏屏幕显示情况"},
    "外观问题": {
        "屏幕及正面外观",
        "副屏外观情况",
        "中框及外壳外观",
    },
    "拆修问题": {"拆修及浸液情况"},
}

OFFICIAL_PRIMARY_CATEGORIES = {
    "基本情况",
    "设备功能情况",
    "屏幕显示情况",
    "副屏屏幕显示情况",
    "屏幕及正面外观",
    "副屏外观情况",
    "中框及外壳外观",
    "拆修及浸液情况",
    "包装及配件",
}

GENERIC_STANDARD_TITLES = {
    "有",
    "无",
    "正常",
    "不支持",
    "无以上问题",
    "不检测",
    "机型",
    "颜色",
}

BROAD_MATCH_TERMS = {
    "标准",
    "标准定义",
    "检测方法",
    "质检",
    "问题",
    "单选",
    "多选",
    "通用",
    *configured_product_names(),
}


class EmbeddingProvider(Protocol):
    def embed_texts(
        self,
        texts: list[str],
        progress_callback: Callable[[int, int], None] | None = None,
    ) -> list[list[float]]: ...


@dataclass(frozen=True)
class HumanAnnotation:
    atomic_id: str
    source_sheet: str
    decision: str
    note: str
    target_cluster_id: str = ""
    source_cluster_id: str = ""


@dataclass(frozen=True)
class _QueryLexicalFeatures:
    raw: str
    full_grams: Counter[str]
    focus_grams: Counter[str]
    normalized: str
    numbers: set[str]
    note_normalized: str


@dataclass(frozen=True)
class _StandardLexicalFeatures:
    full_grams: Counter[str]
    focus_grams: Counter[str]
    candidate_path: str
    title: str
    numbers: set[str]
    normalized_keywords: tuple[tuple[str, str], ...]


def _text(value: Any) -> str:
    return str(value or "").strip()


def _normalize(value: Any) -> str:
    return re.sub(
        r"[^0-9a-z\u4e00-\u9fff%≤≥<>]+",
        "",
        _text(value).lower(),
    )


def _product_from_scope(scope: str) -> str:
    return product_from_scope(scope)


def _char_ngrams(value: Any) -> Counter[str]:
    text = _normalize(value)
    grams: Counter[str] = Counter()
    for size in (2, 3):
        if len(text) < size:
            continue
        grams.update(text[index : index + size] for index in range(len(text) - size + 1))
    return grams


def _cosine(left: Counter[str], right: Counter[str]) -> float:
    if not left or not right:
        return 0.0
    numerator = sum(count * right.get(key, 0) for key, count in left.items())
    if not numerator:
        return 0.0
    left_norm = sqrt(sum(count * count for count in left.values()))
    right_norm = sqrt(sum(count * count for count in right.values()))
    return numerator / (left_norm * right_norm)


def _vector_cosine(left: list[float], right: list[float]) -> float:
    if not left or len(left) != len(right):
        return 0.0
    numerator = sum(a * b for a, b in zip(left, right))
    left_norm = sqrt(sum(value * value for value in left))
    right_norm = sqrt(sum(value * value for value in right))
    if not left_norm or not right_norm:
        return 0.0
    return numerator / (left_norm * right_norm)


def _numbers(value: Any) -> set[str]:
    return {
        number.lstrip("0") or "0"
        for number in re.findall(r"\d+(?:\.\d+)?", _text(value))
    }


def _split_atomic_ids(value: Any) -> list[str]:
    return re.findall(r"S\d{3}-\d{2}", _text(value), flags=re.IGNORECASE)


def load_human_annotations(path: str | Path) -> dict[str, HumanAnnotation]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    annotations: dict[str, HumanAnnotation] = {}

    if "待人工确认" in workbook.sheetnames:
        sheet = workbook["待人工确认"]
        headers = [_text(cell.value) for cell in sheet[1]]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            atomic_id = _text(row.get("原子知识ID"))
            if not atomic_id:
                continue
            annotations[atomic_id] = HumanAnnotation(
                atomic_id=atomic_id,
                source_sheet="待人工确认",
                decision=_text(row.get("人工处理结论")),
                note=_text(row.get("人工说明")),
            )

    if "人工簇审核" in workbook.sheetnames:
        sheet = workbook["人工簇审核"]
        headers = [_text(cell.value) for cell in sheet[1]]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            atomic_ids = _split_atomic_ids(row.get("成员原子ID"))
            for atomic_id in atomic_ids:
                existing = annotations.get(atomic_id)
                note = _text(row.get("拆分/合并说明"))
                if existing and existing.note and note:
                    note = f"{existing.note}\n{note}"
                elif existing and existing.note:
                    note = existing.note
                annotations[atomic_id] = HumanAnnotation(
                    atomic_id=atomic_id,
                    source_sheet="人工簇审核",
                    decision=_text(row.get("人工审核结论")) or (existing.decision if existing else ""),
                    note=note,
                    target_cluster_id=_text(row.get("目标主题簇ID")),
                    source_cluster_id=_text(row.get("主题簇ID")),
                )
    workbook.close()
    return annotations


def _effective_category_l1(
    unit: dict[str, Any],
    annotation: HumanAnnotation | None,
    standards: list[StandardCatalogItem],
) -> str:
    note = annotation.note if annotation else ""
    official_categories = sorted(OFFICIAL_PRIMARY_CATEGORIES, key=len, reverse=True)
    for category in official_categories:
        if category in note:
            return category
    legacy_mentions = [
        legacy
        for legacy in OFFICIAL_CATEGORY_ALIASES
        if legacy in note
    ]
    if len(legacy_mentions) == 1:
        targets = OFFICIAL_CATEGORY_ALIASES[legacy_mentions[0]]
        if len(targets) == 1:
            return next(iter(targets))
    if "拆修" in note and "外观问题" not in note:
        return "拆修及浸液情况"
    return _text(unit.get("category_l1"))


def _effective_scope_type(
    unit: dict[str, Any],
    annotation: HumanAnnotation | None,
) -> str:
    note = annotation.note if annotation else ""
    if any(marker in note for marker in ("全手机机型都适用", "手机都适用", "全机型都适用")):
        return "品类专用"
    if any(marker in note for marker in ("苹果全机型", "品牌专用")):
        return "品牌专用"
    if any(marker in note for marker in ("安卓全部机型", "平台专用")):
        return "平台专用"
    if "机型专用" in note:
        return "机型专用"
    return _text(unit.get("scope_type"))


def _effective_subject(
    unit: dict[str, Any],
    annotation: HumanAnnotation | None,
) -> str:
    note = annotation.note if annotation else ""
    match = re.search(
        r"核心对象(?:应该)?(?:改为|是|为)[:：]?\s*([^\n，,；;。]+)",
        note,
    )
    if match:
        return match.group(1).strip()
    return _text(unit.get("subject"))


def _manual_merge_target(annotation: HumanAnnotation | None) -> str:
    if not annotation:
        return ""
    explicit = annotation.target_cluster_id
    if explicit:
        return explicit
    match = re.search(
        r"(?:与|应该与)\s*((?:T|S)\d{3}(?:-\d{2})?)\s*合并",
        annotation.note,
        flags=re.IGNORECASE,
    )
    return match.group(1).upper() if match else ""


def build_effective_atomic_unit(
    unit: dict[str, Any],
    annotation: HumanAnnotation | None,
    standards: list[StandardCatalogItem],
) -> dict[str, Any]:
    effective = dict(unit)
    effective["category_l1"] = _effective_category_l1(unit, annotation, standards)
    effective["scope_type"] = _effective_scope_type(unit, annotation)
    effective["subject"] = _effective_subject(unit, annotation)
    return effective


def build_atomic_query(
    unit: dict[str, Any],
    annotation: HumanAnnotation | None = None,
) -> str:
    fields = [
        unit.get("normalized_issue"),
        unit.get("product_category"),
        unit.get("scope_type"),
        unit.get("platform"),
        unit.get("brand"),
        unit.get("model_scope"),
        unit.get("category_l1"),
        unit.get("category_l2"),
        unit.get("subject"),
        unit.get("phenomenon"),
        unit.get("judgment_target"),
        unit.get("resolution_mode"),
        unit.get("standard_path"),
        unit.get("threshold_or_exception"),
    ]
    if annotation:
        fields.extend([annotation.decision, annotation.note])
    return "\n".join(_text(value) for value in fields if _text(value))


def _standard_text(item: StandardCatalogItem) -> str:
    return "\n".join(
        value
        for value in (
            item.title,
            item.category_l1,
            item.category_l2,
            item.standard_path,
            item.scope,
            " ".join(item.keywords),
            item.response_snippet[:2400],
        )
        if value
    )


def _embedding_standard_text(item: StandardCatalogItem) -> str:
    return "\n".join(
        value
        for value in (
            item.title,
            item.category_l1,
            item.category_l2,
            item.standard_path,
            " ".join(item.keywords[:20]),
            item.response_snippet[:300],
        )
        if value
    )


def _query_lexical_features(
    unit: dict[str, Any],
    query: str,
    annotation: HumanAnnotation | None,
) -> _QueryLexicalFeatures:
    focus_text = " ".join(
        [
            _text(unit.get("normalized_issue")),
            _text(unit.get("subject")),
            _text(unit.get("phenomenon")),
            _text(unit.get("judgment_target")),
        ]
    )
    return _QueryLexicalFeatures(
        raw=query,
        full_grams=_char_ngrams(query),
        focus_grams=_char_ngrams(focus_text),
        normalized=_normalize(query),
        numbers=_numbers(
            f"{unit.get('threshold_or_exception', '')} "
            f"{unit.get('normalized_issue', '')} "
            f"{annotation.note if annotation else ''}"
        ),
        note_normalized=_normalize(annotation.note if annotation else ""),
    )


def _standard_lexical_features(
    item: StandardCatalogItem,
) -> _StandardLexicalFeatures:
    return _StandardLexicalFeatures(
        full_grams=_char_ngrams(_standard_text(item)),
        focus_grams=_char_ngrams(
            f"{item.title} {item.category_l2} {item.standard_path}"
        ),
        candidate_path=_normalize(
            f"{item.title}{item.category_l2}{item.standard_path}"
        ),
        title=_normalize(item.title),
        numbers=_numbers(
            f"{item.title} {item.standard_path} {item.response_snippet}"
        ),
        normalized_keywords=tuple(
            (keyword, _normalize(keyword))
            for keyword in item.keywords
        ),
    )


def _lexical_score(
    unit: dict[str, Any],
    item: StandardCatalogItem,
    annotation: HumanAnnotation | None,
    query_features: _QueryLexicalFeatures,
    standard_features: _StandardLexicalFeatures,
) -> tuple[float, list[str]]:
    reasons: list[str] = []
    full_similarity = _cosine(
        query_features.full_grams,
        standard_features.full_grams,
    )
    title_similarity = _cosine(
        query_features.focus_grams,
        standard_features.focus_grams,
    )
    score = full_similarity * 58.0 + title_similarity * 28.0

    category_l1 = _text(unit.get("category_l1"))
    if category_l1 and category_l1 == item.category_l1:
        score += 18.0
        reasons.append("正式一级分类一致")
    elif item.category_l1 in OFFICIAL_CATEGORY_ALIASES.get(category_l1, set()):
        score += 5.0
        reasons.append("旧问题大类与正式一级类相容")

    subject = _normalize(unit.get("subject"))
    if (
        subject
        and len(subject) >= 2
        and subject in standard_features.candidate_path
    ):
        score += 12.0
        reasons.append("核心对象命中标准路径")

    title = standard_features.title
    if (
        title
        and len(title) >= 3
        and title in query_features.normalized
    ):
        score += 16.0
        reasons.append("标准程度值/标题直接命中")

    keyword_hits = []
    for keyword, normalized_keyword in standard_features.normalized_keywords:
        if (
            len(normalized_keyword) >= 2
            and keyword not in BROAD_MATCH_TERMS
            and normalized_keyword in query_features.normalized
        ):
            keyword_hits.append(keyword)
    if keyword_hits:
        distinct_hits = list(dict.fromkeys(keyword_hits))
        score += min(22.0, sum(min(len(_normalize(hit)), 6) * 0.8 for hit in distinct_hits))
        reasons.append(f"标准关键词命中：{'、'.join(distinct_hits[:4])}")

    shared_numbers = query_features.numbers & standard_features.numbers
    if shared_numbers:
        score += min(12.0, 6.0 * len(shared_numbers))
        reasons.append(f"阈值数字一致：{','.join(sorted(shared_numbers))}")
    elif query_features.numbers and standard_features.numbers:
        score -= 4.0

    if (
        item.title in GENERIC_STANDARD_TITLES
        and item.title not in query_features.raw
    ):
        score -= 10.0

    if annotation and annotation.note:
        if _normalize(item.category_l1) in query_features.note_normalized:
            score += 10.0
            reasons.append("人工说明明确命中一级分类")
        if _normalize(item.category_l2) in query_features.note_normalized:
            score += 12.0
            reasons.append("人工说明明确命中二级分类")
        if (
            title
            and len(title) >= 3
            and title in query_features.note_normalized
        ):
            score += 14.0
            reasons.append("人工说明明确命中标准项")

    return score, reasons


def _candidate_payload(
    item: StandardCatalogItem,
    *,
    rank: int,
    score: float,
    lexical_score: float,
    embedding_score: float | None,
    reasons: list[str],
) -> dict[str, Any]:
    return {
        "rank": rank,
        "score": round(score, 4),
        "lexical_score": round(lexical_score, 4),
        "embedding_score": (
            round(embedding_score, 6)
            if embedding_score is not None
            else None
        ),
        "match_reasons": reasons,
        **asdict(item),
    }


def map_atomic_units_to_standards(
    atomic_units: list[dict[str, Any]],
    standard_catalog: list[StandardCatalogItem],
    annotations: dict[str, HumanAnnotation] | None = None,
    *,
    top_k: int = 5,
    embedding_client: EmbeddingProvider | None = None,
    progress_callback: Callable[[str, int, int], None] | None = None,
) -> dict[str, Any]:
    annotations = annotations or {}
    requested_products = {
        canonical_product_name(
            unit.get("product_category_code") or unit.get("product_category")
        )
        for unit in atomic_units
        if _text(unit.get("product_category_code") or unit.get("product_category"))
    }
    requested_products.discard("待确认")
    product_standards: dict[str, list[StandardCatalogItem]] = {}
    for item in standard_catalog:
        product = _product_from_scope(item.scope)
        if (
            item.standard_id.startswith("RAW-")
            and item.category_l1 in {product, "模块"}
        ):
            continue
        if product not in requested_products:
            continue
        product_standards.setdefault(product, []).append(item)

    effective_units: list[dict[str, Any]] = []
    queries: list[str] = []
    standard_texts: dict[str, list[str]] = {}
    standard_lexical_features: dict[str, list[_StandardLexicalFeatures]] = {}
    for product, standards in product_standards.items():
        standard_texts[product] = [
            _embedding_standard_text(item)
            for item in standards
        ]
        standard_lexical_features[product] = [
            _standard_lexical_features(item)
            for item in standards
        ]

    for source_unit in atomic_units:
        atomic_id = _text(source_unit.get("unit_id") or source_unit.get("atomic_id"))
        annotation = annotations.get(atomic_id)
        source_product = canonical_product_name(
            source_unit.get("product_category_code")
            or source_unit.get("product_category")
        )
        standards = product_standards.get(source_product, [])
        effective = build_effective_atomic_unit(source_unit, annotation, standards)
        effective_units.append(effective)
        queries.append(build_atomic_query(effective, annotation))

    query_lexical_features = [
        _query_lexical_features(
            effective,
            query,
            annotations.get(
                _text(source_unit.get("unit_id") or source_unit.get("atomic_id"))
            ),
        )
        for source_unit, effective, query in zip(
            atomic_units,
            effective_units,
            queries,
        )
    ]
    lexical_rankings: list[
        list[tuple[StandardCatalogItem, int, float, list[str]]]
    ] = []
    for index, (source_unit, effective) in enumerate(
        zip(atomic_units, effective_units)
    ):
        atomic_id = _text(source_unit.get("unit_id") or source_unit.get("atomic_id"))
        annotation = annotations.get(atomic_id)
        product = canonical_product_name(
            effective.get("product_category_code")
            or effective.get("product_category")
        )
        standards = product_standards.get(product, [])
        lexical_features = standard_lexical_features.get(product, [])
        ranking: list[
            tuple[StandardCatalogItem, int, float, list[str]]
        ] = []
        for standard_index, item in enumerate(standards):
            lexical_score, reasons = _lexical_score(
                effective,
                item,
                annotation,
                query_lexical_features[index],
                lexical_features[standard_index],
            )
            ranking.append(
                (item, standard_index, lexical_score, reasons)
            )
        ranking.sort(key=lambda entry: entry[2], reverse=True)
        lexical_rankings.append(ranking)
        if progress_callback:
            progress_callback(
                "词法候选召回",
                index + 1,
                len(atomic_units),
            )

    embedding_status = "not_configured"
    embedding_error = ""
    unit_vectors: list[list[float]] = []
    standard_vector_by_key: dict[tuple[str, int], list[float]] = {}
    embedding_shortlist_size = max(5, top_k)
    if embedding_client is not None:
        try:
            unit_vectors = embedding_client.embed_texts(
                queries,
                progress_callback=(
                    (
                        lambda completed, total: progress_callback(
                            "原子知识Embedding",
                            completed,
                            total,
                        )
                    )
                    if progress_callback
                    else None
                ),
            )
            shortlist_keys: list[tuple[str, int]] = []
            seen_shortlist_keys: set[tuple[str, int]] = set()
            for effective, ranking in zip(effective_units, lexical_rankings):
                product = canonical_product_name(
                    effective.get("product_category_code")
                    or effective.get("product_category")
                )
                for _item, standard_index, _score, _reasons in ranking[
                    :embedding_shortlist_size
                ]:
                    key = (product, standard_index)
                    if key not in seen_shortlist_keys:
                        seen_shortlist_keys.add(key)
                        shortlist_keys.append(key)
            shortlist_vectors = embedding_client.embed_texts(
                [
                    standard_texts[product][standard_index]
                    for product, standard_index in shortlist_keys
                ],
                progress_callback=(
                    (
                        lambda completed, total: progress_callback(
                            "候选标准Embedding",
                            completed,
                            total,
                        )
                    )
                    if progress_callback
                    else None
                ),
            )
            standard_vector_by_key = dict(
                zip(shortlist_keys, shortlist_vectors)
            )
            embedding_status = "used"
        except EmbeddingError as exc:
            embedding_status = "fallback_lexical"
            embedding_error = str(exc)

    records: list[dict[str, Any]] = []
    for index, (source_unit, effective, query, lexical_ranking) in enumerate(
        zip(
            atomic_units,
            effective_units,
            queries,
            lexical_rankings,
        )
    ):
        atomic_id = _text(source_unit.get("unit_id") or source_unit.get("atomic_id"))
        annotation = annotations.get(atomic_id)
        product = canonical_product_name(
            effective.get("product_category_code")
            or effective.get("product_category")
        )
        ranking_to_score = (
            lexical_ranking[:embedding_shortlist_size]
            if embedding_status == "used"
            else lexical_ranking
        )
        scored: list[
            tuple[StandardCatalogItem, float, float, float | None, list[str]]
        ] = []
        for item, standard_index, lexical_score, reasons in ranking_to_score:
            embedding_score: float | None = None
            combined_score = lexical_score
            if embedding_status == "used":
                embedding_score = _vector_cosine(
                    unit_vectors[index],
                    standard_vector_by_key[(product, standard_index)],
                )
                combined_score = (
                    lexical_score * 0.58
                    + embedding_score * 100.0 * 0.42
                )
                if embedding_score >= 0.7:
                    reasons = [*reasons, "Embedding语义相似度较高"]
            scored.append(
                (
                    item,
                    combined_score,
                    lexical_score,
                    embedding_score,
                    reasons,
                )
            )
        if progress_callback:
            progress_callback(
                "融合排序",
                index + 1,
                len(atomic_units),
            )
        scored.sort(key=lambda entry: entry[1], reverse=True)
        candidates = [
            _candidate_payload(
                item,
                rank=rank,
                score=combined_score,
                lexical_score=lexical_score,
                embedding_score=embedding_score,
                reasons=reasons,
            )
            for rank, (
                item,
                combined_score,
                lexical_score,
                embedding_score,
                reasons,
            ) in enumerate(scored[: max(1, top_k)], start=1)
        ]
        top_gap = (
            candidates[0]["score"] - candidates[1]["score"]
            if len(candidates) > 1
            else candidates[0]["score"]
            if candidates
            else 0.0
        )
        manual_decision = annotation.decision if annotation else ""
        review_required = (
            not candidates
            or manual_decision in {"不确定", "需要拆分"}
            or bool(effective.get("requires_review"))
            or candidates[0]["score"] < 55.0
            or top_gap < 4.0
        )
        records.append(
            {
                "atomic_id": atomic_id,
                "sample_id": _text(source_unit.get("sample_id")),
                "original_unit": source_unit,
                "effective_unit": effective,
                "human_annotation": asdict(annotation) if annotation else {},
                "manual_merge_target": _manual_merge_target(annotation),
                "retrieval_query": query,
                "candidate_count": len(candidates),
                "top_score_gap": round(top_gap, 4),
                "review_required": review_required,
                "recommendation_status": (
                    "candidate_only"
                    if review_required
                    else "high_confidence"
                ),
                "candidates": candidates,
            }
        )

    return {
        "metadata": {
            "atomic_unit_count": len(atomic_units),
            "standard_count": len(standard_catalog),
            "mapped_count": sum(bool(record["candidates"]) for record in records),
            "review_required_count": sum(record["review_required"] for record in records),
            "embedding_status": embedding_status,
            "embedding_error": embedding_error,
            "embedding_shortlist_size": embedding_shortlist_size,
            "embedded_candidate_standard_count": len(
                standard_vector_by_key
            ),
            "top_k": max(1, top_k),
            "rule": "正式品类过滤 + 人工修正字段 + 结构化词组/阈值召回；Embedding可用时融合排序",
        },
        "records": records,
    }
