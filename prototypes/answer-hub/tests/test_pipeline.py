from __future__ import annotations

import json
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from answer_hub.catalog import StandardCatalogItem, load_standard_catalog
from answer_hub.workflow import (
    build_topic_review_rows,
    evaluate_review_rows,
    finalize_topic_review_rows,
    filter_source_rows_by_product_type,
    filter_preprocessed_rows_for_model,
    initial_label_from_workbook,
    preprocess_source_rows,
    initial_label_rows,
    finalize_review_rows,
    retrieve_standard_matches,
    write_candidate_knowledge_workbook,
)


SOURCE_HEADERS = [
    "序号",
    "上传者",
    "分析时间",
    "工单ID",
    "回收单号",
    "聊天内容",
    "图片链接",
    "核心问题",
    "判定结论",
    "判定依据",
    "产品类型",
    "一级分类",
    "二级分类",
    "参考话术",
]


def _write_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "共享数据汇总"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


class WorkflowTests(unittest.TestCase):
    def test_filter_source_rows_by_phone_product_type(self) -> None:
        selected, excluded = filter_source_rows_by_product_type(
            [
                {"序号": 1, "产品类型": "手机"},
                {"序号": 2, "产品类型": "平板"},
                {"序号": 3, "产品类型": ""},
            ],
            "手机",
        )
        self.assertEqual([row["序号"] for row in selected], [1])
        self.assertEqual([row["序号"] for row in excluded], [2, 3])
        self.assertIn("产品类型不匹配", excluded[0]["排除原因"])

    def test_preprocess_source_rows_adds_cleaning_fields(self) -> None:
        rows = [
            {
                "序号": 1,
                "上传者": "测试1 ",
                "聊天内容": "第一行\n\n第二行",
                "图片链接": "a.jpg， b.jpg\nb.jpg",
                "核心问题": "  屏幕这个点怎么判，是色斑吗  ",
                "判定结论": "  该屏幕上的点应被判定为色斑。  ",
                "判定依据": "  平台标准依据：色斑属于显示问题。  ",
                "一级分类": "显示问题",
                "二级分类": "色斑",
                "参考话术": "  根据图片判断为色斑。  ",
            }
        ]
        processed = preprocess_source_rows(rows)
        self.assertEqual(processed[0]["预处理状态"], "preprocessed")
        self.assertEqual(processed[0]["可进入模型初标"], "是")
        self.assertEqual(processed[0]["上传者"], "测试1")
        self.assertEqual(processed[0]["聊天内容"], "第一行\n第二行")
        self.assertEqual(processed[0]["图片链接"], "a.jpg\nb.jpg")
        self.assertEqual(processed[0]["核心问题"], "屏幕这个点怎么判，是色斑吗")
        self.assertEqual(processed[0]["数据ID"], "row-00001")

    def test_missing_critical_fields_are_excluded_from_model_labeling(self) -> None:
        processed = preprocess_source_rows(
            [
                {
                    "序号": 1,
                    "工单ID": "W-MISSING",
                    "聊天内容": "屏幕有异常",
                    "核心问题": "",
                    "判定结论": "待确认",
                    "判定依据": "图片不足",
                    "产品类型": "手机",
                    "一级分类": "显示问题",
                    "二级分类": "色斑",
                }
            ]
        )
        eligible, excluded = filter_preprocessed_rows_for_model(processed)
        self.assertEqual(eligible, [])
        self.assertEqual(excluded[0]["数据ID"], "W-MISSING")
        self.assertIn("核心问题", excluded[0]["排除原因"])

    def test_missing_chat_content_stays_eligible_but_requires_review(self) -> None:
        source = {
            "序号": 1,
            "工单ID": "W-NO-CHAT",
            "聊天内容": "",
            "核心问题": "手机屏幕色斑如何判定",
            "判定结论": "判定为色斑",
            "判定依据": "色斑属于显示问题",
            "产品类型": "手机",
            "一级分类": "显示问题",
            "二级分类": "色斑",
        }
        processed = preprocess_source_rows([source])
        eligible, excluded = filter_preprocessed_rows_for_model(processed)
        self.assertEqual(len(eligible), 1)
        self.assertEqual(excluded, [])
        self.assertIn("缺少原始聊天上下文", processed[0]["预处理备注"])

        standard = load_standard_catalog(
            Path(__file__).resolve().parents[1] / "examples" / "standard_catalog.example.json"
        )[0]
        candidate = initial_label_rows(processed, [standard])[0]
        self.assertEqual(candidate["是否重点复核"], "是")

    def test_initial_label_rows_uses_standard_catalog(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            standard_path = tmp_path / "standards.json"
            standard_path.write_text(
                json.dumps(
                    [
                        {
                            "standard_id": "STD-001",
                            "title": "屏幕色斑如何判定",
                            "category_l1": "显示问题",
                            "category_l2": "色斑",
                            "keywords": ["色斑", "屏幕"],
                            "scope": "适用于屏幕显示异常判断",
                            "response_snippet": "色斑判定",
                        }
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            standards = load_standard_catalog(standard_path)
            rows = [
                {
                    "序号": 1,
                    "上传者": "测试1",
                    "工单ID": "W-001",
                    "聊天内容": "屏幕这个点怎么判，是色斑吗",
                    "核心问题": "屏幕这个点怎么判，是色斑吗",
                    "判定结论": "该屏幕上的点应被判定为色斑。",
                    "判定依据": "平台标准依据：色斑属于显示问题。",
                    "产品类型": "电脑",
                    "一级分类": "显示问题",
                    "二级分类": "色斑",
                    "参考话术": "根据图片判断为色斑。",
                }
            ]
            preprocessed = preprocess_source_rows(rows)
            labeled = initial_label_rows(preprocessed, standards)
            self.assertEqual(labeled[0]["模型一级分类"], "显示问题")
            self.assertEqual(labeled[0]["模型二级分类"], "色斑")
            self.assertEqual(labeled[0]["流程状态"], "review_pending")
            self.assertEqual(labeled[0]["模型阶段状态"], "model_labeled")
            self.assertEqual(labeled[0]["预处理状态"], "preprocessed")
            self.assertEqual(labeled[0]["是否重点复核"], "否")
            self.assertEqual(labeled[0]["模型知识形态"], "流程方法")
            self.assertEqual(labeled[0]["知识分类"], "检测方法")
            self.assertEqual(labeled[0]["生效状态"], "待审核")
            self.assertIn("屏幕显示异常如何通过图片核验", labeled[0]["主标题"])
            self.assertIn("核验流程", labeled[0]["知识内容"])

    def test_candidate_knowledge_card_does_not_repeat_source_narrative(self) -> None:
        standard = load_standard_catalog(
            Path(__file__).resolve().parents[1] / "examples" / "standard_catalog.example.json"
        )[0]
        row = preprocess_source_rows(
            [
                {
                    "工单ID": "PHONE-CARD-001",
                    "核心问题": "回收师在现场遇到屏幕色斑问题，希望获得如何判定的标准指导。",
                    "判定结论": "屏幕色斑判定为显示问题。",
                    "判定依据": "平台标准依据：色斑属于显示问题。事实核查结果：图片显示局部颜色异常。",
                    "产品类型": "手机",
                    "一级分类": "显示问题",
                    "二级分类": "色斑",
                    "参考话术": "按色斑标准处理。",
                }
            ]
        )[0]
        candidate = initial_label_rows([row], [standard])[0]
        self.assertEqual(candidate["主标题"], "屏幕显示异常如何通过图片核验")
        self.assertEqual(candidate["知识来源"], "方向二会话候选")
        self.assertEqual(candidate["关联标准项"], "")
        self.assertNotIn("回收师在现场遇到", candidate["知识内容"])
        self.assertNotIn("屏幕色斑判定为显示问题", candidate["知识内容"])

    def test_candidate_uses_primary_standard_path_and_non_narrative_subtitles(self) -> None:
        standard = StandardCatalogItem(
            standard_id="",
            title="外壳磕碰/掉漆（单选）相关问题如何处理",
            category_l1="中框及外壳外观",
            category_l2="外壳磕碰/掉漆（单选）",
            knowledge_type="场景判定",
            standard_path=(
                "【中框及外壳外观】-【外壳磕碰/掉漆（单选）】\n"
                "【中框及外壳外观】-【外壳磕碰/掉漆（单选）】-【最大直径≤3mm且2mm以上数量≤5】"
            ),
            keywords=["中框及外壳外观", "外壳磕碰/掉漆（单选）"],
            scope="手机",
            response_snippet="场景处理：\n- 中框磨损按外壳磕碰标准核验。",
            status="生效中",
            version="SJ-HSYJBZ-2026009",
        )
        row = preprocess_source_rows(
            [
                {
                    "工单ID": "PHONE-PRIMARY-PATH-001",
                    "聊天内容": "中框磨损怎么判",
                    "核心问题": "回收师在现场回收中遇到中框磨损，希望获得判定指导。",
                    "判定结论": "中框磨损应判定为外壳磕碰。",
                    "判定依据": "平台标准依据：中框磨损按外壳磕碰处理。事实核查结果：图片显示中框磨损。",
                    "产品类型": "手机",
                    "一级分类": "外观问题",
                    "二级分类": "中框及外壳外观",
                }
            ]
        )[0]

        candidate = initial_label_rows([row], [standard])[0]

        self.assertEqual(candidate["主标题"], "外壳磕碰/掉漆如何通过图片核验")
        self.assertEqual(candidate["关联标准项"], "【中框及外壳外观】-【外壳磕碰/掉漆（单选）】")
        self.assertNotIn("\n", candidate["关联标准项"])
        self.assertNotIn("回收师", candidate["副标题"])
        self.assertLessEqual(len(candidate["检索关键词"]), 160)
        self.assertNotIn("事实核查结果", candidate["知识内容"])

    def test_model_query_generates_reusable_confirmation_flow(self) -> None:
        standard = StandardCatalogItem(
            standard_id="",
            title="设备机型应该如何选择",
            category_l1="基本情况",
            category_l2="机型",
            knowledge_type="检测方法",
            standard_path="【基本情况】-【机型】",
            keywords=["设备机型", "型号"],
            scope="通用",
            response_snippet="以实物特征和官方查询结果确认设备机型。",
            status="生效中",
            version="SJ-HSYJBZ-2026009",
        )
        row = preprocess_source_rows(
            [
                {
                    "工单ID": "PHONE-MODEL-QUERY-001",
                    "聊天内容": "这个手机型号怎么确认",
                    "核心问题": "回收师需要确认设备机型和型号。",
                    "判定结论": "该设备型号为某机型。",
                    "判定依据": "平台标准依据：应以官方查询和实物特征为准。",
                    "产品类型": "手机",
                    "一级分类": "基本情况",
                    "二级分类": "机型",
                }
            ]
        )[0]

        candidate = initial_label_rows([row], [standard])[0]

        self.assertEqual(candidate["主标题"], "设备机型如何查询与确认")
        self.assertEqual(candidate["候选知识形态"], "流程方法")
        self.assertIn("查询流程", candidate["知识内容"])
        self.assertNotIn("该设备型号为某机型", candidate["知识内容"])

    def test_function_problem_uses_function_verification_flow(self) -> None:
        standard = StandardCatalogItem(
            standard_id="",
            title="摄像头功能如何核验",
            category_l1="功能问题",
            category_l2="摄像头功能",
            knowledge_type="检测方法",
            standard_path="【功能检测】-【摄像头功能】",
            keywords=["摄像头", "拍照"],
            scope="手机",
            response_snippet="按前后摄像头拍照、录像和对焦结果核验。",
            status="生效中",
            version="SJ-HSYJBZ-2026009",
        )
        row = preprocess_source_rows(
            [
                {
                    "工单ID": "PHONE-CAMERA-001",
                    "聊天内容": "后摄拍照模糊怎么确认",
                    "核心问题": "设备后摄像头拍照模糊，需要确认是否属于摄像头功能异常。",
                    "判定结论": "后摄可能存在异常。",
                    "判定依据": "平台标准依据：需按拍照、录像和对焦结果核验。",
                    "产品类型": "手机",
                    "一级分类": "功能问题",
                    "二级分类": "摄像头功能",
                }
            ]
        )[0]

        candidate = initial_label_rows([row], [standard])[0]

        self.assertEqual(candidate["主标题"], "摄像头功能如何核验")
        self.assertIn("功能核验流程", candidate["知识内容"])
        self.assertNotIn("外观异常", candidate["知识内容"])

    def test_explicit_boundary_case_keeps_specific_judgment(self) -> None:
        standard = StandardCatalogItem(
            standard_id="",
            title="屏幕坏点和漏液如何区分",
            category_l1="显示问题",
            category_l2="屏幕异常",
            knowledge_type="场景判定",
            standard_path="【屏幕及正面外观】-【显示异常】",
            keywords=["坏点", "漏液"],
            scope="手机",
            response_snippet="按坏点和漏液的定义及边界条件进行区分。",
            status="生效中",
            version="SJ-HSYJBZ-2026009",
        )
        row = preprocess_source_rows(
            [
                {
                    "工单ID": "PHONE-BOUNDARY-001",
                    "聊天内容": "这个是坏点还是漏液",
                    "核心问题": "屏幕坏点还是漏液如何区分。",
                    "判定结论": "根据边界条件判定为坏点。",
                    "判定依据": "平台标准依据：坏点与漏液有明确现象定义。",
                    "产品类型": "手机",
                    "一级分类": "显示问题",
                    "二级分类": "屏幕异常",
                }
            ]
        )[0]

        candidate = initial_label_rows([row], [standard])[0]

        self.assertEqual(candidate["候选知识形态"], "具体判定")
        self.assertEqual(candidate["知识分类"], "场景判定")
        self.assertIn("场景结论", candidate["知识内容"])

    def test_candidate_export_contains_only_knowledge_master_columns(self) -> None:
        candidate = {
            "主标题": "设备机型如何查询与确认",
            "副标题": "设备机型怎么确认",
            "知识内容": "查询流程：\n1. 查看关于本机。",
            "知识分类": "检测方法",
            "知识来源": "方向二会话候选",
            "关联标准项": "【基本情况】-【机型】",
            "适用范围": "手机",
            "生效状态": "待审核",
            "来源版本": "SJ-HSYJBZ-2026009",
            "变更类型": "新增",
            "失效原因": "",
            "检索关键词": "设备机型如何查询与确认",
            "校验备注": "需人工重点复核",
            "候选ID": "KC-001",
            "模型运行ID": "run-001",
        }
        expected_headers = [
            "主标题", "副标题", "知识内容", "知识分类", "知识来源", "关联标准项", "适用范围",
            "生效状态", "来源版本", "变更类型", "失效原因", "检索关键词", "校验备注",
        ]
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "candidate_knowledge.xlsx"
            write_candidate_knowledge_workbook([candidate], output_path)
            workbook = load_workbook(output_path, read_only=True, data_only=True)
            worksheet = workbook["候选知识"]
            headers = [cell.value for cell in next(worksheet.iter_rows(values_only=False))]
            values = [cell.value for cell in next(worksheet.iter_rows(min_row=2, values_only=False))]
            workbook.close()

        self.assertEqual(headers, expected_headers)
        self.assertEqual(values[0], candidate["主标题"])
        self.assertEqual(len(values), len(expected_headers))

    def test_candidate_export_groups_duplicate_topics_into_theme_rows(self) -> None:
        rows = [
            {
                "主标题": "设备机型如何查询与确认",
                "副标题": "设备机型怎么确认",
                "知识内容": "查询流程：\n1. 查看关于本机。",
                "知识分类": "检测方法",
                "知识来源": "方向二会话候选",
                "关联标准项": "【基本情况】-【机型】",
                "适用范围": "手机",
                "生效状态": "待审核",
                "来源版本": "SJ-HSYJBZ-2026009",
                "变更类型": "新增",
                "失效原因": "",
                "检索关键词": "设备机型如何查询与确认",
                "校验备注": "来源记录ID：A",
                "来源记录ID": "A",
            },
            {
                "主标题": "设备机型如何查询与确认",
                "副标题": "设备机型怎么确认",
                "知识内容": "查询流程：\n1. 查看关于本机。",
                "知识分类": "检测方法",
                "知识来源": "方向二会话候选",
                "关联标准项": "【基本情况】-【机型】",
                "适用范围": "手机",
                "生效状态": "待审核",
                "来源版本": "SJ-HSYJBZ-2026009",
                "变更类型": "新增",
                "失效原因": "",
                "检索关键词": "设备机型如何查询与确认",
                "校验备注": "来源记录ID：B",
                "来源记录ID": "B",
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "candidate_knowledge.xlsx"
            write_candidate_knowledge_workbook(rows, output_path)
            workbook = load_workbook(output_path, read_only=True, data_only=True)
            worksheet = workbook["候选知识"]
            data = list(worksheet.iter_rows(values_only=True))
            workbook.close()

        self.assertEqual(len(data), 2)
        headers = data[0]
        first_row = data[1]
        self.assertEqual(headers[0], "主标题")
        self.assertIn("主题聚合样本数：2", first_row[12])
        self.assertIn("来源记录ID：A、B", first_row[12])

    def test_topic_review_excludes_structured_only_records(self) -> None:
        base = {
            "产品类型": "手机",
            "一级分类": "基本情况",
            "二级分类": "机型",
            "主标准路径": "【基本情况】-【机型】",
            "问题意图": "信息查询",
            "对象/部位": "机型",
            "异常现象": "机型",
            "解题方式": "官方信息查询与实物核对",
        }
        conversation_row = {
            **base,
            "数据ID": "A",
            "工单ID": "W-A",
            "聊天内容": "这个手机型号怎么确认",
            "图片处理状态": "无图片链接（文本初标）",
            "核心问题": "设备型号如何确认",
        }
        structured_only_row = {
            **base,
            "数据ID": "B",
            "工单ID": "W-B",
            "聊天内容": "",
            "图片处理状态": "无图片链接（文本初标）",
            "核心问题": "设备型号如何确认",
        }
        second_conversation_row = {
            **conversation_row,
            "数据ID": "C",
            "工单ID": "W-C",
        }

        topics, mapping, gaps, pending = build_topic_review_rows(
            [conversation_row, second_conversation_row, structured_only_row],
            use_mimo=False,
        )

        self.assertEqual(len(topics), 1)
        self.assertEqual(topics[0]["主题样本数"], 2)
        self.assertEqual(topics[0]["主题来源记录ID"], "A\nC")
        self.assertEqual(len(mapping), 2)
        self.assertEqual(len(gaps), 1)
        self.assertEqual(pending, [])
        self.assertEqual(gaps[0]["数据ID"], "B")
        self.assertIn("缺少原始聊天内容和可用图片", gaps[0]["证据缺口原因"])

    def test_topic_review_finalization_uses_direct_candidate_edits_and_builds_training_sample(self) -> None:
        topic_row = {
            "主题ID": "TOP-001",
            "主题样本数": 2,
            "主题来源记录ID": "A\nB",
            "主题证据等级": "完整会话",
            "主题证据摘要": "A | 完整会话",
            "主题检索标准Top5": "STD-001 | 机型",
            "主题标准版本": "SJ-HSYJBZ-2026009",
            "主标题": "手机机型如何查询与确认",
            "副标题": "设备型号怎么确认",
            "知识内容": "查询流程：\n1. 查看关于本机。",
            "知识分类": "检测方法",
            "知识来源": "方向二主题候选",
            "关联标准项": "【基本情况】-【机型】",
            "适用范围": "手机",
            "生效状态": "待审核",
            "来源版本": "SJ-HSYJBZ-2026009",
            "变更类型": "新增",
            "失效原因": "",
            "检索关键词": "设备机型 | 型号",
            "校验备注": "主题聚合样本数：2",
            "审核结论": "修改后通过",
            "审核备注": "标题更准确",
            "错误类型": "标题不准",
            "错误原因": "补充手机范围",
            "是否进入训练集": "是",
            "审核人": "operator",
            "审核时间": "2026-07-13 18:00:00",
        }

        final_rows, feedback_rows, training_rows = finalize_topic_review_rows([topic_row])

        self.assertEqual(len(final_rows), 1)
        self.assertEqual(final_rows[0]["主标题"], "手机机型如何查询与确认")
        self.assertEqual(final_rows[0]["生效状态"], "待审核")
        self.assertEqual(len(feedback_rows), 1)
        self.assertEqual(feedback_rows[0]["错误类型"], "标题不准")
        self.assertEqual(len(training_rows), 1)
        self.assertEqual(training_rows[0]["target"]["主标题"], "手机机型如何查询与确认")

    def test_ingest_writes_topic_review_and_theme_candidate_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_path = tmp_path / "source.xlsx"
            standards_path = tmp_path / "standards.json"
            output_dir = tmp_path / "output"
            _write_workbook(
                source_path,
                SOURCE_HEADERS,
                [[
                    1, "tester", "2026-07-13", "W-TOPIC-001", "", "这个手机型号怎么确认",
                    "", "手机机型和型号如何确认", "需要通过系统信息和官方查询确认",
                    "平台标准依据：以实物特征和官方查询结果确认设备机型。",
                    "手机", "基本情况", "机型", "",
                ], [
                    2, "tester", "2026-07-13", "W-TOPIC-002", "", "手机型号怎么查",
                    "", "手机机型和型号如何确认", "需要通过系统信息和官方查询确认",
                    "平台标准依据：以实物特征和官方查询结果确认设备机型。",
                    "手机", "基本情况", "机型", "",
                ]],
            )
            standards_path.write_text(
                json.dumps(
                    [{
                        "主标题": "设备机型应该如何选择",
                        "知识分类": "检测方法",
                        "关联标准项": "【基本情况】-【机型】",
                        "适用范围": "手机",
                        "生效状态": "生效中",
                        "来源版本": "SJ-HSYJBZ-2026009",
                        "检索关键词": "设备机型 型号",
                        "知识内容": "以实物特征和官方查询结果确认设备机型。",
                    }],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            summary = initial_label_from_workbook(
                source_path=source_path,
                standards_path=standards_path,
                output_dir=output_dir,
                product_type="手机",
                use_mimo=False,
                audit_db_path=tmp_path / "audit.db",
            )
            topic_workbook = Path(summary["topic_review_file"])
            candidate_workbook = Path(summary["candidate_output_file"])
            self.assertTrue(topic_workbook.exists())
            self.assertTrue(candidate_workbook.exists())
            self.assertEqual(summary["topic_rows"], 1)
            self.assertEqual(summary["evidence_gap_rows"], 0)
            topic_book = load_workbook(topic_workbook, read_only=True, data_only=True)
            self.assertIn("topic_review_queue", topic_book.sheetnames)
            self.assertIn("topic_source_mapping", topic_book.sheetnames)
            self.assertIn("topic_model_drafts", topic_book.sheetnames)
            topic_book.close()

    def test_retrieval_collapses_standard_chunks_with_same_path(self) -> None:
        standards = [
            StandardCatalogItem(
                standard_id="",
                title="屏幕色斑标准定义",
                category_l1="显示问题",
                category_l2="色斑",
                knowledge_type="标准定义",
                standard_path="【显示问题】-【色斑】",
                keywords=["色斑", "屏幕"],
                scope="手机",
                response_snippet="色斑属于显示问题。",
                status="published",
                version="v1",
            ),
            StandardCatalogItem(
                standard_id="",
                title="屏幕色斑检测方法",
                category_l1="显示问题",
                category_l2="色斑",
                knowledge_type="检测方法",
                standard_path="【显示问题】-【色斑】",
                keywords=["色斑", "屏幕"],
                scope="手机",
                response_snippet="白屏核验。",
                status="published",
                version="v1",
            ),
        ]
        matches = retrieve_standard_matches(
            {
                "核心问题": "手机屏幕色斑如何判定",
                "产品类型": "手机",
                "一级分类": "显示问题",
                "二级分类": "色斑",
            },
            standards,
        )
        self.assertEqual(len(matches), 1)

    def test_finalize_review_rows_exports_feedback_and_published(self) -> None:
        row = {
            "数据ID": "W-001",
            "工单ID": "W-001",
            "模型主标题": "屏幕色斑如何判定",
            "模型一级分类": "显示问题",
            "模型二级分类": "色斑",
            "模型关联标准": "STD-001(2.5)",
            "CZ复核结论": "修改后通过",
            "CZ主标题": "屏幕色斑判定标准",
            "CZ一级分类": "显示问题",
            "CZ二级分类": "色斑",
            "CZ关联标准": "STD-001(2.5)",
            "CZ复核备注": "标题改写",
            "错误类型": "标题不准",
            "错误原因": "标题过于口语化",
            "是否进入再训练样本": "是",
            "审核人": "cz",
            "审核时间": "2026-07-12 10:00:00",
        }
        published_rows, feedback_rows = finalize_review_rows([row])
        self.assertEqual(len(published_rows), 1)
        self.assertEqual(len(feedback_rows), 1)
        self.assertEqual(published_rows[0]["生效状态"], "published")
        self.assertEqual(feedback_rows[0]["错误类型"], "标题不准")

    def test_cz_catalog_filters_inactive_standards_and_parses_path_categories(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "standards.json"
            path.write_text(
                json.dumps(
                    [
                        {
                            "主标题": "屏幕色斑判定",
                            "关联标准项": "【显示问题】-【色斑】",
                            "生效状态": "生效中",
                            "来源版本": "v-active",
                            "检索关键词": "屏幕 色斑",
                        },
                        {
                            "主标题": "旧版屏幕色斑判定",
                            "关联标准项": "【显示问题】-【色斑】",
                            "生效状态": "已失效",
                            "来源版本": "v-old",
                            "检索关键词": "屏幕 色斑",
                        },
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            standards = load_standard_catalog(path)
            self.assertEqual(len(standards), 1)
            self.assertEqual(standards[0].category_l1, "显示问题")
            self.assertEqual(standards[0].category_l2, "色斑")
            self.assertEqual(standards[0].version, "v-active")

    def test_evaluate_review_rows_calculates_review_metrics(self) -> None:
        report = evaluate_review_rows(
            [
                {
                    "CZ复核结论": "修改后通过",
                    "模型关联标准": "STD-001(8.0)",
                    "CZ关联标准": "STD-001",
                    "检索标准Top5": "STD-001 | 屏幕色斑判定 | 分数:8.0",
                    "模型一级分类": "显示问题",
                    "CZ一级分类": "显示问题",
                    "模型二级分类": "色斑",
                    "CZ二级分类": "色斑",
                    "模型主标题": "屏幕色斑如何判定",
                    "CZ主标题": "屏幕色斑判定",
                    "是否重点复核": "否",
                    "错误类型": "标题不准",
                },
                {
                    "CZ复核结论": "标记Bad Case",
                    "模型关联标准": "",
                    "CZ关联标准": "",
                    "检索标准Top5": "未搜索到相关知识（待人工补充）",
                    "模型一级分类": "拆修问题",
                    "CZ一级分类": "拆修问题",
                    "模型二级分类": "屏幕拆修",
                    "CZ二级分类": "屏幕拆修",
                    "模型主标题": "疑似拆修核验流程",
                    "CZ主标题": "疑似拆修核验流程",
                    "是否重点复核": "是",
                    "错误类型": "标准未覆盖/标准召回不足",
                },
            ]
        )
        self.assertEqual(report["reviewed_rows"], 2)
        self.assertEqual(report["standard_top5_hit_rate"]["rate"], 1.0)
        self.assertEqual(report["model_standard_reference_match_rate"]["rate"], 1.0)
        self.assertEqual(report["title_modification_rate"]["rate"], 0.5)
        self.assertEqual(report["rejected_or_bad_case_rate"]["rate"], 0.5)
        self.assertEqual(report["standard_uncovered_rate"]["rate"], 0.5)


if __name__ == "__main__":
    unittest.main()
