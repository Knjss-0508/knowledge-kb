from __future__ import annotations

from openpyxl import Workbook

from scripts import eval_case4_cluster_accuracy as accuracy


def test_duplicate_source_ids_are_matched_within_product_category(tmp_path) -> None:
    workbook_path = tmp_path / "cluster_result.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "聚类结果"
    worksheet.append(["聚类主题ID", "产品类型", "主题来源记录ID"])
    worksheet.append(["TOP-PHONE", "手机", "DUPLICATE-ID"])
    worksheet.append(["TOP-CAMERA", "相机机身", "DUPLICATE-ID"])
    workbook.save(workbook_path)

    gold_rows = [
        {
            "真实会话ID（文本）": "ID=DUPLICATE-ID",
            "品类": "手机",
            "拆分序号": 1,
        },
        {
            "真实会话ID（文本）": "ID=DUPLICATE-ID",
            "品类": "相机机身",
            "拆分序号": 1,
        },
    ]

    predictions = accuracy._model_predictions(
        gold_rows,
        model_workbook=workbook_path,
        model_json=None,
    )

    assert accuracy._prediction_for_row(gold_rows[0], predictions) == "TOP-PHONE"
    assert accuracy._prediction_for_row(gold_rows[1], predictions) == "TOP-CAMERA"


def test_evaluate_keeps_duplicate_source_ids_separate_by_product(tmp_path) -> None:
    model_path = tmp_path / "cluster_result.xlsx"
    model_workbook = Workbook()
    model_sheet = model_workbook.active
    model_sheet.title = "聚类结果"
    model_sheet.append(["聚类主题ID", "产品类型", "主题来源记录ID"])
    model_sheet.append(["TOP-PHONE", "手机", "DUPLICATE-ID"])
    model_sheet.append(["TOP-CAMERA", "相机机身", "DUPLICATE-ID"])
    model_workbook.save(model_path)

    gold_path = tmp_path / "gold.xlsx"
    gold_workbook = Workbook()
    gold_sheet = gold_workbook.active
    gold_sheet.title = "原子问题明细"
    gold_sheet.append(
        [
            "人工主题编号",
            "拆分序号",
            "纳入准确率",
            "真实会话ID（文本）",
            "品类",
        ]
    )
    gold_sheet.append(["H-PHONE", 1, "是", "ID=DUPLICATE-ID", "手机"])
    gold_sheet.append(["H-CAMERA", 1, "是", "ID=DUPLICATE-ID", "相机机身"])
    gold_workbook.save(gold_path)

    summary = accuracy.evaluate(gold_path, model_workbook=model_path)

    assert summary["included_rows"] == 2
    assert summary["unmatched_included_rows"] == 0
    assert summary["purity_coverage"]["purity"] == 1.0
    assert summary["product_boundary"]["cross_product_cluster_count"] == 0


def test_evaluate_reports_cross_product_clusters_from_matched_atomic_rows(
    tmp_path,
) -> None:
    model_path = tmp_path / "cluster_result.xlsx"
    model_workbook = Workbook()
    model_sheet = model_workbook.active
    model_sheet.title = "聚类结果"
    model_sheet.append(["聚类主题ID", "产品类型", "主题来源记录ID"])
    model_sheet.append(["TOP-CROSS", "手机", "PHONE-ID"])
    model_sheet.append(["TOP-CROSS", "相机机身", "CAMERA-ID"])
    model_workbook.save(model_path)

    gold_path = tmp_path / "gold.xlsx"
    gold_workbook = Workbook()
    gold_sheet = gold_workbook.active
    gold_sheet.title = "原子问题明细"
    gold_sheet.append(
        [
            "人工主题编号",
            "拆分序号",
            "纳入准确率",
            "真实会话ID（文本）",
            "品类",
        ]
    )
    gold_sheet.append(["H-PHONE", 1, "是", "ID=PHONE-ID", "手机"])
    gold_sheet.append(["H-CAMERA", 1, "是", "ID=CAMERA-ID", "相机机身"])
    gold_workbook.save(gold_path)

    summary = accuracy.evaluate(gold_path, model_workbook=model_path)

    assert summary["product_boundary"] == {
        "cross_product_cluster_count": 1,
        "cross_product_clusters": [
            {
                "model_topic": "TOP-CROSS",
                "product_categories": ["手机", "相机机身"],
                "included_atoms": 2,
            }
        ],
    }


def test_evaluate_reports_cross_business_line_clusters(tmp_path) -> None:
    model_path = tmp_path / "cluster_result.xlsx"
    model_workbook = Workbook()
    model_sheet = model_workbook.active
    model_sheet.title = "聚类结果"
    model_sheet.append(["聚类主题ID", "产品类型", "主题来源记录ID"])
    model_sheet.append(["TOP-CROSS-LINE", "手机", "SELF-ID"])
    model_sheet.append(["TOP-CROSS-LINE", "手机", "AGGREGATE-ID"])
    model_workbook.save(model_path)

    gold_path = tmp_path / "gold.xlsx"
    gold_workbook = Workbook()
    gold_sheet = gold_workbook.active
    gold_sheet.title = "原子问题明细"
    gold_sheet.append(
        [
            "人工主题编号",
            "拆分序号",
            "纳入准确率",
            "真实会话ID（文本）",
            "品类",
            "回收业务层级",
        ]
    )
    gold_sheet.append(
        ["H-SELF", 1, "是", "ID=SELF-ID", "手机", "自营回收"]
    )
    gold_sheet.append(
        [
            "H-AGGREGATE",
            1,
            "是",
            "ID=AGGREGATE-ID",
            "手机",
            "聚合回收",
        ]
    )
    gold_workbook.save(gold_path)

    summary = accuracy.evaluate(gold_path, model_workbook=model_path)

    assert summary["business_line_boundary"] == {
        "cross_business_line_cluster_count": 1,
        "cross_business_line_clusters": [
            {
                "model_topic": "TOP-CROSS-LINE",
                "business_lines": ["聚合回收", "自营回收"],
                "included_atoms": 2,
            }
        ],
    }
