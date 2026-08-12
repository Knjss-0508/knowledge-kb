from __future__ import annotations

from pathlib import Path
import json

from openpyxl import Workbook

from answer_hub.catalog import load_standard_catalog
from answer_hub.standards_compiler import compile_standard_catalog, read_raw_standard_sheet


def _write_raw_standard(path: Path, sheet_name: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["一级分类", "二级分类", "判定标准"])
    sheet.append(["显示问题", "色斑", "白屏下确认色斑的位置、大小和边界。"])
    sheet.append(["", "坏点", "在标准测试画面下记录坏点数量。"])
    sheet.merge_cells("A2:A3")
    workbook.save(path)


def _excel_safe_name(value: str) -> str:
    return value.replace("/", "、")


def test_read_raw_standard_sheet_propagates_vertical_merged_category(tmp_path: Path) -> None:
    path = tmp_path / "phone.xlsx"
    sheet_name = "SJ-HSYJBZ-2026009【5.13以后】"
    _write_raw_standard(path, sheet_name)

    rows = read_raw_standard_sheet(path, sheet_name)

    third_row = next(row for row in rows if row["row_number"] == 3)
    assert third_row["cells"][1] == "显示问题"
    assert third_row["cells"][2] == "坏点"


def test_compile_standard_catalog_includes_all_products(tmp_path: Path) -> None:
    sources: dict[str, Path] = {}
    active_sheets: dict[str, tuple[str, ...]] = {}
    for product_type in ("手机", "智能手表", "平板电脑", "耳机/耳麦"):
        safe_name = _excel_safe_name(product_type)
        sheet_name = f"{safe_name}现行标准"
        source = tmp_path / f"{safe_name}.xlsx"
        _write_raw_standard(source, sheet_name)
        sources[product_type] = source
        active_sheets[product_type] = (sheet_name,)

    output = tmp_path / "active_standards.json"
    summary = compile_standard_catalog(
        sources,
        output,
        active_sheets=active_sheets,
    )

    assert summary["total_items"] == 8
    payload = json.loads(output.read_text(encoding="utf-8"))
    assert {item["scope"] for item in payload["items"]} == {
        "手机-通用",
        "智能手表-通用",
        "平板电脑-通用",
        "耳机/耳麦-通用",
    }
    assert len(load_standard_catalog(output)) == 8


def test_compile_structured_qc_sheet_uses_real_category_path(tmp_path: Path) -> None:
    sources: dict[str, Path] = {}
    active_sheets: dict[str, tuple[str, ...]] = {}
    for product_type in ("手机", "智能手表", "平板电脑", "耳机/耳麦"):
        safe_name = _excel_safe_name(product_type)
        sheet_name = f"{safe_name}-5.13以后"
        source = tmp_path / f"【{safe_name}】.xlsx"
        workbook = Workbook()
        home = workbook.active
        home.title = "首页"
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(
            [
                "",
                "一级类",
                "二级项",
                "程度值",
                "三级值",
                "标准定义",
                "检测方法",
            ]
        )
        sheet.append(
            [
                "",
                "基本情况*（A）",
                "机型",
                "/",
                "/",
                "按实物特征确认机型。",
                "打开关于本机查看。",
            ]
        )
        sheet.append(
            [
                "",
                "",
                "",
                "",
                "",
                "查询结果与实物不一致时以实物为准。",
                "",
            ]
        )
        sheet.append(
            [
                "",
                "设备功能情况",
                "通话功能-SIM卡2（单选）",
                "异常",
                "卡二功能异常",
                "使用正常SIM卡交叉测试。",
                "换卡后仍无法识别则判定异常。",
            ]
        )
        workbook.save(source)
        sources[product_type] = source
        active_sheets[product_type] = (sheet_name,)

    output = tmp_path / "structured.json"
    summary = compile_standard_catalog(
        sources,
        output,
        active_sheets=active_sheets,
    )

    assert summary["total_items"] == 8
    items = load_standard_catalog(output)
    phone_items = [item for item in items if item.scope.startswith("手机-")]
    model_item = next(item for item in phone_items if item.category_l2 == "机型")
    sim_item = next(
        item
        for item in phone_items
        if item.category_l2 == "通话功能-SIM卡2（单选）"
    )
    assert model_item.category_l1 == "基本情况"
    assert "查询结果与实物不一致" in model_item.response_snippet
    assert model_item.standard_path == "【手机】-【基本情况】-【机型】"
    assert sim_item.title == "卡二功能异常"
    assert "【异常】-【卡二功能异常】" in sim_item.standard_path
    assert sim_item.standard_id.startswith("QC-")


def test_compile_standard_catalog_supports_new_configured_products(tmp_path: Path) -> None:
    sources: dict[str, Path] = {}
    active_sheets: dict[str, tuple[str, ...]] = {}
    for product_type in (
        "笔记本",
        "单电/微单机身",
        "单反机身",
        "相机镜头",
    ):
        safe_name = _excel_safe_name(product_type)
        sheet_name = f"{safe_name}现行标准"
        source = tmp_path / f"{safe_name}.xlsx"
        _write_raw_standard(source, sheet_name)
        sources[product_type] = source
        active_sheets[product_type] = (sheet_name,)

    output = tmp_path / "expanded_standards.json"
    summary = compile_standard_catalog(
        sources,
        output,
        active_sheets=active_sheets,
    )

    assert summary["total_items"] == 8
    assert {item.scope for item in load_standard_catalog(output)} == {
        "笔记本-通用",
        "单电/微单机身-通用",
        "单反机身-通用",
        "相机镜头-通用",
    }


def test_compile_standard_catalog_normalizes_clear_active_sheet_aliases(
    tmp_path: Path,
) -> None:
    source = tmp_path / "tablet.xlsx"
    sheet_name = "平板旧配置工作表"
    _write_raw_standard(source, sheet_name)

    output = tmp_path / "tablet.json"
    summary = compile_standard_catalog(
        {"平板": source},
        output,
        active_sheets={"平板": (sheet_name,)},
    )

    assert summary["total_items"] == 2
    assert {item.scope for item in load_standard_catalog(output)} == {
        "平板电脑-通用",
    }
