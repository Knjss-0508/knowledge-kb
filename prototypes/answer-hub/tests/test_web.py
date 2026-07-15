from __future__ import annotations

from io import BytesIO
import json

from openpyxl import Workbook, load_workbook

from answer_hub.web import create_app
from answer_hub.workflow import REVIEW_COLUMNS


def _source_workbook() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "序号", "上传者", "分析时间", "工单ID", "回收单号", "聊天内容", "图片链接",
        "核心问题", "判定结论", "判定依据", "产品类型", "一级分类", "二级分类", "参考话术",
    ])
    sheet.append([
        1, "测试", "2026-07-12", "PHONE-001", "REC-001", "屏幕有色斑", "a.jpg",
        "屏幕有色斑怎么判", "判定为色斑", "色斑属于显示问题", "手机", "显示问题", "色斑", "请拍清楚色斑",
    ])
    sheet.append([
        2, "测试", "2026-07-12", "PAD-001", "REC-002", "屏幕泛黄", "b.jpg",
        "平板屏幕泛黄怎么判", "判定为老化", "白屏显示泛黄", "平板", "显示问题", "老化", "白屏拍摄",
    ])
    data = BytesIO()
    workbook.save(data)
    data.seek(0)
    return data


def _review_workbook() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "review_queue"
    headers = [
        "数据ID", "工单ID", "核心问题", "模型主标题", "模型知识内容", "模型关联标准",
        "标准检索状态", "模型初标依据", "图片证据摘要", "模型错误", "是否重点复核",
        *REVIEW_COLUMNS,
    ]
    sheet.append(headers)
    sheet.append([
        "PHONE-001", "PHONE-001", "屏幕色斑如何判定", "手机屏幕色斑判定",
        "按标准核验。", "PHONE-DISPLAY-001(7.0)", "已命中相关知识", "分类和标准匹配。",
        "图片已接收。", "", "否",
        *([""] * len(REVIEW_COLUMNS)),
    ])
    data = BytesIO()
    workbook.save(data)
    data.seek(0)
    return data


def test_phone_preview_filters_other_products() -> None:
    app = create_app()
    client = app.test_client()
    standards = json.dumps([
        {
            "standard_id": "PHONE-DISPLAY-001",
            "title": "手机屏幕色斑如何判定",
            "category_l1": "显示问题",
            "category_l2": "色斑",
            "keywords": ["手机", "屏幕", "色斑"],
            "scope": "适用于手机屏幕显示异常",
            "response_snippet": "按色斑处理",
        }
    ], ensure_ascii=False).encode("utf-8")

    response = client.post(
        "/api/preview",
        data={
            "source": (_source_workbook(), "source.xlsx"),
            "standards": (BytesIO(standards), "phone-standards.json"),
            "product_type": "手机",
            "min_confidence": "0.75",
            "use_mimo": "false",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["source_total_rows"] == 2
    assert payload["selected_rows"] == 1
    assert payload["excluded_rows"] == 1
    assert payload["standard_count"] == 1
    assert payload["candidates"][0]["模型二级分类"] == "色斑"


def test_review_workbench_reads_and_exports_cz_changes() -> None:
    app = create_app()
    client = app.test_client()

    preview = client.post(
        "/api/review-queue",
        data={"review_file": (_review_workbook(), "review_queue.xlsx")},
        content_type="multipart/form-data",
    )
    assert preview.status_code == 200
    payload = preview.get_json()
    assert payload["total_rows"] == 1
    assert payload["pending_rows"] == 1
    assert payload["rows"][0]["_review_row_index"] == 2

    changes = [
        {
            "row_index": 2,
            "updates": {
                "CZ复核结论": "修改后通过",
                "CZ主标题": "手机屏幕色斑判定标准",
                "CZ一级分类": "显示问题",
                "CZ二级分类": "色斑",
                "审核人": "cz",
            },
        }
    ]
    export = client.post(
        "/api/review-export",
        data={
            "review_file": (_review_workbook(), "review_queue.xlsx"),
            "changes": json.dumps(changes, ensure_ascii=False),
        },
        content_type="multipart/form-data",
    )
    assert export.status_code == 200
    workbook = load_workbook(BytesIO(export.data), data_only=True)
    sheet = workbook["review_queue"]
    headers = {str(cell.value): cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["CZ复核结论"]).value == "修改后通过"
    assert sheet.cell(2, headers["CZ主标题"]).value == "手机屏幕色斑判定标准"
    assert sheet.cell(2, headers["审核人"]).value == "cz"
