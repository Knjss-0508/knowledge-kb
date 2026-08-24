from __future__ import annotations

from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from urllib.error import HTTPError
import json
import sqlite3
import threading
import time

from openpyxl import Workbook, load_workbook
import pytest

import answer_hub.mimo as mimo_module
import answer_hub.workflow as workflow_module
from answer_hub.audit import AuditStore
from answer_hub.catalog import StandardCatalogItem, load_standard_catalog
from answer_hub.classification_catalog import ClassificationCatalogItem
from answer_hub.images import ImageEvidence
from answer_hub.mimo import (
    MimoClient,
    MimoConfig,
    MimoError,
    MimoLabelResult,
    TOPIC_DISPLAY_QUESTION_PROMPT_VERSION,
    TOPIC_SIGNAL_PROMPT_VERSION,
    TOPIC_STAGE_PROMPT_VERSION,
    _topic_signal_source_payload,
    _validate_topic_display_questions,
    _validate_topic_stage,
    _validate_topic_review,
)
from answer_hub.workflow import (
    _direct_mimo_topic_groups,
    build_cluster_validation_rows,
    build_topic_review_rows,
    cluster_validation_from_workbook,
    evaluate_cluster_validation_rows,
    generate_phone_candidate_rows,
    initial_label_rows,
    preprocess_source_rows,
    write_cluster_only_workbook,
    write_topic_candidate_knowledge_workbook,
)


def _source_rows() -> list[dict[str, object]]:
    return preprocess_source_rows(
        [
            {
                "序号": 1,
                "工单ID": "PHONE-001",
                "聊天内容": "屏幕有色斑，麻烦确认是否属于显示问题。",
                "图片链接": "https://example.com/phone.jpg",
                "视频链接": "https://example.com/phone.mp4",
                "核心问题": "手机屏幕色斑如何判定",
                "判定结论": "判定为色斑",
                "判定依据": "色斑属于显示问题",
                "产品类型": "手机",
                "一级分类": "显示问题",
                "二级分类": "色斑",
                "参考话术": "请拍摄清晰屏幕图片",
            }
        ]
    )


def test_case_only_topic_signal_payload_treats_historical_reply_as_primary_evidence() -> None:
    payload = _topic_signal_source_payload(_source_rows()[0])

    assert payload["primary_evidence"]["historical_actual_reply"] == "请拍摄清晰屏幕图片"
    assert "legacy_script" not in payload["legacy_reference_only"]


def test_failed_topic_transcription_keeps_source_fact_body_for_human_review() -> None:
    rows = preprocess_source_rows(
        [
            {
                "工单ID": "FAILED-FACT-001",
                "聊天内容": "充电孔一直弹出进水提示，需先清理接口再测试充电。",
                "核心问题": "平板充电孔进水提示如何判定",
                "判定结论": "证据不足，需补充接口清理和充电测试。",
                "判定依据": "仅凭提示不能判定浸液。",
                "历史实际回复": "先清理接口，再测试充电功能，证据不足时不要直接判定。",
                "产品类型": "平板",
                "对象/部位": "充电接口",
                "异常现象": "持续进水提示",
            }
        ]
    )
    topic = workflow_module._failed_topic_transcription_row(
        "TOP-FAILED-FACT",
        ("rule", "自营回收", "平板电脑", "充电接口"),
        rows,
        {"knowledge_value": "值得沉淀"},
        provider="mimo",
        model_name="mimo-v2.5",
        prompt_version="test",
        model_run_id="run",
        transcription_status="topic_model_validation_failed",
        model_call_status="model_success",
        error="输出校验失败",
        matches=[],
        use_standard_references=False,
    )

    assert topic["知识内容"]
    assert "清理接口" in topic["知识内容"]
    assert "证据不足" in topic["知识内容"]
    assert topic["推荐回复"] == ""
    assert topic["模型阶段状态"] == "topic_model_validation_failed"


def test_missing_image_measurement_keeps_source_rules_and_adds_boundary() -> None:
    query = {
        "对象/部位": "屏幕",
        "异常现象": "彩色斑点",
    }
    content = "1. 屏幕背景色下可见彩色斑点时，按来源规则核验漏液。\n2. 直径大于1mm的黑点或彩点属于判定条件。"

    guarded = workflow_module._append_measurement_evidence_boundary(
        content,
        query=query,
    )

    assert "直径大于1mm" in guarded
    assert "当前案例图片未提供可核验的尺寸" in guarded
    assert "不能据此直接确定本案例的档位或结论" in guarded


def test_topic_signal_prompt_uses_final_judgment_without_notebook_model_example() -> None:
    prompt = mimo_module._build_topic_signal_prompt(
        {
            "工单ID": "TOPIC-SIGNAL-001",
            "产品类型": "手机",
            "聊天内容": "这个是漏液吗？多颜色和息屏都看了，最后客服说选折痕。",
            "历史实际回复": "选折痕。",
            "核心问题": "屏幕漏液如何判定",
            "一级分类": "显示问题",
            "二级分类": "漏液",
        },
        [],
        [],
        use_standard_references=False,
    )

    assert TOPIC_SIGNAL_PROMPT_VERSION == "multi-category-conversation-topic-signal-v9-12-category-taxonomy"
    assert "以答疑人员的最终判定结论为准" in prompt
    assert "问“是不是漏液”→答“选折痕”" in prompt
    assert "“白光检测”的目的=检查屏幕是否被更换/拆修" in prompt
    assert "术语字典（全流程共享）" in prompt
    assert "偏光检测" in prompt
    assert "[屏幕部件] 偏光膜" in prompt
    assert "[浸液现象] 浸液痕迹" in prompt
    assert "红线检测" in prompt
    assert "一根线" in prompt
    assert "Xray" not in prompt
    assert "X-ray检测" not in prompt
    assert "GSX" not in prompt
    assert "standard_refs 必须输出 []。不得补写、猜测或引用任何质检标准。" in prompt
    assert "笔记本型号查询" not in prompt
    assert "型号查询存在标准化流程" not in prompt


def test_cluster_unit_prompt_prioritizes_chat_and_disambiguates_tool_terms() -> None:
    prompt = mimo_module._build_cluster_unit_prompt(
        {
            "工单ID": "CHAT-PRIMARY-001",
            "产品类型": "平板",
            "聊天内容": (
                "26/07/15 17:50:00:00 问题类型：质检问题 "
                "问题描述：一根线读出用户判断怎么选 "
                "转人工原因：回答内容无法理解\n"
                "26/07/15 17:51:03:03 这个是异常吗\n"
                "26/07/15 17:51:24:24 数据正常 默认那里默认异常"
            ),
            "核心问题": "屏幕出现一根线条怎么判",
            "原始核心问题": "屏幕出现一根线条怎么判",
            "判定结论": "工具读数正常，不属于屏幕物理线条",
            "原始判定结论": "工具读数正常，不属于屏幕物理线条",
        }
    )

    assert "一根线读出用户判断怎么选" not in (
        mimo_module._primary_conversation_evidence(
            "26/07/15 17:50:00:00 问题类型：质检问题 "
            "问题描述：一根线读出用户判断怎么选 "
            "转人工原因：回答内容无法理解\n"
            "26/07/15 17:51:03:03 这个是异常吗"
        )
    )
    assert mimo_module.CLUSTER_UNIT_PROMPT_VERSION == (
        "multi-category-conversation-cluster-units-v29-12-category-taxonomy"
    )
    assert "系统截图/弹窗中的文字" in prompt
    assert "通常指验机工具或检测结果" in prompt
    assert "不代表屏幕上出现物理线条" in prompt
    assert "手机、平板电脑和笔记本场景中的验机工具或工具读数" in prompt
    assert "偏光检测" in prompt
    assert "水渍光斑不要与屏幕亮斑合并" in prompt
    assert "A/B/C/D面" not in prompt
    assert "摄像头镜片与物理开关" not in prompt
    assert "DLC版" not in prompt
    assert "红线检测" in prompt
    assert "Xray" not in prompt
    assert "X-ray检测" not in prompt
    assert "GSX" not in prompt
    assert '"human_corrected_core_problem": "屏幕出现一根线条怎么判"' in prompt
    assert (
        '"human_corrected_judgment_conclusion": '
        '"工具读数正常，不属于屏幕物理线条"'
    ) in prompt
    assert "人工校正后的核心问题和判定结论" in prompt
    assert "与完整聊天明显冲突" in prompt
    assert "默认不拆" in prompt
    assert "命中以下任一必须拆分信号时才拆" in prompt
    assert "信号A — 对象隔离" in prompt
    assert "信号B — 分类隔离" in prompt
    assert "信号C — 答疑分别处理" in prompt
    assert "信号D — 拆修/非原厂多对象" in prompt
    assert "主板有标签，屏幕有贴纸怎么判" in prompt
    assert "信号E — 信息查询多目标" in prompt
    assert "有没有 BIOS 锁→型号是这个吗→硬盘内存品牌吗→支持指纹吗" in prompt
    assert "不要被“同属外观问题”" in prompt
    assert "屏幕划痕" in prompt
    assert "充电口松动" in prompt
    assert "电池健康读取不出" in prompt
    assert "不要因为前后问题在同一个会话里就默认合并" in prompt
    assert "同一对象+同一异常" in prompt
    assert "同一对象的多个描述角度" in prompt
    assert "外观问题（category_l1=\"外观问题\"）覆盖范围很广" in prompt
    assert "不同 subject 的外观问题不可合并" in prompt
    assert "孤立的单个词或客服简短回复" in prompt
    assert "少见案例、特殊案例或标准咨询可以独立形成单成员主题" in prompt
    assert "包装盒防拆标签" in prompt
    assert "已固化的聚类判定口径" in prompt
    assert "平板屏幕显示标准" in prompt
    assert "不同现象值必须拆分" in prompt
    assert "笔记本型号查询" not in prompt
    assert "型号查询存在标准化流程" not in prompt


def test_mimo_client_extracts_multiple_text_cluster_units_in_one_request(
    monkeypatch,
) -> None:
    def batch_result(record_id: str, issue: str) -> dict[str, object]:
        return {
            "record_id": record_id,
            "conversation_type": "single_topic",
            "reason": "该会话只有一个独立问题。",
            "media_analysis": {
                "image_summary": "无图片",
                "video_summary": "无视频",
                "media_relevance": "无媒体",
                "used_for_topic_split": False,
                "requires_review": False,
            },
            "topics": [
                {
                    "normalized_issue": issue,
                    "product_category": "手机",
                    "scope_type": "品类专用",
                    "platform": "通用",
                    "brand": "通用",
                    "model_scope": "通用",
                    "category_l1": "流程操作",
                    "category_l2": "节点核验",
                    "intent": "流程操作",
                    "subject": "节点状态",
                    "phenomenon": "状态待确认",
                    "judgment_target": "确认节点状态",
                    "resolution_mode": "按会话上下文核对",
                    "standard_path": "节点核验流程",
                    "threshold_or_exception": "无明确阈值",
                    "evidence_summary": "完整聊天支持该问题。",
                    "confidence": 0.9,
                    "requires_review": False,
                }
            ],
        }

    payloads: list[dict[str, object]] = []
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="test-model",
        )
    )

    def respond(payload):
        payloads.append(payload)
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "results": [
                                    batch_result("R0001", "节点状态核验A"),
                                    batch_result("R0002", "节点状态核验B"),
                                ]
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    monkeypatch.setattr(client, "_post", respond)
    results = client.analyze_cluster_units_batch(
        [
            {
                "工单ID": "A",
                "聊天内容": "节点状态核验A",
                "产品类型": "手机",
            },
            {
                "工单ID": "B",
                "聊天内容": "节点状态核验B",
                "产品类型": "手机",
            },
        ]
    )

    assert len(payloads) == 1
    assert [result.candidate["topics"][0]["normalized_issue"] for result in results] == [
        "节点状态核验A",
        "节点状态核验B",
    ]
    prompt = payloads[0]["messages"][1]["content"][0]["text"]
    assert '"record_id": "R0001"' in prompt
    assert '"record_id": "R0002"' in prompt


def test_cluster_unit_source_guard_removes_unsourced_threshold_and_standard_path() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="mimo-v2.5-test",
        )
    )
    client._post = lambda _payload: {  # type: ignore[method-assign]
        "choices": [
            {
                "message": {
                    "content": json.dumps(
                        {
                            "conversation_type": "single_topic",
                            "reason": "会话只有一个屏幕现象确认问题。",
                            "topics": [
                                {
                                    "normalized_issue": "手机｜屏幕｜点状现象｜确认类型",
                                    "product_category": "手机",
                                    "scope_type": "品类专用",
                                    "platform": "通用",
                                    "brand": "通用",
                                    "model_scope": "通用",
                                    "category_l1": "显示问题",
                                    "category_l2": "点状现象",
                                    "intent": "标准判定",
                                    "subject": "屏幕",
                                    "phenomenon": "点状现象",
                                    "judgment_target": "确认现象类型",
                                    "resolution_mode": "结合来源事实核验",
                                    "standard_path": "【手机】-【屏幕】-【色斑】",
                                    "threshold_or_exception": "直径大于1mm判为色斑",
                                    "evidence_summary": "聊天只要求确认屏幕点状现象。",
                                    "confidence": 0.9,
                                    "requires_review": False,
                                }
                            ],
                        },
                        ensure_ascii=False,
                    )
                }
            }
        ]
    }

    result = client.analyze_cluster_units(
        {
            "工单ID": "UNSOURCED-CLUSTER-001",
            "产品类型": "手机",
            "聊天内容": "屏幕这个点属于什么现象？",
            "核心问题": "确认屏幕点状现象类型",
        }
    )
    topic = result.candidate["topics"][0]

    assert topic["standard_path"] == "待确认"
    assert topic["threshold_or_exception"] == "无明确阈值"
    assert topic["requires_review"] is True


def test_cluster_fusion_prompt_keeps_media_second_topic_example() -> None:
    prompt = mimo_module._build_cluster_fusion_prompt(
        {
            "工单ID": "MEDIA-SECOND-TOPIC",
            "产品类型": "手机",
            "聊天内容": "这个手机相机倍数正常吗？",
        },
        {
            "conversation_type": "single_topic",
            "topics": [{"normalized_issue": "相机倍数是否正常"}],
        },
        {
            "conversation_type": "multi_topic",
            "media_analysis": {
                "used_for_topic_split": True,
            },
            "topics": [
                {"normalized_issue": "相机倍数是否正常"},
                {"normalized_issue": "屏幕垂直彩色亮线怎么判"},
            ],
        },
        {"images": [], "videos": []},
    )

    assert mimo_module.CLUSTER_FUSION_PROMPT_VERSION == (
        "multi-category-conversation-cluster-fusion-v9-12-category-taxonomy"
    )
    assert "媒体\"未看到\"≠文字问题不存在" in prompt
    assert "明确文字主题" in prompt
    assert "可靠媒体新增主题" in prompt
    assert "术语字典（全流程共享）" in prompt
    assert "笔记本型号查询" not in prompt
    assert "型号查询存在标准化流程" not in prompt


def test_cluster_pair_review_prompt_uses_judgment_rules() -> None:
    prompt = mimo_module._build_cluster_pair_review_prompt(
        {
            "normalized_issue": "手机｜屏幕｜坏点｜判定是否合格",
            "product_category": "手机",
        },
        {
            "normalized_issue": "手机｜屏幕｜亮线｜判定是否合格",
            "product_category": "手机",
        },
        similarity=0.85,
        threshold=0.75,
    )

    assert mimo_module.CLUSTER_PAIR_REVIEW_PROMPT_VERSION == (
        "knowledge-cluster-membership-review-v16-12-category-taxonomy"
    )
    assert "以下情况可以判断为\"同一主题\"（即使机型/品牌不同）" in prompt
    assert "术语字典（全流程共享）" in prompt
    assert "全新机判定标准" in prompt
    assert "外观损伤定性（磕碰/碎裂/划痕的边界）" in prompt
    assert "核心对象相同 + 判定目标相同 + 处理路径相同" in prompt
    assert "一级分类不同（如功能问题 vs 外观问题，显示问题 vs 拆修问题）" in prompt
    assert "一级分类相同但判定对象不同" in prompt
    assert "同属外观问题" in prompt
    assert "品类相同 + 判定对象相同 + 判定标准相同" in prompt
    assert "屏幕坏点判定 vs 屏幕亮线判定" in prompt
    assert "不要根据相似度直接下结论" in prompt
    assert "手机外壳外观标准" in prompt
    assert "同一标准族可作为同主题候选" in prompt
    assert "手机屏幕显示标准" in prompt
    assert "不同现象值必须拆分" in prompt
    assert "二级分类或标准路径的文字不同，不是绝对拆分条件" in prompt
    assert "产品品类不同绝对禁止合并" in prompt
    assert "回收业务层级不同绝对禁止合并" in prompt
    assert "human_corrected_core_problem" in prompt
    assert "human_corrected_judgment_conclusion" in prompt
    assert "上游核心问题摘要和旧分类已从审核输入中移除" not in prompt
    assert "笔记本型号查询" not in prompt
    assert "型号查询存在标准化流程" not in prompt


def test_atomic_topic_cluster_prompt_uses_shared_terminology() -> None:
    prompt = mimo_module._build_atomic_topic_cluster_prompt(
        [
            {
                "unit_id": "S001-U1",
                "normalized_issue": "平板｜一根线工具读出用户判断｜最终判定",
                "source_core_problem": "工具读数显示用户判断，应该如何勾选",
                "source_judgment_conclusion": "以人工复检确认后的结果为准",
                "product_category": "平板",
                "scope_type": "品类专用",
                "category_l1": "流程操作",
                "category_l2": "质检判定流程",
                "intent": "标准判定",
                "subject": "验机工具结果",
                "phenomenon": "工具读出用户判断",
                "judgment_target": "确认最终判定依据",
                "resolution_mode": "以人工复检事实为主",
                "standard_path": "待确认",
                "threshold_or_exception": "无明确阈值",
                "requires_review": True,
            }
        ]
    )

    assert mimo_module.ATOMIC_TOPIC_CLUSTER_PROMPT_VERSION == (
        "atomic-knowledge-topic-clustering-v16-classification-catalog-reviewable-recall"
    )
    assert "术语字典（全流程共享）" in prompt
    assert "一根线" in prompt
    assert "红线检测" in prompt
    assert "已固化的聚类判定口径" in prompt
    assert "平板屏幕显示标准" in prompt
    assert "不同现象值必须拆分" in prompt
    assert "二级分类或标准路径的文字不同，不是绝对拆分条件" in prompt
    assert "review_requests 不等于必须单独成簇" in prompt
    assert "产品品类不同绝对禁止合并" in prompt
    assert '"source_core_problem": "工具读数显示用户判断，应该如何勾选"' in prompt
    assert (
        '"source_judgment_conclusion": "以人工复检确认后的结果为准"'
        in prompt
    )
    assert "少见、特殊或标准咨询类原子问题" in prompt
    assert "classification_catalog_status=classification_ambiguous" in prompt
    assert "不得仅因候选路径不同把同品类案例强制拆开" in prompt
    assert "Xray" not in prompt
    assert "GSX" not in prompt


def test_cluster_review_payload_keeps_human_corrected_evidence() -> None:
    payload = mimo_module._cluster_review_evidence_payload(
        {
            "source_conversation": (
                "问题类型：质检问题 问题描述：乱填内容\n"
                "请结合现场图片确认屏幕上的点"
            ),
            "source_core_problem": "人工确认这是屏幕色斑还是灰尘",
            "source_judgment_conclusion": "人工看图后确认是屏幕色斑",
            "quality_question_description": "乱填内容",
        }
    )

    assert payload["primary_conversation_evidence"] == (
        "请结合现场图片确认屏幕上的点"
    )
    assert payload["human_corrected_core_problem"] == (
        "人工确认这是屏幕色斑还是灰尘"
    )
    assert payload["human_corrected_judgment_conclusion"] == (
        "人工看图后确认是屏幕色斑"
    )
    assert "quality_question_description" not in payload


def test_cluster_review_payload_keeps_atomic_and_human_questions_separate() -> None:
    row = {
        "_原子知识ID": "WO-001-U01",
        "工单ID": "WO-001",
        "回收业务层级": "自营回收",
        "产品类型": "手机",
        "核心问题": "手机｜屏幕｜色斑｜判定",
        "原始核心问题": "人工确认这是屏幕色斑还是灰尘",
        "原始判定结论": "人工看图后确认是屏幕色斑",
    }

    payload = mimo_module._cluster_review_evidence_payload(
        workflow_module._cluster_validation_payload(row)
    )
    member_payload = mimo_module._cluster_review_evidence_payload(
        workflow_module._cluster_membership_member_payload(row)
    )

    for value in (payload, member_payload):
        assert value["atomic_question"] == "手机｜屏幕｜色斑｜判定"
        assert value["human_corrected_core_problem"] == (
            "人工确认这是屏幕色斑还是灰尘"
        )
        assert value["human_corrected_judgment_conclusion"] == (
            "人工看图后确认是屏幕色斑"
        )


def test_direct_mimo_source_signature_changes_after_human_correction() -> None:
    base = {
        "工单ID": "WO-SIGNATURE-001",
        "回收单号": "RO-SIGNATURE-001",
        "回收业务层级": "自营回收",
        "产品类型": "手机",
        "聊天内容": "请结合现场图片确认屏幕上的点",
        "核心问题": "屏幕上的点如何判定",
        "判定结论": "待人工确认",
        "图片链接": "",
        "视频链接": "",
    }
    corrected = {
        **base,
        "核心问题": "人工确认这是屏幕色斑还是灰尘",
        "判定结论": "人工看图后确认是屏幕色斑",
    }

    assert workflow_module._direct_mimo_source_signature(base) != (
        workflow_module._direct_mimo_source_signature(corrected)
    )


def test_direct_mimo_source_signature_tracks_ai_result_conflicts() -> None:
    base = {
        "工单ID": "WO-SIGNATURE-CONFLICT-001",
        "回收单号": "RO-SIGNATURE-CONFLICT-001",
        "回收业务层级": "自营回收",
        "产品类型": "手机",
        "聊天内容": "请结合现场图片确认屏幕上的点",
        "核心问题": "屏幕上的点如何判定",
        "判定结论": "待人工确认",
        "图片链接": "",
        "视频链接": "",
        "AI结果冲突字段": "",
    }
    conflicted = {**base, "AI结果冲突字段": "核心问题"}

    assert workflow_module._direct_mimo_source_signature(base) != (
        workflow_module._direct_mimo_source_signature(conflicted)
    )


def test_direct_mimo_progress_signature_tracks_product_taxonomy(
    monkeypatch,
) -> None:
    monkeypatch.setattr(
        workflow_module,
        "product_taxonomy_metadata",
        lambda: {"version": "taxonomy-v1", "digest": "digest-v1"},
    )
    first = workflow_module._direct_mimo_progress_signatures()
    monkeypatch.setattr(
        workflow_module,
        "product_taxonomy_metadata",
        lambda: {"version": "taxonomy-v2", "digest": "digest-v2"},
    )
    second = workflow_module._direct_mimo_progress_signatures()

    assert first["atomic"]["product_taxonomy_digest"] == "digest-v1"
    assert second["atomic"]["product_taxonomy_digest"] == "digest-v2"
    assert first["cluster"] != second["cluster"]
    assert first["reconcile"] != second["reconcile"]


def test_direct_mimo_dedup_keeps_stable_source_identity() -> None:
    base = {
        "工单ID": "WO-DEDUP-001",
        "回收单号": "RO-DEDUP-001",
        "回收业务层级": "自营回收",
        "产品类型": "手机",
        "聊天内容": "请结合现场图片确认屏幕上的点",
        "核心问题": "屏幕上的点如何判定",
        "判定结论": "待人工确认",
        "图片链接": "",
        "视频链接": "",
    }
    corrected_paraphrase = {
        **base,
        "核心问题": "人工确认这是屏幕色斑还是灰尘",
        "判定结论": "人工看图后确认是屏幕色斑",
    }

    unique_rows, duplicate_rows = (
        workflow_module._direct_mimo_deduplicate_rows(
            [base, corrected_paraphrase]
        )
    )

    assert unique_rows == [base]
    assert len(duplicate_rows) == 1


def test_aggregate_business_line_prompt_does_not_use_self_operated_rules() -> None:
    prompt = mimo_module._clustering_rules_prompt_block(
        {
            "回收业务层级": "聚合回收",
            "产品类型": "手机",
            "对象/部位": "屏幕",
            "异常现象": "碎裂",
        }
    )

    assert "禁止套用自营回收十品类质检聚类口径" in prompt
    assert "手机屏幕外观标准" not in prompt


def test_remaining_mimo_prompts_use_scoped_notebook_terminology() -> None:
    source_row = {
        "产品类型": "笔记本",
        "聊天内容": "屏幕偏光膜破损，主板有飞线。",
    }
    topic = {
        "theme_id": "NB-001",
        "product_category": "笔记本",
        "product_categories": ["笔记本"],
        "normalized_issues": ["笔记本｜屏幕偏光膜｜破损｜判断拆修部件"],
    }
    prompts = (
        mimo_module._build_prompt(source_row, [], []),
        mimo_module._build_topic_prompt(topic, [], use_standard_references=False),
        mimo_module._build_topic_review_prompt(
            topic,
            {"title": "笔记本屏幕偏光膜破损怎么判断"},
            [],
            use_standard_references=False,
        ),
        mimo_module._build_topic_stage_prompt(topic),
        mimo_module._build_topic_display_questions_prompt([topic]),
    )

    assert mimo_module.PROMPT_VERSION == (
        "multi-category-topic-transcription-v13-recall-subtitles"
    )
    assert mimo_module.TOPIC_REVIEW_PROMPT_VERSION == (
        "multi-category-topic-content-quality-review-v9-12-category-taxonomy"
    )
    assert mimo_module.TOPIC_STAGE_PROMPT_VERSION == (
        "multi-category-topic-stage-value-v8-12-category-taxonomy"
    )
    assert mimo_module.TOPIC_DISPLAY_QUESTION_PROMPT_VERSION == (
        "topic-display-question-v3-12-category-taxonomy"
    )
    for prompt in prompts:
        assert "术语字典（全流程共享）" in prompt
        assert "A/B/C/D面" in prompt
        assert "[屏幕部件] 偏光膜" in prompt
        assert "DLC版" not in prompt
        assert "快门数" not in prompt
    assert "来源事实证据包" in prompts[1]
    assert "没有来源的内容不得补写" in prompts[1]
    assert "人工核心问题和人工判定结论" in prompts[1]
    assert "applicable_brands" in prompts[1]
    assert "applicable_models" in prompts[1]
    assert "来源已经明确机型及功能结论" in prompts[1]
    assert "标题、正文、推荐回复和 applicable_models" in prompts[1]
    assert "问题背景、判断对象、来源核验依据" in prompts[1]
    assert "只能回答当前单一原子主题" in prompts[1]
    assert "历史实际回复中出现过" in prompts[1]
    assert "重复问候" in prompts[1]
    assert "content_type" in prompts[1]
    assert "简洁编号格式" in prompts[1]
    assert "不得写空占位" in prompts[1]
    assert "事实ID" in prompts[2]
    assert "不能因为这些内容出现在历史实际回复中就判断一致" in prompts[2]
    assert "异常标点" in prompts[2]


def test_generic_applicable_scope_uses_product_name_for_cz() -> None:
    rows = [{"产品类型": "手机", "聊天内容": "咨询设备信息。"}]

    assert workflow_module._normalized_applicable_scope(
        "手机",
        "手机-通用",
        rows,
    ) == "手机"
    assert workflow_module._normalized_applicable_scope(
        "手机",
        "通用",
        rows,
    ) == "手机"
    assert workflow_module._normalized_applicable_scope(
        "手机",
        "手机-iOS",
        rows,
    ) == "手机"
    assert workflow_module._normalized_applicable_scope(
        "笔记本",
        "笔记本-Windows",
        [{"产品类型": "笔记本", "聊天内容": "Windows设备。"}],
    ) == "笔记本"
    assert workflow_module._supported_topic_applicability_values(
        ["15"],
        [{"聊天内容": "检测持续15秒后结束。"}],
        "适用机型",
        "手机",
    ) == []


def test_candidate_validation_normalizes_alias_scope_and_rejects_non_category_scope() -> None:
    base = {
        "title": "耳机连接如何核验",
        "subtitles": [],
        "content": "1. 根据来源事实核对连接表现。",
        "content_type": "核验型",
        "category_l1": "功能问题",
        "category_l2": "连接功能",
        "layer": "L2",
        "knowledge_form": "流程方法",
        "standard_refs": [],
        "applicable_scope": "耳机",
        "applicable_brands": [],
        "applicable_models": [],
        "confidence": 0.8,
        "reasoning_summary": "来源事实支持该品类。",
        "needs_human_review": True,
        "image_evidence_summary": "无图片。",
    }

    assert (
        mimo_module._validate_candidate(base, set())["applicable_scope"]
        == "耳机/耳麦"
    )

    invalid = dict(base, applicable_scope="手机屏幕")
    with pytest.raises(MimoError, match="applicable_scope"):
        mimo_module._validate_candidate(invalid, set())


def test_direct_mimo_workflow_applies_local_multi_topic_rescue() -> None:
    class SingleTopicMimo:
        config = SimpleNamespace(model="mimo-local-rescue-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "模型误判为单主题。",
                    "topics": [
                        {
                            "normalized_issue": "笔记本拆修与硬件信息确认",
                            "product_category": "笔记本",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "信息查询",
                            "category_l2": "型号/硬件信息",
                            "intent": "信息查询",
                            "subject": "设备信息",
                            "phenomenon": "多个目标连续确认",
                            "judgment_target": "确认多个独立信息目标",
                            "resolution_mode": "人工复核",
                            "standard_path": "待确认",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "模型原始输出为单主题。",
                            "confidence": 0.82,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            assert {unit["unit_id"] for unit in units} <= {
                "F041-U1",
                "F041-U2",
                "F041-U3",
                "F041-U4",
            }
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "两个原子问题的处理目标不同，分别保留。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "F041",
            "工单ID": "F041",
            "产品类型": "笔记本",
            "核心问题": "需要确认BIOS锁状态、型号、硬盘内存品牌和指纹支持",
            "聊天内容": (
                "有没有BIOS锁\n"
                "没锁的\n"
                "型号是这个吗\n"
                "RedmiBook 14 2025\n"
                "硬盘内存品牌吗\n"
                "内存和硬盘都是品牌认证的\n"
                "这个机器支持指纹吗\n"
                "不支持"
            ),
        }
    ]

    topic_groups, meta = _direct_mimo_topic_groups(rows, SingleTopicMimo())

    assert meta["local_multi_topic_rescue"] == 1
    assert meta["local_multi_topic_rescue_samples"] == ["F041"]
    assert len(topic_groups) == 4
    assert {
        group_rows[0]["判定目标"]
        for _key, group_rows in topic_groups
    } == {
        "确认设备是否存在BIOS锁",
        "确认设备具体型号和年款",
        "确认内存和硬盘品牌属性",
        "确认设备是否支持指纹识别",
    }
    statuses = {group_rows[0]["语义标注状态"] for _key, group_rows in topic_groups}
    assert statuses == {"atomic_unit_labeled_local_multi_topic_rescue"}
    review_reasons = {
        group_rows[0]["人工优先复核原因"] for _key, group_rows in topic_groups
    }
    assert all(
        "local_structured_info_query_rescue" in reason
        for reason in review_reasons
    )
    assert all(
        group_rows[0]["_原子需要复核"]
        for _key, group_rows in topic_groups
    )


@pytest.mark.parametrize(
    ("product_type", "conversation", "expected_reason", "expected_targets"),
    [
        (
            "聚合回收",
            (
                "爬虫读出乱码\n"
                "序列号不判\n"
                "这个屏幕需要判碎裂吗\n"
                "要的"
            ),
            "local_serial_plus_screen_damage_rescue",
            ("确认序列号读取乱码的处理方式", "确认屏幕碎裂是否需要判定"),
        ),
        (
            "笔记本",
            (
                "白底这样的印记要判别的吗\n"
                "其他颜色下有吗\n"
                "要判色斑\n"
                "印记也要判"
            ),
            "local_notebook_screen_spot_surface_rescue",
            ("确认是否属于屏幕显示色斑", "确认是否属于屏幕表面印记"),
        ),
        (
            "相机镜头",
            "这些是进灰吗\n镜头进灰异物+消光漆脱落",
            "local_lens_internal_dual_condition_rescue",
            ("确认镜头内部是否进灰或存在异物", "确认镜头内部消光漆是否脱落"),
        ),
        (
            "笔记本",
            (
                "什么机型\n"
                "联想拯救者R9000P 2022款\n"
                "你这个是有膜的\n"
                "撕了看外观"
            ),
            "local_model_plus_screen_film_rescue",
            ("确认设备具体型号和年款", "确认机型出厂是否带屏幕膜"),
        ),
    ],
)
def test_local_multi_topic_rescue_restores_explicit_dual_topics(
    product_type: str,
    conversation: str,
    expected_reason: str,
    expected_targets: tuple[str, str],
) -> None:
    original = {
        "normalized_issue": "模型合并成一个主题",
        "product_category": product_type,
        "scope_type": "品类专用",
        "platform": "通用",
        "brand": "通用",
        "model_scope": "通用",
        "category_l1": "待确认",
        "category_l2": "待确认",
        "intent": "待确认",
        "subject": "待确认",
        "phenomenon": "待确认",
        "judgment_target": "待确认",
        "resolution_mode": "人工复核",
        "standard_path": "待确认",
        "threshold_or_exception": "无明确阈值",
        "evidence_summary": "模型原始只抽出一个主题。",
        "confidence": 0.9,
        "requires_review": False,
    }
    topics, reason = workflow_module._direct_local_multi_topic_rescue_topics(
        {
            "产品类型": product_type,
            "核心问题": "同一会话包含两个独立问题",
            "聊天内容": conversation,
        },
        [original],
    )

    assert reason == expected_reason
    assert tuple(topic["judgment_target"] for topic in topics) == expected_targets
    assert all(topic["requires_review"] for topic in topics)
    assert all(topic["_local_multi_topic_rescue_reason"] == reason for topic in topics)


def test_case_only_topic_generation_rejects_model_standard_references() -> None:
    features, _ = generate_phone_candidate_rows(
        _source_rows(),
        [],
        use_mimo=False,
        image_downloader=_ReadyImageDownloader(),
        use_standard_references=False,
    )
    topics, _mapping, gaps, pending = build_topic_review_rows(
        features,
        [],
        mimo_client=_FakeMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    # 模型成功返回但内容质量门禁未通过，使用独立的质量失败状态。
    assert topics[0]["模型阶段状态"] == "topic_model_quality_failed"
    assert topics[0]["模型调用状态"] == "model_success"
    assert topics[0]["模型输出校验状态"] == "passed"
    assert topics[0]["模型质量状态"] == "failed"
    assert topics[0]["知识草稿状态"] == "blocked"
    assert "标准引用" in topics[0]["校验备注"]
    assert topics[0]["知识内容"]
    assert "色斑" in topics[0]["知识内容"]
    assert "来源未说明的其他情形不得直接套用" in topics[0]["知识内容"]
    assert topics[0]["推荐回复"] == ""
    assert topics[0]["关联标准项"] == ""


def test_case_only_review_ignores_model_request_for_standard_reference() -> None:
    class StandardDependentReviewMimo(_FakeMimo):
        def label_topic(self, _topic, _matches, use_standard_references=False):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "title": "屏幕显示异常如何通过图片核验",
                    "subtitles": [],
                    "content": (
                        "核验流程：\n"
                        "1. 确认异常出现的画面和位置。\n"
                        "2. 补充白屏全景与异常位置近景。\n"
                        "3. 排除反光、贴膜和环境光干扰。\n"
                        "4. 信息不足时补充后再处理。"
                    ),
                    "category_l1": "显示问题",
                    "category_l2": "色斑",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": [],
                    "applicable_scope": "手机-通用",
                    "recommended_reply": "请补充白屏全景和异常位置近景，并排除反光或贴膜影响；信息不足时补充后再处理。",
                    "confidence": 0.82,
                    "reasoning_summary": "依据完整会话和案例图片形成核验流程。",
                    "needs_human_review": False,
                    "image_evidence_summary": "案例包含可用图片。",
                    "requires_images": True,
                    "image_usage_instruction": "保留脱敏案例图说明异常位置。",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(
            self,
            _topic,
            _draft,
            _matches,
            use_standard_references=False,
        ):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "decision": "证据不足待补充",
                    "knowledge_value": "待确认",
                    "error_type": "标准未覆盖/标准召回不足",
                    "reason": "缺少标准引用。",
                    "standard_consistency": "无可信标准",
                    "evidence_sufficiency": "不足",
                    "confidence": 0.9,
                    "priority_review": True,
                },
                request_audit={},
                response_audit={},
            )

    features, _ = generate_phone_candidate_rows(
        _source_rows(),
        [],
        use_mimo=False,
        image_downloader=_ReadyImageDownloader(),
        use_standard_references=False,
    )
    topics, _mapping, gaps, pending = build_topic_review_rows(
        features,
        [],
        mimo_client=StandardDependentReviewMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["模型初标错误类型"] == "场景理解错"
    assert topics[0]["模型初标结论"] == "需修改"
    assert topics[0]["主题无来源内容"]
    assert "已忽略模型提出的标准补充要求" in topics[0]["模型初标原因"]


def _standards() -> list[StandardCatalogItem]:
    return [
        StandardCatalogItem(
            standard_id="PHONE-DISPLAY-001",
            title="手机屏幕色斑判定",
            category_l1="显示问题",
            category_l2="色斑",
            knowledge_type="场景判定",
            standard_path="【显示问题】-【色斑】",
            keywords=["手机", "屏幕", "色斑"],
            scope="适用于手机屏幕色斑异常",
            response_snippet="按色斑标准核验",
            status="published",
            version="v2026.07",
        )
    ]


@pytest.mark.parametrize(
    ("snippet", "question", "expected"),
    [
        (
            "碎裂是指外屏玻璃出现裂纹、断裂或结构性破损。",
            "什么是屏幕碎裂？",
            "定义型",
        ),
        (
            "检测方法：/；磕点直径不超过1mm时按有触感划痕处理，直径大于1mm时按屏幕磕点处理。",
            "屏幕磕点多大按磕点处理？",
            "阈值型",
        ),
        (
            "用指甲横向刮擦划痕，有明显阻滞感为有触感划痕；无阻滞感为无触感划痕。",
            "有触感划痕如何核验？",
            "核验型",
        ),
        (
            "根据结构形态、触感和尺寸区分碎裂、磕点、有触感划痕和无触感划痕；分别计数。",
            "屏幕碎裂、磕点和划痕如何区分？",
            "区分型",
        ),
    ],
)
def test_topic_content_type_uses_source_rule_priority(
    snippet: str,
    question: str,
    expected: str,
) -> None:
    standard = StandardCatalogItem(
        standard_id="SCREEN-APPEARANCE-001",
        title="屏幕外观瑕疵判定",
        category_l1="外观问题",
        category_l2="屏幕",
        knowledge_type="场景判定",
        standard_path="【外观问题】-【屏幕】",
        keywords=["屏幕"],
        scope="手机",
        response_snippet=snippet,
        status="published",
        version="v2026.08",
    )
    query = {
        "核心问题": question,
        "对象/部位": "屏幕外屏玻璃",
        "异常现象": "屏幕外观瑕疵",
    }

    assert workflow_module._classify_topic_content_type(
        query,
        [query],
        [(standard, 1.0)],
    ) == expected


def test_topic_content_type_distinction_overrides_threshold_and_verification() -> None:
    standard = StandardCatalogItem(
        standard_id="SCREEN-APPEARANCE-002",
        title="屏幕碎裂、磕点和划痕区分",
        category_l1="外观问题",
        category_l2="屏幕",
        knowledge_type="场景判定",
        standard_path="【外观问题】-【屏幕】",
        keywords=["屏幕", "划痕"],
        scope="手机",
        response_snippet=(
            "磕点直径不超过1mm时按有触感划痕处理；"
            "用指甲横向刮擦判断触感，有触感和无触感划痕分别计数。"
        ),
        status="published",
        version="v2026.08",
    )

    assert workflow_module._classify_topic_content_type(
        {"核心问题": "屏幕磕点和划痕如何区分？"},
        [{"核心问题": "屏幕磕点和划痕如何区分？"}],
        [(standard, 1.0)],
    ) == "区分型"


def test_short_definition_content_is_allowed_by_topic_quality_gate() -> None:
    assert workflow_module._topic_content_has_complete_short_structure(
        "1. 碎裂是指外屏玻璃出现裂纹、断裂或结构性破损。",
        "定义型",
    )
    assert not workflow_module._topic_content_has_complete_short_structure(
        "1. 碎裂是指外屏玻璃出现裂纹。",
        "区分型",
    )


def test_mimo_content_type_requires_the_matching_number_of_numbered_points() -> None:
    assert mimo_module._content_type_has_expected_numbered_points(
        "1. 磕点是指外屏玻璃受碰撞形成的凹坑或缺口。\n"
        "2. 直径大于1mm时按屏幕磕点处理。",
        "阈值型",
    )
    assert not mimo_module._content_type_has_expected_numbered_points(
        "1. 磕点直径大于1mm时按屏幕磕点处理。",
        "区分型",
    )


def test_new_contract_subtitles_keep_only_recall_questions() -> None:
    subtitles = mimo_module._validate_subtitles(
        [
            "屏幕碎裂、磕点和划痕分别怎么判断？",
            "有触感和无触感划痕如何区分？",
        ],
        "屏幕碎裂、磕点和划痕如何区分？",
        required=True,
    )

    assert len(subtitles) == 2
    assert mimo_module._validate_subtitles(
        [
            "适用部位：外屏玻璃",
            "屏幕碎裂、磕点和划痕如何区分？",
            "外屏玻璃的磕点怎么核验？",
        ],
        "屏幕碎裂、磕点和划痕如何区分？",
        required=True,
    ) == ["外屏玻璃的磕点怎么核验？"]


@pytest.mark.parametrize(
    ("supplied_content_type", "expected_content_type"),
    [
        ("流程型", "核验型"),
        ("查询型", "核验型"),
        ("边界型", "区分型"),
        ("解释型", "定义型"),
    ],
)
def test_candidate_validation_normalizes_common_content_type_aliases(
    supplied_content_type: str,
    expected_content_type: str,
) -> None:
    content = {
        "核验型": "1. 打开设备设置。\n2. 查看电池健康度显示结果。",
        "区分型": "1. 进灰表现为屏幕内部颗粒。\n2. 漏液表现为显示区域异常。",
        "定义型": "1. 电池健康度是设备当前可用容量相对设计容量的状态指标。",
    }[expected_content_type]
    candidate = {
        "title": "平板电池健康度应如何核验？",
        "subtitles": [],
        "content": content,
        "content_type": supplied_content_type,
        "category_l1": "电池",
        "category_l2": "电池健康度",
        "layer": "L2",
        "knowledge_form": "流程方法",
        "standard_refs": [],
        "applicable_scope": "平板电脑",
        "applicable_brands": [],
        "applicable_models": [],
        "confidence": 0.8,
        "reasoning_summary": "来源事实支持该核验方法。",
        "needs_human_review": True,
        "image_evidence_summary": "无图片。",
    }

    validated = mimo_module._validate_candidate(candidate, set())

    assert validated["content_type"] == expected_content_type
    assert validated["subtitles"] == []


def test_candidate_validation_requires_content_type_in_new_contract() -> None:
    candidate = {
        "title": "平板电池健康度应如何核验？",
        "subtitles": ["平板电池健康度怎么查看？"],
        "content": "打开设置查看电池健康度。",
        "category_l1": "电池",
        "category_l2": "电池健康度",
        "layer": "L2",
        "knowledge_form": "流程方法",
        "standard_refs": [],
        "applicable_scope": "平板电脑",
        "applicable_brands": [],
        "applicable_models": [],
        "confidence": 0.8,
        "reasoning_summary": "来源事实支持该核验方法。",
        "needs_human_review": True,
        "image_evidence_summary": "无图片。",
    }

    with pytest.raises(MimoError, match="content_type"):
        mimo_module._validate_candidate(candidate, set())


@pytest.mark.parametrize(
    "error",
    (
        "MiMo 输出的 content_type 必须为 定义型、阈值型、核验型 或 区分型",
        "MiMo 输出的核验型正文必须使用对应数量的编号要点",
    ),
)
def test_content_contract_errors_are_validation_failures(error: str) -> None:
    assert workflow_module._topic_model_failure_status(
        MimoError(error)
    ) == ("topic_model_validation_failed", "model_success")


def test_rule_fallback_writes_compact_numbered_content_by_content_type() -> None:
    standard = StandardCatalogItem(
        standard_id="SCREEN-APPEARANCE-003",
        title="屏幕碎裂、磕点和划痕如何区分？",
        category_l1="外观问题",
        category_l2="屏幕",
        knowledge_type="场景判定",
        standard_path="【外观问题】-【屏幕】",
        keywords=["屏幕", "碎裂", "磕点", "划痕"],
        scope="手机",
        response_snippet=(
            "标准定义：1. 根据结构形态、触感、尺寸和数量区分碎裂、磕点和划痕。"
            "2. 磕点直径不超过1mm时按有触感划痕处理，直径大于1mm时按屏幕磕点处理。"
            "3. 用指甲横向刮擦，有阻滞感为有触感划痕；有触感和无触感划痕分别计数。"
        ),
        status="published",
        version="v2026.08",
    )
    row = {
        "数据ID": "SCREEN-001",
        "工单ID": "SCREEN-001",
        "核心问题": "屏幕碎裂、磕点和划痕如何区分？",
        "判定依据": "按屏幕外观标准核验。",
        "产品类型": "手机",
        "一级分类": "外观问题",
        "二级分类": "屏幕",
        "问题意图": "现象区分",
        "对象/部位": "屏幕外屏玻璃",
        "异常现象": "碎裂、磕点、划痕",
        "解题方式": "对照标准区分",
    }

    draft = workflow_module._topic_rule_draft(
        "TOP-SCREEN-001",
        [row],
        [(standard, 1.0)],
    )

    assert draft["content_type"] == "区分型"
    assert draft["content"].startswith("1. 根据结构形态、触感、尺寸和数量区分")
    assert "2. 磕点直径不超过1mm" in draft["content"]
    assert "3. 用指甲横向刮擦" in draft["content"]
    assert "主问题：" not in draft["content"]
    assert draft["subtitles"][:2] == [
        "屏幕碎裂、磕点和划痕分别怎么判断？",
        "屏幕碎裂、磕点和划痕中的有触感和无触感情况如何区分？",
    ]


class _ReadyImageDownloader:
    def fetch(self, _links: str) -> list[ImageEvidence]:
        return [
            ImageEvidence(
                url="https://example.com/phone.jpg",
                status="ready",
                mime_type="image/jpeg",
                byte_size=8,
                data_url="data:image/jpeg;base64,AA==",
            )
        ]


class _FailedImageDownloader:
    def fetch(self, _links: str) -> list[ImageEvidence]:
        return [ImageEvidence(url="https://example.com/phone.jpg", status="failed", error="timeout")]


class _FakeMimo:
    config = SimpleNamespace(model="mimo-v2.5-test")

    def analyze_topic_signal(self, _source, _matches, _images):
        return MimoLabelResult(
            candidate={
                "intent": "标准判定",
                "subject": "屏幕",
                "phenomenon": "色斑",
                "resolution_mode": "对照标准判定",
                "category_l1": "显示问题",
                "category_l2": "色斑",
                "topic_tags": ["意图:标准判定", "对象:屏幕", "现象:色斑", "处理:对照标准判定"],
                "standard_refs": ["PHONE-DISPLAY-001"],
                "requires_images": True,
                "image_evidence_summary": "图片已接收，仍需人工确认细节。",
                "reasoning_summary": "完整会话在询问屏幕色斑的标准判定。",
                "confidence": 0.91,
                "needs_human_review": False,
            },
            request_audit={"topic_signal": "test"},
            response_audit={"choices": []},
        )

    def label(self, _source, _matches, _images):
        return MimoLabelResult(
            candidate={
                "title": "手机屏幕色斑判定",
                "subtitles": ["屏幕有色斑怎么判"],
                "content": "按色斑标准核验；证据不足时待人工确认。",
                "category_l1": "显示问题",
                "category_l2": "色斑",
                "layer": "L2",
                "knowledge_form": "具体判定",
                "standard_refs": ["PHONE-DISPLAY-001"],
                "applicable_scope": "适用于手机屏幕色斑异常",
                "confidence": 0.91,
                "reasoning_summary": "会话与检索标准的显示问题/色斑一致。",
                "needs_human_review": False,
                "image_evidence_summary": "图片已接收，仍需人工确认细节。",
            },
            request_audit={"source": "test"},
            response_audit={"choices": []},
        )

    def label_topic(self, _topic, _matches):
        return MimoLabelResult(
            candidate={
                "title": "手机屏幕色斑如何通过图片核验",
                "subtitles": ["显示异常", "屏幕 / 显示异常"],
                "content": "核验流程：\n1. 明确异常所在屏幕区域。\n2. 补充清晰近景、全景及不同角度图片。\n3. 对照当前有效显示质检标准。\n4. 证据不足时重点复核并转人工。",
                "category_l1": "显示问题",
                "category_l2": "色斑",
                "layer": "L2",
                "knowledge_form": "流程方法",
                "standard_refs": [],
                "applicable_scope": "手机",
                "confidence": 0.65,
                "reasoning_summary": "显示问题通常需要结合现场图片和生效标准沉淀核验流程。",
                "needs_human_review": True,
                "image_evidence_summary": "聚合案例包含可用图片。",
            },
            request_audit={"topic": "test"},
            response_audit={"choices": []},
        )

    def review_topic(self, _topic, _draft, _matches):
        return MimoLabelResult(
            candidate={
                "decision": "通过",
                "knowledge_value": "值得沉淀",
                "error_type": "",
                "reason": "转写草稿已沉淀为流程型知识，标准引用和证据链可追溯。",
                "standard_consistency": "一致",
                "evidence_sufficiency": "充分",
                "confidence": 0.88,
                "priority_review": False,
            },
            request_audit={"review": "test"},
            response_audit={"choices": []},
        )

    def review_cluster_pair(self, _left, _right, similarity, threshold):
        decision = "同一主题" if similarity >= threshold else "不同主题"
        return MimoLabelResult(
            candidate={
                "decision": decision,
                "topic_label": "测试主题",
                "reason": "根据两条记录的意图、对象和处理目标进行判断。",
                "key_difference": "" if decision == "同一主题" else "处理目标不同",
                "confidence": 0.9,
            },
            request_audit={"cluster_pair": "test"},
            response_audit={"choices": []},
        )


class _FakeEmbedding:
    config = SimpleNamespace(model="semantic-cluster-test")

    def embed_texts(self, texts):
        assert len(texts) == 3
        return [
            [1.0, 0.0],
            [0.99, 0.01],
            [0.0, 1.0],
        ]


def test_mimo_cluster_units_keeps_clear_secondary_topics() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="mimo-v2.5-test",
        )
    )
    client._post = lambda _payload: {
        "choices": [
            {
                "message": {
                    "content": json.dumps(
                        {
                            "conversation_type": "multi_topic",
                            "reason": "会话同时咨询前摄像头状态和按键触发闪屏，需要不同知识回答。",
                            "topics": [
                                {
                                    "normalized_issue": "前摄像头疑似异物或更换的核验",
                                    "product_category": "手机",
                                    "scope_type": "品类专用",
                                    "platform": "通用",
                                    "brand": "通用",
                                    "model_scope": "通用",
                                    "category_l1": "功能问题",
                                    "category_l2": "摄像头功能",
                                    "intent": "检测核验",
                                    "subject": "前摄像头",
                                    "phenomenon": "疑似异物或更换",
                                    "judgment_target": "核验前摄像头是否存在异物或更换",
                                    "resolution_mode": "结合外观证据转人工核验",
                                    "standard_path": "摄像头功能核验",
                                    "threshold_or_exception": "待确认",
                                    "evidence_summary": "聊天明确询问前摄像头是否正常，但没有形成最终结论。",
                                    "confidence": 0.78,
                                    "requires_review": False,
                                },
                                {
                                    "normalized_issue": "按开机键触发屏幕闪烁的判定",
                                    "product_category": "手机",
                                    "scope_type": "品类专用",
                                    "platform": "通用",
                                    "brand": "通用",
                                    "model_scope": "通用",
                                    "category_l1": "显示问题",
                                    "category_l2": "屏幕闪烁",
                                    "intent": "标准判定",
                                    "subject": "屏幕",
                                    "phenomenon": "按开机键时闪烁",
                                    "judgment_target": "判断按键触发的屏幕闪烁是否属于显示异常",
                                    "resolution_mode": "对照闪屏标准判定",
                                    "standard_path": "屏幕显示异常判定",
                                    "threshold_or_exception": "仅按开机键时触发",
                                    "evidence_summary": "聊天和上游视频分析均确认按开机键时出现闪烁。",
                                    "confidence": 0.94,
                                    "requires_review": False,
                                },
                            ],
                        },
                        ensure_ascii=False,
                    )
                }
            }
        ]
    }

    result = client.analyze_cluster_units(
        {
            "工单ID": "PHONE-MULTI-001",
            "产品类型": "手机",
            "聊天内容": "前摄像头正常吗？另外只有按开机键时屏幕才闪。",
            "核心问题": "前摄像头与屏幕闪烁问题",
            "判定结论": "屏幕闪烁按显示问题处理",
            "上游媒体分析摘要": "视频确认按开机键时屏幕闪烁。",
        }
    )

    assert result.candidate["conversation_type"] == "multi_topic"
    assert len(result.candidate["topics"]) == 2
    assert result.candidate["topics"][0]["subject"] == "前摄像头"
    assert result.candidate["topics"][1]["subject"] == "屏幕"


def test_mimo_cluster_units_sends_images_and_videos_to_media_model() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="mimo-v2.5-pro",
            media_model="mimo-v2.5",
            cluster_media_policy="always",
        )
    )
    captured_payloads: list[dict[str, object]] = []

    def fake_post(payload):
        captured_payloads.append(payload)
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "conversation_type": "single_topic",
                                "reason": "聊天、图片和视频均在询问屏幕闪烁问题。",
                                "media_analysis": {
                                    "image_summary": "图片显示设备设置页面。",
                                    "video_summary": "视频显示按下开机键后屏幕闪烁。",
                                    "media_relevance": "高度相关",
                                    "used_for_topic_split": False,
                                    "requires_review": False,
                                },
                                "topics": [
                                    {
                                        "normalized_issue": "手机｜屏幕｜按键触发闪烁｜判断显示异常",
                                        "product_category": "手机",
                                        "scope_type": "品类专用",
                                        "platform": "通用",
                                        "brand": "通用",
                                        "model_scope": "通用",
                                        "category_l1": "显示问题",
                                        "category_l2": "屏幕闪烁",
                                        "intent": "标准判定",
                                        "subject": "屏幕",
                                        "phenomenon": "按下开机键后闪烁",
                                        "judgment_target": "判断是否属于显示异常",
                                        "resolution_mode": "结合视频现象对照标准判定",
                                        "standard_path": "屏幕显示异常判定",
                                        "threshold_or_exception": "仅按下开机键时触发",
                                        "evidence_summary": "视频直接显示按下开机键后屏幕闪烁。",
                                        "confidence": 0.94,
                                        "requires_review": False,
                                    }
                                ],
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    client._post = fake_post  # type: ignore[method-assign]
    result = client.analyze_cluster_units(
        {
            "工单ID": "PHONE-MEDIA-001",
            "产品类型": "手机",
            "聊天内容": "按下开机键后屏幕会闪，请看图片和视频。",
            "图片链接": "https://example.com/screen.jpg",
            "视频链接": "https://example.com/screen.mp4",
        }
    )

    assert captured_payloads[0]["model"] == "mimo-v2.5"
    content = captured_payloads[0]["messages"][1]["content"]
    assert [part["type"] for part in content] == ["text", "image_url", "video_url"]
    assert result.request_audit["media"]["mode"] == "mimo-direct-multimodal"
    assert result.request_audit["media"]["images"][0]["status"] == "attached"
    assert result.request_audit["media"]["videos"][0]["status"] == "attached"
    assert result.candidate["media_analysis"]["media_relevance"] == "相关"
    assert result.candidate["media_analysis"]["video_summary"] == "视频显示按下开机键后屏幕闪烁。"


def test_mimo_cluster_units_drops_corrupted_video_and_keeps_images() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="mimo-v2.5-pro",
            media_model="mimo-v2.5",
            cluster_media_policy="always",
        )
    )
    payloads: list[dict[str, object]] = []

    def fake_post(payload):
        payloads.append(payload)
        if len(payloads) == 1:
            raise MimoError(
                "MiMo HTTP 400: Multimodal data is corrupted or cannot be processed."
            )
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "conversation_type": "single_topic",
                                "reason": "视频不可读，图片和聊天仍指向同一外观问题。",
                                "media_analysis": {
                                    "image_summary": "图片显示相机型号标签。",
                                    "video_summary": "视频无法读取。",
                                    "media_relevance": "相关",
                                    "used_for_topic_split": False,
                                    "requires_review": True,
                                },
                                "topics": [
                                    {
                                        "normalized_issue": "单电/微单机身｜型号标签｜型号不一致｜确认型号",
                                        "product_category": "单电/微单机身",
                                        "scope_type": "品类专用",
                                        "platform": "通用",
                                        "brand": "通用",
                                        "model_scope": "通用",
                                        "category_l1": "基本情况",
                                        "category_l2": "型号确认",
                                        "intent": "检测核验",
                                        "subject": "型号标签",
                                        "phenomenon": "外观型号与标签型号不一致",
                                        "judgment_target": "确认实际型号",
                                        "resolution_mode": "结合图片和查询结果核验",
                                        "standard_path": "相机型号核验",
                                        "threshold_or_exception": "无明确阈值",
                                        "evidence_summary": "图片和聊天显示型号信息不一致。",
                                        "confidence": 0.82,
                                        "requires_review": False,
                                    }
                                ],
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    client._post = fake_post  # type: ignore[method-assign]
    result = client.analyze_cluster_units(
        {
            "工单ID": "CAMERA-MEDIA-001",
                "产品类型": "单电/微单机身",
            "聊天内容": "外观型号和标签型号不一致，请结合图片视频确认。",
            "图片链接": "https://example.com/camera.jpg",
            "视频链接": "https://example.com/broken.mp4",
        }
    )

    first_content = payloads[0]["messages"][1]["content"]
    second_content = payloads[1]["messages"][1]["content"]
    assert [part["type"] for part in first_content] == [
        "text",
        "image_url",
        "video_url",
    ]
    assert [part["type"] for part in second_content] == ["text", "image_url"]
    assert result.request_audit["media"]["mode"] == (
        "mimo-direct-multimodal-video-fallback"
    )
    assert result.request_audit["media"]["videos"][0]["status"] == "unavailable"
    assert result.candidate["media_analysis"]["requires_review"] is True
    assert result.candidate["topics"][0]["requires_review"] is True


def test_mimo_cluster_units_on_demand_defers_images_for_sufficient_chat() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="mimo-v2.5",
            media_model="mimo-v2.5",
            cluster_media_policy="on_demand",
            cluster_media_min_text_chars=120,
        )
    )
    captured_payloads: list[dict[str, object]] = []

    def fake_post(payload):
        captured_payloads.append(payload)
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "conversation_type": "single_topic",
                                "reason": "完整聊天已经明确说明屏幕色斑判定问题。",
                                "topics": [
                                    {
                                        "normalized_issue": "手机｜屏幕｜色斑｜判定",
                                        "product_category": "手机",
                                        "scope_type": "品类专用",
                                        "platform": "通用",
                                        "brand": "通用",
                                        "model_scope": "通用",
                                        "category_l1": "显示问题",
                                        "category_l2": "屏幕色斑",
                                        "intent": "标准判定",
                                        "subject": "屏幕",
                                        "phenomenon": "色斑",
                                        "judgment_target": "判断是否属于色斑",
                                        "resolution_mode": "对照显示标准判定",
                                        "standard_path": "屏幕色斑判定",
                                        "threshold_or_exception": "无明确阈值",
                                        "evidence_summary": "聊天完整描述了色斑现象和判定诉求。",
                                        "confidence": 0.92,
                                        "requires_review": False,
                                    }
                                ],
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    client._post = fake_post  # type: ignore[method-assign]
    result = client.analyze_cluster_units(
        {
            "工单ID": "PHONE-ON-DEMAND-001",
            "产品类型": "手机",
            "聊天内容": (
                "回收师说明手机屏幕在白色和灰色背景下均出现固定的局部彩色斑块，"
                "客服进一步确认异常位置不会随画面移动，双方最终都在咨询该现象"
                "应按屏幕色斑还是其他显示异常进行判定。聊天中还补充说明重新开机、"
                "切换多个纯色背景后异常区域保持不变，客服已明确要求按同一个显示问题"
                "继续核验，因此无需依赖图片才能识别本次聚类主题。"
            ),
            "图片链接": "https://example.com/screen.jpg",
        }
    )

    content = captured_payloads[0]["messages"][1]["content"]
    assert [part["type"] for part in content] == ["text"]
    assert result.request_audit["media"]["mode"] == (
        "mimo-direct-on-demand-text-only"
    )
    assert result.request_audit["media"]["images"][0]["status"] == "deferred"
    assert client.metrics_snapshot()["cluster_media_deferred_rows"] == 1


def test_mimo_cluster_units_on_demand_keeps_explicit_visual_evidence() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="mimo-v2.5",
            media_model="mimo-v2.5",
            cluster_media_policy="on_demand",
            cluster_media_min_text_chars=120,
        )
    )
    source_row = {
        "工单ID": "PHONE-ON-DEMAND-VISUAL-001",
        "产品类型": "手机",
        "聊天内容": (
            "回收师已经详细说明摄像头附近存在异常痕迹，并补充了多个角度的描述。"
            "客服继续追问异常是否会影响拍照功能，双方围绕同一个摄像头问题进行了"
            "多轮确认。请看图中圈出的摄像头位置，最终判定仍需要结合该处实拍细节。"
            "聊天还说明切换拍照模式、重新启动设备并清洁镜片后现象仍然存在，"
            "因此文字背景虽然充分，但具体痕迹形态仍必须查看现场图片。"
        ),
        "图片链接": "https://example.com/camera.jpg",
    }

    assert client.can_batch_cluster_units(source_row) is False


def test_mimo_cluster_units_on_demand_keeps_video_evidence() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="mimo-v2.5",
            media_model="mimo-v2.5",
            cluster_media_policy="on_demand",
            cluster_media_min_text_chars=80,
        )
    )
    captured_payloads: list[dict[str, object]] = []

    def fake_post(payload):
        captured_payloads.append(payload)
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "conversation_type": "single_topic",
                                "reason": "视频展示了屏幕动态闪烁。",
                                "media_analysis": {
                                    "image_summary": "无图片",
                                    "video_summary": "视频显示屏幕持续闪烁。",
                                    "media_relevance": "相关",
                                    "used_for_topic_split": False,
                                    "requires_review": False,
                                },
                                "topics": [
                                    {
                                        "normalized_issue": "手机｜屏幕｜闪屏｜判定",
                                        "product_category": "手机",
                                        "scope_type": "品类专用",
                                        "platform": "通用",
                                        "brand": "通用",
                                        "model_scope": "通用",
                                        "category_l1": "显示问题",
                                        "category_l2": "屏幕闪烁",
                                        "intent": "标准判定",
                                        "subject": "屏幕",
                                        "phenomenon": "动态闪烁",
                                        "judgment_target": "判断是否属于闪屏",
                                        "resolution_mode": "结合视频对照标准判定",
                                        "standard_path": "屏幕闪屏判定",
                                        "threshold_or_exception": "无明确阈值",
                                        "evidence_summary": "聊天和视频均说明屏幕闪烁。",
                                        "confidence": 0.94,
                                        "requires_review": False,
                                    }
                                ],
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    client._post = fake_post  # type: ignore[method-assign]
    client.analyze_cluster_units(
        {
            "工单ID": "PHONE-ON-DEMAND-VIDEO-001",
            "产品类型": "手机",
            "聊天内容": (
                "回收师和客服已经通过多轮问答确认设备只有在亮屏状态下出现持续闪烁，"
                "本次需要结合上传的视频确认该动态现象是否属于闪屏。"
            ),
            "视频链接": "https://example.com/screen.mp4",
        }
    )

    content = captured_payloads[0]["messages"][1]["content"]
    assert [part["type"] for part in content] == ["text", "video_url"]
    assert client.metrics_snapshot()["cluster_media_attached_rows"] == 1


def test_mimo_cluster_units_never_uses_text_only_and_allows_batching() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="mimo-v2.5",
            media_model="mimo-v2.5",
            cluster_media_policy="never",
        )
    )
    captured_payloads: list[dict[str, object]] = []

    def fake_post(payload):
        captured_payloads.append(payload)
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "conversation_type": "single_topic",
                                "reason": "仅根据聊天文本识别屏幕闪烁问题。",
                                "topics": [
                                    {
                                        "normalized_issue": "手机｜屏幕｜闪屏｜判定",
                                        "product_category": "手机",
                                        "scope_type": "品类专用",
                                        "platform": "通用",
                                        "brand": "通用",
                                        "model_scope": "通用",
                                        "category_l1": "显示问题",
                                        "category_l2": "屏幕闪烁",
                                        "intent": "标准判定",
                                        "subject": "屏幕",
                                        "phenomenon": "动态闪烁",
                                        "judgment_target": "判断是否属于闪屏",
                                        "resolution_mode": "根据聊天描述对照标准判定",
                                        "standard_path": "屏幕闪屏判定",
                                        "threshold_or_exception": "无明确阈值",
                                        "evidence_summary": "聊天明确说明屏幕持续闪烁。",
                                        "confidence": 0.92,
                                        "requires_review": False,
                                    }
                                ],
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    client._post = fake_post  # type: ignore[method-assign]
    source_row = {
        "工单ID": "PHONE-TEXT-ONLY-001",
        "产品类型": "手机",
        "聊天内容": "请看图和视频，屏幕一直闪，应该怎么判？",
        "图片链接": "https://example.com/screen.jpg",
        "视频链接": "https://example.com/screen.mp4",
    }

    assert client.can_batch_cluster_units(source_row) is True
    result = client.analyze_cluster_units(source_row)

    content = captured_payloads[0]["messages"][1]["content"]
    assert [part["type"] for part in content] == ["text"]
    assert result.request_audit["media"]["mode"] == "mimo-direct-text-only"
    assert result.request_audit["media"]["images"][0]["status"] == "skipped"
    assert result.request_audit["media"]["videos"][0]["status"] == "skipped"


def test_cluster_unit_validation_normalizes_basic_information_category() -> None:
    payload = {
        "conversation_type": "single_topic",
        "reason": "会话只询问一个基本信息问题。",
        "topics": [
            {
                "normalized_issue": "手机｜包装｜塑封状态｜判断是否全新",
                "product_category": "手机",
                "scope_type": "品类专用",
                "platform": "通用",
                "brand": "通用",
                "model_scope": "通用",
                "category_l1": "基本信息",
                "category_l2": "全新机判定",
                "intent": "标准判定",
                "subject": "包装",
                "phenomenon": "存在塑封",
                "judgment_target": "判断是否为全新机",
                "resolution_mode": "根据包装状态判断",
                "standard_path": "全新机判定",
                "threshold_or_exception": "无明确阈值",
                "evidence_summary": "聊天明确询问塑封机器是否为全新机。",
                "confidence": 0.8,
                "requires_review": False,
            }
        ],
    }

    result = mimo_module._validate_cluster_units(payload)

    assert result["topics"][0]["category_l1"] == "成色与回收标准"


@pytest.mark.parametrize(
    ("conversation_type", "topic_count", "error_message"),
    [
        ("single_topic", 0, "single_topic 必须恰好包含 1 个问题单元"),
        ("single_topic", 2, "single_topic 必须恰好包含 1 个问题单元"),
        ("multi_topic", 1, "multi_topic 必须包含 2 到 3 个问题单元"),
        ("multi_topic", 4, "multi_topic 必须包含 2 到 3 个问题单元"),
        ("uncertain", 2, "uncertain 最多保留 1 个暂定问题单元"),
    ],
)
def test_cluster_unit_validation_enforces_atomic_topic_count_by_conversation_type(
    conversation_type: str,
    topic_count: int,
    error_message: str,
) -> None:
    topic = {
        "normalized_issue": "手机｜屏幕｜色斑｜判断是否属于色斑",
        "product_category": "手机",
        "scope_type": "品类专用",
        "platform": "通用",
        "brand": "通用",
        "model_scope": "通用",
        "category_l1": "显示问题",
        "category_l2": "色斑",
        "intent": "标准判定",
        "subject": "屏幕",
        "phenomenon": "色斑",
        "judgment_target": "判断屏幕异常是否属于色斑",
        "resolution_mode": "结合聊天证据核验",
        "standard_path": "屏幕色斑判定",
        "threshold_or_exception": "无明确阈值",
        "evidence_summary": "聊天支持屏幕色斑问题。",
        "confidence": 0.93,
        "requires_review": False,
    }
    payload = {
        "conversation_type": conversation_type,
        "reason": "测试原子问题数量边界。",
        "topics": [
            {
                **topic,
                "normalized_issue": f"{topic['normalized_issue']}｜{index}",
            }
            for index in range(topic_count)
        ],
    }

    with pytest.raises(MimoError, match=error_message):
        mimo_module._validate_cluster_units(payload)


def test_cluster_fusion_guardrail_keeps_explicit_text_multi_topics() -> None:
    text_candidate = {
        "conversation_type": "multi_topic",
        "reason": "文字明确包含屏幕漏液和电池健康度两个问题。",
        "topics": [
            {
                "normalized_issue": "平板｜屏幕｜漏液｜判定标准",
                "product_category": "平板",
                "requires_review": False,
            },
            {
                "normalized_issue": "平板｜电池｜健康度无法读取｜操作指引",
                "product_category": "平板",
                "requires_review": False,
            },
        ],
    }
    media_candidate = {
        "conversation_type": "single_topic",
        "reason": "图片只展示电池健康度。",
        "media_analysis": {
            "image_summary": "图片展示电池健康度页面。",
            "video_summary": "无视频",
            "media_relevance": "相关",
            "used_for_topic_split": False,
            "requires_review": False,
        },
        "topics": [
            {
                "normalized_issue": "平板｜电池｜健康度无法读取｜操作指引",
                "product_category": "平板",
                "requires_review": False,
            }
        ],
    }
    fused_candidate = {
        "conversation_type": "single_topic",
        "reason": "融合模型错误地只保留电池问题。",
        "topics": media_candidate["topics"],
    }

    result = mimo_module._enforce_cluster_fusion_guardrails(
        fused_candidate,
        text_candidate,
        media_candidate,
        {"images": [], "videos": []},
    )

    assert result["conversation_type"] == "multi_topic"
    assert len(result["topics"]) == 2
    assert "保留MiMo Pro" in result["reason"]


def test_cluster_fusion_guardrail_keeps_new_media_topic_and_flags_conflict() -> None:
    text_candidate = {
        "conversation_type": "single_topic",
        "reason": "文字只询问相机倍数。",
        "topics": [
            {
                "normalized_issue": "手机｜相机｜倍数是否正常｜判定",
                "product_category": "手机",
                "requires_review": False,
            }
        ],
    }
    media_candidate = {
        "conversation_type": "multi_topic",
        "reason": "图片还显示屏幕亮线。",
        "media_analysis": {
            "image_summary": "图片显示相机界面和垂直亮线。",
            "video_summary": "无视频",
            "media_relevance": "相关",
            "used_for_topic_split": True,
            "requires_review": False,
        },
        "topics": [
            {
                "normalized_issue": "手机｜相机｜倍数是否正常｜判定",
                "product_category": "手机",
                "requires_review": False,
            },
            {
                "normalized_issue": "手机｜屏幕｜垂直亮线｜判定",
                "product_category": "手机",
                "requires_review": False,
            },
        ],
    }
    fused_candidate = {
        "conversation_type": "single_topic",
        "reason": "融合模型遗漏媒体新增主题。",
        "topics": text_candidate["topics"],
    }

    result = mimo_module._enforce_cluster_fusion_guardrails(
        fused_candidate,
        text_candidate,
        media_candidate,
        {
            "images": [],
            "videos": [
                {
                    "status": "unavailable",
                    "url": "https://example.com/broken.mp4",
                }
            ],
        },
    )

    assert result["conversation_type"] == "multi_topic"
    assert len(result["topics"]) == 2
    assert result["media_analysis"]["requires_review"] is True
    assert all(topic["requires_review"] for topic in result["topics"])


def test_single_record_only_extracts_features_and_topic_model_saves_audit(tmp_path: Path) -> None:
    audit = AuditStore(tmp_path / "phone_mvp.db")
    features, run_id = generate_phone_candidate_rows(
        _source_rows(),
        _standards(),
        mimo_client=_FakeMimo(),
        audit_store=audit,
        image_downloader=_ReadyImageDownloader(),
    )

    feature = features[0]
    assert run_id
    assert feature["模型阶段状态"] == "topic_signal_labeled"
    assert feature["问题意图"] == "标准判定"
    assert feature["对象/部位"] == "屏幕"
    assert "意图:标准判定" in feature["主题标签"]
    assert not feature.get("模型知识内容")
    feature["关联标准项"] = "历史标准引用需保留"

    second = dict(feature)
    second["数据ID"] = "PHONE-002"
    second["工单ID"] = "PHONE-002"
    second["原始工单ID"] = "PHONE-002"
    topics, mapping, gaps, pending = build_topic_review_rows(
        [feature, second],
        _standards(),
        mimo_client=_FakeMimo(),
        audit_store=audit,
        run_id=run_id,
    )
    assert len(topics) == 1
    assert len(mapping) == 2
    assert not gaps
    assert not pending
    assert topics[0]["主题模型提供方"] == "mimo"
    assert topics[0]["知识分类"] == "质检标准"
    assert topics[0]["是否重点复核"] == "是"
    assert topics[0]["模型初标提供方"] == "mimo"
    assert topics[0]["模型初标结论"] == "通过"
    assert topics[0]["模型初标错误类型"] == ""
    assert topics[0]["模型初标是否值得沉淀"] == "值得沉淀"
    assert "历史标准引用需保留" in topics[0]["关联标准项"]

    connection = sqlite3.connect(audit.path)
    try:
        assert connection.execute("SELECT COUNT(*) FROM ingestion_records").fetchone()[0] == 1
        assert connection.execute("SELECT COUNT(*) FROM model_runs").fetchone()[0] == 3
        statuses = {
            row[0]
            for row in connection.execute("SELECT status FROM model_runs").fetchall()
        }
        assert "topic_stage_legacy_compatible" in statuses
        assert connection.execute("SELECT COUNT(*) FROM candidates").fetchone()[0] == 2
    finally:
        connection.close()


def test_rule_fallback_recomputes_stale_tag_cluster_keys() -> None:
    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": conversation,
            "核心问题": "AirPods 查找功能是否需要检测",
            "产品类型": "耳机",
            "模型主题一级分类": "功能问题",
            "模型主题二级分类": "传感器功能",
            "问题意图": "检测核验",
            "对象/部位": "传感器功能",
            "异常现象": "查找功能检测要求待确认",
            "解题方式": "历史回复核对",
            "标签聚类键": stale_key,
            "语义标注依据": "历史实际回复均指向查找功能是否需要检测。",
        }
        for record_id, conversation, stale_key in (
            ("AIRPODS-001", "AirPods 一代的查找功能要检查吗？", "旧键-AIRPODS-001"),
            ("AIRPODS-002", "AirPods 二代查找功能是否需要质检？", "旧键-AIRPODS-002"),
        )
    ]

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert topics[0]["主题样本数"] == 2
    assert len(mapping) == 2
    assert not gaps
    assert not pending
    assert {row["标签聚类键"] for row in mapping} == {
        "耳机 | 检测核验 | 传感器功能 | 查找功能检测要求待确认 | 历史回复核对"
    }


def test_topic_review_validation_requires_consistent_deposition_value() -> None:
    review = _validate_topic_review(
        {
            "decision": "通过",
            "knowledge_value": "值得沉淀",
            "error_type": "",
            "reason": "问题清楚、处理方式可复用。",
            "standard_consistency": "无可信标准",
            "evidence_sufficiency": "充分",
            "content_consistency": "一致",
            "image_necessity": "不需要",
            "title_quality": "清晰",
            "confidence": 0.9,
            "priority_review": False,
        }
    )

    assert review["knowledge_value"] == "值得沉淀"


def test_topic_review_validation_rejects_unworthy_pass_decision() -> None:
    with pytest.raises(MimoError, match="decision 必须为驳回"):
        _validate_topic_review(
            {
                "decision": "通过",
                "knowledge_value": "不值得沉淀",
                "error_type": "",
                "reason": "纯个案。",
                "standard_consistency": "无可信标准",
                "evidence_sufficiency": "充分",
                "content_consistency": "一致",
                "image_necessity": "不需要",
                "title_quality": "清晰",
                "confidence": 0.9,
                "priority_review": False,
            }
        )


def test_topic_stage_validation_accepts_supported_labels() -> None:
    classification = _validate_topic_stage(
        {
            "topic_stage": "质检流程",
            "knowledge_value": "值得沉淀",
            "stage_reason": "主要诉求是如何读取并核对设备信息。",
            "value_reason": "可形成稳定的检查步骤供后续复用。",
            "reusable_knowledge": "按设备页面、检测工具和实物信息依次核对。",
            "confidence": 0.91,
            "needs_human_review": False,
        }
    )

    assert classification["topic_stage"] == "质检流程"
    assert classification["knowledge_value"] == "值得沉淀"
    assert classification["confidence"] == 0.91


def test_topic_stage_validation_accepts_uncertain_as_manual_review_label() -> None:
    classification = _validate_topic_stage(
        {
            "topic_stage": "不确定",
            "knowledge_value": "值得沉淀",
            "stage_reason": "主题诉求需要人工结合业务背景确认。",
            "value_reason": "可能存在复用价值，但分类不好判断。",
            "reusable_knowledge": "待人工复核后补充分类和可沉淀内容。",
            "confidence": 0.51,
            "needs_human_review": False,
        }
    )

    assert classification["topic_stage"] == "不确定"
    assert classification["needs_human_review"] is True


def test_topic_stage_validation_rejects_unknown_stage() -> None:
    with pytest.raises(MimoError, match="topic_stage 不合法"):
        _validate_topic_stage(
            {
                "topic_stage": "售后服务",
                "knowledge_value": "值得沉淀",
                "stage_reason": "不在允许范围内。",
                "value_reason": "存在复用价值。",
                "reusable_knowledge": "测试内容。",
                "confidence": 0.8,
                "needs_human_review": True,
            }
        )


def test_classify_topic_stage_uses_dedicated_prompt_and_validator() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test-key",
            base_url="https://example.com/v1",
            model="mimo-test",
        )
    )
    captured_payloads: list[dict[str, object]] = []

    def fake_post(payload: dict[str, object]) -> dict[str, object]:
        captured_payloads.append(payload)
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "topic_stage": "案例解析",
                                "knowledge_value": "不值得沉淀",
                                "stage_reason": "最终结论依赖当前案例图片。",
                                "value_reason": "缺少可复用的判定边界或核验步骤。",
                                "reusable_knowledge": "仅记录了单个案例结论，无法提炼通用知识。",
                                "confidence": 0.88,
                                "needs_human_review": True,
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    client._post = fake_post  # type: ignore[method-assign]
    result = client.classify_topic_stage(
        {
            "theme_id": "C001",
            "normalized_issues": ["请看图片判断该处是否属于磕碰"],
            "evidence_summaries": ["当前设备提供了一张现场图片。"],
        }
    )

    assert result.candidate["topic_stage"] == "案例解析"
    assert result.candidate["knowledge_value"] == "不值得沉淀"
    assert result.request_audit["prompt_version"] == TOPIC_STAGE_PROMPT_VERSION
    assert captured_payloads[0]["temperature"] == 0.0
    user_content = captured_payloads[0]["messages"][1]["content"][0]["text"]  # type: ignore[index]
    assert "质检标准、质检流程、案例解析、课外常识、不确定" in user_content
    assert "人工校正后的核心问题、产品类型和判定结论" in user_content
    assert "基础常见问题若已有稳定知识覆盖" in user_content
    assert "少见案例、特殊案例或标准咨询" in user_content
    assert "单成员主题不等于不值得沉淀" in user_content


def test_unworthy_topic_is_classified_before_transcription_and_skips_draft_generation() -> None:
    class UnworthyTopicMimo:
        config = SimpleNamespace(model="mimo-topic-value-test")

        def classify_topic_stage(self, topic):
            assert topic["member_count"] == 1
            assert topic["standard_paths"] == []
            return MimoLabelResult(
                candidate={
                    "topic_stage": "案例解析",
                    "knowledge_value": "不值得沉淀",
                    "stage_reason": "结论依赖当前单个案例。",
                    "value_reason": "缺少可复用的判断边界或操作步骤。",
                    "reusable_knowledge": "当前只有单个案例结论。",
                    "confidence": 0.92,
                    "needs_human_review": True,
                },
                request_audit={"prompt_version": TOPIC_STAGE_PROMPT_VERSION},
                response_audit={},
            )

        def label_topic(self, *_args, **_kwargs):
            raise AssertionError("不值得沉淀的主题不应进入知识转写")

        def review_topic(self, *_args, **_kwargs):
            raise AssertionError("未转写主题不应执行内容质量初标")

    rows = [
        {
            "数据ID": "CASE-001",
            "来源记录ID": "SOURCE-CASE-001",
            "工单ID": "CASE-001",
            "聊天内容": "请看这张图片，这个位置算不算磕碰？客服回复这台机器正常。",
            "核心问题": "当前图片中的位置是否属于磕碰",
            "产品类型": "手机",
            "问题意图": "案例判定",
            "对象/部位": "外壳",
            "异常现象": "疑似磕碰",
            "解题方式": "查看当前案例图片",
            "语义标注依据": "只有当前案例图片和单次结论。",
            "图片链接": "https://example.com/case.jpg",
            "视频链接": "https://example.com/case.mp4",
            "图片处理状态": "可用:1",
            "语义标注图片必要性": "需要",
            "关联标准项": "历史标准引用需保留",
            "主标准路径": "历史标准路径仅保留",
            "来源版本": "qc-old-v1",
        }
    ]

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=UnworthyTopicMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert len(mapping) == 1
    assert not gaps
    assert len(pending) == 1
    assert pending[0]["数据ID"] == "CASE-001"
    assert "只有1条案例" in pending[0]["待聚合原因"]
    assert "后续命中相似主题后再进入沉淀判断" in pending[0]["待聚合原因"]
    assert topics[0]["主题状态"] == "incubating_pending_cluster"
    assert topics[0]["主题问题分类"] == "案例解析"
    assert topics[0]["主题沉淀价值"] == "不值得沉淀"
    assert topics[0]["主题转写状态"] == "skipped_not_worthy"
    assert topics[0]["模型初标状态"] == "topic_initial_review_skipped"
    assert topics[0]["主题图片链接"] == "https://example.com/case.jpg"
    assert topics[0]["主题视频链接"] == "https://example.com/case.mp4"
    assert topics[0]["主题来源记录ID"] == "SOURCE-CASE-001"
    assert "SOURCE-CASE-001" in topics[0]["主题视频来源"]
    assert topics[0]["关联标准项"] == "历史标准引用需保留"
    assert topics[0]["来源版本"] == "qc-old-v1"


def test_single_case_with_reusable_boundary_stays_out_of_pending_cluster() -> None:
    rows = [
        {
            "数据ID": "CASE-THRESHOLD-001",
            "工单ID": "CASE-THRESHOLD-001",
            "聊天内容": "屏幕点状瑕疵直径大于1mm怎么处理？",
            "核心问题": "屏幕点状瑕疵直径大于1mm如何判定",
            "产品类型": "手机",
            "问题意图": "标准判定",
            "对象/部位": "屏幕",
            "异常现象": "点状瑕疵",
            "解题方式": "测量点状瑕疵直径，大于1mm时按异常处理",
            "语义标注依据": "历史回复明确给出大于1mm的可复用边界。",
            "_原子阈值例外": "直径大于1mm",
        }
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        [],
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["主题沉淀价值"] == "值得沉淀"
    assert topics[0]["主题状态"] != "incubating_pending_cluster"


def test_single_case_with_explicit_steps_remains_worthy_despite_missing_threshold() -> None:
    rows = [
        {
            "数据ID": "CASE-STEPS-001",
            "工单ID": "CASE-STEPS-001",
            "聊天内容": "没有明确阈值，但必须先拍摄全景图片，再核对设备型号。",
            "核心问题": "缺少数值阈值时如何补充核验信息",
            "产品类型": "手机",
            "问题意图": "质检流程",
            "对象/部位": "整机",
            "异常现象": "信息不足",
            "解题方式": "没有明确阈值，但必须先拍摄全景图片，再核对设备型号。",
            "语义标注依据": "没有明确阈值，但必须先拍摄全景图片，再核对设备型号。",
        }
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        [],
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["主题沉淀价值"] == "值得沉淀"
    assert topics[0]["主题转写状态"] != "skipped_not_worthy"


def _non_reusable_case_rows(count: int) -> list[dict[str, object]]:
    return [
        {
            "数据ID": f"CASE-NON-REUSABLE-{index:03d}",
            "工单ID": f"CASE-NON-REUSABLE-{index:03d}",
            "聊天内容": "请看当前这台手机外壳是否正常，客服仅回复这台机器正常。",
            "核心问题": "当前机器外壳是否正常",
            "产品类型": "手机",
            "问题意图": "案例判定",
            "对象/部位": "外壳",
            "异常现象": "疑似划痕",
            "解题方式": "查看当前案例",
            "语义标注依据": "客服只回复当前机器正常，未提供其他说明。",
        }
        for index in range(1, count + 1)
    ]


def test_multiple_cases_without_reusable_evidence_do_not_become_worthy_by_count() -> None:
    topics, _mapping, gaps, _pending = build_topic_review_rows(
        _non_reusable_case_rows(2),
        [],
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert topics[0]["主题沉淀价值"] == "不值得沉淀"
    assert topics[0]["主题转写状态"] == "skipped_not_worthy"


def test_negated_rule_terms_do_not_count_as_reusable_evidence() -> None:
    rows = _non_reusable_case_rows(2)
    for row in rows:
        row["语义标注依据"] = "当前案例没有通用边界或核验步骤，只记录了机器正常。"

    topics, _mapping, gaps, _pending = build_topic_review_rows(
        rows,
        [],
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert topics[0]["主题沉淀价值"] == "不值得沉淀"
    assert topics[0]["主题转写状态"] == "skipped_not_worthy"


def test_question_only_threshold_does_not_count_as_source_rule() -> None:
    rows = _non_reusable_case_rows(1)
    rows[0].update(
        {
            "聊天内容": "屏幕点状瑕疵是否超过1mm就算异常？",
            "核心问题": "屏幕点状瑕疵超过1mm是否异常",
            "解题方式": "询问是否超过1mm",
            "语义标注依据": "用户只是在询问是否超过1mm，没有来源给出结论。",
            "_原子阈值例外": "是否超过1mm",
        }
    )

    topics, _mapping, gaps, _pending = build_topic_review_rows(
        rows,
        [],
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert topics[0]["主题沉淀价值"] == "不值得沉淀"
    assert topics[0]["主题转写状态"] == "skipped_not_worthy"


def test_model_worthy_decision_is_not_released_by_multiple_case_count_alone() -> None:
    class UnsupportedWorthyTopicMimo:
        config = SimpleNamespace(model="mimo-topic-value-test")

        def __init__(self) -> None:
            self.label_calls = 0

        def classify_topic_stage(self, topic):
            assert topic["member_count"] == 2
            return MimoLabelResult(
                candidate={
                    "topic_stage": "案例解析",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "存在两条相似案例。",
                    "value_reason": "仅根据案例数量判断存在复用价值。",
                    "reusable_knowledge": "两台机器均被回复为正常。",
                    "confidence": 0.91,
                    "needs_human_review": False,
                },
                request_audit={"prompt_version": TOPIC_STAGE_PROMPT_VERSION},
                response_audit={},
            )

        def label_topic(self, *_args, **_kwargs):
            self.label_calls += 1
            raise AssertionError("来源没有可复用规则时不应进入知识转写")

    client = UnsupportedWorthyTopicMimo()
    topics, _mapping, gaps, _pending = build_topic_review_rows(
        _non_reusable_case_rows(2),
        mimo_client=client,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert client.label_calls == 0
    assert topics[0]["主题沉淀价值"] == "不值得沉淀"
    assert topics[0]["主题转写状态"] == "skipped_not_worthy"


def test_untranscribed_topic_title_uses_issue_description_not_chat_header() -> None:
    class UnworthyTopicMimo:
        config = SimpleNamespace(model="mimo-topic-title-test")

        def classify_topic_stage(self, topic):
            assert topic["member_count"] == 1
            return MimoLabelResult(
                candidate={
                    "topic_stage": "案例解析",
                    "knowledge_value": "不值得沉淀",
                    "stage_reason": "结论依赖当前单个案例。",
                    "value_reason": "缺少可复用的判断边界或操作步骤。",
                    "reusable_knowledge": "当前只有单个案例结论。",
                    "confidence": 0.9,
                    "needs_human_review": True,
                },
                request_audit={"prompt_version": TOPIC_STAGE_PROMPT_VERSION},
                response_audit={},
            )

        def label_topic(self, *_args, **_kwargs):
            raise AssertionError("不值得沉淀的主题不应进入知识转写")

    rows = [
        {
            "数据ID": "CASE-TITLE-001",
            "工单ID": "CASE-TITLE-001",
            "聊天内容": (
                "26/07/15 18:22:53:53 问题类型：质检问题 "
                "问题描述：不确定手机外观是不是碎裂 转人工原因：回答内容无法理解\n"
                "请根据用户上传的设备实物图片，对照系统中的外观选项标准图示。"
            ),
            "核心问题": "",
            "产品类型": "手机",
            "问题意图": "案例判定",
            "对象/部位": "外观",
            "异常现象": "疑似碎裂",
            "解题方式": "查看当前案例图片",
            "语义标注依据": "用户要求根据当前实物图片确认外观是否碎裂。",
            "图片链接": "https://example.com/case-title.jpg",
            "图片处理状态": "可用:1",
        }
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        mimo_client=UnworthyTopicMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert topics[0]["主标题"] == "不确定手机外观是否为碎裂"
    assert "26/07/15" not in topics[0]["主标题"]
    assert "问题类型" not in topics[0]["主标题"]
    assert "转人工原因" not in topics[0]["主标题"]


def test_untranscribed_topic_title_prefers_direct_cluster_theme_name() -> None:
    class DirectClusterTitleMimo:
        config = SimpleNamespace(model="mimo-topic-cluster-title-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "会话和图片均指向外观碎裂判断。",
                    "topics": [
                        {
                            "normalized_issue": "手机｜外观｜疑似碎裂｜判断是否碎裂",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "外观问题",
                            "category_l2": "碎裂",
                            "intent": "案例判定",
                            "subject": "外观",
                            "phenomenon": "疑似碎裂",
                            "judgment_target": "判断外观是否碎裂",
                            "resolution_mode": "结合当前案例图片判断",
                            "standard_path": "待确认",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": f"{row['数据ID']} 的聊天和图片支持该主题。",
                            "confidence": 0.88,
                            "requires_review": True,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机外观碎裂判定",
                            "member_atomic_ids": [unit["unit_id"] for unit in units],
                            "scope_consistent": True,
                            "object_consistent": True,
                            "judgment_target_consistent": True,
                            "standard_path_consistent": True,
                            "threshold_exception_consistent": True,
                            "shared_knowledge_definition": "判断手机外观疑似碎裂是否成立。",
                            "merge_basis": "两条记录的品类、对象、现象和判定目标一致。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

        def classify_topic_stage(self, topic):
            assert topic["member_count"] == 2
            return MimoLabelResult(
                candidate={
                    "topic_stage": "案例解析",
                    "knowledge_value": "不值得沉淀",
                    "stage_reason": "当前只支持案例复核。",
                    "value_reason": "缺少可复用边界。",
                    "reusable_knowledge": "当前只有个案判断线索。",
                    "confidence": 0.91,
                    "needs_human_review": True,
                },
                request_audit={"prompt_version": TOPIC_STAGE_PROMPT_VERSION},
                response_audit={},
            )

        def label_topic(self, *_args, **_kwargs):
            raise AssertionError("不值得沉淀的主题不应进入知识转写")

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": (
                "26/07/15 18:22:53:53 问题类型：质检问题 "
                "问题描述：不确定手机外观是不是碎裂 转人工原因：回答内容无法理解\n"
                "请根据用户上传的设备实物图片判断。"
            ),
            "核心问题": "",
            "产品类型": "手机",
            "问题意图": "案例判定",
            "对象/部位": "外观",
            "异常现象": "疑似碎裂",
            "解题方式": "查看当前案例图片",
            "图片链接": f"https://example.com/{record_id}.jpg",
            "图片处理状态": "可用:1",
        }
        for record_id in ("CASE-TITLE-A", "CASE-TITLE-B")
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        mimo_client=DirectClusterTitleMimo(),
        clustering_mode="direct_mimo",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert topics[0]["主标题"] == "手机外观碎裂如何判定"
    assert "问题描述" not in topics[0]["主标题"]


def test_untranscribed_topic_title_normalizes_duplicate_cluster_label() -> None:
    class UnworthyTopicMimo:
        config = SimpleNamespace(model="mimo-topic-title-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检流程",
                    "knowledge_value": "不值得沉淀",
                    "stage_reason": "当前主题只用于人工价值复核。",
                    "value_reason": "当前批次不生成知识草稿。",
                    "reusable_knowledge": "",
                    "confidence": 0.9,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, *_args, **_kwargs):
            raise AssertionError("不值得沉淀的主题不应进入知识转写")

    rows = [
        {
            "数据ID": "CASE-DUPLICATE-TITLE-001",
            "工单ID": "CASE-DUPLICATE-TITLE-001",
            "聊天内容": "中框及外壳外观应该如何核验？",
            "产品类型": "手机",
            "问题意图": "检测核验",
            "对象/部位": "中框及外壳外观",
            "异常现象": "待确认",
            "解题方式": "现场图片/视频补充与案例证据核验",
            "图片链接": "https://example.com/title.jpg",
            "图片处理状态": "可用:1",
            "_聚类主题标题": (
                "中框及外壳外观 | 中框及外壳外观 | "
                "现场图片/视频补充与案例证据核验"
            ),
        }
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        mimo_client=UnworthyTopicMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert topics[0]["主标题"] == "中框及外壳外观如何核验"


def test_untranscribed_topic_title_keeps_natural_cluster_title() -> None:
    class UnworthyTopicMimo:
        config = SimpleNamespace(model="mimo-topic-title-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检流程",
                    "knowledge_value": "不值得沉淀",
                    "stage_reason": "当前主题只用于人工价值复核。",
                    "value_reason": "当前批次不生成知识草稿。",
                    "reusable_knowledge": "",
                    "confidence": 0.9,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, *_args, **_kwargs):
            raise AssertionError("不值得沉淀的主题不应进入知识转写")

    rows = [
        {
            "数据ID": "CASE-NATURAL-TITLE-001",
            "工单ID": "CASE-NATURAL-TITLE-001",
            "聊天内容": "传感器功能应该如何核验？",
            "产品类型": "手机",
            "问题意图": "检测核验",
            "对象/部位": "传感器功能",
            "异常现象": "待确认",
            "解题方式": "按测试步骤核验",
            "图片链接": "",
            "_聚类主题标题": "传感器功能如何核验",
        }
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        mimo_client=UnworthyTopicMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert topics[0]["主标题"] == "传感器功能如何核验"


def test_unexpected_topic_stage_failure_is_isolated_per_topic() -> None:
    class PartiallyFailingTopicStageMimo:
        config = SimpleNamespace(model="mimo-topic-value-test")

        def __init__(self) -> None:
            self.calls = 0

        def classify_topic_stage(self, _topic):
            self.calls += 1
            if self.calls == 1:
                raise RuntimeError("unexpected topic-stage failure")
            return MimoLabelResult(
                candidate={
                    "topic_stage": "案例解析",
                    "knowledge_value": "不值得沉淀",
                    "stage_reason": "结论依赖当前单个案例。",
                    "value_reason": "缺少可复用规则。",
                    "reusable_knowledge": "当前仅有个案结论。",
                    "confidence": 0.9,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, *_args, **_kwargs):
            raise AssertionError("不值得沉淀主题不应进入转写")

        def review_topic(self, *_args, **_kwargs):
            raise AssertionError("未转写主题不应进入内容质量初标")

    rows = [
        {
            "数据ID": "CASE-FAIL-001",
            "工单ID": "CASE-FAIL-001",
            "聊天内容": "请看这张图，这处算正常吗？",
            "核心问题": "当前案例是否正常",
            "产品类型": "手机",
            "问题意图": "案例判定",
            "对象/部位": "外壳",
            "异常现象": "待确认",
            "解题方式": "查看当前案例",
            "语义标注依据": "只有当前案例信息。",
        },
        {
            "数据ID": "CASE-OK-002",
            "工单ID": "CASE-OK-002",
            "聊天内容": "请看当前图片，这个位置是什么情况？",
            "核心问题": "当前案例是什么情况",
            "产品类型": "平板",
            "问题意图": "案例判定",
            "对象/部位": "屏幕",
            "异常现象": "待确认",
            "解题方式": "查看当前案例",
            "语义标注依据": "只有当前案例信息。",
        },
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        mimo_client=PartiallyFailingTopicStageMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 2
    assert {topic["主题分类状态"] for topic in topics} == {
        "topic_stage_classification_failed",
        "topic_stage_classified_model",
    }
    failed = next(
        topic
        for topic in topics
        if topic["主题分类状态"] == "topic_stage_classification_failed"
    )
    assert failed["主题分类重点复核"] == "是"
    assert "RuntimeError" in failed["主题分类错误"]


def test_worthy_topic_is_transcribed_then_receives_content_quality_review() -> None:
    calls: list[str] = []

    class WorthyTopicMimo:
        config = SimpleNamespace(model="mimo-topic-value-test")

        def classify_topic_stage(self, topic):
            calls.append("classify")
            assert topic["member_count"] == 2
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检流程",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "主要诉求是如何读取并核对设备信息。",
                    "value_reason": "两个案例提供了一致且可复用的检查步骤。",
                    "reusable_knowledge": "先进入设备信息页读取，再用检测工具交叉核对。",
                    "confidence": 0.94,
                    "needs_human_review": False,
                },
                request_audit={"prompt_version": TOPIC_STAGE_PROMPT_VERSION},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, use_standard_references=False):
            calls.append("transcribe")
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "title": "设备信息读取与交叉核对流程",
                    "subtitles": [],
                    "content": "1. 进入设备信息页读取配置。\n2. 使用检测工具交叉核对。\n3. 信息不一致时保留截图并人工复核。",
                    "category_l1": "信息查询",
                    "category_l2": "设备信息核对",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": [],
                    "applicable_scope": "手机-通用",
                    "recommended_reply": "请先进入设备信息页读取配置，再使用检测工具交叉核对；如结果不一致，请保留截图后人工复核。",
                    "confidence": 0.9,
                    "reasoning_summary": "两个案例均提供了相同的核对步骤。",
                    "needs_human_review": False,
                    "image_evidence_summary": "不依赖图片。",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(
            self,
            topic,
            _draft,
            _matches,
            use_standard_references=False,
        ):
            calls.append("quality_review")
            assert topic["topic_stage"] == "质检流程"
            assert topic["knowledge_value"] == "值得沉淀"
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "标题、内容、步骤和推荐回复均与主题证据一致。",
                    "standard_consistency": "无可信标准",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.93,
                    "priority_review": False,
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": conversation,
            "核心问题": "如何读取并核对设备信息",
            "产品类型": "手机",
            "问题意图": "信息查询",
            "对象/部位": "设备信息",
            "异常现象": "读取结果需核对",
            "解题方式": "先进入设备信息页读取，再使用检测工具交叉核对",
            "语义标注依据": conversation,
        }
        for record_id, conversation in (
            ("FLOW-001", "先进入设备信息页读取配置，再用检测工具核对。"),
            ("FLOW-002", "设备页面和检测工具需要依次交叉核对。"),
        )
    ]

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=WorthyTopicMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert calls == ["classify", "transcribe", "quality_review"]
    assert len(topics) == 1
    assert len(mapping) == 2
    assert not gaps
    assert not pending
    assert topics[0]["主题问题分类"] == "质检流程"
    assert topics[0]["知识分类"] == "质检流程"
    assert topics[0]["主题沉淀价值"] == "值得沉淀"
    assert topics[0]["主题转写状态"] == "topic_model_quality_failed"
    assert topics[0]["适用范围"] == "手机"
    assert topics[0]["模型初标结论"] == "需修改"
    assert "保留截图" in topics[0]["主题无来源内容"]


def test_topic_standard_retriever_runs_after_topic_value_gate() -> None:
    calls: list[str] = []
    standard = StandardCatalogItem(
        standard_id="KB-STD-001",
        title="设备信息读取标准",
        category_l1="质检标准",
        category_l2="",
        knowledge_type="业务沉淀标准",
        standard_path="CZ业务沉淀标准：KB-STD-001",
        keywords=["设备信息", "核对"],
        scope="手机",
        response_snippet="先读取设备信息，再使用检测工具交叉核对。",
        status="published",
        version="cz-snapshot-v1",
    )

    class StandardBackedTopicMimo:
        config = SimpleNamespace(model="mimo-standard-backed-test")

        def classify_topic_stage(self, _topic):
            calls.append("classify")
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "主题需要按已生效标准改写。",
                    "value_reason": "来源问题可复用。",
                    "reusable_knowledge": "设备信息读取和交叉核对。",
                    "confidence": 0.94,
                    "needs_human_review": False,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, matches, use_standard_references=False):
            calls.append("transcribe")
            assert use_standard_references is True
            assert [item.standard_id for item, _score in matches] == ["KB-STD-001"]
            return MimoLabelResult(
                candidate={
                    "title": "设备信息读取与交叉核对",
                    "subtitles": [],
                    "content": standard.response_snippet,
                    "category_l1": "质检标准",
                    "category_l2": "设备信息",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": ["KB-STD-001"],
                    "applicable_scope": "手机",
                    "recommended_reply": standard.response_snippet,
                    "confidence": 0.9,
                    "reasoning_summary": "依据已生效标准改写。",
                    "needs_human_review": False,
                    "image_evidence_summary": "不依赖图片。",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(self, _topic, _draft, _matches, use_standard_references=False):
            calls.append("quality_review")
            assert use_standard_references is True
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "内容与已生效标准一致。",
                    "standard_consistency": "一致",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.93,
                    "priority_review": False,
                },
                request_audit={},
                response_audit={},
            )

    def retrieve(topic_id, _rows, query):
        calls.append("retrieve")
        assert calls == ["classify", "retrieve"]
        assert topic_id
        assert query["产品类型"] == "手机"
        return [(standard, 0.91)], {
            "source": "headquarters_standard",
            "status": "success",
            "knowledge_version": "cz-snapshot-v1",
        }

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": conversation,
            "核心问题": "如何读取并核对设备信息",
            "产品类型": "手机",
            "问题意图": "信息查询",
            "对象/部位": "设备信息",
            "异常现象": "读取结果需核对",
            "解题方式": "先进入设备信息页读取，再使用检测工具交叉核对",
            "语义标注依据": conversation,
        }
        for record_id, conversation in (
            ("202608100001", "先进入设备信息页读取配置，再用检测工具核对。"),
            ("202608100002", "设备页面和检测工具需要依次交叉核对。"),
        )
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=StandardBackedTopicMimo(),
        clustering_mode="rule",
        use_standard_references=True,
        topic_standard_retriever=retrieve,
    )

    assert calls == ["classify", "retrieve", "transcribe", "quality_review"]
    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert "KB-STD-001" in topics[0]["关联标准项"]
    assert "cz-snapshot-v1" in topics[0]["主题标准版本"]


def test_missing_topic_standard_transcribes_experience_candidate_for_manual_review() -> None:
    calls: list[str] = []

    class ExperienceSupplementTopicMimo:
        config = SimpleNamespace(model="mimo-standard-required-test")

        def classify_topic_stage(self, _topic):
            calls.append("classify")
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "主题需要按已生效标准改写。",
                    "value_reason": "来源问题可复用。",
                    "reusable_knowledge": "设备信息读取和交叉核对。",
                    "confidence": 0.94,
                    "needs_human_review": False,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, matches, *, use_standard_references=False):
            calls.append("transcribe")
            assert use_standard_references is True
            assert not matches
            return MimoLabelResult(
                candidate={
                    "title": "设备信息读取与交叉核对方法",
                    "subtitles": [],
                    "content": (
                        "判定规则：设备信息不一致时一律判定为异常。\n"
                        "适用范围：手机设备信息核对。\n"
                        "处理步骤：先进入设备信息页读取配置，再使用检测工具交叉核对。\n"
                        "适用边界：当前主题未命中总部标准，其他设备或信息冲突时补充证据后再确认。"
                    ),
                    "category_l1": "基本情况",
                    "category_l2": "设备信息",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": [],
                    "applicable_scope": "手机",
                    "recommended_reply": "您好，请先进入设备信息页读取配置，再使用检测工具交叉核对；信息冲突时补充证据后再确认。",
                    "confidence": 0.9,
                    "reasoning_summary": "依据来源会话中的设备信息核对步骤整理。",
                    "needs_human_review": True,
                    "image_evidence_summary": "不依赖图片。",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(self, _topic, _draft, matches, *, use_standard_references=False):
            calls.append("quality_review")
            assert use_standard_references is True
            assert not matches
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "草稿仅使用来源会话中的核对步骤，作为经验补充候选。",
                    "standard_consistency": "无可信标准",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.9,
                    "priority_review": True,
                },
                request_audit={},
                response_audit={},
            )

    def retrieve(_topic_id, _rows, _query):
        calls.append("retrieve")
        return [], {
            "source": "headquarters_standard",
            "status": "error",
            "error": "RuntimeError: CZ 标准服务不可用",
        }

    rows = [
        {
            "数据ID": "202608100001",
            "工单ID": "202608100001",
            "聊天内容": "如何读取并核对设备信息？",
            "核心问题": "如何读取并核对设备信息",
            "产品类型": "手机",
            "问题意图": "信息查询",
            "对象/部位": "设备信息",
            "异常现象": "读取结果需核对",
            "解题方式": "先进入设备信息页读取，再使用检测工具交叉核对",
            "语义标注依据": "聊天明确询问设备信息核对方法。",
        }
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=ExperienceSupplementTopicMimo(),
        clustering_mode="rule",
        use_standard_references=True,
        topic_standard_retriever=retrieve,
        require_standard_match=True,
    )

    assert calls == ["classify", "retrieve", "transcribe", "quality_review"]
    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["主题转写状态"] == "topic_model_quality_failed"
    assert topics[0]["主题标准检索状态"] == "error"
    assert topics[0]["关联标准项"] == ""
    assert topics[0]["知识来源"] == "方向二经验补充候选"
    assert topics[0]["模型初标重点复核"] == "是"
    assert topics[0]["模型初标结论"] == "需修改"
    assert topics[0]["模型初标是否值得沉淀"] == "待确认"
    assert topics[0]["模型初标错误类型"] == "标准未覆盖/标准召回不足"
    assert topics[0]["推荐回复"] == ""
    assert "信息不一致时一律判定为异常" not in topics[0]["知识内容"]
    assert topics[0]["知识内容"].startswith("当前来源未提供可直接套用的明确规则")
    assert "回收师" not in topics[0]["知识内容"]


def test_rule_fallback_never_promotes_case_threshold_without_standard() -> None:
    rows = [
        {
            "数据ID": "FALLBACK-001",
            "工单ID": "FALLBACK-001",
            "聊天内容": "后摄镜片和保护圈之间有缝隙，自测约0.3-0.4mm。",
            "核心问题": "后摄镜片与保护圈之间的缝隙如何判定",
            "判定结论": "缝隙未达到大于0.5mm的判定阈值，按正常外观状态处理。",
            "判定依据": (
                "外壳组件衔接处缝隙大于0.5mm才判定；"
                "当前自测为0.3-0.4mm，未达到阈值。"
            ),
            "参考话术": (
                "根据图片和测量，缝隙约0.3-0.4mm，未达到0.5mm阈值，"
                "按正常外观状态继续质检。"
            ),
            "产品类型": "平板电脑",
            "一级分类": "外壳外观情况",
            "二级分类": "外壳其他现象",
            "问题意图": "标准判定",
            "对象/部位": "后摄镜片与保护圈衔接处",
            "异常现象": "缝隙",
            "解题方式": "测量缝隙宽度并与0.5mm阈值比较",
            "语义标注依据": "来源明确记录自测尺寸、判定阈值和处理结论。",
        }
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=None,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=True,
        transcribe_all_admitted_topics=True,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["知识来源"] == "方向二经验补充候选"
    assert topics[0]["模型初标结论"] == "需修改"
    assert topics[0]["模型初标错误类型"] == "标准未覆盖/标准召回不足"
    assert topics[0]["知识内容"].startswith("当前来源未提供可直接套用的明确规则")
    assert "测量缝隙宽度" in topics[0]["知识内容"]
    assert all(
        value not in topics[0]["知识内容"]
        for value in ("0.3-0.4mm", "0.5mm", "大于0.5mm", "按正常外观状态")
    )
    assert all(
        value not in topics[0]["推荐回复"]
        for value in ("0.3-0.4mm", "0.5mm", "大于0.5mm", "按正常外观状态")
    )


def test_missing_headquarters_standard_falls_back_to_local_quality_standard() -> None:
    calls: list[str] = []
    selected_standard_ids: list[str] = []
    local_standard = StandardCatalogItem(
        standard_id="LOCAL-PHONE-INFO-001",
        title="手机设备信息读取与交叉核对",
        category_l1="质检流程",
        category_l2="设备信息",
        knowledge_type="本地质检标准",
        standard_path="本地质检标准：手机/设备信息",
        keywords=["手机", "设备信息", "读取", "交叉核对"],
        scope="手机",
        response_snippet="先读取设备信息，再使用检测工具交叉核对。",
        status="published",
        version="local-qc-v20260817",
    )

    class LocalStandardTopicMimo:
        config = SimpleNamespace(model="mimo-local-standard-fallback-test")

        def classify_topic_stage(self, _topic):
            calls.append("classify")
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检流程",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "主题需要按可用质检标准改写。",
                    "value_reason": "来源问题可复用。",
                    "reusable_knowledge": "设备信息读取和交叉核对。",
                    "confidence": 0.94,
                    "needs_human_review": False,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, matches, *, use_standard_references=False):
            calls.append("transcribe")
            assert use_standard_references is True
            selected_standard_ids.extend(
                item.standard_id for item, _score in matches
            )
            assert selected_standard_ids == ["LOCAL-PHONE-INFO-001"]
            return MimoLabelResult(
                candidate={
                    "title": "手机设备信息读取与交叉核对方法",
                    "subtitles": [],
                    "content": local_standard.response_snippet,
                    "category_l1": "质检流程",
                    "category_l2": "设备信息",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": ["LOCAL-PHONE-INFO-UNKNOWN"],
                    "applicable_scope": "手机",
                    "recommended_reply": local_standard.response_snippet,
                    "confidence": 0.9,
                    "reasoning_summary": "依据本地质检标准改写。",
                    "needs_human_review": True,
                    "image_evidence_summary": "不依赖图片。",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(self, _topic, _draft, matches, *, use_standard_references=False):
            calls.append("quality_review")
            assert use_standard_references is True
            assert [item.standard_id for item, _score in matches] == [
                "LOCAL-PHONE-INFO-001"
            ]
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "草稿依据本地质检标准，需人工确认后送审。",
                    "standard_consistency": "一致",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.9,
                    "priority_review": True,
                },
                request_audit={},
                response_audit={},
            )

    def retrieve(_topic_id, _rows, _query):
        calls.append("retrieve")
        return [], {
            "source": "headquarters_standard",
            "status": "error",
            "knowledge_version": "cz-snapshot-v1",
            "error": "RuntimeError: CZ 标准服务不可用",
        }

    rows = [
        {
            "数据ID": "202608100001",
            "工单ID": "202608100001",
            "聊天内容": "如何读取并核对手机设备信息？",
            "核心问题": "如何读取并核对设备信息",
            "产品类型": "手机",
            "问题意图": "信息查询",
            "对象/部位": "设备信息",
            "异常现象": "读取结果需核对",
            "解题方式": "先进入设备信息页读取，再使用检测工具交叉核对",
            "语义标注依据": "聊天明确询问设备信息核对方法。",
        }
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        standard_catalog=[local_standard],
        mimo_client=LocalStandardTopicMimo(),
        clustering_mode="rule",
        use_standard_references=True,
        topic_standard_retriever=retrieve,
        require_standard_match=True,
    )

    assert calls == ["classify", "retrieve", "transcribe", "quality_review"]
    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["知识来源"] == "方向二本地质检标准候选"
    assert "LOCAL-PHONE-INFO-001" in topics[0]["关联标准项"]
    assert "LOCAL-PHONE-INFO-UNKNOWN" not in topics[0]["关联标准项"]
    assert topics[0]["主题标准检索来源"] == "local_quality_standard"
    assert topics[0]["主题标准检索状态"] == "fallback_match"
    assert "CZ 标准服务不可用" in topics[0]["主题标准检索错误"]
    assert topics[0]["模型初标重点复核"] == "是"


def test_experience_only_candidate_is_not_exported_or_added_to_training() -> None:
    topic = {
        "主题ID": "EXPERIENCE-ONLY-001",
        "知识来源": "方向二经验补充候选",
        "关联标准项": "",
        "审核结论": "通过",
        "是否值得沉淀": "值得沉淀",
        "是否可用": "可用",
        "是否进入训练集": "是",
        "主标题": "手机屏幕异常如何核验？",
        "知识内容": "当前来源未提供可直接套用的明确规则，不能据此作出确定结论。",
        "推荐回复": "请补充证据后再判定。",
    }

    final_rows, feedback_rows, training_rows = (
        workflow_module.finalize_topic_review_rows([topic])
    )

    assert final_rows == []
    assert len(feedback_rows) == 1
    assert training_rows == []


def test_experience_content_drops_bare_numeric_thresholds_without_standard() -> None:
    content = workflow_module._build_experience_review_content(
        {
            "对象/部位": "屏幕坏点",
            "异常现象": "数量待确认",
            "解题方式": "记录坏点数量，不少于3时按异常处理，并观察直径≥5",
        }
    )

    assert "不少于3" not in content
    assert "≥5" not in content
    assert "记录坏点数量" in content


def test_default_workflow_without_catalog_keeps_standard_review_mode(tmp_path: Path) -> None:
    source_path = tmp_path / "source.xlsx"
    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.append(["工单ID", "聊天内容", "产品类型"])
    source_sheet.append(["202608210005", "手机屏幕色斑如何判定？", "手机"])
    source_book.save(source_path)

    result = workflow_module.initial_label_from_workbook(
        source_path=source_path,
        standards_path=None,
        output_dir=tmp_path / "outputs",
        use_mimo=False,
        clustering_mode="rule",
    )

    assert result["standard_references_enabled"] is True


def test_case_analysis_draft_is_rewritten_as_reusable_standard_knowledge() -> None:
    calls: list[str] = []
    standard = StandardCatalogItem(
        standard_id="KB-STD-SCREEN-ASH-001",
        title="平板屏幕进灰判定",
        category_l1="屏幕外观情况",
        category_l2="屏幕进灰",
        knowledge_type="业务沉淀标准",
        standard_path="CZ业务沉淀标准：KB-STD-SCREEN-ASH-001",
        keywords=["平板", "屏幕", "进灰", "漏液"],
        scope="平板电脑",
        response_snippet="白底可见灰尘直径≤1mm且数量≤10颗时，判定为屏幕进灰。",
        status="published",
        version="cz-snapshot-v1",
    )

    class CaseAnalysisDraftMimo:
        config = SimpleNamespace(model="mimo-case-analysis-regression-test")

        def classify_topic_stage(self, _topic):
            calls.append("classify")
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "问题询问可复用的屏幕判定标准。",
                    "value_reason": "标准已覆盖该现象。",
                    "reusable_knowledge": "屏幕进灰与漏液的区分方法。",
                    "confidence": 0.94,
                    "needs_human_review": False,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, *, retry_reason="", **_kwargs):
            calls.append("rewrite" if retry_reason else "transcribe")
            if not retry_reason:
                return MimoLabelResult(
                    candidate={
                        "title": "回收师询问平板屏幕漏液的判定方法，并描述手电筒照射时看到白色异物",
                        "subtitles": [],
                        "content": (
                            "适用情形：回收师询问平板屏幕漏液的判定方法。\n"
                            "核验要点：本案例息屏状态下看到屏幕内部白色颗粒。\n"
                            "处理结论：根据本次会话及图片，建议勾选屏幕进灰。"
                        ),
                        "category_l1": "屏幕外观情况",
                        "category_l2": "屏幕进灰",
                        "layer": "L2",
                        "knowledge_form": "具体判定",
                        "standard_refs": ["KB-STD-SCREEN-ASH-001"],
                        "applicable_scope": "平板电脑",
                        "recommended_reply": (
                            "您好，关于回收师本次看到的白色异物，"
                            "根据本次会话及图片建议勾选屏幕进灰。"
                        ),
                        "confidence": 0.9,
                        "reasoning_summary": "根据本案例图片得出结论。",
                        "needs_human_review": False,
                        "image_evidence_summary": "不依赖图片。",
                        "requires_images": False,
                        "image_usage_instruction": "",
                    },
                    request_audit={},
                    response_audit={},
                )
            assert "案例分析" in retry_reason
            return MimoLabelResult(
                candidate={
                    "title": "平板屏幕进灰与漏液的判定方法",
                    "subtitles": ["白色颗粒与有色漏液的区分"],
                    "content": (
                        "适用范围：平板电脑。\n"
                        "判定标准：白底可见灰尘直径≤1mm且数量≤10颗时，判定为屏幕进灰。\n"
                        "核验方法：在规定显示条件下观察异物颜色、形态和数量。\n"
                        "处理边界：有色块并遮挡显示内容时，应按漏液标准另行判定。"
                    ),
                    "category_l1": "屏幕外观情况",
                    "category_l2": "屏幕进灰",
                    "layer": "L2",
                    "knowledge_form": "具体判定",
                    "standard_refs": ["KB-STD-SCREEN-ASH-001"],
                    "applicable_scope": "平板电脑",
                    "recommended_reply": (
                        "您好，按平板屏幕进灰标准，在规定显示条件下观察异物颜色、"
                        "形态和数量；白色细小颗粒满足对应条件时勾选屏幕进灰，"
                        "有色块遮挡显示内容时再按漏液标准处理。"
                    ),
                    "confidence": 0.91,
                    "reasoning_summary": "依据已生效标准整理可复用判定方法。",
                    "needs_human_review": False,
                    "image_evidence_summary": "不依赖图片。",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(self, _topic, draft, _matches, **_kwargs):
            calls.append("quality_review")
            assert draft["title"] == "平板屏幕进灰与漏液如何判定"
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "草稿是标准知识，不是案例复述。",
                    "standard_consistency": "一致",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.94,
                    "priority_review": False,
                },
                request_audit={},
                response_audit={},
            )

    def retrieve(_topic_id, _rows, _query):
        calls.append("retrieve")
        return [(standard, 0.91)], {
            "source": "headquarters_standard",
            "status": "success",
            "knowledge_version": "cz-snapshot-v1",
        }

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": "平板屏幕里面有白色颗粒，这是不是漏液？",
            "核心问题": "平板屏幕进灰与漏液如何判定",
            "产品类型": "平板电脑",
            "问题意图": "标准判定",
            "对象/部位": "屏幕内部",
            "异常现象": "白色颗粒状异物",
            "解题方式": "按屏幕进灰和漏液标准核验",
            "语义标注依据": "会话询问白色异物是否属于屏幕漏液。",
        }
        for record_id in ("202608100010", "202608100011")
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=CaseAnalysisDraftMimo(),
        clustering_mode="rule",
        use_standard_references=True,
        topic_standard_retriever=retrieve,
        require_standard_match=True,
    )

    assert calls == [
        "classify",
        "retrieve",
        "transcribe",
        "rewrite",
        "quality_review",
    ]
    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["主标题"] == "平板屏幕进灰与漏液如何判定"
    assert "回收师" not in topics[0]["知识内容"]
    assert "本次会话" not in topics[0]["推荐回复"]


def test_topic_content_keeps_only_structured_knowledge_and_separates_reply() -> None:
    separated_body, separated_reply = workflow_module._split_embedded_recommended_reply(
        "判定规则：按当前标准核验。答复建议：请按当前标准处理。"
    )
    assert "答复建议" not in separated_body
    assert separated_reply == "请按当前标准处理。"

    standard = StandardCatalogItem(
        standard_id="KB-STD-SCREEN-ASH-002",
        title="手机屏幕进灰判定",
        category_l1="屏幕外观情况",
        category_l2="屏幕进灰",
        knowledge_type="业务沉淀标准",
        standard_path="CZ业务沉淀标准：KB-STD-SCREEN-ASH-002",
        keywords=["手机", "屏幕", "进灰"],
        scope="手机",
        response_snippet="白底可见灰尘直径≤1mm且数量≤10颗时，判定为屏幕进灰。",
        status="published",
        version="cz-snapshot-v1",
    )

    class MixedFieldTopicMimo:
        config = SimpleNamespace(model="mimo-topic-content-separation-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "主题命中可复用的屏幕判定标准。",
                    "value_reason": "标准已覆盖该现象。",
                    "reusable_knowledge": "手机屏幕进灰判定与处理方法。",
                    "confidence": 0.95,
                    "needs_human_review": False,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, **_kwargs):
            return MimoLabelResult(
                candidate={
                    "title": "手机屏幕进灰如何判定",
                    "subtitles": [],
                    "content": (
                        "### 知识标题：手机屏幕进灰如何判定\n"
                        "判定标准：白底可见灰尘直径≤2mm且数量≤20颗时，判定为屏幕进灰。\n"
                        "核验方法：1. 观察到颗粒直径超过2mm时直接判定为屏幕进灰。\n"
                        "处理边界：颗粒超过2mm时可跳过人工审核。答复建议："
                        "请在白底显示条件下核验异物；满足标准条件时按屏幕进灰处理，"
                        "证据不足请补充照片后转人工审核。"
                    ),
                    "category_l1": "屏幕外观情况",
                    "category_l2": "屏幕进灰",
                    "layer": "L2",
                    "knowledge_form": "具体判定",
                    "standard_refs": ["KB-STD-SCREEN-ASH-002"],
                    "applicable_scope": "手机",
                    "recommended_reply": "",
                    "confidence": 0.92,
                    "reasoning_summary": "依据已生效标准整理。",
                    "needs_human_review": False,
                    "image_evidence_summary": "不依赖图片。",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(self, _topic, draft, _matches, **_kwargs):
            assert draft["title"] == "手机屏幕进灰如何判定"
            assert draft["content"].startswith("1. ")
            assert "判定规则：" not in draft["content"]
            assert "处理步骤：" not in draft["content"]
            assert "例外与边界：" not in draft["content"]
            assert "标题：" not in draft["content"]
            assert "知识标题" not in draft["content"]
            assert "推荐回复" not in draft["content"]
            assert "答复建议" not in draft["content"]
            assert "≤2mm" not in draft["content"]
            assert "超过2mm" not in draft["content"]
            assert "≤1mm且数量≤10颗" in draft["content"]
            assert "≤1mm且数量≤10颗" in draft["content"]
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "知识正文和推荐回复字段职责清晰。",
                    "standard_consistency": "一致",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.95,
                    "priority_review": False,
                },
                request_audit={},
                response_audit={},
            )

    def retrieve(_topic_id, _rows, _query):
        return [(standard, 0.95)], {
            "source": "headquarters_standard",
            "status": "success",
            "knowledge_version": "cz-snapshot-v1",
        }

    topics, _mapping, gaps, pending = build_topic_review_rows(
        [
            {
                "数据ID": "SCREEN-STRUCTURED-001",
                "工单ID": "SCREEN-STRUCTURED-001",
                "聊天内容": "手机屏幕里面有白色颗粒，应该怎么判定？",
                "核心问题": "手机屏幕进灰如何判定",
                "产品类型": "手机",
                "问题意图": "标准判定",
                "对象/部位": "屏幕内部",
                "异常现象": "白色颗粒状异物",
                "解题方式": "按屏幕进灰标准核验",
                "语义标注依据": "会话询问手机屏幕白色颗粒的判定。",
            }
        ],
        mimo_client=MixedFieldTopicMimo(),
        clustering_mode="rule",
        use_standard_references=True,
        topic_standard_retriever=retrieve,
        require_standard_match=True,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["知识内容"].startswith("1. ")
    assert "判定规则：" not in topics[0]["知识内容"]
    assert "处理步骤：" not in topics[0]["知识内容"]
    assert "例外与边界：" not in topics[0]["知识内容"]
    assert "标题：" not in topics[0]["知识内容"]
    assert "知识标题" not in topics[0]["知识内容"]
    assert "推荐回复" not in topics[0]["知识内容"]
    assert "答复建议" not in topics[0]["知识内容"]
    assert "≤2mm" not in topics[0]["知识内容"]
    assert "超过2mm" not in topics[0]["知识内容"]
    assert "≤1mm且数量≤10颗" in topics[0]["知识内容"]
    assert "≤1mm且数量≤10颗" in topics[0]["知识内容"]
    assert topics[0]["主题无来源内容"] == ""
    assert topics[0]["模型初标结论"] == "通过"
    assert topics[0]["推荐回复"]


def test_topic_model_call_budget_keeps_remaining_topics_for_manual_review() -> None:
    class BudgetedTopicMimo:
        config = SimpleNamespace(model="mimo-topic-budget-test")

        def __init__(self) -> None:
            self.classify_calls = 0

        def classify_topic_stage(self, _topic):
            self.classify_calls += 1
            return MimoLabelResult(
                candidate={
                    "topic_stage": "案例解析",
                    "knowledge_value": "不值得沉淀",
                    "stage_reason": "当前主题需要人工判断是否可复用。",
                    "value_reason": "暂不自动沉淀。",
                    "reusable_knowledge": "",
                    "confidence": 0.8,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"{product}当前案例怎么处理",
            "核心问题": f"{product}当前案例怎么处理",
            "产品类型": product,
            "问题意图": "案例判定",
            "对象/部位": "待确认",
            "异常现象": "待确认",
            "解题方式": "人工复核",
            "语义标注依据": "当前案例证据不足。",
        }
        for record_id, product in (("A", "手机"), ("B", "平板"))
    ]
    mimo = BudgetedTopicMimo()

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        mimo_client=mimo,
        clustering_mode="rule",
        use_standard_references=False,
        topic_model_call_limit=1,
    )

    assert len(topics) == 2
    assert mimo.classify_calls == 1
    assert {topic["主题分类状态"] for topic in topics} == {
        "topic_stage_classified_model",
        "topic_stage_skipped_model_budget",
    }
    skipped = next(
        topic
        for topic in topics
        if topic["主题分类状态"] == "topic_stage_skipped_model_budget"
    )
    assert skipped["主题分类重点复核"] == "是"


def test_topic_model_call_budget_uses_high_configured_limit(monkeypatch) -> None:
    monkeypatch.setenv("ANSWER_HUB_TOPIC_MODEL_CALL_LIMIT", "1000")

    class ConfiguredBudgetTopicMimo:
        config = SimpleNamespace(model="mimo-topic-configured-budget-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "案例解析",
                    "knowledge_value": "待确认",
                    "stage_reason": "需要人工确认是否形成可复用知识。",
                    "value_reason": "需要人工确认。",
                    "reusable_knowledge": "",
                    "confidence": 0.8,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"{product}当前案例怎么处理",
            "核心问题": f"{product}当前案例怎么处理",
            "产品类型": product,
            "问题意图": "案例判定",
            "对象/部位": "待确认",
            "异常现象": "待确认",
            "解题方式": "人工复核",
            "语义标注依据": "当前案例证据不足。",
        }
        for record_id, product in (("A", "手机"), ("B", "平板"))
    ]
    clustering_meta: dict[str, object] = {}

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        mimo_client=ConfiguredBudgetTopicMimo(),
        clustering_mode="rule",
        use_standard_references=False,
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 2
    assert clustering_meta["topic_model_call_limit"] == 1000
    assert clustering_meta["topic_model_calls"] == 2
    assert clustering_meta["topic_model_budget_skipped"] == 0
    assert all(
        topic["主题分类状态"] != "topic_stage_skipped_model_budget"
        for topic in topics
    )


def test_generic_topic_draft_is_rewritten_from_case_evidence_before_initial_review() -> None:
    calls: list[str] = []

    class GenericDraftMimo:
        config = SimpleNamespace(model="mimo-topic-generic-draft-test")

        def classify_topic_stage(self, _topic):
            calls.append("classify")
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检流程",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "主题属于功能核验流程。",
                    "value_reason": "模型认为可以沉淀。",
                    "reusable_knowledge": "功能核验流程。",
                    "confidence": 0.9,
                    "needs_human_review": False,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, retry_reason="", **_kwargs):
            if retry_reason:
                calls.append("rewrite")
                assert "历史实际回复" in retry_reason
                return MimoLabelResult(
                    candidate={
                        "title": "传感器功能如何核验",
                        "subtitles": [],
                        "content": (
                            "案例结论：AirPods 一代在当前质检流程中不要求检查“查找”功能。\n"
                            "处理方式：确认设备代际为 AirPods 一代后，可直接进入后续质检步骤。\n"
                            "适用边界：设备型号或代际无法确认时，不沿用该结论，应先补充设备信息。"
                        ),
                        "category_l1": "功能问题",
                        "category_l2": "传感器功能",
                        "layer": "L2",
                        "knowledge_form": "流程方法",
                        "standard_refs": [],
                            "applicable_scope": "耳机-AirPods 一代",
                            "applicable_brands": ["Apple"],
                            "applicable_models": ["AirPods 一代"],
                        "recommended_reply": (
                            "AirPods 一代在当前流程中不要求检查“查找”功能；"
                            "确认代际后可直接继续后续质检。"
                        ),
                        "confidence": 0.9,
                        "reasoning_summary": "历史实际回复明确说明了适用代际和处理方式。",
                        "needs_human_review": True,
                        "image_evidence_summary": "",
                        "requires_images": False,
                        "image_usage_instruction": "",
                    },
                    request_audit={},
                    response_audit={},
                )
            calls.append("transcribe")
            return MimoLabelResult(
                candidate={
                    "title": "传感器功能如何核验",
                    "subtitles": [],
                    "content": (
                        "功能核验流程：\n"
                        "1. 明确待核验功能、测试条件和所用配件。\n"
                        "2. 排除电量、网络、权限、保护壳等外部影响。\n"
                        "3. 使用一致的测试条件复测，并记录画面、提示、声音或响应结果。\n"
                        "4. 结果不稳定或无法复现时，补充测试证据后再判定。"
                    ),
                    "category_l1": "功能问题",
                    "category_l2": "传感器功能",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": [],
                    "applicable_scope": "耳机-通用",
                    "recommended_reply": "请按统一条件复测后确认。",
                    "confidence": 0.9,
                    "reasoning_summary": "功能问题需要测试核验。",
                    "needs_human_review": False,
                    "image_evidence_summary": "",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(self, *_args, **_kwargs):
            calls.append("quality_review")
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "正文包含来源案例的代际、检测要求和处理边界。",
                    "standard_consistency": "无可信标准",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.9,
                    "priority_review": True,
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "AIRPODS-FINDMY-001",
            "工单ID": "AIRPODS-FINDMY-001",
            "聊天内容": "AirPods 一代的查找功能需要检测吗？",
            "历史实际回复": (
                "对于 AirPods 一代，标准质检流程中不要求检查“查找”功能，"
                "可以直接进行后续质检步骤。"
            ),
            "核心问题": "AirPods 一代查找功能是否需要检测",
            "产品类型": "耳机",
            "问题意图": "检测核验",
            "对象/部位": "传感器功能",
            "异常现象": "查找功能检测要求待确认",
            "解题方式": "功能测试与结果核对",
            "语义标注依据": "历史实际回复明确说明 AirPods 一代不要求检查查找功能。",
        }
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        mimo_client=GenericDraftMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert calls == ["classify", "transcribe", "rewrite", "quality_review"]
    assert topics[0]["主题沉淀价值"] == "值得沉淀"
    assert topics[0]["主题转写状态"] == "topic_model_rewritten_for_evidence"
    assert topics[0]["模型初标结论"] == "需修改"
    assert topics[0]["模型初标错误类型"] in {"标题不准", "话术不合适"}
    assert "AirPods 一代" in topics[0]["知识内容"]
    assert "不要求检查“查找”功能" in topics[0]["知识内容"]
    assert "明确待核验功能" not in topics[0]["知识内容"]
    assert topics[0]["适用范围"] == "耳机/耳麦"
    assert topics[0]["适用品牌"] == ""
    assert topics[0]["适用机型"] == "AirPods 一代"


def test_generic_topic_draft_stays_pending_when_evidence_rewrite_is_still_generic() -> None:
    calls: list[str] = []

    class StillGenericMimo:
        config = SimpleNamespace(model="mimo-topic-generic-retry-test")

        def classify_topic_stage(self, _topic):
            calls.append("classify")
            return MimoLabelResult(
                candidate={
                    "topic_stage": "案例解析",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "模型尝试沉淀当前案例。",
                    "value_reason": "模型认为存在可复用结论。",
                    "reusable_knowledge": "通用核验流程。",
                    "confidence": 0.9,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, retry_reason="", **_kwargs):
            calls.append("rewrite" if retry_reason else "transcribe")
            return MimoLabelResult(
                candidate={
                    "title": "设备来源如何核验",
                    "subtitles": [],
                    "content": (
                        "适用主题：其他待确认 / 设备来源\n"
                        "核验流程：\n"
                        "1. 先确认需要检查的设备、现象和问题。\n"
                        "2. 提供截图、照片、视频或查询结果作为依据。\n"
                        "3. 参考已有案例梳理适用范围、边界与例外。\n"
                        "4. 暂时不能确认时，完善资料后再处理。"
                    ),
                    "category_l1": "其他待确认",
                    "category_l2": "设备来源",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": [],
                    "applicable_scope": "手机-通用",
                    "recommended_reply": "请补充信息后再处理。",
                    "confidence": 0.8,
                    "reasoning_summary": "当前信息不足。",
                    "needs_human_review": True,
                    "image_evidence_summary": "",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(self, *_args, **_kwargs):
            raise AssertionError("重写后仍为空泛模板时不应进入内容质量初标")

    rows = [
        {
            "数据ID": "SOURCE-PENDING-001",
            "工单ID": "SOURCE-PENDING-001",
            "聊天内容": "这个设备来源怎么确认？",
            "历史实际回复": "当前信息不足，需进一步确认。",
            "核心问题": "设备来源如何确认",
            "产品类型": "手机",
            "问题意图": "信息查询",
            "对象/部位": "设备来源",
            "异常现象": "待确认",
            "解题方式": "先查询平台记录，再核对序列号",
            "语义标注依据": "当前没有具体来源结论。",
        }
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        mimo_client=StillGenericMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert calls == ["classify", "transcribe", "rewrite"]
    assert topics[0]["主题沉淀价值"] == "待确认"
    assert topics[0]["主题转写状态"] == "skipped_generic_draft"
    assert topics[0]["模型初标结论"] == "未执行"


def test_topic_quality_review_does_not_reclassify_knowledge_value() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test-key",
            base_url="https://example.com/v1",
            model="mimo-test",
        )
    )
    captured_payloads: list[dict[str, object]] = []

    def fake_post(payload: dict[str, object]) -> dict[str, object]:
        captured_payloads.append(payload)
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "decision": "通过",
                                "knowledge_value": "值得沉淀",
                                "error_type": "",
                                "reason": "草稿内容与主题证据一致。",
                                "standard_consistency": "无可信标准",
                                "evidence_sufficiency": "充分",
                                "content_consistency": "一致",
                                "image_necessity": "不需要",
                                "title_quality": "清晰",
                                "confidence": 0.9,
                                "priority_review": False,
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    client._post = fake_post  # type: ignore[method-assign]
    client.review_topic(
        {
            "topic_id": "TOP-QUALITY-001",
            "topic_stage": "质检流程",
            "knowledge_value": "值得沉淀",
            "evidence_summary": "两个案例提供了一致的检查步骤。",
        },
        {
            "主标题": "设备信息核对流程",
            "知识内容": "先读取设备信息，再使用检测工具交叉核对。",
        },
        [],
        use_standard_references=False,
    )

    prompt = captured_payloads[0]["messages"][1]["content"][0]["text"]  # type: ignore[index]
    assert "沉淀价值已经在转写前完成" in prompt
    assert "不得重新判断是否值得沉淀" in prompt
    assert "knowledge_value 仅作兼容字段，原样返回主题输入中的 knowledge_value" in prompt
    assert "必须标注知识点是否值得沉淀" not in prompt


def test_candidate_knowledge_export_only_contains_transcribed_worthy_topics(
    tmp_path: Path,
) -> None:
    output = tmp_path / "candidate_knowledge.xlsx"
    write_topic_candidate_knowledge_workbook(
        [
            {
                "主题ID": "TOP-WORTHY",
                "知识ID": "TOP-WORTHY",
                "主题沉淀价值": "值得沉淀",
                "主题转写状态": "topic_model_labeled",
                "模型调用状态": "model_success",
                "模型输出校验状态": "passed",
                "模型质量状态": "passed",
                "知识草稿状态": "ready_for_human_review",
                "模型初标结论": "通过",
                "模型初标错误类型": "",
                "审核结论": "修改后通过",
                "主标题": "手机设备信息应如何核对？",
                "知识内容": "1. 先读取设备信息。\n2. 再使用检测工具交叉核对。",
                "推荐回复": "请先读取设备信息，再使用检测工具交叉核对。",
                "知识分类": "质检流程",
                "适用范围": "手机-通用",
                "关键词": "设备信息；核对",
            },
            {
                "主题ID": "TOP-UNWORTHY",
                "知识ID": "TOP-UNWORTHY",
                "主题沉淀价值": "不值得沉淀",
                "主题转写状态": "skipped_not_worthy",
                "主标题": "当前图片中的位置是否属于磕碰",
                "知识内容": "主题未进入知识转写。",
            },
            {
                "主题ID": "TOP-GENERIC",
                "知识ID": "TOP-GENERIC",
                "主题沉淀价值": "值得沉淀",
                "主题转写状态": "skipped_generic_draft",
                "主标题": "屏幕显示异常如何通过图片核验",
                "知识内容": "适用主题：外观问题 / 屏幕及正面外观\n核验流程：\n1. 先确认异常位置。\n2. 补充图片后再判定。",
            },
        ],
        output,
        use_standard_references=False,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    sheet = workbook["候选知识"]
    values = list(sheet.iter_rows(values_only=True))
    workbook.close()

    assert len(values) == 2
    assert values[1][0] == "TOP-WORTHY"


def test_standard_second_match_rejects_judgment_target_mismatch() -> None:
    standard = StandardCatalogItem(
        standard_id="STD-NB-SSD-QUERY",
        title="固态硬盘查看方法",
        category_l1="硬件信息",
        category_l2="固态硬盘",
        knowledge_type="质检标准",
        standard_path="【笔记本】-【硬件信息】-【固态硬盘】",
        keywords=["硬盘", "查看"],
        scope="笔记本-通用",
        response_snippet="打开设备信息查看固态硬盘型号。",
        status="published",
        version="v1",
    )
    reasons = workflow_module._standard_match_rejection_reasons(
        {
            "产品类型": "笔记本",
            "对象/部位": "硬盘",
            "异常现象": "品牌部件",
            "问题意图": "品牌判定",
            "核心问题": "硬盘是否为品牌部件",
        },
        standard,
    )

    assert "judgment_target_mismatch" in reasons


def test_empty_title_is_rebuilt_from_structured_fields() -> None:
    title = workflow_module._rebuild_title_from_structured_fields(
        {
            "产品类型": "平板电脑",
            "对象/部位": "电池健康度",
            "异常现象": "读取优先级",
            "问题意图": "查询方法",
        }
    )

    assert title == "平板电脑电池健康度应按什么优先级读取？"


def test_title_rebuild_uses_intent_specific_question() -> None:
    assert workflow_module._rebuild_title_from_structured_fields(
        {
            "产品类型": "平板电脑",
            "对象/部位": "序列号",
            "异常现象": "查看位置",
            "问题意图": "信息查询",
        }
    ) == "平板电脑序列号应在哪里查看？"
    assert workflow_module._rebuild_title_from_structured_fields(
        {
            "产品类型": "平板电脑",
            "对象/部位": "屏幕",
            "异常现象": "进灰与漏液边界",
            "问题意图": "边界判定",
        }
    ) == "平板电脑屏幕进灰与漏液应如何区分？"


def test_measurement_gate_blocks_visual_numeric_conclusion_without_measurement() -> None:
    gate = workflow_module._topic_image_measurement_gate(
        [
            {
                "产品类型": "手机",
                "对象/部位": "屏幕",
                "异常现象": "磕点",
                "聊天内容": "图片中看到一个磕点，需要判断是否超过1mm。",
                "图片链接": "https://example.com/case.jpg",
            }
        ],
        {
            "content": "1. 磕点直径超过1mm时判定为异常。",
            "recommended_reply": "该磕点超过1mm，应判定为异常。",
        },
    )

    assert gate["measurement_required"] is True
    assert gate["measurement_available"] is False
    assert gate["visual_conclusion_allowed"] is False
    assert gate["status"] == "required_missing"


def test_battery_health_percentage_does_not_trigger_image_measurement_gate() -> None:
    gate = workflow_module._topic_image_measurement_gate(
        [
            {
                "产品类型": "平板电脑",
                "对象/部位": "电池健康度",
                "异常现象": "健康度为84%",
                "聊天内容": (
                    "截图显示电池健康度为84%，有人认为健康度≤85%需要处理，"
                    "需要确认应读取哪个结果。"
                ),
                "图片链接": "https://example.com/tablet-battery.jpg",
            }
        ],
        {
            "content": "1. 电池健康度低于85%时需要人工确认。",
            "recommended_reply": "电池健康度为84%，请补充量尺照片。",
        },
    )

    assert gate["measurement_required"] is False
    assert gate["measurement_available"] is False
    assert gate["visual_conclusion_allowed"] is True
    assert gate["status"] == "not_needed"


def test_model_draft_cannot_create_camera_gap_measurement_requirement() -> None:
    gate = workflow_module._topic_image_measurement_gate(
        [
            {
                "产品类型": "平板电脑",
                "对象/部位": "后摄镜片",
                "异常现象": "镜片边缘有缝隙",
                "聊天内容": "图片中后摄镜片边缘有缝隙，需要确认如何处理。",
                "图片链接": "https://example.com/tablet-camera-gap.jpg",
            }
        ],
        {
            "content": "1. 缝隙超过0.5mm时按镜片更换处理。",
            "recommended_reply": "该缝隙超过0.5mm，应按镜片更换处理。",
        },
    )

    assert gate["measurement_required"] is False
    assert gate["status"] == "not_needed"


def test_model_failure_and_quality_failure_are_distinct() -> None:
    assert workflow_module._draft_status_for_model_result(
        model_error="MiMo 响应超时",
        quality_issues=[],
        has_standard=True,
    ) == ("model_failed", "standard_rule_fallback")
    assert workflow_module._draft_status_for_model_result(
        model_error="",
        quality_issues=["标题为空"],
        has_standard=True,
    ) == ("model_success", "blocked")


def test_model_validation_failure_does_not_generate_rule_fallback_content() -> None:
    class ValidationFailingMimo:
        config = SimpleNamespace(model="mimo-validation-failure-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检流程",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "来源询问电池健康度读取方式。",
                    "value_reason": "该问题可以沉淀为可复用的读取流程。",
                    "reusable_knowledge": "按来源中已确认的顺序读取电池健康度。",
                    "confidence": 0.92,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, *, use_standard_references):
            assert use_standard_references is False
            raise MimoError(
                "MiMo 主题 JSON 校验失败（已重试一次）："
                "MiMo 输出缺少或为空：title"
            )

    rows = [
        {
            "数据ID": f"TABLET-BATTERY-{index}",
            "工单ID": f"TABLET-BATTERY-{index}",
            "聊天内容": "平板电池健康度应以哪个页面的读取结果为准？",
            "核心问题": "平板电池健康度应以哪个读取结果为准",
            "历史实际回复": (
                "先进入设备设置页面读取本机电池健康度，"
                "再核对检测工具结果；本机不显示时选择无法检测。"
            ),
            "判定结论": "本机不显示电池健康度时选择无法检测。",
            "产品类型": "平板电脑",
            "一级分类": "电池",
            "二级分类": "电池健康度",
            "问题意图": "读取优先级",
            "对象/部位": "电池健康度",
            "异常现象": "读取结果不一致",
            "解题方式": "按读取来源优先级核验",
        }
        for index in (1, 2)
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=ValidationFailingMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    topic = topics[0]
    assert topic["主题转写状态"] == "topic_model_validation_failed"
    assert topic["模型调用状态"] == "model_success"
    assert topic["模型输出校验状态"] == "failed"
    assert topic["模型质量状态"] == "failed"
    assert topic["知识草稿状态"] == "blocked"
    assert topic["知识内容"]
    assert "电池健康度" in topic["知识内容"]
    assert "无法检测" in topic["知识内容"]
    assert topic["推荐回复"] == ""
    assert topic["模型初标结论"] == "未执行"


def test_failed_topic_title_uses_current_atomic_structure_without_cross_topic_text() -> None:
    rows = [
        {
            "数据ID": "TABLET-CAMERA-DUST-001",
            "工单ID": "TABLET-CAMERA-DUST-001",
            "聊天内容": "后摄区域有灰尘，想确认是否属于屏幕进灰。",
            "核心问题": "电池循环次数太少，电池容量又低，要算异常吗",
            "产品类型": "平板电脑",
            "一级分类": "外观问题",
            "二级分类": "后摄区域",
            "问题意图": "归属判定",
            "对象/部位": "后摄区域灰尘",
            "异常现象": "是否属于屏幕进灰",
            "解题方式": "确认灰尘所在部位后判断归属",
        }
    ]
    topic = workflow_module._failed_topic_transcription_row(
        "TOP-TABLET-CAMERA-DUST-001",
        ("平板电脑", "后摄区域灰尘", "是否属于屏幕进灰"),
        rows,
        {
            "topic_stage": "质检流程",
            "knowledge_value": "值得沉淀",
            "confidence": 0.8,
        },
        provider="mimo",
        model_name="mimo-test",
        prompt_version="test",
        model_run_id="run-test",
        transcription_status="topic_model_validation_failed",
        model_call_status="model_success",
        error="副标题不合格",
        matches=[],
        use_standard_references=False,
    )

    assert topic["主标题"] == "平板电脑后摄区域灰尘是否属于屏幕进灰？"
    assert "电池循环次数" not in topic["主标题"]
    assert "是否属于屏幕进灰应如何判定" not in topic["主标题"]
    assert "判定定" not in topic["主标题"]


def test_failed_topic_title_uses_current_topic_category_when_object_is_missing() -> None:
    rows = [
        {
            "数据ID": "TABLET-SN-001",
            "工单ID": "TABLET-SN-001",
            "聊天内容": "同一会话还问了电池问题，但当前主题是序列号查看位置。",
            "核心问题": "电池健康度84%是否异常",
            "产品类型": "平板电脑",
            "一级分类": "信息查询",
            "二级分类": "序列号",
            "问题意图": "位置查询",
            "对象/部位": "",
            "异常现象": "",
            "解题方式": "查看系统信息页面",
        }
    ]
    topic = workflow_module._failed_topic_transcription_row(
        "TOP-TABLET-SN-001",
        workflow_module._topic_group_key(rows[0]),
        rows,
        {
            "topic_stage": "质检流程",
            "knowledge_value": "值得沉淀",
            "confidence": 0.8,
        },
        provider="mimo",
        model_name="mimo-test",
        prompt_version="test",
        model_run_id="run-test",
        transcription_status="topic_model_validation_failed",
        model_call_status="model_success",
        error="content_type 缺失",
        matches=[],
        use_standard_references=False,
    )

    assert "序列号" in topic["主标题"]
    assert "电池健康度" not in topic["主标题"]


def test_natural_title_removes_duplicated_judgment_suffix() -> None:
    assert (
        workflow_module._as_natural_question_title("平板屏幕漏液如何判定定")
        == "平板屏幕漏液如何判定"
    )


def test_recommended_reply_is_generated_from_final_content_only() -> None:
    reply = workflow_module._recommended_reply_from_final_content(
        "1. 先读取本机电池健康度。\n2. 无法获取时选择无法检测。",
        evidence_status="available",
    )

    assert "本机电池健康度" in reply
    assert "回收师" not in reply
    assert "关于“" not in reply


def test_model_standard_mapping_error_blocks_formal_export() -> None:
    issues = workflow_module._topic_candidate_export_gate_issues(
        {
            "审核结论": "修改后通过",
            "主标题": "笔记本硬盘是否为品牌部件？",
            "知识内容": "1. 先查询硬盘型号。",
            "推荐回复": "先查询硬盘型号。",
            "关联标准项": "STD-NB-SSD-QUERY",
            "模型初标错误类型": "标准项映射错",
        },
        use_standard_references=True,
    )

    assert "模型初标标准映射错误" in issues


def test_standard_candidate_export_requires_explicit_review_approval_and_real_standard(
    tmp_path: Path,
) -> None:
    output = tmp_path / "candidate_knowledge.xlsx"
    write_topic_candidate_knowledge_workbook(
        [
            {
                "主题ID": "TOP-APPROVED",
                "知识ID": "TOP-APPROVED",
                "主题沉淀价值": "值得沉淀",
                "主题转写状态": "topic_model_labeled",
                "模型调用状态": "model_success",
                "模型输出校验状态": "passed",
                "模型质量状态": "passed",
                "知识草稿状态": "ready_for_human_review",
                "模型初标结论": "通过",
                "模型初标错误类型": "",
                "审核结论": "修改后通过",
                "主标题": "平板电池健康度应按什么优先级读取？",
                "副标题": "平板电池健康度怎么核验？",
                "知识内容": "1. 先读取本机显示值。\n2. 无法获取时选择无法检测。",
                "推荐回复": "先读取本机显示值，无法获取时选择无法检测。",
                "关联标准项": "STD-TABLET-BATTERY",
                "主题无来源内容": "",
                "主题图片必要性": "辅助图例",
                "图例": "",
            },
            {
                "主题ID": "TOP-REJECTED",
                "知识ID": "TOP-REJECTED",
                "主题沉淀价值": "值得沉淀",
                "主题转写状态": "topic_model_labeled",
                "模型调用状态": "model_success",
                "模型输出校验状态": "passed",
                "模型质量状态": "passed",
                "知识草稿状态": "ready_for_human_review",
                "模型初标结论": "通过",
                "审核结论": "驳回",
                "主标题": "平板电池健康度应按什么优先级读取？",
                "知识内容": "1. 先读取本机显示值。",
                "推荐回复": "先读取本机显示值。",
                "关联标准项": "STD-TABLET-BATTERY",
            },
            {
                "主题ID": "TOP-NO-STANDARD",
                "知识ID": "TOP-NO-STANDARD",
                "主题沉淀价值": "值得沉淀",
                "主题转写状态": "topic_model_labeled",
                "模型调用状态": "model_success",
                "模型输出校验状态": "passed",
                "模型质量状态": "passed",
                "知识草稿状态": "ready_for_human_review",
                "模型初标结论": "通过",
                "审核结论": "修改后通过",
                "主标题": "平板电池健康度应按什么优先级读取？",
                "知识内容": "1. 先读取本机显示值。",
                "推荐回复": "先读取本机显示值。",
                "关联标准项": "",
            },
        ],
        output,
        use_standard_references=True,
    )
    workbook = load_workbook(output, read_only=True, data_only=True)
    values = list(workbook["候选知识"].iter_rows(values_only=True))
    workbook.close()

    assert len(values) == 2
    assert values[1][0] == "平板电池健康度应按什么优先级读取？"


def test_case_only_candidate_export_blocks_unapproved_failed_topics(
    tmp_path: Path,
) -> None:
    output = tmp_path / "candidate_knowledge.xlsx"
    write_topic_candidate_knowledge_workbook(
        [
            {
                "主题ID": "TOP-APPROVED",
                "知识ID": "TOP-APPROVED",
                "主题沉淀价值": "值得沉淀",
                "主题转写状态": "topic_model_labeled",
                "模型调用状态": "model_success",
                "模型输出校验状态": "passed",
                "模型质量状态": "passed",
                "知识草稿状态": "ready_for_human_review",
                "模型初标结论": "通过",
                "模型初标错误类型": "",
                "审核结论": "修改后通过",
                "主标题": "平板电池健康度应如何读取？",
                "知识内容": (
                    "1. 先读取设备本机显示的电池健康度。\n"
                    "2. 本机无法读取时，再按来源中已明确的其他方式核验。"
                ),
                "推荐回复": (
                    "先读取设备本机显示的电池健康度；"
                    "本机无法读取时，再按已明确的其他方式核验。"
                ),
                "主题无来源内容": "",
            },
            {
                "主题ID": "TOP-FAILED",
                "知识ID": "TOP-FAILED",
                "主题沉淀价值": "值得沉淀",
                "主题转写状态": "topic_model_validation_failed",
                "模型调用状态": "model_success",
                "模型输出校验状态": "failed",
                "模型质量状态": "failed",
                "知识草稿状态": "blocked",
                "模型初标结论": "未执行",
                "主标题": "平板后摄镜片缝隙应如何判定？",
                "知识内容": "",
                "推荐回复": "",
            },
        ],
        output,
        use_standard_references=False,
    )
    workbook = load_workbook(output, read_only=True, data_only=True)
    values = list(workbook["候选知识"].iter_rows(values_only=True))
    workbook.close()

    assert len(values) == 2
    assert values[1][0] == "TOP-APPROVED"


def test_standard_content_keeps_rule_points_and_subtitle_is_single_question() -> None:
    standard = StandardCatalogItem(
        standard_id="STD-TABLET-001",
        title="电池健康度",
        category_l1="电池",
        category_l2="电池健康度",
        knowledge_type="质检标准",
        standard_path="【平板电脑】-【电池】-【电池健康度】",
        keywords=["平板", "电池健康度"],
        scope="平板电脑-通用",
        response_snippet=(
            "标准定义：1. 先读取本机显示值。\n"
            "2. 本机无入口时使用验机工具。\n"
            "3. 仍无法获取时选择无法检测。\n"
            "4. 不得用容量字段代替健康度。\n"
            "检测方法：记录实际显示结果。"
        ),
        status="published",
        version="v1",
    )
    content = workflow_module._build_compact_standard_content(
        standard,
        workflow_module.CONTENT_TYPE_VERIFICATION,
    )
    subtitles = workflow_module._finalize_topic_subtitles(
        ["怎么核验？", "有哪些处理条件？"],
        "平板电池健康度应按什么优先级读取？",
        content,
        workflow_module.CONTENT_TYPE_VERIFICATION,
    )

    assert content.count("\n") >= 4
    assert subtitles.count("\n") == 0


def test_finalized_subtitle_drops_a_natural_question_from_another_topic() -> None:
    subtitle = workflow_module._finalize_topic_subtitles(
        ["平板电池健康度怎么查看？"],
        "平板后摄镜片缝隙应如何核验？",
        "1. 检查后摄镜片与保护圈之间是否存在异常缝隙。",
        workflow_module.CONTENT_TYPE_VERIFICATION,
    )

    assert "电池健康度" not in subtitle
    assert "后摄" in subtitle or "镜片" in subtitle


def test_topic_display_questions_validation_requires_short_questions() -> None:
    questions = _validate_topic_display_questions(
        {
            "questions": [
                {"theme_id": "C001", "question": "防水标变红怎么判?"},
                {"theme_id": "C002", "question": "电池健康度读不出来怎么办？"},
            ]
        },
        {"C001", "C002"},
    )

    assert questions == [
        {"theme_id": "C001", "question": "防水标变红怎么判？"},
        {"theme_id": "C002", "question": "电池健康度读不出来怎么办？"},
    ]


def test_topic_display_questions_validation_rejects_missing_theme() -> None:
    with pytest.raises(MimoError, match="缺少 theme_id"):
        _validate_topic_display_questions(
            {
                "questions": [
                    {"theme_id": "C001", "question": "防水标变红怎么判？"}
                ]
            },
            {"C001", "C002"},
        )


def test_topic_display_questions_validation_rejects_two_questions() -> None:
    with pytest.raises(MimoError, match="只能输出一个问句"):
        _validate_topic_display_questions(
            {
                "questions": [
                    {
                        "theme_id": "C001",
                        "question": "存储容量怎么选？主板内存不符是什么意思？",
                    }
                ]
            },
            {"C001"},
        )


def test_rewrite_topic_display_questions_uses_dedicated_prompt() -> None:
    client = MimoClient(
        MimoConfig(
            api_key="test-key",
            base_url="https://example.com/v1",
            model="mimo-test",
        )
    )
    captured_payloads: list[dict[str, object]] = []

    def fake_post(payload: dict[str, object]) -> dict[str, object]:
        captured_payloads.append(payload)
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "questions": [
                                    {
                                        "theme_id": "C001",
                                        "question": "摄像头里面有毛发怎么判？",
                                    }
                                ]
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    client._post = fake_post  # type: ignore[method-assign]
    result = client.rewrite_topic_display_questions(
        [
            {
                "theme_id": "C001",
                "normalized_issues": ["摄像头内部存在毛发异物，判定是否影响质检"],
            }
        ]
    )

    assert result.candidate["questions"][0]["question"] == "摄像头里面有毛发怎么判？"
    assert (
        result.request_audit["prompt_version"]
        == TOPIC_DISPLAY_QUESTION_PROMPT_VERSION
    )
    user_content = captured_payloads[0]["messages"][1]["content"][0]["text"]  # type: ignore[index]
    assert "防水标变红怎么判？" in user_content


def test_topic_signal_uses_conversation_over_legacy_question_and_categories() -> None:
    class ConversationFirstMimo(_FakeMimo):
        def analyze_topic_signal(self, _source, _matches, _images):
            return MimoLabelResult(
                candidate={
                    "intent": "信息查询",
                    "subject": "设备机型",
                    "phenomenon": "机型查询",
                    "resolution_mode": "信息查询与实物核对",
                    "category_l1": "基本情况",
                    "category_l2": "机型",
                    "topic_tags": ["意图:信息查询", "对象:设备机型", "现象:机型查询", "处理:信息查询与实物核对"],
                    "standard_refs": [],
                    "requires_images": False,
                    "image_evidence_summary": "无需依赖图片。",
                    "reasoning_summary": "完整会话在询问设备机型查询方法。",
                    "confidence": 0.9,
                    "needs_human_review": True,
                },
                request_audit={"topic_signal": "conversation-first"},
                response_audit={"choices": []},
            )

    rows = _source_rows()
    rows[0].update(
        {
            "聊天内容": "这台设备的机型怎么查？",
            "核心问题": "手机屏幕色斑如何判定",
            "一级分类": "显示问题",
            "二级分类": "色斑",
        }
    )
    features, _ = generate_phone_candidate_rows(
        rows,
        _standards(),
        mimo_client=ConversationFirstMimo(),
        image_downloader=_ReadyImageDownloader(),
    )

    feature = features[0]
    assert feature["问题意图"] == "信息查询"
    assert feature["对象/部位"] == "设备机型"
    assert feature["模型主题一级分类"] == "基本情况"
    assert feature["核心问题"] == "手机屏幕色斑如何判定"


def test_semantic_clustering_merges_similar_rows_and_keeps_singleton_topic() -> None:
    features, _ = generate_phone_candidate_rows(
        _source_rows(),
        _standards(),
        use_mimo=False,
        image_downloader=_ReadyImageDownloader(),
    )
    first = features[0]
    second = dict(first)
    second["数据ID"] = "PHONE-002"
    second["工单ID"] = "PHONE-002"
    second["原始工单ID"] = "PHONE-002"
    third = dict(first)
    third["数据ID"] = "PHONE-003"
    third["工单ID"] = "PHONE-003"
    third["原始工单ID"] = "PHONE-003"
    third["核心问题"] = "设备机型如何查询"
    third["问题意图"] = "信息查询"
    third["对象/部位"] = "机型"

    clustering_meta = {}
    topics, mapping, gaps, pending = build_topic_review_rows(
        [first, second, third],
        _standards(),
        use_mimo=False,
        clustering_mode="semantic",
        semantic_threshold=0.8,
        embedding_client=_FakeEmbedding(),
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 2
    assert len(mapping) == 3
    assert not gaps
    assert not pending
    singleton = next(topic for topic in topics if topic["主题样本数"] == 1)
    assert singleton["主题来源记录ID"] == "PHONE-003"
    assert singleton["是否重点复核"] == "是"
    assert clustering_meta["effective_mode"] == "semantic"
    assert clustering_meta["model"] == "semantic-cluster-test"


def test_mimo_cluster_gate_rejects_false_merge_and_merges_confirmed_topic() -> None:
    class ClusterGateEmbedding:
        config = SimpleNamespace(model="semantic-cluster-test")

        def embed_texts(self, texts):
            assert len(texts) == 3
            return [
                [1.0, 0.0],
                [0.87, 0.49],
                [0.86, 0.51],
            ]

    class ClusterGateMimo:
        config = SimpleNamespace(model="mimo-cluster-gate-test")

        def review_cluster_pair(self, left, right, _similarity, _threshold):
            pair = {left["数据ID"], right["数据ID"]}
            decision = "同一主题" if pair == {"B", "C"} else "不同主题"
            return MimoLabelResult(
                candidate={
                    "decision": decision,
                    "topic_label": "测试主题",
                    "reason": "根据判定对象、问题意图和处理目标进行判断。",
                    "key_difference": "标准对象不同" if decision == "不同主题" else "",
                    "confidence": 0.9,
                },
                request_audit={"cluster_gate": "test"},
                response_audit={"choices": []},
            )

    rows = [
        {
            "数据ID": "A",
            "工单ID": "A",
            "聊天内容": "主板内部标签异常如何判定",
            "核心问题": "主板内部标签是否属于拆修痕迹",
            "判定依据": "需要结合主板内部痕迹核验",
            "产品类型": "手机",
            "一级分类": "拆修问题",
            "二级分类": "主板拆修",
            "问题意图": "痕迹核验",
            "对象/部位": "主板",
            "异常现象": "内部标签",
            "解题方式": "对照拆修标准",
            "主标准路径": "【拆修问题】-【主板拆修】",
        },
        {
            "数据ID": "B",
            "工单ID": "B",
            "聊天内容": "外壳防水标签变红是否需要判定",
            "核心问题": "外壳防水标签变红是否需要处理",
            "判定依据": "外壳标签不作为浸液判定依据",
            "产品类型": "手机",
            "一级分类": "浸液问题",
            "二级分类": "防水标",
            "问题意图": "异常核验",
            "对象/部位": "外壳",
            "异常现象": "防水标签变红",
            "解题方式": "对照浸液标准",
            "主标准路径": "【浸液问题】-【防水标】",
        },
        {
            "数据ID": "C",
            "工单ID": "C",
            "聊天内容": "手机外壳防水标发红怎么处理",
            "核心问题": "外壳防水标签变红是否需要处理",
            "判定依据": "外壳标签不作为浸液判定依据",
            "产品类型": "手机",
            "一级分类": "浸液问题",
            "二级分类": "防水标",
            "问题意图": "异常核验",
            "对象/部位": "外壳",
            "异常现象": "防水标签变红",
            "解题方式": "对照浸液标准",
            "主标准路径": "【浸液问题】-【防水标】",
        },
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=ClusterGateMimo(),
        clustering_mode="semantic_mimo",
        semantic_threshold=0.84,
        cluster_review_floor=0.75,
        cluster_auto_merge_threshold=0.9999,
        cluster_review_limit=10,
        embedding_client=ClusterGateEmbedding(),
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 2
    assert {row["来源记录ID"] for row in mapping} == {"A", "B", "C"}
    mapping_by_id = {row["来源记录ID"]: row for row in mapping}
    assert mapping_by_id["B"]["聚类决策"] == "业务硬规则冲突后新建主题"
    assert mapping_by_id["B"]["聚类裁决提供方"] == "business-rule"
    assert mapping_by_id["C"]["聚类决策"] == "大模型确认合并"
    assert mapping_by_id["C"]["聚类裁决提供方"] == "mimo"
    assert not gaps
    assert not pending
    assert clustering_meta["effective_mode"] == "semantic_mimo"
    assert clustering_meta["mimo_review_calls"] == 1
    assert clustering_meta["mimo_review_approved"] == 1
    assert clustering_meta["mimo_review_rejected"] == 0
    assert clustering_meta["mimo_hard_rule_rejected"] == 1


def test_direct_mimo_clusters_one_to_many_without_embedding() -> None:
    class DirectMimo:
        config = SimpleNamespace(model="mimo-direct-test")

        def analyze_cluster_units(self, row):
            subject = "屏幕" if row["数据ID"] in {"A", "B"} else "摄像头"
            category_l1 = "显示问题" if subject == "屏幕" else "功能问题"
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "会话包含一个清晰问题。",
                    "topics": [
                        {
                            "normalized_issue": f"手机｜{subject}｜异常｜核验",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": category_l1,
                            "category_l2": f"{subject}异常",
                            "intent": "检测核验",
                            "subject": subject,
                            "phenomenon": "异常",
                            "judgment_target": f"判断{subject}是否异常",
                            "resolution_mode": "对照标准核验",
                            "standard_path": f"对照{subject}标准核验",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            screen_member_ids = [
                unit["unit_id"]
                for unit in units
                if unit["subject"] == "屏幕"
            ]
            camera_member_ids = [
                unit["unit_id"]
                for unit in units
                if unit["subject"] == "摄像头"
            ]
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "屏幕异常核验",
                            "member_atomic_ids": screen_member_ids,
                            "merge_basis": "适用范围、对象、目标、标准路径和阈值一致。",
                        },
                        {
                            "cluster_id": "C002",
                            "theme_name": "摄像头异常核验",
                            "member_atomic_ids": camera_member_ids,
                            "merge_basis": "摄像头与屏幕是不同判定对象。",
                        },
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": conversation,
            "核心问题": conversation,
            "产品类型": "手机",
            "问题意图": "检测核验",
            "对象/部位": "待确认",
            "异常现象": "异常",
            "解题方式": "对照标准核验",
        }
        for record_id, conversation in (
            ("A", "屏幕异常怎么核验"),
            ("B", "屏幕显示异常如何确认"),
            ("C", "摄像头功能异常如何核验"),
        )
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=DirectMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 2
    assert sorted(topic["主题样本数"] for topic in topics) == [1, 2]
    assert len(mapping) == 3
    assert not gaps
    assert not pending
    assert clustering_meta["effective_mode"] == "direct_mimo"
    assert clustering_meta["atomic_unit_count"] == 3
    assert clustering_meta["direct_cluster_calls"] == 1


def test_direct_mimo_clusters_same_product_before_model_label_partitioning() -> None:
    class ProductFirstMimo:
        config = SimpleNamespace(model="mimo-product-first-test")

        def analyze_cluster_units(self, row):
            category_l1 = "信息查询" if row["数据ID"] == "A" else "基本情况"
            intent = "信息查询" if row["数据ID"] == "A" else "检测核验"
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "会话都在确认同一类设备型号信息。",
                    "topics": [
                        {
                            "normalized_issue": "手机｜设备型号｜核验型号信息｜确认型号",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": category_l1,
                            "category_l2": "型号查询",
                            "intent": intent,
                            "subject": "设备型号",
                            "phenomenon": "型号信息待核验",
                            "judgment_target": "确认设备型号",
                            "resolution_mode": "按设备标识核验型号信息",
                            "standard_path": "查看设备标识并核验型号",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "聊天都在询问设备型号如何确认。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            assert {unit["unit_id"] for unit in units} == {"A-U1", "B-U1"}
            assert {
                unit["historical_actual_reply"] for unit in units
            } == {"请按设备标识核验型号信息。"}
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机设备型号核验",
                            "member_atomic_ids": [unit["unit_id"] for unit in units],
                            "merge_basis": "同品类、同对象、同判定目标和同处理路径。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": "手机型号怎么确认",
            "核心问题": "手机型号怎么确认",
            "历史实际回复": "请按设备标识核验型号信息。",
            "产品类型": "手机",
        }
        for record_id in ("A", "B")
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=ProductFirstMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 1
    assert topics[0]["主题样本数"] == 2
    assert len(mapping) == 2
    assert not gaps
    assert not pending
    assert clustering_meta["direct_cluster_calls"] == 1


def test_direct_mimo_source_product_overrides_model_product_mislabel() -> None:
    class MislabelProductMimo:
        config = SimpleNamespace(model="mimo-product-conflict-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "模型错误地把两个来源品类都标为手机。",
                    "topics": [
                        {
                            "normalized_issue": f"{row['聊天内容']}如何判定",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "功能问题",
                            "category_l2": "功能状态",
                            "intent": "标准判定",
                            "subject": "设备功能",
                            "phenomenon": "功能异常",
                            "judgment_target": "判断功能是否异常",
                            "resolution_mode": "按实际功能核验",
                            "standard_path": "设备功能判定",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "设备功能判定",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "模型错误尝试跨品类合并。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "PHONE",
            "工单ID": "PHONE",
            "聊天内容": "手机功能异常",
            "产品类型": "手机",
        },
        {
            "数据ID": "TABLET",
            "工单ID": "TABLET",
            "聊天内容": "平板功能异常",
            "产品类型": "平板",
        },
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=MislabelProductMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
        cluster_only=True,
    )

    assert len(topics) == 2
    assert {topic["产品类型"] for topic in topics} == {"手机", "平板电脑"}
    assert not mapping
    assert clustering_meta["atomic_product_conflicts"] == 1
    assert sum(topic["是否重点复核"] == "是" for topic in topics) >= 1
    assert not gaps
    assert not pending


def test_direct_mimo_unknown_source_product_stays_isolated() -> None:
    class UnknownSourceProductMimo:
        config = SimpleNamespace(model="mimo-unknown-source-product-test")

        def __init__(self) -> None:
            self.cluster_calls = 0

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "流程操作",
                            "category_l2": "节点核验",
                            "intent": "流程操作",
                            "subject": "节点状态",
                            "phenomenon": "状态待确认",
                            "judgment_target": "确认节点状态",
                            "resolution_mode": "按会话上下文核对",
                            "standard_path": "节点核验流程",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, _units):
            self.cluster_calls += 1
            raise AssertionError("源品类待确认的记录必须逐条隔离")

    reviewer = UnknownSourceProductMimo()
    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"聚合回收节点状态核验{record_id}",
            "产品类型": "待确认",
        }
        for record_id in ("A", "B")
    ]

    groups, _meta = _direct_mimo_topic_groups(
        rows,
        reviewer,
        batch_size=2,
    )

    assert len(groups) == 2
    assert reviewer.cluster_calls == 0
    assert {
        group_rows[0]["产品类型"]
        for _key, group_rows in groups
    } == {"待确认"}
    assert all(
        group_rows[0]["_原子需要复核"]
        for _key, group_rows in groups
    )


def test_direct_mimo_repairs_duplicate_atomic_ids_before_clustering() -> None:
    class DuplicateWorkOrderMimo:
        config = SimpleNamespace(model="mimo-duplicate-id-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "会话包含一个清晰问题。",
                    "topics": [
                        {
                            "normalized_issue": f"手机｜屏幕｜{row['聊天内容']}｜判定",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "显示问题",
                            "category_l2": "屏幕异常",
                            "intent": "标准判定",
                            "subject": "屏幕",
                            "phenomenon": row["聊天内容"],
                            "judgment_target": "判断屏幕异常",
                            "resolution_mode": "对照标准核验",
                            "standard_path": "屏幕异常判定标准",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            unit_ids = [unit["unit_id"] for unit in units]
            assert len(unit_ids) == len(set(unit_ids))
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机屏幕异常判定",
                            "member_atomic_ids": unit_ids,
                            "merge_basis": "同品类、同对象和同判定目标。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "工单ID": "WO-DUP",
            "聊天内容": conversation,
            "核心问题": conversation,
            "产品类型": "手机",
        }
        for conversation in ("屏幕色斑怎么判", "屏幕亮斑怎么判")
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=DuplicateWorkOrderMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 2
    assert len(mapping) == 2
    assert not gaps
    assert not pending
    assert clustering_meta["atomic_unit_id_collisions_resolved"] == 1
    assert clustering_meta["direct_cluster_calls"] == 1


def test_direct_mimo_reuses_completed_cluster_batches_from_progress(
    tmp_path: Path,
) -> None:
    def topic_payload(text: str) -> dict[str, object]:
        return {
            "normalized_issue": f"手机｜屏幕｜{text}｜判定",
            "product_category": "手机",
            "scope_type": "品类专用",
            "platform": "通用",
            "brand": "通用",
            "model_scope": "通用",
            "category_l1": "显示问题",
            "category_l2": text,
            "intent": "标准判定",
            "subject": "屏幕",
            "phenomenon": text,
            "judgment_target": f"判断屏幕{text}",
            "resolution_mode": "对照标准核验",
            "standard_path": "屏幕显示异常判定标准",
            "threshold_or_exception": "无明确阈值",
            "evidence_summary": "完整聊天支持该问题。",
            "confidence": 0.9,
            "requires_review": False,
        }

    class CacheMimo:
        config = SimpleNamespace(model="mimo-cluster-cache-test")

        def __init__(self) -> None:
            self.atomic_calls = 0
            self.cluster_calls = 0

        def analyze_cluster_units(self, row):
            self.atomic_calls += 1
            return MimoLabelResult(
                candidate={"topics": [topic_payload(row["聊天内容"])]},
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            self.cluster_calls += 1
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机屏幕异常判定",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "同品类、同对象和同判定目标。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

        def review_cluster_membership(self, *_args, **_kwargs):
            raise AssertionError("已缓存的批次不应进入二次模型复核")

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": text,
            "核心问题": text,
            "产品类型": "手机",
        }
        for record_id, text in (("A", "色斑"), ("B", "色斑"))
    ]
    progress_path = tmp_path / "direct_mimo_progress.json"

    first_client = CacheMimo()
    _direct_mimo_topic_groups(
        rows,
        first_client,
        batch_size=2,
        progress_path=progress_path,
    )
    assert first_client.atomic_calls == 2
    assert first_client.cluster_calls == 1

    class CachedOnlyMimo(CacheMimo):
        def analyze_cluster_units(self, _row):
            raise AssertionError("原子问题应从缓存读取")

        def cluster_atomic_units(self, _units):
            raise AssertionError("主题聚类批次应从缓存读取")

    second_client = CachedOnlyMimo()
    groups, meta = _direct_mimo_topic_groups(
        rows,
        second_client,
        batch_size=2,
        progress_path=progress_path,
    )

    assert len(groups) == 1
    assert meta["atomic_extraction_cache_hits"] == 2
    assert meta["direct_cluster_calls"] == 0


def test_direct_mimo_caches_successful_split_retry_batches(
    tmp_path: Path,
) -> None:
    def topic_payload(row: dict[str, object]) -> dict[str, object]:
        return {
            "normalized_issue": row["聊天内容"],
            "product_category": "手机",
            "scope_type": "品类专用",
            "platform": "通用",
            "brand": "通用",
            "model_scope": "通用",
            "category_l1": "流程操作",
            "category_l2": "节点核验",
            "intent": "流程操作",
            "subject": "节点状态",
            "phenomenon": "状态待确认",
            "judgment_target": "确认节点状态",
            "resolution_mode": "按会话上下文核对",
            "standard_path": "节点核验流程",
            "threshold_or_exception": "无明确阈值",
            "evidence_summary": "完整聊天支持该问题。",
            "confidence": 0.9,
            "requires_review": False,
        }

    class SplitRetryMimo:
        config = SimpleNamespace(model="mimo-split-retry-cache-test")

        def __init__(self) -> None:
            self.cluster_calls = 0

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={"topics": [topic_payload(row)]},
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            self.cluster_calls += 1
            if len(units) == 4:
                raise MimoError("模拟大批次JSON无效")
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "节点状态核验",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "拆半后的小批次可以正常聚类。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": f"R{index}",
            "工单ID": f"R{index}",
            "聊天内容": f"节点状态核验{index}",
            "产品类型": "手机",
        }
        for index in range(4)
    ]
    progress_path = tmp_path / "direct_mimo_progress.json"
    first_client = SplitRetryMimo()

    first_groups, first_meta = _direct_mimo_topic_groups(
        rows,
        first_client,
        batch_size=4,
        progress_path=progress_path,
    )

    assert len(first_groups) == 2
    assert first_client.cluster_calls == 3
    assert first_meta["direct_cluster_retry_splits"] == 1

    class CachedOnlyMimo(SplitRetryMimo):
        def analyze_cluster_units(self, _row):
            raise AssertionError("原子提取应从缓存读取")

        def cluster_atomic_units(self, _units):
            raise AssertionError("拆半重试结果应从缓存读取")

    second_client = CachedOnlyMimo()
    second_groups, second_meta = _direct_mimo_topic_groups(
        rows,
        second_client,
        batch_size=4,
        progress_path=progress_path,
    )

    assert len(second_groups) == 2
    assert second_meta["direct_cluster_calls"] == 0
    assert second_meta["direct_cluster_cache_hits"] == 1


def test_direct_mimo_resumes_atomic_extraction_from_progress_checkpoint(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "1")

    def topic_payload() -> dict[str, object]:
        return {
            "normalized_issue": "手机｜屏幕｜色斑｜判定",
            "product_category": "手机",
            "scope_type": "品类专用",
            "platform": "通用",
            "brand": "通用",
            "model_scope": "通用",
            "category_l1": "显示问题",
            "category_l2": "色斑",
            "intent": "标准判定",
            "subject": "屏幕",
            "phenomenon": "色斑",
            "judgment_target": "判断屏幕色斑",
            "resolution_mode": "对照标准核验",
            "standard_path": "屏幕色斑判定标准",
            "threshold_or_exception": "无明确阈值",
            "evidence_summary": "完整聊天支持该问题。",
            "confidence": 0.9,
            "requires_review": False,
        }

    class InterruptedMimo:
        config = SimpleNamespace(model="mimo-progress-test")

        def analyze_cluster_units(self, row):
            if row["数据ID"] == "B":
                raise KeyboardInterrupt("simulated interruption")
            return MimoLabelResult(
                candidate={"topics": [topic_payload()]},
                request_audit={},
                response_audit={},
            )

    class ResumedMimo:
        config = SimpleNamespace(model="mimo-progress-test")

        def __init__(self) -> None:
            self.analyzed_ids: list[str] = []

        def analyze_cluster_units(self, row):
            self.analyzed_ids.append(row["数据ID"])
            return MimoLabelResult(
                candidate={"topics": [topic_payload()]},
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机屏幕色斑判定",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "同品类、同对象和同判定目标。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"{record_id} 的屏幕色斑怎么判",
            "产品类型": "手机",
        }
        for record_id in ("A", "B", "C")
    ]
    progress_path = tmp_path / "direct_mimo_progress.json"

    with pytest.raises(KeyboardInterrupt, match="simulated interruption"):
        _direct_mimo_topic_groups(
            rows,
            InterruptedMimo(),
            progress_path=progress_path,
        )

    resumed_client = ResumedMimo()
    _groups, meta = _direct_mimo_topic_groups(
        rows,
        resumed_client,
        progress_path=progress_path,
    )

    assert resumed_client.analyzed_ids == ["B", "C"]
    assert meta["atomic_extraction_cache_hits"] == 1


def test_direct_mimo_progress_checkpoint_allows_parallel_atomic_extraction(
    tmp_path: Path,
    monkeypatch,
) -> None:
    class ParallelMimo:
        config = SimpleNamespace(model="mimo-parallel-progress-test")

        def __init__(self) -> None:
            self.lock = threading.Lock()
            self.active = 0
            self.max_active = 0

        def analyze_cluster_units(self, row):
            with self.lock:
                self.active += 1
                self.max_active = max(self.max_active, self.active)
            try:
                time.sleep(0.04)
                return MimoLabelResult(
                    candidate={
                        "topics": [
                            {
                                "normalized_issue": row["聊天内容"],
                                "product_category": "手机",
                                "scope_type": "品类专用",
                                "platform": "通用",
                                "brand": "通用",
                                "model_scope": "通用",
                                "category_l1": "信息查询",
                                "category_l2": "设备信息",
                                "intent": "信息查询",
                                "subject": "设备信息",
                                "phenomenon": "查询设备信息",
                                "judgment_target": "确认设备信息",
                                "resolution_mode": "查询并核对",
                                "standard_path": "设备信息查询",
                                "threshold_or_exception": "无明确阈值",
                                "evidence_summary": "完整聊天支持该问题。",
                                "confidence": 0.9,
                                "requires_review": False,
                            }
                        ]
                    },
                    request_audit={},
                    response_audit={},
                )
            finally:
                with self.lock:
                    self.active -= 1

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "设备信息查询",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "同品类、同查询对象和同处理路径。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "3")
    rows = [
        {
            "数据ID": f"P{index}",
            "工单ID": f"P{index}",
            "聊天内容": f"查询设备信息{index}",
            "产品类型": "手机",
        }
        for index in range(6)
    ]
    client = ParallelMimo()

    _direct_mimo_topic_groups(
        rows,
        client,
        batch_size=6,
        progress_path=tmp_path / "direct_mimo_progress.json",
    )

    assert client.max_active >= 2


def test_direct_mimo_isolates_unexpected_atomic_extraction_failure(
    tmp_path: Path,
) -> None:
    class IsolatedFailureMimo:
        config = SimpleNamespace(model="mimo-isolated-atomic-failure-test")

        def analyze_cluster_units(self, row):
            if row["数据ID"] == "A":
                raise ValueError("模拟单条普通异常")
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "功能问题",
                            "category_l2": "功能状态",
                            "intent": "标准判定",
                            "subject": "设备功能",
                            "phenomenon": "功能异常",
                            "judgment_target": "判断功能是否异常",
                            "resolution_mode": "按实际功能核验",
                            "standard_path": "设备功能判定",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "分别保留。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"手机功能异常{record_id}",
            "产品类型": "手机",
        }
        for record_id in ("A", "B")
    ]

    progress_path = tmp_path / "direct_mimo_progress.json"
    groups, meta = _direct_mimo_topic_groups(
        rows,
        IsolatedFailureMimo(),
        batch_size=2,
        progress_path=progress_path,
    )

    assert len(groups) == 2
    assert meta["atomic_extraction_failed"] == 1
    assert meta["atomic_unit_count"] == 2
    assert meta["atomic_extraction_failure_reasons"][0]["reason"] == (
        "ValueError: 模拟单条普通异常"
    )
    progress = json.loads(progress_path.read_text(encoding="utf-8"))
    failed_results = [
        result
        for result in progress["atomic_results"].values()
        if result.get("failed")
    ]
    assert len(failed_results) == 1
    assert failed_results[0]["topics"]


def test_direct_mimo_recovers_screen_validation_failure_as_review_singleton(
    tmp_path: Path,
) -> None:
    class RecoverableScreenFailureMimo:
        config = SimpleNamespace(model="mimo-screen-validation-recovery-test")

        def analyze_cluster_units(self, _row):
            raise MimoError(
                "MiMo 聚类问题单元 JSON 校验失败（已重试两次）："
                "屏幕显示现象必须归入显示问题，不得误归外观或拆修问题"
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "屏幕进灰如何判定",
                            "member_atomic_ids": [unit["unit_id"] for unit in units],
                            "merge_basis": "保守单主题候选。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "SCREEN-RECOVERY-001",
            "工单ID": "SCREEN-RECOVERY-001",
            "产品类型": "平板电脑",
            "聊天内容": "手电筒照屏幕内部有白色颗粒，想确认是漏液还是进灰。",
            "核心问题": "平板屏幕内部白色颗粒是进灰还是漏液如何判定",
            "对象/部位": "平板屏幕",
            "异常现象": "屏幕内部白色颗粒",
            "解题方式": "根据图片和描述特征区分屏幕进灰与漏液",
            "一级分类": "屏幕外观情况",
            "二级分类": "屏幕进灰（单选）",
        }
    ]

    groups, meta = _direct_mimo_topic_groups(
        rows,
        RecoverableScreenFailureMimo(),
        batch_size=1,
        progress_path=tmp_path / "direct_mimo_progress.json",
    )

    assert len(groups) == 1
    assert meta["atomic_extraction_failed"] == 0
    assert meta["atomic_extraction_recovered"] == 1
    recovered = groups[0][1][0]
    assert recovered["_原子需要复核"] is True
    assert "屏幕显示现象必须归入显示问题" in recovered["人工优先复核原因"]


def test_direct_mimo_locally_splits_notebook_model_and_hardware_brand_queries() -> None:
    class CombinedNotebookQueryMimo:
        config = SimpleNamespace(model="mimo-local-query-split-test")

        def analyze_cluster_units(self, _row):
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": "笔记本型号和内存品牌查询",
                            "product_category": "笔记本",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "信息查询",
                            "category_l2": "设备配置",
                            "intent": "信息查询",
                            "subject": "设备配置",
                            "phenomenon": "型号和内存品牌待确认",
                            "judgment_target": "确认型号和内存品牌",
                            "resolution_mode": "查询并核对",
                            "standard_path": "设备配置查询",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "聊天中分别确认型号和内存品牌。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "本地拆分后的独立查询目标。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "NOTEBOOK-MULTI-001",
            "工单ID": "NOTEBOOK-MULTI-001",
            "产品类型": "笔记本",
            "聊天内容": (
                "帮忙看看型号和内存。小型号看是这一款，"
                "内存和硬盘都是品牌认证的。"
            ),
        }
    ]

    groups, meta = _direct_mimo_topic_groups(
        rows,
        CombinedNotebookQueryMimo(),
        batch_size=4,
    )
    group_rows = [group_rows[0] for _key, group_rows in groups]
    targets = {
        workflow_module._direct_reconcile_fingerprint(row).query_target
        for row in group_rows
    }

    assert len(groups) == 2
    assert targets == {"model_query", "memory_storage_brand"}
    assert meta["local_multi_topic_rescue"] == 1


def test_direct_mimo_batches_text_atomic_extraction(monkeypatch) -> None:
    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "1")
    monkeypatch.setenv("ANSWER_HUB_DIRECT_ATOMIC_BATCH_SIZE", "3")
    monkeypatch.setenv("ANSWER_HUB_DIRECT_ATOMIC_BATCH_MAX_CHARS", "24000")

    def topic_payload(row: dict[str, object]) -> dict[str, object]:
        return {
            "normalized_issue": row["聊天内容"],
            "product_category": "手机",
            "scope_type": "品类专用",
            "platform": "通用",
            "brand": "通用",
            "model_scope": "通用",
            "category_l1": "流程操作",
            "category_l2": "节点核验",
            "intent": "流程操作",
            "subject": "节点状态",
            "phenomenon": "状态待确认",
            "judgment_target": "确认节点状态",
            "resolution_mode": "按会话上下文核对",
            "standard_path": "节点核验流程",
            "threshold_or_exception": "无明确阈值",
            "evidence_summary": "完整聊天支持该问题。",
            "confidence": 0.9,
            "requires_review": False,
        }

    class BatchAtomicMimo:
        config = SimpleNamespace(
            model="mimo-batch-atomic-test",
            media_model="mimo-batch-atomic-test",
            cluster_media_policy="on_demand",
            cluster_media_min_text_chars=220,
        )

        def __init__(self) -> None:
            self.batch_sizes: list[int] = []

        def can_batch_cluster_units(self, _row):
            return True

        def analyze_cluster_units_batch(self, rows):
            self.batch_sizes.append(len(rows))
            return [
                MimoLabelResult(
                    candidate={"topics": [topic_payload(row)]},
                    request_audit={},
                    response_audit={},
                )
                for row in rows
            ]

        def analyze_cluster_units(self, _row):
            raise AssertionError("纯文本记录应使用批量原子提取")

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "分别保留。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    reviewer = BatchAtomicMimo()
    rows = [
        {
            "数据ID": f"R{index}",
            "工单ID": f"R{index}",
            "聊天内容": f"节点状态核验{index}",
            "产品类型": "手机",
        }
        for index in range(6)
    ]

    groups, meta = _direct_mimo_topic_groups(
        rows,
        reviewer,
        batch_size=6,
    )

    assert sum(len(group_rows) for _key, group_rows in groups) == 6
    assert reviewer.batch_sizes == [3, 3]
    assert meta["atomic_extraction_calls"] == 6
    assert meta["atomic_extraction_model_requests"] == 2
    assert meta["atomic_extraction_batch_calls"] == 2


def test_direct_reconcile_fingerprint_ignores_other_atomic_topics_in_chat() -> None:
    fingerprint = workflow_module._direct_reconcile_fingerprint(
        {
            "产品类型": "笔记本",
            "模型主题一级分类": "显示问题",
            "问题意图": "标准判定",
            "对象/部位": "屏幕",
            "异常现象": "色斑",
            "核心问题": "笔记本｜屏幕｜色斑｜判断显示异常",
            "判定目标": "判断是否属于屏幕色斑",
            "解题方式": "按屏幕色斑标准核验",
            "语义标注依据": "当前原子问题只讨论屏幕固定色斑。",
            "聊天内容": (
                "先确认屏幕固定色斑。另外再帮忙看内存和硬盘是不是品牌认证。"
            ),
        }
    )

    assert fingerprint.query_target == ""
    assert fingerprint.standard_family


def test_direct_reconcile_rule_requires_minimum_similarity_for_trusted_target(
    monkeypatch,
) -> None:
    candidate = {
        "产品类型": "笔记本",
        "模型主题一级分类": "信息查询",
        "问题意图": "信息查询",
        "对象/部位": "内存和硬盘",
        "异常现象": "品牌属性待确认",
        "核心问题": "笔记本内存和硬盘是否属于品牌件",
        "判定目标": "确认内存硬盘品牌属性",
        "解题方式": "核对品牌认证信息",
        "_原子平台": "通用",
        "_原子品牌": "通用",
        "_原子机型范围": "通用",
    }
    target = {
        **candidate,
        "核心问题": "如何判断硬盘与内存是不是原厂品牌",
        "异常现象": "原厂或第三方品牌待确认",
        "判定目标": "判断内存硬盘是否为品牌配件",
    }
    monkeypatch.setattr(
        workflow_module,
        "_direct_reconcile_same_topic_family",
        lambda _left, _right: False,
    )
    monkeypatch.setattr(
        workflow_module,
        "_direct_reconcile_same_target",
        lambda _left, _right: True,
    )

    assert (
        workflow_module._direct_reconcile_rule_merge_reason(
            candidate,
            [target],
            0.17,
        )
        == ""
    )
    assert "同一可信业务目标" in (
        workflow_module._direct_reconcile_rule_merge_reason(
            candidate,
            [target],
            0.18,
        )
    )


def test_direct_reconcile_trusted_target_overrides_noisy_rule_family_but_not_product(
    monkeypatch,
) -> None:
    notebook_brand = {
        "产品类型": "笔记本",
        "模型主题一级分类": "信息查询",
        "问题意图": "信息查询",
        "对象/部位": "硬盘",
        "异常现象": "是否为品牌认证硬盘",
        "核心问题": "笔记本硬盘是不是品牌件",
        "判定目标": "确认硬盘品牌属性",
        "解题方式": "通过工具读取品牌信息",
        "_聚类标准族": "笔记本品牌型号机型核验",
        "_聚类合并策略": "separate_by_query_target",
        "_聚类现象值": "品牌型号机型核验",
        "_原子平台": "通用",
        "_原子品牌": "通用",
        "_原子机型范围": "通用",
    }
    notebook_memory = {
        **notebook_brand,
        "对象/部位": "运行内存",
        "异常现象": "品牌未读出时是否按品牌判定",
        "核心问题": "工具没读出内存品牌时怎么判",
        "判定目标": "确认内存品牌属性",
        "_聚类标准族": "笔记本运行内存硬盘CPU显卡配置",
        "_聚类现象值": "运行内存硬盘CPU显卡配置",
    }
    phone_memory = {
        **notebook_memory,
        "产品类型": "手机",
    }

    assert not workflow_module._direct_reconcile_has_hard_conflict(
        notebook_brand,
        [notebook_memory],
    )
    assert workflow_module._direct_reconcile_has_hard_conflict(
        notebook_brand,
        [phone_memory],
    )


def test_direct_reconcile_merges_multi_member_clusters_for_same_trusted_target() -> None:
    def row(record_id: str, subject: str) -> dict[str, object]:
        return {
            "数据ID": record_id,
            "工单ID": record_id,
            "_原子知识ID": f"{record_id}-U01",
            "产品类型": "笔记本",
            "模型主题一级分类": "信息查询",
            "问题意图": "信息查询",
            "对象/部位": subject,
            "异常现象": "是否为品牌认证配件",
            "核心问题": f"笔记本{subject}是不是品牌件",
            "判定目标": f"确认{subject}品牌属性",
            "解题方式": "通过工具读取品牌信息",
            "_聚类标准族": "笔记本品牌型号机型核验",
            "_聚类合并策略": "separate_by_query_target",
            "_聚类现象值": "品牌型号机型核验",
            "_原子平台": "通用",
            "_原子品牌": "通用",
            "_原子机型范围": "通用",
            "_聚类裁决提供方": "mimo-direct",
            "_聚类主题标题": "笔记本内存硬盘品牌属性查询",
        }

    groups = [
        (
            ("direct_mimo", "笔记本", "first"),
            [row("A", "硬盘"), row("B", "内存")],
        ),
        (
            ("direct_mimo", "笔记本", "second"),
            [row("C", "固态硬盘"), row("D", "运行内存")],
        ),
    ]
    meta: dict[str, object] = {}
    reviewer = SimpleNamespace(
        config=SimpleNamespace(model="offline-multi-cluster-test")
    )

    reconciled = workflow_module._reconcile_direct_topic_groups(
        groups,
        reviewer,
        meta,
    )

    assert len(reconciled) == 1
    assert len(reconciled[0][1]) == 4
    assert meta["direct_reconcile_rule_approved"] == 1


def test_unconfigured_product_multi_member_cluster_is_forced_apart() -> None:
    first = {
        "数据ID": "AGG-A",
        "工单ID": "AGG-A",
        "_原子知识ID": "AGG-A-U01",
        "产品类型": "聚合回收",
        "模型主题一级分类": "流程操作",
        "问题意图": "信息查询",
        "对象/部位": "回收节点",
        "异常现象": "节点状态待确认",
        "核心问题": "聚合回收节点状态怎么查询",
    }
    second = {
        **first,
        "数据ID": "AGG-B",
        "工单ID": "AGG-B",
        "_原子知识ID": "AGG-B-U01",
        "对象/部位": "订单状态",
        "核心问题": "聚合回收订单状态怎么查询",
    }

    assert "未配置品类" in workflow_module._direct_cluster_hard_conflict_reason(
        [first, second]
    )


def test_different_business_hierarchy_levels_are_forced_apart() -> None:
    self_operated = {
        "数据ID": "SELF-PHONE",
        "_原子知识ID": "SELF-PHONE-U1",
        "回收业务层级": "自营回收",
        "产品类型": "手机",
        "模型主题一级分类": "外观问题",
        "对象/部位": "屏幕",
        "异常现象": "碎裂",
        "核心问题": "手机屏幕碎裂怎么判",
    }
    aggregate = {
        **self_operated,
        "数据ID": "AGG-PHONE",
        "_原子知识ID": "AGG-PHONE-U1",
        "回收业务层级": "聚合回收",
    }

    reason = workflow_module._direct_cluster_hard_conflict_reason(
        [self_operated, aggregate]
    )

    assert "回收业务层级不同" in reason


def test_aggregate_business_line_does_not_use_self_operated_quality_rules() -> None:
    self_operated = {
        "回收业务层级": "自营回收",
        "产品类型": "手机",
        "对象/部位": "屏幕",
        "异常现象": "碎裂",
        "核心问题": "手机屏幕碎裂怎么判",
    }
    aggregate = {
        **self_operated,
        "回收业务层级": "聚合回收",
    }

    assert workflow_module._direct_clustering_rule_match(self_operated) is not None
    assert workflow_module._direct_clustering_rule_match(aggregate) is None


def test_unconfigured_product_singletons_cannot_be_remerged() -> None:
    def row(record_id: str) -> dict[str, object]:
        return {
            "数据ID": record_id,
            "工单ID": record_id,
            "_原子知识ID": f"{record_id}-U01",
            "产品类型": "聚合回收",
            "模型主题一级分类": "流程操作",
            "问题意图": "信息查询",
            "对象/部位": "回收节点",
            "异常现象": "节点状态待确认",
            "核心问题": "聚合回收节点状态怎么查询",
            "_聚类裁决提供方": "mimo-direct-post-guard",
            "_聚类主题标题": "聚合回收节点状态查询",
        }

    first = row("AGG-A")
    second = row("AGG-B")
    groups = [
        (("direct_mimo", "聚合回收", "first"), [first]),
        (("direct_mimo", "聚合回收", "second"), [second]),
    ]
    meta: dict[str, object] = {}
    reviewer = SimpleNamespace(
        config=SimpleNamespace(model="offline-unconfigured-product-test")
    )

    assert workflow_module._direct_reconcile_has_hard_conflict(first, [second])

    reconciled = workflow_module._reconcile_direct_topic_groups(
        groups,
        reviewer,
        meta,
    )

    assert len(reconciled) == 2
    assert all(len(rows) == 1 for _key, rows in reconciled)


def test_direct_mimo_splits_failed_atomic_batches_to_single_rows(
    monkeypatch,
) -> None:
    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "1")
    monkeypatch.setenv("ANSWER_HUB_DIRECT_ATOMIC_BATCH_SIZE", "4")

    class SplitFailedBatchMimo:
        config = SimpleNamespace(
            model="mimo-split-failed-atomic-batch-test",
            media_model="mimo-split-failed-atomic-batch-test",
            cluster_media_policy="on_demand",
            cluster_media_min_text_chars=220,
        )

        def __init__(self) -> None:
            self.batch_calls = 0
            self.single_calls = 0

        def can_batch_cluster_units(self, _row):
            return True

        def analyze_cluster_units_batch(self, _rows):
            self.batch_calls += 1
            raise MimoError("模拟批量JSON不完整")

        def analyze_cluster_units(self, row):
            self.single_calls += 1
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "流程操作",
                            "category_l2": "节点核验",
                            "intent": "流程操作",
                            "subject": "节点状态",
                            "phenomenon": "状态待确认",
                            "judgment_target": "确认节点状态",
                            "resolution_mode": "按会话上下文核对",
                            "standard_path": "节点核验流程",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "分别保留。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    reviewer = SplitFailedBatchMimo()
    rows = [
        {
            "数据ID": f"R{index}",
            "工单ID": f"R{index}",
            "聊天内容": f"节点状态核验{index}",
            "产品类型": "手机",
        }
        for index in range(4)
    ]

    groups, meta = _direct_mimo_topic_groups(
        rows,
        reviewer,
        batch_size=4,
    )

    assert sum(len(group_rows) for _key, group_rows in groups) == 4
    assert reviewer.batch_calls == 3
    assert reviewer.single_calls == 4
    assert meta["atomic_extraction_model_requests"] == 7
    assert meta["atomic_extraction_batch_splits"] == 3
    assert meta["atomic_extraction_failed"] == 0


def test_direct_mimo_atomic_batches_never_mix_product_categories(
    monkeypatch,
) -> None:
    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "1")
    monkeypatch.setenv("ANSWER_HUB_DIRECT_ATOMIC_BATCH_SIZE", "4")

    class ProductIsolatedBatchMimo:
        config = SimpleNamespace(
            model="mimo-product-isolated-batch-test",
            media_model="mimo-product-isolated-batch-test",
            cluster_media_policy="on_demand",
            cluster_media_min_text_chars=220,
        )

        def __init__(self) -> None:
            self.batch_products: list[set[str]] = []

        def can_batch_cluster_units(self, _row):
            return True

        def analyze_cluster_units_batch(self, rows):
            self.batch_products.append({row["产品类型"] for row in rows})
            return [
                MimoLabelResult(
                    candidate={
                        "topics": [
                            {
                                "normalized_issue": row["聊天内容"],
                                "product_category": row["产品类型"],
                                "scope_type": "品类专用",
                                "platform": "通用",
                                "brand": "通用",
                                "model_scope": "通用",
                                "category_l1": "流程操作",
                                "category_l2": "节点核验",
                                "intent": "流程操作",
                                "subject": "节点状态",
                                "phenomenon": "状态待确认",
                                "judgment_target": "确认节点状态",
                                "resolution_mode": "按会话上下文核对",
                                "standard_path": "节点核验流程",
                                "threshold_or_exception": "无明确阈值",
                                "evidence_summary": "完整聊天支持该问题。",
                                "confidence": 0.9,
                                "requires_review": False,
                            }
                        ]
                    },
                    request_audit={},
                    response_audit={},
                )
                for row in rows
            ]

        def analyze_cluster_units(self, _row):
            raise AssertionError("同品类文本记录应进入批量提取")

        def cluster_atomic_units(self, units):
            assert len({unit["product_category"] for unit in units}) == 1
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "分别保留。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    reviewer = ProductIsolatedBatchMimo()
    rows = [
        {
            "数据ID": f"{product}-{index}",
            "工单ID": f"{product}-{index}",
            "聊天内容": f"{product}节点状态核验{index}",
            "产品类型": product,
        }
        for index in range(2)
        for product in ("手机", "平板")
    ]

    groups, _meta = _direct_mimo_topic_groups(
        rows,
        reviewer,
        batch_size=4,
    )

    assert sum(len(group_rows) for _key, group_rows in groups) == 4
    assert reviewer.batch_products == [{"手机"}, {"平板"}]


def test_direct_mimo_atomic_batches_never_mix_business_hierarchy_levels(
    monkeypatch,
) -> None:
    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "1")
    monkeypatch.setenv("ANSWER_HUB_DIRECT_ATOMIC_BATCH_SIZE", "4")

    class BusinessIsolatedBatchMimo:
        config = SimpleNamespace(model="mimo-business-isolated-batch-test")

        def __init__(self) -> None:
            self.batch_business_lines: list[set[str]] = []

        def can_batch_cluster_units(self, _row):
            return True

        def analyze_cluster_units_batch(self, rows):
            self.batch_business_lines.append(
                {row["回收业务层级"] for row in rows}
            )
            return [
                MimoLabelResult(
                    candidate={
                        "topics": [
                            {
                                "normalized_issue": row["聊天内容"],
                                "product_category": row["产品类型"],
                                "scope_type": "品类专用",
                                "platform": "通用",
                                "brand": "通用",
                                "model_scope": "通用",
                                "category_l1": "外观问题",
                                "category_l2": "屏幕外观",
                                "intent": "标准判定",
                                "subject": "屏幕",
                                "phenomenon": "碎裂",
                                "judgment_target": "确认屏幕碎裂",
                                "resolution_mode": "按当前业务口径核验",
                                "standard_path": "屏幕外观",
                                "threshold_or_exception": "无明确阈值",
                                "evidence_summary": "完整聊天支持该问题。",
                                "confidence": 0.9,
                                "requires_review": False,
                            }
                        ]
                    },
                    request_audit={},
                    response_audit={},
                )
                for row in rows
            ]

        def analyze_cluster_units(self, _row):
            raise AssertionError("同业务层级记录应进入批量提取")

        def cluster_atomic_units(self, units):
            assert len({unit["business_line"] for unit in units}) == 1
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "分别保留。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    reviewer = BusinessIsolatedBatchMimo()
    rows = [
        {
            "数据ID": f"{line}-{index}",
            "工单ID": f"{line}-{index}",
            "聊天内容": f"{line}手机屏幕碎裂核验{index}",
            "回收业务层级": line,
            "产品类型": "手机",
        }
        for line in ("自营回收", "聚合回收")
        for index in range(2)
    ]

    groups, _meta = _direct_mimo_topic_groups(
        rows,
        reviewer,
        batch_size=4,
    )

    assert sum(len(group_rows) for _key, group_rows in groups) == 4
    assert {
        frozenset(batch)
        for batch in reviewer.batch_business_lines
    } == {
        frozenset({"自营回收"}),
        frozenset({"聚合回收"}),
    }


def test_direct_mimo_ignores_member_ids_from_another_product_batch(
    monkeypatch,
) -> None:
    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "1")
    monkeypatch.setenv("ANSWER_HUB_DIRECT_CLUSTER_MAX_WORKERS", "1")

    class ForeignMemberIdMimo:
        config = SimpleNamespace(model="mimo-foreign-member-id-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": row["产品类型"],
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "流程操作",
                            "category_l2": "节点核验",
                            "intent": "流程操作",
                            "subject": "节点状态",
                            "phenomenon": "状态待确认",
                            "judgment_target": "确认节点状态",
                            "resolution_mode": "按会话上下文核对",
                            "standard_path": "节点核验流程",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            if units[0]["product_category"] == "手机":
                clusters = [
                    {
                        "cluster_id": "PHONE-CROSS",
                        "theme_name": "手机节点状态",
                        "member_atomic_ids": [
                            units[0]["unit_id"],
                            "TABLET-1-U1",
                        ],
                        "merge_basis": "模型错误引用了其他品类批次的原子ID。",
                    },
                    {
                        "cluster_id": "PHONE-2",
                        "theme_name": "手机节点状态2",
                        "member_atomic_ids": [units[1]["unit_id"]],
                        "merge_basis": "单独保留。",
                    },
                ]
            else:
                clusters = [
                    {
                        "cluster_id": f"TABLET-{index}",
                        "theme_name": unit["normalized_issue"],
                        "member_atomic_ids": [unit["unit_id"]],
                        "merge_basis": "单独保留。",
                    }
                    for index, unit in enumerate(units, start=1)
                ]
            return MimoLabelResult(
                candidate={
                    "clusters": clusters,
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": f"{prefix}-{index}",
            "工单ID": f"{prefix}-{index}",
            "聊天内容": f"{product}节点状态核验{index}",
            "产品类型": product,
        }
        for prefix, product in (("PHONE", "手机"), ("TABLET", "平板"))
        for index in range(1, 3)
    ]

    groups, meta = _direct_mimo_topic_groups(
        rows,
        ForeignMemberIdMimo(),
        batch_size=4,
    )

    grouped_atomic_ids = [
        row["_原子知识ID"]
        for _key, group_rows in groups
        for row in group_rows
    ]
    assert len(grouped_atomic_ids) == len(set(grouped_atomic_ids)) == 4
    assert all(
        len({row["产品类型"] for row in group_rows}) == 1
        for _key, group_rows in groups
    )
    assert meta["direct_foreign_member_ids_ignored"] == 1


def test_direct_mimo_cluster_batches_run_in_parallel(
    tmp_path: Path,
    monkeypatch,
) -> None:
    class ParallelClusterMimo:
        config = SimpleNamespace(model="mimo-parallel-cluster-test")

        def __init__(self) -> None:
            self.lock = threading.Lock()
            self.active_cluster_calls = 0
            self.max_active_cluster_calls = 0

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "流程操作",
                            "category_l2": "节点核验",
                            "intent": "流程判断",
                            "subject": "节点状态",
                            "phenomenon": "状态待确认",
                            "judgment_target": "确认节点状态",
                            "resolution_mode": "按会话上下文核对",
                            "standard_path": "节点核验流程",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            with self.lock:
                self.active_cluster_calls += 1
                self.max_active_cluster_calls = max(
                    self.max_active_cluster_calls,
                    self.active_cluster_calls,
                )
            try:
                time.sleep(0.04)
                return MimoLabelResult(
                    candidate={
                        "clusters": [
                            {
                                "cluster_id": "C001",
                                "theme_name": "节点状态核验",
                                "member_atomic_ids": [
                                    unit["unit_id"] for unit in units
                                ],
                                "merge_basis": "同品类、同对象和同处理目标。",
                            }
                        ],
                        "split_requests": [],
                        "review_requests": [],
                    },
                    request_audit={},
                    response_audit={},
                )
            finally:
                with self.lock:
                    self.active_cluster_calls -= 1

    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "3")
    rows = [
        {
            "数据ID": f"C{index}",
            "工单ID": f"C{index}",
            "聊天内容": f"节点状态核验{index}",
            "产品类型": "手机",
        }
        for index in range(6)
    ]
    client = ParallelClusterMimo()

    _direct_mimo_topic_groups(
        rows,
        client,
        batch_size=2,
        progress_path=tmp_path / "direct_mimo_progress.json",
    )

    assert client.max_active_cluster_calls >= 2


def test_direct_mimo_skips_low_similarity_model_reconciliation(
    monkeypatch,
) -> None:
    class LowSimilarityMimo:
        config = SimpleNamespace(model="mimo-reconcile-floor-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "",
                            "category_l2": "节点状态",
                            "intent": "标准判定",
                            "subject": "节点位置",
                            "phenomenon": "状态异常",
                            "judgment_target": "",
                            "resolution_mode": "按节点状态核验",
                            "standard_path": "节点状态核验",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "第一轮保守保留为单成员主题。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

        def review_cluster_membership(self, *_args, **_kwargs):
            raise AssertionError("低于模型复核门槛时不应调用 MiMo")

    monkeypatch.setattr(
        workflow_module,
        "_direct_reconcile_similarity",
        lambda _left, _right: 0.55,
    )
    monkeypatch.setattr(
        workflow_module,
        "_direct_reconcile_rule_merge_reason",
        lambda _candidate, _target, _similarity: "",
    )
    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": question,
            "产品类型": "手机",
        }
        for record_id, question in (
            ("A", "节点状态甲怎么判"),
            ("B", "节点状态乙怎么判"),
        )
    ]

    groups, meta = _direct_mimo_topic_groups(
        rows,
        LowSimilarityMimo(),
        batch_size=2,
    )

    assert len(groups) == 2
    assert meta["direct_reconcile_calls"] == 0
    assert meta["direct_reconcile_model_floor_skipped"] >= 1


def test_direct_mimo_default_reconcile_limit_caps_model_calls(
    monkeypatch,
) -> None:
    class LimitedReconcileMimo:
        config = SimpleNamespace(model="mimo-reconcile-limit-test")

        def __init__(self) -> None:
            self.review_calls = 0

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "",
                            "category_l2": "节点状态",
                            "intent": "标准判定",
                            "subject": "节点位置",
                            "phenomenon": "状态异常",
                            "judgment_target": "",
                            "resolution_mode": "按节点状态核验",
                            "standard_path": "节点状态核验",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "第一轮保守保留为单成员主题。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

        def review_cluster_membership(self, *_args, **_kwargs):
            self.review_calls += 1
            return MimoLabelResult(
                candidate={
                    "decision": "不同主题",
                    "topic_label": "不同节点状态",
                    "reason": "模拟模型拒绝合并。",
                    "key_difference": "节点不同",
                    "confidence": 0.9,
                },
                request_audit={},
                response_audit={},
            )

    monkeypatch.delenv("ANSWER_HUB_DIRECT_RECONCILE_LIMIT", raising=False)
    monkeypatch.setattr(
        workflow_module,
        "_direct_reconcile_similarity",
        lambda _left, _right: 0.9,
    )
    monkeypatch.setattr(
        workflow_module,
        "_direct_reconcile_rule_merge_reason",
        lambda _candidate, _target, _similarity: "",
    )
    monkeypatch.setattr(
        workflow_module,
        "_direct_reconcile_has_hard_conflict",
        lambda _candidate, _target: False,
    )
    rows = [
        {
            "数据ID": f"L{index:02d}",
            "工单ID": f"L{index:02d}",
            "聊天内容": f"节点状态核验{index:02d}",
            "产品类型": "手机",
        }
        for index in range(25)
    ]
    client = LimitedReconcileMimo()

    _groups, meta = _direct_mimo_topic_groups(
        rows,
        client,
        batch_size=25,
    )

    assert client.review_calls == 24
    assert meta["direct_reconcile_limit"] == 24
    assert meta["direct_reconcile_limit_reached"] == 1


def test_direct_mimo_reconcile_cache_avoids_repeat_model_review(
    tmp_path: Path,
    monkeypatch,
) -> None:
    def topic_payload(row: dict[str, object]) -> dict[str, object]:
        return {
            "normalized_issue": row["聊天内容"],
            "product_category": "手机",
            "scope_type": "品类专用",
            "platform": "通用",
            "brand": "通用",
            "model_scope": "通用",
            "category_l1": "流程操作",
            "category_l2": "节点核验",
            "intent": "流程判断",
            "subject": "节点状态",
            "phenomenon": "状态待确认",
            "judgment_target": "确认节点状态",
            "resolution_mode": "按会话上下文核对",
            "standard_path": "节点核验流程",
            "threshold_or_exception": "无明确阈值",
            "evidence_summary": "完整聊天支持该问题。",
            "confidence": 0.9,
            "requires_review": False,
        }

    class ReconcileCacheMimo:
        config = SimpleNamespace(model="mimo-reconcile-cache-test")

        def __init__(self) -> None:
            self.review_calls = 0

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={"topics": [topic_payload(row)]},
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "节点状态核验",
                            "member_atomic_ids": [
                                units[0]["unit_id"],
                                units[1]["unit_id"],
                            ],
                            "merge_basis": "前两条可以共用同一处理知识。",
                        },
                        {
                            "cluster_id": "C002",
                            "theme_name": "节点状态核验待归并",
                            "member_atomic_ids": [units[2]["unit_id"]],
                            "merge_basis": "第一轮保守保留。",
                        },
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

        def review_cluster_membership(self, *_args, **_kwargs):
            self.review_calls += 1
            return MimoLabelResult(
                candidate={
                    "decision": "同一主题",
                    "topic_label": "节点状态核验",
                    "reason": "三条记录可以共用同一处理知识。",
                    "key_difference": "",
                    "confidence": 0.93,
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"节点状态核验{record_id}",
            "产品类型": "手机",
        }
        for record_id in ("A", "B", "C")
    ]
    monkeypatch.setattr(
        workflow_module,
        "_direct_reconcile_rule_merge_reason",
        lambda _candidate, _target, _similarity: "",
    )
    progress_path = tmp_path / "direct_mimo_progress.json"
    progress_events: list[tuple[str, dict[str, object]]] = []
    first_client = ReconcileCacheMimo()

    first_groups, first_meta = _direct_mimo_topic_groups(
        rows,
        first_client,
        batch_size=3,
        progress_path=progress_path,
        progress_callback=lambda detail, metrics: progress_events.append(
            (detail, dict(metrics))
        ),
    )

    assert len(first_groups) == 1
    assert first_client.review_calls == 1
    assert first_meta["direct_reconcile_cache_hits"] == 0
    assert any("二次归并" in detail for detail, _metrics in progress_events)

    class CachedOnlyMimo(ReconcileCacheMimo):
        def analyze_cluster_units(self, _row):
            raise AssertionError("原子提取应从缓存读取")

        def cluster_atomic_units(self, _units):
            raise AssertionError("首轮聚类应从缓存读取")

        def review_cluster_membership(self, *_args, **_kwargs):
            raise AssertionError("二次复核应从缓存读取")

    second_client = CachedOnlyMimo()
    second_groups, second_meta = _direct_mimo_topic_groups(
        rows,
        second_client,
        batch_size=3,
        progress_path=progress_path,
    )

    assert len(second_groups) == 1
    assert second_meta["direct_reconcile_calls"] == 0
    assert second_meta["direct_reconcile_cache_hits"] == 1


def test_clustering_rule_lookup_is_cached_per_stable_row(
    monkeypatch,
) -> None:
    real_match = workflow_module.match_clustering_judgment_rule
    match_calls = 0

    def counted_match(**kwargs):
        nonlocal match_calls
        match_calls += 1
        return real_match(**kwargs)

    monkeypatch.setattr(
        workflow_module,
        "match_clustering_judgment_rule",
        counted_match,
    )
    left = {
        "产品类型": "手机",
        "核心问题": "自定义节点甲状态怎么判",
        "聊天内容": "咨询自定义节点甲状态。",
        "模型主题一级分类": "流程操作",
        "模型主题二级分类": "节点状态",
        "对象/部位": "自定义节点甲",
        "异常现象": "状态待确认",
        "主标准路径": "节点状态核验",
    }
    right = {
        "产品类型": "手机",
        "核心问题": "自定义节点乙状态怎么判",
        "聊天内容": "咨询自定义节点乙状态。",
        "模型主题一级分类": "流程操作",
        "模型主题二级分类": "节点状态",
        "对象/部位": "自定义节点乙",
        "异常现象": "状态待确认",
        "主标准路径": "节点状态核验",
    }

    for _index in range(5):
        workflow_module._has_topic_merge_conflict(left, right)

    assert match_calls == 2


def test_direct_reconcile_candidate_scan_caches_fingerprint_per_row(
    monkeypatch,
) -> None:
    real_build_fingerprint = workflow_module.build_clustering_fingerprint
    fingerprint_calls = 0

    def counted_build_fingerprint(**kwargs):
        nonlocal fingerprint_calls
        fingerprint_calls += 1
        return real_build_fingerprint(**kwargs)

    monkeypatch.setattr(
        workflow_module,
        "build_clustering_fingerprint",
        counted_build_fingerprint,
    )
    monkeypatch.setattr(
        workflow_module,
        "_direct_reconcile_similarity",
        lambda _left, _right: 0.0,
    )

    groups = [
        (
            ("direct_mimo", "手机", f"candidate-{index}"),
            [
                {
                    "数据ID": f"CANDIDATE-{index:02d}",
                    "工单ID": f"CANDIDATE-{index:02d}",
                    "_原子知识ID": f"CANDIDATE-{index:02d}-U01",
                    "产品类型": "手机",
                    "模型主题一级分类": "流程操作",
                    "问题意图": "信息查询",
                    "对象/部位": "自定义节点",
                    "异常现象": "状态待确认",
                    "核心问题": f"自定义节点{index:02d}状态如何核验",
                    "_聚类裁决提供方": "mimo-direct",
                }
            ],
        )
        for index in range(12)
    ]
    meta: dict[str, object] = {}
    reviewer = SimpleNamespace(config=SimpleNamespace(model="offline-test"))

    reconciled = workflow_module._reconcile_direct_topic_groups(
        groups,
        reviewer,
        meta,
    )

    assert len(reconciled) == len(groups)
    assert fingerprint_calls == len(groups)


def test_direct_mimo_review_request_can_rejoin_compatible_family() -> None:
    class ReviewRequestMimo:
        config = SimpleNamespace(model="mimo-review-request-test")

        def analyze_cluster_units(self, row):
            phenomenon = "碎裂" if row["数据ID"] == "A" else "掉漆"
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": f"手机外壳{phenomenon}如何判定",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "外观问题",
                            "category_l2": f"外壳{phenomenon}",
                            "intent": "标准判定",
                            "subject": "外壳",
                            "phenomenon": phenomenon,
                            "judgment_target": "判断外壳损伤",
                            "resolution_mode": "对照外观标准判定",
                            "standard_path": f"外壳{phenomenon}判定",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "聊天明确咨询外壳损伤。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机外壳外观判定",
                            "member_atomic_ids": [units[0]["unit_id"]],
                            "merge_basis": "第一条主题清晰。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [
                        {
                            "atomic_id": units[1]["unit_id"],
                            "review_type": "标准路径",
                            "reason": "标准路径文字与第一条不同，需要保留复核标记。",
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def review_cluster_membership(self, *_args, **_kwargs):
            raise AssertionError("同一标准族应由本地规则归并，不应额外调用模型")

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": question,
            "核心问题": question,
            "产品类型": "手机",
        }
        for record_id, question in (
            ("A", "手机外壳碎裂怎么判"),
            ("B", "手机后壳掉漆怎么判"),
        )
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=ReviewRequestMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 1
    assert topics[0]["主题样本数"] == 2
    assert not gaps
    assert not pending
    assert clustering_meta["direct_reconcile_rule_approved"] == 1
    assert any(
        row["聚类裁决提供方"] == "mimo-direct-reconcile-rule"
        for row in mapping
    )


def test_direct_mimo_cluster_prompt_change_keeps_atomic_cache_only(
    tmp_path: Path,
    monkeypatch,
) -> None:
    class CacheMimo:
        config = SimpleNamespace(model="mimo-stage-cache-test")

        def __init__(self) -> None:
            self.atomic_calls = 0
            self.cluster_calls = 0

        def analyze_cluster_units(self, row):
            self.atomic_calls += 1
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "信息查询",
                            "category_l2": "设备信息",
                            "intent": "信息查询",
                            "subject": "设备信息",
                            "phenomenon": "查询设备信息",
                            "judgment_target": "确认设备信息",
                            "resolution_mode": "查询并核对",
                            "standard_path": "设备信息查询",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            self.cluster_calls += 1
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "设备信息查询",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "同品类、同查询对象和同处理路径。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "1")
    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"查询设备信息{record_id}",
            "产品类型": "手机",
        }
        for record_id in ("A", "B")
    ]
    progress_path = tmp_path / "direct_mimo_progress.json"
    first = CacheMimo()
    _direct_mimo_topic_groups(
        rows,
        first,
        batch_size=2,
        progress_path=progress_path,
    )
    assert first.atomic_calls == 2
    assert first.cluster_calls == 1

    monkeypatch.setattr(
        workflow_module,
        "ATOMIC_TOPIC_CLUSTER_PROMPT_VERSION",
        "atomic-cluster-test-upgrade",
    )
    second = CacheMimo()
    _direct_mimo_topic_groups(
        rows,
        second,
        batch_size=2,
        progress_path=progress_path,
    )

    assert second.atomic_calls == 0
    assert second.cluster_calls == 1


def test_direct_mimo_model_change_invalidates_model_caches(
    tmp_path: Path,
) -> None:
    class ModelAwareCacheMimo:
        def __init__(self, model: str) -> None:
            self.config = SimpleNamespace(model=model)
            self.atomic_calls = 0
            self.cluster_calls = 0

        def analyze_cluster_units(self, row):
            self.atomic_calls += 1
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "流程操作",
                            "category_l2": "节点核验",
                            "intent": "流程操作",
                            "subject": "节点状态",
                            "phenomenon": "状态待确认",
                            "judgment_target": "确认节点状态",
                            "resolution_mode": "按会话上下文核对",
                            "standard_path": "节点核验流程",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            self.cluster_calls += 1
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "节点状态核验",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "同品类、同对象和同处理目标。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"节点状态核验{record_id}",
            "产品类型": "手机",
        }
        for record_id in ("A", "B")
    ]
    progress_path = tmp_path / "direct_mimo_progress.json"
    first = ModelAwareCacheMimo("model-a")
    _direct_mimo_topic_groups(
        rows,
        first,
        batch_size=2,
        progress_path=progress_path,
    )
    assert first.atomic_calls == 2
    assert first.cluster_calls == 1

    second = ModelAwareCacheMimo("model-b")
    _direct_mimo_topic_groups(
        rows,
        second,
        batch_size=2,
        progress_path=progress_path,
    )

    assert second.atomic_calls == 2
    assert second.cluster_calls == 1


def test_direct_mimo_atomic_content_change_invalidates_cluster_cache(
    tmp_path: Path,
) -> None:
    class ContentAwareCacheMimo:
        config = SimpleNamespace(model="mimo-content-aware-cache-test")

        def __init__(self, issue_suffix: str) -> None:
            self.issue_suffix = issue_suffix
            self.atomic_calls = 0
            self.cluster_calls = 0

        def analyze_cluster_units(self, row):
            self.atomic_calls += 1
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": (
                                f"{row['聊天内容']}-{self.issue_suffix}"
                            ),
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "流程操作",
                            "category_l2": "节点核验",
                            "intent": "流程操作",
                            "subject": "节点状态",
                            "phenomenon": "状态待确认",
                            "judgment_target": "确认节点状态",
                            "resolution_mode": "按会话上下文核对",
                            "standard_path": "节点核验流程",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            self.cluster_calls += 1
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": units[0]["normalized_issue"],
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "同品类节点核验。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"节点状态核验{record_id}",
            "产品类型": "手机",
        }
        for record_id in ("A", "B")
    ]
    progress_path = tmp_path / "direct_mimo_progress.json"
    first = ContentAwareCacheMimo("旧语义")
    _direct_mimo_topic_groups(
        rows,
        first,
        batch_size=2,
        progress_path=progress_path,
    )
    assert first.atomic_calls == 2
    assert first.cluster_calls == 1

    progress = json.loads(progress_path.read_text(encoding="utf-8"))
    progress["atomic_results"] = {}
    progress_path.write_text(
        json.dumps(progress, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    second = ContentAwareCacheMimo("新语义")
    groups, _meta = _direct_mimo_topic_groups(
        rows,
        second,
        batch_size=2,
        progress_path=progress_path,
    )

    assert second.atomic_calls == 2
    assert second.cluster_calls == 1
    assert any(
        "新语义" in group_rows[0]["核心问题"]
        for _key, group_rows in groups
    )


def test_direct_mimo_failed_atomic_result_is_retried_on_resume(
    tmp_path: Path,
) -> None:
    class RetryAtomicFailureMimo:
        config = SimpleNamespace(model="mimo-retry-atomic-failure-test")

        def __init__(self, fail_a: bool) -> None:
            self.fail_a = fail_a
            self.atomic_ids: list[str] = []

        def analyze_cluster_units(self, row):
            self.atomic_ids.append(row["数据ID"])
            if self.fail_a and row["数据ID"] == "A":
                raise ValueError("模拟临时原子提取失败")
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "流程操作",
                            "category_l2": "节点核验",
                            "intent": "流程操作",
                            "subject": "节点状态",
                            "phenomenon": "状态待确认",
                            "judgment_target": "确认节点状态",
                            "resolution_mode": "按会话上下文核对",
                            "standard_path": "节点核验流程",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "分别保留。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"节点状态核验{record_id}",
            "产品类型": "手机",
        }
        for record_id in ("A", "B")
    ]
    progress_path = tmp_path / "direct_mimo_progress.json"
    first = RetryAtomicFailureMimo(fail_a=True)
    _direct_mimo_topic_groups(
        rows,
        first,
        batch_size=2,
        progress_path=progress_path,
    )
    assert first.atomic_ids == ["A", "B"]

    second = RetryAtomicFailureMimo(fail_a=False)
    _direct_mimo_topic_groups(
        rows,
        second,
        batch_size=2,
        progress_path=progress_path,
    )

    assert second.atomic_ids == ["A"]


def test_direct_mimo_failed_cluster_result_is_retried_on_resume(
    tmp_path: Path,
) -> None:
    class RetryClusterFailureMimo:
        config = SimpleNamespace(model="mimo-retry-cluster-failure-test")

        def __init__(self, fail_cluster: bool) -> None:
            self.fail_cluster = fail_cluster
            self.atomic_calls = 0
            self.cluster_calls = 0

        def analyze_cluster_units(self, row):
            self.atomic_calls += 1
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "流程操作",
                            "category_l2": "节点核验",
                            "intent": "流程操作",
                            "subject": "节点状态",
                            "phenomenon": "状态待确认",
                            "judgment_target": "确认节点状态",
                            "resolution_mode": "按会话上下文核对",
                            "standard_path": "节点核验流程",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            self.cluster_calls += 1
            if self.fail_cluster:
                raise MimoError("模拟临时聚类失败")
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "节点状态核验",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "同品类节点核验。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"节点状态核验{record_id}",
            "产品类型": "手机",
        }
        for record_id in ("A", "B")
    ]
    progress_path = tmp_path / "direct_mimo_progress.json"
    first = RetryClusterFailureMimo(fail_cluster=True)
    _direct_mimo_topic_groups(
        rows,
        first,
        batch_size=2,
        progress_path=progress_path,
    )
    assert first.atomic_calls == 2
    assert first.cluster_calls == 1

    second = RetryClusterFailureMimo(fail_cluster=False)
    _direct_mimo_topic_groups(
        rows,
        second,
        batch_size=2,
        progress_path=progress_path,
    )

    assert second.atomic_calls == 0
    assert second.cluster_calls == 1


def test_direct_mimo_cluster_failure_marks_rows_for_priority_review() -> None:
    class FailedClusterMimo:
        config = SimpleNamespace(model="mimo-failed-cluster-review-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "流程操作",
                            "category_l2": "节点核验",
                            "intent": "流程操作",
                            "subject": "节点状态",
                            "phenomenon": "状态待确认",
                            "judgment_target": "确认节点状态",
                            "resolution_mode": "按会话上下文核对",
                            "standard_path": "节点核验流程",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, _units):
            raise MimoError("模拟聚类服务临时失败")

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"节点状态核验{record_id}",
            "产品类型": "手机",
        }
        for record_id in ("A", "B")
    ]

    groups, meta = _direct_mimo_topic_groups(
        rows,
        FailedClusterMimo(),
        batch_size=2,
    )

    assert len(groups) == 2
    assert meta["direct_cluster_failed"] == 1
    assert all(
        group_rows[0]["_聚类需要复核"]
        for _key, group_rows in groups
    )
    assert all(
        "聚类调用失败" in group_rows[0]["人工优先复核原因"]
        for _key, group_rows in groups
    )


def test_cluster_only_workbook_has_one_sheet_and_skips_downstream_models(
    tmp_path: Path,
) -> None:
    class ClusterOnlyMimo:
        config = SimpleNamespace(model="cluster-only-test")

        def analyze_cluster_units(self, _row):
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": "手机｜屏幕｜色斑｜判定",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "显示问题",
                            "category_l2": "色斑",
                            "intent": "标准判定",
                            "subject": "屏幕",
                            "phenomenon": "色斑",
                            "judgment_target": "判断色斑",
                            "resolution_mode": "对照聊天核验",
                            "standard_path": "待确认",
                            "evidence_summary": "聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": True,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机屏幕色斑判定",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "同品类、同对象和同判定目标。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

        def classify_topic_stage(self, *_args, **_kwargs):
            raise AssertionError("cluster-only 不应调用主题价值模型")

        def label_topic(self, *_args, **_kwargs):
            raise AssertionError("cluster-only 不应调用知识转写模型")

        def review_topic(self, *_args, **_kwargs):
            raise AssertionError("cluster-only 不应调用内容初审模型")

    work_order_ids = (
        "002077208618029027906",
        "002077209776885862912",
    )
    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "原始工单ID": record_id,
            "聊天内容": "屏幕色斑怎么判",
            "核心问题": "结合现场图片确认屏幕上的点是否属于色斑",
            "原始核心问题": "旧模型认为屏幕上的点属于坏点",
            "判定结论": "结合图片判定为屏幕色斑",
            "原始判定结论": "旧模型判断为屏幕坏点",
            "产品类型": "手机",
        }
        for record_id in work_order_ids
    ]
    output = tmp_path / "cluster_result.xlsx"

    summary = write_cluster_only_workbook(
        rows,
        output,
        mimo_client=ClusterOnlyMimo(),
        use_mimo=True,
        clustering_mode="direct_mimo",
    )

    assert summary["cluster_only"] is True
    assert summary["topic_model_calls"] == 0
    workbook = load_workbook(output, read_only=False)
    assert workbook.sheetnames == ["聚类结果"]
    worksheet = workbook["聚类结果"]
    header_index = {
        str(cell.value): index
        for index, cell in enumerate(worksheet[1], start=1)
    }
    work_order_cell = worksheet.cell(
        row=2,
        column=header_index["主题工单ID"],
    )
    member_question_cell = worksheet.cell(
        row=2,
        column=header_index["成员核心问题"],
    )
    assert set(str(work_order_cell.value).splitlines()) == set(work_order_ids)
    assert work_order_cell.number_format == "@"
    member_question = str(member_question_cell.value)
    assert "原子问题：手机｜屏幕｜色斑｜判定" in member_question
    assert (
        "上游核心问题：结合现场图片确认屏幕上的点是否属于色斑"
        in member_question
    )
    assert "上游判定结论：结合图片判定为屏幕色斑" in member_question


def test_cluster_only_rejects_modified_work_order_id(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="工单ID在聚类过程中发生变化"):
        write_cluster_only_workbook(
            [
                {
                    "数据ID": "ROW-001",
                    "原始工单ID": "WO-ORIGINAL-001",
                    "工单ID": "WO-MODIFIED-001",
                    "聊天内容": "屏幕色斑怎么判",
                    "核心问题": "屏幕色斑如何判定",
                    "产品类型": "手机",
                }
            ],
            tmp_path / "cluster_result.xlsx",
            use_mimo=False,
            clustering_mode="rule",
        )


def test_direct_mimo_low_confidence_multi_member_cluster_stays_priority_review() -> None:
    class LowConfidenceMimo:
        config = SimpleNamespace(model="mimo-low-confidence-review-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "功能问题",
                            "category_l2": "功能状态",
                            "intent": "标准判定",
                            "subject": "设备功能",
                            "phenomenon": "功能异常",
                            "judgment_target": "判断功能是否异常",
                            "resolution_mode": "按实际功能核验",
                            "standard_path": "设备功能判定",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "聊天证据不够充分。",
                            "confidence": 0.6,
                            "requires_review": False,
                        }
                    ]
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "设备功能判定",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "问题对象和处理目标一致。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": f"手机功能异常{record_id}",
            "产品类型": "手机",
        }
        for record_id in ("A", "B")
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=LowConfidenceMimo(),
        clustering_mode="direct_mimo",
        cluster_only=True,
    )

    assert len(topics) == 1
    assert topics[0]["主题样本数"] == 2
    assert topics[0]["是否重点复核"] == "是"
    assert not gaps
    assert not pending


def test_cluster_admission_allows_clear_high_confidence_singleton_into_topic_stage() -> None:
    class HighConfidenceSingletonMimo:
        config = SimpleNamespace(model="mimo-cluster-admission-singleton-test")

        def __init__(self) -> None:
            self.topic_stage_calls = 0

        def analyze_cluster_units(self, _row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "会话只有一个清晰的屏幕色斑判定问题。",
                    "topics": [
                        {
                            "normalized_issue": "手机｜屏幕｜色斑｜判断是否属于色斑",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "显示问题",
                            "category_l2": "色斑",
                            "intent": "标准判定",
                            "subject": "屏幕",
                            "phenomenon": "色斑",
                            "judgment_target": "判断屏幕异常是否属于色斑",
                            "resolution_mode": "结合聊天证据核验",
                            "standard_path": "屏幕色斑判定",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天只支持屏幕色斑判定。",
                            "confidence": 0.93,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, _units):
            raise AssertionError("清晰单原子问题不需要额外调用聚类模型")

        def classify_topic_stage(self, _topic):
            self.topic_stage_calls += 1
            return MimoLabelResult(
                candidate={
                    "topic_stage": "案例解析",
                    "knowledge_value": "不值得沉淀",
                    "stage_reason": "当前只有单个案例结论。",
                    "value_reason": "缺少可复用边界。",
                    "reusable_knowledge": "保留案例供人工复核。",
                    "confidence": 0.9,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, *_args, **_kwargs):
            raise AssertionError("不值得沉淀主题不应进入知识转写")

        def review_topic(self, *_args, **_kwargs):
            raise AssertionError("未转写主题不应进入内容初审")

    client = HighConfidenceSingletonMimo()
    clustering_meta: dict[str, object] = {}
    topics, mapping, gaps, pending = build_topic_review_rows(
        [
            {
                "数据ID": "ADMISSION-SINGLE-001",
                "工单ID": "ADMISSION-SINGLE-001",
                "聊天内容": "手机屏幕这个点是色斑吗？",
                "核心问题": "手机屏幕色斑如何判定",
                "产品类型": "手机",
            }
        ],
        mimo_client=client,
        clustering_mode="direct_mimo",
        use_standard_references=False,
        enforce_cluster_admission=True,
        clustering_meta=clustering_meta,
    )

    assert client.topic_stage_calls == 1
    assert len(topics) == 1
    assert len(mapping) == 1
    assert not gaps
    assert len(pending) == 1
    assert pending[0]["待聚合状态"] == "incubating_pending_cluster"
    assert pending[0]["聚类准入状态"] == "已自动放行"
    assert pending[0]["聚类准入置信度"] == pytest.approx(0.93)
    assert topics[0]["聚类准入状态"] == "已自动放行"
    assert topics[0]["聚类准入置信度"] == pytest.approx(0.93)
    assert clustering_meta["cluster_admission_admitted_topics"] == 1
    assert clustering_meta["cluster_admission_pending_topics"] == 0


def test_cluster_admission_creates_provisional_singleton_candidates_for_low_confidence_cluster() -> None:
    class LowConfidenceClusterMimo:
        config = SimpleNamespace(model="mimo-cluster-admission-low-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "每条会话各自只有一个问题。",
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "显示问题",
                            "category_l2": "色斑",
                            "intent": "标准判定",
                            "subject": "屏幕",
                            "phenomenon": "色斑",
                            "judgment_target": "判断屏幕异常是否属于色斑",
                            "resolution_mode": "结合聊天证据核验",
                            "standard_path": "屏幕色斑判定",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "聊天支持屏幕色斑问题。",
                            "confidence": 0.94,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机屏幕色斑判定",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "scope_consistent": True,
                            "object_consistent": True,
                            "judgment_target_consistent": True,
                            "standard_path_consistent": True,
                            "threshold_exception_consistent": True,
                            "shared_knowledge_definition": "判断手机屏幕异常是否属于色斑。",
                            "merge_basis": "成员对象和判定目标一致，但合并把握不足。",
                            "confidence": 0.68,
                            "requires_review": False,
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    clustering_meta: dict[str, object] = {}
    topics, mapping, gaps, pending = build_topic_review_rows(
        [
            {
                "数据ID": record_id,
                "工单ID": record_id,
                "聊天内容": f"手机屏幕色斑如何判定{record_id}",
                "核心问题": "手机屏幕色斑如何判定",
                "产品类型": "手机",
            }
            for record_id in ("ADMISSION-LOW-001", "ADMISSION-LOW-002")
        ],
        mimo_client=LowConfidenceClusterMimo(),
        clustering_mode="direct_mimo",
        use_standard_references=False,
        enforce_cluster_admission=True,
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 2
    assert len(mapping) == 2
    assert not gaps
    assert not pending
    assert {topic["主题样本数"] for topic in topics} == {1}
    assert all(topic["主题状态"] == "provisional_singleton_review_pending" for topic in topics)
    assert all(topic["聚类准入状态"] == "暂定单主题候选" for topic in topics)
    assert all(topic["是否重点复核"] == "是" for topic in topics)
    assert all(topic["模型初标重点复核"] == "是" for topic in topics)
    assert all("低于自动放行阈值" in topic["聚类准入原因"] for topic in topics)
    assert clustering_meta["cluster_admission_admitted_topics"] == 0
    assert clustering_meta["cluster_admission_pending_topics"] == 0
    assert clustering_meta["cluster_admission_provisional_topics"] == 1
    assert clustering_meta["cluster_admission_provisional_candidates"] == 2


def test_cluster_admission_creates_provisional_singleton_candidates_when_model_requests_review() -> None:
    class ReviewRequiredClusterMimo:
        config = SimpleNamespace(model="mimo-cluster-admission-review-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "每条会话各自只有一个问题。",
                    "topics": [
                        {
                            "normalized_issue": row["聊天内容"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "显示问题",
                            "category_l2": "色斑",
                            "intent": "标准判定",
                            "subject": "屏幕",
                            "phenomenon": "色斑",
                            "judgment_target": "判断屏幕异常是否属于色斑",
                            "resolution_mode": "结合聊天证据核验",
                            "standard_path": "屏幕色斑判定",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "聊天支持屏幕色斑问题。",
                            "confidence": 0.95,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机屏幕色斑判定",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "scope_consistent": True,
                            "object_consistent": True,
                            "judgment_target_consistent": True,
                            "standard_path_consistent": True,
                            "threshold_exception_consistent": True,
                            "shared_knowledge_definition": "判断手机屏幕异常是否属于色斑。",
                            "merge_basis": "字段看似一致，但模型要求人工复核。",
                            "confidence": 0.95,
                            "requires_review": True,
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    topics, mapping, gaps, pending = build_topic_review_rows(
        [
            {
                "数据ID": record_id,
                "工单ID": record_id,
                "聊天内容": f"手机屏幕色斑如何判定{record_id}",
                "核心问题": "手机屏幕色斑如何判定",
                "产品类型": "手机",
            }
            for record_id in ("ADMISSION-REVIEW-001", "ADMISSION-REVIEW-002")
        ],
        mimo_client=ReviewRequiredClusterMimo(),
        clustering_mode="direct_mimo",
        use_standard_references=False,
        enforce_cluster_admission=True,
    )

    assert len(topics) == 2
    assert len(mapping) == 2
    assert not gaps
    assert not pending
    assert {topic["主题样本数"] for topic in topics} == {1}
    assert all(topic["主题状态"] == "provisional_singleton_review_pending" for topic in topics)
    assert all(topic["聚类准入状态"] == "暂定单主题候选" for topic in topics)
    assert all(topic["是否重点复核"] == "是" for topic in topics)
    assert all("模型要求人工复核" in topic["聚类准入原因"] for topic in topics)


def test_cluster_admission_creates_provisional_candidate_for_direct_mimo_rule_fallback(
    monkeypatch,
) -> None:
    class DownstreamMustNotRun:
        config = SimpleNamespace(model="mimo-cluster-admission-fallback-test")

    def fail_direct_clustering(*_args, **_kwargs):
        raise MimoError("模拟 direct_mimo 整体不可用")

    monkeypatch.setattr(
        workflow_module,
        "_direct_mimo_topic_groups",
        fail_direct_clustering,
    )
    clustering_meta: dict[str, object] = {}
    topics, mapping, gaps, pending = build_topic_review_rows(
        [
            {
                "数据ID": "ADMISSION-FALLBACK-001",
                "工单ID": "ADMISSION-FALLBACK-001",
                "聊天内容": "手机屏幕色斑如何判定",
                "核心问题": "手机屏幕色斑如何判定",
                "产品类型": "手机",
                "问题意图": "标准判定",
                "对象/部位": "屏幕",
                "异常现象": "色斑",
                "解题方式": "结合聊天证据核验",
                "语义标注置信度": 0.95,
            }
        ],
        mimo_client=DownstreamMustNotRun(),
        clustering_mode="direct_mimo",
        use_standard_references=False,
        enforce_cluster_admission=True,
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 1
    assert len(mapping) == 1
    assert not gaps
    assert not pending
    assert topics[0]["主题状态"] == "provisional_singleton_review_pending"
    assert topics[0]["聚类准入状态"] == "暂定单主题候选"
    assert topics[0]["是否重点复核"] == "是"
    assert "规则或其他模式降级" in topics[0]["聚类准入原因"]
    assert clustering_meta["effective_mode"] == "rule"
    assert clustering_meta["cluster_admission_pending_topics"] == 0
    assert clustering_meta["cluster_admission_provisional_topics"] == 1


def test_direct_mimo_deduplicates_repeated_source_rows_before_clustering() -> None:
    class DuplicateSourceMimo:
        config = SimpleNamespace(model="mimo-direct-dedup-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "会话包含一个清晰问题。",
                    "topics": [
                        {
                            "normalized_issue": "手机屏幕色斑如何判定",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "显示问题",
                            "category_l2": "色斑",
                            "intent": "标准判定",
                            "subject": "屏幕",
                            "phenomenon": "色斑",
                            "judgment_target": "判断屏幕色斑",
                            "resolution_mode": "对照标准核验",
                            "standard_path": "屏幕色斑判定标准",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, _units):
            raise AssertionError("重复源样本不应进入批内聚类")

    rows = [
        {
            "数据ID": "A",
            "工单ID": "WO-001",
            "聊天内容": "屏幕色斑怎么判",
            "核心问题": "手机屏幕色斑如何判定",
            "产品类型": "手机",
            "问题意图": "标准判定",
            "对象/部位": "屏幕",
            "异常现象": "色斑",
            "解题方式": "对照标准核验",
        },
        {
            "数据ID": "B",
            "工单ID": "WO-001",
            "聊天内容": "屏幕色斑怎么判",
            "核心问题": "手机屏幕色斑如何判定",
            "产品类型": "手机",
            "问题意图": "标准判定",
            "对象/部位": "屏幕",
            "异常现象": "色斑",
            "解题方式": "对照标准核验",
        },
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=DuplicateSourceMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 1
    assert len(mapping) == 1
    assert not gaps
    assert not pending
    assert clustering_meta["duplicate_source_count"] == 1
    assert clustering_meta["atomic_unit_count"] == 1
    assert clustering_meta["direct_cluster_calls"] == 0


def test_direct_mimo_auto_merges_same_object_same_phenomenon_singletons() -> None:
    class AutoMergeMimo:
        config = SimpleNamespace(model="mimo-direct-auto-merge-test")

        def analyze_cluster_units(self, row):
            if row["数据ID"] == "A":
                subject = "屏幕"
                category_l2 = "漏液"
                normalized_issue = "平板｜屏幕｜点状瑕疵｜判定为漏液"
                resolution_mode = "根据点状瑕疵的直径大小进行判定"
                standard_path = "平板点状瑕疵，整体测量直径，大于1mm算漏液，小于等于1mm算坏点"
                threshold = "直径大于1mm算漏液，小于等于1mm算坏点"
                judgment_target = "判定是否为漏液"
                evidence = "屏幕点状瑕疵，按直径判定。"
            else:
                subject = "屏幕"
                category_l2 = "坏点与漏液"
                normalized_issue = "平板｜屏幕｜点状瑕疵（大于1mm为漏液，小于等于1mm为坏点）｜判定标准"
                resolution_mode = "按尺寸阈值判定"
                standard_path = "整体测量瑕疵直径，大于1mm判定为漏液，小于等于1mm判定为坏点"
                threshold = "大于1mm为漏液，小于等于1mm为坏点"
                judgment_target = "区分漏液与坏点"
                evidence = "屏幕点状瑕疵，按尺寸阈值判定。"
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "会话围绕屏幕点状瑕疵判定。",
                    "topics": [
                        {
                            "normalized_issue": normalized_issue,
                            "product_category": "平板",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "显示问题",
                            "category_l2": category_l2,
                            "intent": "标准判定",
                            "subject": subject,
                            "phenomenon": "点状瑕疵",
                            "judgment_target": judgment_target,
                            "resolution_mode": resolution_mode,
                            "standard_path": standard_path,
                            "threshold_or_exception": threshold,
                            "evidence_summary": evidence,
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "点状瑕疵判定A",
                            "member_atomic_ids": [units[0]["unit_id"]],
                            "merge_basis": "第一轮保守保留为单成员主题。",
                        },
                        {
                            "cluster_id": "C002",
                            "theme_name": "点状瑕疵判定B",
                            "member_atomic_ids": [units[1]["unit_id"]],
                            "merge_basis": "第一轮保守保留为单成员主题。",
                        },
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

        def review_cluster_membership(self, *_args, **_kwargs):
            raise AssertionError("规则自动合并时不应调用大模型复核")

    rows = [
        {
            "数据ID": "A",
            "工单ID": "A",
            "聊天内容": "平板屏幕点状瑕疵大于1mm怎么判",
            "核心问题": "平板屏幕点状瑕疵如何判定",
            "产品类型": "平板",
            "问题意图": "标准判定",
            "对象/部位": "屏幕",
            "异常现象": "点状瑕疵",
            "解题方式": "对照标准核验",
        },
        {
            "数据ID": "B",
            "工单ID": "B",
            "聊天内容": "平板屏幕点状瑕疵怎么按阈值判",
            "核心问题": "平板屏幕点状瑕疵如何判定",
            "产品类型": "平板",
            "问题意图": "标准判定",
            "对象/部位": "屏幕",
            "异常现象": "点状瑕疵",
            "解题方式": "对照标准核验",
        },
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=AutoMergeMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 1
    assert topics[0]["主题样本数"] == 2
    assert not gaps
    assert not pending
    assert clustering_meta["direct_reconcile_approved"] == 1
    assert clustering_meta["direct_reconcile_rule_approved"] == 1
    assert any(
        row["聚类裁决提供方"] == "mimo-direct-reconcile-rule"
        for row in mapping
    )


def test_direct_mimo_post_guard_splits_remerged_multi_topic_atoms() -> None:
    class ReMergeMultiTopicMimo:
        config = SimpleNamespace(model="mimo-direct-post-guard-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "multi_topic",
                    "reason": "会话同时咨询面容识别和摄像头功能。",
                    "topics": [
                        {
                            "normalized_issue": "面容识别实际不能使用怎么判定",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "功能问题",
                            "category_l2": "生物识别功能",
                            "intent": "标准判定",
                            "subject": "面容识别",
                            "phenomenon": "工具正常但实际不能使用",
                            "judgment_target": "判断面容识别是否异常",
                            "resolution_mode": "按实际使用结果判定",
                            "standard_path": "待确认",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "答疑明确面容判异常。",
                            "confidence": 0.9,
                            "requires_review": False,
                        },
                        {
                            "normalized_issue": "摄像头无画面怎么判定",
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "功能问题",
                            "category_l2": "摄像头功能",
                            "intent": "标准判定",
                            "subject": "摄像头",
                            "phenomenon": "能打开相机但无画面",
                            "judgment_target": "判断摄像头是否异常",
                            "resolution_mode": "按实际成像结果判定",
                            "standard_path": "待确认",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "答疑明确摄像头另外判异常。",
                            "confidence": 0.9,
                            "requires_review": False,
                        },
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "实际不能使用按异常判定",
                            "member_atomic_ids": [unit["unit_id"] for unit in units],
                            "merge_basis": "错误地按上位原则合并。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "CASE-001",
            "工单ID": "CASE-001",
            "聊天内容": "面容实际不能用；摄像头无画面。",
            "核心问题": "面容和摄像头怎么判",
            "产品类型": "手机",
        }
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=ReMergeMultiTopicMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 2
    assert sorted(topic["主题样本数"] for topic in topics) == [1, 1]
    assert not gaps
    assert not pending
    assert clustering_meta["direct_post_guard_split_clusters"] == 1
    assert clustering_meta["direct_post_guard_singletons"] == 2
    assert {
        row["聚类裁决提供方"]
        for row in mapping
    } == {"mimo-direct-post-guard"}


def test_direct_cluster_guard_rejects_cross_component_same_standard_family(
    monkeypatch,
) -> None:
    def same_standard_family_match(_row):
        return SimpleNamespace(
            standard_family="手机外壳外观标准",
            merge_policy="same_standard_family",
            phenomenon_value="外观损伤",
        )

    monkeypatch.setattr(
        workflow_module,
        "_direct_clustering_rule_match",
        same_standard_family_match,
    )
    rows = [
        {
            "数据ID": "2084839625209156193",
            "工单ID": "2084839625209156193",
            "产品类型": "手机",
            "模型主题二级分类": "中框及外壳外观",
            "对象/部位": "后置摄像头镜片",
            "异常现象": "磕点",
            "主标准路径": "按磕点直径划分外观成色等级",
        },
        {
            "数据ID": "2086417168668299863",
            "工单ID": "2086417168668299863",
            "产品类型": "手机",
            "模型主题二级分类": "中框及外壳外观",
            "对象/部位": "外壳边缘",
            "异常现象": "线性痕迹（非掉漆）",
            "主标准路径": "中框及外壳外观-外壳划痕（单选）",
        },
    ]

    assert workflow_module._direct_cluster_hard_conflict_reason(rows) == "判定对象不同"


def test_direct_mimo_reconciles_high_similarity_singletons() -> None:
    class DirectReconcileMimo:
        config = SimpleNamespace(model="mimo-direct-reconcile-test")

        def analyze_cluster_units(self, row):
            subject = "电池健康度" if row["数据ID"] == "A" else "电池健康值"
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "会话包含一个清晰问题。",
                    "topics": [
                        {
                            "normalized_issue": f"华为手机｜{subject}｜显示评估中｜判定标准",
                            "product_category": "手机",
                            "scope_type": "品牌专用",
                            "platform": "待确认",
                            "brand": "华为",
                            "model_scope": "通用",
                            "category_l1": "功能问题",
                            "category_l2": "电池健康度",
                            "intent": "标准判定",
                            "subject": subject,
                            "phenomenon": "显示评估中",
                            "judgment_target": "判断电池健康显示评估中时如何处理",
                            "resolution_mode": "提供电池健康显示评估中时的判定标准和处理方法",
                            "standard_path": "待确认",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": f"C{index:03d}",
                            "theme_name": unit["normalized_issue"],
                            "member_atomic_ids": [unit["unit_id"]],
                            "merge_basis": "第一轮保守保留为单成员主题。",
                        }
                        for index, unit in enumerate(units, start=1)
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

        def review_cluster_membership(self, candidate, cluster_members, _similarity, _threshold):
            assert candidate["_原子知识ID"] != cluster_members[0]["_原子知识ID"]
            return MimoLabelResult(
                candidate={
                    "decision": "同一主题",
                    "topic_label": "华为电池健康显示评估中",
                    "reason": "对象、现象、判定目标和处理方式一致，仅表述不同。",
                    "key_difference": "",
                    "confidence": 0.92,
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": question,
            "核心问题": question,
            "产品类型": "手机",
            "问题意图": "标准判定",
            "对象/部位": "电池",
            "异常现象": "显示评估中",
            "解题方式": "对照标准判定",
            "标签聚类键": "旧错误键",
        }
        for record_id, question in (
            ("A", "华为手机电池健康度显示评估中怎么处理"),
            ("B", "华为手机电池健康值显示评估中如何判定"),
        )
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=DirectReconcileMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 1
    assert topics[0]["主题样本数"] == 2
    assert not gaps
    assert not pending
    assert clustering_meta["direct_reconcile_calls"] == 0
    assert clustering_meta["direct_reconcile_approved"] == 1
    assert clustering_meta["direct_reconcile_rule_approved"] == 1
    assert any(
        row["聚类裁决提供方"] == "mimo-direct-reconcile-rule"
        for row in mapping
    )
    assert all(row["标签聚类键"] != "旧错误键" for row in mapping)


def test_direct_mimo_marks_failed_batches_and_skips_reconciliation() -> None:
    class FailedDirectMimo:
        config = SimpleNamespace(model="mimo-direct-failed-test")

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "reason": "会话包含一个清晰问题。",
                    "topics": [
                        {
                            "normalized_issue": row["核心问题"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "功能问题",
                            "category_l2": "电池健康度",
                            "intent": "标准判定",
                            "subject": "电池健康度",
                            "phenomenon": "显示评估中",
                            "judgment_target": "判断显示评估中时如何处理",
                            "resolution_mode": "对照标准判定",
                            "standard_path": "待确认",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, _units):
            raise MimoError("模拟整批聚类失败")

        def review_cluster_membership(self, *_args, **_kwargs):
            raise AssertionError("失败批次不应参与二次自动合并")

    rows = [
        {
            "数据ID": record_id,
            "工单ID": record_id,
            "聊天内容": question,
            "核心问题": question,
            "产品类型": "手机",
            "问题意图": "标准判定",
            "对象/部位": "电池",
            "异常现象": "显示评估中",
            "解题方式": "对照标准判定",
        }
        for record_id, question in (
            ("A", "电池健康度显示评估中怎么处理"),
            ("B", "电池健康值显示评估中如何判定"),
        )
    ]
    clustering_meta: dict[str, object] = {}

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=FailedDirectMimo(),
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
    )

    assert len(topics) == 2
    assert not gaps
    assert not pending
    assert clustering_meta["direct_cluster_failed"] == 1
    assert clustering_meta["direct_cluster_failure_reasons"][0]["reason"] == (
        "MimoError: 模拟整批聚类失败"
    )
    assert clustering_meta["direct_reconcile_calls"] == 0
    assert {
        row["聚类裁决提供方"]
        for row in mapping
    } == {"mimo-direct-failed"}


def test_local_clustering_rule_preclassifies_source_and_blocks_model_boundary_conflict() -> None:
    class WrongPhenomenonMimo:
        config = SimpleNamespace(model="mimo-rule-preclass-conflict-test")

        def __init__(self) -> None:
            self.cluster_inputs: list[list[str]] = []

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "topics": [
                        {
                            "normalized_issue": row["核心问题"],
                            "product_category": "手机",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "显示问题",
                            "category_l2": "屏幕显示",
                            "intent": "标准判定",
                            "subject": "屏幕",
                            # Deliberately return the same phenomenon for both
                            # source records; the source rule must catch this.
                            "phenomenon": "漏液",
                            "judgment_target": "判断屏幕显示异常",
                            "resolution_mode": "按屏幕标准核验",
                            "standard_path": "手机屏幕显示标准",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.95,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            self.cluster_inputs.append([unit["unit_id"] for unit in units])
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "手机屏幕显示异常核验",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "同一批次内的模型标签相同。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "LEAKAGE",
            "工单ID": "LEAKAGE",
            "聊天内容": "手机屏幕漏液如何判定",
            "核心问题": "手机屏幕漏液如何判定",
            "判定结论": "需要按屏幕显示标准核验",
            "产品类型": "手机",
        },
        {
            "数据ID": "COLOR",
            "工单ID": "COLOR",
            "聊天内容": "手机屏幕色斑如何判定",
            "核心问题": "手机屏幕色斑如何判定",
            "判定结论": "需要按屏幕显示标准核验",
            "产品类型": "手机",
        },
    ]
    clustering_meta: dict[str, object] = {}
    reviewer = WrongPhenomenonMimo()

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        mimo_client=reviewer,
        clustering_mode="direct_mimo",
        clustering_meta=clustering_meta,
        cluster_only=True,
    )

    assert len(topics) == 2
    assert not mapping
    assert not gaps
    assert not pending
    assert clustering_meta["clustering_rule_pre_match_count"] == 2
    assert clustering_meta["clustering_rule_model_conflict_count"] == 1
    # The conflict must survive through the cluster guard and prevent a
    # multi-member topic, even if the mock reviewer receives both records.
    assert all(topic["主题样本数"] == 1 for topic in topics)


def test_classification_catalog_candidate_path_separates_uncovered_topics() -> None:
    class CatalogBoundaryMimo:
        config = SimpleNamespace(model="mimo-classification-catalog-boundary-test")

        def __init__(self) -> None:
            self.cluster_inputs: list[list[str]] = []

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "topics": [
                        {
                            "normalized_issue": row["核心问题"],
                            "product_category": "平板电脑",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "基本情况",
                            "category_l2": "待确认",
                            "intent": "信息查询",
                            "subject": "设备信息",
                            "phenomenon": "待确认",
                            "judgment_target": "确认信息",
                            "resolution_mode": "按来源证据核验",
                            "standard_path": "待确认",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "完整聊天支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            self.cluster_inputs.append([unit["unit_id"] for unit in units])
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "平板设备信息查询",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "分类库候选路径相同。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    def item(class_id: str, path: tuple[str, ...], term: str):
        return ClassificationCatalogItem(
            class_id=class_id,
            category="平板电脑",
            category_id=class_id,
            path=path,
            path_str=" > ".join(path),
            leaf_level=3,
            leaf=path[-1],
            upper=path[1:3],
            aliases=(term,),
            keywords=(term,),
            degree="",
            definition="",
            detection_method="",
            is_negative=False,
            search_text=term,
        )

    catalog = (
        item(
            "PB-SN",
            ("平板", "基本情况", "序列号", "查看位置"),
            "序列号",
        ),
        item(
            "PB-SENSOR",
            ("平板", "设备功能情况", "传感器功能", "距离感应"),
            "距离感应",
        ),
    )
    rows = [
        {
            "数据ID": "SN",
            "工单ID": "SN",
            "聊天内容": "平板序列号在哪里查看",
            "核心问题": "平板序列号在哪里查看",
            "产品类型": "平板电脑",
        },
        {
            "数据ID": "SENSOR",
            "工单ID": "SENSOR",
            "聊天内容": "平板距离感应器是否支持",
            "核心问题": "平板距离感应器是否支持",
            "产品类型": "平板电脑",
        },
    ]
    reviewer = CatalogBoundaryMimo()

    groups, meta = _direct_mimo_topic_groups(
        rows,
        reviewer,
        classification_catalog=catalog,
        batch_size=4,
    )

    assert len(groups) == 2
    assert sorted(len(member_rows) for _key, member_rows in groups) == [1, 1]
    assert reviewer.cluster_inputs == []
    assert meta["classification_catalog_enabled"] is True
    assert meta["classification_catalog_match_count"] == 2


def test_classification_catalog_ambiguity_is_reviewable_without_forcing_singleton() -> None:
    class AmbiguousCatalogMimo:
        config = SimpleNamespace(model="mimo-classification-catalog-ambiguous-test")

        def __init__(self) -> None:
            self.cluster_inputs: list[list[str]] = []

        def analyze_cluster_units(self, row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "topics": [
                        {
                            "normalized_issue": row["核心问题"],
                            "product_category": "平板电脑",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "屏幕外观",
                            "category_l2": "灰尘或异物",
                            "intent": "标准判定",
                            "subject": "屏幕内部",
                            "phenomenon": "疑似异物",
                            "judgment_target": "判断是否属于屏幕进灰",
                            "resolution_mode": "补充清晰图片后按标准核验",
                            "standard_path": "待确认",
                            "threshold_or_exception": "无明确阈值",
                            "evidence_summary": "聊天内容支持该问题。",
                            "confidence": 0.9,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            self.cluster_inputs.append([unit["unit_id"] for unit in units])
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "平板屏幕异物是否属于进灰",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "两条案例对象和判定目标一致，但分类库候选存在歧义。",
                            "scope_consistent": True,
                            "object_consistent": True,
                            "judgment_target_consistent": True,
                            "standard_path_consistent": True,
                            "threshold_exception_consistent": True,
                            "shared_knowledge_definition": "平板屏幕内部疑似异物是否按屏幕进灰标准核验。",
                            "confidence": 0.82,
                            "requires_review": False,
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    def item(class_id: str, path: tuple[str, ...], term: str):
        return ClassificationCatalogItem(
            class_id=class_id,
            category="平板电脑",
            category_id=class_id,
            path=path,
            path_str=" > ".join(path),
            leaf_level=3,
            leaf=path[-1],
            upper=path[1:3],
            aliases=(term,),
            keywords=(term,),
            degree="",
            definition="",
            detection_method="",
            is_negative=False,
            search_text=term,
        )

    catalog = (
        item("PB-DUST", ("平板", "屏幕外观情况", "屏幕进灰", "异物"), "异物"),
        item("PB-LEAK", ("平板", "屏幕显示情况", "屏幕漏液", "异物"), "异物"),
    )
    rows = [
        {
            "数据ID": "DUST-A",
            "工单ID": "DUST-A",
            "聊天内容": "平板屏幕里面有异物，是否属于进灰",
            "核心问题": "平板屏幕里面有异物，是否属于进灰",
            "产品类型": "平板电脑",
        },
        {
            "数据ID": "DUST-B",
            "工单ID": "DUST-B",
            "聊天内容": "平板屏幕内部看到小异物，怎么判断",
            "核心问题": "平板屏幕内部看到小异物，怎么判断",
            "产品类型": "平板电脑",
        },
    ]
    reviewer = AmbiguousCatalogMimo()

    groups, _meta = _direct_mimo_topic_groups(
        rows,
        reviewer,
        classification_catalog=catalog,
        batch_size=4,
    )

    assert reviewer.cluster_inputs == [["DUST-A-U1", "DUST-B-U1"]]
    assert len(groups) == 1
    assert len(groups[0][1]) == 2
    assert all(row["_原子需要复核"] for row in groups[0][1])
    assert all("分类库候选存在歧义" in row["人工优先复核原因"] for row in groups[0][1])
    assert all(row["_聚类需要复核"] for row in groups[0][1])


def test_direct_mimo_reports_progress_for_each_atomic_batch(
    monkeypatch,
) -> None:
    class BatchProgressMimo:
        config = SimpleNamespace(model="mimo-batch-progress-test")

        def can_batch_cluster_units(self, _row):
            return True

        def analyze_cluster_units_batch(self, rows):
            return [
                MimoLabelResult(
                    candidate={
                        "topics": [
                            {
                                "normalized_issue": row["核心问题"],
                                "product_category": "手机",
                                "scope_type": "品类专用",
                                "platform": "通用",
                                "brand": "通用",
                                "model_scope": "通用",
                                "category_l1": "显示问题",
                                "category_l2": "色斑",
                                "intent": "标准判定",
                                "subject": "屏幕",
                                "phenomenon": "色斑",
                                "judgment_target": "判断屏幕色斑",
                                "resolution_mode": "对照标准判定",
                                "standard_path": "屏幕显示",
                                "threshold_or_exception": "无",
                                "evidence_summary": "聊天内容支持该问题。",
                                "confidence": 0.9,
                                "requires_review": False,
                            }
                        ]
                    },
                    request_audit={},
                    response_audit={},
                )
                for row in rows
            ]

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {"member_atomic_ids": [unit["unit_id"]]}
                        for unit in units
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    monkeypatch.setenv("ANSWER_HUB_MIMO_MAX_WORKERS", "1")
    monkeypatch.setenv("ANSWER_HUB_DIRECT_ATOMIC_BATCH_SIZE", "2")
    monkeypatch.setenv("ANSWER_HUB_DIRECT_MIMO_BATCH_SIZE", "2")
    monkeypatch.setenv("ANSWER_HUB_DIRECT_PROGRESS_FLUSH_EVERY", "5")
    events: list[tuple[str, dict[str, object]]] = []
    rows = [
        {
            "数据ID": f"PROGRESS-{index}",
            "工单ID": f"PROGRESS-{index}",
            "聊天内容": f"第 {index} 条手机屏幕色斑如何判定",
            "核心问题": "手机屏幕色斑如何判定",
            "产品类型": "手机",
        }
        for index in range(1, 5)
    ]

    _groups, meta = _direct_mimo_topic_groups(
        rows,
        BatchProgressMimo(),
        progress_callback=lambda detail, metrics: events.append(
            (detail, dict(metrics))
        ),
    )

    atomic_counts = [
        event[1]["atomic_extraction_completed"]
        for event in events
        if "atomic_extraction_completed" in event[1]
    ]
    assert atomic_counts == [0, 2, 4]
    assert meta["atomic_extraction_batches_completed"] == 2
    assert meta["atomic_extraction_batches_total"] == 2


def test_cluster_validation_compares_clustering_model_and_human_labels() -> None:
    features, _ = generate_phone_candidate_rows(
        _source_rows(),
        _standards(),
        use_mimo=False,
        image_downloader=_ReadyImageDownloader(),
    )
    first = features[0]
    second = dict(first)
    second["数据ID"] = "PHONE-002"
    second["工单ID"] = "PHONE-002"
    third = dict(first)
    third["数据ID"] = "PHONE-003"
    third["工单ID"] = "PHONE-003"
    third["核心问题"] = "设备机型如何查询"
    third["问题意图"] = "信息查询"
    third["对象/部位"] = "机型"

    rows, summary = build_cluster_validation_rows(
        [first, second, third],
        semantic_threshold=0.8,
        max_pairs=3,
        embedding_client=_FakeEmbedding(),
        mimo_client=_FakeMimo(),
    )

    assert len(rows) == 3
    assert summary["validation_pairs"] == 3
    assert {row["聚类预测"] for row in rows} == {"同一主题", "不同主题"}
    assert all(row["大模型状态"] == "已标注" for row in rows)
    assert all("人工错误类型" in row for row in rows)
    for row in rows:
        row["人工判断"] = row["聚类预测"]
    evaluation = evaluate_cluster_validation_rows(rows)
    assert evaluation["clustering_accuracy"] == 1.0
    assert evaluation["large_model_accuracy"] == 1.0
    assert evaluation["v1_release_ready"] is False


def test_cluster_validation_reuses_mimo_media_signals_from_workbook(tmp_path: Path) -> None:
    class CapturingMimo(_FakeMimo):
        def __init__(self) -> None:
            self.signal_sources: list[dict[str, object]] = []
            self.pair_payloads: list[tuple[dict[str, str], dict[str, str]]] = []

        def analyze_topic_signal(self, source, matches, images):
            self.signal_sources.append(dict(source))
            return super().analyze_topic_signal(source, matches, images)

        def review_cluster_pair(self, left, right, similarity, threshold):
            self.pair_payloads.append((dict(left), dict(right)))
            return super().review_cluster_pair(left, right, similarity, threshold)

    source_path = tmp_path / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "方向二"
    headers = [
        "工单ID",
        "聊天内容",
        "图片链接",
        "视频链接",
        "核心问题",
        "判定结论",
        "判定依据",
        "产品类型",
        "一级分类",
        "二级分类",
        "参考话术",
    ]
    worksheet.append(headers)
    for index in range(3):
        worksheet.append(
            [
                f"PHONE-{index + 1:03d}",
                "屏幕有色斑，请结合现场图片确认。",
                "https://example.com/phone.jpg",
                "https://example.com/phone.mp4",
                "手机屏幕色斑如何判定",
                "待确认",
                "需要结合现场证据",
                "手机",
                "显示问题",
                "色斑",
                "请补充清晰图片",
            ]
        )
    workbook.save(source_path)

    mimo = CapturingMimo()
    rows, summary = cluster_validation_from_workbook(
        source_path,
        product_type="手机",
        semantic_threshold=0.8,
        max_pairs=3,
        embedding_client=_FakeEmbedding(),
        mimo_client=mimo,
        image_downloader=_ReadyImageDownloader(),
    )

    assert len(mimo.signal_sources) == 3
    assert mimo.pair_payloads
    assert summary["conversation_signal_model_enabled"] is True
    assert all(row["记录A_图片证据摘要"] == "图片已接收，仍需人工确认细节。" for row in rows)
    assert all(row["记录A_主题标签"] for row in rows)
    assert all(row["记录A_语义标注依据"] for row in rows)
    assert all(row["记录A_视频链接"] == "https://example.com/phone.mp4" for row in rows)
    assert all(
        row["记录A_视频处理状态"] == "存在视频，当前未解析视频内容（1个）"
        for row in rows
    )
    left_payload, _right_payload = mimo.pair_payloads[0]
    assert left_payload["图片证据摘要"] == "图片已接收，仍需人工确认细节。"
    assert left_payload["视频处理状态"] == "存在视频，当前未解析视频内容（1个）"


def test_cluster_validation_evaluation_tracks_annotation_risks() -> None:
    rows = [
        {
            "聚类预测": "同一主题",
            "大模型判断": "同一主题",
            "人工判断": "不同主题",
        },
        {
            "聚类预测": "不同主题",
            "大模型判断": "同一主题",
            "人工判断": "同一主题",
        },
        {
            "聚类预测": "不同主题",
            "大模型判断": "不同主题",
            "人工判断": "不确定",
        },
        {
            "聚类预测": "不同主题",
            "大模型判断": "不同主题",
            "人工判断": "",
        },
    ]

    evaluation = evaluate_cluster_validation_rows(rows)

    assert evaluation["reviewed_pairs"] == 3
    assert evaluation["pending_pairs"] == 1
    assert evaluation["uncertain_pairs"] == 1
    assert evaluation["decisive_pairs"] == 2
    assert evaluation["clustering_accuracy"] == 0.0
    assert evaluation["large_model_accuracy"] == 0.5
    assert evaluation["false_merge_pairs"] == 1
    assert evaluation["false_merge_rate"] == 1.0
    assert evaluation["false_split_pairs"] == 1
    assert evaluation["false_split_rate"] == 1.0


def test_cluster_validation_releases_v1_at_eighty_percent_with_enough_labels() -> None:
    rows = [
        {
            "聚类预测": "同一主题",
            "大模型判断": "同一主题",
            "人工判断": "同一主题" if index < 16 else "不同主题",
        }
        for index in range(20)
    ]

    evaluation = evaluate_cluster_validation_rows(rows)

    assert evaluation["clustering_accuracy"] == 0.8
    assert evaluation["v1_release_ready"] is True
    assert evaluation["v1_release_status"] == "可上线第一版"


def test_cluster_validation_scales_to_hundreds_without_materializing_all_pairs() -> None:
    class BulkEmbedding:
        config = SimpleNamespace(model="bulk-semantic-test")

        def embed_texts(self, texts, progress_callback=None):
            vectors = []
            theme_vectors = (
                [1.0, 0.0],
                [0.75, 0.6614378],
                [0.75, -0.6614378],
            )
            for index, _text in enumerate(texts):
                vectors.append(theme_vectors[index % len(theme_vectors)])
            if progress_callback:
                progress_callback(len(texts), len(texts))
            return vectors

    rows = [
        {
            "数据ID": f"BULK-{index:04d}",
            "工单ID": f"BULK-{index:04d}",
            "聊天内容": f"第 {index} 条脱敏测试会话",
            "核心问题": f"测试主题 {index % 3}",
            "产品类型": "手机",
            "一级分类": "批量测试",
            "二级分类": f"主题 {index % 3}",
        }
        for index in range(500)
    ]
    progress_events: list[tuple[str, int, int]] = []

    validation_rows, summary = build_cluster_validation_rows(
        rows,
        semantic_threshold=0.8,
        max_pairs=20,
        embedding_client=BulkEmbedding(),
        use_mimo=False,
        progress_callback=lambda stage, completed, total: progress_events.append(
            (stage, completed, total)
        ),
    )

    assert len(validation_rows) == 20
    assert summary["eligible_rows"] == 500
    assert summary["candidate_pairs"] == 124750
    assert {row["聚类预测"] for row in validation_rows} == {"同一主题", "不同主题"}
    assert any(stage == "embedding" for stage, _completed, _total in progress_events)
    assert any(stage == "pair_sampling" for stage, _completed, _total in progress_events)


def test_unavailable_image_without_chat_goes_to_evidence_gap() -> None:
    rows = _source_rows()
    rows[0]["聊天内容"] = ""
    features, _ = generate_phone_candidate_rows(
        rows,
        _standards(),
        mimo_client=_FakeMimo(),
        image_downloader=_FailedImageDownloader(),
    )
    topics, _mapping, gaps, pending = build_topic_review_rows(features, _standards(), use_mimo=False)
    assert not topics
    assert not pending
    assert len(gaps) == 1
    assert "不可用:1" in gaps[0]["图片处理状态"]


def test_mimo_client_retries_invalid_json_once() -> None:
    client = MimoClient(MimoConfig(api_key="test", base_url="https://example.com/v1", model="mimo-v2.5-test"))
    responses = iter(
        [
            {"choices": [{"message": {"content": "not-json"}}]},
            {
                "choices": [
                    {
                        "message": {
                            "content": json.dumps(
                                {
                                    "title": "手机屏幕色斑判定",
                                    "subtitles": ["色斑"],
                                    "content": "1. 按标准核验。",
                                    "category_l1": "显示问题",
                                    "category_l2": "色斑",
                                    "layer": "L2",
                                    "knowledge_form": "具体判定",
                                    "content_type": "核验型",
                                    "standard_refs": ["PHONE-DISPLAY-001"],
                                    "applicable_scope": "手机",
                                    "confidence": 0.9,
                                    "reasoning_summary": "匹配色斑标准。",
                                    "needs_human_review": False,
                                    "image_evidence_summary": "无图片。",
                                },
                                ensure_ascii=False,
                            )
                        }
                    }
                ]
            },
        ]
    )
    client._post = lambda _payload: next(responses)  # type: ignore[method-assign]

    result = client.label(_source_rows()[0], [( _standards()[0], 7.5)], [])
    assert result.request_audit["attempt"] == 2
    assert result.candidate["standard_refs"] == ["PHONE-DISPLAY-001"]
    assert result.candidate["applicable_scope"] == "手机"


def test_cz_rag_master_schema_is_read_as_standard_content(tmp_path: Path) -> None:
    path = tmp_path / "cz-phone-master.json"
    path.write_text(
        json.dumps(
            [
                {
                    "主标题": "设备机型是什么意思",
                    "知识内容": "按实物特征确认设备机型。",
                    "知识分类": "标准定义",
                    "关联标准项": "【基本情况】-【机型】",
                    "适用范围": "通用",
                    "生效状态": "生效中",
                    "来源版本": "SJ-HSYJBZ-2026009",
                    "检索关键词": "设备机型 | 机型怎么确认",
                }
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    item = load_standard_catalog(path)[0]
    assert item.standard_path == "【基本情况】-【机型】"
    assert item.knowledge_type == "标准定义"
    assert item.response_snippet == "按实物特征确认设备机型。"
    assert item.version == "SJ-HSYJBZ-2026009"


def test_ambiguous_repair_trace_becomes_process_not_irrelevant_answer() -> None:
    rows = preprocess_source_rows(
        [
            {
                "工单ID": "REPAIR-001",
                "聊天内容": "这个位置是胶吗？没有看出什么。",
                "核心问题": "设备某个位置疑似胶状物，不确定是否为维修痕迹。",
                "判定结论": "现有图片未识别出明显异常。",
                "判定依据": "证据不足，需结合清晰图片和质检标准确认。",
                "产品类型": "手机",
                "一级分类": "拆修问题",
                "二级分类": "屏幕拆修",
            }
        ]
    )
    weak_standard = StandardCatalogItem(
        standard_id="STD-SCREEN",
        title="屏幕检测方法",
        category_l1="",
        category_l2="",
        knowledge_type="检测方法",
        standard_path="【屏幕】",
        keywords=["屏幕"],
        scope="通用",
        response_snippet="检查屏幕。",
        status="published",
        version="v1",
    )
    candidate = initial_label_rows(rows, [weak_standard])[0]
    assert candidate["模型知识形态"] == "流程方法"
    assert "核验" in candidate["模型主标题"]
    assert candidate["模型关联标准"] == ""
    assert candidate["标准检索状态"] == "未搜索到相关知识（待人工补充）"
    assert candidate["是否重点复核"] == "是"


def test_mimo_cannot_override_uncertainty_process_guardrail() -> None:
    rows = preprocess_source_rows(
        [
            {
                "工单ID": "UNCERTAIN-001",
                "聊天内容": "屏幕上疑似有色斑，不确定是否符合标准。",
                "核心问题": "手机屏幕疑似色斑如何确认",
                "判定结论": "现有图片暂无法确认。",
                "判定依据": "证据不足，需要补拍白屏图片。",
                "产品类型": "手机",
                "一级分类": "显示问题",
                "二级分类": "色斑",
                "参考话术": "请补拍清晰图片。",
            }
        ]
    )
    features, _ = generate_phone_candidate_rows(
        rows,
        _standards(),
        mimo_client=_FakeMimo(),
        image_downloader=_ReadyImageDownloader(),
    )
    second = dict(features[0])
    second["数据ID"] = "UNCERTAIN-002"
    second["工单ID"] = "UNCERTAIN-002"
    second["原始工单ID"] = "UNCERTAIN-002"
    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        [features[0], second],
        _standards(),
        mimo_client=_FakeMimo(),
    )
    candidate = topics[0]
    assert candidate["主题模型提供方"] == "mimo"
    assert candidate["知识分类"] == "质检标准"
    assert candidate["是否重点复核"] == "是"
    assert "强制降级为流程方法" in candidate["校验备注"]


def test_mimo_client_records_usage_latency_and_cost(monkeypatch) -> None:
    class Response:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self) -> bytes:
            return json.dumps(
                {
                    "choices": [{"message": {"content": "{}"}}],
                    "usage": {
                        "prompt_tokens": 1000,
                        "completion_tokens": 500,
                        "total_tokens": 1500,
                    },
                }
            ).encode("utf-8")

    monkeypatch.setattr(mimo_module, "urlopen", lambda *_args, **_kwargs: Response())
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="test-model",
            max_requests_per_second=50,
            input_cost_per_million_tokens=2,
            output_cost_per_million_tokens=4,
        )
    )

    response = client._post({"model": "test-model", "messages": []})
    metrics = client.metrics_snapshot()

    assert response["_answer_hub_metrics"]["attempt"] == 1
    assert metrics["model_calls"] == 1
    assert metrics["model_total_tokens"] == 1500
    assert metrics["model_estimated_cost"] == pytest.approx(0.004)


def test_mimo_client_aborts_slow_streaming_response(monkeypatch) -> None:
    response_closed = threading.Event()

    class SlowStreamingResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self) -> bytes:
            response_closed.wait(1)
            return b'{"choices":[{"message":{"content":"{}"}}]}'

        def close(self) -> None:
            response_closed.set()

    monkeypatch.setattr(mimo_module, "urlopen", lambda *_args, **_kwargs: SlowStreamingResponse())
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="test-model",
            max_retries=0,
            max_requests_per_second=50,
            response_read_timeout_seconds=0.1,
            response_timeout_circuit_breaker_failures=1,
        )
    )
    monkeypatch.setattr(client, "_throttle", lambda: None)

    with pytest.raises(MimoError, match="响应总超时"):
        client._post({"model": "test-model", "messages": []})
    assert response_closed.wait(0.2)
    with pytest.raises(MimoError, match="本次运行已熔断"):
        client._post({"model": "test-model", "messages": []})


def test_mimo_client_allows_one_response_timeout_before_opening_circuit(
    monkeypatch,
) -> None:
    response_closed = threading.Event()
    requests = 0

    class SlowStreamingResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self) -> bytes:
            response_closed.wait(1)
            return b'{"choices":[{"message":{"content":"{}"}}]}'

        def close(self) -> None:
            response_closed.set()

    class FastResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self) -> bytes:
            return b'{"choices":[{"message":{"content":"{}"}}]}'

    def respond(*_args, **_kwargs):
        nonlocal requests
        requests += 1
        return SlowStreamingResponse() if requests == 1 else FastResponse()

    monkeypatch.setattr(mimo_module, "urlopen", respond)
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="test-model",
            max_retries=0,
            max_requests_per_second=50,
            response_read_timeout_seconds=0.1,
            response_timeout_circuit_breaker_failures=2,
        )
    )
    monkeypatch.setattr(client, "_throttle", lambda: None)

    with pytest.raises(MimoError, match="响应总超时"):
        client._post({"model": "test-model", "messages": []})
    assert response_closed.wait(0.2)

    response = client._post({"model": "test-model", "messages": []})

    assert requests == 2
    assert response["choices"][0]["message"]["content"] == "{}"


def test_mimo_client_disables_thinking_and_caps_completion_tokens(
    monkeypatch,
) -> None:
    captured_payload: dict[str, object] = {}

    class Response:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self) -> bytes:
            return b'{"choices":[{"message":{"content":"{}"}}]}'

    def respond(request, **_kwargs):
        captured_payload.update(json.loads(request.data.decode("utf-8")))
        return Response()

    monkeypatch.setattr(mimo_module, "urlopen", respond)
    client = MimoClient(
        MimoConfig(
            api_key="test",
            base_url="https://example.com/v1",
            model="test-model",
            max_retries=0,
            max_requests_per_second=50,
            thinking_type="disabled",
            max_completion_tokens=1536,
        )
    )
    monkeypatch.setattr(client, "_throttle", lambda: None)

    client._post({"model": "test-model", "messages": []})

    assert captured_payload["thinking"] == {"type": "disabled"}
    assert captured_payload["max_completion_tokens"] == 1536


def test_mimo_config_loads_ordered_backup_api_keys(monkeypatch) -> None:
    monkeypatch.setenv("MIMO_API_KEY", "primary-test-key")
    monkeypatch.setenv(
        "MIMO_API_KEYS",
        "backup-test-key-1, primary-test-key; backup-test-key-2",
    )
    monkeypatch.setenv("MIMO_BASE_URL", "https://example.com/v1")
    monkeypatch.setenv("MIMO_MODEL", "test-model")

    config = MimoConfig.from_env()

    assert config is not None
    assert config.all_api_keys() == (
        "primary-test-key",
        "backup-test-key-1",
        "backup-test-key-2",
    )


def test_env_example_defaults_to_cost_efficient_multimodal_mimo() -> None:
    env_example = (
        Path(__file__).resolve().parents[1] / ".env.example"
    ).read_text(encoding="utf-8")
    entries: dict[str, str] = {}
    for raw_line in env_example.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        assert key not in entries, f".env.example 存在重复配置：{key}"
        entries[key] = value

    assert entries["MIMO_API_KEY"] == ""
    assert entries["MIMO_MODEL"] == "mimo-v2.5"
    assert entries["MIMO_MEDIA_MODEL"] == "mimo-v2.5"
    assert entries["MIMO_THINKING_TYPE"] == "disabled"
    assert entries["MIMO_MAX_COMPLETION_TOKENS"] == "2048"
    assert entries["ANSWER_HUB_MIMO_MAX_WORKERS"] == "4"
    assert entries["ANSWER_HUB_DIRECT_CLUSTER_MAX_WORKERS"] == "4"
    assert entries["ANSWER_HUB_DIRECT_ATOMIC_BATCH_SIZE"] == "4"
    assert entries["ANSWER_HUB_DIRECT_ATOMIC_BATCH_MAX_CHARS"] == "16000"
    assert entries["ANSWER_HUB_DIRECT_MIMO_BATCH_SIZE"] == "6"
    assert workflow_module.DEFAULT_DIRECT_MIMO_MAX_WORKERS == 4
    assert workflow_module.DEFAULT_DIRECT_ATOMIC_BATCH_SIZE == 4
    assert workflow_module.DEFAULT_DIRECT_ATOMIC_BATCH_MAX_CHARS == 16000
    assert workflow_module.DEFAULT_DIRECT_MIMO_BATCH_SIZE == 6
    assert entries["MIMO_INPUT_COST_PER_MILLION_TOKENS"] == "0.14"
    assert entries["MIMO_OUTPUT_COST_PER_MILLION_TOKENS"] == "0.28"
    assert "mimo-v2.5-pro" not in {
        entries["MIMO_MODEL"],
        entries["MIMO_MEDIA_MODEL"],
    }


def test_mimo_v25_text_model_defaults_media_requests_to_same_model(
    monkeypatch,
) -> None:
    monkeypatch.setattr(mimo_module, "load_dotenv", lambda: None)
    monkeypatch.setenv("MIMO_API_KEY", "test-key")
    monkeypatch.setenv("MIMO_BASE_URL", "https://api.xiaomimimo.com/v1")
    monkeypatch.setenv("MIMO_MODEL", "mimo-v2.5")
    monkeypatch.delenv("MIMO_MEDIA_MODEL", raising=False)

    config = MimoConfig.from_env()

    assert config is not None
    assert config.model == "mimo-v2.5"
    assert config.media_model == "mimo-v2.5"


def test_mimo_client_switches_key_when_balance_is_exhausted(monkeypatch) -> None:
    authorizations: list[str] = []

    class Response:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self) -> bytes:
            return b'{"choices":[{"message":{"content":"{}"}}]}'

    def respond(request, **_kwargs):
        authorizations.append(request.get_header("Authorization"))
        if len(authorizations) == 1:
            raise HTTPError(
                request.full_url,
                402,
                "Payment Required",
                hdrs=None,
                fp=BytesIO(
                    b'{"error":{"message":"insufficient balance"}}'
                ),
            )
        return Response()

    monkeypatch.setattr(mimo_module, "urlopen", respond)
    client = MimoClient(
        MimoConfig(
            api_key="primary-test-key",
            api_keys=("backup-test-key",),
            base_url="https://example.com/v1",
            model="test-model",
            max_retries=0,
            max_requests_per_second=50,
        )
    )
    monkeypatch.setattr(client, "_throttle", lambda: None)

    response = client._post({"model": "test-model", "messages": []})
    metrics = client.metrics_snapshot()

    assert authorizations == [
        "Bearer primary-test-key",
        "Bearer backup-test-key",
    ]
    assert response["_answer_hub_metrics"]["attempt"] == 2
    assert response["_answer_hub_metrics"]["key_switches"] == 1
    assert metrics["model_key_switches"] == 1


def test_mimo_client_keeps_key_for_plain_rate_limit(monkeypatch) -> None:
    authorizations: list[str] = []

    class Response:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self) -> bytes:
            return b'{"choices":[{"message":{"content":"{}"}}]}'

    def respond(request, **_kwargs):
        authorizations.append(request.get_header("Authorization"))
        if len(authorizations) == 1:
            raise HTTPError(
                request.full_url,
                429,
                "Too Many Requests",
                hdrs=None,
                fp=BytesIO(b'{"error":{"message":"rate limit exceeded"}}'),
            )
        return Response()

    monkeypatch.setattr(mimo_module, "urlopen", respond)
    monkeypatch.setattr(mimo_module.time, "sleep", lambda _seconds: None)
    client = MimoClient(
        MimoConfig(
            api_key="primary-test-key",
            api_keys=("backup-test-key",),
            base_url="https://example.com/v1",
            model="test-model",
            max_retries=1,
            max_requests_per_second=50,
        )
    )
    monkeypatch.setattr(client, "_throttle", lambda: None)

    response = client._post({"model": "test-model", "messages": []})

    assert authorizations == [
        "Bearer primary-test-key",
        "Bearer primary-test-key",
    ]
    assert response["_answer_hub_metrics"]["attempt"] == 2
    assert response["_answer_hub_metrics"]["key_switches"] == 0


def test_rule_topic_clustering_does_not_merge_unrelated_other_appearance_cases() -> None:
    rows = []
    for index, (record_id, core_problem, reply) in enumerate(
        [
            (
                "MIXED-APPEARANCE-001",
                "散热器区域是断裂还是出厂对称设计",
                "该机型此处散热器区域为出厂对称设计，属于正常外观状态。",
            ),
            (
                "MIXED-APPEARANCE-002",
                "笔记本键帽缺失应判什么",
                "设备缺失键帽的情况应判定为需要更换键盘。",
            ),
            (
                "MIXED-APPEARANCE-003",
                "现场提交多张图片等待后台确认设备状态",
                "图片材料是判定关键，请确保图片清晰、全面。",
            ),
        ],
        start=1,
    ):
        rows.append(
            {
                "数据ID": record_id,
                "工单ID": record_id,
                "聊天内容": f"第{index}条完整会话：{core_problem}",
                "核心问题": core_problem,
                "历史实际回复": reply,
                "产品类型": "笔记本",
                "一级分类": "外观问题",
                "二级分类": "其他外观",
                "模型主题一级分类": "外观问题",
                "模型主题二级分类": "其他外观",
                "问题意图": "异常核验",
                "对象/部位": "其他外观",
                "异常现象": "其他外观",
                "解题方式": "现场图片/视频补充与案例证据核验",
                "图片链接": f"https://example.com/mixed-{index}.jpg",
                "图片处理状态": "可用:1",
            }
        )

    topics, mapping, _gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 3
    assert len(mapping) == 3
    assert len(pending) == 1
    assert pending[0]["数据ID"] == "MIXED-APPEARANCE-003"
    topics_by_source = {topic["主题来源记录ID"]: topic for topic in topics}
    assert (
        topics_by_source["MIXED-APPEARANCE-001"]["主题沉淀价值"]
        == "值得沉淀"
    )
    assert (
        topics_by_source["MIXED-APPEARANCE-002"]["主题沉淀价值"]
        == "值得沉淀"
    )
    assert (
        topics_by_source["MIXED-APPEARANCE-003"]["主题沉淀价值"]
        == "不值得沉淀"
    )
    assert {topic["主题样本数"] for topic in topics} == {1}
    assert all("设备外观异常如何通过图片核验" != topic["主标题"] for topic in topics)


def test_rule_topic_title_prefers_core_problem_over_chat_timestamp_excerpt() -> None:
    rows = [
        {
            "数据ID": "TABLET-BATTERY-001",
            "工单ID": "TABLET-BATTERY-001",
            "聊天内容": (
                "26/07/15 10:04:38:38 问题类型：质检问题 "
                "问题描述：平板屏幕漏液如何判定 转人工原因：该问题没有相关知识\n"
                "电池健康度数据无法读取怎么办？"
            ),
            "核心问题": "iPad 电池健康度数据无法读取时应如何判定",
            "历史实际回复": "当电池健康度数据无法读取时，请以设置-通用-关于本机页面显示的电池信息为准。",
            "产品类型": "平板",
            "一级分类": "基本信息问题",
            "二级分类": "电池健康度",
            "问题意图": "异常核验",
            "对象/部位": "电池健康度",
            "异常现象": "数据无法读取",
            "解题方式": "以关于本机页面信息为准",
            "图片链接": "https://example.com/tablet-battery.jpg",
            "图片处理状态": "可用:1",
        }
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert "电池健康度" in topics[0]["主标题"]
    assert "漏液" not in topics[0]["主标题"]
    assert "26/07/15" not in topics[0]["主标题"]
    assert "电池健康度数据无法读取时，请以设置-通用-关于本机" in topics[0]["知识内容"]


def test_rule_topic_boundary_title_uses_natural_issue_terms() -> None:
    rows = [
        {
            "数据ID": "SCREEN-BOUNDARY-001",
            "工单ID": "SCREEN-BOUNDARY-001",
            "聊天内容": (
                "26/07/15 09:53:25:25 问题类型：质检问题 "
                "问题描述：询问手机内屏是否存在漏液情况 转人工原因：回答内容无法理解"
            ),
            "核心问题": (
                "回收师在回收手机时，发现屏幕存在一个异常点，"
                "无法自行准确区分该点属于“坏点”还是“漏液”，因此发起咨询。"
            ),
            "历史实际回复": "请使用菲林卡实测，点位直径大于0.5mm按漏液处理，小于等于0.5mm按坏点处理。",
            "产品类型": "手机",
            "一级分类": "显示问题",
            "二级分类": "漏液",
            "问题意图": "边界判定",
            "对象/部位": "屏幕",
            "异常现象": "坏点/漏液边界",
            "解题方式": "定义与边界条件对照",
            "图片链接": "https://example.com/screen-boundary.jpg",
            "图片处理状态": "可用:1",
        }
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert topics[0]["主标题"] == "手机屏幕坏点和漏液如何区分"
    assert "26/07/15" not in topics[0]["主标题"]
    assert "回收师在回收" not in topics[0]["主标题"]


def test_case_only_rule_fallback_blocks_generic_visual_check_template() -> None:
    rows = [
        {
            "数据ID": "GENERIC-VISUAL-001",
            "工单ID": "GENERIC-VISUAL-001",
            "聊天内容": "屏幕边缘胶条破损，需要看图片确认。",
            "核心问题": "屏幕边缘胶条破损如何核验",
            "历史实际回复": "请补充清晰图片后确认。",
            "产品类型": "手机",
            "一级分类": "外观问题",
            "二级分类": "屏幕及正面外观",
            "问题意图": "痕迹核验",
            "对象/部位": "屏幕",
            "异常现象": "疑似拆修痕迹",
            "解题方式": "现场图片补充与痕迹核验",
            "图片链接": "https://example.com/screen-edge.jpg",
            "图片处理状态": "可用:1",
        },
        {
            "数据ID": "GENERIC-VISUAL-002",
            "工单ID": "GENERIC-VISUAL-002",
            "聊天内容": "折叠屏支架缺口，需要看图片确认。",
            "核心问题": "折叠屏支架缺口如何核验",
            "历史实际回复": "请补充清晰图片后确认。",
            "产品类型": "手机",
            "一级分类": "外观问题",
            "二级分类": "屏幕及正面外观",
            "问题意图": "痕迹核验",
            "对象/部位": "屏幕",
            "异常现象": "疑似拆修痕迹",
            "解题方式": "现场图片补充与痕迹核验",
            "图片链接": "https://example.com/foldable-bracket.jpg",
            "图片处理状态": "可用:1",
        },
    ]

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert {topic["主题转写状态"] for topic in topics} == {"skipped_generic_draft"}
    assert all(topic["主题沉淀价值"] == "待确认" for topic in topics)
    assert all(topic["模型初标结论"] == "未执行" for topic in topics)


def test_topic_transcription_uses_representative_source_facts_and_matching_case_image() -> None:
    captured: dict[str, object] = {}

    class EvidencePackageMimo:
        config = SimpleNamespace(model="mimo-evidence-package-test")

        def classify_topic_stage(self, topic):
            captured["stage_topic"] = topic
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "多个案例描述同一转轴异响判断问题。",
                    "value_reason": "人工判定结论包含可复用的触发情形。",
                    "reusable_knowledge": "闭合瞬间的单次异响也属于转轴异响。",
                    "confidence": 0.91,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, topic, _matches, *, use_standard_references):
            assert use_standard_references is False
            captured["topic"] = topic
            return MimoLabelResult(
                candidate={
                    "title": "笔记本转轴产生的声音如何判定为转轴异响",
                    "subtitles": ["闭合瞬间单次异响如何判定"],
                    "content": (
                        "判断对象：笔记本转轴开合过程中的异常声音。\n"
                        "来源现象：对应案例在闭合瞬间出现单次异响。\n"
                        "处理结论：人工判定结论明确该情形属于转轴异响。\n"
                        "适用边界：当前只覆盖来源案例明确记录的闭合瞬间单次异响，"
                        "其他声音情形仍需补充对应来源证据。"
                    ),
                    "category_l1": "其他问题",
                    "category_l2": "特殊问题",
                    "layer": "L2",
                    "knowledge_form": "具体判定",
                    "standard_refs": [],
                    "applicable_scope": "笔记本-通用",
                    "recommended_reply": (
                        "您好，本案例记录的是笔记本闭合瞬间出现单次异响，"
                        "人工判定结论明确该情形属于转轴异响。其他不同声音或出现时机，"
                        "还需要结合对应案例证据再确认。"
                    ),
                    "confidence": 0.86,
                    "reasoning_summary": "内容只使用主题来源事实，并保留适用边界。",
                    "needs_human_review": True,
                    "image_evidence_summary": "案例图来自人工判定结论对应的来源记录。",
                    "requires_images": True,
                    "image_usage_instruction": "展示对应案例中的转轴部位图片。",
                },
                request_audit={},
                response_audit={},
            )

    rows = []
    for index in range(1, 11):
        record_id = f"HINGE-{index:02d}"
        is_boundary_case = index == 10
        rows.append(
            {
                "数据ID": record_id,
                "工单ID": record_id,
                "聊天内容": (
                    "笔记本开合时有声音，咨询如何判断。"
                    if not is_boundary_case
                    else "请看案例图，笔记本闭合瞬间只响了一次，这种情况怎么判断？"
                ),
                "核心问题": "笔记本转轴产生的声音如何判定",
                "原始核心问题": (
                    "人工确认闭合瞬间单次异响是否属于转轴异响"
                    if is_boundary_case
                    else "人工确认笔记本转轴声音如何判断"
                ),
                "判定结论": (
                    "闭合瞬间的单次异响也属于转轴异响"
                    if is_boundary_case
                    else "需要结合具体出现时机判断"
                ),
                "原始判定结论": (
                    "闭合瞬间的单次异响也属于转轴异响"
                    if is_boundary_case
                    else "需要结合具体出现时机判断"
                ),
                "历史实际回复": (
                    "闭合瞬间的单次异响也属于转轴异响。"
                    if is_boundary_case
                    else "请说明声音出现的具体时机。"
                ),
                "产品类型": "笔记本",
                "一级分类": "其他问题",
                "二级分类": "特殊问题",
                "模型主题一级分类": "其他问题",
                "模型主题二级分类": "特殊问题",
                "问题意图": "边界判定",
                "对象/部位": "转轴",
                "异常现象": "开合异响",
                "解题方式": "结合出现时机判断",
                "图片链接": (
                    "https://example.com/hinge-boundary.jpg"
                    if is_boundary_case
                    else ""
                ),
                "视频链接": (
                    "https://example.com/hinge-boundary.mp4"
                    if is_boundary_case
                    else ""
                ),
                "图片处理状态": "可用:1" if is_boundary_case else "无图片链接（文本初标）",
                "语义标注图片必要性": "需要" if is_boundary_case else "不需要",
                "语义标注依据": (
                    "人工核心问题、判定结论和案例图共同确认闭合瞬间单次异响。"
                    if is_boundary_case
                    else "当前案例只提供一般咨询。"
                ),
            }
        )

    topics, mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=EvidencePackageMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert len(mapping) == 10
    assert not gaps
    assert not pending
    model_topic = captured["topic"]
    evidence_package = model_topic["evidence_package"]
    stage_evidence_package = captured["stage_topic"]["evidence_package"]
    assert evidence_package["fact_count"] == 10
    assert "HINGE-10" in evidence_package["representative_source_ids"]
    boundary_fact = next(
        fact
        for fact in evidence_package["facts"]
        if fact["source_record_id"] == "HINGE-10"
    )
    assert boundary_fact["human_core_problem"] == "人工确认闭合瞬间单次异响是否属于转轴异响"
    assert boundary_fact["human_judgment_conclusion"] == "闭合瞬间的单次异响也属于转轴异响"
    assert boundary_fact["image_urls"] == ["https://example.com/hinge-boundary.jpg"]
    assert all(
        "video_urls" not in fact
        for fact in evidence_package["representative_facts"]
    )
    assert all(
        "video_urls" not in fact
        for fact in stage_evidence_package["representative_facts"]
    )

    topic = topics[0]
    assert "HINGE-10" in topic["主题代表性记录ID"]
    assert "HINGE-10" in topic["主题事实引用"]
    assert topic["图例"] == "https://example.com/hinge-boundary.jpg"
    assert "HINGE-10" in topic["主题图例来源"]
    assert topic["主题视频链接"] == "https://example.com/hinge-boundary.mp4"


def test_topic_evidence_package_json_remains_valid_when_source_text_is_large() -> None:
    long_text = "来源事实" * 2000
    evidence_package = {
        "fact_count": 8,
        "representative_fact_ids": [f"F{index:02d}" for index in range(1, 9)],
        "representative_source_ids": [f"R-{index:03d}" for index in range(1, 9)],
        "representative_facts": [
            {
                "fact_id": f"F{index:02d}",
                "source_record_id": f"R-{index:03d}",
                "conversation_excerpt": long_text,
                "human_judgment_conclusion": long_text,
                "image_urls": [f"https://example.com/case-{index}.jpg"],
                "video_urls": [f"https://example.com/case-{index}.mp4"],
            }
            for index in range(1, 9)
        ],
        "source_fact_refs": [
            f"[F{index:02d}] 来源记录=R-{index:03d}"
            for index in range(1, 9)
        ],
    }

    encoded = workflow_module._topic_evidence_package_json(evidence_package)
    decoded = json.loads(encoded)

    assert len(encoded) <= 30000
    assert decoded["representative_facts"][0]["image_urls"] == [
        "https://example.com/case-1.jpg"
    ]
    assert decoded["representative_facts"][0]["video_urls"] == [
        "https://example.com/case-1.mp4"
    ]


def test_model_review_cannot_approve_an_isolated_short_conclusion() -> None:
    class ShortConclusionMimo:
        config = SimpleNamespace(model="mimo-short-conclusion-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "两个案例讨论同一转轴异响问题。",
                    "value_reason": "人工判定结论一致。",
                    "reusable_knowledge": "闭合瞬间的单次异响也属于转轴异响。",
                    "confidence": 0.9,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, *, use_standard_references):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "title": "笔记本转轴产生的声音如何判定",
                    "subtitles": [],
                    "content": "闭合瞬间的单次异响也属于转轴异响。",
                    "category_l1": "其他问题",
                    "category_l2": "特殊问题",
                    "layer": "L2",
                    "knowledge_form": "具体判定",
                    "standard_refs": [],
                    "applicable_scope": "笔记本-通用",
                    "recommended_reply": "您好，闭合瞬间的单次异响也属于转轴异响。",
                    "confidence": 0.9,
                    "reasoning_summary": "复述人工判定结论。",
                    "needs_human_review": False,
                    "image_evidence_summary": "",
                    "requires_images": False,
                    "image_usage_instruction": "不需要图片。",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(
            self,
            _topic,
            _draft,
            _matches,
            *,
            use_standard_references,
            **_kwargs,
        ):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "模型认为结论与来源一致。",
                    "standard_consistency": "无可信标准",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.95,
                    "priority_review": False,
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": f"SHORT-{index}",
            "工单ID": f"SHORT-{index}",
            "聊天内容": "笔记本闭合瞬间出现单次异响，这种情况是否属于转轴异响？",
            "核心问题": "笔记本闭合瞬间单次异响如何判定",
            "原始核心问题": "人工确认闭合瞬间单次异响是否属于转轴异响",
            "判定结论": "闭合瞬间的单次异响也属于转轴异响",
            "原始判定结论": "闭合瞬间的单次异响也属于转轴异响",
            "产品类型": "笔记本",
            "一级分类": "其他问题",
            "二级分类": "特殊问题",
            "模型主题一级分类": "其他问题",
            "模型主题二级分类": "特殊问题",
            "问题意图": "边界判定",
            "对象/部位": "转轴",
            "异常现象": "闭合瞬间单次异响",
            "解题方式": "结合出现时机判断",
            "语义标注依据": "人工判定结论说明该情形属于转轴异响。",
        }
        for index in (1, 2)
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=ShortConclusionMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["模型初标结论"] == "需修改"
    assert "知识正文过短" in topics[0]["模型初标原因"]
    assert topics[0]["模型初标重点复核"] == "是"


def test_standard_mode_model_review_cannot_override_deterministic_content_gate() -> None:
    standard = StandardCatalogItem(
        standard_id="STD-HINGE-001",
        title="笔记本转轴异响判定",
        category_l1="其他问题",
        category_l2="特殊问题",
        knowledge_type="质检标准",
        standard_path="【笔记本】-【其他问题】-【转轴异响】",
        keywords=["笔记本", "转轴", "异响"],
        scope="笔记本-通用",
        response_snippet="闭合瞬间的单次异响也属于转轴异响。",
        status="published",
        version="v1",
    )
    guarded = workflow_module._apply_topic_initial_review_guard(
        {
            "decision": "通过",
            "knowledge_value": "值得沉淀",
            "error_type": "",
            "reason": "模型认为内容可以通过。",
            "standard_consistency": "一致",
            "evidence_sufficiency": "充分",
            "content_consistency": "一致",
            "title_quality": "清晰",
            "confidence": 0.95,
            "priority_review": False,
        },
        {
            "主标题": "笔记本转轴异响如何判定",
            "知识内容": "闭合瞬间的单次异响也属于转轴异响。",
            "关联标准项": standard.standard_path,
            "主题图片必要性": "无案例图",
            "主题图片链接": "",
            "主题证据等级": "完整会话",
            "是否重点复核": "否",
            "主题无来源内容": "",
        },
        [(standard, 0.95)],
        use_standard_references=True,
    )

    assert guarded["decision"] == "需修改"
    assert "确定性内容门禁不允许覆盖" in guarded["reason"]


def test_model_review_cannot_pass_reply_that_still_switches_to_another_topic() -> None:
    guarded = workflow_module._apply_topic_initial_review_guard(
        {
            "decision": "通过",
            "knowledge_value": "值得沉淀",
            "error_type": "",
            "reason": "模型认为推荐回复与整段历史回复一致。",
            "standard_consistency": "无可信标准",
            "evidence_sufficiency": "充分",
            "content_consistency": "一致",
            "title_quality": "清晰",
            "confidence": 0.95,
            "priority_review": False,
        },
        {
            "主标题": "屏幕显示异常如何通过图片核验",
            "知识内容": (
                "1. 拍摄屏幕全景和异常点近景。\n"
                "2. 排除反光、贴膜和环境光干扰。\n"
                "3. 记录异常位置和复现情况。"
            ),
            "推荐回复": (
                "您好，请拍摄屏幕全景和异常点近景；"
                "针对卡顿：请继续测试系统和摄像头。"
            ),
            "关联标准项": "",
            "主题图片必要性": "无案例图",
            "主题图片链接": "",
            "主题证据等级": "完整会话",
            "是否重点复核": "否",
            "主题无来源内容": "",
        },
        [],
        use_standard_references=False,
    )

    assert guarded["decision"] == "需修改"
    assert guarded["error_type"] == "话术不合适"
    assert "切换到其他主题：卡顿" in guarded["reason"]
    assert "确定性内容门禁不允许覆盖" in guarded["reason"]


def test_recommended_reply_quality_detects_unlabeled_topic_drift_and_hidden_fragments() -> None:
    title = "屏幕显示异常如何通过图片核验"
    content = (
        "1. 拍摄屏幕全景和异常点近景。\n"
        "2. 排除反光、贴膜和环境光干扰。\n"
        "3. 记录异常位置和复现情况。"
    )

    unlabeled_issues = workflow_module._recommended_reply_quality_issues(
        (
            "您好，请拍摄屏幕全景和异常点近景。"
            "小米10是直板机型，没有副屏。"
            "请进入设置查看型号。"
            "该设备系统和摄像头存在卡顿。"
        ),
        title=title,
        content=content,
    )
    overlapping_label_issues = workflow_module._recommended_reply_quality_issues(
        (
            "您好，请拍摄屏幕全景和异常点近景；"
            "针对屏幕卡顿：请继续测试系统和摄像头。"
        ),
        title=title,
        content=content,
    )
    generic_other_label_issues = workflow_module._recommended_reply_quality_issues(
        (
            "您好，请拍摄屏幕全景和异常点近景；"
            "关于其他问题：请进入设置查看型号。"
        ),
        title=title,
        content=content,
    )
    fragment_issues = workflow_module._recommended_reply_quality_issues(
        "您好，请查看后盖上的。",
        title=title,
        content=content,
    )

    assert any(issue.startswith("疑似主题外语句：") for issue in unlabeled_issues)
    assert "切换到其他主题：屏幕卡顿" in overlapping_label_issues
    assert "切换到其他主题：其他问题" in generic_other_label_issues
    assert "句意残缺" in fragment_issues


def test_rule_fallback_organizes_human_judgment_and_keeps_missing_standard_blank() -> None:
    rows = [
        {
            "数据ID": f"RULE-HINGE-{index}",
            "工单ID": f"RULE-HINGE-{index}",
            "聊天内容": "笔记本闭合瞬间出现单次异响，这种情况怎么判断？",
            "核心问题": "笔记本闭合瞬间单次异响如何判定",
            "原始核心问题": "人工确认闭合瞬间单次异响是否属于转轴异响",
            "判定结论": "闭合瞬间的单次异响也属于转轴异响",
            "原始判定结论": "闭合瞬间的单次异响也属于转轴异响",
            "判定依据": "声音出现时机为闭合瞬间，案例中只出现一次。",
            "历史实际回复": "闭合瞬间的单次异响也属于转轴异响。",
            "产品类型": "笔记本",
            "一级分类": "其他问题",
            "二级分类": "特殊问题",
            "模型主题一级分类": "其他问题",
            "模型主题二级分类": "特殊问题",
            "问题意图": "边界判定",
            "对象/部位": "转轴",
            "异常现象": "闭合瞬间单次异响",
            "解题方式": "结合出现时机判断",
            "语义标注依据": "人工判定结论明确该案例属于转轴异响。",
        }
        for index in (1, 2)
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    content = topics[0]["知识内容"]
    assert "1. 人工确认闭合瞬间单次异响是否属于转轴异响" in content
    assert "声音出现时机为闭合瞬间" in content
    assert "闭合瞬间的单次异响也属于转轴异响" in content
    assert "来源未说明的其他情形不得直接套用" in content
    assert all(
        marker not in content
        for marker in ("问题背景：", "判断对象：", "来源核验依据：", "人工处理结论：")
    )
    assert topics[0]["推荐回复"] == ""
    assert topics[0]["模型质量状态"] == "failed"
    assert topics[0]["知识草稿状态"] == "evidence_review_only"
    assert topics[0]["关联标准项"] == ""


def test_specific_model_function_fact_keeps_model_in_applicability_only() -> None:
    class GenericModelLookupMimo:
        config = SimpleNamespace(model="mimo-generic-model-lookup-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "课外常识",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "来源明确询问单一机型的指纹支持情况。",
                    "value_reason": "人工判定结论给出了可复用的机型事实。",
                    "reusable_knowledge": "拯救者 Y7000P 不支持指纹。",
                    "confidence": 0.94,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, *, use_standard_references):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "title": "笔记本是否支持指纹功能",
                    "subtitles": [],
                    "content": (
                        "拯救者 Y7000P 需查询官网配置页面核对是否支持指纹。\n"
                        "信息不足时补充设备标签或系统截图。"
                    ),
                    "category_l1": "信息查询",
                    "category_l2": "指纹支持查询",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": [],
                    "applicable_scope": "笔记本",
                    "applicable_brands": [],
                    "applicable_models": [],
                    "recommended_reply": "您好，请查询官网配置确认该机型是否支持指纹。",
                    "confidence": 0.82,
                    "reasoning_summary": "按机型信息查询流程整理。",
                    "needs_human_review": True,
                    "image_evidence_summary": "",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": f"Y7000P-{index}",
            "工单ID": f"Y7000P-{index}",
            "聊天内容": "拯救者 Y7000P 是否支持指纹？",
            "核心问题": "拯救者 Y7000P 是否支持指纹",
            "原始核心问题": "拯救者 Y7000P 是否支持指纹",
            "判定结论": "拯救者 Y7000P 不支持指纹。",
            "原始判定结论": "拯救者 Y7000P 不支持指纹。",
            "历史实际回复": "拯救者 Y7000P 不支持指纹。",
            "产品类型": "笔记本",
            "机型": "拯救者 Y7000P",
            "一级分类": "信息查询",
            "二级分类": "指纹支持查询",
            "模型主题一级分类": "信息查询",
            "模型主题二级分类": "指纹支持查询",
            "问题意图": "信息查询",
            "对象/部位": "指纹识别",
            "异常现象": "是否支持指纹",
            "解题方式": "直接使用已确认的机型功能结论",
            "_原子机型范围": "拯救者 Y7000P",
            "语义标注依据": "人工判定结论明确该机型不支持指纹。",
        }
        for index in (1, 2)
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=GenericModelLookupMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    topic = topics[0]
    assert topic["主标题"] == "笔记本是否支持指纹功能"
    assert "拯救者 Y7000P" not in topic["主标题"]
    assert topic["适用机型"] == "拯救者 Y7000P", {
        field: topic.get(field)
        for field in (
            "主题转写状态",
            "主题沉淀价值",
            "模型阶段状态",
            "校验备注",
        )
    }
    assert "拯救者 Y7000P" in topic["知识内容"]
    assert "不支持指纹" in topic["知识内容"]
    assert "拯救者 Y7000P" not in topic["推荐回复"]
    assert topic["推荐回复"] == ""
    assert topic["模型质量状态"] == "failed"
    assert topic["知识草稿状态"] == "blocked"


def test_single_switch_lite_case_does_not_expand_to_all_game_consoles() -> None:
    class GenericGameConsoleMimo:
        config = SimpleNamespace(model="mimo-generic-console-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "来源明确记录限定机型与包装盒缺失的处理结论。",
                    "value_reason": "来源包含可复用的机型范围和配件缺失边界。",
                    "reusable_knowledge": "限定款日版机型可回收，包装盒缺失只影响估价。",
                    "confidence": 0.94,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, *, use_standard_references):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "title": "日版游戏机是否在回收范围内及包装盒缺失如何处理",
                    "subtitles": [],
                    "content": (
                        "平台对游戏机回收区分购买渠道，部分日版限定机型可以回收。\n"
                        "包装盒缺失不影响回收资格，估价时按配件缺失处理。"
                    ),
                    "category_l1": "其他问题",
                    "category_l2": "版本与配件",
                    "layer": "L2",
                    "knowledge_form": "具体判定",
                    "standard_refs": [],
                    "applicable_scope": "游戏机",
                    "applicable_brands": [],
                    "applicable_models": [],
                    "recommended_reply": (
                        "您好，日版游戏机需要先确认具体版本；包装盒缺失只影响估价。"
                    ),
                    "confidence": 0.91,
                    "reasoning_summary": "根据当前案例整理。",
                    "needs_human_review": True,
                    "image_evidence_summary": "",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": f"SWITCH-LITE-{index}",
            "工单ID": f"SWITCH-LITE-{index}",
            "聊天内容": "日版任天堂 Switch Lite 宝可梦剑盾限定款没有包装盒，是否可以回收？",
            "核心问题": "日版 Switch Lite 限定款是否可回收及包装盒缺失如何处理",
            "原始核心问题": "日版 Switch Lite 限定款是否可回收及包装盒缺失如何处理",
            "判定结论": (
                "日版 Switch Lite 限定款属于可回收机型；"
                "包装盒缺失不影响回收资格，但估价时按配件缺失扣减。"
            ),
            "原始判定结论": (
                "日版 Switch Lite 限定款属于可回收机型；"
                "包装盒缺失不影响回收资格，但估价时按配件缺失扣减。"
            ),
            "历史实际回复": (
                "日版 Switch Lite 限定款可以回收，包装盒缺失会按配件缺失扣减。"
            ),
            "产品类型": "游戏机",
            "一级分类": "其他问题",
            "二级分类": "版本与配件",
            "模型主题一级分类": "其他问题",
            "模型主题二级分类": "版本与配件",
            "问题意图": "范围判定",
            "对象/部位": "购买版本与包装盒",
            "异常现象": "日版限定款且包装盒缺失",
            "解题方式": "按来源确认机型范围和配件缺失影响",
            "语义标注依据": "人工判定明确限定机型可回收及包装盒缺失处理方式。",
        }
        for index in (1, 2)
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=GenericGameConsoleMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    topic = topics[0]
    assert "Switch Lite" not in topic["主标题"]
    assert "游戏机" in topic["主标题"]
    assert "回收范围" in topic["主标题"]
    assert "包装盒" in topic["主标题"]
    assert "Switch Lite" in topic["适用机型"]
    assert "Switch Lite" not in topic["推荐回复"]
    assert topic["适用机型"] == "Switch Lite"
    assert "日版 Switch Lite 限定款" in topic["知识内容"]
    assert topic["推荐回复"] == ""
    assert topic["模型质量状态"] == "failed"
    assert topic["知识草稿状态"] == "blocked"
    assert "属于可回收机型" in topic["知识内容"]
    assert "平台对游戏机回收" not in topic["知识内容"]


def test_unavailable_case_image_is_not_attached_and_requires_manual_supplement() -> None:
    rows = [
        {
            "数据ID": f"BROKEN-IMAGE-{index}",
            "工单ID": f"BROKEN-IMAGE-{index}",
            "聊天内容": "请看图确认屏幕异常位置。",
            "核心问题": "屏幕异常如何通过案例图核验",
            "原始核心问题": "人工确认屏幕异常位置",
            "判定结论": "需要结合清晰案例图确认",
            "原始判定结论": "需要结合清晰案例图确认",
            "产品类型": "手机",
            "一级分类": "显示问题",
            "二级分类": "屏幕异常",
            "模型主题一级分类": "显示问题",
            "模型主题二级分类": "屏幕异常",
            "问题意图": "检测核验",
            "对象/部位": "屏幕",
            "异常现象": "显示异常",
            "解题方式": "案例图核验",
            "图片链接": "https://example.com/unavailable-case.jpg",
            "图片处理状态": "可用:0；不可用:1；failed:下载失败",
            "语义标注图片必要性": "需要",
            "语义标注依据": "结论依赖案例图，但当前图片下载失败。",
        }
        for index in (1, 2)
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["图例"] == ""
    assert topics[0]["主题图例来源"] == ""
    assert topics[0]["模型初标结论"] == "证据不足待补充"
    assert topics[0]["模型初标错误类型"] == "图片判断失误"


def test_model_cannot_add_threshold_missing_from_source_facts() -> None:
    class InventedThresholdMimo:
        config = SimpleNamespace(model="mimo-invented-threshold-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "多个案例讨论同一转轴异响问题。",
                    "value_reason": "人工结论具备复用价值。",
                    "reusable_knowledge": "闭合瞬间的单次异响属于转轴异响。",
                    "confidence": 0.9,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, *, use_standard_references):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "title": "笔记本转轴异响如何判定",
                    "subtitles": [],
                    "content": (
                        "判断对象：笔记本转轴开合时出现的异常声音。\n"
                        "触发条件：闭合瞬间出现单次异响时需要记录。\n"
                        "新增阈值：声音持续超过3秒时必须判定为严重转轴异响。\n"
                        "处理结论：闭合瞬间的单次异响属于转轴异响。\n"
                        "适用边界：其他声音情形需要补充对应来源事实。"
                    ),
                    "category_l1": "其他问题",
                    "category_l2": "特殊问题",
                    "layer": "L2",
                    "knowledge_form": "具体判定",
                    "standard_refs": [],
                    "applicable_scope": "笔记本-通用",
                    "recommended_reply": (
                        "您好，闭合瞬间的单次异响属于转轴异响；"
                        "声音持续超过3秒时按严重转轴异响处理。"
                    ),
                    "confidence": 0.91,
                    "reasoning_summary": "根据人工结论整理。",
                    "needs_human_review": False,
                    "image_evidence_summary": "",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(
            self,
            _topic,
            _draft,
            _matches,
            *,
            use_standard_references,
            **_kwargs,
        ):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "模型认为正文完整。",
                    "standard_consistency": "无可信标准",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.95,
                    "priority_review": False,
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": f"NO-THRESHOLD-{index}",
            "工单ID": f"NO-THRESHOLD-{index}",
            "聊天内容": "笔记本闭合瞬间只响了一次，这种情况怎么判断？",
            "核心问题": "闭合瞬间单次异响如何判定",
            "原始核心问题": "人工确认闭合瞬间单次异响是否属于转轴异响",
            "判定结论": "闭合瞬间的单次异响也属于转轴异响",
            "原始判定结论": "闭合瞬间的单次异响也属于转轴异响",
            "判定依据": "案例只记录闭合瞬间出现单次异响。",
            "产品类型": "笔记本",
            "一级分类": "其他问题",
            "二级分类": "特殊问题",
            "模型主题一级分类": "其他问题",
            "模型主题二级分类": "特殊问题",
            "问题意图": "边界判定",
            "对象/部位": "转轴",
            "异常现象": "闭合瞬间单次异响",
            "解题方式": "结合出现时机判断",
            "语义标注依据": "人工结论未提供持续时间阈值。",
        }
        for index in (1, 2)
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=InventedThresholdMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["模型初标结论"] == "需修改"
    assert "来源事实不支持" in topics[0]["模型初标原因"]
    assert "3秒" in topics[0]["主题无来源内容"]
    assert topics[0]["主题转写状态"] == "topic_model_quality_failed"
    assert topics[0]["模型调用状态"] == "model_success"
    assert topics[0]["模型输出校验状态"] == "passed"
    assert topics[0]["模型质量状态"] == "failed"
    assert topics[0]["知识草稿状态"] == "blocked"
    assert topics[0]["推荐回复"] == ""


def test_source_claim_guard_does_not_treat_model_annotations_as_fact() -> None:
    unsupported = workflow_module._topic_unsupported_source_claims(
        {
            "content": "声音持续超过3秒时必须判定为严重转轴异响。",
            "recommended_reply": "",
        },
        {
            "facts": [
                {
                    "human_judgment_conclusion": (
                        "闭合瞬间的单次异响属于转轴异响。"
                    ),
                    "semantic_basis": (
                        "声音持续超过3秒时必须判定为严重转轴异响。"
                    ),
                    "threshold_or_exception": "持续超过3秒",
                }
            ]
        },
    )

    assert unsupported == [
        "声音持续超过3秒时必须判定为严重转轴异响"
    ]


def test_topic_reply_rebuilds_when_historical_reply_contains_other_atomic_topics() -> None:
    polluted_reply = (
        "您好，关于“屏幕显示异常如何通过图片核验”，建议你好，小米10是直板机型，没有副屏。"
        "如果你是遇到了其他带副屏的机型（如折叠屏）副屏不亮的情况，请测试点亮副屏，；"
        "关于小型号问题：请进入手机「设置」->「我的设备」->「全部参数与信息」->「认证信息」；"
        "针对卡顿：该设备可正常开机使用，您描述的系统和摄像头卡顿现象，。"
    )

    class PollutedReplyMimo:
        config = SimpleNamespace(model="mimo-topic-reply-purity-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检流程",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "主题询问如何通过图片核验屏幕显示异常。",
                    "value_reason": "来源提供了可复用的拍摄与记录步骤。",
                    "reusable_knowledge": "拍摄屏幕全景和异常点近景并记录显示现象。",
                    "confidence": 0.94,
                    "needs_human_review": False,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(
            self,
            _topic,
            _matches,
            *,
            use_standard_references,
            **_kwargs,
        ):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "title": "屏幕显示异常如何通过图片核验",
                    "subtitles": [],
                    "content": (
                        "核验流程：\n"
                        "1. 先确认当前案例中的屏幕右上角紫色圆斑是否能在白屏画面复现。\n"
                        "2. 拍摄屏幕正面全景和异常点近景，排除反光、贴膜和环境光干扰。\n"
                        "3. 记录异常的颜色、位置、数量、直径或面积及复现情况。\n"
                        "4. 现场无法复现或图片不清晰时，补充证据后再判定。"
                    ),
                    "category_l1": "显示问题",
                    "category_l2": "显示异常",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": [],
                    "applicable_scope": "手机",
                    "recommended_reply": polluted_reply,
                    "confidence": 0.92,
                    "reasoning_summary": "根据当前屏幕显示异常主题整理核验流程。",
                    "needs_human_review": False,
                    "image_evidence_summary": "来源说明了图片拍摄要求，本测试不读取实际图片。",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

        def review_topic(
            self,
            _topic,
            _draft,
            _matches,
            *,
            use_standard_references,
            **_kwargs,
        ):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "decision": "通过",
                    "knowledge_value": "值得沉淀",
                    "error_type": "",
                    "reason": "模型认为推荐回复与来源会话一致。",
                    "standard_consistency": "无可信标准",
                    "evidence_sufficiency": "充分",
                    "content_consistency": "一致",
                    "image_necessity": "不需要",
                    "title_quality": "清晰",
                    "confidence": 0.95,
                    "priority_review": False,
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "DISPLAY-IMAGE-001",
            "工单ID": "DISPLAY-IMAGE-001",
            "聊天内容": (
                "回收师：屏幕右上角出现紫色圆斑，如何通过图片核验？"
                "答疑：先确认右上角紫色圆斑能否在白屏画面复现；"
                "拍摄屏幕正面全景和异常点近景，排除反光、贴膜和环境光干扰；"
                "记录异常的颜色、位置、数量、直径或面积及复现情况；"
                "现场无法复现或图片不清晰时，补充证据后再判定。"
            ),
            "历史实际回复": (
                "屏幕右上角紫色圆斑需在白屏画面复现后拍摄全景和近景。"
                f"{polluted_reply}"
            ),
            "核心问题": "屏幕右上角紫色圆斑如何通过图片核验",
            "原始核心问题": "屏幕右上角紫色圆斑如何通过图片核验",
            "判定结论": "需要在白屏画面复现后拍摄全景和异常点近景。",
            "原始判定结论": "需要在白屏画面复现后拍摄全景和异常点近景。",
            "产品类型": "手机",
            "一级分类": "显示问题",
            "二级分类": "显示异常",
            "模型主题一级分类": "显示问题",
            "模型主题二级分类": "显示异常",
            "问题意图": "检测核验",
            "对象/部位": "屏幕",
            "异常现象": "屏幕右上角紫色圆斑",
            "解题方式": "通过图片核验异常表现",
            "语义标注依据": "当前原子主题只讨论屏幕显示异常的图片核验方法。",
        }
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=PollutedReplyMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    reply = topics[0]["推荐回复"]
    assert reply, {
        field: topics[0].get(field)
        for field in (
            "主题状态",
            "主题沉淀价值",
            "主题价值原因",
            "主题分类状态",
            "主题分类错误",
            "主题转写状态",
        )
    }
    assert "屏幕" in reply
    assert "当前案例" not in reply
    assert "案例证据" not in reply
    assert all(term not in reply for term in ("副屏", "型号", "卡顿"))
    assert "建议你好" not in reply
    assert "，；" not in reply
    assert "，。" not in reply
    assert "推荐回复包含主题外内容，已按当前主题正文重建" in topics[0]["校验备注"]
    assert topics[0]["模型初标重点复核"] == "是"


def test_source_fact_only_trusts_model_threshold_when_raw_source_contains_it() -> None:
    unsupported_fact = workflow_module._topic_source_fact(
        {
            "数据ID": "THRESHOLD-UNSUPPORTED",
            "聊天内容": "闭合瞬间只响了一次。",
            "判定结论": "闭合瞬间单次异响属于转轴异响。",
            "_原子阈值例外": "持续超过3秒",
        },
        1,
    )
    supported_fact = workflow_module._topic_source_fact(
        {
            "数据ID": "THRESHOLD-SUPPORTED",
            "聊天内容": "案例记录声音持续超过3秒。",
            "判定结论": "持续超过3秒时按严重转轴异响处理。",
            "_原子阈值例外": "持续超过3秒",
        },
        2,
    )
    opposite_fact = workflow_module._topic_source_fact(
        {
            "数据ID": "THRESHOLD-OPPOSITE",
            "聊天内容": "案例要求声音持续不超过3秒。",
            "判定结论": "持续不超过3秒时不按严重异响处理。",
            "_原子阈值例外": "持续超过3秒",
        },
        3,
    )

    assert unsupported_fact["threshold_source_supported"] is False
    assert (
        unsupported_fact["source_supported_threshold_or_exception"]
        == ""
    )
    assert supported_fact["threshold_source_supported"] is True
    assert (
        supported_fact["source_supported_threshold_or_exception"]
        == "持续超过3秒"
    )
    assert opposite_fact["threshold_source_supported"] is False
    assert (
        opposite_fact["source_supported_threshold_or_exception"]
        == ""
    )


def test_source_claim_guard_keeps_negation_and_case_facts_separate() -> None:
    evidence_package = {
        "facts": [
            {
                "human_judgment_conclusion": (
                    "闭合瞬间的单次异响不属于转轴异响。"
                ),
            },
            {
                "human_judgment_conclusion": (
                    "声音持续超过3秒时需要记录。"
                ),
            },
            {
                "human_judgment_conclusion": (
                    "严重转轴异响必须单独判定。"
                ),
            },
        ]
    }

    unsupported = workflow_module._topic_unsupported_source_claims(
        {
            "content": (
                "闭合瞬间的单次异响属于转轴异响。\n"
                "声音持续超过3秒时必须判定为严重转轴异响。"
            ),
            "recommended_reply": "",
        },
        evidence_package,
    )

    assert any("属于转轴异响" in claim for claim in unsupported)
    assert any("超过3秒" in claim for claim in unsupported)


def test_source_claim_guard_rejects_unsourced_actions_but_accepts_sourced_rule() -> None:
    evidence_package = {
        "facts": [
            {
                "human_judgment_conclusion": (
                    "声音持续超过3秒时必须判定为严重转轴异响。"
                ),
            }
        ]
    }

    supported = workflow_module._topic_unsupported_source_claims(
        {
            "content": "声音持续超过3秒时必须判定为严重转轴异响。",
            "recommended_reply": "",
        },
        evidence_package,
    )
    unsupported = workflow_module._topic_unsupported_source_claims(
        {
            "content": "需要拆机检查主板后再处理。",
            "recommended_reply": "",
        },
        evidence_package,
    )

    assert supported == []
    assert unsupported == ["需要拆机检查主板后再处理"]


def test_source_claim_guard_rejects_entity_swap_and_plain_invented_fact() -> None:
    evidence_package = {
        "facts": [
            {
                "human_judgment_conclusion": (
                    "屏幕显示异常，建议送修屏幕。"
                ),
            }
        ]
    }

    unsupported = workflow_module._topic_unsupported_source_claims(
        {
            "content": (
                "主板故障。\n"
                "建议维修主板。\n"
                "故障原因是主板进水。"
            ),
            "recommended_reply": "",
        },
        evidence_package,
    )

    assert unsupported == [
        "主板故障",
        "建议维修主板",
        "故障原因是主板进水",
    ]


def test_source_claim_guard_only_exempts_pure_evidence_gap_statements() -> None:
    evidence_package = {
        "facts": [
            {
                "human_judgment_conclusion": (
                    "闭合瞬间单次异响属于转轴异响。"
                ),
            }
        ]
    }

    safe_gap = workflow_module._topic_unsupported_source_claims(
        {
            "content": (
                "来源未说明的其他情形不得直接套用，"
                "需要补充对应来源事实。"
            ),
            "recommended_reply": "",
        },
        evidence_package,
    )
    hidden_invention = workflow_module._topic_unsupported_source_claims(
        {
            "content": "来源未说明，因此必须拆机检查主板。",
            "recommended_reply": "",
        },
        evidence_package,
    )

    assert safe_gap == []
    assert hidden_invention == [
        "来源未说明，因此必须拆机检查主板"
    ]


def test_initial_review_rejects_meta_document_style_title() -> None:
    guarded = workflow_module._apply_topic_initial_review_guard(
        {
            "decision": "通过",
            "knowledge_value": "值得沉淀",
            "error_type": "",
            "reason": "模型认为标题和正文可以通过。",
            "standard_consistency": "无可信标准",
            "evidence_sufficiency": "充分",
            "content_consistency": "一致",
            "image_necessity": "不需要",
            "title_quality": "清晰",
            "confidence": 0.95,
            "priority_review": False,
        },
        {
            "主标题": "相机镜头自带无法拆除转接环的质检记录规则",
            "知识内容": (
                "判断对象：镜头自带且无法拆除的转接环。\n"
                "处理结论：按来源记录核对镜头本体型号。\n"
                "适用边界：其他转接环情形需要补充来源事实。"
            ),
            "关联标准项": "",
            "主题图片必要性": "无案例图",
            "主题图片链接": "",
            "主题证据等级": "完整会话",
            "是否重点复核": "否",
            "主题无来源内容": "",
        },
        [],
        use_standard_references=False,
    )

    assert guarded["decision"] == "需修改"
    assert guarded["error_type"] == "标题不准"
    assert guarded["title_quality"] == "需修改"
    assert "质检记录规则" in guarded["reason"]


def test_topic_candidate_replaces_pipe_segmented_title_with_natural_question() -> None:
    class PipeTitleMimo:
        config = SimpleNamespace(model="mimo-pipe-title-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "来源明确记录屏幕漏液问题。",
                    "value_reason": "来源包含可复用的对象、现象和处理结论。",
                    "reusable_knowledge": "当前案例明确给出屏幕漏液处理结论。",
                    "confidence": 0.92,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, *, use_standard_references):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "title": (
                        "手机｜屏幕｜漏液（显示异常）｜"
                        "坏点与漏液边界判定；手机｜屏幕｜"
                    ),
                    "subtitles": [],
                    "content": (
                        "iPhone 11 Pro 屏幕异常点使用菲林卡实测。\n"
                        "处理结论：点位直径大于 0.5mm 时按漏液处理。\n"
                        "适用边界：小于等于 0.5mm 时按坏点处理。"
                    ),
                    "category_l1": "显示问题",
                    "category_l2": "漏液",
                    "layer": "L2",
                    "knowledge_form": "具体判定",
                    "standard_refs": [],
                    "applicable_scope": "手机",
                    "applicable_brands": ["Apple"],
                    "applicable_models": ["iPhone 11 Pro"],
                    "recommended_reply": (
                        "您好，iPhone 11 Pro 屏幕异常点需使用菲林卡实测；"
                        "点位直径大于 0.5mm 时按漏液处理。"
                    ),
                    "confidence": 0.91,
                    "reasoning_summary": "只使用当前来源事实。",
                    "needs_human_review": True,
                    "image_evidence_summary": "",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "PIPE-TITLE-001",
            "工单ID": "PIPE-TITLE-001",
            "聊天内容": (
                "回收师咨询 iPhone 11 Pro 屏幕坏点和漏液如何区分。"
                "使用菲林卡实测，点位直径大于 0.5mm 按漏液处理，"
                "小于等于 0.5mm 按坏点处理。"
            ),
            "核心问题": (
                "手机｜屏幕｜漏液（显示异常）｜"
                "判定是否符合回收标准；手机｜屏幕｜"
            ),
            "原始核心问题": "iPhone 11 Pro 屏幕漏液是否可以回收",
            "判定结论": "点位直径大于 0.5mm 按漏液处理，小于等于 0.5mm 按坏点处理。",
            "原始判定结论": "点位直径大于 0.5mm 按漏液处理，小于等于 0.5mm 按坏点处理。",
            "判定依据": "使用菲林卡实测点位直径。",
            "历史实际回复": "iPhone 11 Pro 屏幕异常点大于 0.5mm 按漏液处理。",
            "产品类型": "手机",
            "机型": "iPhone 11 Pro",
            "一级分类": "显示问题",
            "二级分类": "漏液",
            "模型主题一级分类": "显示问题",
            "模型主题二级分类": "漏液",
            "问题意图": "边界判定",
            "对象/部位": "屏幕",
            "异常现象": "坏点与漏液边界",
            "解题方式": "使用菲林卡测量点位直径",
            "语义标注依据": "来源会话和人工结论均明确给出 0.5mm 边界。",
        }
    ]
    rows.append(
        {
            **rows[0],
            "数据ID": "PIPE-TITLE-002",
            "工单ID": "PIPE-TITLE-002",
        }
    )

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=PipeTitleMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["主标题"] == "手机屏幕漏液是否可以回收", {
        field: topics[0].get(field)
        for field in (
            "主题转写状态",
            "主题沉淀价值",
            "模型阶段状态",
            "校验备注",
        )
    }
    assert "｜" not in topics[0]["主标题"]
    assert "|" not in topics[0]["主标题"]
    assert "iPhone 11 Pro" not in topics[0]["主标题"]


def test_topic_candidate_rebuilds_internal_tag_or_statement_title_as_natural_question() -> None:
    class InternalTagTitleMimo:
        config = SimpleNamespace(model="mimo-internal-tag-title-test")

        def classify_topic_stage(self, _topic):
            return MimoLabelResult(
                candidate={
                    "topic_stage": "质检标准",
                    "knowledge_value": "值得沉淀",
                    "stage_reason": "来源明确记录充电部件外观异常的判定问题。",
                    "value_reason": "来源包含可复用的对象、现象和处理方式。",
                    "reusable_knowledge": "需要按现有证据判定充电部件外观异常。",
                    "confidence": 0.92,
                    "needs_human_review": True,
                },
                request_audit={},
                response_audit={},
            )

        def label_topic(self, _topic, _matches, *, use_standard_references):
            assert use_standard_references is False
            return MimoLabelResult(
                candidate={
                    "title": "手机充电部件外观异常判定",
                    "subtitles": [],
                    "content": (
                        "判定规则：根据当前案例中可核实的充电部件外观证据判断。\n"
                        "处理步骤：确认异常部位并补充清晰近景。\n"
                        "例外与边界：证据不足时不能直接套用，需要转人工审核。"
                    ),
                    "category_l1": "外观问题",
                    "category_l2": "充电部件",
                    "layer": "L2",
                    "knowledge_form": "流程方法",
                    "standard_refs": [],
                    "applicable_scope": "手机",
                    "applicable_brands": [],
                    "applicable_models": [],
                    "recommended_reply": "您好，请先补充充电部件外观近景，再按现有证据核验。",
                    "confidence": 0.91,
                    "reasoning_summary": "只使用当前来源事实。",
                    "needs_human_review": True,
                    "image_evidence_summary": "",
                    "requires_images": False,
                    "image_usage_instruction": "",
                },
                request_audit={},
                response_audit={},
            )

    rows = [
        {
            "数据ID": "INTERNAL-TAG-TITLE-001",
            "工单ID": "INTERNAL-TAG-TITLE-001",
            "聊天内容": "咨询手机充电部件外观异常应如何判定。",
            "核心问题": (
                "边界判定 | 充电部件 | 外观异常 | 意图：边界判定 | "
                "对象：充电部件 | 现象：外观异常 | 处理：定义与边界条件对照 | "
                "标准：QC-2D976D3C5294"
            ),
            "原始核心问题": "",
            "人工核心问题": "",
            "判定结论": "需要根据清晰的充电部件外观证据再判断。",
            "判定依据": "当前案例图片和人工复核记录。",
            "历史实际回复": "请补充充电部件外观近景后再核验。",
            "产品类型": "手机",
            "一级分类": "外观问题",
            "二级分类": "充电部件",
            "模型主题一级分类": "外观问题",
            "模型主题二级分类": "充电部件",
            "问题意图": "边界判定",
            "对象/部位": "充电部件",
            "异常现象": "外观异常",
            "解题方式": "定义与边界条件对照",
            "语义标注依据": "当前案例需核验充电部件外观异常。",
        }
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        mimo_client=InternalTagTitleMimo(),
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    title = topics[0]["主标题"]
    assert "手机" in title
    assert "充电部件" in title
    assert "外观异常" in title
    assert "如何" in title
    for marker in (
        "|", "意图:", "对象:", "现象:", "处理:", "标准:QC-",
        "意图：", "对象：", "现象：", "处理：", "标准：QC-",
    ):
        assert marker not in title


def test_untranscribed_topic_title_skips_internal_cluster_tags() -> None:
    internal_tag_title = (
        "边界判定 | 充电部件 | 外观异常 | 意图：边界判定 | "
        "对象：充电部件 | 现象：外观异常 | 处理：定义与边界条件对照 | "
        "标准：QC-2D976D3C5294"
    )

    title = workflow_module._untranscribed_topic_title(
        {
            "产品类型": "手机",
            "对象/部位": "充电部件",
            "异常现象": "外观异常",
        },
        [
            {
                "产品类型": "手机",
                "_聚类主题标题": internal_tag_title,
                "核心问题": internal_tag_title,
                "聊天内容": "",
            }
        ],
    )

    assert title == "手机充电部件外观异常如何判定"
    for marker in (
        "|", "意图:", "对象:", "现象:", "处理:", "标准:QC-",
        "意图：", "对象：", "现象：", "处理：", "标准：QC-",
    ):
        assert marker not in title


def test_cluster_only_title_rebuilds_when_model_label_belongs_to_another_topic() -> None:
    row = {
        "数据ID": "TITLE-GUARD-001",
        "工单ID": "TITLE-GUARD-001",
        "产品类型": "平板电脑",
        "核心问题": "平板电池健康度和后置摄像头区域灰尘分别如何判定",
        "对象/部位": "后置摄像头区域",
        "异常现象": "灰尘颗粒",
        "判定目标": "判断是否属于屏幕进灰",
        "解题方式": "补充屏幕显示区域证据后核验",
        "_聚类主题标题": "电池循环次数太少，电池容量又低，要算异常吗",
        "_聚类决策": "纯大模型1-N聚类",
        "_聚类裁决提供方": "mimo-direct",
        "_聚类裁决原因": "模型输出。",
    }
    topic = workflow_module._cluster_only_topic_row(
        "TOP-TITLE-GUARD",
        ("direct_mimo", "平板电脑"),
        [row],
    )

    assert "电池循环次数" not in topic["聚类主题"]
    assert "后置摄像头区域" in topic["聚类主题"]
    assert "如何" in topic["聚类主题"]


def test_cluster_only_title_uses_atomic_fields_for_multi_target_source() -> None:
    row = {
        "数据ID": "TITLE-GUARD-002",
        "工单ID": "TITLE-GUARD-002",
        "产品类型": "平板电脑",
        "核心问题": (
            "回收师对两个问题有疑问：1. 平板屏幕漏液如何判定；"
            "2. 后置摄像头镜片区域缝隙如何归类"
        ),
        "对象/部位": "后置摄像头镜片区域",
        "异常现象": "缝隙",
        "判定目标": "判断缝隙是否达到外壳缝隙标准",
        "_聚类主题标题": "平板屏幕漏液如何判定",
        "_聚类决策": "纯大模型1-N聚类",
        "_聚类裁决提供方": "mimo-direct",
        "_聚类裁决原因": "模型输出。",
    }
    topic = workflow_module._cluster_only_topic_row(
        "TOP-TITLE-GUARD-2",
        ("direct_mimo", "平板电脑"),
        [row],
    )

    assert "屏幕漏液" not in topic["聚类主题"]
    assert "后置摄像头镜片区域" in topic["聚类主题"]
    assert "缝隙" in topic["聚类主题"]


def test_cluster_only_title_rebuilds_when_atomic_core_question_is_from_other_target() -> None:
    row = {
        "数据ID": "TITLE-GUARD-003",
        "工单ID": "TITLE-GUARD-003",
        "产品类型": "平板电脑",
        "核心问题": "平板屏幕漏液如何判定",
        "对象/部位": "摄像头镜片区域外壳",
        "异常现象": "缝隙（脱胶）",
        "判定目标": "判断缝隙是否达到外壳缝隙标准",
        "_聚类主题标题": "",
        "_聚类决策": "纯大模型1-N聚类",
        "_聚类裁决提供方": "mimo-direct",
        "_聚类裁决原因": "原子问题字段冲突。",
    }
    topic = workflow_module._cluster_only_topic_row(
        "TOP-TITLE-GUARD-3",
        ("direct_mimo", "平板电脑"),
        [row],
    )

    assert "屏幕漏液" not in topic["聚类主题"]
    assert "摄像头镜片区域外壳" in topic["聚类主题"]
    assert "缝隙" in topic["聚类主题"]


def test_rule_title_uses_structured_atomic_question_before_case_narrative() -> None:
    rows = [
        {
            "数据ID": f"SWITCH-TITLE-{index}",
            "工单ID": f"SWITCH-TITLE-{index}",
            "聊天内容": (
                "回收师咨询日版 Switch Lite 宝可梦剑盾限定款是否可回收，"
                "并说明包装盒缺失。"
            ),
            "核心问题": (
                "游戏机｜日版Switch Lite限定款｜"
                "日版机型是否在回收范围内及包装盒缺失是否影响回收判定｜"
                "确认回收资格及配件缺失处理方式"
            ),
            "原始核心问题": (
                "游戏机｜日版Switch Lite限定款｜"
                "日版机型是否在回收范围内及包装盒缺失是否影响回收判定｜"
                "确认回收资格及配件缺失处理方式"
            ),
            "人工核心问题": (
                "回收师在回收“任天堂 Switch Lite 宝可梦剑盾限定款”时，"
                "对日版机型是否在回收范围内以及包装盒缺失是否影响回收判定存在疑问，"
                "因系统无相关知识转入人工咨询。"
            ),
            "判定结论": (
                "该机为日版游戏机，属于可回收机型，包装盒缺失不影响回收判定，"
                "但在最终估价时应按配件缺失处理。"
            ),
            "原始判定结论": (
                "该机为日版游戏机，属于可回收机型，包装盒缺失不影响回收判定，"
                "但在最终估价时应按配件缺失处理。"
            ),
            "历史实际回复": (
                "日版 Switch Lite 限定款可以回收，包装盒缺失按配件缺失扣减。"
            ),
            "产品类型": "游戏机",
            "一级分类": "其他问题",
            "二级分类": "版本与配件",
            "模型主题一级分类": "其他问题",
            "模型主题二级分类": "版本与配件",
            "问题意图": "标准判定",
            "对象/部位": "整机",
            "异常现象": "日版机型回收资格咨询及包装盒缺失影响咨询",
            "解题方式": "确认回收资格及配件缺失处理方式",
            "语义标注依据": "人工判定结论明确日版限定款的回收资格与包装盒缺失处理。",
        }
        for index in (1, 2)
    ]

    topics, _mapping, gaps, pending = build_topic_review_rows(
        rows,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=False,
    )

    assert len(topics) == 1
    assert not gaps
    assert not pending
    assert topics[0]["主标题"] == (
        "游戏机日版限定款是否可回收，包装盒缺失如何处理"
    )


@pytest.mark.parametrize(
    ("core_problem", "category_l1", "category_l2", "standard_path", "expected_title"),
    (
        (
            "帮忙看一下图片",
            "显示问题",
            "色斑",
            "本地质检标准：【手机】【显示问题】【1】",
            "手机屏幕色斑如何通过图片核验",
        ),
        (
            "指南针怎么核验",
            "功能问题",
            "指南针",
            "本地质检标准：【手机】【功能问题】【指南针】",
            "手机指南针功能如何核验",
        ),
        (
            "设备机型如何查询与确认",
            "基本情况",
            "机型",
            "本地质检标准：【手机】【基本情况】【0】",
            "手机设备机型如何查询与确认",
        ),
    ),
)
def test_rule_process_title_uses_product_and_rejects_numeric_standard_leaf(
    core_problem: str,
    category_l1: str,
    category_l2: str,
    standard_path: str,
    expected_title: str,
) -> None:
    standard = StandardCatalogItem(
        standard_id="LOCAL-TITLE-001",
        title="1",
        category_l1=category_l1,
        category_l2=category_l2,
        knowledge_type="本地质检标准",
        standard_path=standard_path,
        keywords=[category_l2],
        scope="手机",
        response_snippet="按当前标准逐项核验。",
        status="published",
        version="local-title-test-v1",
    )

    assert workflow_module._process_title(
        core_problem,
        category_l1,
        category_l2,
        standard,
        product_type="手机",
    ) == expected_title


def test_guess_title_rejects_generic_image_request() -> None:
    assert workflow_module._guess_title("帮忙看一下图片") == ""


def test_guess_title_removes_generic_review_tail_and_normalizes_lid_closure() -> None:
    assert workflow_module._guess_title(
        "拯救者R7000P合盖合不上看一下"
    ) == "拯救者R7000P合盖无法闭合如何处理"


@pytest.mark.parametrize(
    ("structured_question", "expected_title"),
    (
        (
            "手机｜后置摄像头镜片玻璃｜磕点｜根据直径尺寸判定外观成色等级",
            "后置摄像头镜片玻璃磕点如何按直径判定",
        ),
        (
            "游戏机｜日版Switch Lite限定款｜"
            "日版机型是否在回收范围内及包装盒缺失是否影响回收判定｜"
            "确认回收资格及配件缺失处理方式",
            "日版 Switch Lite 限定款是否可回收，包装盒缺失如何处理",
        ),
    ),
)
def test_structured_atomic_question_becomes_natural_search_title(
    structured_question: str,
    expected_title: str,
) -> None:
    assert workflow_module._natural_title_from_structured_atomic_question(
        structured_question
    ) == expected_title


def test_source_title_skips_case_narrative_before_structured_atomic_question() -> None:
    title = workflow_module._natural_topic_title_from_source(
        {
            "产品类型": "手机",
            "对象/部位": "屏幕",
            "异常现象": "漏液",
        },
        [
            {
                "产品类型": "手机",
                "原始核心问题": (
                    "回收师咨询屏幕漏液严重的iPhone 11 Pro设备是否符合回收标准"
                ),
                "核心问题": "手机｜屏幕｜漏液（显示异常）｜判定是否符合回收标准",
            }
        ],
    )

    assert "回收师" not in title
    assert "屏幕" in title
    assert "漏液" in title
    assert not title.endswith("符合回")


def test_candidate_validation_retries_meta_document_style_title() -> None:
    with pytest.raises(MimoError, match="公文式"):
        mimo_module._validate_candidate(
            {
                "title": "相机镜头自带无法拆除转接环的质检记录规则",
                "subtitles": [],
                "content": "1. 根据来源事实核对镜头本体标识，并保留当前案例边界。",
                "category_l1": "其他问题",
                "category_l2": "特殊问题",
                "layer": "L2",
                "knowledge_form": "流程方法",
                "content_type": "核验型",
                "standard_refs": [],
                "applicable_scope": "相机镜头",
                "applicable_brands": [],
                "applicable_models": [],
                "recommended_reply": "您好，请根据来源事实核对镜头本体标识。",
                "confidence": 0.9,
                "reasoning_summary": "内容来自当前主题证据。",
                "needs_human_review": True,
                "image_evidence_summary": "无需图片。",
                "requires_images": False,
                "image_usage_instruction": "",
            },
            set(),
        )


@pytest.mark.parametrize(
    "title",
    (
        "相机镜头转接环质检记录规则。",
        "相机镜头转接环质检记录规则（试行）",
    ),
)
def test_candidate_validation_retries_meta_title_with_trailing_status(
    title: str,
) -> None:
    with pytest.raises(MimoError, match="公文式"):
        mimo_module._validate_candidate(
            {
                "title": title,
                "subtitles": [],
                "content": "1. 根据来源事实核对镜头本体标识，并保留当前案例边界。",
                "category_l1": "其他问题",
                "category_l2": "特殊问题",
                "layer": "L2",
                "knowledge_form": "流程方法",
                "content_type": "核验型",
                "standard_refs": [],
                "applicable_scope": "相机镜头",
                "applicable_brands": [],
                "applicable_models": [],
                "recommended_reply": "您好，请根据来源事实核对镜头本体标识。",
                "confidence": 0.9,
                "reasoning_summary": "内容来自当前主题证据。",
                "needs_human_review": True,
                "image_evidence_summary": "无需图片。",
                "requires_images": False,
                "image_usage_instruction": "",
            },
            set(),
        )


def test_candidate_validation_accepts_natural_record_requirement_question() -> None:
    candidate = mimo_module._validate_candidate(
        {
            "title": "质检工单中需要记录哪些信息",
            "subtitles": [],
            "content": "1. 根据来源事实记录当前对象、现象和处理结果，其他情形需补充证据。",
            "category_l1": "其他问题",
            "category_l2": "特殊问题",
            "layer": "L2",
            "knowledge_form": "流程方法",
            "content_type": "核验型",
            "standard_refs": [],
            "applicable_scope": "相机镜头",
            "applicable_brands": [],
            "applicable_models": [],
            "recommended_reply": "您好，请按来源事实记录当前对象、现象和处理结果。",
            "confidence": 0.9,
            "reasoning_summary": "内容来自当前主题证据。",
            "needs_human_review": True,
            "image_evidence_summary": "无需图片。",
            "requires_images": False,
            "image_usage_instruction": "",
        },
        set(),
    )

    assert candidate["title"] == "质检工单中需要记录哪些信息"


def test_initial_review_rejects_policy_template_body_in_case_only_mode() -> None:
    guarded = workflow_module._apply_topic_initial_review_guard(
        {
            "decision": "通过",
            "knowledge_value": "值得沉淀",
            "error_type": "",
            "reason": "模型认为标题和正文可以通过。",
            "standard_consistency": "无可信标准",
            "evidence_sufficiency": "充分",
            "content_consistency": "一致",
            "image_necessity": "不需要",
            "title_quality": "清晰",
            "confidence": 0.95,
            "priority_review": False,
        },
        {
            "主标题": "镜头自带不可拆转接环时如何确认型号",
            "知识内容": (
                "1. 适用对象：相机镜头在质检时发现自带且无法拆除的转接环。\n"
                "2. 处理原则：该情况属于镜头的固有特征，不影响镜头主体型号的认定。\n"
                "3. 记录要求：提交质检工单时必须备注具体卡口类型。\n"
                "4. 边界说明：后加装且可拆卸的转接环不属于此情况。"
            ),
            "关联标准项": "",
            "主题图片必要性": "无案例图",
            "主题图片链接": "",
            "主题证据等级": "完整会话",
            "是否重点复核": "否",
            "主题无来源内容": "",
        },
        [],
        use_standard_references=False,
    )

    assert guarded["decision"] == "需修改"
    assert guarded["error_type"] == "话术不合适"
    assert "制度条款" in guarded["reason"]


def test_policy_template_detection_handles_markdown_and_halfwidth_colons() -> None:
    markers = workflow_module._topic_policy_template_markers(
        "**适用对象**: 镜头自带转接环\n"
        "2、**处理原则**: 核对来源事实\n"
        "3. **记录要求**: 记录当前案例\n"
        "4、**边界说明**: 其他情形需补充证据"
    )

    assert markers == ["适用对象", "处理原则", "记录要求", "边界说明"]


def test_policy_template_detection_handles_list_and_heading_prefixes() -> None:
    markers = workflow_module._topic_policy_template_markers(
        "- **适用对象**: 镜头自带转接环\n"
        "### 处理原则: 核对来源事实\n"
        "* **记录要求**: 记录当前案例\n"
        "> **边界说明**: 其他情形需补充证据"
    )

    assert markers == ["适用对象", "处理原则", "记录要求", "边界说明"]


def test_policy_template_detection_handles_heading_only_and_alt_numbering() -> None:
    markers = workflow_module._topic_policy_template_markers(
        "### 适用对象\n镜头自带转接环\n"
        "1) __处理原则__:\n核对来源事实\n"
        "（2）__记录要求__：\n记录当前案例\n"
        "### __边界说明__\n其他情形需补充证据"
    )

    assert markers == ["适用对象", "处理原则", "记录要求", "边界说明"]


def test_policy_template_detection_handles_combined_markdown_prefixes() -> None:
    markers = workflow_module._topic_policy_template_markers(
        "> ### 适用对象 ###\n镜头自带转接环\n"
        "> ### 处理原则 ###\n核对来源事实\n"
        "> ### 记录要求 ###\n记录当前案例\n"
        "> ### 边界说明 ###\n其他情形需补充证据"
    )

    assert markers == ["适用对象", "处理原则", "记录要求", "边界说明"]


def test_initial_review_rejects_internal_analysis_report_structure() -> None:
    review = workflow_module._rule_topic_initial_review(
        {
            "主标题": "镜头自带不可拆转接环时如何确认型号",
            "知识内容": (
                "判断对象：镜头自带且无法拆除的转接环。\n"
                "来源核验依据：按来源记录核对镜头本体标识。\n"
                "处理结论：当前案例按镜头本体型号处理。\n"
                "适用边界：其他情形需要补充对应来源事实。"
            ),
            "关联标准项": "",
            "主题图片必要性": "无案例图",
            "主题图片链接": "",
            "主题证据等级": "完整会话",
            "是否重点复核": "否",
            "主题无来源内容": "",
        },
        [],
        use_standard_references=False,
    )

    assert review["decision"] == "需修改"
    assert review["error_type"] == "话术不合适"
    assert "内部分析标签" in review["reason"]


def test_recommended_reply_is_direct_without_title_or_case_intro() -> None:
    reply = workflow_module._recommended_reply(
        "平板电池健康度应按什么优先级读取？",
        (
            "1. 苹果平板按本机、验机工具、苹果支持 App 诊断的顺序读取。\n"
            "2. 都无法获取时选电池健康度无法检测。"
        ),
        use_standard_references=True,
    )

    assert "您好，关于" not in reply
    assert "平板电池健康度应按什么优先级读取" not in reply
    assert "本机、验机工具、苹果支持 App 诊断" in reply


def test_topic_candidate_final_gate_removes_model_and_pipe_subtitle() -> None:
    rows = [
        {
            "数据ID": "MODEL-TITLE-001",
            "工单ID": "202608210001",
            "聊天内容": "笔记本屏幕出现色斑，如何判定？",
            "核心问题": "笔记本屏幕色斑如何判定",
            "产品类型": "笔记本",
            "问题意图": "标准判定",
            "对象/部位": "屏幕",
            "异常现象": "色斑",
            "解题方式": "按显示标准核验",
            "语义标注依据": "来源明确询问屏幕色斑判定。",
        }
    ]
    result = workflow_module._topic_candidate_row(
        "TOP-MODEL-TITLE-001",
        ("笔记本", "屏幕", "色斑"),
        rows,
        [],
        {
            "title": "拯救者R7000P屏幕色斑如何判定",
            "subtitles": ["笔记本｜拯救者R7000P｜屏幕｜色斑怎么核验？"],
            "content": "1. 回收师在现场上传图片；2. 平台标准依据：暂无。",
            "recommended_reply": "您好，关于“拯救者R7000P屏幕色斑如何判定”，回收师上传图片。",
            "applicable_models": ["拯救者R7000P"],
            "confidence": 0.8,
            "knowledge_form": "具体判定",
        },
        "mimo",
        "test-model",
        "test-prompt",
        "test-run",
        "",
        "topic_model_labeled",
        0.75,
        use_standard_references=True,
    )

    assert "拯救者R7000P" not in result["主标题"]
    assert "｜" not in result["副标题"]
    assert "|" not in result["副标题"]
    assert "您好，关于" not in result["推荐回复"]


def test_no_standard_candidate_does_not_emit_standard_language() -> None:
    rows = [
        {
            "数据ID": "NO-STANDARD-001",
            "工单ID": "202608210002",
            "聊天内容": "设备接口插入困难，无法确认原因。",
            "核心问题": "笔记本接口插入困难如何核验",
            "产品类型": "笔记本",
            "问题意图": "检测核验",
            "对象/部位": "USB接口",
            "异常现象": "插入困难",
            "解题方式": "补充接口近景并检查异物或针脚",
            "语义标注依据": "来源仅说明需要补充证据。",
        }
    ]
    result = workflow_module._topic_candidate_row(
        "TOP-NO-STANDARD-001",
        ("笔记本", "USB接口", "插入困难"),
        rows,
        [],
        {
            "title": "笔记本USB接口插入困难如何核验",
            "subtitles": [],
            "content": "回收师反馈接口插入困难；平台标准依据：当前标准要求检查接口。",
            "recommended_reply": "您好，请按平台标准判定该接口异常。",
            "confidence": 0.8,
            "knowledge_form": "流程方法",
            "content_type": "核验型",
        },
        "mimo",
        "test-model",
        "test-prompt",
        "test-run",
        "",
        "topic_model_labeled",
        0.75,
        use_standard_references=True,
    )

    assert result["关联标准项"] == ""
    assert "平台标准依据" not in result["知识内容"]
    assert "当前有效标准" not in result["知识内容"]
    assert "回收师" not in result["知识内容"]
    assert "平台标准" not in result["推荐回复"]


def test_no_standard_reference_gate_ignores_descriptive_standard_topic_language() -> None:
    assert not workflow_module._candidate_contains_standard_reference(
        {
            "title": "平板屏幕内部白色异物应如何区分？",
            "subtitles": [],
            "content": "1. 先确认异物位于屏幕内部还是表面。",
            "recommended_reply": "先确认异物位置，再补充亮屏和息屏照片。",
            "reasoning_summary": "该问题属于质检标准咨询，但本次未引用标准。",
            "standard_refs": [],
        }
    )
    # 可验证的标准标识（编号、路径、字段名）才算标准引用。
    assert workflow_module._candidate_contains_standard_reference(
        {
            "content": "1. 标准编号：QC-TABLET-001 要求按屏幕漏液判定。",
            "standard_refs": [],
        }
    )
    assert workflow_module._candidate_contains_standard_reference(
        {
            "content": "1. 按 QC-TABLET-001 的要求执行。",
            "standard_refs": [],
        }
    )
    assert workflow_module._candidate_contains_standard_reference(
        {
            "content": "1. 标准路径：\n【电池】-【电池健康度】要求按顺序读取。",
            "standard_refs": [],
        }
    )
    # 历史回复里的口语“标准”是案例事实，不是标准引用，不能误拦截。
    for content in (
        "1. 按质检标准判为屏幕漏液。",
        "1. 按平台标准，选择功能异常。",
        "1. 平台标准：电池健康度低于80%算异常。",
        "1. 回收标准中电池健康度低于80%算异常。",
        "1. 按平台口径，电池健康度低于80%选异常。",
        "1. 根据平台标准，该现象必须判定为屏幕漏液。",
    ):
        assert not workflow_module._candidate_contains_standard_reference(
            {"content": content, "standard_refs": []}
        )


def test_handling_options_extracted_from_standard_snippet_and_source_conclusion() -> None:
    assert workflow_module._extract_handling_options_from_text(
        "2. 出厂机型与实物机型不符，需要勾选【设备机况不支持回收】。"
    ) == ["勾选【设备机况不支持回收】"]
    assert workflow_module._extract_handling_options_from_text(
        "屏幕漏液按【屏幕异常】判定；进水按【进液】处理。"
    ) == ["按【屏幕异常】判定", "按【进液】处理"]
    assert workflow_module._extract_handling_options_from_text(
        "电池健康度无法读取时判定为【无法检测】。"
    ) == ["判定为【无法检测】"]
    assert workflow_module._extract_handling_options_from_text("") == []


def test_topic_candidate_row_exports_handling_options_when_standard_hits() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-BATTERY",
        title="平板电池健康度判定方式",
        category_l1="电池",
        category_l2="电池健康度",
        knowledge_type="场景判定",
        standard_path="【电池】-【电池健康度】",
        keywords=["平板", "电池健康度", "读取"],
        scope="平板电脑-通用",
        response_snippet=(
            "1. 优先读取本机电池健康值。\n"
            "2. 无法读取时使用指定检测工具。\n"
            "3. 仍无法获取时勾选【电池健康度无法检测】。"
        ),
        status="published",
        version="飞书平板标准-20260817",
    )
    rows = [
        {
            "数据ID": "TABLET-OPT-001",
            "工单ID": "202608220011",
            "聊天内容": "平板电池健康度无法读取怎么选？",
            "核心问题": "平板电池健康度无法读取怎么选",
            "产品类型": "平板电脑",
            "一级分类": "电池",
            "二级分类": "电池健康度",
            "问题意图": "检测核验",
            "对象/部位": "电池健康度",
            "异常现象": "无法读取",
            "解题方式": "按标准优先级读取",
            "语义标注依据": "来源明确咨询电池健康度读取方式。",
        }
    ]
    result = workflow_module._topic_candidate_row(
        "TOP-OPT-001",
        ("平板电脑", "电池", "电池健康度"),
        rows,
        [(standard, 0.95)],
        {
            "title": "平板电池健康度应按什么优先级读取？",
            "subtitles": [],
            "content": "1. 优先读取本机电池健康值。",
            "recommended_reply": "按本机、工具、人工的顺序读取。",
            "confidence": 0.9,
            "knowledge_form": "具体判定",
            "content_type": "判定型",
        },
        "mimo",
        "test-model",
        "test-prompt",
        "test-run",
        "",
        "topic_model_labeled",
        0.75,
        use_standard_references=True,
    )

    assert result["关联标准项"] != ""
    assert "勾选【电池健康度无法检测】" in result["候选项/处理项"]


def test_topic_candidate_row_exports_source_options_when_standard_missing() -> None:
    rows = [
        {
            "数据ID": "TABLET-OPT-002",
            "工单ID": "202608220012",
            "聊天内容": "平板耳机孔有灰怎么处理？",
            "核心问题": "平板耳机孔有灰怎么处理",
            "产品类型": "平板电脑",
            "一级分类": "耳机孔",
            "二级分类": "有灰",
            "问题意图": "处理建议",
            "对象/部位": "耳机孔",
            "异常现象": "有灰",
            "人工判定结论": "判定为耳机孔进灰，按【清洁处理】。",
            "判定结论": "耳机孔进灰",
            "历史实际回复": "您好，耳机孔进灰按【清洁处理】即可。",
            "解题方式": "清理耳机孔灰尘",
            "语义标注依据": "来源明确说明耳机孔进灰处理方式。",
        }
    ]
    result = workflow_module._topic_candidate_row(
        "TOP-OPT-002",
        ("平板电脑", "耳机孔", "有灰"),
        rows,
        [],
        {
            "title": "平板耳机孔进灰应如何处理？",
            "subtitles": [],
            "content": "1. 判定为耳机孔进灰。",
            "recommended_reply": "您好，耳机孔进灰按【清洁处理】即可。",
            "confidence": 0.8,
            "knowledge_form": "具体判定",
            "content_type": "判定型",
        },
        "mimo",
        "test-model",
        "test-prompt",
        "test-run",
        "",
        "topic_model_labeled",
        0.75,
        use_standard_references=True,
    )

    assert result["标准引用标签"] == "未引用标准-人工重点复核"
    assert result["关联标准项"] == ""
    assert "按【清洁处理】" in result["候选项/处理项"]


def test_topic_candidate_row_keeps_handling_options_empty_without_evidence() -> None:
    rows = [
        {
            "数据ID": "TABLET-OPT-003",
            "工单ID": "202608220013",
            "聊天内容": "平板屏幕有横线需要补充什么证据？",
            "核心问题": "平板屏幕有横线需要补充什么证据",
            "产品类型": "平板电脑",
            "一级分类": "显示问题",
            "二级分类": "横线",
            "问题意图": "补充证据",
            "对象/部位": "屏幕",
            "异常现象": "横线",
            "解题方式": "补充亮屏和息屏照片",
            "语义标注依据": "来源仅说明需要补充证据。",
        }
    ]
    result = workflow_module._topic_candidate_row(
        "TOP-OPT-003",
        ("平板电脑", "屏幕", "横线"),
        rows,
        [],
        {
            "title": "平板屏幕横线应补充什么证据？",
            "subtitles": [],
            "content": "1. 补充亮屏和息屏照片。",
            "recommended_reply": "请补充亮屏和息屏照片。",
            "confidence": 0.7,
            "knowledge_form": "具体判定",
            "content_type": "核验型",
        },
        "mimo",
        "test-model",
        "test-prompt",
        "test-run",
        "",
        "topic_model_labeled",
        0.75,
        use_standard_references=True,
    )

    assert result["候选项/处理项"] == ""


def test_standard_mapping_failure_revokes_reference_before_export() -> None:
    topic = {
        "标准引用标签": "已引用标准知识点",
        "标准引用门禁状态": "accepted",
        "关联标准项": "CZ-HQ-TABLET-WRONG | 外壳辅件缺损",
        "候选项/处理项": "【外壳辅件缺损】",
        "主题标准版本": "v1",
        "来源版本": "v1",
        "知识来源": "方向二总部标准候选",
        "主题对象/部位": "外壳",
        "主题异常现象": "缝隙",
        "主题解题方式": "补充局部照片后核对",
        "知识内容": (
            "1. 满足以下任一条件时，勾选"
            "【外壳外观情况】-【外壳其他现象】-【外壳辅件缺损】：\n\n"
            "镜片整体缺失。"
        ),
        "推荐回复": "勾选【外壳外观情况】-【外壳其他现象】-【外壳辅件缺损】。",
        "模型初标错误类型": "标准项映射错",
        "模型初标标准一致性": "不一致",
        "模型初标原因": "来源事实是外壳缝隙，不是辅件缺损。",
    }

    guarded = workflow_module._enforce_standard_reference_consistency(
        topic,
        use_standard_references=True,
    )

    assert guarded["标准引用标签"] == "未引用标准-人工重点复核"
    assert guarded["标准引用门禁状态"] == "retrieved_mapping_rejected"
    assert guarded["关联标准项"] == ""
    assert guarded["候选项/处理项"] == ""
    assert guarded["主题标准版本"] == ""
    assert guarded["来源版本"] == ""
    assert guarded["知识来源"] == "方向二经验补充候选"
    assert "【外壳外观情况】" not in guarded["知识内容"]
    assert "勾选" not in guarded["知识内容"]
    assert guarded["推荐回复"] == ""


def test_standard_label_without_association_is_rejected() -> None:
    guarded = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "标准引用门禁状态": "accepted",
            "关联标准项": "",
            "候选项/处理项": "",
        },
        use_standard_references=True,
    )

    assert guarded["标准引用标签"] == "未引用标准-人工重点复核"
    assert guarded["标准引用门禁状态"] == "rejected_or_missing"


def test_no_standard_content_does_not_claim_standard_selection() -> None:
    cleaned = workflow_module._strip_unverified_standard_language(
        "依据标准判定为屏幕进灰，并勾选相应等级。"
    )

    assert "依据标准" not in cleaned
    assert "勾选相应等级" not in cleaned
    assert "人工" in cleaned


def test_recommended_reply_from_final_content_drops_full_standard_path() -> None:
    reply = workflow_module._recommended_reply_from_final_content(
        "1. 【拆修及浸液情况】-【电池拆修】-【电池-工具读出异常】\n"
        "2. 上门门店场景：苹果机型使用验机工具（一根线）验机报告读出电池结果为“异常”。"
    )

    assert reply.startswith("选择【电池-工具读出异常】")
    assert "【拆修及浸液情况】-【电池拆修】" not in reply


def test_case_narrative_title_requires_structured_rebuild() -> None:
    assert workflow_module._title_requires_structured_rebuild(
        "回收师询问平板屏幕漏液的判定方法，并描述用手电筒照射屏幕内部看到如何判定"
    )
    rebuilt = workflow_module._rebuild_title_from_structured_fields(
        {
            "产品类型": "平板电脑",
            "对象/部位": "屏幕",
            "异常现象": "漏液",
            "问题意图": "标准判定",
        }
    )
    assert rebuilt == "平板电脑屏幕漏液应如何判定？"


def test_product_conflict_is_a_hard_cluster_admission_block() -> None:
    admission = workflow_module._cluster_topic_admission(
        [
            {
                "产品类型": "平板电脑",
                "_原子品类冲突": True,
                "语义标注置信度": 0.95,
                "_聚类裁决置信度": 0.95,
            }
        ],
        {
            "requested_mode": "direct_mimo",
            "effective_mode": "direct_mimo",
        },
        enabled=True,
        min_confidence=0.75,
    )

    assert admission["admitted"] is False
    assert admission["hard_blocked"] is True
    assert "品类冲突" in admission["reason"]


def test_failed_topic_transcription_body_is_knowledge_like_not_case_analysis() -> None:
    rows = [
        {
            "数据ID": "TABLET-FAIL-BODY-001",
            "工单ID": "202608230001",
            "产品类型": "平板电脑",
            "一级分类": "拆修及浸液情况",
            "二级分类": "屏幕拆修",
            "问题意图": "标准判定",
            "对象/部位": "屏幕",
            "异常现象": "验机工具读出异常",
            "核心问题": "平板屏幕工具读出异常如何判定",
            "判定结论": "按【屏幕-工具读出异常】处理",
            "判定依据": (
                "关键事实：一根线读出屏幕结果为异常；"
                "匹配口径：屏幕拆修项；"
                "定义：工具读出异常时按屏幕维修处理。"
            ),
            "历史实际回复": (
                "老师，这种情况优先看工具结果，"
                "一根线读出异常就勾选【屏幕-工具读出异常】。"
            ),
            "解题方式": "按工具结果选择拆修项",
            "语义标注依据": "来源明确记录工具结果和处理选项。",
        }
    ]
    topic = workflow_module._failed_topic_transcription_row(
        "TOP-FAIL-BODY-001",
        ("平板电脑", "拆修及浸液情况", "屏幕拆修"),
        rows,
        {"knowledge_value": "值得沉淀"},
        provider="mimo",
        model_name="mimo-v2.5",
        prompt_version="test",
        model_run_id="run",
        transcription_status="topic_model_validation_failed",
        model_call_status="model_success",
        error="模型草稿包含分析过程",
        matches=[],
        use_standard_references=True,
    )

    content = topic["知识内容"]
    assert content
    assert "关键事实" not in content
    assert "匹配口径" not in content
    assert "定义：" not in content
    assert "老师" not in content
    assert "回收师" not in content
    assert "【屏幕-工具读出异常】" in content
    assert topic["推荐回复"] == ""


def test_user_judgment_tool_result_rejects_battery_health_standard_target() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-BATTERY-HEALTH",
        title="平板电池健康度判定方式",
        category_l1="基本情况",
        category_l2="电池健康度（单选）",
        knowledge_type="场景判定",
        standard_path="【基本情况】-【电池健康度（单选）】",
        keywords=["平板", "电池健康度", "最大容量"],
        scope="平板电脑-苹果",
        response_snippet="按本机、工具、支持APP的顺序读取电池健康度。",
        status="published",
        version="test",
    )
    query = {
        "产品类型": "平板电脑",
        "一级分类": "基本情况",
        "二级分类": "电池健康度（单选）",
        "问题意图": "标准判定",
        "对象/部位": "电池健康度",
        "异常现象": "用户判断",
        "核心问题": "一根线工具读出电池用户判断怎么选",
        "人工核心问题": "电池序列号工具读出用户判断应如何处理",
        "人工判定结论": "用户判断不能作为电池健康度的勾选依据。",
        "判定依据": "应按电池拆修项判断工具结果。",
        "平台": "iOS",
        "品牌": "Apple",
    }

    reasons = workflow_module._standard_match_rejection_reasons(query, standard)

    assert "judgment_target_mismatch" in reasons


def test_user_judgment_tool_result_can_bridge_to_battery_repair_standard() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-BATTERY-REPAIR",
        title="电池工具读出用户判断如何处理",
        category_l1="拆修及浸液情况",
        category_l2="电池拆修（单选）",
        knowledge_type="场景判定",
        standard_path="【拆修及浸液情况】-【电池拆修（单选）】",
        keywords=["电池", "用户判断", "工具读出异常"],
        scope="平板电脑-苹果",
        response_snippet="电池工具读出异常时，勾选【电池-工具读出异常】。",
        status="published",
        version="test",
    )
    query = {
        "产品类型": "平板电脑",
        "一级分类": "基本情况",
        "二级分类": "电池健康度（单选）",
        "问题意图": "标准判定",
        "对象/部位": "电池健康度",
        "异常现象": "用户判断",
        "核心问题": "一根线工具读出电池用户判断怎么选",
        "人工核心问题": "电池序列号工具读出用户判断应如何处理",
        "人工判定结论": "按电池拆修项处理。",
        "判定依据": "用户判断属于工具结果，需按电池拆修判断。",
        "平台": "iOS",
        "品牌": "Apple",
    }

    reasons = workflow_module._standard_match_rejection_reasons(query, standard)

    assert "object_mismatch" not in reasons
    assert "standard_path_mismatch" not in reasons


def test_user_judgment_without_observed_repair_evidence_cannot_select_tool_abnormal() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-BATTERY-TOOL-ABNORMAL",
        title="电池-工具读出异常是什么意思",
        category_l1="拆修及浸液情况",
        category_l2="电池拆修（单选）",
        knowledge_type="标准定义",
        standard_path=(
            "【拆修及浸液情况】-【电池拆修】-【电池-工具读出异常】"
        ),
        keywords=["电池", "工具读出异常"],
        scope="平板电脑-苹果",
        response_snippet="验机工具读出电池异常时，勾选【电池-工具读出异常】。",
        status="published",
        version="test",
    )
    query = {
        "产品类型": "平板电脑",
        "一级分类": "拆修及浸液情况",
        "二级分类": "电池拆修",
        "问题意图": "标准判定",
        "对象/部位": "电池",
        "异常现象": "验机工具读出用户判断",
        "核心问题": "平板电池验机工具显示用户判断时如何处理",
        "人工判定结论": "用户判断不算问题。",
        "判定依据": "需现场查看电池是否存在拆修现象。",
        "平台": "iOS",
        "品牌": "Apple",
    }

    reasons = workflow_module._standard_match_rejection_reasons(query, standard)

    assert "user_judgment_observation_required" in reasons


def test_user_judgment_repair_knowledge_requires_observation_before_tool_abnormal() -> None:
    content = workflow_module._failed_topic_source_knowledge_content(
        {
            "产品类型": "平板电脑",
            "一级分类": "拆修及浸液情况",
            "二级分类": "电池拆修",
            "对象/部位": "电池",
            "异常现象": "验机工具读出用户判断",
            "核心问题": "平板电池验机工具显示用户判断时如何处理",
            "解题方式": "现场核验电池是否存在拆修现象",
        }
    )

    assert "不能仅凭该提示直接勾选" in content
    assert "未发现对应部位的拆修现象时，不处理该提示" in content
    assert "发现对应部位存在明确拆修现象时，按对应部位的工具读出异常项处理" in content


def test_cross_product_atomic_question_is_not_exported_as_review_candidate() -> None:
    class CrossProductMimo:
        config = SimpleNamespace(model="mimo-cross-product-export-test")

        def analyze_cluster_units(self, _row):
            return MimoLabelResult(
                candidate={
                    "conversation_type": "single_topic",
                    "topics": [
                        {
                            "normalized_issue": "笔记本C面涂鸦是否拒收",
                            "product_category": "笔记本",
                            "scope_type": "品类专用",
                            "platform": "通用",
                            "brand": "通用",
                            "model_scope": "通用",
                            "category_l1": "外观状态",
                            "category_l2": "外观异常",
                            "intent": "标准判定",
                            "subject": "C面",
                            "phenomenon": "涂鸦",
                            "judgment_target": "判断是否拒收",
                            "resolution_mode": "按外观标准核验",
                            "standard_path": "外观状态",
                            "threshold_or_exception": "",
                            "evidence_summary": "聊天中夹带了笔记本问题。",
                            "confidence": 0.95,
                            "requires_review": False,
                        }
                    ],
                },
                request_audit={},
                response_audit={},
            )

        def cluster_atomic_units(self, units):
            return MimoLabelResult(
                candidate={
                    "clusters": [
                        {
                            "cluster_id": "C001",
                            "theme_name": "笔记本外观异常",
                            "member_atomic_ids": [
                                unit["unit_id"] for unit in units
                            ],
                            "merge_basis": "单成员。",
                        }
                    ],
                    "split_requests": [],
                    "review_requests": [],
                },
                request_audit={},
                response_audit={},
            )

    topics, _mapping, _gaps, pending = build_topic_review_rows(
        [
            {
                "数据ID": "TABLET-WITH-LAPTOP-CHAT",
                "工单ID": "202608230099",
                "聊天内容": "平板WiFi版怎么区分？笔记本C面涂鸦能回收吗？",
                "核心问题": "平板WiFi版和蜂窝版如何区分",
                "产品类型": "平板电脑",
            }
        ],
        use_mimo=False,
        mimo_client=CrossProductMimo(),
        clustering_mode="direct_mimo",
        enforce_cluster_admission=True,
    )

    assert topics == []
    assert len(pending) == 1
    assert pending[0]["待聚合状态"] == "pending_product_conflict_review"
    assert "品类冲突" in pending[0]["待聚合原因"]


def test_battery_user_judgment_query_is_retargeted_to_repair_scope() -> None:
    corrected = workflow_module._retarget_battery_user_judgment_query(
        {
            "产品类型": "平板电脑",
            "一级分类": "基本情况",
            "二级分类": "电池健康度（单选）",
            "问题意图": "标准判定",
            "对象/部位": "电池健康度",
            "异常现象": "用户判断",
            "核心问题": "一根线工具读出电池用户判断怎么选",
            "解题方式": "按健康度流程选择无法检测",
            "历史实际回复": "工具读出用户判断，应按电池拆修项核验。",
        }
    )

    assert corrected["一级分类"] == "拆修及浸液情况"
    assert corrected["二级分类"] == "电池拆修"
    assert corrected["对象/部位"] == "电池"
    assert corrected["异常现象"] == "验机工具读出用户判断"
    assert "健康度流程" not in corrected["解题方式"]


def test_battery_user_judgment_title_variant_is_retargeted_to_same_repair_topic() -> None:
    title_variant = workflow_module._retarget_battery_user_judgment_query(
        {
            "产品类型": "平板电脑",
            "一级分类": "基本情况",
            "二级分类": "电池健康度（单选）",
            "问题意图": "标准判定",
            "对象/部位": "电池健康度",
            "异常现象": "用户判断",
            "核心问题": "平板电脑电池健康度质检选项用户判断如何选择如何处理",
        }
    )
    tool_variant = workflow_module._retarget_battery_user_judgment_query(
        {
            "产品类型": "平板电脑",
            "一级分类": "基本情况",
            "二级分类": "电池健康度（单选）",
            "问题意图": "标准判定",
            "对象/部位": "电池健康度",
            "异常现象": "用户判断",
            "核心问题": "平板电脑电池验机工具读出用户判断应如何判定",
        }
    )

    assert title_variant["一级分类"] == "拆修及浸液情况"
    assert title_variant["二级分类"] == "电池拆修"
    assert title_variant["核心问题"] == tool_variant["核心问题"]
    assert title_variant["异常现象"] == tool_variant["异常现象"]
    assert (
        workflow_module._rebuild_title_from_structured_fields(title_variant)
        == "平板电脑电池拆修检测显示“用户判断”时如何处理？"
    )


def test_user_judgment_candidate_forces_repair_title_over_natural_model_title() -> None:
    rows = [
        {
            "数据ID": "TABLET-USER-JUDGMENT-001",
            "工单ID": "202608230001",
            "产品类型": "平板电脑",
            "一级分类": "基本情况",
            "二级分类": "电池健康度（单选）",
            "问题意图": "标准判定",
            "对象/部位": "电池健康度",
            "异常现象": "用户判断",
            "核心问题": "平板电脑电池健康度质检选项用户判断如何选择如何处理",
            "聊天内容": "一根线工具显示电池用户判断，应怎么处理？",
            "历史实际回复": "工具显示用户判断时，现场核验电池拆修现象。",
        }
    ]

    result = workflow_module._topic_candidate_row(
        "TOP-TABLET-USER-JUDGMENT-001",
        ("平板电脑", "电池", "用户判断"),
        rows,
        [],
        {
            "title": "平板电脑电池健康度质检选项用户判断如何选择？",
            "subtitles": [],
            "content": "请根据工具提示处理。",
            "recommended_reply": "",
            "confidence": 0.8,
            "knowledge_form": "具体判定",
        },
        "mimo",
        "test-model",
        "test-prompt",
        "test-run",
        "",
        "topic_model_labeled",
        0.75,
        use_standard_references=True,
    )

    assert result["主标题"] == "平板电脑电池拆修检测显示“用户判断”时如何处理？"
    assert "电池健康度" not in result["主标题"]


def test_non_tablet_battery_user_judgment_is_not_retargeted_or_merged() -> None:
    query = {
        "产品类型": "手机",
        "一级分类": "基本情况",
        "二级分类": "电池健康度（单选）",
        "问题意图": "标准判定",
        "对象/部位": "电池健康度",
        "异常现象": "用户判断",
        "核心问题": "手机电池验机工具读出用户判断应如何判定",
    }

    corrected = workflow_module._retarget_battery_user_judgment_query(query)
    merged = workflow_module._merge_known_equivalent_topic_groups(
        [
            (
                ("direct_mimo", "自营回收", "手机", "cluster-1"),
                [dict(query)],
            )
        ]
    )

    assert corrected == query
    assert merged[0][0] == ("direct_mimo", "自营回收", "手机", "cluster-1")
    assert merged[0][1][0]["核心问题"] == query["核心问题"]


def test_equivalent_battery_user_judgment_topic_groups_are_merged() -> None:
    topic_groups = [
        (
            ("direct_mimo", "自营回收", "平板电脑", "cluster-1"),
            [
                {
                    "产品类型": "平板电脑",
                    "核心问题": "平板电脑电池健康度质检选项用户判断如何选择如何处理",
                    "异常现象": "用户判断",
                }
            ],
        ),
        (
            ("direct_mimo", "自营回收", "平板电脑", "cluster-2"),
            [
                {
                    "产品类型": "平板电脑",
                    "核心问题": "平板电脑电池验机工具读出用户判断应如何判定",
                    "异常现象": "用户判断",
                }
            ],
        ),
    ]

    merged = workflow_module._merge_known_equivalent_topic_groups(topic_groups)

    assert len(merged) == 1
    assert len(merged[0][1]) == 2
    assert all(
        row["核心问题"] == "平板电池验机工具显示用户判断时如何处理"
        for row in merged[0][1]
    )


def test_single_battery_health_percentage_is_not_a_reusable_topic() -> None:
    assert workflow_module._is_single_battery_health_observation_topic(
        {
            "产品类型": "平板电脑",
            "核心问题": "平板电脑电池健康度84%如何处理",
            "对象/部位": "电池健康度",
            "异常现象": "84%",
        }
    )
    assert not workflow_module._is_single_battery_health_observation_topic(
        {
            "产品类型": "平板电脑",
            "核心问题": "平板电池健康度本机和验机工具均无法检测如何处理",
            "对象/部位": "电池健康度",
            "异常现象": "无法检测",
        }
    )
    assert workflow_module._is_single_battery_health_observation_topic(
        {
            "产品类型": "平板电脑",
            "核心问题": "平板电脑电池健康度84%如何处理",
            "对象/部位": "电池健康度",
            "异常现象": "验机工具读数84%",
            "解题方式": "本机与验机工具均显示84%",
        }
    )


def test_revoked_user_judgment_standard_keeps_special_review_body_and_status() -> None:
    topic = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "标准引用门禁状态": "accepted",
            "关联标准项": (
                "【拆修及浸液情况】-【电池拆修】-【电池-工具读出异常】"
            ),
            "主题标准版本": "v1",
            "来源版本": "v1",
            "知识来源": "方向二总部标准候选",
            "主题对象/部位": "电池",
            "主题异常现象": "验机工具读出用户判断",
            "主题解题方式": "现场核验电池是否存在明确拆修现象",
            "核心问题": "平板电池验机工具显示用户判断时如何处理",
            "模型初标错误类型": "标准项映射错误",
            "模型初标原因": "标准项映射错误：不能仅凭用户判断直接勾选工具异常。",
            "模型初标标准一致性": "不一致",
        },
        use_standard_references=True,
    )

    assert topic["标准引用标签"] == "未引用标准-人工重点复核"
    assert topic["标准引用门禁状态"] == "retrieved_mapping_rejected"
    assert topic["关联标准项"] == ""
    assert topic["候选项/处理项"] == ""
    assert "不能仅凭该提示直接勾选" in topic["知识内容"]
    assert "未发现对应部位的拆修现象时，不处理该提示" in topic["知识内容"]


def test_revoked_user_judgment_standard_uses_topic_export_fields() -> None:
    topic = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "关联标准项": (
                "【拆修及浸液情况】-【电池拆修】-【电池-工具读出异常】"
            ),
            "适用范围": "平板电脑",
            "主标题": "平板电脑电池拆修检测显示“用户判断”时如何处理？",
            "主题对象/部位": "电池",
            "主题异常现象": "验机工具读出用户判断",
            "主题解题方式": "核对工具报告和对应电池拆修结论",
            "模型初标错误类型": "标准项映射错误",
        },
        use_standard_references=True,
    )

    assert topic["标准引用门禁状态"] == "retrieved_mapping_rejected"
    assert "不能仅凭该提示直接勾选" in topic["知识内容"]
    assert "现场核验电池是否存在明确拆修现象" in topic["知识内容"]
    assert "未发现对应部位的拆修现象时，不处理该提示" in topic["知识内容"]


def test_revoked_standard_keeps_source_backed_wifi_process_and_reply() -> None:
    topic = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "关联标准项": "CZ-HQ-TABLET-WIFI | WIFI版应该如何查看",
            "适用范围": "平板电脑",
            "主题对象/部位": "网络制式",
            "主题异常现象": "区分WiFi版和蜂窝版",
            "主题解题方式": "检查SIM卡槽和IMEI信息",
            "模型初标错误类型": "内容不完整、标准召回不足",
            "模型初标原因": "标准只覆盖WiFi版，遗漏蜂窝版标准项。",
            "模型初标标准一致性": "不一致",
            "主题事实证据包": json.dumps(
                {
                    "representative_facts": [
                        {
                            "atomic_question": "平板电脑｜网络制式｜区分WiFi版和蜂窝版",
                            "human_core_problem": "如何判断平板是WiFi版还是蜂窝版。",
                            "human_judgment_conclusion": (
                                "有SIM卡槽且系统内可查到IMEI为蜂窝版，否则为WiFi版。"
                            ),
                            "historical_actual_reply": (
                                "先看机身有没有SIM卡卡槽；有卡槽时，到设置-关于本机"
                                "查看是否有IMEI，能插卡且有IMEI就是蜂窝版；确认后"
                                "蜂窝版填写IMEI，WiFi版填写PSN。"
                            ),
                            "judgment_basis": "来源明确记录卡槽和IMEI的区分方法。",
                        }
                    ]
                },
                ensure_ascii=False,
            ),
        },
        use_standard_references=True,
    )

    assert topic["标准引用标签"] == "未引用标准-人工重点复核"
    assert topic["标准引用门禁状态"] == "retrieved_mapping_rejected"
    assert "SIM卡卡槽" in topic["知识内容"]
    assert "IMEI" in topic["知识内容"]
    assert "SIM卡卡槽" in topic["推荐回复"]
    assert "IMEI" in topic["推荐回复"]


def test_revoked_standard_keeps_source_backed_battery_reading_process() -> None:
    topic = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "关联标准项": "CZ-HQ-TABLET-BATTERY | 96%≤电池健康度≤100%",
            "适用范围": "平板电脑",
            "主题对象/部位": "电池健康度",
            "主题异常现象": "本机和验机工具无法读取",
            "模型初标错误类型": "标准项映射错",
            "模型初标原因": "错误映射为96%≤电池健康度≤100%。",
            "模型初标标准一致性": "不一致",
            "主题事实证据包": json.dumps(
                {
                    "representative_facts": [
                        {
                            "historical_actual_reply": (
                                "本机不显示电池健康度且工具检测失败时，请打开设备上的"
                                "支持APP，进入设备性能-电池性能-联系技术支持，等待诊断"
                                "报告返回后根据结果填写；此方法也无效时，勾选"
                                "【电池健康度无法检测】。"
                            ),
                            "judgment_basis": "来源明确记录支持APP诊断是下一步。",
                        }
                    ]
                },
                ensure_ascii=False,
            ),
        },
        use_standard_references=True,
    )

    assert "支持APP" in topic["知识内容"]
    assert "无法检测" in topic["知识内容"]
    assert "96%" not in topic["知识内容"]
    assert "支持APP" in topic["推荐回复"]
    assert "【电池健康度无法检测】" in topic["候选项/处理项"]


def test_revoked_standard_keeps_source_backed_strong_light_boundary() -> None:
    topic = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "关联标准项": "CZ-HQ-TABLET-SCREEN | 碎裂/磕点如何分级判定",
            "适用范围": "平板电脑",
            "主题对象/部位": "屏幕边缘",
            "主题异常现象": "强光下可见的不规则起伏",
            "模型初标错误类型": "标准引用错误",
            "模型初标原因": "标准路径属于碎裂/磕点，与屏幕脱胶主题不一致。",
            "模型初标标准一致性": "不一致",
            "主题事实证据包": json.dumps(
                {
                    "representative_facts": [
                        {
                            "historical_actual_reply": (
                                "请补充按压检测：将屏幕朝地面、后盖朝上进行按压，"
                                "观察屏幕盖板是否有上下起伏现象。"
                            ),
                            "source_supported_threshold_or_exception": (
                                "不打强光不可见的不用判。"
                            ),
                        }
                    ]
                },
                ensure_ascii=False,
            ),
        },
        use_standard_references=True,
    )

    assert "不打强光不可见的不用判" in topic["知识内容"]
    assert "屏幕朝地面、后盖朝上" in topic["知识内容"]
    assert "不打强光不可见的不用判" in topic["推荐回复"]


def test_revoked_standard_rebuilds_camera_gap_measurement_rule() -> None:
    topic = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "关联标准项": "CZ-HQ-TABLET-WRONG | 外壳辅件缺损如何处理",
            "候选项/处理项": "【外壳外观情况】-【外壳其他现象】-【外壳辅件缺损】",
            "主题标准版本": "飞书平板标准-20260817",
            "来源版本": "飞书平板标准-20260817",
            "知识来源": "方向二总部标准候选",
            "适用范围": "平板电脑",
            "主标题": "平板电脑外壳辅件缺损功能如何核验",
            "副标题": "平板电脑外壳辅件缺损功能怎么核验？",
            "主题对象/部位": "后置摄像头镜片与保护圈衔接处",
            "主题异常现象": "缝隙",
            "主题解题方式": "根据尺寸判定",
            "模型初标错误类型": "主题内容与来源事实不匹配",
            "模型初标原因": (
                "引用标准路径为外壳辅件缺损，但来源事实是后摄镜片衔接处缝隙的"
                "尺寸判定，标准与来源事实不一致。"
            ),
            "模型初标标准一致性": "不一致",
            "主题事实证据包": json.dumps(
                {
                    "representative_facts": [
                        {
                            "atomic_question": (
                                "平板电脑｜后置摄像头镜片与保护圈衔接处｜"
                                "缝隙（自测0.3-0.4mm）｜未达到>0.5mm判定标准，不判异常"
                            ),
                            "human_judgment_conclusion": (
                                "自测缝隙约0.3-0.4mm，未达到>0.5mm的判定标准，"
                                "当前不满足外壳缝隙的判定条件。"
                            ),
                            "historical_actual_reply": (
                                "根据图片和测量，后摄镜片与保护圈之间的缝隙约0.3-0.4mm，"
                                "未达到0.5mm的判定阈值，建议按正常外观状态处理。"
                            ),
                            "source_supported_threshold_or_exception": "缝隙＞0.5mm",
                        }
                    ]
                },
                ensure_ascii=False,
            ),
        },
        use_standard_references=True,
    )

    assert topic["标准引用标签"] == "未引用标准-人工重点复核"
    assert topic["标准引用门禁状态"] == "retrieved_mapping_rejected"
    assert topic["关联标准项"] == ""
    assert topic["候选项/处理项"] == ""
    assert "外壳辅件缺损" not in topic["主标题"]
    assert "后置摄像头镜片与保护圈衔接处" in topic["主标题"]
    assert "缝隙" in topic["主标题"]
    assert "外壳辅件缺损" not in topic["副标题"]
    assert "后置摄像头镜片与保护圈衔接处" in topic["副标题"]
    assert "外壳辅件缺损" not in topic["知识内容"]
    assert "后置摄像头镜片与保护圈衔接处" in topic["知识内容"]
    assert "测量" in topic["知识内容"]
    assert "缝隙＞0.5mm" in topic["知识内容"]
    assert "不满足上述条件" in topic["知识内容"]
    assert "补充清晰图片和测量证据" in topic["知识内容"]
    assert "缝隙＞0.5mm" in topic["推荐回复"]


def test_revoked_standard_does_not_recover_multi_topic_source_reply() -> None:
    topic = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "关联标准项": "CZ-HQ-TABLET-WRONG | 自动重启/关机",
            "适用范围": "平板电脑",
            "主标题": "苹果平板验机工具读出有重启记录时如何判定？",
            "主题对象/部位": "系统",
            "主题异常现象": "验机工具读出有重启记录iOS",
            "模型初标错误类型": "正文内容与主题核心问题不匹配",
            "模型初标原因": "标准与来源事实不一致。",
            "模型初标标准一致性": "不一致",
            "主题事实证据包": json.dumps(
                {
                    "representative_facts": [
                        {
                            "historical_actual_reply": (
                                "1. 关于屏幕检测，先查看工具是否显示正常，再按屏幕结果处理。\n"
                                "2. 关于重启记录，再查看验机工具是否读出有重启记录iOS。"
                            ),
                        }
                    ]
                },
                ensure_ascii=False,
            ),
        },
        use_standard_references=True,
    )

    assert "当前标准引用已撤销" in topic["知识内容"]
    assert topic["推荐回复"] == ""


@pytest.mark.parametrize(
    ("row", "included", "excluded"),
    [
        (
            {
                "数据ID": "ATOMIC-REPLY-REBOOT",
                "工单ID": "ATOMIC-REPLY-REBOOT",
                "产品类型": "平板电脑",
                "核心问题": "平板电脑｜系统｜验机工具读出有重启记录iOS｜判定",
                "对象/部位": "系统",
                "异常现象": "验机工具读出有重启记录iOS",
                "判定目标": "有重启记录-ios",
                "解题方式": "按验机工具重启记录处理",
                "历史实际回复": (
                    "1. 关于屏幕检测，工具读出正常即可勾选无问题。\n"
                    "2. 关于重启记录，苹果机型验机工具读出有重启记录时，"
                    "按有重启记录-ios处理。"
                ),
            },
            "重启记录",
            "屏幕检测",
        ),
        (
            {
                "数据ID": "ATOMIC-REPLY-DUST",
                "工单ID": "ATOMIC-REPLY-DUST",
                "产品类型": "平板电脑",
                "核心问题": "平板电脑｜后置摄像头区域｜灰尘｜是否属于屏幕进灰",
                "对象/部位": "后置摄像头区域",
                "异常现象": "灰尘",
                "判定目标": "确认是否属于屏幕进灰",
                "解题方式": "检查屏幕显示区域",
                "历史实际回复": (
                    "1. 关于电池健康度，工具读数84%按对应区间处理。\n"
                    "2. 关于进灰，需确认灰尘是否位于点亮屏幕后可见的屏幕显示区域。"
                ),
            },
            "关于进灰",
            "电池健康度",
        ),
    ],
)
def test_atomic_source_fact_scopes_numbered_multi_topic_reply(
    row: dict[str, object],
    included: str,
    excluded: str,
) -> None:
    fact = workflow_module._topic_source_fact(row, 1)

    assert included in fact["historical_actual_reply"]
    assert excluded not in fact["historical_actual_reply"]


def test_atomic_source_fact_scopes_human_fields_and_revoked_content() -> None:
    row = {
        "数据ID": "ATOMIC-HUMAN-REBOOT",
        "工单ID": "ATOMIC-HUMAN-REBOOT",
        "产品类型": "平板电脑",
        "核心问题": "平板电脑｜系统｜验机工具读出有重启记录iOS｜判定",
        "对象/部位": "系统",
        "异常现象": "有重启记录iOS",
        "判定目标": "有重启记录-ios",
        "解题方式": "验机工具读出有重启记录时按有重启记录-ios处理",
        "原始核心问题": (
            "回收师询问屏幕工具读出异常是否需要看闪光图；"
            "同时询问验机工具有重启记录iOS是否需要判定。"
        ),
        "原始判定结论": (
            "1. 屏幕问题：工具读出正常即可按正常处理，无需看闪光图。\n"
            "2. 重启记录问题：验机工具读出有重启记录iOS时，"
            "判定为有重启记录-ios。"
        ),
        "历史实际回复": (
            "1. 关于屏幕检测，工具读出正常即可勾选无问题，无需通过闪光图判断。\n"
            "2. 关于重启记录，苹果机型验机工具读出有重启记录时，"
            "按有重启记录-ios处理。"
        ),
    }

    fact = workflow_module._topic_source_fact(row, 1)
    assert "屏幕" not in fact["human_core_problem"]
    assert "屏幕" not in fact["human_judgment_conclusion"]
    assert "屏幕" not in fact["historical_actual_reply"]
    assert "重启记录" in fact["human_judgment_conclusion"]

    topic = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "关联标准项": "CZ-HQ-TABLET-WRONG | 屏幕异常",
            "适用范围": "平板电脑",
            "主标题": "苹果平板验机工具读出有重启记录时如何判定？",
            "主题对象/部位": "系统",
            "主题异常现象": "有重启记录iOS",
            "主题解题方式": "验机工具读出有重启记录时按有重启记录-ios处理",
            "模型初标错误类型": "标准项映射错误",
            "主题事实证据包": json.dumps(
                {"representative_facts": [fact]},
                ensure_ascii=False,
            ),
        },
        use_standard_references=True,
    )

    assert "屏幕" not in topic["知识内容"]
    assert "闪光图" not in topic["知识内容"]
    assert "重启记录" in topic["知识内容"]
    assert "屏幕" not in topic["推荐回复"]


def test_revoked_standard_content_cleans_unverified_claim_and_nested_numbering() -> None:
    topic = workflow_module._enforce_standard_reference_consistency(
        {
            "标准引用标签": "已引用标准知识点",
            "关联标准项": "CZ-HQ-TABLET-WRONG | 屏幕异常",
            "适用范围": "平板电脑",
            "主标题": "苹果平板验机工具读出有重启记录时如何判定？",
            "主题对象/部位": "系统",
            "主题异常现象": "有重启记录iOS",
            "主题解题方式": "验机工具读出有重启记录时按有重启记录-ios处理",
            "模型初标错误类型": "标准项映射错误",
            "主题事实证据包": json.dumps(
                {
                    "representative_facts": [
                        {
                            "historical_actual_reply": (
                                "2. 重启记录问题：请确认验机工具已读出“有重启记录iOS”。"
                                "依据质检标准，此情况应直接判定为“有重启记录-ios”。"
                            ),
                        }
                    ]
                },
                ensure_ascii=False,
            ),
        },
        use_standard_references=True,
    )

    assert topic["关联标准项"] == ""
    assert "依据质检标准" not in topic["知识内容"]
    assert "依据质检标准" not in topic["推荐回复"]
    assert topic["知识内容"].startswith("1. 重启记录问题：")
    assert "1. 2." not in topic["知识内容"]


def test_standard_handling_options_for_unreadable_battery_health_keep_only_final_option() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-BATTERY-HEALTH",
        title="苹果电池健康度如何分级判定",
        category_l1="基本情况",
        category_l2="电池健康度",
        knowledge_type="场景判定",
        standard_path=(
            "【基本情况】-【电池健康度】-【96%≤电池健康度≤100%】\n"
            "【基本情况】-【电池健康度】-【电池健康度无法检测】"
        ),
        keywords=["电池健康度", "无法检测"],
        scope="平板电脑-苹果",
        response_snippet=(
            "【判断与勾选】\n"
            "- 96%≤电池健康度≤100%：勾选【基本情况】-【电池健康度】-【96%≤电池健康度≤100%】。\n"
            "- 本机、验机工具和 Apple 支持均无法取得电池健康度时，"
            "勾选【基本情况】-【电池健康度】-【电池健康度无法检测】。"
        ),
        status="published",
        version="test",
    )

    options = workflow_module._standard_handling_options(
        standard,
        {
            "产品类型": "平板电脑",
            "核心问题": "本机和验机工具均无法读取电池健康度如何处理",
            "对象/部位": "电池健康度",
            "异常现象": "本机和工具无法读取",
        },
    )

    assert options == ["勾选【基本情况】-【电池健康度】-【电池健康度无法检测】"]


def test_standard_handling_options_exclude_section_headings_and_platform_labels() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-REBOOT",
        title="苹果有重启记录-ios判定方式",
        category_l1="基本情况",
        category_l2="开机情况",
        knowledge_type="场景判定",
        standard_path="【基本情况】-【开机情况】-【有重启记录-ios】",
        keywords=["重启记录"],
        scope="平板电脑-苹果",
        response_snippet=(
            "勾选项：【基本情况】-【开机情况】-【有重启记录-ios】\n"
            "【标准说明】\n"
            "- 【苹果】机型使用验机工具读出有重启记录。\n"
            "联动处理：有自动重启时，勾选"
            "【基本情况】-【开机情况】-【自动重启/关机】。"
        ),
        status="published",
        version="test",
    )

    assert workflow_module._standard_handling_options(standard) == [
        "勾选【基本情况】-【开机情况】-【有重启记录-ios】",
        "勾选【基本情况】-【开机情况】-【自动重启/关机】",
    ]


def test_multilevel_standard_does_not_use_single_option_condition_template() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-DAMAGE",
        title="最大直径如何分级判定",
        category_l1="外壳外观情况",
        category_l2="外壳磕碰/掉漆",
        knowledge_type="场景判定",
        standard_path="【外壳外观情况】-【外壳磕碰/掉漆】",
        keywords=["掉漆", "直径"],
        scope="平板电脑-通用",
        response_snippet=(
            "勾选项：【外壳外观情况】-【外壳磕碰/掉漆】-"
            "【最大直径≤6mm且2mm以上数量≤10】\n"
            "以下图片对应勾选项：【外壳外观情况】-【外壳磕碰/掉漆】-"
            "【最大直径>10mm或2mm以上数量>15】\n"
            "以下图片对应勾选项：【外壳外观情况】-【外壳磕碰/掉漆】-"
            "【最大直径≤10mm且2mm以上数量≤15】"
        ),
        status="published",
        version="test",
    )

    content = workflow_module._build_compact_standard_content(
        standard,
        workflow_module.CONTENT_TYPE_THRESHOLD,
    )

    assert not content.startswith("1. 满足以下任一条件时")
    assert "勾选【外壳外观情况】-【外壳磕碰/掉漆】-" not in content


def test_compact_knowledge_content_keeps_blank_lines_between_condition_items() -> None:
    content = workflow_module._compact_knowledge_content(
        "1. 满足以下任一条件时，勾选【外壳】-【其他】-【辅件缺损】：\n\n"
        "摄像头镜片碎裂或镜片整体缺失。\n\n"
        "防尘网罩破损。"
    )

    assert (
        content
        == "1. 满足以下任一条件时，勾选【外壳】-【其他】-【辅件缺损】：\n\n"
        "摄像头镜片碎裂或镜片整体缺失。\n\n"
        "防尘网罩破损。"
    )


def test_standard_content_turns_option_path_and_conditions_into_readable_rule() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-ACCESSORY",
        title="平板外壳辅件缺损如何处理",
        category_l1="外壳外观情况",
        category_l2="外壳其他现象",
        knowledge_type="场景判定",
        standard_path=(
            "【外壳外观情况】-【外壳其他现象】-"
            "【外壳辅件(镜片/防尘网罩等)缺损】"
        ),
        keywords=["镜片", "防尘网罩", "缺损"],
        scope="平板电脑-通用",
        response_snippet=(
            "勾选项：【外壳外观情况】-【外壳其他现象】-"
            "【外壳辅件(镜片/防尘网罩等)缺损】\n"
            "【标准说明】\n"
            "- 摄像头镜片碎裂或镜片整体缺失。\n"
            "- （注意：iPad镜头区域碎裂按图例\n"
            "- 防尘网罩（含扬声器、麦克风、听筒网）破损。\n"
            "- 闪光灯镜片碎裂或镜片整体缺失。\n"
            "- 机身按键键帽碎裂、缺失、脱落（包含home键）。"
        ),
        status="published",
        version="test",
    )

    content = workflow_module._build_compact_standard_content(
        standard,
        workflow_module.CONTENT_TYPE_DEFINITION,
    )

    assert content.startswith(
        "1. 满足以下任一条件时，勾选"
        "【外壳外观情况】-【外壳其他现象】-"
        "【外壳辅件(镜片/防尘网罩等)缺损】：\n\n"
    )
    assert "\n\n摄像头镜片碎裂或镜片整体缺失。" in content
    assert "\n\n防尘网罩（含扬声器、麦克风、听筒网）破损。" in content
    assert "\n2. 摄像头镜片碎裂或镜片整体缺失。" not in content
    assert "按图例" not in content
    assert "\n1. 【外壳外观情况】" not in content


def test_standard_content_uses_a_bare_leading_path_as_the_selectable_option() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-ACCESSORY-BARE",
        title="平板外壳辅件缺损如何处理",
        category_l1="外壳外观情况",
        category_l2="外壳其他现象",
        knowledge_type="场景判定",
        standard_path=(
            "【外壳外观情况】-【外壳其他现象】-"
            "【外壳辅件(镜片/防尘网罩等)缺损】"
        ),
        keywords=["镜片", "防尘网罩", "缺损"],
        scope="平板电脑-通用",
        response_snippet=(
            "【外壳外观情况】-【外壳其他现象】-"
            "【外壳辅件(镜片/防尘网罩等)缺损】\n"
            "摄像头镜片碎裂或镜片整体缺失。\n"
            "防尘网罩（含扬声器、麦克风、听筒网）破损。"
        ),
        status="published",
        version="test",
    )

    content = workflow_module._build_compact_standard_content(
        standard,
        workflow_module.CONTENT_TYPE_DEFINITION,
    )

    assert content.startswith(
        "1. 满足以下任一条件时，勾选"
        "【外壳外观情况】-【外壳其他现象】-"
        "【外壳辅件(镜片/防尘网罩等)缺损】：\n\n"
    )
    assert "\n\n摄像头镜片碎裂或镜片整体缺失。" in content
    assert "\n2. 摄像头镜片碎裂或镜片整体缺失。" not in content


def test_standard_content_for_unreadable_battery_health_keeps_reading_chain_only() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-BATTERY-HEALTH",
        title="苹果电池健康度如何分级判定",
        category_l1="基本情况",
        category_l2="电池健康度",
        knowledge_type="场景判定",
        standard_path="【基本情况】-【电池健康度】-【电池健康度无法检测】",
        keywords=["电池健康度", "无法检测", "Apple 支持"],
        scope="平板电脑-苹果",
        response_snippet=(
            "【标准说明】\n"
            "- 平板电池健康度优先以本机显示值为准，本机不显示则以验机工具读取值为准。\n"
            "- 若本机和验机工具均无法获取，则以 Apple 支持 App 诊断结果填写。\n"
            "- 以上方法均无效则勾选【基本情况】-【电池健康度】-【电池健康度无法检测】。\n"
            "【判断与勾选】\n"
            "- 96%≤电池健康度≤100%：勾选【基本情况】-【电池健康度】-【96%≤电池健康度≤100%】。\n"
            "- 81%≤电池健康度≤85%：勾选【基本情况】-【电池健康度】-【81%≤电池健康度≤85%】。\n"
            "- 本机、验机工具和 Apple 支持均无法取得电池健康度时，勾选【基本情况】-【电池健康度】-【电池健康度无法检测】。"
        ),
        status="published",
        version="test",
    )

    content = workflow_module._build_compact_standard_content(
        standard,
        workflow_module.CONTENT_TYPE_VERIFICATION,
        query={
            "产品类型": "平板电脑",
            "核心问题": "本机和验机工具均无法读取电池健康度如何处理",
            "对象/部位": "电池健康度",
            "异常现象": "本机和工具无法读取",
        },
    )

    assert "Apple 支持 App 诊断结果" in content
    assert "电池健康度无法检测" in content
    assert content.count("电池健康度无法检测") == 1
    assert "96%≤" not in content
    assert "81%≤" not in content


def test_configured_empty_standard_catalog_stops_before_transcription(tmp_path: Path) -> None:
    source_path = tmp_path / "source.xlsx"
    standard_path = tmp_path / "empty_standards.xlsx"

    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.append(["工单ID", "聊天内容", "产品类型"])
    source_sheet.append(["202608210003", "平板电池健康度如何读取？", "平板电脑"])
    source_book.save(source_path)

    standard_book = Workbook()
    standard_sheet = standard_book.active
    standard_sheet.append(["标准ID", "标准标题", "生效状态", "知识内容"])
    standard_book.save(standard_path)

    with pytest.raises(ValueError, match="没有读取到有效的生效标准"):
        workflow_module.initial_label_from_workbook(
            source_path=source_path,
            standards_path=standard_path,
            output_dir=tmp_path / "outputs",
            use_mimo=False,
            clustering_mode="rule",
        )


def test_configured_incomplete_standard_record_stops_before_transcription(
    tmp_path: Path,
) -> None:
    source_path = tmp_path / "source.xlsx"
    standard_path = tmp_path / "incomplete_standards.xlsx"

    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.append(["工单ID", "聊天内容", "产品类型"])
    source_sheet.append(["202608210004", "手机屏幕色斑如何判定？", "手机"])
    source_book.save(source_path)

    standard_book = Workbook()
    standard_sheet = standard_book.active
    standard_sheet.append(
        ["标准ID", "标准标题", "生效状态", "标准路径", "适用范围", "知识内容"]
    )
    standard_sheet.append(
        ["QC-INCOMPLETE-001", "手机屏幕色斑", "published", "", "手机", ""]
    )
    standard_book.save(standard_path)

    with pytest.raises(ValueError, match="字段不完整"):
        workflow_module.initial_label_from_workbook(
            source_path=source_path,
            standards_path=standard_path,
            output_dir=tmp_path / "outputs",
            use_mimo=False,
            clustering_mode="rule",
        )


def test_cz_tablet_knowledge_workbook_loads_as_referenceable_tablet_standards(
    tmp_path: Path,
) -> None:
    standard_path = tmp_path / "tablet.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "知识库主表"
    sheet.append(
        [
            "主标题",
            "副标题",
            "知识内容",
            "生效状态",
            "知识分类",
            "知识来源",
            "关联标准项",
            "适用范围",
            "适用类目",
            "来源版本",
            "主题键",
        ]
    )
    sheet.append(
        [
            "苹果平板电池健康度判定方式",
            "苹果平板电池健康度怎么读取",
            "1. 优先读取本机电池健康值。\n2. 无法读取时使用指定检测工具。",
            "生效中",
            "场景判定",
            "总部标准",
            "【电池】-【电池健康度】",
            "苹果",
            "平板电脑",
            "飞书平板标准-20260817",
            "电池-电池健康度::苹果",
        ]
    )
    sheet.append(
        [
            "待审核知识点",
            "",
            "待审核内容",
            "待审核",
            "场景判定",
            "总部标准",
            "【电池】-【待审核】",
            "通用",
            "平板电脑",
            "飞书平板标准-20260817",
            "电池-待审核::通用",
        ]
    )
    workbook.save(standard_path)

    items = load_standard_catalog(standard_path)

    assert len(items) == 1
    assert items[0].standard_id.startswith("CZ-HQ-TABLET-")
    assert items[0].standard_path == "【电池】-【电池健康度】"
    assert items[0].scope == "平板电脑-苹果"
    assert items[0].version == "飞书平板标准-20260817"


def test_tablet_review_queue_labels_standard_hit_and_missing_standard_for_review() -> None:
    standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-BATTERY",
        title="平板电池健康度判定方式",
        category_l1="电池",
        category_l2="电池健康度",
        knowledge_type="场景判定",
        standard_path="【电池】-【电池健康度】",
        keywords=["平板", "电池健康度", "读取"],
        scope="平板电脑-通用",
        response_snippet=(
            "1. 优先读取本机电池健康值。\n"
            "2. 无法读取时使用指定检测工具。\n"
            "3. 仍无法获取时转人工确认。"
        ),
        status="published",
        version="飞书平板标准-20260817",
    )
    source_row = {
        "数据ID": "TABLET-LABEL-001",
        "工单ID": "202608220001",
        "聊天内容": "平板电池健康度应该按什么顺序读取？",
        "核心问题": "平板电池健康度应该按什么顺序读取",
        "产品类型": "平板电脑",
        "一级分类": "电池",
        "二级分类": "电池健康度",
        "问题意图": "检测核验",
        "对象/部位": "电池健康度",
        "异常现象": "读取顺序不明确",
        "解题方式": "按总部标准知识点核验读取顺序",
        "语义标注依据": "来源明确咨询平板电池健康度读取顺序。",
    }

    hit_topics, _mapping, _gaps, _pending = build_topic_review_rows(
        [source_row],
        standard_catalog=[standard],
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=True,
        transcribe_all_admitted_topics=True,
    )
    missing_topics, _mapping, _gaps, _pending = build_topic_review_rows(
        [{**source_row, "数据ID": "TABLET-LABEL-002", "工单ID": "202608220002"}],
        standard_catalog=[],
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=True,
        transcribe_all_admitted_topics=True,
    )

    assert hit_topics[0]["标准引用标签"] == "已引用标准知识点"
    assert hit_topics[0]["关联标准项"]
    assert missing_topics[0]["标准引用标签"] == "未引用标准-人工重点复核"
    assert missing_topics[0]["关联标准项"] == ""
    assert missing_topics[0]["是否重点复核"] == "是"


def test_tablet_placeholder_taxonomy_still_matches_only_compatible_platform_standard() -> None:
    apple_standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-SN-APPLE",
        title="苹果SN查看方式",
        category_l1="基本情况",
        category_l2="SN",
        knowledge_type="检测方法",
        standard_path="【基本情况】-【SN】",
        keywords=["苹果", "SN", "序列号"],
        scope="平板电脑-苹果",
        response_snippet="苹果平板打开设置、通用、关于本机查看序列号。",
        status="published",
        version="飞书平板标准-20260817",
    )
    android_standard = StandardCatalogItem(
        standard_id="CZ-HQ-TABLET-SN-ANDROID",
        title="安卓SN查看方式",
        category_l1="基本情况",
        category_l2="SN",
        knowledge_type="检测方法",
        standard_path="【基本情况】-【SN】",
        keywords=["安卓", "SN", "序列号", "努比亚", "红魔"],
        scope="平板电脑-安卓",
        response_snippet="安卓平板打开设置、关于本机查看序列号。",
        status="published",
        version="飞书平板标准-20260817",
    )
    source_row = {
        "数据ID": "TABLET-SN-001",
        "工单ID": "202608220003",
        "聊天内容": "努比亚红魔电竞平板的序列号在哪里查看？",
        "核心问题": "努比亚红魔电竞平板的SN在哪里查看",
        "产品类型": "平板电脑",
        "一级分类": "待确认",
        "二级分类": "待确认",
        "问题意图": "信息查询",
        "对象/部位": "SN",
        "异常现象": "SN",
        "解题方式": "进入系统设置查看序列号",
        "语义标注依据": "来源明确咨询努比亚红魔平板SN查看路径。",
        "_原子平台": "Android",
        "_原子品牌": "努比亚",
    }

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        [source_row],
        standard_catalog=[apple_standard, android_standard],
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=True,
        transcribe_all_admitted_topics=True,
    )

    assert topics[0]["标准引用标签"] == "已引用标准知识点"
    assert "CZ-HQ-TABLET-SN-ANDROID" in topics[0]["关联标准项"]
    assert "CZ-HQ-TABLET-SN-APPLE" not in topics[0]["关联标准项"]


def test_tablet_apple_android_and_universal_scope_semantics() -> None:
    def build_with_scope(
        scope: str,
        *,
        platform: str = "",
        brand: str = "",
        model: str = "",
        question: str = "平板SN在哪里查看",
    ) -> dict[str, object]:
        qualifier = scope.rsplit("-", 1)[-1]
        standard = StandardCatalogItem(
            standard_id=f"CZ-HQ-TABLET-SCOPE-{qualifier}",
            title="平板SN查看方式",
            category_l1="基本情况",
            category_l2="SN",
            knowledge_type="检测方法",
            standard_path="【基本情况】-【SN】",
            keywords=["平板", "SN", "序列号"],
            scope=scope,
            response_snippet="进入系统设置中的关于本机页面查看序列号。",
            status="published",
            version="飞书平板标准-20260817",
        )
        source_row = {
            "数据ID": f"TABLET-SCOPE-{qualifier}-{brand or platform or 'UNKNOWN'}",
            "工单ID": "202608220005",
            "聊天内容": f"{question}？",
            "核心问题": question,
            "产品类型": "平板电脑",
            "一级分类": "待确认",
            "二级分类": "待确认",
            "问题意图": "信息查询",
            "对象/部位": "SN",
            "异常现象": "SN",
            "解题方式": "进入系统设置查看序列号",
            "语义标注依据": "来源明确咨询平板SN查看路径。",
            "_原子平台": platform,
            "_原子品牌": brand,
            "_原子机型范围": model,
        }
        topics, _mapping, _gaps, _pending = build_topic_review_rows(
            [source_row],
            standard_catalog=[standard],
            use_mimo=False,
            clustering_mode="rule",
            use_standard_references=True,
            transcribe_all_admitted_topics=True,
        )
        return topics[0]

    apple_on_apple = build_with_scope(
        "平板电脑-苹果",
        platform="iOS",
        brand="Apple",
        model="iPad Air",
    )
    android_on_apple = build_with_scope(
        "平板电脑-安卓",
        platform="iOS",
        brand="Apple",
        model="iPad Air",
    )
    android_on_unlisted_non_apple = build_with_scope(
        "平板电脑-安卓",
        platform="Android",
        brand="星河",
        model="GalaxyPad X1",
    )
    android_on_explicit_non_apple_text = build_with_scope(
        "平板电脑-安卓",
        question="回收星河 Nebula X 电竞平板时如何查看SN",
    )
    universal_on_apple = build_with_scope(
        "平板电脑-通用",
        platform="iOS",
        brand="Apple",
        model="iPad Air",
    )
    universal_on_non_apple = build_with_scope(
        "平板电脑-通用",
        platform="Android",
        brand="星河",
        model="GalaxyPad X1",
    )
    apple_unknown = build_with_scope("平板电脑-苹果")
    android_unknown = build_with_scope("平板电脑-安卓")

    assert apple_on_apple["标准引用标签"] == "已引用标准知识点"
    assert android_on_apple["标准引用标签"] == "未引用标准-人工重点复核"
    assert android_on_unlisted_non_apple["标准引用标签"] == "已引用标准知识点"
    assert android_on_explicit_non_apple_text["标准引用标签"] == "已引用标准知识点"
    assert universal_on_apple["标准引用标签"] == "已引用标准知识点"
    assert universal_on_non_apple["标准引用标签"] == "已引用标准知识点"
    assert apple_unknown["标准引用标签"] == "未引用标准-人工重点复核"
    assert android_unknown["标准引用标签"] == "未引用标准-人工重点复核"


def test_tablet_camera_gap_prefers_matching_repair_standard_over_generic_damage(
    tmp_path: Path,
) -> None:
    standard_path = tmp_path / "tablet.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "知识库主表"
    sheet.append(
        [
            "主标题",
            "副标题",
            "知识内容",
            "生效状态",
            "知识分类",
            "知识来源",
            "关联标准项",
            "适用范围",
            "适用类目",
            "来源版本",
            "主题键",
        ]
    )
    sheet.append(
        [
            "外壳辅件缺损如何处理",
            "外壳镜片缺损怎么处理",
            "外壳辅件缺失或破损时按外壳缺损处理。",
            "生效中",
            "场景判定",
            "总部标准",
            "【外壳外观情况】-【外壳其他现象】-【外壳辅件缺损】",
            "通用",
            "平板电脑",
            "飞书平板标准-20260817",
            "外壳-辅件缺损::通用",
        ]
    )
    sheet.append(
        [
            "平板更换摄像头镜片",
            "平板后置摄像头镜片边缘有明显缝隙\n平板后置摄像头镜片边缘有溢胶",
            (
                "后置摄像头镜片边缘有明显缝隙、溢胶或尺寸不匹配时，"
                "勾选后摄镜片更换。"
            ),
            "生效中",
            "场景判定",
            "总部标准",
            "【拆修及浸液情况】-【其他零部件拆修】-【后摄镜片更换(有溢胶、缝隙)】",
            "通用",
            "平板电脑",
            "飞书平板手写-20260817",
            "拆修-后摄镜片更换::通用",
        ]
    )
    workbook.save(standard_path)
    catalog = load_standard_catalog(standard_path)
    source_row = {
        "数据ID": "TABLET-CAMERA-001",
        "工单ID": "202608220004",
        "聊天内容": "平板后置摄像头镜片边缘有缝隙，像胶水，应该怎么判？",
        "核心问题": "平板后置摄像头镜片边缘有缝隙如何判定",
        "产品类型": "平板电脑",
        "一级分类": "待确认",
        "二级分类": "待确认",
        "问题意图": "边界判定",
        "对象/部位": "摄像头",
        "异常现象": "疑似拆修痕迹",
        "解题方式": "核验后摄镜片边缘是否有缝隙或溢胶",
        "语义标注依据": "来源明确询问后置摄像头镜片缝隙。",
    }

    topics, _mapping, _gaps, _pending = build_topic_review_rows(
        [source_row],
        standard_catalog=catalog,
        use_mimo=False,
        clustering_mode="rule",
        use_standard_references=True,
        transcribe_all_admitted_topics=True,
    )

    assert topics[0]["标准引用标签"] == "已引用标准知识点"
    assert "平板更换摄像头镜片" in topics[0]["关联标准项"]
    assert "外壳辅件缺损如何处理" not in topics[0]["关联标准项"]
