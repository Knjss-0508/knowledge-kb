from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
import json

from openpyxl import Workbook, load_workbook
from streamlit.testing.v1 import AppTest

from answer_hub.transfer_analysis.analysis import (
    _assess_capability,
    associate_conversation,
    import_source_file,
    run_weekly_analysis,
    stratified_sample,
)
from answer_hub.transfer_analysis.collectors import (
    EndpointProfile,
    _ConfiguredPlaywrightSession,
)
from answer_hub.transfer_analysis.schema import (
    ANALYSIS_COLUMNS,
    DEFAULT_CAPABILITY_REGISTRY,
)
from answer_hub.transfer_analysis.store import TransferAnalysisStore


def _write_workbook(path: Path, rows: list[dict]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    headers = list(rows[0])
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    workbook.save(path)


def test_stratified_sample_is_exact_and_marks_buckets() -> None:
    records = []
    for index in range(120):
        records.append(
            {
                "transfer_id": f"t-{index}",
                "event_time": f"2026-07-{13 + index % 7:02d}T10:00:00",
                "transfer_reason": (
                    "答非所问" if index % 5 == 0 else "该问题没有相关知识"
                ),
                "category": "手机" if index % 2 == 0 else "笔记本",
            }
        )

    first = stratified_sample(records, 40, seed="week")
    second = stratified_sample(records, 40, seed="week")

    assert len(first) == 40
    assert [row["transfer_id"] for row in first] == [
        row["transfer_id"] for row in second
    ]
    assert {row["_sample_bucket"] for row in first} <= {
        "比例分层",
        "重点场景",
        "随机补充",
    }
    assert any(row["transfer_reason"] == "答非所问" for row in first)


def test_association_prefers_same_engineer_and_prior_session() -> None:
    transfer = {
        "event_time": "2026-07-16T12:00:00",
        "engineer": "张三",
    }
    candidates = [
        {
            "source_id": "wrong-engineer",
            "engineer": "李四",
            "ended_at": "2026-07-16T11:55:00",
        },
        {
            "source_id": "same-engineer",
            "engineer": "张三",
            "ended_at": "2026-07-16T11:30:00",
        },
        {
            "source_id": "future",
            "engineer": "张三",
            "ended_at": "2026-07-16T13:00:00",
        },
    ]

    result = associate_conversation(transfer, candidates)

    assert result["conversation"]["source_id"] == "same-engineer"
    assert result["confidence"] == "中"
    assert result["candidate_count"] == 3


def test_capability_does_not_treat_generic_images_as_supported() -> None:
    generic = _assess_capability(
        "请看图片判断这个接口是否损坏",
        [],
        ["https://example.invalid/a.jpg"],
        [],
        DEFAULT_CAPABILITY_REGISTRY,
    )
    memory_tool = _assess_capability(
        "请识别内存品牌",
        [],
        ["https://example.invalid/memory.jpg"],
        [],
        DEFAULT_CAPABILITY_REGISTRY,
    )
    laptop_tool = _assess_capability(
        "请识别笔记本型号",
        [],
        ["https://example.invalid/laptop.jpg"],
        [],
        DEFAULT_CAPABILITY_REGISTRY,
    )

    assert generic["solvable"] == "否"
    assert generic["tool_attribution"] == "百晓生能力不支持"
    assert memory_tool["solvable"] == "是"
    assert memory_tool["required_tool"] == "内存硬盘品牌识别工具"
    assert memory_tool["tool_attribution"] == "应调用工具未调用"
    assert laptop_tool["solvable"] == "不确定"
    assert laptop_tool["uncertain_tool_scope"] is True


def test_configured_collector_paginates_without_playwright_runtime() -> None:
    profile = EndpointProfile.from_dict(
        {
            "system": "manhattan",
            "base_url": "https://example.invalid",
            "operations": {
                "list": {
                    "method": "POST",
                    "path": "/records",
                    "items_path": "data.records",
                    "total_path": "data.total",
                    "page_size": 100,
                }
            },
        }
    )
    session = _ConfiguredPlaywrightSession(profile)
    calls = []

    def fake_request(operation, variables):
        calls.append(variables["page"])
        if variables["page"] == 1:
            return {"data": {"records": list(range(100)), "total": 205}}
        if variables["page"] == 2:
            return {"data": {"records": list(range(100, 200)), "total": 205}}
        return {"data": {"records": list(range(200, 205)), "total": 205}}

    session.request = fake_request
    rows = session.fetch_items("list", {}, paginate=True)

    assert len(rows) == 205
    assert calls == [1, 2, 3]


def test_rule_workflow_writes_diagnostics_only_in_remark(tmp_path: Path) -> None:
    manhattan_path = tmp_path / "manhattan.xlsx"
    bxs_path = tmp_path / "bxs.xlsx"
    standards_path = tmp_path / "standards.json"
    store = TransferAnalysisStore(tmp_path / "transfer.db")
    week_start = datetime(2026, 7, 13)

    manhattan_rows = []
    bxs_rows = []
    for index in range(4):
        work_order = f"WO-{index}"
        transfer_id = f"T-{index}"
        bxs_question = "手机电池健康度怎么判断"
        manual_question = (
            "屏幕破损应该怎么判定"
            if index == 0
            else "手机电池健康度低应该如何处理"
        )
        event_time = week_start + timedelta(days=index, hours=12)
        manhattan_rows.append(
            {
                "转人工ID": transfer_id,
                "工单ID": work_order,
                "转人工时间": event_time.isoformat(timespec="seconds"),
                "工程师": "工程师A",
                "转人工原因": "答非所问",
                "类目": "手机",
                "机型": "测试机型",
                "订单所处状态": "检测中",
                "聊天内容": f"工程师：{manual_question}\n客服：请检查相关标准。",
            }
        )
        bxs_rows.append(
            {
                "会话id": f"B-{index}",
                "工单ID": work_order,
                "开始时间": (event_time - timedelta(hours=1)).isoformat(timespec="seconds"),
                "结束时间": (event_time - timedelta(minutes=30)).isoformat(timespec="seconds"),
                "工程师": "工程师A",
                "聊天内容": (
                    f"工程师：{bxs_question}\n"
                    "百晓生：请根据电池健康度标准判断。"
                ),
                "意图识别结果": "电池检测",
                "召回知识": "电池健康度判断",
                "Top相似度": 0.9,
                "生产阈值": 0.75,
            }
        )
    _write_workbook(manhattan_path, manhattan_rows)
    _write_workbook(bxs_path, bxs_rows)
    standards_path.write_text(
        json.dumps(
            [
                {
                    "标准ID": "STD-BATTERY",
                    "主标题": "手机电池健康度判断",
                    "一级分类": "手机",
                    "检索关键词": "电池 健康度",
                    "参考话术": "根据电池健康度和充放电状态判断。",
                    "状态": "已发布",
                }
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    import_source_file(manhattan_path, "manhattan", store)
    import_source_file(bxs_path, "baixiaosheng", store)
    summary = run_weekly_analysis(
        store,
        "2026-07-13",
        standards_path,
        tmp_path / "output",
        sample_size=4,
        use_mimo=False,
    )

    rows = store.list_annotation_rows("2026-07-13")
    mismatch = next(row for row in rows if row["工单ID"] == "WO-0")
    workbook = load_workbook(summary["report_file"], read_only=True)

    assert len(rows) == 4
    assert mismatch["转人工原因(校正)"] == "其他"
    assert "前后问题不一致" in mismatch["备注"]
    assert "【诊断】" in mismatch["备注"]
    assert "诊断标签" not in ANALYSIS_COLUMNS
    assert workbook.sheetnames == [
        "转人工分析明细",
        "人工复核队列",
        "badcase清单",
        "知识补充候选",
        "召回质量分析",
        "工具调用问题",
        "周度统计",
        "责任方优化清单",
    ]
    workbook.close()


def test_streamlit_transfer_workspace_renders_without_exception() -> None:
    app = AppTest.from_file("streamlit_app.py")
    app.session_state["workspace_page"] = "转人工分析"
    app.run(timeout=30)

    assert not app.exception
    assert any(item.value == "数据导入与采集" for item in app.subheader)
    assert any(
        item.options == ["数据采集", "周度分析", "人工复核", "分析报告"]
        for item in app.segmented_control
    )
