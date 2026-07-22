from __future__ import annotations

from datetime import datetime
from html import escape
from io import BytesIO
import json
from pathlib import Path
import re
import sys
import tempfile
from typing import Any, Callable

import streamlit as st
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT / "src"))

from answer_hub.workflow import (  # noqa: E402
    CASE_KNOWLEDGE_COLUMNS,
    CLUSTER_VALIDATION_COLUMNS,
    ERROR_TYPES,
    KNOWLEDGE_MASTER_COLUMNS,
    KNOWLEDGE_REVIEW_EXTENSION_COLUMNS,
    REVIEW_DECISIONS,
    TOPIC_CANDIDATE_COLUMNS,
    TOPIC_REVIEW_COLUMNS,
    build_case_knowledge_rows,
    cluster_validation_from_workbook,
    evaluate_cluster_validation_rows,
    finalize_topic_review_rows,
    initial_label_from_workbook,
)
from answer_hub.standards_compiler import (  # noqa: E402
    compile_standard_catalog,
    discover_default_standard_sources,
    load_standard_source_manifest,
)
from answer_hub.product_taxonomy import configured_product_names  # noqa: E402
from answer_hub.auto_review import (  # noqa: E402
    AutoReviewPolicy,
    evaluate_auto_review_validation,
    partition_auto_review_candidates,
    teammate_validation_decision,
)
from answer_hub.cz_integration import (  # noqa: E402
    CzIntegrationAdapter,
    select_submittable_candidates,
)
from answer_hub.embedding import EmbeddingClient  # noqa: E402
from answer_hub.images import split_image_urls  # noqa: E402
from answer_hub.automation import (  # noqa: E402
    AUTOMATION_STAGES,
    list_automation_runs,
    run_automation_pipeline,
)
from answer_hub.transfer_analysis.ui import render_transfer_analysis  # noqa: E402


st.set_page_config(
    page_title="答疑知识自动化工作台",
    layout="wide",
    initial_sidebar_state="collapsed",
)


@st.cache_resource(show_spinner=False)
def _shared_embedding_client() -> EmbeddingClient | None:
    return EmbeddingClient.from_env()


@st.cache_resource(show_spinner=False)
def _shared_cz_adapter() -> CzIntegrationAdapter:
    return CzIntegrationAdapter()


ACTIVE_STANDARD_CATALOG = ROOT / "data" / "compiled_standards" / "active_standards.json"
STANDARD_SOURCE_MANIFEST = ROOT / "config" / "standard_sources.json"
PRODUCT_OPTIONS = ["全部", *configured_product_names()]


def _latest_existing_knowledge_path() -> Path | None:
    candidates = sorted(
        ROOT.glob("outputs/*/人工答疑知识库_*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def _compile_active_standard_catalog() -> dict[str, Any]:
    if STANDARD_SOURCE_MANIFEST.is_file():
        sources, active_sheets = load_standard_source_manifest(STANDARD_SOURCE_MANIFEST)
    else:
        sources = discover_default_standard_sources()
        active_sheets = None
    return compile_standard_catalog(
        sources,
        ACTIVE_STANDARD_CATALOG,
        active_sheets=active_sheets,
        existing_knowledge_path=_latest_existing_knowledge_path(),
    )


CLUSTER_REVIEW_DECISIONS = ("同一主题", "不同主题", "不确定")
CLUSTER_REVIEW_ERROR_TYPES = (
    "问题意图不同",
    "对象/部位不同",
    "异常现象不同",
    "判定标准不同",
    "处理方式不同",
    "证据不足",
    "原分类噪声",
    "其他",
)


st.markdown(
    """
    <style>
    :root {
      --ink: #172033;
      --muted: #667085;
      --line: #d9e1ea;
      --line-strong: #c5d0dc;
      --canvas: #f3f6fa;
      --surface: #ffffff;
      --soft: #f7f9fc;
      --accent: #176b87;
      --accent-dark: #12566d;
      --accent-soft: #e8f4f7;
      --indigo: #4f5d95;
      --warning: #a15c07;
      --warning-soft: #fff5e6;
      --danger: #b42318;
      --success: #087443;
    }
    #MainMenu, footer, header { visibility: hidden; }
    .stApp {
      background: var(--canvas);
    }
    .block-container {
      max-width: 1660px;
      padding: 1.2rem 2.1rem 3rem;
    }
    [data-testid="stMainBlockContainer"] {
      width: calc(100vw - 2rem) !important;
      max-width: 1900px !important;
      padding: 1rem 1.1rem 3rem !important;
    }
    [data-testid="stSidebar"] {
      background: var(--surface);
      border-right: 1px solid var(--line);
    }
    h1, h2, h3 {
      color: var(--ink);
      letter-spacing: 0;
    }
    h1 { font-size: 1.7rem !important; margin-bottom: 0.15rem !important; }
    h2 { font-size: 1.15rem !important; }
    h3 { font-size: 1rem !important; }
    .stCaption {
      color: var(--muted) !important;
      font-size: 0.82rem !important;
      line-height: 1.45;
    }
    .workspace-header {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 1.5rem;
      background: var(--surface);
      border: 1px solid var(--line);
      border-left: 4px solid var(--accent);
      border-radius: 8px;
      padding: 1.15rem 1.35rem 1.05rem;
      margin-bottom: 0.8rem;
      box-shadow: 0 3px 12px rgba(23, 32, 51, 0.04);
    }
    .workspace-header-copy {
      min-width: 0;
    }
    .workspace-header-meta {
      display: flex;
      flex-wrap: wrap;
      justify-content: flex-end;
      gap: 0.45rem;
      color: var(--muted);
      font-size: 0.76rem;
      white-space: nowrap;
    }
    .meta-pill {
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 0.32rem 0.58rem;
      background: #fbfcfe;
    }
    .page-heading {
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      gap: 1rem;
      margin: 1.35rem 0 0.85rem;
    }
    .page-heading h2 {
      margin: 0;
      font-size: 1.25rem !important;
    }
    .page-heading p {
      margin: 0.22rem 0 0;
      color: var(--muted);
      font-size: 0.88rem;
    }
    .section-label {
      color: var(--muted);
      font-size: 0.74rem;
      font-weight: 750;
      letter-spacing: 0.02em;
      margin: 0.7rem 0 0.42rem;
      text-transform: uppercase;
    }
    .section-label::before {
      content: "";
      display: inline-block;
      width: 4px;
      height: 12px;
      margin-right: 7px;
      vertical-align: -1px;
      border-radius: 2px;
      background: var(--accent);
    }
    .status-strip {
      display: flex;
      align-items: center;
      gap: 0.55rem;
      border: 1px solid var(--line);
      border-radius: 7px;
      background: var(--surface);
      padding: 0.65rem 0.8rem;
      color: var(--muted);
      font-size: 0.8rem;
      line-height: 1.35;
    }
    .status-dot {
      width: 8px;
      height: 8px;
      flex: 0 0 auto;
      border-radius: 50%;
      background: var(--accent);
    }
    .status-dot.warning { background: #d28a12; }
    .status-dot.success { background: var(--success); }
    .status-dot.danger { background: var(--danger); }
    .action-panel {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--surface);
      padding: 0.95rem 1rem 0.8rem;
      box-shadow: 0 2px 8px rgba(23, 32, 51, 0.03);
    }
    .action-panel .section-label {
      margin-top: 0;
    }
    [data-testid="stMetric"] {
      min-height: 92px;
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 0.8rem 0.9rem;
      box-shadow: 0 2px 8px rgba(23, 32, 51, 0.03);
    }
    [data-testid="stMetricValue"] { color: var(--ink); font-size: 1.5rem; }
    [data-testid="stMetricLabel"] { color: var(--muted); font-size: 0.82rem; }
    .stButton > button, .stDownloadButton > button {
      min-height: 2.45rem;
      border-radius: 6px;
      border-color: var(--line-strong);
      font-weight: 600;
      box-shadow: none;
    }
    .stButton > button[kind="primary"] {
      background: var(--accent);
      border-color: var(--accent);
    }
    .stButton > button[kind="primary"]:hover {
      background: var(--accent-dark);
      border-color: var(--accent-dark);
    }
    .stDownloadButton > button {
      background: var(--surface);
      color: var(--ink);
    }
    [data-testid="stFileUploader"] {
      border: 1px dashed var(--line-strong);
      border-radius: 7px;
      background: #fbfcfe;
      padding: 0.35rem 0.55rem 0.5rem;
    }
    [data-testid="stFileUploader"] section {
      padding: 0.45rem 0.55rem;
      border: 0;
      background: transparent;
    }
    [data-testid="stFileUploaderDropzone"] {
      min-height: 104px;
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 6px;
    }
    [data-testid="stDataFrame"] {
      border: 1px solid var(--line);
      border-radius: 7px;
      overflow: hidden;
      background: var(--surface);
      box-shadow: 0 2px 8px rgba(23, 32, 51, 0.03);
    }
    [data-testid="stTextInput"] input,
    [data-testid="stTextArea"] textarea,
    [data-testid="stSelectbox"] [data-baseweb="select"] > div {
      border-color: var(--line-strong);
      border-radius: 6px;
      background: var(--surface);
    }
    [data-testid="stDataFrame"] {
      margin-top: 0.25rem;
    }
    [data-testid="stTextArea"] textarea {
      background: #fbfcfe;
      border-color: var(--line);
      font-size: 0.94rem;
      line-height: 1.55;
    }
    [data-testid="stWidgetLabel"] p {
      font-size: 0.88rem;
    }
    [data-testid="stCheckbox"] label,
    [data-testid="stRadio"] label {
      color: var(--ink);
      font-size: 0.86rem;
    }
    [data-testid="stRadio"] > div {
      gap: 0;
      padding: 0.2rem;
      border: 1px solid var(--line);
      border-radius: 7px;
      background: var(--surface);
      width: fit-content;
    }
    [data-testid="stRadio"] label {
      padding: 0.35rem 0.85rem;
      border-radius: 5px;
    }
    [data-testid="stRadio"] label:has(input:checked) {
      background: var(--accent-soft);
      color: var(--accent-dark);
      font-weight: 700;
    }
    [data-testid="stTabs"] [data-baseweb="tab-list"] {
      gap: 0.15rem;
      border-bottom: 1px solid var(--line);
    }
    [data-testid="stTabs"] button[role="tab"] {
      height: 2.35rem;
      border-radius: 5px 5px 0 0;
      color: var(--muted);
      font-weight: 600;
    }
    [data-testid="stTabs"] button[aria-selected="true"] {
      color: var(--accent);
      border-bottom-color: var(--accent);
      background: var(--accent-soft);
    }
    [data-testid="stExpander"] {
      border-color: var(--line);
      border-radius: 7px;
      background: var(--surface);
    }
    .workspace-kicker {
      color: var(--accent);
      font-size: 0.72rem;
      font-weight: 700;
      letter-spacing: 0;
      text-transform: uppercase;
      margin-bottom: 0.35rem;
    }
    .workspace-title {
      color: var(--ink);
      font-size: 1.55rem;
      font-weight: 700;
      line-height: 1.2;
      margin: 0;
    }
    .workspace-subtitle {
      color: var(--muted);
      margin: 0.45rem 0 1.35rem;
      font-size: 0.92rem;
    }
    .cluster-progress-card {
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 0.85rem 0.95rem;
      margin-bottom: 0.75rem;
    }
    .cluster-progress-title {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 0.75rem;
      color: var(--ink);
      font-size: 0.96rem;
      font-weight: 700;
      margin-bottom: 0.35rem;
    }
    .cluster-progress-meta {
      color: var(--muted);
      font-size: 0.82rem;
      line-height: 1.45;
    }
    .cluster-pair-heading {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 0.75rem;
      margin-bottom: 0.55rem;
    }
    .cluster-pair-id {
      color: var(--ink);
      font-size: 1.12rem;
      font-weight: 750;
    }
    .cluster-pair-position {
      color: var(--muted);
      font-size: 0.8rem;
    }
    .cluster-tag-row {
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      gap: 0.4rem;
      margin-bottom: 0.65rem;
    }
    .cluster-neutral-tag {
      display: inline-flex;
      align-items: center;
      min-height: 1.65rem;
      padding: 0.18rem 0.55rem;
      border: 1px solid #dfe3e8;
      border-radius: 999px;
      background: #ffffff;
      color: #4b5563;
      font-size: 0.82rem;
      font-weight: 650;
    }
    .cluster-neutral-tag.status::before {
      content: "";
      width: 6px;
      height: 6px;
      margin-right: 0.38rem;
      border-radius: 50%;
      background: #6b7280;
    }
    .cluster-facts {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 0.55rem;
      margin-bottom: 0.75rem;
    }
    .cluster-fact {
      min-width: 0;
      border: 1px solid #e2e5e9;
      border-radius: 7px;
      background: #ffffff;
      padding: 0.62rem 0.72rem;
    }
    .cluster-fact span {
      display: block;
      color: #7a818b;
      font-size: 0.78rem;
      margin-bottom: 0.15rem;
    }
    .cluster-fact strong {
      display: block;
      overflow: hidden;
      color: #27313d;
      font-size: 1.06rem;
      font-weight: 700;
      text-overflow: ellipsis;
      white-space: nowrap;
    }
    .cluster-case-label {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 0.5rem;
      color: var(--ink);
      font-size: 0.9rem;
      font-weight: 700;
      margin-bottom: 0.3rem;
    }
    .cluster-case-meta {
      color: var(--muted);
      font-size: 0.74rem;
      line-height: 1.4;
      margin-bottom: 0.55rem;
    }
    .cluster-review-guide {
      border: 1px solid #e2e5e9;
      background: #f8f9fa;
      color: #4b5563;
      border-radius: 7px;
      padding: 0.7rem 0.8rem;
      font-size: 0.88rem;
      line-height: 1.55;
      margin-bottom: 0.75rem;
    }
    .cluster-model-card {
      border: 1px solid #e2e5e9;
      border-radius: 8px;
      background: #ffffff;
      padding: 0.85rem 0.9rem;
      margin-bottom: 0.75rem;
    }
    .cluster-model-card strong {
      color: #27313d;
    }
    .cluster-model-card .model-meta {
      color: #7a818b;
      font-size: 0.78rem;
      line-height: 1.45;
      margin: 0.3rem 0 0.7rem;
    }
    .cluster-model-card .model-block {
      border-top: 1px solid #edf0f2;
      padding-top: 0.62rem;
      margin-top: 0.62rem;
    }
    .cluster-model-card .model-label {
      color: #7a818b;
      font-size: 0.76rem;
      font-weight: 650;
      margin-bottom: 0.2rem;
    }
    .cluster-model-card p {
      color: #394452;
      font-size: 0.88rem;
      line-height: 1.6;
      margin: 0;
      white-space: normal;
      overflow-wrap: anywhere;
    }
    .conversation-card {
      border: 1px solid #dfe3e8;
      border-radius: 9px;
      background: #ffffff;
      padding: 0.8rem 0.85rem 0.85rem;
      margin-bottom: 0.75rem;
    }
    .conversation-card-header {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 0.8rem;
      padding-bottom: 0.62rem;
      border-bottom: 1px solid #edf0f2;
    }
    .conversation-card-title {
      color: #27313d;
      font-size: 1.02rem;
      font-weight: 750;
    }
    .conversation-card-meta {
      color: #7a818b;
      font-size: 0.78rem;
      line-height: 1.4;
      text-align: right;
    }
    .conversation-question {
      margin: 0.65rem 0 0.7rem;
      padding: 0.58rem 0.68rem;
      border: 1px solid #e7e9ec;
      border-radius: 7px;
      background: #fafafa;
      color: #394452;
      font-size: 0.92rem;
      line-height: 1.55;
    }
    .conversation-question span {
      color: #7a818b;
      font-size: 0.78rem;
      font-weight: 650;
      margin-right: 0.45rem;
    }
    .conversation-lane-labels {
      display: flex;
      justify-content: space-between;
      color: #8a9099;
      font-size: 0.76rem;
      font-weight: 650;
      padding: 0 0.2rem 0.4rem;
    }
    .conversation-transcript {
      max-height: 520px;
      overflow-y: auto;
      border-radius: 7px;
      background: #f7f7f6;
      padding: 0.9rem 0.85rem;
      scrollbar-color: #cfd4da transparent;
      scrollbar-width: thin;
    }
    .conversation-row {
      display: flex;
      width: 100%;
      margin: 0.3rem 0;
    }
    .conversation-row.engineer {
      justify-content: flex-start;
      padding-right: 18%;
    }
    .conversation-row.answer {
      justify-content: flex-end;
      padding-left: 18%;
    }
    .conversation-row.system {
      justify-content: center;
      padding: 0 8%;
    }
    .conversation-bubble {
      max-width: 100%;
      border: 1px solid #dde1e5;
      border-radius: 9px;
      background: #ffffff;
      color: #303a46;
      padding: 0.48rem 0.62rem 0.52rem;
      font-size: 0.94rem;
      line-height: 1.65;
      overflow-wrap: anywhere;
      box-shadow: 0 1px 2px rgba(31, 41, 51, 0.025);
    }
    .conversation-row.answer .conversation-bubble {
      background: #f0f1f2;
      border-color: #d8dce0;
    }
    .conversation-row.system .conversation-bubble {
      max-width: 92%;
      border: 0;
      border-radius: 7px;
      background: #eceeed;
      color: #747b84;
      padding: 0.38rem 0.62rem;
      font-size: 0.78rem;
      text-align: center;
      box-shadow: none;
    }
    .conversation-time {
      color: #9aa0a8;
      font-size: 0.7rem;
      line-height: 1.2;
      margin-bottom: 0.16rem;
    }
    .conversation-row.answer .conversation-time {
      text-align: right;
    }
    .conversation-empty {
      color: #8a9099;
      font-size: 0.86rem;
      padding: 1rem;
      text-align: center;
    }
    .cluster-session-warning {
      color: var(--warning);
      background: var(--warning-soft);
      border: 1px solid #f2d7aa;
      border-radius: 6px;
      padding: 0.6rem 0.7rem;
      font-size: 0.82rem;
      line-height: 1.45;
      margin-top: 0.65rem;
    }
    [data-testid="stBaseButton-primary"] {
      background: #46515f !important;
      border-color: #46515f !important;
      color: #ffffff !important;
    }
    [data-testid="stBaseButton-primary"]:hover {
      background: #343e49 !important;
      border-color: #343e49 !important;
    }
    @media (max-width: 900px) {
      .block-container,
      [data-testid="stMainBlockContainer"] {
        width: 100% !important;
        padding: 0.9rem 0.8rem 2rem !important;
      }
      .workspace-header { display: block; }
      .workspace-header-meta {
        justify-content: flex-start;
        margin-top: 0.75rem;
      }
      .page-heading { display: block; }
      .workspace-title { font-size: 1.35rem; }
      .cluster-facts { grid-template-columns: 1fr; }
      .conversation-row.engineer { padding-right: 8%; }
      .conversation-row.answer { padding-left: 8%; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(item) for item in value if item not in (None, ""))
    return str(value)


def _load_topic_workbook(
    workbook_bytes: bytes,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    if "topic_review_queue" not in workbook.sheetnames:
        raise ValueError("工作簿缺少 topic_review_queue 工作表，请上传主题级候选工作簿。")

    def read_sheet(sheet_name: str) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            return []
        worksheet = workbook[sheet_name]
        values = list(worksheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value).strip() if value is not None else "" for value in values[0]]
        rows = []
        for row_index, values_row in enumerate(values[1:], start=2):
            row = {
                header: values_row[column_index] if column_index < len(values_row) else None
                for column_index, header in enumerate(headers)
                if header
            }
            if any(value not in (None, "") for value in row.values()):
                row["_review_row_index"] = row_index
                rows.append(row)
        return rows

    topic_rows = read_sheet("topic_review_queue")
    expected = TOPIC_CANDIDATE_COLUMNS + TOPIC_REVIEW_COLUMNS
    if topic_rows:
        missing = [field for field in expected if field not in topic_rows[0]]
        if missing:
            raise ValueError(f"主题复核工作簿缺少字段：{', '.join(missing)}")
    return (
        topic_rows,
        read_sheet("topic_source_mapping"),
        read_sheet("evidence_gap_rows"),
        read_sheet("pending_cluster_rows"),
        read_sheet("topic_model_drafts"),
    )


def _update_topic_workbook(workbook_bytes: bytes, changes: dict[int, dict[str, str]]) -> bytes:
    workbook = load_workbook(BytesIO(workbook_bytes))
    worksheet = workbook["topic_review_queue"]
    header_map = {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }
    for row_index, updates in changes.items():
        for field, value in updates.items():
            worksheet.cell(row=row_index, column=header_map[field], value=value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _rows_to_xlsx_bytes(sheet_name: str, columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(columns)
    for row in rows:
        worksheet.append(
            [
                "\n".join(str(item) for item in value) if isinstance((value := row.get(column, "")), list) else value
                for column in columns
            ]
        )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load_cluster_validation_workbook(workbook_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    if "聚类验证" not in workbook.sheetnames:
        raise ValueError("工作簿缺少“聚类验证”工作表。")
    worksheet = workbook["聚类验证"]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return []

    headers = [str(value).strip() if value is not None else "" for value in values[0]]
    required = {"验证对ID", "聚类预测", "记录A_核心问题", "记录B_核心问题"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"聚类验证工作簿缺少字段：{', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for values_row in values[1:]:
        source = {
            header: values_row[index] if index < len(values_row) else None
            for index, header in enumerate(headers)
            if header
        }
        if not any(value not in (None, "") for value in source.values()):
            continue
        row = {column: source.get(column, "") for column in CLUSTER_VALIDATION_COLUMNS}
        decision = _text(row.get("人工判断")).strip()
        row["人工判断"] = decision if decision in CLUSTER_REVIEW_DECISIONS else ""
        row["人工错误类型"] = _text(row.get("人工错误类型")).strip()
        for field in ("人工备注", "审核人", "审核时间"):
            value = row.get(field)
            row[field] = "" if isinstance(value, (int, float)) else _text(value).strip()
        rows.append(row)
    return rows


def _cluster_import_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    record_ids = {
        _text(row.get(field)).strip()
        for row in rows
        for field in ("记录A_ID", "记录B_ID")
        if _text(row.get(field)).strip()
    }
    thresholds = [
        row.get("聚类阈值")
        for row in rows
        if row.get("聚类阈值") not in (None, "")
    ]
    model_names = [
        _text(row.get("大模型名称")).strip()
        for row in rows
        if _text(row.get("大模型名称")).strip()
    ]
    return {
        "eligible_rows": len(record_ids),
        "cluster_count": "-",
        "candidate_pairs": len(rows),
        "validation_pairs": len(rows),
        "threshold": thresholds[0] if thresholds else "-",
        "embedding_model": "从验证工作簿导入",
        "large_model_enabled": bool(model_names),
    }


def _cluster_review_status(row: dict[str, Any]) -> str:
    decision = _text(row.get("人工判断")).strip()
    if decision not in CLUSTER_REVIEW_DECISIONS:
        return "待标注"
    if decision == "不确定":
        return "不确定"
    cluster_decision = _text(row.get("聚类预测")).strip()
    model_decision = _text(row.get("大模型判断")).strip()
    cluster_conflict = cluster_decision in CLUSTER_REVIEW_DECISIONS[:2] and cluster_decision != decision
    model_conflict = model_decision in CLUSTER_REVIEW_DECISIONS[:2] and model_decision != decision
    if cluster_conflict and model_conflict:
        return "双重冲突"
    if cluster_conflict:
        return "与聚类不一致"
    if model_conflict:
        return "与模型不一致"
    return "已标注"


def _cluster_review_indices(
    rows: list[dict[str, Any]],
    keyword: str,
    status_filter: str,
    sample_type_filter: str,
) -> list[int]:
    query = keyword.strip().lower()
    result: list[int] = []
    for index, row in enumerate(rows):
        status = _cluster_review_status(row)
        if status_filter == "待标注" and status != "待标注":
            continue
        if status_filter == "已标注" and status == "待标注":
            continue
        if status_filter == "不确定" and status != "不确定":
            continue
        if status_filter == "与聚类不一致" and status not in {"与聚类不一致", "双重冲突"}:
            continue
        if status_filter == "与模型不一致" and status not in {"与模型不一致", "双重冲突"}:
            continue
        if sample_type_filter != "全部类型" and _text(row.get("样本类型")) != sample_type_filter:
            continue
        if query:
            search_text = " ".join(
                _text(row.get(field))
                for field in (
                    "验证对ID",
                    "记录A_ID",
                    "记录B_ID",
                    "记录A_核心问题",
                    "记录B_核心问题",
                    "记录A_一级分类",
                    "记录A_二级分类",
                    "记录B_一级分类",
                    "记录B_二级分类",
                    "记录A_图片证据摘要",
                    "记录B_图片证据摘要",
                    "记录A_主题标签",
                    "记录B_主题标签",
                    "记录A_语义标注依据",
                    "记录B_语义标注依据",
                    "大模型主题",
                    "人工备注",
                )
            ).lower()
            if query not in search_text:
                continue
        result.append(index)
    return result


def _split_cluster_error_types(value: Any) -> list[str]:
    parts = re.split(r"[、,，;\n]+", _text(value))
    return [part.strip() for part in parts if part.strip() in CLUSTER_REVIEW_ERROR_TYPES]


def _next_unreviewed_cluster_index(rows: list[dict[str, Any]], current_index: int) -> int:
    if not rows:
        return 0
    for offset in range(1, len(rows) + 1):
        index = (current_index + offset) % len(rows)
        if _cluster_review_status(rows[index]) == "待标注":
            return index
    return min(current_index + 1, len(rows) - 1)


def _jsonl_bytes(rows: list[dict[str, Any]]) -> bytes:
    return "".join(json.dumps(row, ensure_ascii=False, default=str) + "\n" for row in rows).encode("utf-8")


def _read_only_text(label: str, value: Any, height: int = 160, key: str | None = None) -> None:
    st.text_area(
        label,
        value=_text(value),
        height=height,
        disabled=True,
        label_visibility="visible",
        key=key,
    )


_CHAT_LINE_PATTERN = re.compile(
    r"^(?P<time>\d{2}/\d{2}/\d{2}\s+\d{2}:\d{2}:\d{2}(?::\d{2})?)\s*(?P<content>.*)$"
)


def _is_chat_system_message(content: str) -> bool:
    normalized = content.strip()
    if not normalized:
        return True
    return (
        normalized.startswith(("问题类型：", "Play Video", "Current Time", "Loaded:"))
        or normalized in {"预览", "[图片]", "[视频]", "图片", "视频"}
        or bool(re.fullmatch(r"\[图片\](?:\s*[×xX]\s*\d+)?", normalized))
        or "快捷回复" in normalized
    )


def _infer_chat_role(content: str, previous_role: str | None) -> str:
    normalized = content.strip()
    answer_phrases = (
        "稍等",
        "我看下",
        "我看一下",
        "看一下哈",
        "什么问题，描述一下",
        "什么问题 描述一下",
        "遇到了什么问题",
        "您好",
        "你好同事",
        "建议",
        "判定为",
        "按掉漆",
        "按划痕",
        "按标准",
        "选权益机",
        "不影响",
        "正常的",
        "可以正常",
        "无法激活不收",
        "对的",
        "是的",
        "需要录入",
        "已备注",
    )
    engineer_phrases = (
        "老师",
        "麻烦",
        "辛苦",
        "帮忙",
        "帮我",
        "请问",
        "怎么判",
        "判什么",
        "选哪个",
        "能否",
        "可以回收吗",
        "是什么情况",
        "算有",
        "正常吗",
        "可以吗",
        "是吧",
    )
    engineer_acknowledgements = {"好的", "好", "OK", "ok", "收到", "明白", "谢谢", "行"}
    short_answers = {"可以", "不可以", "正常", "不正常", "需要", "不用", "不收", "收", "对", "是"}

    if any(phrase in normalized for phrase in answer_phrases):
        return "answer"
    if any(phrase in normalized for phrase in engineer_phrases):
        return "engineer"
    if (
        normalized in engineer_acknowledgements
        or normalized.startswith(("好的", "好嘞"))
        or "谢谢" in normalized
    ):
        return "engineer"
    if normalized in short_answers and previous_role == "engineer":
        return "answer"
    if normalized.endswith(("？", "?")) or re.search(r"(吗|么|呢|呀|啊)$", normalized):
        return "engineer"
    return previous_role or "engineer"


def _parse_conversation_messages(value: Any) -> list[dict[str, str]]:
    messages: list[dict[str, str]] = []
    for raw_line in _text(value).splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = _CHAT_LINE_PATTERN.match(line)
        if match:
            messages.append(
                {
                    "time": match.group("time"),
                    "content": match.group("content").strip(),
                }
            )
        else:
            messages.append({"time": "", "content": line})

    previous_role: str | None = None
    for message in messages:
        content = message["content"]
        if _is_chat_system_message(content):
            message["role"] = "system"
            continue
        role = _infer_chat_role(content, previous_role)
        message["role"] = role
        previous_role = role
    return messages


def _conversation_html_text(value: Any) -> str:
    return escape(_text(value), quote=True).replace("\n", "<br>")


def _render_conversation_case(
    *,
    label: str,
    record_id: Any,
    category_l1: Any,
    category_l2: Any,
    core_question: Any,
    chat_content: Any,
) -> None:
    messages = _parse_conversation_messages(chat_content)
    message_html: list[str] = []
    for message in messages:
        role = message.get("role", "system")
        time_html = (
            f"<div class='conversation-time'>{_conversation_html_text(message.get('time'))}</div>"
            if message.get("time")
            else ""
        )
        message_html.append(
            (
                f"<div class='conversation-row {role}'>"
                "<div class='conversation-bubble'>"
                f"{time_html}{_conversation_html_text(message.get('content'))}"
                "</div></div>"
            )
        )
    if not message_html:
        message_html.append("<div class='conversation-empty'>当前记录没有可展示的聊天内容</div>")

    category = " / ".join(
        value
        for value in (
            _text(category_l1).strip(),
            _text(category_l2).strip(),
        )
        if value
    )
    st.markdown(
        (
            "<div class='conversation-card'>"
            "<div class='conversation-card-header'>"
            f"<div class='conversation-card-title'>{_conversation_html_text(label)}</div>"
            "<div class='conversation-card-meta'>"
            f"{_conversation_html_text(record_id)}"
            f"{' · ' + _conversation_html_text(category) if category else ''}"
            "</div></div>"
            "<div class='conversation-question'>"
            "<span>核心问题</span>"
            f"{_conversation_html_text(core_question) or '未填写'}"
            "</div>"
            "<div class='conversation-lane-labels'><span>工程师</span><span>答疑回复</span></div>"
            f"<div class='conversation-transcript'>{''.join(message_html)}</div>"
            "</div>"
        ),
        unsafe_allow_html=True,
    )


def _render_cluster_media_evidence(row: dict[str, Any], prefix: str) -> None:
    image_urls = split_image_urls(_text(row.get(f"{prefix}_图片链接")))
    video_urls = split_image_urls(_text(row.get(f"{prefix}_视频链接")))
    image_status = _text(row.get(f"{prefix}_图片处理状态")) or "未记录"
    video_status = _text(row.get(f"{prefix}_视频处理状态")) or (
        "存在视频，当前未解析视频内容" if video_urls else "无视频链接"
    )
    image_summary = _text(row.get(f"{prefix}_图片证据摘要"))
    topic_tags = _text(row.get(f"{prefix}_主题标签"))
    semantic_reason = _text(row.get(f"{prefix}_语义标注依据"))
    image_necessity = _text(row.get(f"{prefix}_图片必要性")) or "待确认"

    with st.expander(
        "媒体证据与模型语义",
        expanded=bool(image_urls or video_urls),
        icon=":material/perm_media:",
    ):
        st.caption(
            f"图片处理：{image_status} · 视频处理：{video_status} · 图片必要性：{image_necessity}"
        )
        if image_urls:
            try:
                st.image(
                    image_urls[:4],
                    caption=[f"现场图片 {index}" for index in range(1, min(len(image_urls), 4) + 1)],
                    width=220,
                )
            except Exception as exc:
                st.caption(f"图片加载失败：{exc}")
        else:
            st.caption("没有可展示的图片链接。")

        if video_urls:
            st.warning("视频可供人工播放，但当前 MiMo 聚类链路尚未解析视频画面或声音。")
            for index, video_url in enumerate(video_urls[:2], start=1):
                st.caption(f"现场视频 {index}")
                try:
                    st.video(video_url)
                except Exception as exc:
                    st.caption(f"视频加载失败：{exc}")

        if image_summary:
            st.markdown("**图片证据摘要**")
            st.write(image_summary)
        if topic_tags:
            st.markdown("**模型主题标签**")
            st.write(topic_tags)
        if semantic_reason:
            st.markdown("**语义标注依据**")
            st.write(semantic_reason)


def _reset_editor(row: dict[str, Any]) -> None:
    st.session_state.selected_topic_row = row["_review_row_index"]
    for field in [
        *KNOWLEDGE_MASTER_COLUMNS,
        *KNOWLEDGE_REVIEW_EXTENSION_COLUMNS,
        *TOPIC_REVIEW_COLUMNS,
    ]:
        st.session_state[f"topic_{field}"] = _text(row.get(field))
    st.session_state["topic_知识ID"] = _text(row.get("知识ID") or row.get("主题ID"))
    st.session_state["topic_图例"] = _text(row.get("图例") or row.get("主题图片链接"))
    st.session_state["topic_关键词"] = _text(row.get("关键词") or row.get("检索关键词"))
    st.session_state["topic_是否值得沉淀"] = (
        _text(row.get("是否值得沉淀")) or "未标注"
    )


def _selected_row(rows: list[dict[str, Any]], row_index: int | None) -> dict[str, Any] | None:
    return next((row for row in rows if row["_review_row_index"] == row_index), None)


def _filtered_rows(rows: list[dict[str, Any]], keyword: str, decision: str, focus_only: bool) -> list[dict[str, Any]]:
    query = keyword.strip().lower()
    result = []
    for row in rows:
        row_decision = teammate_validation_decision(row)
        if decision == "未审核" and row_decision:
            continue
        if decision and decision != "未审核" and row_decision != decision:
            continue
        if focus_only and _text(row.get("模型初标重点复核")) != "是":
            continue
        if query and not any(
            query in _text(row.get(field)).lower()
            for field in ("主题ID", "主标题", "主题来源记录ID", "关联标准项", "知识分类")
        ):
            continue
        result.append(row)
    return result


def _mapping_for_topic(mapping_rows: list[dict[str, Any]], topic_id: str) -> list[dict[str, Any]]:
    return [row for row in mapping_rows if _text(row.get("主题ID")) == topic_id]


def _page_heading(title: str, description: str) -> None:
    st.markdown(
        f"""
        <div class="page-heading">
          <div>
            <h2>{title}</h2>
            <p>{description}</p>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _automation_stage_status(status: str) -> tuple[str, str]:
    return {
        "pending": ("等待执行", "gray"),
        "running": ("执行中", "blue"),
        "completed": ("已完成", "green"),
        "failed": ("失败", "red"),
    }.get(status, (status or "未知", "gray"))


def _automation_metrics_text(metrics: dict[str, Any]) -> str:
    labels = {
        "source_rows": "原始记录",
        "standards": "有效标准",
        "selected_rows": "品类入选",
        "eligible_rows": "可处理",
        "excluded_rows": "排除",
        "feature_rows": "语义特征",
        "model_labeled_rows": "模型标注",
        "topic_rows": "主题候选",
        "evidence_gap_rows": "证据缺口",
        "pending_cluster_rows": "待聚合",
    }
    parts = [
        f"{labels[key]} {value}"
        for key, value in metrics.items()
        if key in labels
    ]
    return " · ".join(parts)


def _automation_datetime(value: Any) -> datetime | None:
    text = _text(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _render_automation_stages(manifest: dict[str, Any]) -> None:
    stages = manifest.get("stages") or []
    for start in range(0, len(stages), 3):
        columns = st.columns(3, gap="large")
        for column, stage in zip(columns, stages[start : start + 3]):
            with column:
                with st.container(border=True, height="stretch"):
                    status_label, color = _automation_stage_status(_text(stage.get("status")))
                    st.badge(status_label, color=color)
                    st.markdown(f"**{stage.get('label', '')}**")
                    st.caption(_text(stage.get("detail")) or "等待上游阶段完成。")
                    metrics_text = _automation_metrics_text(stage.get("metrics") or {})
                    if metrics_text:
                        st.caption(metrics_text)


def _automation_artifact_path(
    manifest: dict[str, Any],
    artifact_key: str,
) -> Path | None:
    raw_path = _text((manifest.get("artifacts") or {}).get(artifact_key))
    if not raw_path:
        return None
    path = Path(raw_path)
    if not path.is_file():
        return None
    run_dir = Path(_text(manifest.get("run_dir"))).resolve()
    resolved = path.resolve()
    if not resolved.is_relative_to(run_dir):
        return None
    return resolved


def _render_automation_artifacts(manifest: dict[str, Any]) -> None:
    artifact_specs = [
        ("topic_review", "下载主题审核队列", "topic_review_queue.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("candidate_knowledge", "下载组员标注表", "candidate_knowledge.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("summary", "下载运行摘要", "summary.json", "application/json"),
    ]
    available = [
        (artifact_key, label, file_name, mime, path)
        for artifact_key, label, file_name, mime in artifact_specs
        if (path := _automation_artifact_path(manifest, artifact_key)) is not None
    ]
    if not available:
        return
    columns = st.columns(len(available))
    for column, (_artifact_key, label, file_name, mime, path) in zip(columns, available):
        column.download_button(
            label,
            data=path.read_bytes(),
            file_name=file_name,
            mime=mime,
            width="stretch",
        )


def _render_automation() -> None:
    _page_heading(
        "自动化看板",
        "将脱敏会话自动完成清洗、语义标注、主题聚类和知识转写，最终进入人工审核队列。",
    )
    automation_root = ROOT / "outputs" / "automation-runs"
    history = list_automation_runs(automation_root, limit=30)
    completed_runs = sum(run.get("status") == "review_pending" for run in history)
    failed_runs = sum(run.get("status") == "failed" for run in history)
    latest_topics = int((history[0].get("summary") or {}).get("topic_rows", 0)) if history else 0
    with st.container(horizontal=True):
        st.metric("自动化运行", len(history), border=True)
        st.metric("进入待审核", completed_runs, border=True)
        st.metric("运行失败", failed_runs, border=True)
        st.metric("最近主题候选", latest_topics, border=True)

    st.info(
        "当前自动化边界：生成待审核知识并保留全链路记录；人工确认后可提交 cz 待审核队列，但不会自动发布。",
        icon=":material/info:",
    )
    st.success(
        "当前使用无标准引用模式：候选知识只依据第二部分脱敏会话、历史回复和案例图生成。",
        icon=":material/check_circle:",
    )
    st.caption("原标准关联流程已保留在代码中，但本轮运行不会读取、检索或输出质检标准。")

    st.markdown("<div class='section-label'>启动一次自动化运行</div>", unsafe_allow_html=True)
    with st.form("automation_run_form"):
        product_label = st.selectbox(
            "处理品类",
            PRODUCT_OPTIONS,
            key="automation_product_type",
            help="选择“全部”时，系统仍会按产品类型隔离聚类，不会跨品类合并。",
        )
        upload_left, upload_right = st.columns([1.35, 0.65], gap="large")
        with upload_left:
            source_file = st.file_uploader(
                "脱敏会话 Excel",
                type=["xlsx"],
                key="automation_source_file",
                help="字段沿用方向二共享数据格式。",
            )
        with upload_right:
            st.markdown("**知识生成依据**")
            st.write("完整会话、历史实际回复、案例图片")
            st.caption("当前流程不新增标准关联；已有值会保留并单独搁置。")

        config_left, config_middle, config_right = st.columns([1, 1, 1.2], gap="large")
        with config_left:
            clustering_mode = st.selectbox(
                "聚类方式",
                ["direct_mimo", "semantic_mimo", "semantic", "rule"],
                format_func=lambda value: {
                    "direct_mimo": "纯大模型 1～N 聚类（推荐）",
                    "semantic_mimo": "语义标签 + 大模型裁决",
                    "semantic": "语义候选召回",
                    "rule": "本地规则回退",
                }[value],
                key="automation_clustering_mode",
            )
        with config_middle:
            semantic_threshold = st.slider(
                "语义相似度阈值",
                min_value=0.60,
                max_value=0.98,
                value=0.84,
                step=0.01,
                disabled=clustering_mode in {"direct_mimo", "rule"},
                key="automation_semantic_threshold",
            )
        with config_right:
            use_mimo = st.toggle(
                "调用 MiMo 进行语义标注和知识转写",
                value=True,
                key="automation_use_mimo",
                help="未配置模型时会自动回退为规则草稿并进入重点人工复核。",
            )

        with st.expander("高级聚类配置", icon=":material/tune:"):
            advanced_left, advanced_middle, advanced_right = st.columns(3)
            with advanced_left:
                cluster_review_floor = st.slider(
                    "进入裁决的最低相似度",
                    min_value=0.60,
                    max_value=0.90,
                    value=0.75,
                    step=0.01,
                    key="automation_cluster_review_floor",
                )
            with advanced_middle:
                cluster_auto_merge_threshold = st.slider(
                    "高置信自动合并阈值",
                    min_value=0.80,
                    max_value=0.98,
                    value=0.92,
                    step=0.01,
                    key="automation_cluster_auto_merge_threshold",
                )
            with advanced_right:
                cluster_review_limit = int(
                    st.number_input(
                        "单次最多裁决次数",
                        min_value=10,
                        max_value=500,
                        value=100,
                        step=10,
                        key="automation_cluster_review_limit",
                    )
                )
        submitted = st.form_submit_button(
            "启动自动化运行",
            type="primary",
            icon=":material/play_arrow:",
            disabled=not source_file,
            width="stretch",
        )

    if submitted and source_file:
        progress = st.progress(0.0, text="准备启动自动化流程")
        status_box = st.status("自动化流程运行中", expanded=True)
        seen_events: set[tuple[str, str, str]] = set()

        def on_progress(manifest: dict[str, Any]) -> None:
            stages = manifest.get("stages") or []
            completed = sum(stage.get("status") == "completed" for stage in stages)
            failed = any(stage.get("status") == "failed" for stage in stages)
            active = next(
                (
                    stage
                    for stage in stages
                    if stage.get("status") in {"running", "failed"}
                ),
                stages[-1] if stages else {},
            )
            ratio = min(1.0, completed / max(1, len(stages)))
            progress.progress(
                ratio,
                text=f"{active.get('label', '处理中')} · {active.get('detail', '')}",
            )
            event = (
                _text(active.get("id")),
                _text(active.get("status")),
                _text(active.get("detail")),
            )
            if event not in seen_events:
                status_label, _color = _automation_stage_status(event[1])
                status_box.write(f"**{active.get('label', '流程')}** · {status_label}：{event[2]}")
                seen_events.add(event)
            if failed:
                progress.progress(ratio, text="自动化流程执行失败")

        with tempfile.TemporaryDirectory(prefix="answer-hub-automation-") as temporary_dir:
            temporary_root = Path(temporary_dir)
            source_path = temporary_root / Path(source_file.name).name
            source_path.write_bytes(source_file.getvalue())
            manifest = run_automation_pipeline(
                source_path=source_path,
                standards_path=None,
                output_root=automation_root,
                product_type="" if product_label == "全部" else product_label,
                use_mimo=use_mimo,
                clustering_mode=clustering_mode,
                semantic_threshold=semantic_threshold,
                cluster_review_floor=cluster_review_floor,
                cluster_auto_merge_threshold=cluster_auto_merge_threshold,
                cluster_review_limit=cluster_review_limit,
                embedding_client=_shared_embedding_client(),
                progress_callback=on_progress,
            )
        st.session_state.automation_manifest = manifest
        if manifest.get("status") == "review_pending":
            progress.progress(1.0, text="自动化流程完成，已进入人工审核队列")
            status_box.update(label="自动化流程已完成", state="complete", expanded=False)
            topic_review_path = _automation_artifact_path(manifest, "topic_review")
            if topic_review_path:
                st.session_state.generated_topic_workbook = topic_review_path.read_bytes()
                st.session_state.generated_topic_summary = manifest.get("summary") or {}
                st.session_state.generated_topic_run_dir = _text(manifest.get("run_dir"))
            st.success("知识候选已进入待审核队列，可切换到“审核与反馈”继续处理。")
        else:
            status_box.update(label="自动化流程执行失败", state="error", expanded=True)
            st.error(f"运行失败：{_text(manifest.get('error')) or '请查看阶段记录'}")
        history = list_automation_runs(automation_root, limit=30)

    current_manifest = st.session_state.get("automation_manifest")
    if current_manifest:
        st.markdown("<div class='section-label'>本次运行链路</div>", unsafe_allow_html=True)
        _render_automation_stages(current_manifest)
        summary = current_manifest.get("summary") or {}
        if summary:
            with st.container(horizontal=True):
                st.metric("原始记录", summary.get("source_total_rows", 0), border=True)
                st.metric("可处理记录", summary.get("eligible_rows", 0), border=True)
                st.metric("主题候选", summary.get("topic_rows", 0), border=True)
                st.metric("证据缺口", summary.get("evidence_gap_rows", 0), border=True)
                st.metric("排除记录", summary.get("excluded_rows", 0), border=True)
        _render_automation_artifacts(current_manifest)
        st.caption(f"运行目录：{_text(current_manifest.get('run_dir'))}")

    st.markdown("<div class='section-label'>历史运行</div>", unsafe_allow_html=True)
    if history:
        st.dataframe(
            [
                {
                    "运行 ID": run.get("run_id", ""),
                    "创建时间": _automation_datetime(run.get("created_at")),
                    "状态": run.get("status_label", run.get("status", "")),
                    "会话文件": run.get("source_name", ""),
                    "可处理记录": (run.get("summary") or {}).get("eligible_rows", 0),
                    "主题候选": (run.get("summary") or {}).get("topic_rows", 0),
                    "证据缺口": (run.get("summary") or {}).get("evidence_gap_rows", 0),
                    "错误": run.get("error", ""),
                }
                for run in history
            ],
            hide_index=True,
            height=320,
            column_config={
                "运行 ID": st.column_config.TextColumn("运行 ID", pinned=True),
                "创建时间": st.column_config.DatetimeColumn("创建时间", format="YYYY-MM-DD HH:mm:ss"),
            },
        )
    else:
        st.caption("尚无自动化运行记录。")


def _run_generation(
    source_file: Any,
    standards_input: Any | None,
    product_type: str,
    use_mimo: bool,
    clustering_mode: str,
    semantic_threshold: float,
    cluster_review_floor: float,
    cluster_auto_merge_threshold: float,
    cluster_review_limit: int,
) -> tuple[bytes, dict[str, Any], Path]:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = ROOT / "outputs" / "topic-workbench" / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)
    source_path = run_dir / source_file.name
    source_path.write_bytes(source_file.getvalue())
    standards_path: Path | None
    if standards_input is None:
        standards_path = None
    elif isinstance(standards_input, Path):
        standards_path = standards_input
    else:
        standards_path = run_dir / standards_input.name
        standards_path.write_bytes(standards_input.getvalue())
    summary = initial_label_from_workbook(
        source_path=source_path,
        standards_path=standards_path,
        output_dir=run_dir,
        product_type=product_type,
        use_mimo=use_mimo,
        clustering_mode=clustering_mode,
        semantic_threshold=semantic_threshold,
        cluster_review_floor=cluster_review_floor,
        cluster_auto_merge_threshold=cluster_auto_merge_threshold,
        cluster_review_limit=cluster_review_limit,
        embedding_client=_shared_embedding_client(),
        use_standard_references=False,
    )
    review_path = Path(summary["topic_review_file"])
    return review_path.read_bytes(), summary, run_dir


def _run_cluster_validation(
    source_file: Any,
    product_type: str,
    semantic_threshold: float,
    max_pairs: int,
    use_mimo: bool,
    progress_callback: Callable[[str, int, int], None] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any], Path]:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = ROOT / "outputs" / "cluster-validation" / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)
    source_path = run_dir / source_file.name
    source_path.write_bytes(source_file.getvalue())
    rows, summary = cluster_validation_from_workbook(
        source_path=source_path,
        product_type=product_type,
        semantic_threshold=semantic_threshold,
        max_pairs=max_pairs,
        use_mimo=use_mimo,
        embedding_client=_shared_embedding_client(),
        progress_callback=progress_callback,
    )
    return rows, summary, run_dir


def _render_cluster_validation() -> None:
    _page_heading(
        "聚类标注工作台",
        "生成或导入聚类边界样本，在同一页面完成会话对比、模型复核、人工标注和结果导出。",
    )
    notice = st.session_state.pop("cluster_validation_notice", "")
    if notice:
        st.toast(notice, icon=":material/check_circle:")

    st.markdown("<div class='section-label'>任务来源</div>", unsafe_allow_html=True)
    task_source = st.segmented_control(
        "任务来源",
        ["生成新任务", "继续已有标注"],
        default="生成新任务",
        required=True,
        key="cluster_validation_task_source",
        label_visibility="collapsed",
        width="stretch",
        persist_state="session",
    )

    with st.container(border=True):
        if task_source == "生成新任务":
            source_file = st.file_uploader(
                "方向二数据 Excel",
                type=["xlsx"],
                key="cluster_validation_source",
                help="上传已脱敏的数据。该验证流程不生成知识正文。",
            )
            validation_product_type = st.selectbox(
                "验证品类",
                configured_product_names(),
                key="cluster_validation_product_type",
                help="聚类验证按叶子品类隔离，不进行跨品类合并。",
            )
            threshold_column, count_column, model_column = st.columns(
                [1, 0.85, 1.3],
                gap="large",
                vertical_alignment="bottom",
            )
            with threshold_column:
                semantic_threshold = st.slider(
                    "语义相似度阈值",
                    min_value=0.60,
                    max_value=0.98,
                    value=0.84,
                    step=0.01,
                    key="cluster_validation_threshold",
                )
            with count_column:
                max_pairs = st.number_input(
                    "验证样本对数量",
                    min_value=2,
                    max_value=100,
                    value=20,
                    step=2,
                    key="cluster_validation_pair_count",
                )
            with model_column:
                use_mimo = st.toggle(
                    "调用 MiMo 多模态语义标注",
                    value=True,
                    key="cluster_validation_use_mimo",
                    help="先复用第二部分的会话与图片分析，再独立判断两条记录是否属于同一主题；视频目前仅供人工播放。",
                )

            with st.container(horizontal=True, horizontal_alignment="right", vertical_alignment="center"):
                st.caption("生成任务需要 Embedding 服务；调用 MiMo 时会分析聊天与图片证据，视频暂由人工查看。")
                run_clicked = st.button(
                    "生成聚类验证任务",
                    type="primary",
                    disabled=source_file is None,
                    icon=":material/play_arrow:",
                )

            if run_clicked:
                progress_bar = st.progress(0, text="正在准备聚类验证数据...")
                stage_ranges = {
                    "semantic_labeling": (0, 48, "正在提取会话语义与图片证据"),
                    "embedding": (48, 76, "正在生成语义向量"),
                    "clustering": (76, 84, "正在执行语义聚类"),
                    "pair_sampling": (84, 92, "正在抽取边界验证样本"),
                    "large_model": (92, 99, "正在进行大模型独立标注"),
                }

                def update_progress(stage: str, completed: int, total: int) -> None:
                    start, end, label = stage_ranges.get(stage, (0, 99, "正在处理"))
                    fraction = min(1.0, max(0.0, completed / total)) if total else 1.0
                    percent = int(start + (end - start) * fraction)
                    progress_bar.progress(percent, text=f"{label}：{completed}/{total}")

                with st.spinner("正在生成向量、执行语义聚类并标注边界样本..."):
                    try:
                        rows, summary, run_dir = _run_cluster_validation(
                            source_file,
                            validation_product_type,
                            semantic_threshold,
                            int(max_pairs),
                            use_mimo,
                            progress_callback=update_progress,
                        )
                    except Exception as exc:
                        progress_bar.empty()
                        st.error(f"聚类验证生成失败：{exc}")
                        return
                progress_bar.progress(100, text="聚类验证任务生成完成")
                st.session_state.cluster_validation_rows = rows
                st.session_state.cluster_validation_summary = summary
                st.session_state.cluster_validation_run_dir = str(run_dir)
                st.session_state.cluster_validation_selected = 0
                st.session_state.cluster_validation_source_name = source_file.name
                st.session_state.cluster_validation_notice = "聚类验证任务已生成，可以开始人工标注。"
                st.rerun()
        else:
            review_file = st.file_uploader(
                "聚类验证工作簿",
                type=["xlsx"],
                key="cluster_validation_review_source",
                help="上传之前下载的 cluster_validation_review.xlsx，可继续上次标注。",
            )
            with st.container(horizontal=True, horizontal_alignment="right", vertical_alignment="center"):
                st.caption("导入已有任务不需要启动 Embedding 或 MiMo 服务。无效人工判断值会恢复为待标注。")
                import_clicked = st.button(
                    "载入标注任务",
                    type="primary",
                    disabled=review_file is None,
                    icon=":material/upload_file:",
                )
            if import_clicked:
                try:
                    imported_rows = _load_cluster_validation_workbook(review_file.getvalue())
                except Exception as exc:
                    st.error(f"载入聚类验证工作簿失败：{exc}")
                    return
                if not imported_rows:
                    st.warning("工作簿中没有可标注的验证样本。")
                    return
                st.session_state.cluster_validation_rows = imported_rows
                st.session_state.cluster_validation_summary = _cluster_import_summary(imported_rows)
                st.session_state.cluster_validation_selected = 0
                st.session_state.cluster_validation_source_name = review_file.name
                st.session_state.cluster_validation_notice = f"已载入 {len(imported_rows)} 组聚类验证样本。"
                st.rerun()

    rows: list[dict[str, Any]] = st.session_state.get("cluster_validation_rows", [])
    summary: dict[str, Any] = st.session_state.get("cluster_validation_summary", {})
    if not rows:
        st.info(
            "可以生成新的聚类验证任务，也可以导入已有 `cluster_validation_review.xlsx` 直接开始标注。",
            icon=":material/info:",
        )
        return

    for row in rows:
        row.setdefault("人工错误类型", "")

    evaluation = evaluate_cluster_validation_rows(rows)
    reviewed_pairs = int(evaluation.get("reviewed_pairs", 0))
    pending_pairs = int(evaluation.get("pending_pairs", len(rows)))
    progress_value = reviewed_pairs / len(rows) if rows else 0.0

    st.markdown("<div class='section-label'>标注进度</div>", unsafe_allow_html=True)
    metrics = st.columns(4)
    metrics[0].metric("验证样本对", len(rows))
    metrics[1].metric("已标注", reviewed_pairs)
    metrics[2].metric("待标注", pending_pairs)
    metrics[3].metric(
        "模型一致率",
        f"{evaluation['large_model_accuracy']:.1%}"
        if evaluation.get("large_model_accuracy") is not None
        else "-",
    )
    st.progress(progress_value, text=f"人工标注完成度：{reviewed_pairs}/{len(rows)}")
    accuracy_parts = [
        f"来源：{st.session_state.get('cluster_validation_source_name', '当前任务')}",
        f"Embedding：{summary.get('embedding_model', '-')}",
        f"阈值：{summary.get('threshold', '-')}",
        (
            f"聚类准确率：{evaluation['clustering_accuracy']:.1%}"
            if evaluation.get("clustering_accuracy") is not None
            else "聚类准确率：待标注"
        ),
        f"误合并：{evaluation.get('false_merge_pairs', 0)}",
        f"误拆分：{evaluation.get('false_split_pairs', 0)}",
        f"不确定：{evaluation.get('uncertain_pairs', 0)}",
    ]
    st.caption(" · ".join(accuracy_parts))

    st.markdown("<div class='section-label'>人工标注</div>", unsafe_allow_html=True)
    st.session_state.setdefault("cluster_validation_reviewer_name", "")
    sample_types = sorted({_text(row.get("样本类型")) for row in rows if _text(row.get("样本类型"))})
    with st.container(border=True):
        keyword_column, status_column, type_column, reviewer_column = st.columns(
            [1.35, 1, 1, 0.9],
            gap="medium",
            vertical_alignment="bottom",
        )
        with keyword_column:
            keyword = st.text_input(
                "搜索",
                key="cluster_validation_keyword",
                placeholder="验证对、工单、核心问题或分类",
                icon=":material/search:",
            )
        with status_column:
            status_filter = st.selectbox(
                "标注状态",
                ["全部", "待标注", "已标注", "不确定", "与聚类不一致", "与模型不一致"],
                key="cluster_validation_status_filter",
            )
        with type_column:
            sample_type_filter = st.selectbox(
                "样本类型",
                ["全部类型", *sample_types],
                key="cluster_validation_type_filter",
            )
        with reviewer_column:
            reviewer_name = st.text_input(
                "默认审核人",
                key="cluster_validation_reviewer_name",
                placeholder="姓名或工号",
            )

    filtered_indices = _cluster_review_indices(rows, keyword, status_filter, sample_type_filter)
    if not filtered_indices:
        st.warning("当前筛选条件下没有样本。")
        return

    selected = min(int(st.session_state.get("cluster_validation_selected", 0)), len(rows) - 1)
    if selected not in filtered_indices:
        selected = filtered_indices[0]
        st.session_state.cluster_validation_selected = selected

    queue_column, evidence_column, review_column = st.columns(
        [0.52, 2.45, 0.9],
        gap="large",
    )

    with queue_column:
        position = filtered_indices.index(selected)
        st.markdown(
            (
                "<div class='cluster-progress-card'>"
                "<div class='cluster-progress-title'><span>样本队列</span>"
                f"<span>{position + 1}/{len(filtered_indices)}</span></div>"
                f"<div class='cluster-progress-meta'>当前筛选 {len(filtered_indices)} 组；"
                f"全部任务已完成 {reviewed_pairs}/{len(rows)}。</div></div>"
            ),
            unsafe_allow_html=True,
        )
        queue_rows = [
            {
                "序号": index + 1,
                "状态": _cluster_review_status(rows[index]),
                "类型": _text(rows[index].get("样本类型")),
                "相似度": rows[index].get("语义相似度"),
                "模型": _text(rows[index].get("大模型判断")) or "-",
                "人工": _text(rows[index].get("人工判断")) or "待标注",
                "核心问题": (
                    f"{_text(rows[index].get('记录A_核心问题'))[:36]} ↔ "
                    f"{_text(rows[index].get('记录B_核心问题'))[:36]}"
                ),
            }
            for index in filtered_indices
        ]
        queue_key = f"cluster_validation_queue_{abs(hash((keyword, status_filter, sample_type_filter, tuple(filtered_indices))))}"
        queue_event = st.dataframe(
            queue_rows,
            key=queue_key,
            on_select="rerun",
            selection_mode="single-row",
            width="stretch",
            height=510,
            hide_index=True,
            row_height=44,
            column_order=["状态", "相似度", "人工", "核心问题"],
            column_config={
                "状态": st.column_config.TextColumn("状态", pinned=True, width="small"),
                "相似度": st.column_config.NumberColumn("相似度", format="%.3f", width="small"),
                "人工": st.column_config.TextColumn("人工", width="small"),
                "核心问题": st.column_config.TextColumn("核心问题", width="large"),
            },
        )
        if queue_event.selection.rows:
            queue_selected = filtered_indices[queue_event.selection.rows[0]]
            if queue_selected != selected:
                st.session_state.cluster_validation_selected = queue_selected
                st.rerun()

        previous_column, next_column = st.columns(2)
        if previous_column.button(
            "上一条",
            icon=":material/arrow_back:",
            disabled=position == 0,
            width="stretch",
        ):
            st.session_state.cluster_validation_selected = filtered_indices[position - 1]
            st.rerun()
        if next_column.button(
            "下一条",
            icon=":material/arrow_forward:",
            disabled=position >= len(filtered_indices) - 1,
            width="stretch",
        ):
            st.session_state.cluster_validation_selected = filtered_indices[position + 1]
            st.rerun()

        st.download_button(
            "下载当前标注进度",
            data=_rows_to_xlsx_bytes("聚类验证", CLUSTER_VALIDATION_COLUMNS, rows),
            file_name="cluster_validation_review.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            icon=":material/download:",
            width="stretch",
        )
        st.markdown(
            "<div class='cluster-session-warning'>标注暂存在当前浏览器会话中。关闭页面或重启服务前，请下载当前进度。</div>",
            unsafe_allow_html=True,
        )

    row = rows[selected]
    pair_id = _text(row.get("验证对ID"))
    status = _cluster_review_status(row)

    with evidence_column:
        st.markdown(
            (
                "<div class='cluster-pair-heading'>"
                f"<div class='cluster-pair-id'>{_conversation_html_text(pair_id)}</div>"
                f"<div class='cluster-pair-position'>全部样本第 {selected + 1}/{len(rows)} 组</div>"
                "</div>"
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            (
                "<div class='cluster-tag-row'>"
                f"<span class='cluster-neutral-tag status'>{_conversation_html_text(status)}</span>"
                f"<span class='cluster-neutral-tag'>{_conversation_html_text(row.get('样本类型'))}</span>"
                "</div>"
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            (
                "<div class='cluster-facts'>"
                "<div class='cluster-fact'><span>语义相似度</span>"
                f"<strong>{_conversation_html_text(row.get('语义相似度') or '-')}</strong></div>"
                "<div class='cluster-fact'><span>聚类预测</span>"
                f"<strong>{_conversation_html_text(row.get('聚类预测') or '-')}</strong></div>"
                "<div class='cluster-fact'><span>大模型判断</span>"
                f"<strong>{_conversation_html_text(row.get('大模型判断') or '-')}</strong></div>"
                "</div>"
            ),
            unsafe_allow_html=True,
        )

        _render_conversation_case(
            label="样本 A",
            record_id=row.get("记录A_ID"),
            category_l1=row.get("记录A_一级分类"),
            category_l2=row.get("记录A_二级分类"),
            core_question=row.get("记录A_核心问题"),
            chat_content=row.get("记录A_聊天内容"),
        )
        _render_cluster_media_evidence(row, "记录A")

        _render_conversation_case(
            label="样本 B",
            record_id=row.get("记录B_ID"),
            category_l1=row.get("记录B_一级分类"),
            category_l2=row.get("记录B_二级分类"),
            core_question=row.get("记录B_核心问题"),
            chat_content=row.get("记录B_聊天内容"),
        )
        _render_cluster_media_evidence(row, "记录B")

        with st.expander("标注判断规则", icon=":material/rule:"):
            st.markdown(
                """
                - **同一主题**：两条会话可以由同一篇知识准确回答，且问题意图、对象、异常现象和判定标准基本一致。
                - **不同主题**：仅共享“外观、拆修”等宽泛概念，或判断对象、标准边界、处理流程存在关键差异。
                - **不确定**：聊天不完整、关键图片缺失或证据冲突，暂时无法负责任地下结论。
                """
            )

    with review_column:
        st.markdown("<div class='section-label'>模型复核</div>", unsafe_allow_html=True)
        model_decision = _text(row.get("大模型判断")) or "未标注"
        model_reason = "\n".join(
            part
            for part in (
                _text(row.get("大模型原因")),
                _text(row.get("大模型关键差异")),
            )
            if part
        )
        st.markdown(
            (
                "<div class='cluster-model-card'>"
                f"<strong>结论：{_conversation_html_text(model_decision)}</strong>"
                "<div class='model-meta'>"
                f"置信度：{_conversation_html_text(row.get('大模型置信度') or '-')} · "
                f"模型：{_conversation_html_text(row.get('大模型名称') or '-')} · "
                f"状态：{_conversation_html_text(row.get('大模型状态') or '-')}"
                "</div>"
                "<div class='model-block'><div class='model-label'>建议主题</div>"
                f"<p>{_conversation_html_text(row.get('大模型主题') or '未提供')}</p></div>"
                "<div class='model-block'><div class='model-label'>判断原因与关键差异</div>"
                f"<p>{_conversation_html_text(model_reason or '未提供')}</p></div>"
                "</div>"
            ),
            unsafe_allow_html=True,
        )

        st.markdown("<div class='section-label'>人工结论</div>", unsafe_allow_html=True)
        st.markdown(
            "<div class='cluster-review-guide'>先独立阅读 A/B 会话，再参考模型建议。最终标准是：两条会话能否由同一篇知识准确回答。</div>",
            unsafe_allow_html=True,
        )
        with st.form(f"cluster_validation_form_{pair_id}", border=True):
            current_decision = _text(row.get("人工判断")).strip()
            human_decision = st.segmented_control(
                "人工判断",
                list(CLUSTER_REVIEW_DECISIONS),
                default=current_decision if current_decision in CLUSTER_REVIEW_DECISIONS else None,
                required=False,
                key=f"{pair_id}_human_decision",
                width="stretch",
            )
            error_types = st.pills(
                "关键差异或不确定原因",
                list(CLUSTER_REVIEW_ERROR_TYPES),
                selection_mode="multi",
                default=_split_cluster_error_types(row.get("人工错误类型")),
                key=f"{pair_id}_error_types",
                width="stretch",
                help="选择最主要的原因，可多选。",
            )
            human_note = st.text_area(
                "人工备注",
                value=_text(row.get("人工备注")),
                height=118,
                placeholder="例如：对象不同，A 是外壳，B 是镜头；需要引用不同的判定标准。",
                key=f"{pair_id}_human_note",
            )
            existing_reviewer = _text(row.get("审核人")).strip()
            st.caption(
                f"审核人：{reviewer_name or existing_reviewer or '未填写'} · "
                f"上次保存：{_text(row.get('审核时间')) or '尚未保存'}"
            )
            save_column, save_next_column = st.columns(2)
            save_only = save_column.form_submit_button(
                "保存",
                icon=":material/save:",
                width="stretch",
            )
            save_and_next = save_next_column.form_submit_button(
                "保存并下一条",
                type="primary",
                icon=":material/arrow_forward:",
                width="stretch",
            )

        if save_only or save_and_next:
            if human_decision not in CLUSTER_REVIEW_DECISIONS:
                st.warning("请先选择人工判断。")
            else:
                row["人工判断"] = human_decision
                row["人工错误类型"] = "、".join(error_types or [])
                row["人工备注"] = human_note.strip()
                row["审核人"] = reviewer_name or existing_reviewer
                row["审核时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                st.session_state.cluster_validation_rows = rows
                if save_and_next:
                    st.session_state.cluster_validation_selected = _next_unreviewed_cluster_index(rows, selected)
                st.session_state.cluster_validation_notice = f"已保存 {pair_id} 的人工标注。"
                st.rerun()

    current_evaluation = evaluate_cluster_validation_rows(rows)
    st.markdown("<div class='section-label'>导出与评估</div>", unsafe_allow_html=True)
    downloads = st.columns(3)
    downloads[0].download_button(
        "下载聚类验证工作簿",
        data=_rows_to_xlsx_bytes("聚类验证", CLUSTER_VALIDATION_COLUMNS, rows),
        file_name="cluster_validation_review.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        icon=":material/table_view:",
        width="stretch",
    )
    downloads[1].download_button(
        "下载人工反馈 JSONL",
        data=_jsonl_bytes(rows),
        file_name="cluster_validation_feedback.jsonl",
        mime="application/jsonl",
        icon=":material/data_object:",
        width="stretch",
    )
    downloads[2].download_button(
        "下载验证指标",
        data=json.dumps(current_evaluation, ensure_ascii=False, indent=2).encode("utf-8"),
        file_name="cluster_validation_metrics.json",
        mime="application/json",
        icon=":material/analytics:",
        width="stretch",
    )


def _render_generation() -> None:
    _page_heading(
        "生成主题候选",
        "以完整会话、历史实际回复和案例图生成主题标签，再聚合为无标准引用的可审核知识草稿。",
    )
    st.markdown("<div class='section-label'>输入材料</div>", unsafe_allow_html=True)
    product_label = st.selectbox(
        "处理品类",
        PRODUCT_OPTIONS,
        key="generation_product_type",
        help="选择“全部”时按当前生效品类分别聚类和转写。",
    )
    upload_left, upload_right = st.columns([1.35, 0.65], gap="large")
    with upload_left:
        source_file = st.file_uploader(
            "方向二数据 Excel",
            type=["xlsx"],
            key="source_file",
            help="上传已完成脱敏的方向二会话数据。",
        )
    with upload_right:
        st.markdown("**当前生成模式**")
        st.write("无标准引用")
        st.caption("不会读取标准目录；新候选不新增标准关联，已有值保留。")

    st.markdown("<div class='section-label'>执行配置</div>", unsafe_allow_html=True)
    mode_column, threshold_column, mimo_column = st.columns([0.9, 1.1, 1.35], gap="large")
    with mode_column:
        clustering_mode = st.selectbox(
            "聚类方式",
            ["direct_mimo", "semantic_mimo", "semantic", "rule"],
            format_func=lambda value: {
                "direct_mimo": "纯大模型 1～N 聚类（推荐）",
                "semantic_mimo": "模型语义标签 + 大模型裁决",
                "semantic": "模型语义标签 + 语义候选召回",
                "rule": "规则回退",
            }[value],
            help="推荐模式直接由 MiMo 拆分原子问题并做 1～N 聚类，不依赖 Embedding；原语义召回模式保留为备用。",
        )
    with threshold_column:
        semantic_threshold = st.slider(
            "语义相似度阈值",
            min_value=0.60,
            max_value=0.98,
            value=0.84,
            step=0.01,
            disabled=clustering_mode in {"direct_mimo", "rule"},
            help="阈值越高，主题拆分越细；阈值越低，越容易合并为同一主题。",
        )
    with mimo_column:
        use_mimo = st.checkbox(
            "调用 MiMo：会话语义标注、聚类裁决、主题转写与模型初标",
            value=True,
            help="会话语义标注会优先阅读完整聊天和可用图片；需要在本机 .env 配置 MiMo。",
        )
    cluster_review_floor = 0.75
    cluster_auto_merge_threshold = 0.92
    cluster_review_limit = 100
    if clustering_mode == "semantic_mimo":
        st.markdown("<div class='section-label'>大模型聚类裁决</div>", unsafe_allow_html=True)
        review_floor_column, auto_merge_column, review_limit_column = st.columns([1, 1, 1], gap="large")
        with review_floor_column:
            cluster_review_floor = st.slider(
                "进入大模型裁决的最低相似度",
                min_value=0.60,
                max_value=0.90,
                value=0.75,
                step=0.01,
                help="低于该值直接新建主题；达到该值后，才会进入候选主题裁决。",
            )
        with auto_merge_column:
            cluster_auto_merge_threshold = st.slider(
                "高置信自动合并阈值",
                min_value=0.80,
                max_value=0.98,
                value=0.92,
                step=0.01,
                help="只有标签聚类键一致且模型特征无明显冲突时才允许自动合并；其他边界候选仍交给大模型判断。",
            )
        with review_limit_column:
            cluster_review_limit = int(st.number_input(
                "单次最多大模型裁决次数",
                min_value=10,
                max_value=500,
                value=100,
                step=10,
                help="达到上限后，剩余边界记录保守地新建主题，避免超出调用成本。",
            ))
        if not use_mimo:
            st.warning("当前模式需要启用 MiMo；否则会自动回退为纯语义聚类。")
    config_left, config_right = st.columns([1.4, 0.6], gap="large")
    with config_left:
        if source_file:
            status_text = "会话数据已就绪，可以开始生成无标准引用的主题候选。"
            status_class = "success"
        else:
            status_text = "请上传方向二脱敏会话数据。"
            status_class = "warning"
        st.markdown(
            f"<div class='status-strip'><span class='status-dot {status_class}'></span>{status_text}</div>",
            unsafe_allow_html=True,
        )
    with config_right:
        generate_clicked = st.button(
            "生成主题候选",
            type="primary",
            disabled=not source_file,
            width="stretch",
        )

    if generate_clicked:
        with st.spinner("正在读取会话、分析图片、生成模型标签并聚合主题..."):
            try:
                workbook_bytes, summary, run_dir = _run_generation(
                    source_file,
                    None,
                    "" if product_label == "全部" else product_label,
                    use_mimo,
                    clustering_mode,
                    semantic_threshold,
                    cluster_review_floor,
                    cluster_auto_merge_threshold,
                    cluster_review_limit,
                )
            except Exception as exc:
                st.error(f"主题候选生成失败：{exc}")
                return
        st.session_state.generated_topic_workbook = workbook_bytes
        st.session_state.generated_topic_summary = summary
        st.session_state.generated_topic_run_dir = str(run_dir)
        st.success("主题候选已生成，可切换到“审核与反馈”继续审核。")

    summary = st.session_state.get("generated_topic_summary")
    workbook_bytes = st.session_state.get("generated_topic_workbook")
    if summary and workbook_bytes:
        st.markdown("<div class='section-label'>本次运行结果</div>", unsafe_allow_html=True)
        metrics = st.columns(5)
        metrics[0].metric("可处理记录", summary.get("eligible_rows", 0))
        metrics[1].metric("模型语义标注", summary.get("topic_signal_labeled_rows", 0))
        metrics[2].metric("主题候选", summary.get("topic_rows", 0))
        metrics[3].metric("证据缺口", summary.get("evidence_gap_rows", 0))
        metrics[4].metric("字段/品类排除", summary.get("excluded_rows", 0))
        effective_mode = summary.get("clustering_effective_mode", "")
        requested_mode = summary.get("clustering_requested_mode", "")
        clustering_model = summary.get("clustering_model", "")
        clustering_error = summary.get("clustering_error", "")
        if requested_mode == "direct_mimo" and effective_mode == "direct_mimo":
            st.markdown(
                (
                    "<div class='status-strip'><span class='status-dot success'></span>"
                    f"纯大模型 1～N 聚类已生效；模型：{clustering_model or '-'}；"
                    f"原子问题：{summary.get('atomic_unit_count', 0)}，"
                    f"聚类调用：{summary.get('direct_cluster_calls', 0)}，"
                    f"提取失败回退：{summary.get('atomic_extraction_failed', 0)}，"
                    f"聚类失败回退：{summary.get('direct_cluster_failed', 0)}。"
                    "</div>"
                ),
                unsafe_allow_html=True,
            )
        elif requested_mode == "direct_mimo":
            st.markdown(
                (
                    "<div class='status-strip'><span class='status-dot warning'></span>"
                    f"纯大模型聚类未生效，已回退为规则聚类：{clustering_error or 'MiMo 未配置'}。</div>"
                ),
                unsafe_allow_html=True,
            )
        elif requested_mode == "semantic_mimo" and effective_mode == "semantic_mimo":
            st.markdown(
                (
                    "<div class='status-strip'><span class='status-dot success'></span>"
                    f"模型会话标注 + 标签聚合 + 语义候选召回 + 大模型裁决已生效；Embedding：{clustering_model or '-'}；"
                    f"裁决：{summary.get('clustering_review_calls', 0)} 次，"
                    f"标签合并：{summary.get('clustering_tag_auto_merged', 0)}，"
                    f"合并：{summary.get('clustering_review_approved', 0)}，"
                    f"拒绝：{summary.get('clustering_review_rejected', 0)}，"
                    f"自动合并：{summary.get('clustering_auto_merged', 0)}。"
                    "</div>"
                ),
                unsafe_allow_html=True,
            )
        elif requested_mode == "semantic_mimo" and effective_mode == "semantic":
            st.markdown(
                (
                    "<div class='status-strip'><span class='status-dot warning'></span>"
                    f"大模型聚类裁决未生效，当前使用模型标签 + 语义候选召回：{clustering_error or 'MiMo 未配置'}。</div>"
                ),
                unsafe_allow_html=True,
            )
        elif requested_mode == "semantic" and effective_mode == "semantic":
            st.markdown(
                (
                    "<div class='status-strip'><span class='status-dot success'></span>"
                    f"模型标签 + 语义候选召回已生效；Embedding：{clustering_model or '-'}；"
                    f"阈值：{summary.get('clustering_threshold', '-')}。</div>"
                ),
                unsafe_allow_html=True,
            )
        elif requested_mode in {"semantic", "semantic_mimo"}:
            st.markdown(
                (
                    "<div class='status-strip'><span class='status-dot warning'></span>"
                    f"语义聚类未生效，已回退规则聚类：{clustering_error or 'Embedding 服务不可用'}。</div>"
                ),
                unsafe_allow_html=True,
            )
        st.caption(f"本次运行目录：{st.session_state.get('generated_topic_run_dir', '')}")
        st.download_button(
            "下载主题复核工作簿",
            data=workbook_bytes,
            file_name="topic_review_queue.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )


def _render_review() -> None:
    _page_heading(
        "自动标注验证与例外处理",
        "组员标注用于验证模型自动审核准确率；生产启用后，仅风险候选进入人工例外处理。",
    )
    cz_adapter = _shared_cz_adapter()
    api_state = cz_adapter.readiness()
    api_configured = bool(api_state["configured"])
    auto_policy = AutoReviewPolicy.from_env()
    review_mode = "生产自动审核模式" if auto_policy.enabled else "组员验证模式"
    st.markdown(
        f"""
        <div class="status-strip">
          <span class="status-dot {'success' if api_configured else 'warning'}"></span>
          <span>cz 对接状态：{api_state['status']} · 当前为{review_mode}。{'模型或验证通过后可提交待审核队列。' if api_configured else '请先在服务端配置接口地址和集成密钥。'}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("<div class='section-label'>审核数据源</div>", unsafe_allow_html=True)
    generated_bytes = st.session_state.get("generated_topic_workbook")
    uploaded_file = st.file_uploader("上传 topic_review_queue.xlsx", type=["xlsx"], key="topic_review_upload")
    use_generated = st.checkbox(
        "使用本次生成的主题候选",
        value=bool(generated_bytes) and uploaded_file is None,
        disabled=not generated_bytes,
    )
    workbook_bytes = generated_bytes if use_generated else (uploaded_file.getvalue() if uploaded_file else None)
    if workbook_bytes is None:
        st.info("先在“生成主题候选”中运行，或上传已有的 topic_review_queue.xlsx。")
        return

    file_token = f"{len(workbook_bytes)}:{hash(workbook_bytes)}"
    if st.session_state.get("topic_review_file_token") != file_token:
        try:
            rows, mapping_rows, evidence_gap_rows, pending_cluster_rows, model_draft_rows = _load_topic_workbook(workbook_bytes)
        except ValueError as exc:
            st.error(str(exc))
            return
        st.session_state.topic_review_file_token = file_token
        st.session_state.topic_review_rows = rows
        st.session_state.topic_mapping_rows = mapping_rows
        st.session_state.topic_evidence_gap_rows = evidence_gap_rows
        st.session_state.topic_pending_cluster_rows = pending_cluster_rows
        st.session_state.topic_model_draft_rows = model_draft_rows
        st.session_state.topic_review_changes = {}
        st.session_state.cz_submit_result = None
        draft_by_topic = {
            _text(draft.get("主题ID")): draft
            for draft in model_draft_rows
        }
        st.session_state.topic_initial_drafts = {
            row["_review_row_index"]: {
                field: _text(draft_by_topic.get(_text(row.get("主题ID")), {}).get(field) or row.get(field))
                for field in KNOWLEDGE_MASTER_COLUMNS
            }
            for row in rows
        }
        st.session_state.selected_topic_row = None

    rows: list[dict[str, Any]] = st.session_state.topic_review_rows
    mapping_rows: list[dict[str, Any]] = st.session_state.topic_mapping_rows
    evidence_gap_rows: list[dict[str, Any]] = st.session_state.topic_evidence_gap_rows
    pending_cluster_rows: list[dict[str, Any]] = st.session_state.topic_pending_cluster_rows
    model_draft_rows: list[dict[str, Any]] = st.session_state.get("topic_model_draft_rows", [])
    changes: dict[int, dict[str, str]] = st.session_state.topic_review_changes
    initial_drafts: dict[int, dict[str, str]] = st.session_state.get("topic_initial_drafts", {})
    potential_auto_rows, manual_exception_rows = partition_auto_review_candidates(
        rows,
        auto_policy,
    )
    validation_report = evaluate_auto_review_validation(rows, auto_policy)

    st.markdown("<div class='section-label'>筛选与自动审核概览</div>", unsafe_allow_html=True)
    filter_keyword, filter_decision, filter_focus = st.columns([1.65, 1, 0.72])
    keyword = filter_keyword.text_input("搜索主题、标准或记录 ID", placeholder="例如：屏幕、机型、TOP-")
    decision = filter_decision.selectbox(
        "组员验证状态",
        ["", "未审核", *REVIEW_DECISIONS],
        format_func=lambda value: value or "全部",
    )
    focus_only = filter_focus.checkbox("仅重点复核")
    filtered_rows = _filtered_rows(rows, keyword, decision, focus_only)

    pending_count = sum(not teammate_validation_decision(row) for row in rows)
    with st.container(horizontal=True):
        st.metric("主题候选", len(rows), border=True)
        st.metric("模型可自动放行", len(potential_auto_rows), border=True)
        st.metric("人工例外", len(manual_exception_rows), border=True)
        st.metric("组员已验证", validation_report["validated_rows"], border=True)
        st.metric(
            "验证准确率",
            (
                f"{validation_report['accuracy']:.1%}"
                if validation_report["accuracy"] is not None
                else "-"
            ),
            border=True,
        )
        st.metric("待验证", pending_count, border=True)

    with st.container(border=True):
        st.markdown("#### 模型自动审核验证")
        with st.container(horizontal=True):
            st.metric(
                "自动放行精确率",
                (
                    f"{validation_report['pass_precision']:.1%}"
                    if validation_report["pass_precision"] is not None
                    else "-"
                ),
                border=True,
            )
            st.metric(
                "自动放行召回率",
                (
                    f"{validation_report['pass_recall']:.1%}"
                    if validation_report["pass_recall"] is not None
                    else "-"
                ),
                border=True,
            )
            st.metric("错误放行", validation_report["false_pass"], border=True)
            st.metric("错误拦截", validation_report["false_reject"], border=True)
        if validation_report["gate_ready"]:
            st.success("当前验证结果已达到配置门槛，可将该模型和 Prompt 版本绑定后启用生产自动审核。")
        else:
            reasons = "；".join(validation_report["gate_reasons"])
            st.warning(f"尚未达到生产自动审核门槛：{reasons or '等待组员验证数据'}。")
        if validation_report["versions"]:
            st.dataframe(
                validation_report["versions"],
                hide_index=True,
                width="stretch",
            )
        if validation_report["by_product"]:
            st.dataframe(
                validation_report["by_product"],
                hide_index=True,
                width="stretch",
                column_config={
                    "accuracy": st.column_config.NumberColumn(
                        "准确率",
                        format="percent",
                    )
                },
            )
        if auto_policy.enabled and not auto_policy.deployment_ready:
            st.error("已开启生产自动审核，但尚未配置已验证的模型名称和 Prompt 版本，系统不会自动放行。")

    with st.expander(f"人工例外队列（{len(manual_exception_rows)}）", expanded=False):
        st.dataframe(
            [
                {
                    "主题ID": _text(item.get("主题ID")),
                    "产品类型": _text(item.get("产品类型")) or _text(item.get("适用范围")),
                    "主标题": _text(item.get("主标题")),
                    "模型结论": _text(item.get("模型初标结论")),
                    "模型沉淀价值": _text(item.get("模型初标是否值得沉淀")),
                    "置信度": item.get("模型初标置信度"),
                    "例外原因": _text(item.get("自动审核原因")),
                }
                for item in manual_exception_rows
            ],
            hide_index=True,
            width="stretch",
        )

    with st.expander(f"证据缺口记录（{len(evidence_gap_rows)}）", expanded=False):
        st.dataframe(
            [
                {
                    "数据ID": _text(row.get("数据ID")),
                    "工单ID": _text(row.get("工单ID")),
                    "核心问题": _text(row.get("核心问题")),
                    "证据缺口原因": _text(row.get("证据缺口原因")),
                }
                for row in evidence_gap_rows
            ],
            width="stretch",
            hide_index=True,
            height=220,
        )
    with st.expander(f"待聚合记录（{len(pending_cluster_rows)}）", expanded=False):
        st.dataframe(
            [
                {
                    "数据ID": _text(row.get("数据ID")),
                    "工单ID": _text(row.get("工单ID")),
                    "核心问题": _text(row.get("核心问题")),
                    "主题特征": " / ".join(
                        _text(row.get(field))
                        for field in ("问题意图", "对象/部位", "异常现象", "解题方式")
                        if _text(row.get(field))
                    ),
                    "待聚合原因": _text(row.get("待聚合原因")),
                }
                for row in pending_cluster_rows
            ],
            width="stretch",
            hide_index=True,
            height=220,
        )

    st.markdown("<div class='section-label'>主题审核工作区</div>", unsafe_allow_html=True)
    left, right = st.columns([0.86, 1.34], gap="large")
    with left:
        st.markdown(f"<div class='section-label'>主题队列 · {len(filtered_rows)} / {len(rows)}</div>", unsafe_allow_html=True)
        options = [row["_review_row_index"] for row in filtered_rows]
        if not options:
            st.warning("没有符合筛选条件的主题。")
            return
        selected = st.selectbox(
            "当前主题",
            options,
            index=options.index(st.session_state.selected_topic_row)
            if st.session_state.selected_topic_row in options
            else 0,
            format_func=lambda row_index: next(
                f"{_text(row.get('主题ID'))} · {_text(row.get('主标题'))}"
                for row in filtered_rows
                if row["_review_row_index"] == row_index
            ),
            label_visibility="collapsed",
        )
        selected_row = _selected_row(rows, selected)
        if selected_row is not None and selected != st.session_state.selected_topic_row:
            _reset_editor(selected_row)
        st.dataframe(
            [
                {
                    "主题": _text(row.get("主标题")),
                    "样本": _text(row.get("主题样本数")),
                    "证据": _text(row.get("主题证据等级")),
                    "模型初标": _text(row.get("模型初标结论")) or "待初标",
                    "沉淀价值": _text(row.get("模型初标是否值得沉淀")) or "待判断",
                    "组员验证": teammate_validation_decision(row) or "待验证",
                    "自动审核": _text(row.get("自动审核状态")) or "待判断",
                }
                for row in filtered_rows
            ],
            width="stretch",
            hide_index=True,
            height=580,
        )

    with right:
        row = _selected_row(rows, st.session_state.selected_topic_row)
        if row is None:
            st.info("请选择一个主题。")
            return
        st.markdown(f"### {_text(row.get('主标题'))}")
        context_left, context_right, context_status = st.columns([1.1, 1.1, 0.85])
        context_left.caption(f"主题 ID · {_text(row.get('主题ID'))}")
        context_right.caption("标准引用 · 不新增；已有值保留并搁置")
        context_status.caption(
            f"模型初标：{_text(row.get('模型初标结论')) or '待初标'} · "
            f"组员验证：{teammate_validation_decision(row) or '待验证'}"
        )

        workbench_tab, evidence_tab = st.tabs(["审核工作区", "案例证据"])
        with workbench_tab:
            preview_column, edit_column = st.columns([0.9, 1.1], gap="large")
            with preview_column:
                model_draft = next(
                    (
                        item
                        for item in model_draft_rows
                        if _text(item.get("主题ID")) == _text(row.get("主题ID"))
                    ),
                    {},
                )
                initial_draft = initial_drafts.get(
                    row["_review_row_index"],
                    {
                        field: _text(model_draft.get(field) or row.get(field))
                        for field in KNOWLEDGE_MASTER_COLUMNS
                    },
                )
                provider = _text(model_draft.get("转写提供方") or row.get("主题模型提供方"))
                transcription_status = _text(row.get("模型阶段状态"))
                if provider == "mimo" and transcription_status == "topic_model_labeled":
                    st.success("已完成 MiMo 主题转写")
                elif provider == "mimo":
                    st.error("MiMo 主题转写失败，已使用规则草稿")
                else:
                    st.warning("当前为规则转写草稿，未完成 MiMo 主题转写")
                trace_left, trace_right = st.columns(2)
                trace_left.caption(f"转写模型 · {_text(model_draft.get('转写模型名称') or row.get('主题模型名称')) or '-'}")
                trace_right.caption(f"转写置信度 · {_text(model_draft.get('转写置信度') or row.get('主题置信度')) or '-'}")
                st.caption(
                    f"Prompt · {_text(model_draft.get('转写Prompt版本') or row.get('主题Prompt版本')) or '-'}  |  "
                    f"运行 ID · {_text(model_draft.get('转写模型运行ID') or row.get('主题模型运行ID')) or '-'}"
                )
                _read_only_text("主题转写主标题", initial_draft.get("主标题"), height=66)
                _read_only_text("主题转写副标题", initial_draft.get("副标题"), height=72)
                _read_only_text("主题转写知识内容", initial_draft.get("知识内容"), height=220)
                _read_only_text(
                    "推荐回复",
                    model_draft.get("推荐回复") or row.get("推荐回复"),
                    height=120,
                )
                _read_only_text(
                    "图例",
                    row.get("图例") or row.get("主题图片链接") or "当前知识不需要案例图",
                    height=76,
                )

                st.markdown("<div class='section-label'>模型初标审核</div>", unsafe_allow_html=True)
                initial_provider = _text(row.get("模型初标提供方"))
                initial_status = _text(row.get("模型初标状态"))
                if initial_provider == "mimo" and initial_status == "topic_initial_reviewed_model":
                    st.success("已完成 MiMo 模型初标审核")
                elif initial_provider == "mimo":
                    st.error("MiMo 模型初标失败，已回退为规则初标")
                else:
                    st.warning("当前为规则模型初标，待用 MiMo 验证")
                review_left, review_middle, review_right = st.columns(3)
                review_left.metric("初标结论", _text(row.get("模型初标结论")) or "-")
                review_middle.metric(
                    "是否值得沉淀",
                    _text(row.get("模型初标是否值得沉淀")) or "-",
                )
                review_right.metric("初标置信度", _text(row.get("模型初标置信度")) or "-")
                st.caption(
                    f"{_text(row.get('模型初标模型名称')) or '-'} · "
                    f"{_text(row.get('模型初标Prompt版本')) or '-'} · "
                    f"{_text(row.get('模型初标运行ID')) or '-'}"
                )
                st.caption(
                    f"证据充分性：{_text(row.get('模型初标证据充分性')) or '-'}；"
                    f"内容一致性：{_text(row.get('模型初标内容一致性')) or '-'}；"
                    f"重点复核：{_text(row.get('模型初标重点复核')) or '-'}"
                )
                _read_only_text("模型初标原因", row.get("模型初标原因"), height=96)

            with edit_column:
                editor_tab, feedback_tab = st.tabs(["候选内容", "组员验证"])
                with editor_tab:
                    st.caption("右侧内容是无标准引用的10项候选知识草稿；保存后会直接写入复核工作簿。")
                    with st.form("topic_content_edit_form"):
                        st.markdown("<div class='section-label'>基础信息</div>", unsafe_allow_html=True)
                        st.text_input("知识ID", key="topic_知识ID", disabled=True)
                        basic_left, basic_right = st.columns(2)
                        basic_left.text_input("主标题", key="topic_主标题")
                        basic_right.text_input("副标题", key="topic_副标题")
                        basic_left.text_input("知识分类", key="topic_知识分类")
                        basic_right.text_input("适用范围", key="topic_适用范围")

                        st.markdown("<div class='section-label'>知识正文</div>", unsafe_allow_html=True)
                        st.text_area("知识内容", height=260, key="topic_知识内容")
                        st.text_area(
                            "推荐回复",
                            height=120,
                            key="topic_推荐回复",
                            help="给人工答疑直接使用，建议控制在 80～180 字。",
                        )
                        st.text_area(
                            "图例",
                            height=88,
                            key="topic_图例",
                            help="填写已脱敏案例图链接或媒体ID；不需要图片时可留空。",
                        )

                        st.markdown("<div class='section-label'>检索信息</div>", unsafe_allow_html=True)
                        st.text_area("关键词", height=88, key="topic_关键词")
                        st.caption("当前模式不新增关联标准项；已有值不会被清空。")
                        edit_submitted = st.form_submit_button("保存候选草稿", type="primary")
                    if edit_submitted:
                        updates = {
                            field: _text(st.session_state.get(f"topic_{field}"))
                            for field in KNOWLEDGE_MASTER_COLUMNS
                        }
                        updates.update(
                            {
                                "知识ID": _text(st.session_state.get("topic_知识ID"))
                                or _text(row.get("主题ID")),
                                "图例": _text(st.session_state.get("topic_图例")),
                                "关键词": _text(st.session_state.get("topic_关键词")),
                                "推荐回复": _text(st.session_state.get("topic_推荐回复")),
                                "知识来源": "方向二案例沉淀",
                                "生效状态": "待审核",
                                "变更类型": "新增",
                                "失效原因": "",
                                "检索关键词": _text(st.session_state.get("topic_关键词")),
                            }
                        )
                        row.update(updates)
                        changes[row["_review_row_index"]] = {
                            **changes.get(row["_review_row_index"], {}),
                            **updates,
                        }
                        st.session_state.topic_review_changes = changes
                        st.success(f"已保存主题 {_text(row.get('主题ID'))} 的10项候选草稿。")

                with feedback_tab:
                    st.caption("组员填写：是否值得沉淀、是否可用、如何修改、问题反馈。不值得沉淀的纯个案知识不会进入批量送审。")
                    with st.form("topic_review_form"):
                        value_left, value_right = st.columns(2)
                        with value_left:
                            st.segmented_control(
                                "是否值得沉淀",
                                ["未标注", "是", "否"],
                                key="topic_是否值得沉淀",
                            )
                        value_right.selectbox(
                            "是否可用",
                            ["", "是", "否"],
                            format_func=lambda value: value or "未标注",
                            key="topic_是否可用",
                        )
                        st.text_area("如何修改", height=72, key="topic_如何修改")
                        st.text_area("问题反馈", height=72, key="topic_问题反馈")
                        st.markdown("<div class='section-label'>人工例外处理（仅风险候选，可选）</div>", unsafe_allow_html=True)
                        review_top_left, review_top_right, review_top_train = st.columns([1, 1, 0.85])
                        decision_value = st.session_state.get("topic_审核结论", "")
                        review_top_left.selectbox(
                            "审核结论",
                            ["", *REVIEW_DECISIONS],
                            index=(["", *REVIEW_DECISIONS].index(decision_value) if decision_value in ["", *REVIEW_DECISIONS] else 0),
                            format_func=lambda value: value or "未审核",
                            key="topic_审核结论",
                        )
                        review_top_right.selectbox("错误类型", ["", *ERROR_TYPES], key="topic_错误类型")
                        review_top_train.selectbox("进入训练集", ["", "是", "否"], key="topic_是否进入训练集")
                        st.text_area("审核备注", height=68, key="topic_审核备注")
                        st.text_area("错误原因", height=68, key="topic_错误原因")
                        reviewer_left, reviewer_right = st.columns(2)
                        reviewer_left.text_input("审核人", key="topic_审核人")
                        reviewer_right.text_input(
                            "审核时间",
                            placeholder="YYYY-MM-DD HH:mm:ss",
                            key="topic_审核时间",
                        )
                        submitted = st.form_submit_button("保存验证结果", type="primary")

                    if submitted:
                        updates = {field: _text(st.session_state.get(f"topic_{field}")) for field in TOPIC_REVIEW_COLUMNS}
                        updates.update(
                            {
                                field: _text(st.session_state.get(f"topic_{field}"))
                                for field in ("是否值得沉淀", "是否可用", "如何修改", "问题反馈")
                            }
                        )
                        if updates["是否值得沉淀"] == "未标注":
                            updates["是否值得沉淀"] = ""
                        if not updates["审核时间"] and updates["审核结论"]:
                            updates["审核时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        row.update(updates)
                        changes[row["_review_row_index"]] = {
                            **changes.get(row["_review_row_index"], {}),
                            **updates,
                        }
                        st.session_state.topic_review_changes = changes
                        st.success(f"已暂存主题 {_text(row.get('主题ID'))} 的验证结果。")

        with evidence_tab:
            evidence_meta_left, evidence_meta_right, evidence_meta_score = st.columns(3)
            evidence_meta_left.metric("聚合样本", _text(row.get("主题样本数")) or "0")
            evidence_meta_right.metric("模型置信度", _text(row.get("主题置信度")) or "-")
            evidence_meta_score.metric("证据等级", _text(row.get("主题证据等级")) or "-")
            _read_only_text("主题聚类键", row.get("主题聚类键"), height=88)
            _read_only_text("主题证据摘要", row.get("主题证据摘要"), height=160)
            _read_only_text(
                "候选图例",
                row.get("图例") or row.get("主题图片链接") or "当前知识不需要案例图",
                height=110,
            )
            topic_mapping = _mapping_for_topic(mapping_rows, _text(row.get("主题ID")))
            st.markdown("<div class='section-label'>模型语义标注</div>", unsafe_allow_html=True)
            st.dataframe(
                [
                    {
                        "记录 ID": _text(item.get("来源记录ID")),
                        "模型分类": " / ".join(
                            part
                            for part in (
                                _text(item.get("模型主题一级分类")),
                                _text(item.get("模型主题二级分类")),
                            )
                            if part
                        ) or "待确认",
                        "主题标签": _text(item.get("主题标签")),
                        "模型依据": _text(item.get("语义标注依据")),
                        "图片": _text(item.get("语义标注图片必要性")),
                        "状态": _text(item.get("语义标注状态")),
                    }
                    for item in topic_mapping
                ],
                width="stretch",
                hide_index=True,
            )
            st.markdown("<div class='section-label'>原始聊天内容</div>", unsafe_allow_html=True)
            chat_rows = [
                item
                for item in topic_mapping
                if _text(item.get("聊天内容")).strip()
            ]
            if chat_rows:
                for chat_index, item in enumerate(chat_rows, start=1):
                    source_id = _text(item.get("来源记录ID")) or _text(item.get("工单ID")) or f"记录 {chat_index}"
                    question = _text(item.get("核心问题"))
                    with st.expander(f"{chat_index:02d} · {source_id}", expanded=chat_index == 1):
                        if question:
                            st.caption(f"上游问题描述（仅参考）：{question}")
                        st.caption(
                            "主题标签："
                            f"{_text(item.get('主题标签')) or '待确认'} · "
                            f"图片判断：{_text(item.get('语义标注图片必要性')) or '待确认'}"
                        )
                        _read_only_text(
                            "会话内容",
                            item.get("聊天内容"),
                            height=180,
                            key=f"topic_chat_{_text(row.get('主题ID'))}_{chat_index}",
                        )
                        historical_reply = _text(item.get("历史实际回复")).strip()
                        if historical_reply:
                            _read_only_text(
                                "历史实际回复",
                                historical_reply,
                                height=110,
                                key=f"topic_reply_{_text(row.get('主题ID'))}_{chat_index}",
                            )
                        image_urls = split_image_urls(_text(item.get("图片链接")))
                        if image_urls:
                            try:
                                st.image(
                                    image_urls[:4],
                                    caption=[f"现场图片 {index}" for index, _url in enumerate(image_urls[:4], start=1)],
                                    width=180,
                                )
                            except Exception as exc:
                                st.caption(f"图片加载失败：{exc}")
            else:
                st.info("当前工作簿没有携带原始聊天内容；可以查看上方证据摘要或重新生成主题工作簿。")
            st.markdown("<div class='section-label'>来源记录</div>", unsafe_allow_html=True)
            st.dataframe(
                [
                    {
                        "记录 ID": _text(item.get("来源记录ID")),
                        "工单 ID": _text(item.get("工单ID")),
                        "核心问题": _text(item.get("核心问题")),
                        "证据": _text(item.get("证据等级")),
                    }
                    for item in topic_mapping
                ],
                width="stretch",
                hide_index=True,
                height=210,
            )

    reviewed_workbook = (
        _update_topic_workbook(workbook_bytes, changes)
        if changes
        else workbook_bytes
    )
    _reviewed_rows, feedback_rows, training_rows = finalize_topic_review_rows(rows)
    submittable_rows = select_submittable_candidates(rows, auto_policy)
    final_rows, _submission_feedback, _submission_training = finalize_topic_review_rows(
        submittable_rows
    )
    st.divider()
    st.markdown("<div class='section-label'>导出与提交</div>", unsafe_allow_html=True)
    downloads = st.columns(3)
    downloads[0].download_button(
        "下载审核工作簿",
        data=reviewed_workbook,
        file_name="topic_review_queue_reviewed.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
    downloads[1].download_button(
        "下载10项候选",
        data=_rows_to_xlsx_bytes(
            "候选知识",
            CASE_KNOWLEDGE_COLUMNS,
            build_case_knowledge_rows(final_rows),
        ),
        file_name="candidate_knowledge_for_submission.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
    downloads[2].download_button(
        "下载训练反馈",
        data=_jsonl_bytes(training_rows),
        file_name="topic_training_samples.jsonl",
        mime="application/jsonl",
        width="stretch",
    )
    st.caption(
        f"组员验证反馈 {len(feedback_rows)} 条；可进入训练集 {len(training_rows)} 条；"
        f"已确认值得沉淀且可提交 cz 待审核队列 {len(submittable_rows)} 条。"
    )

    with st.container(border=True):
        st.markdown("#### 提交 cz 待审核队列")
        if auto_policy.enabled:
            st.caption("无标准引用模式优先提交已完成案例证据核验且通过人工验证的候选；风险候选继续留在人工例外队列。")
        else:
            st.caption("当前为验证模式，仅提交组员确认“值得沉淀”且验证通过的试运行候选；准确率达标并启用生产策略后，将改为模型自动放行。")
        submit_clicked = st.button(
            "提交模型自动通过候选" if auto_policy.enabled else "提交验证通过候选",
            type="primary",
            icon=":material/send:",
            width="stretch",
            key="submit_reviewed_candidates_to_cz",
            disabled=not api_configured or not submittable_rows,
        )
        if submit_clicked:
            with st.spinner("正在读取 cz 分类字典并提交候选..."):
                try:
                    st.session_state.cz_submit_result = cz_adapter.submit_candidates(
                        submittable_rows
                    )
                except Exception as exc:
                    st.session_state.cz_submit_result = None
                    st.error(f"提交 cz 失败：{exc}")
                else:
                    st.success("候选提交完成，结果如下。")

        submit_result = st.session_state.get("cz_submit_result")
        if submit_result:
            with st.container(horizontal=True):
                st.metric("进入待审核", int(submit_result.get("accepted") or 0), border=True)
                st.metric("Qwen3疑似重复拦截", int(submit_result.get("intercepted") or 0), border=True)
                st.metric("明确重复阻断", int(submit_result.get("blocked") or 0), border=True)
                st.metric("幂等复用", int(submit_result.get("reused") or 0), border=True)
                st.metric("已拒绝", int(submit_result.get("rejected") or 0), border=True)
            result_rows = submit_result.get("results") or []
            if result_rows:
                st.dataframe(
                    [
                        {
                            "事件 ID": _text(item.get("event_id")),
                            "状态": _text(item.get("status")),
                            "知识 ID": _text(item.get("knowledge_id")),
                            "入库 ID": _text(item.get("ingestion_id")),
                            "Qwen3拦截": _text(
                                (item.get("deduplication") or {}).get("action")
                            ),
                            "错误码": _text(item.get("error_code")),
                            "错误信息": _text(item.get("error_message")),
                        }
                        for item in result_rows
                    ],
                    hide_index=True,
                    width="stretch",
                )


st.markdown(
    """
    <div class="workspace-header">
      <div class="workspace-header-copy">
        <div class="workspace-kicker">Answer Hub · Knowledge Automation</div>
        <div class="workspace-title">答疑知识自动化工作台</div>
        <div class="workspace-subtitle">会话清洗、语义标注、主题沉淀、证据核验和人工审核在同一工作区完成；正式发布仍由 cz 知识库网站处理。</div>
      </div>
      <div class="workspace-header-meta">
        <span class="meta-pill">自动化编排</span>
        <span class="meta-pill">主题级沉淀</span>
        <span class="meta-pill">人工复核</span>
        <span class="meta-pill">可追溯导出</span>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)
page = st.segmented_control(
    "工作区",
    ["自动化看板", "转人工分析", "聚类验证", "生成主题候选", "审核与反馈"],
    default="自动化看板",
    key="workspace_page",
    label_visibility="collapsed",
    persist_state="session",
)
if page == "自动化看板":
    _render_automation()
elif page == "转人工分析":
    render_transfer_analysis()
elif page == "聚类验证":
    _render_cluster_validation()
elif page == "生成主题候选":
    _render_generation()
else:
    _render_review()
