from __future__ import annotations

import argparse
from collections import Counter, defaultdict
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from answer_hub.business_taxonomy import (
    AGGREGATE_BUSINESS_LINE_NAME,
    SELF_OPERATED_BUSINESS_LINE_NAME,
)


INCLUDED_VALUE = "是"
PLACEHOLDER_REMARK = "58"
PREDICTION_KEY_SEPARATOR = "\x1f"


def _text(value: Any) -> str:
    return str(value or "").strip()


def _parse_int(value: Any) -> int | None:
    text = _text(value)
    if not text:
        return None
    match = re.search(r"\d+", text)
    return int(match.group()) if match else None


def _read_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "原子问题明细" not in workbook.sheetnames:
        raise ValueError("金标准工作簿缺少“原子问题明细”工作表")
    worksheet = workbook["原子问题明细"]
    rows = worksheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(rows, ())]
    return [
        {
            header: value
            for header, value in zip(headers, row)
            if header
        }
        for row in rows
    ]


def _normalize_id(value: Any) -> str:
    text = _text(value)
    if text.startswith("ID="):
        return text[3:].strip()
    return text


def _split_cell_values(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[\r\n,，]+", text) if part.strip()]


def _gold_key(row: dict[str, Any]) -> str:
    return (
        _normalize_id(row.get("真实会话ID（文本）"))
        or _text(row.get("来源数据原始行号"))
        or _text(row.get("来源原始行号"))
    )


def _gold_product(row: dict[str, Any]) -> str:
    return _text(row.get("品类") or row.get("产品类型"))


def _gold_business_line(row: dict[str, Any]) -> str:
    explicit = _text(
        row.get("回收业务层级")
        or row.get("回收业务")
        or row.get("业务线")
    )
    if explicit:
        return explicit
    product = _gold_product(row)
    if product == AGGREGATE_BUSINESS_LINE_NAME:
        return AGGREGATE_BUSINESS_LINE_NAME
    return SELF_OPERATED_BUSINESS_LINE_NAME


def _prediction_key(source_id: Any, product_category: Any = "") -> str:
    normalized_id = _normalize_id(source_id)
    product = _text(product_category)
    if not normalized_id or not product:
        return normalized_id
    return f"{normalized_id}{PREDICTION_KEY_SEPARATOR}{product}"


def _append_prediction(
    result: dict[str, list[str]],
    source_id: Any,
    topic: Any,
    product_category: Any = "",
) -> None:
    normalized_id = _normalize_id(source_id)
    normalized_topic = _text(topic)
    if not normalized_id or not normalized_topic:
        return
    keys = [normalized_id]
    product_key = _prediction_key(normalized_id, product_category)
    if product_key != normalized_id:
        keys.append(product_key)
    for key in keys:
        if normalized_topic not in result[key]:
            result[key].append(normalized_topic)


def _prediction_values(
    predictions: dict[str, list[str]],
    source_id: Any,
    product_category: Any = "",
) -> list[str]:
    normalized_id = _normalize_id(source_id)
    if not normalized_id:
        return []
    product_key = _prediction_key(normalized_id, product_category)
    if product_key != normalized_id and product_key in predictions:
        return predictions[product_key]
    return predictions.get(normalized_id, [])


def _read_model_workbook(workbook_path: Path) -> dict[str, list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "原子问题明细" in workbook.sheetnames:
        worksheet = workbook["原子问题明细"]
        rows = worksheet.iter_rows(values_only=True)
        headers = [_text(value) for value in next(rows, ())]
        header_index = {header: index for index, header in enumerate(headers)}
        topic_index = header_index.get("主题编号")
        source_index = header_index.get("原始行号")
        product_index = header_index.get("产品类型")
        if product_index is None:
            product_index = header_index.get("品类")
        if topic_index is None or source_index is None:
            raise ValueError("模型原子问题明细缺少“主题编号”或“原始行号”")
        result: dict[str, list[str]] = defaultdict(list)
        for row in rows:
            if len(row) <= max(topic_index, source_index):
                continue
            source_id = _text(row[source_index])
            topic = _text(row[topic_index])
            product = (
                _text(row[product_index])
                if product_index is not None and len(row) > product_index
                else ""
            )
            _append_prediction(result, source_id, topic, product)
        return dict(result)

    if "聚类结果" not in workbook.sheetnames:
        raise ValueError("模型工作簿缺少“原子问题明细”或“聚类结果”工作表")
    worksheet = workbook["聚类结果"]
    rows = worksheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(rows, ())]
    header_index = {header: index for index, header in enumerate(headers)}
    topic_index = header_index.get("聚类主题ID")
    id_index = header_index.get("主题来源记录ID") or header_index.get("主题工单ID")
    product_index = header_index.get("产品类型")
    if topic_index is None or id_index is None:
        raise ValueError("模型聚类结果缺少主题 ID 或来源记录 ID")
    result: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        if len(row) <= max(topic_index, id_index):
            continue
        topic = _text(row[topic_index])
        product = (
            _text(row[product_index])
            if product_index is not None and len(row) > product_index
            else ""
        )
        for source_id in _split_cell_values(row[id_index]):
            _append_prediction(result, source_id, topic, product)
    return dict(result)


def _read_model_json(json_path: Path) -> dict[str, list[str]]:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    atomic_units = payload.get("atomic_units") if isinstance(payload, dict) else None
    clusters = payload.get("clusters") if isinstance(payload, dict) else None
    if not isinstance(atomic_units, list) or not isinstance(clusters, list):
        raise ValueError("模型 JSON 必须包含 atomic_units 和 clusters 数组")
    cluster_by_atomic_id: dict[str, str] = {}
    for cluster in clusters:
        if not isinstance(cluster, dict):
            continue
        cluster_id = _text(cluster.get("cluster_id"))
        for atomic_id in cluster.get("member_atomic_ids") or []:
            if cluster_id and _text(atomic_id):
                cluster_by_atomic_id[_text(atomic_id)] = cluster_id

    records: list[tuple[str, str, str, str]] = []
    for unit in atomic_units:
        if not isinstance(unit, dict):
            continue
        atomic_id = _text(unit.get("unit_id") or unit.get("atomic_id"))
        cluster_id = _text(unit.get("cluster_id")) or cluster_by_atomic_id.get(atomic_id, "")
        source_id = _normalize_id(
            unit.get("source_record_key") or unit.get("work_order_id") or unit.get("source_id")
        )
        product = _text(
            unit.get("product_category")
            or unit.get("产品类型")
            or unit.get("category")
        )
        if source_id and cluster_id:
            records.append((source_id, product, atomic_id, cluster_id))
    result: dict[str, list[str]] = defaultdict(list)
    for source_id, product, _atomic_id, cluster_id in sorted(
        records,
        key=lambda item: (item[0], item[1], item[2]),
    ):
        _append_prediction(result, source_id, cluster_id, product)
    return dict(result)


def _model_predictions(
    gold_rows: list[dict[str, Any]],
    *,
    model_workbook: Path | None,
    model_json: Path | None,
) -> dict[str, list[str]]:
    if bool(model_workbook) == bool(model_json):
        raise ValueError("必须且只能提供 --model-workbook 或 --model-json")
    predictions = (
        _read_model_workbook(model_workbook)
        if model_workbook is not None
        else _read_model_json(model_json)  # type: ignore[arg-type]
    )
    by_source_row: dict[str, list[str]] = {}
    for row in gold_rows:
        session_id = _gold_key(row)
        product = _gold_product(row)
        output_key = _prediction_key(session_id, product)
        predicted = _prediction_values(predictions, session_id, product)
        if predicted:
            by_source_row[output_key] = predicted
        else:
            source_row = _text(row.get("来源原始行号"))
            predicted = _prediction_values(predictions, source_row, product)
            if predicted:
                by_source_row[output_key] = predicted
    return by_source_row


def _prediction_for_row(
    row: dict[str, Any],
    predictions: dict[str, list[str]],
) -> str:
    product = _gold_product(row)
    values = _prediction_values(predictions, _gold_key(row), product)
    if not values:
        values = _prediction_values(
            predictions,
            _text(row.get("来源原始行号")),
            product,
        )
    if not values:
        return ""
    sequence = _parse_int(row.get("拆分序号")) or 1
    return values[sequence - 1] if sequence <= len(values) else values[0]


def _is_included(row: dict[str, Any]) -> bool:
    return _text(row.get("纳入准确率")) == INCLUDED_VALUE


def _pairwise_metrics(
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    total_pairs = 0
    predicted_same = 0
    gold_same = 0
    true_positive = 0
    for index, left in enumerate(rows):
        left_gold = _text(left.get("人工主题编号"))
        left_predicted = _text(left.get("模型主题编号"))
        for right in rows[index + 1 :]:
            right_gold = _text(right.get("人工主题编号"))
            right_predicted = _text(right.get("模型主题编号"))
            if not left_gold or not right_gold or not left_predicted or not right_predicted:
                continue
            total_pairs += 1
            same_gold = left_gold == right_gold
            same_predicted = left_predicted == right_predicted
            gold_same += int(same_gold)
            predicted_same += int(same_predicted)
            true_positive += int(same_gold and same_predicted)
    precision = true_positive / predicted_same if predicted_same else 0.0
    recall = true_positive / gold_same if gold_same else 0.0
    f1 = (
        2 * precision * recall / (precision + recall)
        if precision + recall
        else 0.0
    )
    return {
        "total_pairs": total_pairs,
        "predicted_same_pairs": predicted_same,
        "gold_same_pairs": gold_same,
        "true_positive_pairs": true_positive,
        "false_merge_pairs": predicted_same - true_positive,
        "false_split_pairs": gold_same - true_positive,
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "f1": round(f1, 4),
    }


def _purity_and_coverage(
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    predicted_to_gold: dict[str, Counter[str]] = defaultdict(Counter)
    gold_to_predicted: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows:
        gold = _text(row.get("人工主题编号"))
        predicted = _text(row.get("模型主题编号"))
        if not gold or not predicted:
            continue
        predicted_to_gold[predicted][gold] += 1
        gold_to_predicted[gold][predicted] += 1

    total = len(rows)
    purity_numerator = sum(
        max(counts.values())
        for counts in predicted_to_gold.values()
        if counts
    )
    coverage_numerator = sum(
        max(counts.values())
        for counts in gold_to_predicted.values()
        if counts
    )
    purity = purity_numerator / total if total else 0.0
    coverage = coverage_numerator / total if total else 0.0
    f1 = (
        2 * purity * coverage / (purity + coverage)
        if purity + coverage
        else 0.0
    )
    return {
        "included_atoms": total,
        "predicted_cluster_count": len(predicted_to_gold),
        "gold_topic_count": len(gold_to_predicted),
        "purity": round(purity, 4),
        "coverage": round(coverage, 4),
        "f1": round(f1, 4),
    }


def _multi_topic_split_stats(
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    by_session: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        session_id = _text(row.get("真实会话ID（文本）"))
        if session_id:
            by_session[session_id].append(row)

    split_sessions = [
        session_rows
        for session_rows in by_session.values()
        if len({_text(row.get("人工主题编号")) for row in session_rows}) > 1
    ]
    affected_atoms = sum(len(session_rows) for session_rows in split_sessions)
    model_kept_together = sum(
        int(len({_text(row.get("模型主题编号")) for row in session_rows}) == 1)
        for session_rows in split_sessions
    )
    return {
        "included_sessions": len(by_session),
        "multi_topic_sessions": len(split_sessions),
        "multi_topic_atoms": affected_atoms,
        "model_kept_multi_topic_session_together": model_kept_together,
        "model_split_multi_topic_session": (
            len(split_sessions) - model_kept_together
        ),
    }


def _category_metrics(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[_text(row.get("品类")) or "未知品类"].append(row)
    result = []
    for category, category_rows in sorted(grouped.items()):
        purity = _purity_and_coverage(category_rows)
        pairwise = _pairwise_metrics(category_rows)
        result.append(
            {
                "category": category,
                "included_atoms": len(category_rows),
                "gold_topics": purity["gold_topic_count"],
                "predicted_clusters": purity["predicted_cluster_count"],
                "purity": purity["purity"],
                "coverage": purity["coverage"],
                "f1": purity["f1"],
                "pairwise_precision": pairwise["precision"],
                "pairwise_recall": pairwise["recall"],
                "pairwise_f1": pairwise["f1"],
            }
        )
    return result


def _product_boundary_metrics(
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    topic_products: dict[str, set[str]] = defaultdict(set)
    topic_atom_counts: Counter[str] = Counter()
    for row in rows:
        predicted = _text(row.get("模型主题编号"))
        product = _gold_product(row)
        if not predicted or not product:
            continue
        topic_products[predicted].add(product)
        topic_atom_counts[predicted] += 1

    cross_product_clusters = [
        {
            "model_topic": topic,
            "product_categories": sorted(products),
            "included_atoms": topic_atom_counts[topic],
        }
        for topic, products in sorted(topic_products.items())
        if len(products) > 1
    ]
    return {
        "cross_product_cluster_count": len(cross_product_clusters),
        "cross_product_clusters": cross_product_clusters,
    }


def _business_line_boundary_metrics(
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    topic_business_lines: dict[str, set[str]] = defaultdict(set)
    topic_atom_counts: Counter[str] = Counter()
    for row in rows:
        predicted = _text(row.get("模型主题编号"))
        business_line = _gold_business_line(row)
        if not predicted or not business_line:
            continue
        topic_business_lines[predicted].add(business_line)
        topic_atom_counts[predicted] += 1

    cross_business_line_clusters = [
        {
            "model_topic": topic,
            "business_lines": sorted(business_lines),
            "included_atoms": topic_atom_counts[topic],
        }
        for topic, business_lines in sorted(topic_business_lines.items())
        if len(business_lines) > 1
    ]
    return {
        "cross_business_line_cluster_count": len(
            cross_business_line_clusters
        ),
        "cross_business_line_clusters": cross_business_line_clusters,
    }


def _error_examples(
    rows: list[dict[str, Any]],
    *,
    limit: int,
) -> dict[str, list[dict[str, Any]]]:
    predicted_to_gold: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows:
        predicted = _text(row.get("模型主题编号"))
        gold = _text(row.get("人工主题编号"))
        if predicted and gold:
            predicted_to_gold[predicted][gold] += 1

    merge_candidates = []
    for predicted, counts in predicted_to_gold.items():
        if len(counts) <= 1:
            continue
        merge_candidates.append(
            {
                "model_topic": predicted,
                "included_atoms": sum(counts.values()),
                "gold_topic_count": len(counts),
                "gold_topic_distribution": dict(counts.most_common()),
            }
        )
    merge_candidates.sort(
        key=lambda item: (
            -item["gold_topic_count"],
            -item["included_atoms"],
            item["model_topic"],
        )
    )

    gold_to_predicted: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows:
        predicted = _text(row.get("模型主题编号"))
        gold = _text(row.get("人工主题编号"))
        if predicted and gold:
            gold_to_predicted[gold][predicted] += 1
    split_candidates = []
    for gold, counts in gold_to_predicted.items():
        if len(counts) <= 1:
            continue
        split_candidates.append(
            {
                "gold_topic": gold,
                "included_atoms": sum(counts.values()),
                "model_topic_count": len(counts),
                "model_topic_distribution": dict(counts.most_common()),
            }
        )
    split_candidates.sort(
        key=lambda item: (
            -item["model_topic_count"],
            -item["included_atoms"],
            item["gold_topic"],
        )
    )

    return {
        "error_merges": merge_candidates[:limit],
        "error_splits": split_candidates[:limit],
    }


def evaluate(
    workbook_path: Path,
    *,
    error_limit: int = 30,
    model_workbook: Path | None = None,
    model_json: Path | None = None,
) -> dict[str, Any]:
    all_rows = _read_rows(workbook_path)
    predictions: dict[str, list[str]] | None = None
    if model_workbook is not None or model_json is not None:
        predictions = _model_predictions(
            all_rows,
            model_workbook=model_workbook,
            model_json=model_json,
        )
    included_rows = [
        {
            **row,
            "模型主题编号": (
                _prediction_for_row(row, predictions)
                if predictions is not None
                else _text(row.get("模型主题编号"))
            ),
        }
        for row in all_rows
        if _is_included(row)
        and _text(row.get("人工主题编号"))
        and (
            (
                _prediction_for_row(row, predictions)
                if predictions is not None
                else _text(row.get("模型主题编号"))
            )
        )
    ]
    matched_rows = len(included_rows)
    excluded_rows = [row for row in all_rows if not _is_included(row)]
    summary = {
        "workbook": str(workbook_path),
        "model_result": (
            str(model_workbook)
            if model_workbook is not None
            else str(model_json)
            if model_json is not None
            else "人工校正表中的历史模型主题编号"
        ),
        "all_detail_rows": len(all_rows),
        "included_rows": len(included_rows),
        "unmatched_included_rows": (
            sum(
                1
                for row in all_rows
                if _is_included(row)
                and _text(row.get("人工主题编号"))
                and not (
                    _prediction_for_row(row, predictions)
                    if predictions is not None
                    else _text(row.get("模型主题编号"))
                )
            )
            if predictions is not None
            else 0
        ),
        "excluded_or_unlabeled_rows": len(excluded_rows),
        "purity_coverage": _purity_and_coverage(included_rows),
        "pairwise": _pairwise_metrics(included_rows),
        "multi_topic_split": _multi_topic_split_stats(included_rows),
        "product_boundary": _product_boundary_metrics(included_rows),
        "business_line_boundary": _business_line_boundary_metrics(
            included_rows
        ),
        "by_category": _category_metrics(included_rows),
        "errors": _error_examples(included_rows, limit=error_limit),
        "notes": [
            "指标只使用“纳入准确率=是”的原子问题。",
            "人工同一会话拆出的多个原子问题保留为独立金标准原子。",
            "模型预测优先按“真实会话ID+品类”匹配，品类缺失时才回退单ID。",
            "回收业务层级优先读取金标准字段；旧表缺失时，聚合回收按品类标记识别，其余按自营回收统计。",
            "“58”仅作为人工备注占位值，不作为错误判定依据。",
        ],
    }
    return summary


def _write_markdown(path: Path, summary: dict[str, Any]) -> None:
    purity = summary["purity_coverage"]
    pairwise = summary["pairwise"]
    split = summary["multi_topic_split"]
    product_boundary = summary["product_boundary"]
    business_line_boundary = summary["business_line_boundary"]
    lines = [
        "# Case4 聚类准确率评估",
        "",
        f"- 纳入人工原子问题：{summary['included_rows']}",
        f"- 排除或未纳入：{summary['excluded_or_unlabeled_rows']}",
        "",
        "## 主指标",
        "",
        "| 指标 | 数值 |",
        "|---|---:|",
        f"| 预测簇纯度（Precision） | {purity['purity']:.2%} |",
        f"| 金标准主题覆盖（Recall） | {purity['coverage']:.2%} |",
        f"| 纯度-覆盖 F1 | {purity['f1']:.2%} |",
        f"| 成对 Precision | {pairwise['precision']:.2%} |",
        f"| 成对 Recall | {pairwise['recall']:.2%} |",
        f"| 成对 F1 | {pairwise['f1']:.2%} |",
        "",
        "## 多主题会话",
        "",
        f"- 纳入会话：{split['included_sessions']}",
        f"- 人工拆成多个主题的会话：{split['multi_topic_sessions']}",
        f"- 受影响人工原子问题：{split['multi_topic_atoms']}",
        f"- 模型仍保持在同一主题的会话：{split['model_kept_multi_topic_session_together']}",
        f"- 模型已拆开的会话：{split['model_split_multi_topic_session']}",
        "",
        "## 品类硬边界",
        "",
        f"- 跨品类预测簇：{product_boundary['cross_product_cluster_count']}",
        "",
        "## 回收业务层级硬边界",
        "",
        (
            "- 跨回收业务层级预测簇："
            f"{business_line_boundary['cross_business_line_cluster_count']}"
        ),
        "",
        "## 品类指标",
        "",
        "| 品类 | 原子数 | 预测簇 | 人工主题 | 纯度 | 覆盖 | F1 | 成对F1 |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for item in summary["by_category"]:
        lines.append(
            "| {category} | {included_atoms} | {predicted_clusters} | "
            "{gold_topics} | {purity:.2%} | {coverage:.2%} | {f1:.2%} | "
            "{pairwise_f1:.2%} |".format(**item)
        )
    lines.extend(["", "## 高频错误合并", ""])
    for item in summary["errors"]["error_merges"]:
        lines.append(
            f"- 模型主题 `{item['model_topic']}`："
            f"{item['included_atoms']} 个原子，混入 {item['gold_topic_count']} 个人工主题；"
            f"分布 `{item['gold_topic_distribution']}`。"
        )
    lines.extend(["", "## 高频错误拆分", ""])
    for item in summary["errors"]["error_splits"]:
        lines.append(
            f"- 人工主题 `{item['gold_topic']}`："
            f"{item['included_atoms']} 个原子被拆到 {item['model_topic_count']} 个模型主题；"
            f"分布 `{item['model_topic_distribution']}`。"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="在本地使用Case4人工校正表评估主题聚类准确性。",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path(
            "outputs/case4-gpt-topic-clusters-reviewed-20260802/"
            "case4-gpt-topic-clusters-人工校正.xlsx"
        ),
    )
    parser.add_argument("--output-json", type=Path)
    parser.add_argument("--output-md", type=Path)
    parser.add_argument(
        "--model-workbook",
        type=Path,
        help="待评估的新模型工作簿，可为原子问题明细或cluster-only结果",
    )
    parser.add_argument(
        "--model-json",
        type=Path,
        help="待评估的direct_mimo结果JSON",
    )
    parser.add_argument("--error-limit", type=int, default=30)
    args = parser.parse_args()
    summary = evaluate(
        args.workbook,
        error_limit=max(1, args.error_limit),
        model_workbook=args.model_workbook,
        model_json=args.model_json,
    )
    if args.output_json:
        args.output_json.parent.mkdir(parents=True, exist_ok=True)
        args.output_json.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    if args.output_md:
        args.output_md.parent.mkdir(parents=True, exist_ok=True)
        _write_markdown(args.output_md, summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
