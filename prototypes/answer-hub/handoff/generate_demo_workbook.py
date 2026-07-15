from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = [
    "序号",
    "上传者",
    "分析时间",
    "工单ID",
    "回收单号",
    "聊天内容",
    "图片链接",
    "核心问题",
    "判定结论",
    "判定依据",
    "产品类型",
    "一级分类",
    "二级分类",
    "参考话术",
]

ROWS = [
    [
        1,
        "demo",
        "2026-07-15 10:00:00",
        "DEMO-001",
        "ORDER-001",
        "回收师：屏幕白色背景下有一块颜色不均匀，怎么确认是不是色斑？\n"
        "答疑：请补拍白屏和灰屏照片，并确认异常区域是否固定。",
        "",
        "手机屏幕出现颜色不均匀时如何确认是否属于色斑",
        "需要补充图片后核验",
        "应确认异常位置、范围以及在不同纯色背景下是否持续存在",
        "手机",
        "显示问题",
        "色斑",
        "请提供白屏和灰屏下的清晰照片。",
    ],
    [
        2,
        "demo",
        "2026-07-15 10:05:00",
        "DEMO-002",
        "ORDER-002",
        "现场人员：这台手机亮屏以后局部发黄，算不算屏幕色斑？\n"
        "答疑：先切换纯白和浅灰背景，观察发黄区域是否保持不变。",
        "",
        "亮屏局部发黄应如何检查是否为色斑",
        "暂不直接判定",
        "需要通过纯色背景检查异常区域是否固定并补充清晰证据",
        "手机",
        "显示问题",
        "屏幕异常",
        "请切换纯色背景后重新拍摄屏幕。",
    ],
    [
        3,
        "demo",
        "2026-07-15 10:10:00",
        "DEMO-003",
        "ORDER-003",
        "回收师：设置里显示的名称和外壳看起来不一致，怎么确定具体型号？\n"
        "答疑：需要核对系统型号、设备标识和实物特征。",
        "",
        "系统名称与实物不一致时如何确认设备机型",
        "需要多项信息交叉核对",
        "单独依赖设备名称可能不准确，应结合系统信息和实物特征",
        "手机",
        "基本信息",
        "机型",
        "请提供设置页型号和设备外观信息。",
    ],
    [
        4,
        "demo",
        "2026-07-15 10:15:00",
        "DEMO-004",
        "ORDER-004",
        "现场人员：客户说这是某个型号，但系统里只看到通用名称，应该怎么查？\n"
        "答疑：先查系统型号，再结合机身标识确认。",
        "",
        "系统只显示通用名称时怎样查询手机具体型号",
        "按查询流程核对",
        "需要结合系统型号、机身标识和外观差异进行确认",
        "手机",
        "设备信息",
        "型号查询",
        "请核对系统型号和机身标识。",
    ],
    [
        5,
        "demo",
        "2026-07-15 10:20:00",
        "DEMO-005",
        "ORDER-005",
        "回收师：后盖有几条浅浅的线，在正常灯光下能看到，怎么判断划痕程度？\n"
        "答疑：需要确认长度、深度和正常视距下是否明显。",
        "",
        "手机后盖浅划痕应如何检查程度",
        "按外观检查流程确认",
        "检查划痕位置、长度、深度和正常光线下的可见程度",
        "手机",
        "外观问题",
        "后盖划痕",
        "请在正常光线下拍摄后盖整体和划痕近照。",
    ],
    [
        6,
        "demo",
        "2026-07-15 10:25:00",
        "DEMO-006",
        "ORDER-006",
        "现场人员：手机背面有一条细痕，侧光才能看见，需要怎么记录？\n"
        "答疑：先清洁表面，再检查细痕是否真实存在以及是否有明显深度。",
        "",
        "手机背面细痕在侧光下可见时如何检查",
        "需要清洁后复查",
        "应排除污渍并检查痕迹的长度、深度和正常视角可见性",
        "手机",
        "外观问题",
        "背板磨损",
        "请先清洁表面，再补充正常光线和侧光照片。",
    ],
]


def main() -> None:
    output_path = Path(__file__).resolve().parents[1] / "examples" / "semantic_clustering_demo.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "source_data"
    worksheet.append(HEADERS)
    for row in ROWS:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="176B87")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 8,
        "B": 10,
        "C": 20,
        "D": 14,
        "E": 14,
        "F": 58,
        "G": 18,
        "H": 40,
        "I": 24,
        "J": 42,
        "K": 12,
        "L": 16,
        "M": 16,
        "N": 38,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
